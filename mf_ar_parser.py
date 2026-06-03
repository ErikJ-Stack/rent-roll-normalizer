"""
MF AR-aging parser + Bldg-Unit join.

Parses an operator Aged-Receivables report into per-resident aging rows and
joins them to parsed `MFUnit`s on a normalized Bldg-Unit key (decision
SPEC-MF §2.7.3): two-way unmatched reporting (never silently drop or
fabricate). The joined aging lands in `Rent Roll Analysis` cols Q–T.

Public API:
    parse_mf_ar(source) -> MFARResult
    join_ar_to_units(units, ar_result) -> ARJoinReport
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

import openpyxl

from mf_normalizer import MFUnit, unit_key

_HEADER_MAP = [
    ("bldg", "bldg_unit"), ("unit", "bldg_unit"),
    ("resident", "resident"), ("tenant", "resident"),
    # NB: _norm() strips hyphens, so "0-30 Days" -> "030 days". Needles match the
    # normalized (hyphen-stripped) header text. Order matters: 6190 before 90.
    ("030", "ar_0_30"), ("0 30", "ar_0_30"),
    ("3160", "ar_31_60"), ("31 60", "ar_31_60"),
    ("6190", "ar_61_90"), ("61 90", "ar_61_90"),
    ("over 90", "ar_90_plus"), ("90", "ar_90_plus"),
    ("pre-payment", "prepay"), ("prepayment", "prepay"), ("pre payment", "prepay"),
    ("balance", "balance"),
    ("note", "note"), ("delinquency note", "note"),
]


@dataclass
class MFARRow:
    bldg_unit: str
    resident: str
    ar_0_30: float = 0.0
    ar_31_60: float = 0.0
    ar_61_90: float = 0.0
    ar_90_plus: float = 0.0
    prepay: float = 0.0
    balance: float = 0.0
    note: str = ""


@dataclass
class MFARResult:
    rows: list[MFARRow]
    period_hint: str = ""
    warnings: list[str] = field(default_factory=list)

    @property
    def total_ar(self) -> float:
        return sum(r.balance for r in self.rows)


@dataclass
class ARJoinReport:
    matched: int = 0
    unmatched_ar: list[str] = field(default_factory=list)        # AR rows with no RR unit
    units_balance_no_ar: list[str] = field(default_factory=list)  # RR balance but no AR detail
    warnings: list[str] = field(default_factory=list)


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9 ]", "", str(s).strip().lower()) if s is not None else ""


def _num(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("$", "").strip()
        neg = s.startswith("(") and s.endswith(")")
        s = s.strip("()")
        try:
            return -float(s) if neg else float(s)
        except ValueError:
            return 0.0
    return 0.0


def _load_ws(source):
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    return wb, wb.worksheets[0]


def parse_mf_ar(source) -> MFARResult:
    wb, ws = _load_ws(source)
    rows = list(ws.iter_rows(values_only=True))
    maxc = ws.max_column

    # period hint from the small header band (e.g. "Mar 2026")
    period = ""
    for row in rows[:6]:
        for c in row or ():
            if isinstance(c, str) and re.match(
                r"^\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{4}\s*$",
                c, re.IGNORECASE):
                period = c.strip()
                break
        if period:
            break

    # header row = contains Bldg-Unit + an aging/balance label
    header_row = None
    for i, row in enumerate(rows[:15]):
        joined = " | ".join(_norm(c) for c in row)
        if ("bldg" in joined or "unit" in joined) and ("0-30" in joined or "0 30" in joined
                                                        or "balance" in joined or "aging" in joined):
            header_row = i
            break
    if header_row is None:
        wb.close()
        raise ValueError("Could not locate the AR-aging header row.")

    col: dict[str, int] = {}
    for c in range(maxc):
        h = _norm(rows[header_row][c])
        if not h:
            continue
        for needle, fname in _HEADER_MAP:
            if needle in h and fname not in col:
                col[fname] = c
                break

    def g(row, fname):
        idx = col.get(fname)
        return row[idx] if idx is not None and idx < len(row) else None

    out: list[MFARRow] = []
    for row in rows[header_row + 1:]:
        a = g(row, "bldg_unit")
        if a in (None, ""):
            continue
        if _norm(a).startswith(("total", "grand", "report")):
            break
        out.append(MFARRow(
            bldg_unit=str(a).strip(),
            resident=str(g(row, "resident") or "").strip(),
            ar_0_30=_num(g(row, "ar_0_30")), ar_31_60=_num(g(row, "ar_31_60")),
            ar_61_90=_num(g(row, "ar_61_90")), ar_90_plus=_num(g(row, "ar_90_plus")),
            prepay=_num(g(row, "prepay")), balance=_num(g(row, "balance")),
            note=str(g(row, "note") or "").strip(),
        ))
    wb.close()

    warnings = []
    if not out:
        warnings.append("No AR rows parsed — check the aging-report layout.")
    return MFARResult(rows=out, period_hint=period, warnings=warnings)


def join_ar_to_units(units: list[MFUnit], ar: MFARResult) -> ARJoinReport:
    """Apply AR aging to matching units (by Bldg-Unit). Mutates units in place;
    returns a two-way unmatched report."""
    by_key = {unit_key(u.bldg_unit): u for u in units}
    rep = ARJoinReport()
    matched_keys = set()
    for r in ar.rows:
        k = unit_key(r.bldg_unit)
        u = by_key.get(k)
        if u is None:
            rep.unmatched_ar.append(r.bldg_unit)
            continue
        u.ar_0_30, u.ar_31_60 = r.ar_0_30, r.ar_31_60
        u.ar_61_90, u.ar_90_plus = r.ar_61_90, r.ar_90_plus
        if r.note and not u.notes:
            u.notes = r.note
        matched_keys.add(k)
        rep.matched += 1

    for u in units:
        if u.balance and unit_key(u.bldg_unit) not in matched_keys:
            rep.units_balance_no_ar.append(u.bldg_unit)

    if rep.unmatched_ar:
        rep.warnings.append(
            f"{len(rep.unmatched_ar)} AR row(s) did not match any unit "
            f"(e.g. {', '.join(rep.unmatched_ar[:5])}).")
    if rep.units_balance_no_ar:
        rep.warnings.append(
            f"{len(rep.units_balance_no_ar)} unit(s) carry a balance but had no AR-aging "
            "detail (aging report may be from a different period).")
    return rep


if __name__ == "__main__":
    import sys
    from mf_normalizer import parse_mf_rr
    ar = parse_mf_ar(sys.argv[1])
    print(f"AR rows={len(ar.rows)} | period={ar.period_hint!r} | total AR ${ar.total_ar:,.2f}")
    for w in ar.warnings:
        print("  WARN:", w)
    if len(sys.argv) > 2:
        rr = parse_mf_rr(sys.argv[2])
        rep = join_ar_to_units(rr.units, ar)
        print(f"JOIN: matched {rep.matched}/{len(ar.rows)} | unmatched AR {len(rep.unmatched_ar)} "
              f"| units w/ balance no AR {len(rep.units_balance_no_ar)}")
        joined = sum(u.ar_0_30 + u.ar_31_60 + u.ar_61_90 + u.ar_90_plus for u in rr.units)
        print(f"  joined aging sum ${joined:,.2f}")
        for w in rep.warnings:
            print("  WARN:", w)
