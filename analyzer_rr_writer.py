"""
Analyzer RR Writer
==================

Loads the destination Analyzer workbook, writes the translated rent roll
into the "Rent Roll Input" sheet starting at row 7 (cols A through S), and
returns the modified workbook as bytes.

Module name was previously `t12_writer.py` — a historical artifact from
when the destination workbook was a standalone "T12 intake template"
(predates the bundled-Analyzer-by-default flow that shipped in RR v1.12.0).
Renamed for clarity in 2026-05-10. The companion T12-side writer that
populates the Analyzer's `T12 Input` sheet lives in
`t12_normalizer_writer.py`.

Critical preservation guarantees:
  - All other tabs are untouched
  - Cols T and U (formulas) are untouched
  - Data validations on cols D, E, F, K, P remain in place
  - Header/formatting on rows 1-6 unchanged
  - Old data in A7:S606 from previous runs is cleared before writing
    (so the file is idempotent — re-running on a different rent roll
    won't leave ghost rows from the prior run)

Limit: max 600 bed rows per run (matches the formula extent in cols T-U).

Companion renames shipped 2026-05-15 (UW-BACKLOG BL-0011):
  - `populate_t12()` -> `populate_rr_input()` (matches what the function
    actually populates — the Rent Roll Input sheet, not the T12 Input
    sheet — and mirrors the partner `populate_t12_input()` on
    `t12_normalizer_writer.py`)
  - `T12CapacityError` -> `AnalyzerRRCapacityError` (matches the file
    rename done on 2026-05-10)
"""

from __future__ import annotations

import datetime as dt
import io
from typing import Optional

import openpyxl
import pandas as pd

from property_name import derive_property_name


# Layout constants — match the Analyzer substrate.
# Substrate v0.1.9 layout (pre-v1.16.0 / pre-substrate-v0.1.10):
#   A-R = 18 source cols from Condensed_RR
#   S   = Period Date (set by this writer)
#   T   = Total LOC $ formula (=IFERROR(L7+M7+N7+O7,0))
#   U   = Total Monthly Rev formula (=IFERROR(H7+IFERROR(I7,0)+T7,0))
# Substrate v0.1.10 layout (this writer + RR v1.16.0):
#   V   = 2nd Person Rent $          (new at v1.16.0)
#   W   = Move-out Date              (new at v1.16.0)
#   X   = Balance                    (new at v1.16.0)
#   Y   = Notes                      (new at v1.16.0)
#   Z   = Market PSF                 (new at v1.16.0)
#   AA  = Actual PSF                 (new at v1.16.0)
#   AB  = ACH                        (new at v1.16.0)
#   U formula extended to include +V (2nd Person Rent) in Total Monthly Rev.
# Substrate v0.1.13 layout (per-fee ancillary block):
#   AC  = Meal Plan $                (new at v1.17.0)
#   AD  = Scooter Fee $              (new at v1.17.0)
#   AE  = Housekeeping $             (new at v1.17.0)
#   AF  = Laundry $                  (new at v1.17.0)
#   AG  = Pet $                      (new at v1.17.0)
#   AH  = Total Ancillary $ formula  (substrate v0.1.13)
# Substrate v0.2.13 layout (this writer + RR v1.18.0):
#   AI  = Preleased Date             (new at v1.18.0 — BL-0025, exposure)
SHEET_NAME = "Rent Roll Input"
DATA_START_ROW = 7
DATA_END_ROW = 606  # also the formula extent in cols T, U
COL_A_TO_R_COUNT = 18  # 18 cols from Condensed_RR mapped to A-R
COL_S_INDEX = 19       # Period Date column
COL_V_INDEX = 22       # 2nd Person Rent $ column (start of v1.16.0 extension)
COL_AB_INDEX = 28      # ACH column (end of v1.16.0 extension)
COL_AC_INDEX = 29      # Meal Plan $ (start of v1.17.0 per-fee ancillary block)
COL_AG_INDEX = 33      # Pet $ (end of v1.17.0 per-fee ancillary block)
COL_AI_INDEX = 35      # Preleased Date (v1.18.0; AH=34 holds the Total Ancillary $ formula)

# The 18 source columns in the order the Analyzer's Rent Roll Input expects them
# at cols A-R. These names must match the Condensed_RR column names exactly.
SOURCE_COLUMNS_A_TO_R = [
    "Unit #",          # A
    "Room #",          # B
    "Sq Ft",           # C
    "Care Type",       # D
    "Status",          # E
    "Apt Type",        # F
    "Market Rate",     # G
    "Actual Rate",     # H
    "Concession $",    # I
    "Concession End Date",  # J
    "Care Level",      # K
    "Care Level $",    # L
    "Med Mgmt $",      # M
    "Pharmacy $",      # N
    "Other LOC $",     # O
    "Payer Type",      # P
    "Move-in Date",    # Q
    "Resident Name",   # R
]

# 7 new source columns at v1.16.0, mapped to Rent Roll Input cols V-AB
# (S=Period Date, T-U=formulas remain in place; new data sits after).
SOURCE_COLUMNS_V_TO_AB = [
    "2nd Person Rent $",  # V
    "Move-out Date",      # W
    "Balance",            # X
    "Notes",              # Y
    "Market PSF",         # Z
    "Actual PSF",         # AA
    "ACH",                # AB
]

# 5 new source columns at v1.17.0 (UW-BACKLOG BL-0003), mapped to Rent Roll
# Input cols AC-AG. Per-fee ancillary breakdown — splits what previously
# lumped into "Other LOC $" (col O) into named buckets so Section M2/M4 on
# the Analyzer can compute per-fee capture rates against the operator's
# published schedule. Other LOC $ (col O) remains as the catchall for
# unmatched care headers (Diabetes, Misc, anything not in the named
# buckets). Total LOC $ (col T formula) extended in substrate v0.1.13 to
# include AC-AG so the per-resident total is unchanged.
SOURCE_COLUMNS_AC_TO_AG = [
    "Meal Plan $",        # AC
    "Scooter Fee $",      # AD
    "Housekeeping $",     # AE
    "Laundry $",          # AF
    "Pet $",              # AG
]

# 1 new source column at v1.18.0 (UW-BACKLOG BL-0025), mapped to Rent Roll
# Input col AI. AH=34 is reserved for the Total Ancillary $ formula
# (substrate v0.1.13) so we skip it. Preleased Date pairs with
# Status="Preleased" (also new at v1.18.0) — Section N on Rent Roll Recon
# uses both for the exposure rollup (point-in-time net exposure + forward
# NTV departure buckets).
SOURCE_COLUMNS_AI = [
    "Preleased Date",     # AI
]


class AnalyzerRRCapacityError(Exception):
    """Raised when the rent roll has more rows than Rent Roll Input can hold."""


def _coerce_value(v):
    """Convert pandas/numpy scalars to native Python types openpyxl can write.

    Handles: NaN/NaT → None, numpy ints/floats → int/float, pandas Timestamps
    → datetime.date, anything else → unchanged.
    """
    if v is None:
        return None
    # Pandas null sentinels
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        # pd.isna on some array-like objects raises; skip for those
        pass

    # Pandas/numpy datetime-like → date
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime().date()
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v

    # Numpy scalars
    if hasattr(v, "item"):
        try:
            return v.item()
        except (ValueError, AttributeError):
            pass

    return v


def populate_rr_input(
    analyzer_bytes: bytes,
    translated_df: pd.DataFrame,
    period_date: Optional[dt.date],
    *,
    source_filename: str = "",
) -> bytes:
    """Populate the Analyzer's Rent Roll Input sheet with translated rent roll data and return as bytes.

    Args:
        analyzer_bytes: Raw bytes of the destination Analyzer .xlsx file
        translated_df: DataFrame from analyzer_rr_translator.translate_for_t12()
            — must have the 18 source columns in SOURCE_COLUMNS_A_TO_R
        period_date: Date written to col S on every populated row. Required.
        source_filename: original RR filename. When non-empty, the derived
            property name (via property_name.derive_property_name) is
            written to `Rent Roll Input!A3` per substrate v0.1.8. Empty
            string leaves A3 untouched.

    Raises:
        AnalyzerRRCapacityError: if the rent roll exceeds DATA_END_ROW - DATA_START_ROW + 1 rows
        ValueError: if the Analyzer doesn't contain the expected sheet
    """
    if period_date is None:
        raise ValueError("period_date is required to populate the Analyzer.")

    n_rows = len(translated_df)
    max_rows = DATA_END_ROW - DATA_START_ROW + 1
    if n_rows > max_rows:
        raise AnalyzerRRCapacityError(
            f"Rent roll has {n_rows} bed rows, but the Analyzer 'Rent Roll Input' "
            f"sheet's formulas only extend to row {DATA_END_ROW} "
            f"(max {max_rows} rows). Either trim the rent roll or extend "
            f"the Analyzer formulas to additional rows."
        )

    # Load the destination Analyzer (preserves formulas, formatting, validations, other tabs)
    wb = openpyxl.load_workbook(io.BytesIO(analyzer_bytes), data_only=False)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Analyzer workbook is missing the required '{SHEET_NAME}' sheet. "
            f"Found sheets: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]

    # --- Step 1: Clear any pre-existing data ------------------------------
    # Idempotency: clearing happens before writing so re-running with a
    # different RR doesn't leave ghost rows.
    #   - Cols A-S: source data + period date (always clear)
    #   - Cols T-U: formulas — DO NOT clear (preserved by the Analyzer substrate)
    #   - Cols V-AB: v1.16.0 extension fields (always clear)
    #   - Cols AC-AG: v1.17.0 per-fee ancillary fields (always clear)
    #   - Col AH: Total Ancillary $ formula — DO NOT clear (substrate v0.1.13)
    #   - Col AI: v1.18.0 Preleased Date (always clear)
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        for c in range(1, COL_S_INDEX + 1):  # cols 1-19 = A-S
            ws.cell(row=r, column=c).value = None
        for c in range(COL_V_INDEX, COL_AB_INDEX + 1):  # cols 22-28 = V-AB
            ws.cell(row=r, column=c).value = None
        for c in range(COL_AC_INDEX, COL_AG_INDEX + 1):  # cols 29-33 = AC-AG
            ws.cell(row=r, column=c).value = None
        ws.cell(row=r, column=COL_AI_INDEX).value = None  # col 35 = AI

    # --- Step 2: Write the translated rent roll ---------------------------
    # Match by column NAME (not position) — defensive against future
    # Condensed_RR column reordering.
    missing = [c for c in SOURCE_COLUMNS_A_TO_R if c not in translated_df.columns]
    if missing:
        raise ValueError(
            f"Translated DataFrame is missing required columns: {missing}"
        )
    # v1.16.0 extension cols are optional — if a translated_df was produced
    # by a pre-v1.16.0 normalizer they won't be present. Detect + skip.
    has_v116_cols = all(c in translated_df.columns for c in SOURCE_COLUMNS_V_TO_AB)
    # v1.17.0 per-fee ancillary cols are optional too.
    has_v117_cols = all(c in translated_df.columns for c in SOURCE_COLUMNS_AC_TO_AG)
    # v1.18.0 Preleased Date is optional too.
    has_v118_cols = all(c in translated_df.columns for c in SOURCE_COLUMNS_AI)

    for i, (_, row) in enumerate(translated_df.iterrows()):
        excel_row = DATA_START_ROW + i
        # Cols A-R (1-18) ← the 18 source columns in fixed order
        for col_idx, src_col in enumerate(SOURCE_COLUMNS_A_TO_R, start=1):
            value = _coerce_value(row[src_col])
            ws.cell(row=excel_row, column=col_idx).value = value
        # Col S (19) ← period date on every populated row, formatted as date
        s_cell = ws.cell(row=excel_row, column=COL_S_INDEX)
        s_cell.value = period_date
        s_cell.number_format = "mm/dd/yyyy"
        # Cols V-AB (22-28) ← v1.16.0 extension fields, when available
        if has_v116_cols:
            for offset, src_col in enumerate(SOURCE_COLUMNS_V_TO_AB):
                col_idx = COL_V_INDEX + offset
                value = _coerce_value(row[src_col])
                ws.cell(row=excel_row, column=col_idx).value = value
            # Date formats for the date cells in this group
            ws.cell(row=excel_row, column=COL_V_INDEX + 1).number_format = "mm/dd/yyyy"  # W: Move-out Date
        # Cols AC-AG (29-33) ← v1.17.0 per-fee ancillary fields, when available
        if has_v117_cols:
            for offset, src_col in enumerate(SOURCE_COLUMNS_AC_TO_AG):
                col_idx = COL_AC_INDEX + offset
                value = _coerce_value(row[src_col])
                ws.cell(row=excel_row, column=col_idx).value = value
        # Col AI (35) ← v1.18.0 Preleased Date, when available. Skips AH=34
        # which holds the Total Ancillary $ formula from substrate v0.1.13.
        if has_v118_cols:
            ai_cell = ws.cell(row=excel_row, column=COL_AI_INDEX)
            ai_cell.value = _coerce_value(row["Preleased Date"])
            ai_cell.number_format = "mm/dd/yyyy"

    # --- Step 3: Stamp property name into A3 (substrate v0.1.8 source cell)
    # Only writes when derivation produces something non-empty, so a bad
    # filename doesn't clobber an analyst-typed value carried in from a
    # prior session.
    if source_filename:
        derived = derive_property_name(source_filename)
        if derived:
            ws["A3"].value = derived

    # --- Step 4: Save and return ------------------------------------------
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
