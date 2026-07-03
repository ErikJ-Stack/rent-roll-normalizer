"""
uw_template_writer.py — Track 4 / Phase 2

Pure function that populates the ALF UW Template from a populated Analyzer
workbook, driven by the mapping registry at `tools/uw_template/registry.json`.

Public API:
    populate_uw_template(analyzer_bytes, template_bytes, *,
                         template_version='v4',
                         scenario='normalized',
                         registry_path=None) -> (bytes, PopulateReport)

Conventions:

  - **Analyzer** is loaded with `data_only=True` because UW Output cells are
    formula references into T12 Analytics. The Analyzer must have been
    saved through Excel (or LibreOffice) at least once so the cached values
    are present. The reader does not evaluate formulas itself.
  - **Template** is loaded with `data_only=False` so formula columns on
    Rent Roll Analysis (V, X, Y, Z, AA, AB, AS — see registry's
    intake_targets_unmapped) and all of the downstream output sheets
    (Cover, Scenarios, P&L, Waterfall, ...) preserve their formulas
    untouched. The writer only mutates cells corresponding to mapped
    concepts.
  - **Scenarios** parameter selects T-12 path source column:
      'normalized'  → UW Output col F (default; contract §8 says this is
                      the underwriting figure)
      't12_actual'  → UW Output col E
  - **Bad debt placement:** by default the writer writes the Analyzer's
    bad-debt value to T-12 Analysis!N62 (revenue contra-line — template's
    structural choice). The Analyzer's opex Bad Debt Expense (UW Output
    row 57) is mapped to template N106 in the registry but flagged as a
    duplicate target. The writer skips N106 to avoid double-counting; see
    `_SPECIAL_SKIP_KEYS` below. Override by passing the key in
    `allow_special_keys`.

The writer does NOT modify the Analyzer (read-only) and does NOT mutate
the input bytes (returns new bytes via `io.BytesIO`).
"""
from __future__ import annotations

import io
import json
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Defaults
# ──────────────────────────────────────────────────────────────────────────────

DEFAULT_REGISTRY = Path(__file__).resolve().parent / "tools" / "uw_template" / "registry.json"

# Statuses for which the writer skips by default. Override by passing
# `include_statuses=(...)` to populate_uw_template.
_DEFAULT_SKIP_STATUSES = frozenset({
    "gap_source",
    "gap_target",
    "header_only",
    "derived",
    "manual",
    "substrate_ready_parser_pending",
    "decided_pending_upstream",
})

# Hard-coded skip set — concepts the writer deliberately does not populate
# because including them would double-count or conflict with another concept.
# Each key maps to a one-line reason for the report.
_SPECIAL_SKIP_KEYS: dict[str, str] = {
    "opex_bad_debt_expense": (
        "Bad Debt Expense (UW Output row 57) is already written to N62 "
        "(revenue contra-line) via `bad_debt_writeoffs_revenue`. Writing "
        "to N106 too would double-count. Override with allow_special_keys."
    ),
}

# ── T-12 Analysis Layer-3 model (v5 / v6 templates) ────────────────────────────
# The Layer-3 income section has TOTAL rows that are live template formulas
# (e.g. v5 N63 =N58+…, N69 =N63+…; v6 N61 =SUM(N58:N60), N77 EGI =N61+N65+…).
# The writer must NOT paste values over those formulas, and the line items it
# does paste must carry the sign the additive formulas expect.
#
# Total-row concepts: skipped by the generic loop (keyed by concept, version-
# agnostic) and handled by `_finalize_t12_layer3`, which preserves/authors the
# col-N formula and mirrors it across the monthly grid B–M. Per-version row
# positions live in `_T12_LAYOUT` below (v6 rebuilt INCOME so every total row
# shifted — EGI N69→N77, EBITDARM N116→N131, etc.).
_T12_TOTAL_CONCEPTS: frozenset[str] = frozenset({
    "base_rent_normalized",    # v5 Net Rent Revenue (formula); nulled in v6
    "egi",                     # EGI               (v5 N69 / v6 N77)
    "labor_total",             # Total Labor       (v5 N85 / v6 N99)
    "opex_nonlabor_total",     # Total Non-Labor   (v5 N111 / v6 N126)
    "opex_total_incl_mgmt",    # Total Op Ex       (v5 N114 / v6 N129)
    "opex_total_excl_mgmt",    # Op Ex excl mgmt   (v5 N115 / v6 N130)
    "ebitdarm",                # EBITDARM          (v5 N116 / v6 N131)
    "ebitdar",                 # EBITDAR (NOI)     (v5 N117 / v6 N132, authored)
    "ebitda",                  # EBITDA            (v5 N118 / v6 N133, authored)
})

# Income-waterfall contra lines: the Analyzer reports vacancy/bad-debt as
# positive magnitudes and loss-to-lease as a signed gap, but the template's
# additive Net Rent formula treats them as reductions. Negate on write so the
# waterfall subtracts them (GPR + LtL + Vacancy + Concessions + BadDebt = Net
# Rent). Concessions is already signed negative in the T12 GL, so it's excluded.
_T12_CONTRA_KEYS: frozenset[str] = frozenset({
    "loss_to_lease",
    "physical_vacancy_loss",
    "bad_debt_writeoffs_revenue",
})

# T-12 Analysis Layer-3 monthly grid spans cols B(2)..M(13); col N(14) = annual.
_T12_MONTH_COLS = tuple(range(2, 14))  # B..M
_N_REF_RE = re.compile(r"\bN(\d+)")

# Per-version T-12 Analysis Layer-3 + Section I/J row map. The finalize and
# Section-I passes dispatch on this rather than hardcoding v5 rows.
#
#   net_rent_row     v5 income Net Rent (GPR waterfall) — monthly pasted
#                    explicitly as base − concessions − bad debt. None for v6,
#                    whose income is an actual-T12 build (no GPR-derived line).
#   ebitdar_row/_formula, ebitda_row/_formula
#                    rows the template leaves as a value/blank — authored here.
#                    v5: EBITDAR = EBITDARM(N116) − mgmt(N113); EBITDA = EBITDAR.
#                    v6: EBITDAR = EBITDARM(N131) − mgmt(N128); EBITDA = EBITDAR.
#   mirror_rows      total rows whose col-N formula is mirrored across B–M so
#                    monthly totals reconcile to the annual (net_rent_row is
#                    excluded — it's pasted directly).
#   section_i_*/j_*  Section I (Layer 1 — Raw T-12) data band + Section J
#                    raw-totals reconciliation rows. v6 shifted these +15 (the
#                    raw header row moved from 122 to 137).
_T12_LAYOUT: dict[str, dict] = {
    "v5": {
        "net_rent_row": 63,
        "ebitdar_row": 117, "ebitdar_formula": "=N116-N113",
        "ebitda_row": 118,  "ebitda_formula": "=N117",
        "mirror_rows": (69, 85, 111, 114, 115, 116, 117, 118),
        "section_i_start": 123, "section_i_end": 172,
        "section_j_rev": 176, "section_j_opex": 177, "section_j_ebitdar": 178,
    },
    # v6 rev2 (operator Other-Care revision, canonical 2026-06-03): the income
    # restructure shifted Layer-3 totals +3 and Section I/J +3 vs the prior v6.
    # EGI N77→N80, Total Labor N99→N102, Total Non-Labor N126→N129,
    # EBITDARM N131→N134, EBITDAR N132→N135, EBITDA N133→N136. rev2 already
    # ships N135/N136 as formulas, so the author step is a no-op (gated on
    # "not already a formula") — formulas kept here for consistency + mirroring.
    "v6": {
        "net_rent_row": None,
        "ebitdar_row": 135, "ebitdar_formula": "=N134-N131",
        "ebitda_row": 136,  "ebitda_formula": "=N135-N128",
        "mirror_rows": (62, 67, 80, 102, 129, 132, 133, 134, 135, 136),
        "section_i_start": 141, "section_i_end": 190,
        "section_j_rev": 194, "section_j_opex": 195, "section_j_ebitdar": 196,
    },
    # v8 (operator template 2026-06-11, absorbed 2026-06-12): T-12 Analysis
    # layout verified cell-identical to v6 rev2 (all anchor rows checked
    # against the binary) — same entry. v8's changes live on Rent Roll
    # Analysis (paste grid re-anchored to header 213 / data 214+, new NER
    # col AV) and analyst-side sheets (Waterfall, Scenarios col F).
    "v8": {
        "net_rent_row": None,
        "ebitdar_row": 135, "ebitdar_formula": "=N134-N131",
        "ebitda_row": 136,  "ebitda_formula": "=N135-N128",
        "mirror_rows": (62, 67, 80, 102, 129, 132, 133, 134, 135, 136),
        "section_i_start": 141, "section_i_end": 190,
        "section_j_rev": 194, "section_j_opex": 195, "section_j_ebitdar": 196,
    },
    # v11 (operator template 2026-06-16, absorbed 2026-06-16): T-12 Analysis
    # layout verified cell-identical to v8 (197×20, all anchor rows checked) —
    # same entry. v11's changes live on Rent Roll Analysis (paste grid
    # re-anchored to header 223 / data 224+, new Section S concessions-audit
    # block, fill-downs extended to the full band) and Prop Info market-data.
    "v11": {
        "net_rent_row": None,
        "ebitdar_row": 135, "ebitdar_formula": "=N134-N131",
        "ebitda_row": 136,  "ebitda_formula": "=N135-N128",
        "mirror_rows": (62, 67, 80, 102, 129, 132, 133, 134, 135, 136),
        "section_i_start": 141, "section_i_end": 190,
        "section_j_rev": 194, "section_j_opex": 195, "section_j_ebitdar": 196,
    },
}

# Status colours mirror the generator — duplicated here so the writer can
# tag the report with consistent labels.
_STATUS_LABELS: dict[str, str] = {
    "mapped": "mapped",
    "proposed": "proposed",
    "gap_source": "gap_source",
    "gap_target": "gap_target",
    "header_only": "header_only",
    "manual": "manual",
    "derived": "derived",
    "decided_pending_upstream": "decided_pending_upstream",
    "substrate_ready_parser_pending": "substrate_ready_parser_pending",
}


# ──────────────────────────────────────────────────────────────────────────────
# Exceptions
# ──────────────────────────────────────────────────────────────────────────────

class UWTemplateWriterError(Exception):
    """Base exception for the UW Template writer."""


class TemplateVersionMissing(UWTemplateWriterError):
    """Raised when the registry has no entry for the requested template version."""


# ──────────────────────────────────────────────────────────────────────────────
# Report dataclasses
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class ConceptResult:
    """Per-concept outcome of a single populate run."""
    key: str
    path: str  # 't12' | 'rent_roll' | 'ar'
    status: str  # registry status
    outcome: str  # 'written' | 'skipped' | 'error' | 'no_source' | 'no_target' | 'written_list'
    target_address: str = ""  # 'Sheet!A1' when applicable
    cells_written: int = 0
    notes: str = ""
    sample_value: Any = None  # first / sample value for spot-checking
    computed_fallback: bool = False  # value came from the in-Python evaluator


@dataclass
class PopulateReport:
    """Structured outcome of populate_uw_template.

    Inspect `summary` for counts and `results` for per-concept detail.
    """
    template_version: str
    scenario: str
    summary: dict[str, int] = field(default_factory=dict)
    results: list[ConceptResult] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def write_count(self) -> int:
        return sum(r.cells_written for r in self.results)

    def by_outcome(self) -> dict[str, list[ConceptResult]]:
        out: dict[str, list[ConceptResult]] = {}
        for r in self.results:
            out.setdefault(r.outcome, []).append(r)
        return out

    def to_dict(self) -> dict:
        return {
            "template_version": self.template_version,
            "scenario": self.scenario,
            "summary": self.summary,
            "warnings": self.warnings,
            "results": [
                {
                    "key": r.key, "path": r.path, "status": r.status,
                    "outcome": r.outcome,
                    "target_address": r.target_address,
                    "cells_written": r.cells_written,
                    "computed_fallback": r.computed_fallback,
                    "notes": r.notes,
                    "sample_value": (
                        str(r.sample_value) if r.sample_value is not None else None
                    ),
                }
                for r in self.results
            ],
        }


# ──────────────────────────────────────────────────────────────────────────────
# Source resolution
# ──────────────────────────────────────────────────────────────────────────────

_ADDR_RE = re.compile(r"^\$?([A-Z]+)\$?(\d+)$")


def _split_qualified(qualified: str) -> tuple[str, str]:
    """Split 'Sheet!$A$5' into ('Sheet', 'A5').

    Handles single-quoted sheet names: "'Rent Roll Recon'!$B$2" → ('Rent Roll Recon', 'B2').
    """
    if "!" not in qualified:
        raise ValueError(f"Not a qualified address: {qualified!r}")
    sheet, addr = qualified.rsplit("!", 1)
    sheet = sheet.strip().strip("'")
    addr = addr.replace("$", "")
    return sheet, addr


def _resolve_uw_output_value(wb_analyzer, src: dict, scenario: str) -> Any:
    """Read a scalar value from UW Output at (row, column-by-scenario)."""
    sheet = src.get("sheet", "UW Output")
    row = src.get("row")
    col_spec = src.get("column", "")
    if row is None:
        return None
    if col_spec == "E_or_F":
        col = "F" if scenario == "normalized" else "E"
    else:
        col = col_spec  # 'B' / 'C' / 'D' for IL/AL/MC splits, etc.
    return wb_analyzer[sheet][f"{col}{row}"].value


def _resolve_named_range_value(wb_analyzer, src: dict) -> Any:
    """Read a scalar value from a named range, e.g. Property_Name → Cover!$B$5."""
    name = src.get("name")
    if not name or name not in wb_analyzer.defined_names:
        return None
    qualified = wb_analyzer.defined_names[name].value
    sheet, addr = _split_qualified(qualified)
    if sheet not in wb_analyzer.sheetnames:
        return None
    return wb_analyzer[sheet][addr].value


def _resolve_cell_value(wb_analyzer, src: dict) -> Any:
    """Read a scalar value from a literal Sheet!Cell address."""
    sheet = src.get("sheet")
    addr = src.get("address")
    if not sheet or not addr or sheet not in wb_analyzer.sheetnames:
        return None
    return wb_analyzer[sheet][addr].value


def _resolve_rr_input_column(wb_analyzer, src: dict) -> list[Any]:
    """Read a full column from Rent Roll Input, rows 7..606.

    Returns a list of length 600, with None for empty cells. The writer
    truncates to populated rows at write time.
    """
    sheet = src.get("sheet", "Rent Roll Input")
    col = src.get("column", "")
    if not col or sheet not in wb_analyzer.sheetnames:
        return []
    ws = wb_analyzer[sheet]
    return [ws[f"{col}{r}"].value for r in range(7, 607)]


def _compute_derived(wb_analyzer, concept: dict, scenario: str) -> Any:
    """Compute a derived value for a concept whose source.system == 'derived'.

    The set of derived keys is small and hard-coded.
    """
    key = concept.get("key")
    # NB: return None (not 0) when the UW Output source cells are all blank, so
    # the writer's computed-fallback supplies the in-Python value. A fresh
    # openpyxl-built Analyzer has no cached UW Output values → these reads come
    # back blank; summing them to 0 would otherwise mask the fallback (0 is not
    # "blank"), e.g. Prop Info!B15 Licensed Total showing 0 instead of the
    # RR-derived total.
    if key == "licensed_beds_total":
        # SUM of UW Output B70 + C70 + D70 (IL + AL + MC).
        ws = wb_analyzer["UW Output"]
        nums = [v for v in (ws[f"{c}70"].value for c in ("B", "C", "D"))
                if isinstance(v, (int, float))]
        return sum(nums) if nums else None
    if key == "opex_total_incl_mgmt":
        # SUM of UW Output row 63 (Total opex excl. mgmt) + row 64 (Mgmt fee)
        # in the active scenario column.
        col = "F" if scenario == "normalized" else "E"
        ws = wb_analyzer["UW Output"]
        nums = [v for v in (ws[f"{col}63"].value, ws[f"{col}64"].value)
                if isinstance(v, (int, float))]
        return sum(nums) if nums else None
    if key == "second_person_revenue":
        # Source says "Computed elsewhere" — for now, return None and let
        # the writer skip. Could later sum Rent Roll Input col V × 12.
        return None
    return None


def _resolve_source(wb_analyzer, concept: dict, scenario: str) -> Any:
    """Dispatch on concept.source.system to read the source value."""
    src = concept.get("source") or {}
    system = src.get("system")
    if system == "uw_output":
        return _resolve_uw_output_value(wb_analyzer, src, scenario)
    if system == "named_range":
        return _resolve_named_range_value(wb_analyzer, src)
    if system == "cell":
        return _resolve_cell_value(wb_analyzer, src)
    if system == "rr_input":
        return _resolve_rr_input_column(wb_analyzer, src)
    if system == "derived":
        return _compute_derived(wb_analyzer, concept, scenario)
    if system == "gap":
        return None
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Target write
# ──────────────────────────────────────────────────────────────────────────────

def _is_blank(value) -> bool:
    """Treat empty strings and pandas-NaN-likes as blank. Native None too."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _parse_target_addr(addr: str) -> tuple[str, int, bool]:
    """Return (col_letter, row, is_rowstride).

    Address forms:
      'N69'   → ('N', 69, False)
      'A211+' → ('A', 211, True)  — row-stride paste anchor
    """
    is_stride = addr.endswith("+")
    core = addr.rstrip("+")
    m = _ADDR_RE.match(core)
    if not m:
        raise ValueError(f"Unparseable target address: {addr!r}")
    return m.group(1), int(m.group(2)), is_stride


def _write_scalar(ws, addr: str, value: Any) -> None:
    """Write a scalar value to a single cell."""
    ws[addr] = value


def _write_monthly_grid(
    ws, row: int, monthly_values: list, start_col: int = 2, n_months: int = 12
) -> int:
    """Write 12 monthly values across the UW Template Layer-3 grid (cols B–M).

    The `T-12 Analysis` Layer-3 rows lay out 12 months in columns B..M with the
    annual total in column N. Those monthly cells are literal-`0` paste targets
    (not formulas — even the subtotal rows), so the writer must paste each
    month explicitly. Returns the count of cells written.
    """
    written = 0
    for i in range(min(n_months, len(monthly_values))):
        v = monthly_values[i]
        if v is None:
            continue
        ws.cell(row=row, column=start_col + i).value = v
        written += 1
    return written


def _cell_is_formula(ws, row: int, col: int) -> bool:
    """True if the cell holds an Excel formula (data_type 'f' or '='-string)."""
    c = ws.cell(row=row, column=col)
    v = c.value
    return c.data_type == "f" or (isinstance(v, str) and v.startswith("="))


def _mirror_n_formula(formula: str, col_letter: str) -> str:
    """Rewrite a column-N formula to another column.

    The Layer-3 total formulas only reference column N (same-column sums), e.g.
    `=N63+N64+N65+N66+N67+N68` → (col 'B') → `=B63+B64+B65+B66+B67+B68`,
    `=SUM(N71:N84)` → `=SUM(B71:B84)`.
    """
    return _N_REF_RE.sub(col_letter + r"\1", formula)


def _finalize_t12_layer3(ws, monthly: dict[str, list], layout: dict) -> int:
    """Make the T-12 Analysis Layer-3 total rows live formulas (col N + B–M).

    Runs after the generic concept loop (which writes the line items and skips
    the total-row concepts). Driven by the per-version `layout` row map. For
    each total row:
      - col N: preserve the template's formula; author EBITDAR / EBITDA where
        the template left them as a value/blank.
      - cols B–M: mirror the col-N formula across the 12 month columns — except
        v5's Net Rent (net_rent_row), whose monthly value is pasted directly
        (base rent − concessions − bad debt) because the GPR waterfall feeding
        it has no monthly dimension. v6 income is an actual-T12 build with no
        such line (net_rent_row=None), so the step is skipped.

    Returns the count of formula/value cells written.
    """
    from openpyxl.utils import get_column_letter

    written = 0

    # 1) Author the two missing P&L formulas (EBITDAR = EBITDARM − mgmt fee;
    #    EBITDA = EBITDAR − depreciation(0)) if the template left them as values.
    er, ed = layout["ebitdar_row"], layout["ebitda_row"]
    if not _cell_is_formula(ws, er, 14):
        ws.cell(row=er, column=14).value = layout["ebitdar_formula"]
        written += 1
    if not _cell_is_formula(ws, ed, 14):
        ws.cell(row=ed, column=14).value = layout["ebitda_formula"]
        written += 1

    # 2) v5 Net Rent monthly: base rent − concessions − bad debt, per month, so
    #    the monthly sum reconciles to the annual GPR-waterfall total.
    #    (concessions is already negative; bad debt is positive → subtract.)
    net_rent_row = layout.get("net_rent_row")
    if net_rent_row is not None:
        base = monthly.get("base_rent_normalized") or [0.0] * 12
        conc = monthly.get("concessions_specials") or [0.0] * 12
        bd = monthly.get("bad_debt_writeoffs_revenue") or [0.0] * 12
        for i, col in enumerate(_T12_MONTH_COLS):
            if i >= 12:
                break
            ws.cell(row=net_rent_row, column=col).value = (
                (base[i] or 0.0) + (conc[i] or 0.0) - (bd[i] or 0.0)
            )
            written += 1

    # 3) Mirror each total row's col-N formula across B–M. (Runs after step 1 so
    #    the just-authored EBITDAR/EBITDA formulas get mirrored too.)
    for row in layout["mirror_rows"]:
        n_formula = ws.cell(row=row, column=14).value
        if not (isinstance(n_formula, str) and n_formula.startswith("=")):
            continue
        for col in _T12_MONTH_COLS:
            ws.cell(row=row, column=col).value = _mirror_n_formula(
                n_formula, get_column_letter(col)
            )
            written += 1

    return written


def _write_section_i_raw(ws, raw_lines: list, layout: dict) -> tuple[int, list[str]]:
    """Populate Section I (Layer 1 — Raw T-12) from the summarized raw lines.

    Rebuilds the section: clears the pre-filled skeleton (the data band given by
    `layout`, cols A–P) and writes one row per Analyzer label —
    B = matched GL account names, C–N = 12 monthly values, O = T-12 total,
    P = the standardized bucket (label). Authors the Section J raw-totals
    reconciliation (Total Revenue / Total OpEx / EBITDAR) as live SUM formulas
    over the rows just written. Row positions come from `layout` (v6 shifted the
    whole band +15 vs v5).

    Returns (cells_written, warnings).
    """
    si_start = layout["section_i_start"]
    si_end = layout["section_i_end"]
    sj_rev = layout["section_j_rev"]
    sj_opex = layout["section_j_opex"]
    sj_ebitdar = layout["section_j_ebitdar"]

    warnings: list[str] = []
    capacity = si_end - si_start + 1
    if len(raw_lines) > capacity:
        warnings.append(
            f"Section I holds {capacity} raw lines but the T-12 has "
            f"{len(raw_lines)} — truncating."
        )
        raw_lines = raw_lines[:capacity]

    # Clear the skeleton (A..P) across the full data band.
    for r in range(si_start, si_end + 1):
        for c in range(1, 17):
            ws.cell(row=r, column=c).value = None

    written = 0
    rev_rows: list[int] = []
    opex_rows: list[int] = []
    for i, line in enumerate(raw_lines):
        r = si_start + i
        ws.cell(row=r, column=2).value = " | ".join(line.get("descriptions") or []) or line["label"]
        mvals = line.get("monthly") or []
        # Section I months are cols C..N (3..14); O(15)=total, P(16)=bucket.
        for j in range(min(12, len(mvals))):
            ws.cell(row=r, column=3 + j).value = mvals[j]
        ws.cell(row=r, column=15).value = line.get("total")
        ws.cell(row=r, column=16).value = line["label"]
        written += 3 + min(len(mvals), 12)  # B + O + P + months
        (rev_rows if line.get("section") == "Revenue" else opex_rows).append(r)

    # Section J — raw-totals reconciliation (live SUM formulas over O-column).
    def _sum_o(rows: list[int]) -> str:
        if not rows:
            return "0"
        return "=" + "+".join(f"O{r}" for r in rows) if len(rows) <= 3 else (
            f"=SUM(O{rows[0]}:O{rows[-1]})"
        )

    ws.cell(row=sj_rev, column=15).value = _sum_o(rev_rows)
    ws.cell(row=sj_opex, column=15).value = _sum_o(opex_rows)
    ws.cell(row=sj_ebitdar, column=15).value = f"=O{sj_rev}-O{sj_opex}"
    written += 3
    return written, warnings


def _write_column_stride(
    ws, col_letter: str, start_row: int, values: list[Any],
    max_rows: int = 600,
) -> int:
    """Write a column of values starting at (col_letter)(start_row).

    Returns the number of non-blank cells written. Blank source rows are
    skipped (not overwritten with None — preserves any cells the template
    might have pre-styled or pre-populated).
    """
    written = 0
    for i, v in enumerate(values):
        if i >= max_rows:
            break
        if _is_blank(v):
            continue
        ws[f"{col_letter}{start_row + i}"] = v
        written += 1
    return written


# ──────────────────────────────────────────────────────────────────────────────
# Main entry
# ──────────────────────────────────────────────────────────────────────────────

_REGISTRY_CACHE: dict = {}


def _load_registry(path: str | Path | None) -> dict:
    p = Path(path) if path else DEFAULT_REGISTRY
    if not p.exists():
        raise UWTemplateWriterError(f"Registry not found: {p}")
    # Cache keyed on (path, mtime): the registry is constant across populate
    # calls within a session, but an edited registry.json (dev workflow)
    # still reloads without a process restart.
    key = (str(p.resolve()), p.stat().st_mtime_ns)
    cached = _REGISTRY_CACHE.get(key)
    if cached is None:
        with p.open(encoding="utf-8") as f:
            cached = json.load(f)
        _REGISTRY_CACHE.clear()  # keep at most one registry resident
        _REGISTRY_CACHE[key] = cached
    return cached


# ──────────────────────────────────────────────────────────────────────────────
# Dynamic-array repair (openpyxl quirk #6)
# ──────────────────────────────────────────────────────────────────────────────
#
# openpyxl's `wb.save()` silently drops `xl/metadata.xml` (the XLDAPR /
# `fDynamic="1"` block) and the per-cell `cm="N"` markers that tell Excel a
# formula is a *dynamic array* (SORT / UNIQUE / FILTER / ANCHORARRAY with
# spill) rather than a legacy CSE array. The formula TEXT survives verbatim,
# but without the metadata Excel reads `<f t="array" ref="Z173">=SORT(...)`
# as a single-cell CSE array → returns only the top-left value → Section R /
# Section S on the UW Template collapse to one row (silently wrong, no error).
#
# The committed blank template carries this metadata; the corruption happens
# at writer-save time. This restores it via direct zip/XML surgery (no lxml
# dependency — keeps the footprint flat). The writer never edits the
# dynamic-array anchor cells, so re-applying the original `cm` markers to the
# exact cells that carried them is faithful by construction.

_METADATA_PART = "xl/metadata.xml"
_METADATA_CT_OVERRIDE = (
    '<Override PartName="/xl/metadata.xml" '
    'ContentType="application/vnd.openxmlformats-officedocument.'
    'spreadsheetml.sheetMetadata+xml"/>'
)
_METADATA_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/"
    "relationships/sheetMetadata"
)
# A worksheet part, e.g. 'xl/worksheets/sheet8.xml'.
_WS_PART_RE = re.compile(r"^xl/worksheets/sheet\d+\.xml$")


def _attr(el: str, name: str) -> str | None:
    """Extract a single attribute value from an XML element string."""
    m = re.search(rf'\b{re.escape(name)}="([^"]*)"', el)
    return m.group(1) if m else None


def _sheet_part_to_name(zf: zipfile.ZipFile) -> dict[str, str]:
    """Map 'xl/worksheets/sheetN.xml' → sheet display name for a workbook zip.

    Robust to attribute ordering and `/`-prefixed targets.
    """
    try:
        wb = zf.read("xl/workbook.xml").decode("utf-8")
        rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    except KeyError:
        return {}
    rid_to_target: dict[str, str] = {}
    for rel in re.findall(r"<Relationship\b[^>]*/>", rels):
        rid, tgt = _attr(rel, "Id"), _attr(rel, "Target")
        if rid and tgt:
            rid_to_target[rid] = tgt
    out: dict[str, str] = {}
    for sh in re.findall(r"<sheet\b[^>]*/>", wb):
        name, rid = _attr(sh, "name"), _attr(sh, "r:id")
        if name and rid and rid in rid_to_target:
            # Targets come in two forms: relative ("worksheets/sheet8.xml",
            # original template) and package-absolute ("/xl/worksheets/
            # sheet1.xml", openpyxl output). Normalize both to "xl/...".
            tgt = rid_to_target[rid].lstrip("/")
            if not tgt.startswith("xl/"):
                tgt = "xl/" + tgt
            out[tgt] = name
    return out


def _collect_cm_cells(sheet_xml: str) -> dict[str, str]:
    """Return {cell_ref: cm_value} for every cell carrying a `cm=` marker."""
    out: dict[str, str] = {}
    for m in re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"[^>]*\bcm="(\d+)"', sheet_xml):
        out[m.group(1)] = m.group(2)
    return out


def _inject_cm(sheet_xml: str, ref_to_cm: dict[str, str]) -> tuple[str, int]:
    """Re-add `cm="N"` to the formula cells named in ref_to_cm.

    Only touches cells that (a) are named in ref_to_cm, (b) still have a
    formula (`<f`) in this output, and (c) don't already carry a `cm=`.
    Returns (patched_xml, count_injected).
    """
    n = 0
    for ref, cmv in ref_to_cm.items():
        # Match the cell's opening tag up to '>' followed immediately by '<f'
        # (whitespace allowed). Captures existing attributes to preserve them.
        pat = re.compile(
            r'(<c\b\s+r="' + re.escape(ref) + r'")([^>]*)>(\s*<f\b)'
        )
        m = pat.search(sheet_xml)
        if not m:
            continue
        attrs = m.group(2)
        if "cm=" in attrs:
            continue  # already marked — idempotent
        replacement = f'{m.group(1)}{attrs} cm="{cmv}">{m.group(3)}'
        sheet_xml = sheet_xml[: m.start()] + replacement + sheet_xml[m.end():]
        n += 1
    return sheet_xml, n


def _restore_dynamic_arrays(output_bytes: bytes, template_bytes: bytes) -> bytes:
    """Restore dynamic-array semantics openpyxl dropped on save.

    Re-injects `xl/metadata.xml` from the original template, wires its
    content-type Override + workbook relationship, and re-applies the per-cell
    `cm` markers to the dynamic-array anchor cells. No-op (returns the input
    unchanged) when the template has no `xl/metadata.xml` — e.g. v4 templates
    or any workbook without dynamic arrays.
    """
    with zipfile.ZipFile(io.BytesIO(template_bytes)) as ztpl:
        if _METADATA_PART not in ztpl.namelist():
            return output_bytes  # nothing dynamic to restore
        metadata_xml = ztpl.read(_METADATA_PART)
        tpl_part_name = _sheet_part_to_name(ztpl)
        # {sheet_name: {ref: cm}} from the original template
        cm_by_sheet: dict[str, dict[str, str]] = {}
        for part, name in tpl_part_name.items():
            try:
                xml = ztpl.read(part).decode("utf-8")
            except KeyError:
                continue
            cells = _collect_cm_cells(xml)
            if cells:
                cm_by_sheet[name] = cells

    if not cm_by_sheet and not metadata_xml:
        return output_bytes

    with zipfile.ZipFile(io.BytesIO(output_bytes)) as zout:
        out_names = zout.namelist()
        out_part_name = _sheet_part_to_name(zout)
        parts: dict[str, bytes] = {n: zout.read(n) for n in out_names}

    # 1) Re-apply cm markers per sheet (matched by sheet name).
    for part, name in out_part_name.items():
        ref_to_cm = cm_by_sheet.get(name)
        if not ref_to_cm or part not in parts:
            continue
        xml = parts[part].decode("utf-8")
        xml, _n = _inject_cm(xml, ref_to_cm)
        parts[part] = xml.encode("utf-8")

    # 2) Add metadata.xml.
    parts[_METADATA_PART] = metadata_xml

    # 3) Content-type Override for metadata.xml.
    ct = parts.get("[Content_Types].xml", b"").decode("utf-8")
    if ct and "/xl/metadata.xml" not in ct:
        ct = ct.replace("</Types>", _METADATA_CT_OVERRIDE + "</Types>")
        parts["[Content_Types].xml"] = ct.encode("utf-8")

    # 4) Workbook relationship → metadata.xml (unique rId).
    rels_part = "xl/_rels/workbook.xml.rels"
    rels = parts.get(rels_part, b"").decode("utf-8")
    if rels and "sheetMetadata" not in rels:
        used = [int(x) for x in re.findall(r'Id="rId(\d+)"', rels)]
        next_id = (max(used) + 1) if used else 1
        rel = (
            f'<Relationship Id="rId{next_id}" Type="{_METADATA_REL_TYPE}" '
            f'Target="metadata.xml"/>'
        )
        rels = rels.replace("</Relationships>", rel + "</Relationships>")
        parts[rels_part] = rels.encode("utf-8")

    # 5) Repackage.
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zw:
        for name, data in parts.items():
            zw.writestr(name, data)
    return buf.getvalue()


def populate_uw_template(
    analyzer_bytes: bytes,
    template_bytes: bytes,
    *,
    template_version: str = "v11",
    scenario: str = "normalized",
    registry_path: str | Path | None = None,
    include_statuses: frozenset[str] | None = None,
    allow_special_keys: frozenset[str] | None = None,
    computed_values: dict[str, Any] | None = None,
    computed_monthly: dict[str, list] | None = None,
    raw_t12_lines: list | None = None,
) -> tuple[bytes, PopulateReport]:
    """Populate the UW Template from a populated Analyzer workbook.

    Parameters
    ----------
    analyzer_bytes : bytes
        Raw bytes of a populated `ALF_Financial_Analyzer_Only.xlsx`. The
        Analyzer must have been saved through Excel at least once so the
        UW Output formula values are cached.
    template_bytes : bytes
        Raw bytes of the UW Template (e.g. `ALF_UW_Template_v4.xlsx`).
        Formula cells on the template side are preserved.
    template_version : str, default 'v6'
        Which version-keyed target to use from the registry. v6 (T-12 income
        restructure) is the binding default as of 2026-05-29; v5 and v4 remain
        supported (pass `template_version='v5'` / `'v4'`).
    scenario : str, default 'normalized'
        'normalized' (UW Output col F) or 't12_actual' (col E). Controls
        which annual column the t12-path writes.
    registry_path : str | Path | None
        Override the default registry location.
    include_statuses : frozenset[str] | None
        Statuses to INCLUDE. Defaults to `{'mapped', 'proposed'}`. Anything
        not in this set is skipped (along with the default skip set).
    allow_special_keys : frozenset[str] | None
        Concept keys to NOT special-skip even if they're in `_SPECIAL_SKIP_KEYS`.
        Default is None (honor all special-skips).
    computed_values : dict[str, Any] | None
        Optional `{concept_key: value}` fallback (e.g. from
        `uw_output_model.compute_uw_output_values`). Used **only when the
        Analyzer's cached cell read is blank** — this kills the "cache caveat"
        for Analyzers freshly built in-memory (openpyxl doesn't evaluate
        formulas, so `UW Output` cells have no cached values). An analyst-saved
        Analyzer with real cached values still wins; the fallback only fills
        gaps. Concepts filled this way are counted in
        `report.summary['computed_in_python']` and noted per-concept.
    computed_monthly : dict[str, list] | None
        Optional `{concept_key: [12 monthly values]}` (e.g. from
        `uw_output_model.compute_uw_output_monthly`). For any concept whose
        target is a `T-12 Analysis` column-N scalar (the annual T-12 Total),
        the 12 values are pasted across the Layer-3 monthly grid (cols B–M).
        Those grid cells are literal paste targets (not formulas), so monthly
        is written whenever available — no Analyzer-cached monthly to defer
        to. Counted in `report.summary['monthly_cells_written']`.

    Returns
    -------
    (populated_bytes, PopulateReport)
    """
    computed = computed_values or {}
    monthly = computed_monthly or {}
    monthly_cells = 0
    reg = _load_registry(registry_path)
    templates = reg.get("templates", {})
    if template_version not in templates:
        raise TemplateVersionMissing(
            f"Registry has no template version {template_version!r}. "
            f"Available: {sorted(templates.keys())}"
        )

    include = include_statuses or frozenset({"mapped", "proposed"})
    allow_special = allow_special_keys or frozenset()

    # ── Load workbooks ────────────────────────────────────────────────────────
    wb_analyzer = openpyxl.load_workbook(
        io.BytesIO(analyzer_bytes), data_only=True
    )
    wb_template = openpyxl.load_workbook(
        io.BytesIO(template_bytes), data_only=False
    )

    report = PopulateReport(template_version=template_version, scenario=scenario)

    # Pre-flight: confirm intake sheets exist on the template
    tv = templates[template_version]
    expected_intakes = tv.get("intake_sheets", [])
    for s in expected_intakes:
        if s not in wb_template.sheetnames:
            report.warnings.append(
                f"Template missing expected intake sheet: {s!r}"
            )

    # Per-template-version capacity for the rent_roll path. Defaults to a
    # high number when the registry doesn't pin it (legacy v4 fallback).
    rr_data_end_row = (
        tv.get("rent_roll_data_end_row")
        or tv.get("data_end_row")
        or 610
    )
    # Paste anchor comes from the registry templates block — v8 re-anchored
    # the grid to A214 (header 213; every RR Analysis aggregate reads
    # $214:$613), v4-v6 use A211. Fallback 211 when the block doesn't pin it.
    rr_paste_start = 211
    _anchor_m = re.search(r"!\$?[A-Z]+\$?(\d+)$", tv.get("rent_roll_paste_anchor") or "")
    if _anchor_m:
        rr_paste_start = int(_anchor_m.group(1))
    rr_header_row = tv.get("rent_roll_header_row") or (rr_paste_start - 1)
    rr_max_rows = rr_data_end_row - rr_paste_start + 1

    # Pre-flight: confirm the Rent Roll Analysis header row (version-specific)
    # if rent_roll concepts will be written — older / working-copy templates
    # may not have it, in which case the writer still writes (per the
    # registry's paste anchor) but flags a warning.
    rr_concepts = [c for c in reg.get("concepts", []) if c.get("path") == "rent_roll"]
    if rr_concepts and "Rent Roll Analysis" in wb_template.sheetnames:
        ra = wb_template["Rent Roll Analysis"]
        hdr = ra[f"A{rr_header_row}"].value
        if not hdr:
            report.warnings.append(
                f"Rent Roll Analysis!A{rr_header_row} is blank — this template "
                f"version expects a header row at {rr_header_row} with paste "
                f"anchor at row {rr_paste_start}. Writing to row "
                f"{rr_paste_start}+ anyway; the template may be a working copy "
                "that predates the registry's row layout."
            )

    # ── Iterate concepts ──────────────────────────────────────────────────────
    for concept in reg.get("concepts", []):
        key = concept.get("key", "")
        path = concept.get("path", "t12")
        status = concept.get("status", "")

        result = ConceptResult(
            key=key, path=path, status=status, outcome="skipped",
        )

        # Skip on status filter
        if status in _DEFAULT_SKIP_STATUSES:
            result.outcome = "skipped"
            result.notes = f"status={status} in default skip set"
            report.results.append(result)
            continue
        if status not in include:
            result.outcome = "skipped"
            result.notes = f"status={status} not in include set"
            report.results.append(result)
            continue

        # Skip on special-key list (e.g. opex_bad_debt_expense duplicate)
        if key in _SPECIAL_SKIP_KEYS and key not in allow_special:
            result.outcome = "skipped"
            result.notes = _SPECIAL_SKIP_KEYS[key]
            report.results.append(result)
            continue

        # Resolve target for this template version
        tgt = (concept.get("targets") or {}).get(template_version)
        if not tgt:
            result.outcome = "no_target"
            result.notes = f"no target for template version {template_version!r}"
            report.results.append(result)
            continue

        target_sheet = tgt.get("sheet")
        target_addr = tgt.get("address", "")
        result.target_address = f"{target_sheet}!{target_addr}"

        if target_sheet not in wb_template.sheetnames:
            result.outcome = "error"
            result.notes = f"target sheet {target_sheet!r} not in template"
            report.results.append(result)
            continue

        ws = wb_template[target_sheet]

        # T-12 Analysis Layer-3 TOTAL rows are live template formulas, handled
        # by `_finalize_t12_layer3` after this loop. Skip them here so the
        # generic writer never pastes a value over the formula.
        if key in _T12_TOTAL_CONCEPTS and target_sheet == "T-12 Analysis":
            result.outcome = "skipped"
            result.notes = "total row — col-N formula preserved (finalize pass)"
            report.results.append(result)
            continue

        # Resolve source
        try:
            value = _resolve_source(wb_analyzer, concept, scenario)
        except Exception as e:  # defensive — never crash on a single concept
            result.outcome = "error"
            result.notes = f"source resolve failed: {e}"
            report.results.append(result)
            continue

        # In-Python computed fallback — only when the Analyzer cell came back
        # blank (cache caveat: freshly-built Analyzer has no cached formula
        # values). A saved-through-Excel Analyzer keeps its cached value.
        used_computed_fallback = False
        if key in computed and not isinstance(value, list) and _is_blank(value):
            cv = computed[key]
            if not _is_blank(cv):
                value = cv
                used_computed_fallback = True

        # Income-waterfall sign convention: negate the contra lines so the
        # template's additive Net Rent formula subtracts them.
        if (
            key in _T12_CONTRA_KEYS
            and target_sheet == "T-12 Analysis"
            and isinstance(value, (int, float))
        ):
            value = -value

        # Decide scalar vs row-stride
        try:
            col_letter, start_row, is_stride = _parse_target_addr(target_addr)
        except ValueError as e:
            result.outcome = "error"
            result.notes = f"unparseable target: {e}"
            report.results.append(result)
            continue

        if isinstance(value, list):
            # Rent-roll / column-stride source
            if not is_stride:
                result.outcome = "error"
                result.notes = (
                    "source returned a list but target is not a row-stride "
                    f"(target={target_addr!r}); registry inconsistency"
                )
                report.results.append(result)
                continue
            # Cap stride at the template version's data_end_row. Truncate
            # source if it would overflow the template's data block; surface
            # overflow as a warning so the analyst can see a deal-too-big
            # case clearly.
            effective_max = rr_max_rows
            populated_count = sum(1 for v in value if not _is_blank(v))
            if populated_count > effective_max:
                report.warnings.append(
                    f"{concept.get('key')!r}: source has {populated_count} "
                    f"populated rows but template {template_version!r} only "
                    f"holds {effective_max} ({rr_paste_start}..{rr_data_end_row}). "
                    f"Truncating."
                )
            written = _write_column_stride(
                ws, col_letter, start_row, value, max_rows=effective_max
            )
            result.outcome = "written" if written > 0 else "no_source"
            result.cells_written = written
            # sample value: first non-None
            sample = next((v for v in value if not _is_blank(v)), None)
            result.sample_value = sample
            if written == 0:
                result.notes = "source column is empty in Analyzer"
        else:
            # Scalar
            if _is_blank(value):
                result.outcome = "no_source"
                result.notes = "source resolved to blank/None"
                report.results.append(result)
                continue
            if is_stride:
                # Treat as scalar at start_row even though address ends in '+'
                target_addr_eff = f"{col_letter}{start_row}"
            else:
                target_addr_eff = f"{col_letter}{start_row}"
            _write_scalar(ws, target_addr_eff, value)
            result.outcome = "written"
            result.cells_written = 1
            result.sample_value = value

            # Monthly grid (UW Template Layer-3): paste the 12 monthly values
            # across cols B–M for T-12 Analysis column-N (annual) targets.
            if (
                key in monthly
                and target_sheet == "T-12 Analysis"
                and col_letter == "N"
            ):
                mvals = monthly[key]
                if isinstance(mvals, (list, tuple)) and mvals:
                    mlist = list(mvals)
                    # Same contra-sign convention as the annual value.
                    if key in _T12_CONTRA_KEYS:
                        mlist = [(-m if isinstance(m, (int, float)) else m) for m in mlist]
                    monthly_cells += _write_monthly_grid(ws, start_row, mlist)

            if used_computed_fallback:
                result.computed_fallback = True
                result.notes = (
                    "value computed in-Python (Analyzer cell was blank — "
                    "cache caveat fallback)"
                )

        report.results.append(result)

    # ── Finalize T-12 Analysis Layer-3 totals (v5 / v6) ────────────────────────
    # Make the total rows live formulas (col N preserved/authored + mirrored
    # across the monthly grid B–M), so they recompute from the line items and
    # tie month-to-annual. Runs after the line items are written. Row positions
    # come from `_T12_LAYOUT` per template version (v6 rebuilt the income rows).
    t12_finalized = 0
    section_i_cells = 0
    t12_layout = _T12_LAYOUT.get(template_version)
    if t12_layout and "T-12 Analysis" in wb_template.sheetnames:
        ws_t12 = wb_template["T-12 Analysis"]
        try:
            t12_finalized = _finalize_t12_layer3(ws_t12, monthly, t12_layout)
        except Exception as e:  # never fail the populate over the finalize pass
            report.warnings.append(f"T-12 Analysis total-formula finalize skipped ({e}).")
        # Section I (Layer 1 — Raw T-12): rebuild from the summarized raw lines.
        if raw_t12_lines:
            try:
                section_i_cells, si_warnings = _write_section_i_raw(
                    ws_t12, raw_t12_lines, t12_layout
                )
                report.warnings.extend(si_warnings)
            except Exception as e:
                report.warnings.append(f"Section I (raw T-12) population skipped ({e}).")

    # ── Summary ───────────────────────────────────────────────────────────────
    by = report.by_outcome()
    report.summary = {k: len(v) for k, v in by.items()}
    report.summary["total_concepts"] = len(report.results)
    report.summary["cells_written"] = report.write_count()
    report.summary["computed_in_python"] = sum(
        1 for r in report.results if r.computed_fallback
    )
    report.summary["monthly_cells_written"] = monthly_cells
    report.summary["t12_totals_finalized"] = t12_finalized
    report.summary["section_i_raw_cells"] = section_i_cells

    # ── Serialize ─────────────────────────────────────────────────────────────
    out = io.BytesIO()
    wb_template.save(out)
    output_bytes = out.getvalue()

    # ── Restore dynamic-array metadata openpyxl dropped on save ────────────────
    # (openpyxl quirk #6). Without this, the template's Section R / S
    # SORT/UNIQUE/FILTER spills demote to single-cell CSE arrays in Excel and
    # silently collapse to one row. Faithful: re-applies the original `cm`
    # markers to the exact anchor cells the writer never edits.
    try:
        repaired = _restore_dynamic_arrays(output_bytes, template_bytes)
        output_bytes = repaired
        report.summary["dynamic_arrays_restored"] = 1
    except Exception as e:  # never fail the populate over a metadata repair
        report.warnings.append(
            f"Dynamic-array metadata repair skipped ({e}); Section R/S spills "
            f"may need a manual re-entry in Excel."
        )
        report.summary["dynamic_arrays_restored"] = 0

    return output_bytes, report


# ──────────────────────────────────────────────────────────────────────────────
# CLI for ad-hoc runs
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    import sys

    ap = argparse.ArgumentParser(description="Populate the ALF UW Template.")
    ap.add_argument("analyzer", help="path to populated Analyzer .xlsx")
    ap.add_argument("template", help="path to UW Template .xlsx")
    ap.add_argument("output", help="output path for populated template")
    ap.add_argument(
        "--scenario", default="normalized",
        choices=("normalized", "t12_actual"),
    )
    ap.add_argument("--template-version", default="v11")
    ap.add_argument("--registry", default=None)
    args = ap.parse_args()

    analyzer_bytes = Path(args.analyzer).read_bytes()
    template_bytes = Path(args.template).read_bytes()

    populated, report = populate_uw_template(
        analyzer_bytes, template_bytes,
        template_version=args.template_version,
        scenario=args.scenario,
        registry_path=args.registry,
    )

    Path(args.output).write_bytes(populated)

    print(f"\nPopulated → {args.output}")
    print(f"\n=== Summary ===")
    for k, v in sorted(report.summary.items()):
        print(f"  {k:25s}  {v}")
    if report.warnings:
        print(f"\n=== Warnings ({len(report.warnings)}) ===")
        for w in report.warnings:
            print(f"  ! {w}")

    written = [r for r in report.results if r.outcome == "written"]
    if written:
        print(f"\n=== First 10 written cells ===")
        for r in written[:10]:
            print(
                f"  {r.target_address:35s}  ← {r.key:30s}  "
                f"({r.cells_written} cell{'s' if r.cells_written != 1 else ''})  "
                f"sample={r.sample_value!r}"
            )

    sys.exit(0)
