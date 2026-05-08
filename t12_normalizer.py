"""
T12 Normalizer — Track 2 parser
================================

Reads a raw T12 export from a senior-housing operator and returns clean GL
detail rows + detected month labels + the list of descriptions not present in
the destination workbook's `Description_Map`.

Format-registry pattern: each supported format is a class with `detect(wb)` and
`extract(wb)` methods. v0.2.0 ships three formats: `YardiIncomeToBudgetFormat`,
`MriR12mincsFormat`, and `BrokerFinancialSummaryFormat`. Adding a format is a
one-class-plus-one-list-entry change.

Drop-rules applied in order during extraction:
  1. Drop rows with no $ value (all 12 months and T12 total are zero/empty).
  2. Drop rows whose description matches a grand-total pattern
     (`TOTAL `, `NET `, `EBITDA`, `EBITDAR`, `EBITDARM`, `Subtotal,`,
     exact `NET INCOME`, `NET OPERATING INCOME`).
  3. Drop rows whose description is in the explicit drop-list
     (initially: `Other Non Operating Revenue & Expense`).

UNMATCHED detection runs after drop-rules. A description is UNMATCHED if it
doesn't appear in the destination workbook's `Description_Map` (col A, rows
5+). The destination's col P formula is the runtime source of truth — this
list is for surfacing them in the UI.

Cluster B (v0.2.0):
  - `_check_sign_convention()` produces defensive sign warnings on
    known-negative descriptions (concession / vacancy / L2L / bad debt).
  - `_count_populated_months()` returns how many of the 12 columns have any
    non-zero data; values <12 indicate a partial-year T12.
  - `parse_t12(..., annualize_partial_year=True)` multiplies monthly values by
    12/N where N is populated months. Disabled by default.

See SPEC-T12.md §"Parser data flow" for the full design rationale.
"""

from __future__ import annotations

import datetime as dt
import io
import re
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from typing import List, Optional, Set, Tuple

import openpyxl


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class UnknownT12FormatError(Exception):
    """No registered format detector matched the uploaded workbook."""


# T12NormalizerCapacityError lives in t12_normalizer_writer (capacity is a
# writer concern, not a parser concern), but re-exported here for convenience.


# ---------------------------------------------------------------------------
# Drop-rule constants
# ---------------------------------------------------------------------------

# Rule 2: grand-total patterns. All comparisons are case-insensitive.
GRAND_TOTAL_PREFIXES: Tuple[str, ...] = (
    "TOTAL ",
    "TOTAL-",       # rare formatting variant ("Total-Operating Expense")
    "TOTAL - ",     # Yardi style: "Total - Rent Revenue"
    "NET ",
    # Broker convention: "Subtotal, Room & Board", "Subtotal, Care Level".
    # Added in v0.2.0 alongside BrokerFinancialSummaryFormat.
    "SUBTOTAL,",
    "SUBTOTAL ",
)

GRAND_TOTAL_KEYWORDS: Tuple[str, ...] = (
    "EBITDA",       # also catches EBITDAR, EBITDARM
)

GRAND_TOTAL_EXACT: Tuple[str, ...] = (
    "NET INCOME",
    "NET OPERATING INCOME",
)

# Rule 3: explicit drop-list. Compared case-insensitive after TRIM.
EXPLICIT_DROP_LIST: Tuple[str, ...] = (
    "Other Non Operating Revenue & Expense",
    # Yardi reports a "Non-Operating Expenses" subtotal banner that repeats
    # the section-header text and sums the GL rows immediately above it.
    # Without this entry the parser would double-count those rows once they
    # appear individually as GL detail and again as the banner subtotal.
    # Added in v0.1.1 (Salem Management Fees fix). See CHANGELOG-T12 [0.1.1].
    "Non-Operating Expenses",
    # Broker convention: a single-row restatement of the broker's published
    # NOI ("NOI on Statement"). Carries the same dollar value as the
    # synthesized "Total Net Operating Income" row; we drop both. Added in
    # v0.2.0 alongside BrokerFinancialSummaryFormat (Homestead Pensacola).
    "NOI on Statement",
    "Check",  # broker reconciliation row (sum=0 anyway; defensive)
)


def _matches_grand_total(desc: str) -> bool:
    """Return True if `desc` matches any grand-total pattern (Rule 2)."""
    if not desc:
        return False
    upper = desc.upper().strip()
    if upper in (s.upper() for s in GRAND_TOTAL_EXACT):
        return True
    for prefix in GRAND_TOTAL_PREFIXES:
        if upper.startswith(prefix):
            return True
    for kw in GRAND_TOTAL_KEYWORDS:
        if kw in upper:
            return True
    return False


def _matches_drop_list(desc: str) -> bool:
    """Return True if `desc` matches any entry in the explicit drop-list (Rule 3)."""
    if not desc:
        return False
    norm = desc.strip().casefold()
    for entry in EXPLICIT_DROP_LIST:
        if norm == entry.strip().casefold():
            return True
    return False


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------

@dataclass
class GLRow:
    """One cleaned GL detail row from a raw T12."""

    account: str          # may be "" for MRI; trimmed if present
    description: str      # trimmed
    monthly: List[float]  # exactly 12 values, in chronological order matching month_labels
    total: float          # T12 total (may differ from sum(monthly) by rounding)

    def __post_init__(self) -> None:
        if len(self.monthly) != 12:
            raise ValueError(
                f"GLRow expects 12 monthly values, got {len(self.monthly)} "
                f"(description={self.description!r})"
            )


@dataclass
class T12ParseResult:
    gl_rows: List[GLRow]
    month_labels: List[str]            # 12 entries, normalized "MMM YYYY" (or "" for padded slots in partial-year)
    unmatched: List[str]               # descriptions not in Description_Map
    format_name: str                   # human-readable, e.g. "Yardi (Income to Budget)"
    sheet_name: str                    # which sheet the parser pulled from
    # Cluster B fields (v0.2.0)
    sign_warnings: List[str] = field(default_factory=list)
    populated_months: int = 0          # months with at least one non-zero GL value, 0..12
    was_annualized: bool = False       # True if parser scaled values by 12/populated_months


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_float_or_zero(v) -> float:
    """Coerce a cell value to float, returning 0.0 for None / empty / non-numeric."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    # Strip currency / commas / parens (Excel sometimes presents accounting strings)
    cleaned = s.replace(",", "").replace("$", "")
    neg = False
    if cleaned.startswith("(") and cleaned.endswith(")"):
        neg = True
        cleaned = cleaned[1:-1]
    try:
        f = float(cleaned)
        return -f if neg else f
    except ValueError:
        return 0.0


def _has_any_dollar_value(monthly: List[float], total: float) -> bool:
    """Drop-rule 1: a row passes if at least one of the 13 numbers is non-zero."""
    if any(v != 0 for v in monthly):
        return True
    return total != 0


def _normalize_yardi_date_string(s: str) -> str:
    """Yardi row 9 ships strings like '02/28/2025' → 'Feb 2025'.

    Falls back to returning the trimmed input if parsing fails.
    """
    if s is None:
        return ""
    s = str(s).strip()
    if not s:
        return ""
    if isinstance(s, dt.datetime):
        return s.strftime("%b %Y")
    # Most common: "MM/DD/YYYY"
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%b %Y")
        except ValueError:
            continue
    return s


def _normalize_mri_date_string(s) -> str:
    """MRI row 11 ships strings like '01/25' → 'Jan 2025'."""
    if s is None:
        return ""
    if isinstance(s, dt.datetime):
        return s.strftime("%b %Y")
    s = str(s).strip()
    if not s:
        return ""
    for fmt in ("%m/%y", "%m/%Y", "%b %Y", "%B %Y"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%b %Y")
        except ValueError:
            continue
    return s


# ---------------------------------------------------------------------------
# Format registry — abstract base
# ---------------------------------------------------------------------------

class T12Format(ABC):
    """One supported T12 export format."""

    name: str

    @abstractmethod
    def detect(self, wb: openpyxl.Workbook) -> Optional[str]:
        """Return the matching sheet name if this format applies, else None.

        Returning the sheet name (rather than just True) avoids re-doing the
        detection work inside extract() and is robust to multi-sheet workbooks
        where only one sheet is the actual T12.
        """

    @abstractmethod
    def extract(
        self, wb: openpyxl.Workbook, sheet_name: str
    ) -> Tuple[List[GLRow], List[str]]:
        """Walk the matching sheet and return (gl_rows, month_labels).

        Drop-rules are applied here. The returned list is post-filter.
        """


# ---------------------------------------------------------------------------
# Format: Yardi "Income to Budget"
# ---------------------------------------------------------------------------

class YardiIncomeToBudgetFormat(T12Format):
    """Yardi 'Income to Budget' — Salem (Oaks at Salem Road) is the reference.

    Layout:
      - Single visible sheet (sheet name varies per property: "Income to Budget",
        sometimes prefixed/suffixed). Detection by content, not name.
      - Row 8: "Month Ending" banner across cols C-N.
      - Row 9: 12 month-end date strings (e.g., "02/28/2025") in cols C-N.
      - Row 10: "Account number" / "Actual" header banner.
      - Row 11+: indented hierarchy. GL detail rows have a numeric account
        number in col A, description in col B, 12 monthly amounts in cols C-N,
        T12 total in col O. Section headers and subtotals have non-numeric
        col A.
    """

    name = "Yardi (Income to Budget)"

    BODY_START_ROW = 11
    LABEL_ROW = 9
    LABEL_COL_START = 3   # C
    LABEL_COL_END = 14    # N
    DATA_COL_START = 3    # C (first month)
    DATA_COL_END = 14     # N (last month)
    TOTAL_COL = 15        # O

    def _scan_sheet_for_yardi_signal(self, ws) -> bool:
        """Heuristic: at least 3 rows in the body have a numeric account # in col A."""
        hits = 0
        last = min(ws.max_row, self.BODY_START_ROW + 80)
        for r in range(self.BODY_START_ROW, last + 1):
            v = ws.cell(r, 1).value
            if v is None:
                continue
            s = str(v).strip()
            if s.isdigit():
                hits += 1
                if hits >= 3:
                    return True
        return False

    def detect(self, wb: openpyxl.Workbook) -> Optional[str]:
        # Fast path: standard Yardi sheet name
        for s in wb.sheetnames:
            if s.strip().lower() == "income to budget":
                ws = wb[s]
                if self._scan_sheet_for_yardi_signal(ws):
                    return s
        # Slow path: scan all sheets for the numeric-account-# signal
        for s in wb.sheetnames:
            ws = wb[s]
            # Skip sheets that are clearly MRI (handled by MriR12mincsFormat)
            if s.strip().upper() == "MRI_R12MINCS":
                continue
            if self._scan_sheet_for_yardi_signal(ws):
                return s
        return None

    def extract(
        self, wb: openpyxl.Workbook, sheet_name: str
    ) -> Tuple[List[GLRow], List[str]]:
        ws = wb[sheet_name]

        # --- Month labels from row 9 ---
        labels: List[str] = []
        for c in range(self.LABEL_COL_START, self.LABEL_COL_END + 1):
            labels.append(_normalize_yardi_date_string(ws.cell(self.LABEL_ROW, c).value))

        # --- GL detail body ---
        # NOTE: We do NOT filter on numeric account # here. Yardi sometimes
        # reports legitimate single-line expenses (e.g., Salem's "Management
        # Fees" line at row 128) as section-banner-style rows with no account
        # number. The three drop-rules below are sufficient to filter section
        # headers (rule 1: no $), TOTAL/NET/EBITDA subtotals (rule 2), and
        # the banner-subtotal pattern (rule 3 explicit drop-list). v0.1.0
        # had this filter and dropped Salem's $131,579.65 management fee
        # silently — fixed in v0.1.1.
        rows: List[GLRow] = []
        for r in range(self.BODY_START_ROW, ws.max_row + 1):
            a_raw = ws.cell(r, 1).value
            b_raw = ws.cell(r, 2).value

            if b_raw is None:
                continue
            desc = str(b_raw).strip()
            if not desc:
                continue
            # Yardi sometimes ships literal "None" in spacer rows
            if desc.lower() == "none":
                continue

            monthly = [
                _to_float_or_zero(ws.cell(r, c).value)
                for c in range(self.DATA_COL_START, self.DATA_COL_END + 1)
            ]
            total = _to_float_or_zero(ws.cell(r, self.TOTAL_COL).value)

            # Rule 1: drop if no $ value
            if not _has_any_dollar_value(monthly, total):
                continue
            # Rule 2: grand-total pattern
            if _matches_grand_total(desc):
                continue
            # Rule 3: explicit drop-list
            if _matches_drop_list(desc):
                continue

            # Account # is preserved when present (Yardi reports it for
            # standard GL rows) and stored as "" when absent (banner-style
            # expense rows that survived the drop-rules).
            account = ""
            if a_raw is not None:
                a_str = str(a_raw).strip()
                if a_str.isdigit():
                    account = a_str

            rows.append(GLRow(
                account=account,
                description=desc,
                monthly=monthly,
                total=total,
            ))

        return rows, labels


# ---------------------------------------------------------------------------
# Format: MRI "R12MINCS"
# ---------------------------------------------------------------------------

class MriR12mincsFormat(T12Format):
    """MRI 'R12MINCS' — Briar Glen is the reference.

    Layout:
      - Sheet named exactly `MRI_R12MINCS`.
      - Row 11: 12 period codes in cols B-M (e.g., "01/25" → Jan 2025).
      - Row 14+: flat structure. Description in col A (no account #), 12 monthly
        amounts in cols B-M, T12 total in col N. Embedded subtotals
        ("Total Effective Gross Rents", "TOTAL OPERATING REVENUE", etc.) are
        filtered by drop-rule 2.
    """

    name = "MRI R12MINCS"
    SHEET_NAME = "MRI_R12MINCS"

    BODY_START_ROW = 14
    LABEL_ROW = 11
    LABEL_COL_START = 2   # B
    LABEL_COL_END = 13    # M
    DATA_COL_START = 2    # B
    DATA_COL_END = 13     # M
    TOTAL_COL = 14        # N

    def detect(self, wb: openpyxl.Workbook) -> Optional[str]:
        for s in wb.sheetnames:
            if s.strip().upper() == self.SHEET_NAME:
                return s
        return None

    def extract(
        self, wb: openpyxl.Workbook, sheet_name: str
    ) -> Tuple[List[GLRow], List[str]]:
        ws = wb[sheet_name]

        # --- Month labels from row 11 ---
        labels: List[str] = []
        for c in range(self.LABEL_COL_START, self.LABEL_COL_END + 1):
            labels.append(_normalize_mri_date_string(ws.cell(self.LABEL_ROW, c).value))

        # --- GL detail body ---
        rows: List[GLRow] = []
        for r in range(self.BODY_START_ROW, ws.max_row + 1):
            a_raw = ws.cell(r, 1).value
            if a_raw is None:
                continue
            desc = str(a_raw).strip()
            if not desc:
                continue
            # Briar Glen ships some literal "None" strings as section dividers
            # — these have no $ value so they're caught by Rule 1, but we also
            # short-circuit obvious section markers here as a small efficiency.
            if desc.lower() == "none":
                continue

            monthly = [
                _to_float_or_zero(ws.cell(r, c).value)
                for c in range(self.DATA_COL_START, self.DATA_COL_END + 1)
            ]
            total = _to_float_or_zero(ws.cell(r, self.TOTAL_COL).value)

            # Rule 1: drop if no $ value
            if not _has_any_dollar_value(monthly, total):
                continue
            # Rule 2: grand-total pattern
            if _matches_grand_total(desc):
                continue
            # Rule 3: explicit drop-list
            if _matches_drop_list(desc):
                continue

            rows.append(GLRow(
                account="",
                description=desc,
                monthly=monthly,
                total=total,
            ))

        return rows, labels


# ---------------------------------------------------------------------------
# Format: Broker "Financial Summary" (Historical Performance)
# ---------------------------------------------------------------------------

class BrokerFinancialSummaryFormat(T12Format):
    """Broker-prepared T12 summaries with `Historical Performance` header at A4.

    Reference files:
      - 2026-03 Homestead Village Pensacola Financial Summary.xlsx (multi-section
        dashboard; T12 block is one of several time-window views)
      - March 2026 T12.xlsx (single-sheet 12-month T12 only)

    Layout:
      - Variable sheet name. Detection by content: A4 contains "Historical
        Perform[ance]". When multiple sheets match, parser picks the first.
      - Row 4 carries datetime cells across one or more time-window sections
        (T12 monthly, T6M monthly, T2M monthly, plus annual totals like CY).
      - Parser locates the rightmost contiguous run of monotonic monthly
        datetimes (up to 12 cells) and uses those columns as the data block.
        For Homestead Summary that picks cols AB:AM (Apr 2025 → Mar 2026, the
        T12 block), even though the file has 39 datetime cells in row 4 total.
      - Row 5 onward is body. A "banner" row has text in col A and **all 12
        monthly columns are None** (truly blank, not zero). A row with col A
        text and zero values is a GL row that happens to be zero — NOT a
        banner. Banners track the most-recent section name.
      - Pre-financial preamble (Census / ADC / Room Rates summary banners)
        sits before the first banner that matches `Revenue` / `Revenues`.
        Parser drops everything before that banner.
      - Descriptions are prefixed with `<most-recent-banner> | <description>`
        to disambiguate descriptions like `Payroll - Wages` that appear in 8
        different department sections.
      - "Totals" column to the right of the 12 monthly cols is optional and
        ignored — parser computes its own T12 total from the 12 monthly values.
      - Sign convention is standard (revenue +, expense +, concession -). No
        per-format override.
    """

    name = "Broker Financial Summary"

    LABEL_ROW = 4
    BODY_START_ROW = 5
    FINGERPRINT_RE = re.compile(r"historical\s+perform", re.IGNORECASE)
    PREAMBLE_TRIGGER_RE = re.compile(r"\brevenue[s]?\b", re.IGNORECASE)
    # Banners that signal end of the operating P&L. Extraction stops here.
    # `Non-Operating` covers below-NOI items (depreciation, ground lease,
    # non-op income/expense). `Wages Analysis` and `Payroll Summary` are
    # broker-only analytical sections (decomposed wage % ratios) that aren't
    # GL lines and would pollute the descmap if extracted.
    TERMINATE_BANNER_RE = re.compile(
        r"non[-\s]?operating|wages\s+analysis|payroll\s+summary",
        re.IGNORECASE,
    )
    SUBTOTAL_RE = re.compile(r"^\s*subtotal[\s,]", re.IGNORECASE)

    def _matches_fingerprint(self, ws) -> bool:
        v = ws.cell(self.LABEL_ROW, 1).value
        if v is None:
            return False
        return bool(self.FINGERPRINT_RE.search(str(v)))

    def detect(self, wb: openpyxl.Workbook) -> Optional[str]:
        # Per spec: when multiple sheets match, pick the first.
        for s in wb.sheetnames:
            ws = wb[s]
            if self._matches_fingerprint(ws):
                return s
        return None

    def _find_rightmost_monthly_run(self, ws) -> List[Tuple[int, dt.datetime]]:
        """Walk row 4, return rightmost contiguous run of monotonic-by-month
        datetime cells (length up to 12). For partial-year files the run can
        be shorter; caller pads with empty months on the left."""
        cells: List[Tuple[int, dt.datetime]] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(self.LABEL_ROW, c).value
            if isinstance(v, dt.datetime):
                cells.append((c, v))
        if not cells:
            return []
        # Build runs walking right-to-left. Each run is a stretch where each
        # cell is exactly 1 column AND 1 calendar month after the previous.
        runs: List[List[Tuple[int, dt.datetime]]] = []
        cur: List[Tuple[int, dt.datetime]] = [cells[-1]]
        for i in range(len(cells) - 2, -1, -1):
            prev_col, prev_date = cells[i]
            next_col, next_date = cur[0]
            delta_col = next_col - prev_col
            delta_months = (next_date.year - prev_date.year) * 12 + (
                next_date.month - prev_date.month
            )
            if delta_col == 1 and delta_months == 1:
                cur.insert(0, (prev_col, prev_date))
            else:
                runs.append(cur)
                cur = [(prev_col, prev_date)]
        runs.append(cur)
        # Prefer rightmost (first in list) run with >= 12 months; else rightmost run.
        for run in runs:
            if len(run) >= 12:
                return run[-12:]
        return runs[0] if runs else []

    def extract(
        self, wb: openpyxl.Workbook, sheet_name: str
    ) -> Tuple[List[GLRow], List[str]]:
        ws = wb[sheet_name]

        run = self._find_rightmost_monthly_run(ws)
        if not run:
            raise UnknownT12FormatError(
                f"BrokerFinancialSummaryFormat matched sheet {sheet_name!r} "
                f"on the A4 fingerprint, but found no datetime cells in row "
                f"{self.LABEL_ROW}. Expected 1-12 monthly date headers."
            )

        month_cols = [c for c, _ in run]
        labels = [d.strftime("%b %Y") for _, d in run]
        n_detected = len(run)

        # Pad LEFT with empty labels if partial-year. GLRow contract requires
        # exactly 12 monthly values per row; partial-year detection happens
        # downstream via _count_populated_months.
        if n_detected < 12:
            pad = 12 - n_detected
            labels = [""] * pad + labels

        rows: List[GLRow] = []
        # Banner stack:
        #   top_banner = the preamble-trigger banner (e.g. "Revenues") set
        #     once when extraction starts. Acts as the "no-prefix" sentinel —
        #     GL rows with current_banner == top_banner are kept unprefixed
        #     (mirrors the populated_analyzer convention: top-level revenue
        #     siblings like "Concessions" or "Move-In Fees" carry no prefix).
        #   current_banner = most-recent banner seen. Mutated by sub-section
        #     banners and popped back to top_banner on Subtotal, rows.
        top_banner: Optional[str] = None
        current_banner: Optional[str] = None

        for r in range(self.BODY_START_ROW, ws.max_row + 1):
            a_raw = ws.cell(r, 1).value
            if a_raw is None:
                continue
            desc = str(a_raw).strip()
            if not desc:
                continue
            # Skip the row-4 fingerprint string in case max_row scan picks it up
            if self.FINGERPRINT_RE.search(desc):
                continue

            monthly_raw = [ws.cell(r, c).value for c in month_cols]
            all_none = all(v is None for v in monthly_raw)

            # Banner: text in col A, all 12 monthly cells truly None (not 0).
            if all_none:
                # Post-P&L cutoff: stop on banners that signal end of operating P&L.
                if top_banner is not None and self.TERMINATE_BANNER_RE.search(desc):
                    break
                # First banner matching the preamble trigger (Revenue/Revenues)
                # establishes the top-level. Earlier banners (Census / ADC /
                # Room Rates) are silently dropped along with their GL rows.
                if top_banner is None:
                    if self.PREAMBLE_TRIGGER_RE.search(desc):
                        top_banner = desc
                        current_banner = desc
                    # else: still in preamble; ignore.
                else:
                    current_banner = desc
                continue

            # GL row from here on — only kept if we've crossed the Revenue preamble.
            if top_banner is None:
                continue

            monthly = [_to_float_or_zero(v) for v in monthly_raw]
            if len(monthly) < 12:
                monthly = [0.0] * (12 - len(monthly)) + monthly

            # Parser computes its own T12 total — broker "Totals" column is ignored.
            total = sum(monthly)

            # Drop rules. Subtotal-pop fires when Rule 2 catches a "Subtotal,"
            # row: the sub-section is closing, so the next row is a sibling of
            # the parent. Rewind current_banner to top_banner.
            if not _has_any_dollar_value(monthly, total):
                continue
            if _matches_grand_total(desc):
                if self.SUBTOTAL_RE.match(desc):
                    current_banner = top_banner
                continue
            if _matches_drop_list(desc):
                continue

            # Prefix only when we're under a SUB-banner. Top-level rows
            # (current_banner == top_banner) emit unprefixed descriptions.
            if current_banner is not None and current_banner != top_banner:
                final_desc = f"{current_banner} | {desc}"
            else:
                final_desc = desc

            rows.append(GLRow(
                account="",
                description=final_desc,
                monthly=monthly,
                total=total,
            ))

        return rows, labels


# ---------------------------------------------------------------------------
# Cluster B helpers — sign-convention guards + partial-year detection (v0.2.0)
# ---------------------------------------------------------------------------

# Description keywords that should normally produce negative T12 totals. If a
# row's description contains one of these and total > 0, surface a warning.
# Defensive — none of the four verified fixtures (Salem, Briar Glen, Homestead,
# March_2026) trip these on standard signs.
#
# Limited to "CONCESSION" because that's the only line item with universally
# negative convention. Vacancy / Loss-to-Lease / Bad Debt vary by operator:
# some report them as contra-revenue (negative), others as expense (positive).
# Both are valid; substrate v0.1.4's Monthly Trending R10/R11 already absorbs
# either sign. Bad Debt as positive expense is the broker convention.
SIGN_NEGATIVE_KEYWORDS: Tuple[str, ...] = (
    "CONCESSION",
)


def _check_sign_convention(gl_rows: List[GLRow]) -> List[str]:
    """Return a list of human-readable warnings for sign anomalies.

    Convention assumed: revenue +, expense +, concession -. Caller surfaces
    warnings to the operator; no row is dropped.

    The check examines the description SUFFIX (after the most-recent " | "
    separator if any), not the full prefixed string — this avoids false
    positives when a banner happens to contain a keyword (e.g. broker banner
    "Management Fee & Bad Debt" contains "BAD DEBT" but its child rows are
    legitimately positive expense lines).
    """
    warnings: List[str] = []
    for row in gl_rows:
        # Strip parser prefix (banner | description) for keyword matching.
        suffix = row.description.split(" | ", 1)[-1].upper()
        for kw in SIGN_NEGATIVE_KEYWORDS:
            if kw in suffix and row.total > 0:
                warnings.append(
                    f"Sign convention: '{row.description}' has positive total "
                    f"(${row.total:,.2f}); expected negative. Review source."
                )
                break
    return warnings


def _count_populated_months(gl_rows: List[GLRow]) -> int:
    """Return the count of month columns (0..12) where at least one GL row
    has a non-zero value. Drives partial-year T12 detection."""
    if not gl_rows:
        return 0
    populated = [False] * 12
    for row in gl_rows:
        for i, v in enumerate(row.monthly):
            if v != 0:
                populated[i] = True
    return sum(populated)


def _annualize_rows(gl_rows: List[GLRow], populated_months: int) -> List[GLRow]:
    """Multiply every GLRow's monthly + total by 12/populated_months. Returns
    new GLRow instances; does not mutate inputs."""
    if populated_months <= 0 or populated_months >= 12:
        return gl_rows
    multiplier = 12.0 / populated_months
    out: List[GLRow] = []
    for row in gl_rows:
        out.append(GLRow(
            account=row.account,
            description=row.description,
            monthly=[v * multiplier for v in row.monthly],
            total=row.total * multiplier,
        ))
    return out


# ---------------------------------------------------------------------------
# Registry — order matters. More-specific (sheet-name) matchers go first so
# they win over the content-based scanners. BrokerFinancialSummaryFormat
# matches on a content fingerprint (A4 = "Historical Performance"), and broker
# files do not collide with Yardi or MRI sheet names — but it's last in the
# list as a defensive ordering choice (specific names before content scans).
# ---------------------------------------------------------------------------

REGISTRY: List[T12Format] = [
    MriR12mincsFormat(),
    YardiIncomeToBudgetFormat(),
    BrokerFinancialSummaryFormat(),
]


# ---------------------------------------------------------------------------
# Description_Map reader (UNMATCHED detection)
# ---------------------------------------------------------------------------

DESCMAP_SHEET = "Description_Map"
DESCMAP_DATA_START_ROW = 5
DESCMAP_DESC_COL = 1


def read_descmap_descriptions(analyzer_wb: openpyxl.Workbook) -> Set[str]:
    """Return the set of TRIMmed raw descriptions from the destination's
    `Description_Map` sheet, used to detect UNMATCHED rows.

    Comparison key matches the col P formula: TRIM applied, case-sensitive.
    """
    if DESCMAP_SHEET not in analyzer_wb.sheetnames:
        raise ValueError(
            f"Destination workbook missing required sheet '{DESCMAP_SHEET}'. "
            f"Found: {analyzer_wb.sheetnames}"
        )
    ws = analyzer_wb[DESCMAP_SHEET]
    out: Set[str] = set()
    for r in range(DESCMAP_DATA_START_ROW, ws.max_row + 1):
        v = ws.cell(r, DESCMAP_DESC_COL).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            out.add(s)
    return out


# ---------------------------------------------------------------------------
# Top-level entry point
# ---------------------------------------------------------------------------

def parse_t12(
    t12_bytes: bytes,
    descmap_descriptions: Set[str],
    annualize_partial_year: bool = False,
) -> T12ParseResult:
    """Parse a raw T12 file (bytes) against a destination Description_Map.

    Args:
        t12_bytes: contents of an uploaded T12 .xlsx file
        descmap_descriptions: set of trimmed raw descriptions from the
            destination workbook's Description_Map (use
            `read_descmap_descriptions(analyzer_wb)`)
        annualize_partial_year: if True AND fewer than 12 months are populated,
            scale every row's monthly + total by 12/populated_months. Default
            False — partial-year flagging happens via the surfaced warning and
            populated_months field. Caller (app.py) controls this via checkbox.

    Returns:
        T12ParseResult with gl_rows, month_labels, unmatched, format_name,
        sheet_name, sign_warnings, populated_months, was_annualized.

    Raises:
        UnknownT12FormatError: if no registered format detector matches.
    """
    wb = openpyxl.load_workbook(io.BytesIO(t12_bytes), data_only=True)

    matched_format: Optional[T12Format] = None
    matched_sheet: Optional[str] = None
    for fmt in REGISTRY:
        sheet = fmt.detect(wb)
        if sheet is not None:
            matched_format = fmt
            matched_sheet = sheet
            break

    if matched_format is None or matched_sheet is None:
        raise UnknownT12FormatError(
            f"No registered T12 format matched. "
            f"Sheets seen: {wb.sheetnames}. "
            f"Supported formats: {[f.name for f in REGISTRY]}. "
            f"Adding a new format = subclass T12Format, register in REGISTRY."
        )

    gl_rows, month_labels = matched_format.extract(wb, matched_sheet)

    # Cluster B — partial-year detection (B-2)
    populated_months = _count_populated_months(gl_rows)
    was_annualized = False
    if annualize_partial_year and 0 < populated_months < 12:
        gl_rows = _annualize_rows(gl_rows, populated_months)
        was_annualized = True

    # Cluster B — sign-convention guards (B-1). Run AFTER any annualization so
    # warning $ values reflect what the user will see in the workbook.
    sign_warnings = _check_sign_convention(gl_rows)

    # UNMATCHED detection (Python-side; col P formula in workbook is the
    # runtime source of truth)
    unmatched: List[str] = []
    seen: Set[str] = set()
    for row in gl_rows:
        if row.description not in descmap_descriptions and row.description not in seen:
            unmatched.append(row.description)
            seen.add(row.description)

    return T12ParseResult(
        gl_rows=gl_rows,
        month_labels=month_labels,
        unmatched=unmatched,
        format_name=matched_format.name,
        sheet_name=matched_sheet,
        sign_warnings=sign_warnings,
        populated_months=populated_months,
        was_annualized=was_annualized,
    )
