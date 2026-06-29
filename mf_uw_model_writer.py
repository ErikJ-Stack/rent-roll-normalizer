"""
MF UW Model writer — paste normalized RR / AR / T-12 into MF_UW_Model_v25.xlsx.

Writes only the two intake grids (the diagnostic/analytic layers are formula-
driven and left untouched):
  - T-12 → `T-12 Analysis` Layer 1 (anchor A106): Acct# / Name / 12 months /
    O=SUM formula / P=`_StdCOA` bucket.
  - RR (+joined AR) → `Rent Roll Analysis` grid (anchor A273, cols A–T core).

All four write-target sheets are layout-identical across v15 / v20 / v25 (T-12
Analysis Layer 1 @106, Rent Roll Analysis grid @273 cols A–AK / data 273–1772,
Prop Info col A label + col B value, Rental Comps @8), so the same anchors drive
every version — `template_version` is informational only. v25 reshuffled the
non-target sheets (`Dashboard` renamed `Dash`, `Data Refresh` removed → 23
sheets) and trimmed the now-blank `Rent Roll Analysis` helper columns to a
single AL (chart helper, still outside the writer's A–AK clear band → preserved)
and Prop Info to cols A–D; none of that touches the write anchors.

Like v20 (but unlike v15), v25 carries an `xl/metadata.xml` part for Excel-365
dynamic-array semantics (7 `cm`-marked cells). The `_restore_dynamic_arrays`
call below — a no-op on v15 — re-injects `metadata.xml` + the `cm` markers after
openpyxl's save, which would otherwise drop them (openpyxl quirk #6). Verified
7→7 `cm` markers preserved.

openpyxl still drops auxiliary parts on save — **cell comments + their
indicators, the Claude-for-Excel add-in, custom doc properties, and (new in
v25) the extended (x14) data-validation dropdowns on `Rent Roll Analysis`** (no
data/formulas/charts affected; the writer fills those cells with real values
regardless) — surfaced in the report as a warning; open + re-save in Excel only
if you need those annotations / dropdowns back.

Public API:
    populate_mf_model(model_bytes, *, t12=None, rr=None, property_name=None,
                      property_units=None) -> (bytes, dict)
"""
from __future__ import annotations

import io
import re

import openpyxl

from uw_template_writer import _restore_dynamic_arrays

# Rent Roll Analysis grid (header row 272, data 273+). 1-based column indices.
_RR = dict(bldg=1, unit=2, unit_type=3, sqft=4, status=5, resident=6, legal=7,
           move_in=8, lease_start=9, lease_end=10, exp_move_out=11,
           market_rent=12, actual_charges=13, scheduled_charges=14, balance=15,
           deposit=16, ar_0_30=17, ar_31_60=18, ar_61_90=19, ar_90_plus=20,
           notes=22)
_RR_ANCHOR = 273
_RR_END = 1772
_RR_COLS = 37  # A–AK
# Per-unit ancillary breakout buckets -> Rent Roll Analysis cols W–AK.
_ANCILLARY_COL = {
    "mtm": 23, "application": 24, "late": 25, "utility_reimb": 26, "pet": 27,
    "parking": 28, "amenity": 29, "admin": 30, "insurance_passthru": 31,
    "misc": 32, "storage": 33, "package": 34, "lease_lock": 35, "valet": 36,
    "lease_break": 37,
}

# T-12 Analysis Layer 1 (header row 105, data 106+).
_T12_ANCHOR = 106
_T12_END = 255
_T12_HEADER = 105

# Prop Info — label in col A, value to col B (manual-input cells, all blank in
# v15). Populated from an OM extraction (mf_om_extractor.MFOMResult).
_PROP = dict(
    property_name=4, address=5, units=6, num_buildings=7, num_stories=8,
    year_built=9, lot_acres=10, gross_sqft=11, parking_spaces=12,
    building_class=13, amenity_tier=14, total_rentable_sf=15,
    studio_units=16, br1_units=17, br2_units=18,
    city_market=20, msa_name=21, msa_population=22, city_population=23,
    population_growth=24, avg_household_income=27, median_income=28,
    school_rating=29, br3_units=31, blended_rent=32, avg_unit_sf=33,
    market_vacancy=34, market_rent_growth=35, new_supply=36, renter_pct=37,
    electric_meter=39, water_meter=40, gas=41, trash=42,
    value_add_thesis=47,
)

# Rental Comps — comp set rows 8–22 (15 max). SUBJECT row 7 and the Z/AA
# (eff-rent, $/SF) formulas are left untouched. 1-based column indices.
_COMP = dict(distance=17, name=18, building_class=19, vintage=20, units=21,
             unit_type=22, avg_sf=23, asking_rent=24, concession=25,
             occ=28, comment=30)
_COMP_ANCHOR = 8
_COMP_MAX = 15


def _occ_fraction(s):
    """'96%' / '0.96' / 96 -> 0.96; '' -> None. Rental Comps AB expects a fraction."""
    if s in (None, ""):
        return None
    m = re.search(r"[\d.]+", str(s))
    if not m:
        return None
    v = float(m.group(0))
    return v / 100 if v > 1.5 else v


def _comp_vintage(s):
    """First 4-digit year in a vintage string ('1975/2005' -> 1975)."""
    m = re.search(r"(19|20)\d{2}", str(s or ""))
    return int(m.group(0)) if m else None


class MFModelWriterError(Exception):
    pass


def _set(ws, r, c, v):
    if v is not None and v != "":
        ws.cell(r, c).value = v


def _clear_block(ws, r0, r1, c0, c1):
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            cell = ws.cell(r, c)
            if cell.value is not None:
                cell.value = None


def _bedroom_bucket(unit_type: str) -> str | None:
    """Map a floorplan label to studio/br1/br2/br3 for the Prop Info counts."""
    s = (unit_type or "").lower()
    if re.search(r"\bstudio\b|\bstu\b|\befficien", s):
        return "studio_units"
    m = re.search(r"(\d+)\s*(?:x|bd|br|bed|/)", s) or re.search(r"^\s*(\d+)\b", s)
    if m:
        n = int(m.group(1))
        return {1: "br1_units", 2: "br2_units", 3: "br3_units"}.get(n)
    return None


def _write_prop_info(pi, om) -> int:
    """Write the OM property-details + market block to Prop Info col B. Count cells."""
    info, mk = om.prop_info, om.market
    addr = info.address
    if info.city and info.city.lower() not in addr.lower():
        addr = ", ".join(p for p in [addr, info.city, info.state, info.zip_code] if p)
    vals = {
        "property_name": info.property_name, "address": addr,
        "units": info.units_total, "num_buildings": info.num_buildings,
        "num_stories": info.num_stories, "year_built": info.year_built,
        "lot_acres": info.lot_acres, "gross_sqft": info.gross_sqft,
        "parking_spaces": info.parking_spaces,
        "building_class": info.building_class, "amenity_tier": info.amenity_tier,
        "total_rentable_sf": info.total_rentable_sf,
        "studio_units": info.studio_units, "br1_units": info.br1_units,
        "br2_units": info.br2_units, "br3_units": info.br3_units,
        "avg_unit_sf": info.avg_unit_sf,
        "electric_meter": info.electric_meter, "water_meter": info.water_meter,
        "gas": info.gas, "trash": info.trash,
        "value_add_thesis": info.value_add_thesis,
        # market block
        "city_market": mk.city_market, "msa_name": mk.msa_name,
        "msa_population": mk.msa_population, "city_population": mk.city_population,
        "population_growth": mk.population_growth_rate,
        "avg_household_income": mk.avg_household_income,
        "median_income": mk.median_income, "school_rating": mk.school_rating,
        "market_vacancy": mk.market_vacancy_rate,
        "market_rent_growth": mk.market_rent_growth,
        "new_supply": mk.new_supply_units, "renter_pct": mk.renter_pct,
    }
    # derive bedroom counts from unit_mix when the explicit fields are absent
    if not any(vals[k] for k in ("studio_units", "br1_units", "br2_units", "br3_units")):
        agg: dict[str, int] = {}
        for row in info.unit_mix:
            bucket = _bedroom_bucket(row.unit_type)
            if bucket and row.count:
                agg[bucket] = agg.get(bucket, 0) + row.count
        for k, v in agg.items():
            vals[k] = v
    written = 0
    for key, val in vals.items():
        if val in (None, "") or key not in _PROP:
            continue
        pi.cell(_PROP[key], 2).value = val
        written += 1
    return written


def _write_rental_comps(rc, comps) -> int:
    """Write up to 15 comps into Rental Comps rows 8–22 (cols Q–AD). Count comps."""
    n = min(len(comps), _COMP_MAX)
    for i, c in enumerate(comps[:n]):
        r = _COMP_ANCHOR + i
        _set(rc, r, _COMP["distance"], c.distance_mi)
        _set(rc, r, _COMP["name"], c.name)
        _set(rc, r, _COMP["building_class"], c.building_class)
        _set(rc, r, _COMP["vintage"], _comp_vintage(c.year_built))
        _set(rc, r, _COMP["units"], c.units)
        _set(rc, r, _COMP["unit_type"], c.unit_type)
        _set(rc, r, _COMP["avg_sf"], c.avg_sf)
        _set(rc, r, _COMP["asking_rent"], c.asking_rent)
        if c.concession_weeks is not None:
            rc.cell(r, _COMP["concession"]).value = c.concession_weeks
        _set(rc, r, _COMP["occ"], _occ_fraction(c.occupancy))
        _set(rc, r, _COMP["comment"], c.comment)
    return n


def populate_mf_model(model_bytes: bytes, *, t12=None, rr=None, om=None,
                      property_name: str | None = None,
                      property_units: int | None = None,
                      progress=None):
    """Populate the MF UW Model. Returns (xlsx_bytes, report_dict).

    om: an mf_om_extractor.MFOMResult — writes the Prop Info details/market
        block and the Rental Comps comp set. RR-derived units/name take
        precedence over OM where they overlap (the rent roll is authoritative).
    progress: optional callable(frac: float in 0..1) invoked at build
        milestones (load / T-12 / RR / OM / save) so a caller can drive a
        determinate progress UI. The openpyxl load and save are single opaque
        calls, so the fraction steps between milestones rather than streaming.
    """
    _p = progress if callable(progress) else (lambda *_a, **_k: None)
    wb = openpyxl.load_workbook(io.BytesIO(model_bytes), data_only=False)
    _p(0.15)   # model loaded (one of the two slow openpyxl calls)
    report = {"t12_lines": 0, "t12_cells": 0, "rr_units": 0, "rr_cells": 0,
              "om_prop_cells": 0, "om_comps": 0, "warnings": []}

    # --- T-12 Analysis Layer 1 ---
    if t12 is not None:
        if "T-12 Analysis" not in wb.sheetnames:
            raise MFModelWriterError("Model missing 'T-12 Analysis' sheet.")
        ws = wb["T-12 Analysis"]
        _clear_block(ws, _T12_ANCHOR, _T12_END, 1, 16)  # A–P
        # month headers (C–N) aligned to the operator's actual months
        for i, lab in enumerate(t12.month_labels[:12]):
            _set(ws, _T12_HEADER, 3 + i, lab)
        n = min(len(t12.lines), _T12_END - _T12_ANCHOR + 1)
        if len(t12.lines) > n:
            report["warnings"].append(
                f"T-12 has {len(t12.lines)} lines; Layer 1 holds {n} — extra truncated.")
        for i, ln in enumerate(t12.lines[:n]):
            r = _T12_ANCHOR + i
            _set(ws, r, 1, ln.acct or "")
            _set(ws, r, 2, ln.name)
            for m, val in enumerate(ln.monthly[:12]):
                _set(ws, r, 3 + m, val)
            ws.cell(r, 15).value = f"=SUM(C{r}:N{r})"   # O = T-12 Total
            _set(ws, r, 16, ln.bucket)                   # P = → MAPPING
            report["t12_cells"] += 16
        report["t12_lines"] = n
    _p(0.45)   # T-12 Layer 1 written

    # --- Rent Roll Analysis grid ---
    if rr is not None:
        if "Rent Roll Analysis" not in wb.sheetnames:
            raise MFModelWriterError("Model missing 'Rent Roll Analysis' sheet.")
        ws = wb["Rent Roll Analysis"]
        _clear_block(ws, _RR_ANCHOR, _RR_END, 1, _RR_COLS)
        n = min(len(rr.units), _RR_END - _RR_ANCHOR + 1)
        if len(rr.units) > n:
            report["warnings"].append(
                f"RR has {len(rr.units)} units; grid holds {n} — extra truncated.")
        for i, u in enumerate(rr.units[:n]):
            r = _RR_ANCHOR + i
            _set(ws, r, _RR["bldg"], u.bldg)
            _set(ws, r, _RR["unit"], u.unit)
            _set(ws, r, _RR["unit_type"], u.unit_type)
            _set(ws, r, _RR["sqft"], u.sqft)
            _set(ws, r, _RR["status"], u.status)
            _set(ws, r, _RR["resident"], u.resident)
            ws.cell(r, _RR["legal"]).value = bool(u.legal)   # Excel boolean
            _set(ws, r, _RR["move_in"], u.move_in)
            _set(ws, r, _RR["lease_start"], u.lease_start)
            _set(ws, r, _RR["lease_end"], u.lease_end)
            _set(ws, r, _RR["exp_move_out"], u.exp_move_out)
            _set(ws, r, _RR["market_rent"], u.market_rent)
            _set(ws, r, _RR["actual_charges"], u.actual_charges)
            _set(ws, r, _RR["scheduled_charges"], u.scheduled_charges)
            _set(ws, r, _RR["balance"], u.balance)
            _set(ws, r, _RR["deposit"], u.deposit)
            _set(ws, r, _RR["ar_0_30"], u.ar_0_30)
            _set(ws, r, _RR["ar_31_60"], u.ar_31_60)
            _set(ws, r, _RR["ar_61_90"], u.ar_61_90)
            _set(ws, r, _RR["ar_90_plus"], u.ar_90_plus)
            _set(ws, r, _RR["notes"], u.notes)
            report["rr_cells"] += 21
            # per-unit ancillary breakout (W–AK), e.g. Amenity Rent -> Amenity Fees
            for bucket, amount in (u.ancillary or {}).items():
                col = _ANCILLARY_COL.get(bucket)
                if col and amount:
                    _set(ws, r, col, amount)
                    report["rr_cells"] += 1
        report["rr_units"] = n
    _p(0.80)   # Rent Roll grid written

    # --- Prop Info: OM details + market block ---
    if om is not None:
        if "Prop Info" not in wb.sheetnames:
            raise MFModelWriterError("Model missing 'Prop Info' sheet.")
        report["om_prop_cells"] = _write_prop_info(wb["Prop Info"], om)
        # --- Rental Comps: OM comp set ---
        if "Rental Comps" in wb.sheetnames and om.comps:
            report["om_comps"] = _write_rental_comps(wb["Rental Comps"], om.comps)
        for w in om.warnings:
            report["warnings"].append(f"OM: {w}")

    # --- Prop Info name/units (RR-authoritative; overrides OM) ---
    if "Prop Info" in wb.sheetnames:
        pi = wb["Prop Info"]
        if property_name:
            _set(pi, 4, 2, property_name)        # B4 Property Name
        units = property_units if property_units is not None else (
            rr.unit_count if rr is not None else None)
        if units is not None:
            pi.cell(6, 2).value = units          # B6 # Units
    _p(0.90)   # all sheets written; saving next (the other slow openpyxl call)

    buf = io.BytesIO()
    wb.save(buf)
    out = buf.getvalue()
    out = _restore_dynamic_arrays(out, model_bytes)   # no-op on v15 (no metadata.xml)
    _p(1.0)    # workbook saved
    report["warnings"].append(
        "Cell comments, the Claude-for-Excel add-in, custom doc properties, and "
        "the extended (x14) data-validation dropdowns on Rent Roll Analysis are "
        "dropped by the Excel writer (no data/formulas/charts affected) — open + "
        "re-save in Excel if you need those annotations / dropdowns back.")
    return out, report
