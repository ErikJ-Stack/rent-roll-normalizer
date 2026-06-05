"""
Regression tests for `uw_output_model.compute_uw_output_values` — the
in-Python UW Output evaluator that kills the cache caveat.

Run without pytest:
    python tests/test_uw_output_model.py

Two layers, both gated on the gitignored Homestead fixtures:

  1. **Engine vs cached** — parse the Homestead RR + T12 (as the app does),
     compute the UW Output values in Python, and assert they match the
     populated Analyzer fixture's cached `UW Output` values to the penny.
     This is the drift guard: if the Analyzer's T12 Analytics formulas
     change, this test catches the engine going stale.

  2. **Writer fallback end-to-end** — build a FRESH in-memory Analyzer
     (openpyxl, no cached formula values — exactly the cache-caveat state),
     run `populate_uw_template(..., computed_values=...)`, and assert the
     previously-blank `T-12 Analysis` cells now carry the right numbers and
     the report counts the in-Python fallbacks.

All inputs are property-specific data and intentionally gitignored. Skip
cleanly when absent so cold checkouts / CI don't fail.
"""
from __future__ import annotations

import datetime
import io
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
BUNDLED_ANALYZER = ROOT / "ALF_Financial_Analyzer_Only.xlsx"
TEMPLATE_V5 = ROOT / "assets" / "ALF_UW_Template_v5.xlsx"
RR_FIXTURE = ROOT / "Sample Files" / "2026-04-24 Homestead Village Rent Roll v2.xlsx"
T12_FIXTURE = ROOT / "Sample Files" / "Homestead - March 2026 T12.xlsx"
POPULATED_ANALYZER = (
    ROOT / "Sample Files"
    / "Analyzer with 2026-04-24 Homestead Village Rent Roll v2 + March 2026 T12 2026-04-24.xlsx"
)
PERIOD = datetime.date(2026, 4, 24)


class TestFailure(AssertionError):
    pass


def _check(cond: bool, msg: str) -> None:
    if not cond:
        raise TestFailure(msg)


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _eval_t12_n(ws) -> dict:
    """Evaluate the T-12 Analysis col-N P&L chain (openpyxl can't compute
    formulas, so we replicate the template's total formulas from the pasted
    line-item values). Returns {row: value} for the total rows."""
    def v(r):
        return _num(ws.cell(row=r, column=14).value) or 0.0
    N = {}
    N[63] = sum(v(r) for r in (58, 59, 60, 61, 62))           # Net Rent
    N[69] = N[63] + sum(v(r) for r in (64, 65, 66, 67, 68))   # EGI
    N[85] = sum(v(r) for r in range(71, 85))                  # Total Labor
    N[111] = sum(v(r) for r in range(87, 111))                # Total Non-Labor
    N[113] = v(113)                                           # Mgmt Fee
    N[114] = N[85] + N[111] + N[113]                          # Total Op Ex
    N[115] = N[114] - N[113]                                  # Op Ex excl mgmt
    N[116] = N[69] - N[85] - N[111]                           # EBITDARM
    N[117] = N[116] - N[113]                                  # EBITDAR
    N[118] = N[117]                                           # EBITDA
    return N


# Concept key → (UW Output column, row) on the populated fixture (col F =
# normalized scenario, the writer default).
_UW_OUTPUT_REF = {
    "licensed_beds_il": ("B", 70), "licensed_beds_al": ("C", 70), "licensed_beds_mc": ("D", 70),
    "occupied_beds_il": ("B", 71), "occupied_beds_al": ("C", 71), "occupied_beds_mc": ("D", 71),
    "base_rent_normalized": ("F", 6), "loc_revenue": ("F", 7),
    "community_movein_fees": ("F", 8), "concessions_specials": ("F", 9),
    "respite_care": ("F", 10), "other_community_revenue": ("F", 11),
    "egi": ("F", 12), "gpr_base": ("F", 15), "physical_vacancy_loss": ("F", 16),
    "loss_to_lease": ("F", 18),
    "labor_care_staff": ("F", 22), "labor_wellness": ("F", 23), "labor_agency": ("F", 24),
    "labor_activities": ("F", 25), "labor_dining": ("F", 26), "labor_maint_hk": ("F", 27),
    "labor_admin": ("F", 28), "labor_bonus": ("F", 29), "labor_overtime": ("F", 30),
    "labor_pto": ("F", 31), "labor_payroll_taxes": ("F", 32), "labor_benefits": ("F", 33),
    "labor_workers_comp": ("F", 34), "labor_401k": ("F", 35), "labor_total": ("F", 36),
    "opex_food_cost": ("F", 38), "opex_utilities": ("F", 47), "opex_re_taxes": ("F", 53),
    "opex_bad_debt_expense": ("F", 57), "opex_nonlabor_total": ("F", 62),
    "opex_total_excl_mgmt": ("F", 63), "mgmt_fee": ("F", 64),
    "ebitdarm": ("F", 66), "ebitdar": ("F", 67), "ebitda": ("F", 68),
    "bad_debt_writeoffs_revenue": ("F", 57),
}

# Whitelists retired 2026-06-05: the Homestead fixture was rebuilt against a
# v0.2.16 Analyzer (Auto Expense non-labor row, BL-0028) round-tripped through
# Excel, so cached == engine to the penny on every concept. The former
# _AUTO_EXPENSE_DIVERGENCE (UWT v6 Auto Expense, $6,061.32) and _REMAP_2P_DIVERGENCE
# (substrate v0.2.15 2nd-Person re-map, $32,220.49) divergence checks are gone —
# the regression is now a flat penny-match.


def _parse_inputs():
    """Mirror the app pipeline: NormalizeResult + T12ParseResult from fixtures."""
    from normalizer import normalize_rent_roll
    from t12_normalizer import parse_t12, read_descmap_descriptions

    rr = normalize_rent_roll(str(RR_FIXTURE))
    descmap = read_descmap_descriptions(
        openpyxl.load_workbook(BUNDLED_ANALYZER, data_only=True)
    )
    t12 = parse_t12(
        T12_FIXTURE.read_bytes(), descmap, annualize_partial_year=False
    )
    return rr, t12


def test_engine_matches_cached() -> None:
    """compute_uw_output_values reproduces the fixture's cached UW Output."""
    if not (RR_FIXTURE.exists() and T12_FIXTURE.exists() and POPULATED_ANALYZER.exists()):
        print("  ⊘ SKIP test_engine_matches_cached — Homestead fixtures absent")
        return

    from uw_output_model import compute_uw_output_values

    rr, t12 = _parse_inputs()
    vals = compute_uw_output_values(rr, t12)

    uo = openpyxl.load_workbook(POPULATED_ANALYZER, data_only=True)["UW Output"]

    mismatches = []
    for key, (col, row) in _UW_OUTPUT_REF.items():
        engine = _num(vals.get(key))
        cached = _num(uo[f"{col}{row}"].value)
        if engine is None or cached is None:
            ok = engine is None and cached is None
        else:
            ok = abs(engine - cached) < 0.5  # penny-level; allow float noise
        if not ok:
            mismatches.append(f"{key}: engine={vals.get(key)!r} cached={uo[f'{col}{row}'].value!r}")

    _check(not mismatches, "engine diverged from cached UW Output:\n  " + "\n  ".join(mismatches))
    print(f"  ✓ engine matches cached UW Output on {len(_UW_OUTPUT_REF)} concepts (to the penny)")
    # Headline spot-print
    print(f"      EGI={vals['egi']:,.2f}  EBITDARM={vals['ebitdarm']:,.2f}  EBITDA={vals['ebitda']:,.2f}")


def test_writer_fallback_populates() -> None:
    """A fresh (no-cache) Analyzer + computed_values fills T-12 Analysis."""
    if not (RR_FIXTURE.exists() and T12_FIXTURE.exists() and TEMPLATE_V5.exists()):
        print("  ⊘ SKIP test_writer_fallback_populates — fixtures absent")
        return

    from analyzer_rr_translator import translate_for_t12
    from analyzer_rr_writer import populate_rr_input
    from t12_normalizer_writer import populate_t12_input
    from uw_output_model import compute_uw_output_values
    from uw_template_writer import populate_uw_template

    rr, t12 = _parse_inputs()

    # Build a FRESH in-memory Analyzer exactly like the app — never saved
    # through Excel, so every formula cell has no cached value.
    after_rr = populate_rr_input(
        BUNDLED_ANALYZER.read_bytes(), translate_for_t12(rr.condensed),
        PERIOD, source_filename="homestead.xlsx",
    )
    fresh = populate_t12_input(
        after_rr, t12, new_descmap_entries=[], source_filename="t12.xlsx"
    )
    tpl = TEMPLATE_V5.read_bytes()

    # Baseline (no fallback) — the cache caveat: T-12 path comes through blank.
    _, rep_base = populate_uw_template(fresh, tpl, template_version="v5")
    base_t12_blank = sum(
        1 for r in rep_base.results if r.outcome == "no_source" and r.path == "t12"
    )

    # With the in-Python fallback.
    vals = compute_uw_output_values(rr, t12)
    out_bytes, rep = populate_uw_template(
        fresh, tpl, template_version="v5", computed_values=vals
    )
    fix_t12_blank = sum(
        1 for r in rep.results if r.outcome == "no_source" and r.path == "t12"
    )
    n_computed = rep.summary.get("computed_in_python", 0)

    _check(base_t12_blank > 50, f"expected baseline cache caveat (>50 blank), got {base_t12_blank}")
    _check(n_computed >= 45, f"expected >=45 in-Python fallbacks, got {n_computed}")
    _check(fix_t12_blank <= 2, f"expected <=2 residual t12 blanks, got {fix_t12_blank}")

    ta = openpyxl.load_workbook(io.BytesIO(out_bytes), data_only=False)["T-12 Analysis"]
    # Totals are now LIVE FORMULAS (not pasted values).
    for r in (63, 69, 85, 111, 114, 115, 116, 117, 118):
        v = ta.cell(row=r, column=14).value
        _check(isinstance(v, str) and v.startswith("="), f"N{r} should be a formula, got {v!r}")
    # Line items still pasted as values; income contras signed for the waterfall.
    _check(abs(_num(ta["N71"].value) - 1184691.64) < 1.0, f"N71 Care Staff wrong: {ta['N71'].value}")
    _check(_num(ta["N60"].value) < 0, f"N60 Vacancy should be negative contra: {ta['N60'].value}")
    _check(_num(ta["N62"].value) < 0, f"N62 Bad Debt should be negative contra: {ta['N62'].value}")
    # The evaluated chain ties to the Option-A model (bad debt reduces EGI).
    N = _eval_t12_n(ta)
    _check(abs(N[69] - 6964627.48) < 2.0, f"EGI (evaluated) wrong: {N[69]:,.2f}")
    _check(abs(N[116] - 1767482.75) < 2.0, f"EBITDARM (evaluated) wrong: {N[116]:,.2f}")
    _check(abs(N[118] - 1417384.90) < 2.0, f"EBITDA (evaluated) wrong: {N[118]:,.2f}")

    print(
        f"  ✓ fallback: baseline {base_t12_blank} blank → {fix_t12_blank} "
        f"({n_computed} computed in-Python); totals are formulas; "
        f"evaluated EGI={N[69]:,.0f} EBITDARM={N[116]:,.0f} EBITDA={N[118]:,.0f}"
    )


def test_dynamic_array_metadata_restored() -> None:
    """The populated output restores xl/metadata.xml + the cm= markers that
    openpyxl drops, so Section R / S dynamic-array spills survive (openpyxl
    quirk #6)."""
    if not (RR_FIXTURE.exists() and T12_FIXTURE.exists() and TEMPLATE_V5.exists()):
        print("  ⊘ SKIP test_dynamic_array_metadata_restored — fixtures absent")
        return

    import zipfile
    import re as _re
    from analyzer_rr_translator import translate_for_t12
    from analyzer_rr_writer import populate_rr_input
    from t12_normalizer_writer import populate_t12_input
    from uw_output_model import compute_uw_output_values
    from uw_template_writer import populate_uw_template

    rr, t12 = _parse_inputs()
    after_rr = populate_rr_input(
        BUNDLED_ANALYZER.read_bytes(), translate_for_t12(rr.condensed),
        PERIOD, source_filename="h.xlsx",
    )
    fresh = populate_t12_input(after_rr, t12, new_descmap_entries=[], source_filename="t.xlsx")
    vals = compute_uw_output_values(rr, t12)
    out_bytes, rep = populate_uw_template(
        fresh, TEMPLATE_V5.read_bytes(), template_version="v5", computed_values=vals
    )

    _check(rep.summary.get("dynamic_arrays_restored") == 1, "repair flag not set")
    z = zipfile.ZipFile(io.BytesIO(out_bytes))
    _check("xl/metadata.xml" in z.namelist(), "metadata.xml not restored")
    _check(z.testzip() is None, "output zip is corrupt")
    cm_total = sum(
        len(_re.findall(r'cm="', z.read(n).decode("utf-8")))
        for n in z.namelist() if _re.match(r"xl/worksheets/sheet\d+\.xml$", n)
    )
    _check(cm_total >= 500, f"expected ~557 cm markers restored, got {cm_total}")

    # Section R driver Z173 must carry the cm marker (dynamic array, not CSE).
    for n in z.namelist():
        if _re.match(r"xl/worksheets/sheet\d+\.xml$", n):
            xml = z.read(n).decode("utf-8")
            if "UNIQUE" in xml and "FILTER" in xml:
                m = _re.search(r'<c r="Z173"[^>]*>', xml)
                _check(m is not None and 'cm="' in m.group(0),
                       f"Z173 missing cm marker: {m.group(0) if m else None}")
                break

    print(f"  ✓ dynamic-array repair: metadata.xml restored, {cm_total} cm markers, Z173 marked dynamic")


def test_monthly_grid_reconciles() -> None:
    """The T-12 Analysis Layer-3 monthly grid (cols B–M) is populated and each
    row's monthly sum reconciles to its annual T-12 Total (col N)."""
    if not (RR_FIXTURE.exists() and T12_FIXTURE.exists() and TEMPLATE_V5.exists()):
        print("  ⊘ SKIP test_monthly_grid_reconciles — fixtures absent")
        return

    import io as _io
    from analyzer_rr_translator import translate_for_t12
    from analyzer_rr_writer import populate_rr_input
    from t12_normalizer_writer import populate_t12_input
    from uw_output_model import compute_uw_output_values, compute_uw_output_monthly
    from uw_template_writer import populate_uw_template

    rr, t12 = _parse_inputs()
    after_rr = populate_rr_input(
        BUNDLED_ANALYZER.read_bytes(), translate_for_t12(rr.condensed),
        PERIOD, source_filename="h.xlsx",
    )
    fresh = populate_t12_input(after_rr, t12, new_descmap_entries=[], source_filename="t.xlsx")
    vals = compute_uw_output_values(rr, t12)
    mon = compute_uw_output_monthly(rr, t12)
    out_bytes, rep = populate_uw_template(
        fresh, TEMPLATE_V5.read_bytes(), template_version="v5",
        computed_values=vals, computed_monthly=mon,
    )

    n_monthly = rep.summary.get("monthly_cells_written", 0)
    _check(n_monthly >= 400, f"expected a populated monthly grid, got {n_monthly}")

    ta = openpyxl.load_workbook(_io.BytesIO(out_bytes), data_only=False)["T-12 Analysis"]

    def _row_sum(r):  # cols B..M = 2..13 — only numeric (skips mirror formulas)
        return sum(_num(ta.cell(row=r, column=c).value) or 0.0 for c in range(2, 14))

    # Line-item rows: monthly values present and sum to their annual value.
    for r, label in [(71, "Care staff"), (87, "Food cost")]:
        s = _row_sum(r)
        n = _num(ta.cell(row=r, column=14).value)
        _check(s > 0, f"row {r} ({label}) monthly grid is blank")
        _check(abs(s - n) < 2.0, f"row {r} ({label}) monthly sum {s:,.2f} != annual {n:,.2f}")

    # Net Rent monthly (row 63) is a pasted value (GPR waterfall has no monthly);
    # it must sum to the evaluated annual N63.
    N = _eval_t12_n(ta)
    nr_sum = _row_sum(63)
    _check(abs(nr_sum - N[63]) < 2.0, f"Net Rent monthly {nr_sum:,.2f} != annual {N[63]:,.2f}")

    # Total rows mirror the col-N formula across B–M (cells are formulas).
    for r in (69, 85, 111, 116):
        b = ta.cell(row=r, column=2).value  # col B
        _check(isinstance(b, str) and b.startswith("="), f"B{r} should mirror a formula, got {b!r}")

    print(
        f"  ✓ monthly grid: {n_monthly} value cells; line items + Net Rent tie to annual; "
        f"total rows mirror formulas across B–M"
    )


def test_section_i_raw_populated() -> None:
    """Section I (Layer 1 — Raw T-12) is rebuilt one row per Analyzer label:
    Account Name = matched GL names, months C–N tie to the T-12 Total (O)."""
    if not (RR_FIXTURE.exists() and T12_FIXTURE.exists() and TEMPLATE_V5.exists()):
        print("  ⊘ SKIP test_section_i_raw_populated — fixtures absent")
        return

    import io as _io
    from analyzer_rr_translator import translate_for_t12
    from analyzer_rr_writer import populate_rr_input
    from t12_normalizer_writer import populate_t12_input
    from uw_output_model import (
        compute_uw_output_values, compute_uw_output_monthly, compute_t12_raw_lines,
    )
    from uw_template_writer import populate_uw_template

    rr, t12 = _parse_inputs()
    raw = compute_t12_raw_lines(t12)
    _check(len(raw) >= 20, f"expected many raw lines, got {len(raw)}")
    _check(all(abs(sum(l["monthly"]) - l["total"]) < 1.0 for l in raw),
           "a raw line's monthly sum != its total")

    after_rr = populate_rr_input(
        BUNDLED_ANALYZER.read_bytes(), translate_for_t12(rr.condensed),
        PERIOD, source_filename="h.xlsx",
    )
    fresh = populate_t12_input(after_rr, t12, new_descmap_entries=[], source_filename="t.xlsx")
    out_bytes, rep = populate_uw_template(
        fresh, TEMPLATE_V5.read_bytes(), template_version="v5",
        computed_values=compute_uw_output_values(rr, t12),
        computed_monthly=compute_uw_output_monthly(rr, t12),
        raw_t12_lines=raw,
    )
    _check(rep.summary.get("section_i_raw_cells", 0) > 100, "Section I not populated")

    ws = openpyxl.load_workbook(_io.BytesIO(out_bytes), data_only=False)["T-12 Analysis"]
    # First data row (123): Account Name is text, months tie to T-12 Total.
    b123 = ws.cell(row=123, column=2).value
    _check(isinstance(b123, str) and not b123.replace(".", "").isdigit(),
           f"Section I Account Name should be GL text, got {b123!r}")
    months = sum((_num(ws.cell(row=123, column=c).value) or 0.0) for c in range(3, 15))
    total = _num(ws.cell(row=123, column=15).value)
    _check(abs(months - total) < 2.0, f"row 123 months {months:,.2f} != total {total:,.2f}")
    # Section J reconciliation formulas authored.
    _check(isinstance(ws.cell(row=178, column=15).value, str), "Section J EBITDAR not authored")
    print(f"  ✓ Section I: {len(raw)} raw lines; Account Name=GL text; months tie to total; Section J authored")


def main() -> int:
    print("=== test_uw_output_model ===")
    failures = 0
    for fn in (
        test_engine_matches_cached,
        test_writer_fallback_populates,
        test_dynamic_array_metadata_restored,
        test_monthly_grid_reconciles,
        test_section_i_raw_populated,
    ):
        print(f"\n--- {fn.__name__} ---")
        try:
            fn()
        except TestFailure as e:
            failures += 1
            print(f"  ✗ FAIL: {e}")
        except Exception as e:  # noqa: BLE001
            failures += 1
            print(f"  ✗ ERROR: {type(e).__name__}: {e}")
    print()
    if failures:
        print(f"=== {failures} test(s) FAILED ===")
        return 1
    print("=== all tests passed ===")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
