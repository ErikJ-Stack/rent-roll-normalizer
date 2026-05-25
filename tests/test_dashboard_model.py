"""
Track 5 dashboard regression test.

Validates that `dashboard_model.compute_dashboard()` produces values matching
the bundled Analyzer's `data_only=True` Dashboard cells when fed reconstructed
NormalizeResult + T12ParseResult inputs derived from the same populated
workbook.

Fixture
-------
Expects a populated Analyzer at `Sample Files/dashboard/regression_v0211.xlsx`
(gitignored — contains real operator financials per repo convention). The
fixture used during initial development was a Homestead Village populated
Analyzer (RR 2026-04-24 + March 2026 T12). If the fixture is absent, the
test skips with a clear message.

Known xlsx divergences
----------------------
The Dashboard headline tiles `B6` (occupancy), `F20` (blended ADR), and `K6`
(blended RevPOR) reference single-care-type cells (`T12 Analytics!F134/F140/F143`,
each of which is AL- or MC-specific) while being labeled as blended/community
on the Dashboard. The Python model computes structurally-correct blended
values, so these three metrics do NOT match the xlsx in the regression test —
they're whitelisted. A substrate-side fix is tracked as a spawned task.
"""

from __future__ import annotations

import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from dashboard_model import compute_dashboard
from normalizer import NormalizeResult
from t12_normalizer import GLRow, T12ParseResult


REPO_ROOT = Path(__file__).resolve().parent.parent
FIXTURE_PATH = REPO_ROOT / "Sample Files" / "dashboard" / "regression_v0211.xlsx"

PURCHASE_PRICE_HOMESTEAD = 18_000_000  # baked into the Homestead fixture's T12 Analytics!E117


def _reconstruct_inputs(populated_path: Path) -> tuple[NormalizeResult, T12ParseResult]:
    """Read populated Analyzer's Rent Roll Input + T12 Input → fake parser outputs."""
    wb = load_workbook(populated_path, data_only=True)

    rri = wb["Rent Roll Input"]
    headers = []
    for c in range(1, rri.max_column + 1):
        h = rri.cell(4, c).value
        headers.append(str(h).replace("\n", " ").strip() if h else None)
    rows = []
    for r in range(7, rri.max_row + 1):
        row = {h: rri.cell(r, c).value for c, h in enumerate(headers, 1) if h}
        if row.get("Unit #") or row.get("Resident Name"):
            rows.append(row)
    rr_df = pd.DataFrame(rows)
    nr = NormalizeResult(
        normalized=rr_df, condensed=rr_df, mapping_audit=pd.DataFrame(),
        source_headers=list(rr_df.columns), header_row_idx=0,
        care_groups=[], unmapped={},
    )

    ti = wb["T12 Input"]
    month_labels = [str(ti.cell(11, c).value or "") for c in range(3, 15)]
    gl_rows: list[GLRow] = []
    for r in range(12, ti.max_row + 1):
        desc = ti.cell(r, 2).value
        if not desc:
            continue
        monthly = [
            float(v) if isinstance(v, (int, float)) else 0.0
            for v in (ti.cell(r, c).value for c in range(3, 15))
        ]
        total_v = ti.cell(r, 15).value
        total = float(total_v) if isinstance(total_v, (int, float)) else sum(monthly)
        try:
            gl_rows.append(GLRow(account="", description=str(desc).strip(), monthly=monthly, total=total))
        except ValueError:
            continue
    t12 = T12ParseResult(
        gl_rows=gl_rows, month_labels=month_labels, unmatched=[],
        format_name="reconstructed", sheet_name="T12 Input",
        populated_months=12, was_annualized=False,
    )
    return nr, t12


@unittest.skipUnless(FIXTURE_PATH.exists(), f"fixture absent: {FIXTURE_PATH}")
class TestDashboardModelRegression(unittest.TestCase):

    @classmethod
    def setUpClass(cls) -> None:
        cls.nr, cls.t12 = _reconstruct_inputs(FIXTURE_PATH)
        cls.model = compute_dashboard(
            rr_result=cls.nr,
            t12_result=cls.t12,
            property_name="Homestead Village",
            period_label=cls.t12.month_labels[-1],
            purchase_price=PURCHASE_PRICE_HOMESTEAD,
        )
        cls.wb = load_workbook(FIXTURE_PATH, data_only=True)
        cls.dash = cls.wb["Dashboard"]

    # ── Capacity (exact match) ──────────────────────────────────────────────

    def test_licensed_total(self):
        self.assertEqual(self.model.licensed_total, self.dash["C18"].value)

    def test_occupied_total(self):
        self.assertEqual(self.model.occupied_total, self.dash["C20"].value)

    # ── Margins / ratios (penny match) ──────────────────────────────────────

    def _assert_close(self, py: float, xl: float, name: str, rel_tol: float = 0.005):
        self.assertIsNotNone(py, f"{name} should not be None")
        self.assertIsNotNone(xl, f"{name} xlsx value should not be None")
        rel = abs(py - xl) / abs(xl) if xl else 0
        self.assertLess(rel, rel_tol, f"{name} drift {rel*100:.3f}%: py={py} xl={xl}")

    def test_egi(self):
        # B11 is TEXT-formatted; read raw EGI from F19 (uses F52)
        self._assert_close(self.model.egi, self.dash["F19"].value, "egi")

    def test_ebitdar(self):
        # E11 is TEXT-formatted; read raw EBITDAR from I20 (=F163 raw)
        self._assert_close(self.model.ebitdar, self.dash["I20"].value, "ebitdar")

    def test_ebitdarm(self):
        self._assert_close(self.model.ebitdarm, self.dash["I18"].value, "ebitdarm")

    def test_total_opex(self):
        self._assert_close(self.model.total_opex, self.dash["I24"].value, "total_opex")

    def test_total_labor_pct(self):
        self._assert_close(self.model.total_labor_pct, self.dash["L18"].value, "total_labor_pct")

    def test_direct_labor_pct(self):
        self._assert_close(self.model.direct_labor_pct, self.dash["L19"].value, "direct_labor_pct")

    def test_overtime_pct(self):
        self._assert_close(self.model.overtime_pct, self.dash["L20"].value, "overtime_pct")

    def test_agency_pct(self):
        self._assert_close(self.model.agency_pct, self.dash["L21"].value, "agency_pct")

    def test_mgmt_fee_pct(self):
        self._assert_close(self.model.mgmt_fee_pct, self.dash["L23"].value, "mgmt_fee_pct")

    def test_insurance_pct(self):
        self._assert_close(self.model.insurance_pct, self.dash["L24"].value, "insurance_pct")

    def test_bad_debt_pct(self):
        self._assert_close(self.model.bad_debt_pct, self.dash["F24"].value, "bad_debt_pct")

    def test_ebitdar_margin(self):
        self._assert_close(self.model.ebitdar_margin, self.dash["I21"].value, "ebitdar_margin")

    def test_ebitdarm_margin(self):
        # E6 displays the same number as F162 (cached); use I19 (formula F162 raw)
        self._assert_close(self.model.ebitdarm_margin, self.dash["I19"].value, "ebitdarm_margin")

    def test_egi_per_occupied_bed(self):
        self._assert_close(self.model.egi_per_occupied_bed, self.dash["I22"].value, "egi_per_occ")

    def test_opex_per_occupied_bed(self):
        self._assert_close(self.model.opex_per_occupied_bed, self.dash["I23"].value, "opex_per_occ")

    def test_food_ppd(self):
        self._assert_close(self.model.food_ppd, self.dash["L22"].value, "food_ppd")

    def test_purchase_price(self):
        self.assertEqual(self.model.purchase_price, self.dash["D30"].value)

    def test_price_per_bed(self):
        self._assert_close(self.model.price_per_bed, self.dash["D31"].value, "price_per_bed")

    def test_going_in_cap_ebitdarm(self):
        self._assert_close(self.model.going_in_cap, self.dash["D32"].value, "going_in_cap")

    def test_ebitdar_cap(self):
        self._assert_close(self.model.ebitdar_cap, self.dash["D33"].value, "ebitdar_cap")

    # ── Payer mix counts (exact) ───────────────────────────────────────────

    def test_payer_counts(self):
        for i, row in enumerate(self.model.payer_mix):
            xl_cnt = self.dash.cell(43 + i, 3).value or 0
            self.assertEqual(row.resident_count, xl_cnt, f"{row.payer} count")

    def test_payer_revenue_pct(self):
        for i, row in enumerate(self.model.payer_mix):
            xl_pct = self.dash.cell(43 + i, 6).value or 0
            py_pct = row.revenue_pct or 0
            self.assertLess(abs(py_pct - xl_pct), 0.02,
                            f"{row.payer} revenue% drift")

    # ── Monthly EGI series (exact within float tolerance) ──────────────────

    def test_monthly_egi(self):
        self.assertEqual(len(self.model.monthly_egi), 12)
        for i, ms in enumerate(self.model.monthly_egi):
            xl = self.dash.cell(97 + i, 3).value
            self.assertIsNotNone(xl, f"month {i} xlsx value missing")
            self._assert_close(ms.egi, xl, f"monthly_egi[{i}] {ms.month_label}", rel_tol=0.001)

    # ── Known divergences from xlsx (Python correct, xlsx cross-ref bug) ────

    def test_known_divergence_blended_occupancy(self):
        """Dashboard!B6 pulls F134 (=C11/C6 = AL-only). Python computes blended."""
        py = self.model.occupancy_pct
        self.assertIsNotNone(py)
        # py is the *correct* blended occupancy — should differ from xlsx B6 by
        # the AL/blended gap. Just assert py is sensible.
        self.assertGreater(py, 0.5)
        self.assertLess(py, 1.0)

    def test_known_divergence_blended_revpor(self):
        """Dashboard!K6 pulls F143 (= MC-specific). Python computes blended."""
        self.assertIsNotNone(self.model.revpor)
        self.assertGreater(self.model.revpor, 1000)


if __name__ == "__main__":
    unittest.main()
