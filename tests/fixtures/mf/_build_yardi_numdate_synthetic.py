"""
Authoring script for `yardi_numdate_synthetic.xlsx` — a committed, synthetic
MF T-12 fixture that exercises the two Yardi-export wrinkles the real Verona at
Silver Hill T-12 surfaced (MF v0.5.2), without shipping real financials:

  1. **Numeric date-string headers** — the "Month Ending" row renders the
     period as text "MM/DD/YYYY" (e.g. "03/25/2025"), not Excel dates or the
     "Mar 2025" style. The header detector must recognize these.
  2. **Combined "ACCT - Name" cells** — account number and name share one col-A
     cell ("41000 - Market Rent"), so the parser must split the leading account
     number out for acct-bearing-chart logic to work.

It also includes four of the newly-seeded COA accounts (43130 / 52025 / 54130)
so a CI run proves the seed additions classify.

Run from repo root:  python tests/fixtures/mf/_build_yardi_numdate_synthetic.py
"""
from pathlib import Path

import openpyxl

OUT = Path(__file__).resolve().parent / "yardi_numdate_synthetic.xlsx"

# 12 "month ending" dates as TEXT strings (the wrinkle under test).
MONTH_ENDING = [
    "03/25/2025", "04/25/2025", "05/25/2025", "06/25/2025", "07/25/2025",
    "08/25/2025", "09/25/2025", "10/25/2025", "11/25/2025", "12/31/2025",
    "01/25/2026", "02/25/2026",
]

# (col-A label, per-month amount). Combined "ACCT - Name" cells throughout.
# Leaves only (rollup/header rows added separately so we test filtering too).
INCOME = [
    ("      41000 - Market Rent", 100_000.0),          # Gross Potential Rent
    ("      41100 - Vacancy Loss", -8_000.0),          # Vacancy Loss
    ("      43130 - Interest Paid - Security Deposits", -200.0),  # NEW -> Misc Other Income
]
EXPENSE = [
    ("  61030 - Management Fees", 5_000.0),            # Management Fee
    ("  52025 - Backflow Inspections", 250.0),         # NEW -> Contract Services
    ("  54130 - Signage", 100.0),                      # NEW -> Leasing & Marketing
]
EXCLUDED = [
    ("  71020 - Carpet Replacement", 1_000.0),         # 70000-89999 -> EXCLUDED
]


def _annual(amt):
    return amt * 12


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trailing Twelve Months - Detail"

    ws["A1"] = "Synthetic Yardi T-12 (numeric-date + combined-acct) — TEST FIXTURE"
    ws["A4"] = "As of Date:"
    ws["B4"] = "02/25/2026"

    # Header band: row 7 "Month Ending", row 8 numeric dates, row 9 Actual/Summary
    for i, _ in enumerate(MONTH_ENDING):
        ws.cell(7, 2 + i, "Month Ending")
        ws.cell(8, 2 + i, MONTH_ENDING[i])
        ws.cell(9, 2 + i, "Actual")
    ws.cell(9, 14, "Summary")           # total column header (col N)

    r = 10

    def write_leaf(label, amt):
        nonlocal r
        ws.cell(r, 1, label)
        for i in range(12):
            ws.cell(r, 2 + i, amt)
        ws.cell(r, 14, _annual(amt))    # N = annual total
        r += 1

    def write_rollup(label, total):
        nonlocal r
        ws.cell(r, 1, label)
        ws.cell(r, 14, total)
        r += 1

    ws.cell(r, 1, "INCOME"); r += 1
    inc_total = 0.0
    for label, amt in INCOME:
        write_leaf(label, amt); inc_total += _annual(amt)
    write_rollup("Total Income", inc_total)
    r += 1

    ws.cell(r, 1, "EXPENSES"); r += 1
    exp_total = 0.0
    for label, amt in EXPENSE:
        write_leaf(label, amt); exp_total += _annual(amt)
    write_rollup("Total Operating Expense", exp_total)
    r += 1

    write_rollup("Net Operating Income", inc_total - exp_total)
    r += 2

    ws.cell(r, 1, "BELOW THE LINE"); r += 1
    for label, amt in EXCLUDED:
        write_leaf(label, amt)

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  income leaves total  = {inc_total:,.0f}")
    print(f"  expense leaves total = {exp_total:,.0f}")
    print(f"  NOI                  = {inc_total - exp_total:,.0f}")


if __name__ == "__main__":
    main()
