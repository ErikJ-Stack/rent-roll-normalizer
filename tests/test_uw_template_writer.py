"""
Smoke tests for `uw_template_writer.populate_uw_template`.

These run without pytest if you invoke the file directly:
    python tests/test_uw_template_writer.py

Two test layers:

  1. **Empty-Analyzer smoke** — uses the bundled `ALF_Financial_Analyzer_Only.xlsx`
     (substrate v0.2.14, no deal data). Verifies the writer doesn't crash,
     skip-status rules are honored, the PopulateReport has the expected
     shape, and the produced bytes round-trip via openpyxl.

  2. **Populated-Analyzer end-to-end** (skipped if fixture missing) — uses
     `Sample Files/Analyzer with 2026-04-24 Homestead Village Rent Roll v2
     + March 2026 T12 2026-04-24.xlsx`. Spot-checks a few high-confidence
     concepts (EGI, GPR, EBITDARM) and confirms the rent_roll path wrote
     a column-stride of values starting at row 211.

The populated fixture is property-specific data and is intentionally
gitignored. Skip cleanly when absent.
"""
from __future__ import annotations

import io
import json
from pathlib import Path

import openpyxl

from uw_template_writer import (
    PopulateReport,
    populate_uw_template,
    _SPECIAL_SKIP_KEYS,
)

ROOT = Path(__file__).resolve().parent.parent
BUNDLED_ANALYZER = ROOT / "ALF_Financial_Analyzer_Only.xlsx"
# v4 lives in Sample Files (gitignored — working-copy reference only).
TEMPLATE_V4 = ROOT / "Sample Files" / "ALF_UW_Template_v4.xlsx"
# v5 + v6 both live at committed asset paths so CI / cold checkouts can run the
# smoke tests without needing the gitignored Sample Files copy. v6 (T-12 income
# restructure) is the binding default as of 2026-05-29 (UWT v0.8.0); v5 is kept
# as an explicit-version regression.
TEMPLATE_V5 = ROOT / "assets" / "ALF_UW_Template_v5.xlsx"
TEMPLATE_V6 = ROOT / "assets" / "ALF_UW_Template_v6.xlsx"
TEMPLATE = TEMPLATE_V5  # v5 regression tests below pin template_version="v5"
POPULATED_ANALYZER = (
    ROOT / "Sample Files"
    / "Analyzer with 2026-04-24 Homestead Village Rent Roll v2 + March 2026 T12 2026-04-24.xlsx"
)
REGISTRY = ROOT / "tools" / "uw_template" / "registry.json"


# ── small assert helpers (avoid pytest dep) ───────────────────────────────────

class TestFailure(AssertionError):
    pass


def _check(cond: bool, msg: str) -> None:
    if not cond:
        raise TestFailure(msg)


def _section(title: str) -> None:
    print(f"\n--- {title} ---")


# ── test 1: empty-Analyzer smoke ──────────────────────────────────────────────

def test_empty_analyzer_smoke() -> None:
    _section("test_empty_analyzer_smoke (v5 regression — pinned)")

    _check(BUNDLED_ANALYZER.exists(), f"missing bundled Analyzer: {BUNDLED_ANALYZER}")
    _check(TEMPLATE.exists(), f"missing template: {TEMPLATE}")

    analyzer_bytes = BUNDLED_ANALYZER.read_bytes()
    template_bytes = TEMPLATE.read_bytes()

    populated, report = populate_uw_template(
        analyzer_bytes, template_bytes,
        template_version="v5",  # pinned — default is now v6
        scenario="normalized",
    )

    _check(isinstance(populated, bytes), "populated should be bytes")
    _check(len(populated) > 0, "populated bytes are empty")
    _check(isinstance(report, PopulateReport), "report wrong type")

    # report summary keys
    summary = report.summary
    _check("total_concepts" in summary, "summary missing total_concepts")
    _check(summary["total_concepts"] == 137, f"expected 137 concepts (registry v0.5.0 — v6 absorption added 14), got {summary['total_concepts']}")
    _check("cells_written" in summary, "summary missing cells_written")

    by = report.by_outcome()

    # gap_target and gap_source concepts must be skipped, never written
    skipped = {r.key for r in by.get("skipped", [])}
    for r in report.results:
        if r.status in ("gap_target", "gap_source", "header_only", "derived", "manual",
                        "substrate_ready_parser_pending", "decided_pending_upstream"):
            _check(
                r.outcome in ("skipped", "no_source", "no_target"),
                f"concept {r.key!r} (status={r.status}) should be skipped, got {r.outcome}",
            )

    # opex_bad_debt_expense must be in _SPECIAL_SKIP_KEYS and must have been skipped
    bd = next((r for r in report.results if r.key == "opex_bad_debt_expense"), None)
    _check(bd is not None, "opex_bad_debt_expense missing from report")
    _check(bd.outcome == "skipped", f"opex_bad_debt_expense not skipped: {bd.outcome}")
    _check(
        "Bad Debt Expense" in bd.notes and "double-count" in bd.notes,
        f"opex_bad_debt_expense skip note unexpected: {bd.notes!r}",
    )

    # Round-trip the populated bytes back through openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(populated), data_only=False)
    _check("Prop Info" in wb.sheetnames, "Prop Info sheet missing from output")
    _check("T-12 Analysis" in wb.sheetnames, "T-12 Analysis sheet missing from output")
    _check("Rent Roll Analysis" in wb.sheetnames, "Rent Roll Analysis sheet missing")
    # v5 dropped "Additional Fees" — sheet count 17 → 16
    _check(len(wb.sheetnames) == 16, f"expected 16 sheets (v5), got {len(wb.sheetnames)}")
    _check(
        "Additional Fees" not in wb.sheetnames,
        "v5 should not have 'Additional Fees' sheet",
    )

    # v5 has the row 210 header band — no A210 warning expected on v5.
    a210_warning = any("A210 is blank" in w for w in report.warnings)
    _check(
        not a210_warning,
        f"v5 unexpectedly emitted A210-blank warning: {report.warnings}",
    )

    # Bundled Analyzer is empty — t12-path concepts read None from UW Output,
    # so the writer should mostly emit no_source. A small number of concepts
    # may write if Cover!B5 has been auto-filled (it's a formula in v0.2.8+;
    # resolves to "" when both A3 and A10 are empty — that's _is_blank,
    # so writer skips).
    no_source_count = len(by.get("no_source", []))
    _check(no_source_count > 50, f"expected most concepts to be no_source on empty analyzer, got {no_source_count}")

    print(f"  ✓ {len(report.results)} concepts processed")
    print(f"  ✓ {summary.get('cells_written', 0)} cells written on empty Analyzer")
    print(f"  ✓ outcomes: {dict(summary)}")
    if report.warnings:
        print(f"  ! warnings: {len(report.warnings)}")
        for w in report.warnings:
            print(f"    - {w[:90]}")


# ── test 2: populated-Analyzer end-to-end ─────────────────────────────────────

def test_populated_analyzer_e2e() -> None:
    _section("test_populated_analyzer_e2e")

    if not POPULATED_ANALYZER.exists():
        print(f"  SKIP: populated fixture not found at {POPULATED_ANALYZER}")
        print(f"  (this is fine — fixture is gitignored)")
        return

    analyzer_bytes = POPULATED_ANALYZER.read_bytes()
    template_bytes = TEMPLATE.read_bytes()

    populated, report = populate_uw_template(
        analyzer_bytes, template_bytes,
        template_version="v5",  # pinned — default is now v6
        scenario="normalized",
    )

    summary = report.summary
    written_count = summary.get("written", 0)
    cells_written = summary.get("cells_written", 0)

    # v5 expectations: 95 mapped concepts (was 88 on v4), with the
    # Homestead fixture populating most of them. Cells written should
    # grow by hundreds because rent_roll path now also writes cols
    # AP (Care Level Tier) and AR (Preleased Date) at 176 rows each.
    _check(written_count > 35, f"expected > 35 written concepts on v5, got {written_count}")
    _check(cells_written > 100, f"expected > 100 cells written, got {cells_written}")

    # Round-trip and spot-check a few high-confidence cells
    wb_out = openpyxl.load_workbook(io.BytesIO(populated), data_only=False)

    # EGI (normalized) → T-12 Analysis!N69. As of UWT v0.6.3 the Layer-3 TOTAL
    # rows are preserved as live template formulas (the writer no longer pastes
    # values over them), so N69 is a formula string regardless of source.
    ws_t12 = wb_out["T-12 Analysis"]
    egi_cell = ws_t12["N69"].value
    print(f"  T-12 Analysis!N69 (EGI normalized) = {egi_cell!r}")
    _check(
        isinstance(egi_cell, str) and egi_cell.startswith("="),
        f"EGI total should be a preserved formula, got {type(egi_cell).__name__}: {egi_cell!r}",
    )

    # v5: EBITDARM shifted from N115 (v4) to N116
    ebitdarm = ws_t12["N116"].value
    print(f"  T-12 Analysis!N116 (EBITDARM, v5 location) = {ebitdarm!r}")

    # v5: new row N118 EBITDA
    ebitda = ws_t12["N118"].value
    print(f"  T-12 Analysis!N118 (EBITDA, NEW in v5) = {ebitda!r}")

    # v5: new row N115 Total OpEx excl. mgmt
    opex_excl = ws_t12["N115"].value
    print(f"  T-12 Analysis!N115 (Total OpEx excl. mgmt, NEW in v5) = {opex_excl!r}")

    # GPR → T-12 Analysis!N58
    gpr = ws_t12["N58"].value
    print(f"  T-12 Analysis!N58 (GPR) = {gpr!r}")

    # Rent roll path — check row 211 (first data row)
    ws_rr = wb_out["Rent Roll Analysis"]
    print(f"  Rent Roll Analysis row 211 sample:")
    for col in ("A", "B", "C", "D", "E"):
        v = ws_rr[f"{col}211"].value
        print(f"    {col}211 = {v!r}")

    # Count populated rows on the rent roll target
    populated_rows = sum(
        1 for r in range(211, 811) if ws_rr[f"A{r}"].value is not None
    )
    print(f"  Rent Roll Analysis populated rows starting at 211: {populated_rows}")
    _check(populated_rows > 0, "rent roll target has no populated rows")

    # Property name should have landed on Prop Info!B4 (from named range)
    ws_pi = wb_out["Prop Info"]
    prop_name = ws_pi["B4"].value
    print(f"  Prop Info!B4 (Property Name) = {prop_name!r}")

    # v5: Occupied Beds rows added at Prop Info!B20-B22
    print(f"  Prop Info!B20 (Occupied Beds — IL, NEW in v5) = {ws_pi['B20'].value!r}")
    print(f"  Prop Info!B21 (Occupied Beds — AL, NEW in v5) = {ws_pi['B21'].value!r}")
    print(f"  Prop Info!B22 (Occupied Beds — MC, NEW in v5) = {ws_pi['B22'].value!r}")

    # v5.1: column restructure shifted these cols left by 1 (closing AC hole):
    #   Care Level Tier:    AP → AO
    #   Total Ancillary $:  AQ → AP (template formula =SUM(AJ:AN))
    #   Preleased Date:     AR → AQ
    #   ACH:                AS → AR
    # AP (Total Ancillary) should NOT be written — derived in template.
    print(f"  Rent Roll Analysis row 211 v5.1 column positions:")
    expected = {
        "D":  ("Unit Type",         "1 Bedroom (writer paste from Analyzer col F)"),
        "E":  ("Status",            "Occupied (right-shifted from D)"),
        "AO": ("Care Level Tier",   "left-shifted from AP"),
        "AP": ("Total Ancillary $", "left-shifted from AQ — template formula"),
        "AQ": ("Preleased Date",    "left-shifted from AR — None for non-preleased"),
        "AR": ("ACH",               "left-shifted from AS"),
        "AU": ("Effective Conc $",  "last col, left-shifted from AV"),
    }
    for col, (label, why) in expected.items():
        v = ws_rr[f"{col}211"].value
        print(f"    {col}211 ({label}): {v!r}  — {why}")

    # Spot assertions
    _check(
        ws_rr["D211"].value == "1 Bedroom",
        f"D211 should = '1 Bedroom' (writer paste from Analyzer col F); got {ws_rr['D211'].value!r}",
    )
    _check(
        ws_rr["E211"].value == "Occupied",
        f"E211 should = 'Occupied' (right-shifted from D); got {ws_rr['E211'].value!r}",
    )
    # AP should still hold the Total Ancillary template formula (left-shifted from AQ)
    ap211 = ws_rr["AP211"].value
    _check(
        isinstance(ap211, str) and ap211.startswith("=SUM"),
        f"AP211 should hold Total Ancillary template formula =SUM(...) (left-shifted from v5's AQ); got {ap211!r}",
    )

    # Print outcome rollup
    print(f"  ✓ Summary: {dict(summary)}")

    # Surface the first few written t12-path concepts
    written = [r for r in report.results if r.outcome == "written"]
    t12_written = [r for r in written if r.path == "t12"][:5]
    print(f"  First t12 cells written:")
    for r in t12_written:
        print(f"    {r.target_address:35s} ← {r.key:30s} sample={r.sample_value!r}")

    rr_written = [r for r in written if r.path == "rent_roll"][:5]
    print(f"  First rent_roll concepts written:")
    for r in rr_written:
        print(f"    {r.target_address:35s} ← {r.key:30s} ({r.cells_written} cells) sample={r.sample_value!r}")


# ── test 3: v6 empty-Analyzer smoke (new default) ─────────────────────────────

def test_empty_analyzer_smoke_v6() -> None:
    _section("test_empty_analyzer_smoke_v6 (v6 — new default)")

    _check(BUNDLED_ANALYZER.exists(), f"missing bundled Analyzer: {BUNDLED_ANALYZER}")
    _check(TEMPLATE_V6.exists(), f"missing v6 template: {TEMPLATE_V6}")

    populated, report = populate_uw_template(
        BUNDLED_ANALYZER.read_bytes(), TEMPLATE_V6.read_bytes(),
        # template_version omitted → exercises the new v6 default
        scenario="normalized",
    )

    _check(report.template_version == "v6", f"expected v6 default, got {report.template_version!r}")
    _check(isinstance(populated, bytes) and len(populated) > 0, "populated bytes empty")
    _check(report.summary["total_concepts"] == 137, f"expected 137 concepts, got {report.summary['total_concepts']}")

    wb = openpyxl.load_workbook(io.BytesIO(populated), data_only=False)
    _check(len(wb.sheetnames) == 16, f"expected 16 sheets (v6), got {len(wb.sheetnames)}")
    _check("T-12 Analysis" in wb.sheetnames, "T-12 Analysis missing")
    ws = wb["T-12 Analysis"]

    # v6 income layout: EGI is a preserved SUM formula at N77 (not v5's N69).
    egi = ws["N77"].value
    _check(
        isinstance(egi, str) and egi.startswith("="),
        f"v6 EGI should be a preserved formula at N77, got {egi!r}",
    )
    # EBITDAR (N132) + EBITDA (N133) are authored by the finalize pass
    # (template ships N132 as literal 0 / N133 blank).
    ebitdar = ws["N132"].value
    _check(
        isinstance(ebitdar, str) and ebitdar.startswith("="),
        f"v6 EBITDAR (N132) should be authored as a formula, got {ebitdar!r}",
    )
    ebitda = ws["N133"].value
    _check(
        isinstance(ebitda, str) and ebitda.startswith("="),
        f"v6 EBITDA (N133) should be authored as a formula, got {ebitda!r}",
    )
    # Monthly mirror: B77 (EGI, col B) should carry the mirrored formula.
    b77 = ws["B77"].value
    _check(
        isinstance(b77, str) and b77.startswith("="),
        f"v6 EGI monthly mirror at B77 should be a formula, got {b77!r}",
    )
    _check(report.summary.get("t12_totals_finalized", 0) > 0, "v6 finalize wrote nothing")

    print(f"  ✓ v6 default: {report.summary['total_concepts']} concepts, {len(wb.sheetnames)} sheets")
    print(f"  ✓ N77 (EGI)={egi!r}  N132 (EBITDAR)={ebitdar!r}  N133 (EBITDA)={ebitda!r}")
    print(f"  ✓ B77 monthly mirror={b77!r}")
    print(f"  ✓ outcomes: {dict(report.summary)}")


# ── test 4: v6 populated-Analyzer end-to-end ──────────────────────────────────

def test_populated_analyzer_e2e_v6() -> None:
    _section("test_populated_analyzer_e2e_v6")

    if not POPULATED_ANALYZER.exists():
        print(f"  SKIP: populated fixture not found at {POPULATED_ANALYZER}")
        return

    populated, report = populate_uw_template(
        POPULATED_ANALYZER.read_bytes(), TEMPLATE_V6.read_bytes(),
        scenario="normalized",
    )
    _check(report.template_version == "v6", f"expected v6, got {report.template_version!r}")

    wb_out = openpyxl.load_workbook(io.BytesIO(populated), data_only=False)
    ws_t12 = wb_out["T-12 Analysis"]
    for addr, label in (("N77", "EGI"), ("N131", "EBITDARM"), ("N132", "EBITDAR"), ("N133", "EBITDA")):
        v = ws_t12[addr].value
        print(f"  T-12 Analysis!{addr} ({label}) = {v!r}")
        _check(
            isinstance(v, str) and v.startswith("="),
            f"v6 {label} at {addr} should be a formula, got {v!r}",
        )

    # Rent Roll Analysis is unchanged v5→v6 — first data row should populate.
    ws_rr = wb_out["Rent Roll Analysis"]
    _check(
        ws_rr["D211"].value == "1 Bedroom",
        f"D211 should = '1 Bedroom'; got {ws_rr['D211'].value!r}",
    )
    populated_rows = sum(1 for r in range(211, 811) if ws_rr[f"A{r}"].value is not None)
    print(f"  Rent Roll Analysis populated rows from 211: {populated_rows}")
    _check(populated_rows > 0, "rent roll target has no populated rows")
    print(f"  ✓ Summary: {dict(report.summary)}")


# ── runner ────────────────────────────────────────────────────────────────────

def main() -> int:
    failures: list[tuple[str, str]] = []
    for fn in (
        test_empty_analyzer_smoke,
        test_populated_analyzer_e2e,
        test_empty_analyzer_smoke_v6,
        test_populated_analyzer_e2e_v6,
    ):
        try:
            fn()
        except TestFailure as e:
            failures.append((fn.__name__, str(e)))
            print(f"  ✗ FAIL: {e}")
        except Exception as e:
            failures.append((fn.__name__, f"unhandled {type(e).__name__}: {e}"))
            print(f"  ✗ ERROR: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()

    print()
    if failures:
        print(f"=== {len(failures)} test(s) failed ===")
        for name, msg in failures:
            print(f"  ✗ {name}: {msg}")
        return 1
    print("=== all tests passed ===")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
