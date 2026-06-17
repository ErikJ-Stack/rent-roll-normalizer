"""
Rent Roll Normalization App — Streamlit entry point.

Run locally:
    streamlit run app.py

Deploy to Streamlit Cloud:
    Push this folder to a GitHub repo and connect at https://streamlit.io/cloud

This entry point orchestrates two normalizer modules that share a destination
workbook:
  - RR Normalizer (Track 1, see SPEC-RR.md) — writes to `Rent Roll Input`
  - T12 Normalizer (Track 2, see SPEC-T12.md) — writes to `T12 Input`

Analyzer template loading (v1.12.0):
  - The bundled `ALF_Financial_Analyzer_Only.xlsx` from the repo root is
    loaded silently as the default destination workbook on every run.
  - Users can override via the "Advanced — override Analyzer template"
    expander in the intake panel; uploaded files win when present.
  - The bundled file is the canonical source of `Description_Map` for
    UNMATCHED matching. Resolutions are baked into each download but do
    NOT propagate back to the repo; bundled-file edits go through git.

A single run can produce: standalone Normalized RR workbook + populated
Analyzer with both RR and T12 data, when both required uploads are present.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import os
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

from auth import allowed_modes, render_user_controls, require_login
from branding import inject_brand_css, inject_cockpit_css
from mappings import MappingSet, load_mapping_workbook
from normalizer import CONDENSED_COLUMNS, normalize_rent_roll
from period_date import detect_period_date
from property_name import derive_property_name
from reports import build_by_type, build_exceptions, build_summary
from t12_normalizer import (
    UnknownT12FormatError,
    parse_t12,
    read_descmap_descriptions,
)
from t12_normalizer_writer import (
    T12NormalizerCapacityError,
    populate_t12_input,
)
from analyzer_rr_translator import translate_for_t12
from analyzer_rr_writer import AnalyzerRRCapacityError, populate_rr_input
from ar_normalizer import parse_ar_file
from mf_t12_normalizer import parse_mf_t12
from mf_normalizer import parse_mf_rr
from mf_ar_parser import join_ar_to_units, parse_mf_ar
from mf_om_extractor import MFOMExtractorError, parse_mf_om
from mf_uw_model_writer import populate_mf_model
from mf_dashboard import compute_mf_dashboard, render_mf_dashboard
from ar_writer import AROutputError, populate_ar_collections
from dashboard_model import compute_dashboard
from dashboard_ui import render_dashboard
from uw_template_writer import (
    UWTemplateWriterError,
    populate_uw_template,
)
from uw_output_model import (
    compute_uw_output_values,
    compute_uw_output_monthly,
    compute_t12_raw_lines,
)
from writer import write_output


# ---------------------------------------------------------------------------
# Version constants — independent streams per SPEC-T12 §"How the version
# stream relates to Track 1"
# ---------------------------------------------------------------------------
APP_VERSION = "1.14.0"            # alias for RR_VERSION; kept for back-compat
APP_LAST_UPDATED = "2026-05-08"   # alias for RR_LAST_UPDATED

RR_VERSION = "1.19.0"
RR_LAST_UPDATED = "2026-05-27"

UWT_VERSION = "0.11.0"            # Track 4 — operator template v11 absorbed (registry 0.8.0, 196 concepts): RR Analysis paste grid re-anchored 214+ → 224+ (header 223; new Section S concessions-audit block; aggregates read $224:$623), fill-downs now cover full band (v8 >176-bed quirk fixed), T-12 Analysis unchanged. Paired with Analyzer substrate v0.3.0 (UW Output +1 row, T-12 Analytics +OTH col, Description_Map +155 GL descriptions). Bundled default v8 → v11.
UWT_LAST_UPDATED = "2026-06-16"

T12_VERSION = "0.2.1"
T12_LAST_UPDATED = "2026-05-11"

AR_VERSION = "0.1.0"
AR_LAST_UPDATED = "2026-05-23"

T5_VERSION = "0.1.10"             # Track 5 — Webapp Dashboard Surface
T5_LAST_UPDATED = "2026-05-25"

# Bundled Analyzer substrate (stamped at Cover!B8). Hand-maintained like the
# RR/T12 constants above — bump when the bundled workbook is updated. The
# runtime "Using Analyzer ..." line still shows the value detected from the
# file actually in use (handles uploaded overrides).
ANALYZER_SUBSTRATE_VERSION = "0.3.0"
ANALYZER_LAST_UPDATED = "2026-06-16"


# ---------------------------------------------------------------------------
# Bundled Analyzer — loaded silently from repo root by default
# ---------------------------------------------------------------------------
BUNDLED_ANALYZER_PATH = Path(__file__).parent / "ALF_Financial_Analyzer_Only.xlsx"

# Bundled UW Template — loaded silently from assets/ by default. Mirrors the
# Analyzer pattern: operator can override via the Advanced expander for a
# session-specific template (e.g. v8 / v6 / v5 / v4 for a legacy deal). The
# bundled file is the binding v11 default as of 2026-06-16 (UWT v0.11.0):
# operator's Excel-native v11 binary. v11 re-anchors the Rent Roll Analysis
# paste grid to header 223 / data 224+ (every RR aggregate reads $224:$623) —
# a new Section S concessions-audit block pushed the grid down 10 rows vs v8 —
# and extends the K/L/V/W fill-downs to the full band (the v8 >176-bed quirk is
# fixed). T-12 Analysis layout is unchanged vs v8/v6. v8/v6 retained at assets/
# for override use.
BUNDLED_UW_TEMPLATE_PATH = Path(__file__).parent / "assets" / "ALF_UW_Template_v11.xlsx"
BUNDLED_UW_TEMPLATE_VERSION = "v11"

# MF (multifamily) UW Model — committed reference template (Track 4-MF).
BUNDLED_MF_MODEL_PATH = Path(__file__).parent / "assets" / "MF_UW_Model_v15.xlsx"


# ---------------------------------------------------------------------------
# Description_Map dropdown options — sourced from the v0.1.5 substrate
# ---------------------------------------------------------------------------
# Section is bounded; CareType is bounded; Flag has 8 substrate values + blank;
# Label is the existing 55-vocabulary as of v0.1.5 (free-text override allowed
# but discouraged — see SPEC-T12 §"Closed Label vocabulary").
DESCMAP_SECTIONS = ["Revenue", "Labor", "Non-Labor", "Excluded"]
DESCMAP_CARETYPES = ["-", "IL", "AL", "MC"]
DESCMAP_FLAGS = [
    "",  # blank/None default
    "Volatile",
    "Normalize to $0",
    "Normalize at stabilization",
    "Stabilize annually",
    "Flag if >5% of wages",
    "Annualize",
    "Verify assessed value",
    "Normalize to 1-2%",
]


def _build_output_name(source_filename: str) -> str:
    """Build output filename: <source_stem> Normalized YYYY-MM-DD.xlsx"""
    import re
    stem = Path(source_filename).stem
    stem = re.sub(r"\s+Normalized\s+\d{4}-\d{2}-\d{2}\s*$", "", stem, flags=re.IGNORECASE)
    today = dt.date.today().isoformat()
    return f"{stem} Normalized {today}.xlsx"


def _read_descmap_labels(analyzer_bytes: bytes) -> list[str]:
    """Pull the existing Labels from the Analyzer's Description_Map for the
    matcher's Label combobox. Falls back to an empty list on any read error.
    """
    try:
        wb = openpyxl.load_workbook(pd.io.common.BytesIO(analyzer_bytes), data_only=True)
        ws = wb["Description_Map"]
        labels: set[str] = set()
        for r in range(5, ws.max_row + 1):
            v = ws.cell(r, 2).value  # col B = Label
            if v and str(v).strip():
                labels.add(str(v).strip())
        return sorted(labels)
    except Exception:
        return []


def _detect_substrate_version(analyzer_bytes: bytes) -> str:
    """Detect the substrate version of an Analyzer.

    Closes UW-BACKLOG BL-0008. The prior implementation (v1.12.0-v1.17.0)
    only knew the v0.1.4 and v0.1.5 Description_Map markers, so any
    Analyzer at v0.1.6 through v0.1.14 reported as `v0.1.5`. This rewrite
    reads `Cover!B8` (the canonical version stamp set by every migration
    since v0.1.4) as the primary source, then falls back to the legacy
    Label-based heuristic for older Analyzers that predate `Cover!B8` or
    for files where that cell has been damaged.

    Strategy:
        1. Try `Cover!B8`. If it matches `vN.N.N` (any leading-zero
           components allowed), return as-is.
        2. Fall back to substrate-distinctive sentinel cells (in order
           from newest to oldest):
             - Rent Roll Input!AH4 contains "Total" + "Ancillary" → v0.2.2+
             - T12 Raw Data!B16 == "Meal Income" → v0.2.1+
             - Workbook contains "UW Export" sheet → v0.2.0+
             - Rent Roll Recon!I87 contains "Actual" + "PSF" → v0.1.14+
             - T12 Analytics!A168 contains "Reconciliation" → v0.1.14+
             - Rent Roll Input!AC4 contains "Meal Plan" → v0.1.13+
             - Rent Roll Recon!A119 contains "M " → v0.1.12+
             - Rent Roll Input!V4 contains "2nd Person" → v0.1.10+
             - Description_Map contains "2nd Person Revenue" Label → v0.1.5+
             - Description_Map contains "Auto Expense" + "Lease / ground lease" → v0.1.4
        3. If nothing matches: `"pre-v0.1.4"`.

    Returns a string like `"v0.1.14"` or `"(unknown)"` on any read error.
    Used for the version caption only — never gates functionality.
    """
    try:
        wb = openpyxl.load_workbook(pd.io.common.BytesIO(analyzer_bytes), data_only=True)

        # 1. Primary: Cover!B8 stamp (canonical since v0.1.4). Pattern
        # widened in v1.17.3 to accept v0.2.x and any future major/minor —
        # the prior `v0\.1\.\d+` regex misreported v0.2.0 / v0.2.1 because
        # only the patch component was variable.
        try:
            cover_b8 = wb["Cover"]["B8"].value
            if isinstance(cover_b8, str):
                import re
                m = re.match(r"^v\d+\.\d+\.\d+$", cover_b8.strip())
                if m:
                    return cover_b8.strip()
        except Exception:
            pass

        # 2. Fallback: substrate-distinctive sentinel cells (newest → oldest)
        try:
            rri = wb["Rent Roll Input"]
            ah4 = rri.cell(4, 34).value
            if isinstance(ah4, str) and "Total" in ah4 and "Ancillary" in ah4:
                return "v0.2.2+"
        except Exception:
            pass
        try:
            trd = wb["T12 Raw Data"]
            b16 = trd.cell(16, 2).value
            if isinstance(b16, str) and b16.strip() == "Meal Income":
                return "v0.2.1+"
        except Exception:
            pass
        try:
            if "UW Export" in wb.sheetnames:
                return "v0.2.0+"
        except Exception:
            pass
        try:
            rr = wb["Rent Roll Recon"]
            i87 = rr.cell(87, 9).value
            if isinstance(i87, str) and "Actual" in i87 and "PSF" in i87:
                return "v0.1.14+"
        except Exception:
            pass
        try:
            ta = wb["T12 Analytics"]
            a168 = ta.cell(168, 1).value
            if isinstance(a168, str) and "Reconciliation" in a168:
                return "v0.1.14+"
        except Exception:
            pass
        try:
            rri = wb["Rent Roll Input"]
            ac4 = rri.cell(4, 29).value
            if isinstance(ac4, str) and "Meal Plan" in ac4:
                return "v0.1.13+"
        except Exception:
            pass
        try:
            rr = wb["Rent Roll Recon"]
            a119 = rr.cell(119, 1).value
            if isinstance(a119, str) and a119.startswith("M "):
                return "v0.1.12+"
        except Exception:
            pass
        try:
            rri = wb["Rent Roll Input"]
            v4 = rri.cell(4, 22).value
            if isinstance(v4, str) and "2nd Person" in v4:
                return "v0.1.10+"
        except Exception:
            pass

        # 3. Legacy Description_Map Label heuristic (pre-v0.1.10 fallback)
        ws = wb["Description_Map"]
        labels: set[str] = set()
        for r in range(5, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if v and str(v).strip():
                labels.add(str(v).strip())
        if "2nd Person Revenue" in labels:
            return "v0.1.5+"
        if "Auto Expense" in labels and "Lease / ground lease" in labels:
            return "v0.1.4"
        return "pre-v0.1.4"
    except Exception:
        return "(unknown)"


def _load_analyzer(uploaded_file) -> tuple[bytes, str, str]:
    """Resolve the Analyzer source — uploaded file wins over bundled default.

    Returns: (analyzer_bytes, source_label, substrate_version)
      - analyzer_bytes: the raw .xlsx bytes
      - source_label: "uploaded: <filename>" or "bundled (repo)"
      - substrate_version: detected version string (e.g., "v0.1.5")

    Raises FileNotFoundError if neither uploaded file nor bundled file exists.
    """
    if uploaded_file is not None:
        b = uploaded_file.getvalue()
        return b, f"uploaded: {getattr(uploaded_file, 'name', 'analyzer.xlsx')}", _detect_substrate_version(b)
    if BUNDLED_ANALYZER_PATH.exists():
        b = BUNDLED_ANALYZER_PATH.read_bytes()
        return b, "bundled (repo)", _detect_substrate_version(b)
    raise FileNotFoundError(
        f"Bundled Analyzer not found at {BUNDLED_ANALYZER_PATH}. "
        "Either restore the file in the repo root or upload a custom Analyzer "
        "via the Advanced expander in the intake panel."
    )


def _detect_uw_template_version(template_bytes: bytes) -> str:
    """Best-effort UW Template version detection.

    Probe order (newest first; each stage returns on a hit):

    1. **v11** — v11 re-anchored the RR Analysis paste grid: the operative
       "Unit/Bed" header moved to row 223 (data 224+), and a new
       "S. CONCESSIONS AUDIT" block sits at ~A205. Either marker → v11.
       (Must precede v8: v11 also carries the NER column.)
    2. **v8** — v8 added the "NER $/mo (amort)" column at Rent Roll Analysis
       AV (header carried at both 210 and 213). Either AV cell carrying "NER"
       (and not v11 above) marks v8.
    3. **v4 vs v5+** — v5 introduced a "Care Level Tier" column on Rent Roll
       Analysis row 210 that v4 lacks. The v5.1 column restructure moved it
       from AP → AO, so probe AO210 first (current position), AP210 as a
       pre-v5.1 fallback. Either carrying "Care Level Tier" marks v5+.
    4. **v5 vs v6** — v6 rebuilt the T-12 Analysis INCOME section into an
       actual-T12 build. v6 rev2 (canonical since 2026-06-03) carries EGI at
       A80 and Auto Expense at A117; the pre-rev2 v6 had them at A77/A114 —
       probe both pairs (the old A77/A114-only probe mis-detected rev2
       uploads as v5). Any hit marks v6.

    Falls back to "v11" (the binding default) if the file lacks the expected
    sheets (rare — indicates a non-ALF template).
    """
    try:
        import io as _io
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(_io.BytesIO(template_bytes), data_only=False)
        ws_rr = wb["Rent Roll Analysis"] if "Rent Roll Analysis" in wb.sheetnames else None
        # Stage 1 — v11 (paste grid re-anchored: header at 223, Section S at ~205)
        if ws_rr is not None:
            a223 = ws_rr["A223"].value
            if isinstance(a223, str) and "Unit/Bed" in a223:
                return "v11"
            a205 = ws_rr["A205"].value
            if isinstance(a205, str) and "CONCESSION" in a205.upper():
                return "v11"
        # Stage 2 — v8 (NER column at AV; v11 already excluded above)
        if ws_rr is not None:
            for addr in ("AV210", "AV213"):
                v = ws_rr[addr].value
                if isinstance(v, str) and "NER" in v:
                    return "v8"
        # Stage 2 — v4 vs v5+ (Care Level Tier at AO210 in v5.1+/v6, AP210 pre-v5.1)
        is_v5_plus = False
        if ws_rr is not None:
            for addr in ("AO210", "AP210"):
                v = ws_rr[addr].value
                if isinstance(v, str) and "Care Level" in v and "Tier" in v:
                    is_v5_plus = True
                    break
        if not is_v5_plus:
            return "v4"
        # Stage 3 — v5 vs v6 (decided on T-12 Analysis income layout; rev2
        # rows first, pre-rev2 rows as fallback)
        if "T-12 Analysis" in wb.sheetnames:
            ws_t12 = wb["T-12 Analysis"]
            for addr, needle in (
                ("A80", "EFFECTIVE GROSS INCOME"),
                ("A117", "AUTO EXPENSE"),
                ("A77", "EFFECTIVE GROSS INCOME"),
                ("A114", "AUTO EXPENSE"),
            ):
                v = ws_t12[addr].value
                if isinstance(v, str) and needle in v.upper():
                    return "v6"
        return "v5"
    except Exception:
        return "v11"  # default to v11 — the registry's binding template


def _load_uw_template(uploaded_file) -> tuple[bytes, str, str]:
    """Resolve the UW Template source — uploaded file wins over bundled default.

    Mirrors `_load_analyzer` so the operator gets the same load behavior:
    bundled `assets/ALF_UW_Template_v11.xlsx` is used by default; an
    upload via Advanced → "UW Template override" replaces it for the session.

    Returns: (template_bytes, source_label, template_version)
      - template_bytes: raw .xlsx bytes
      - source_label: "uploaded: <filename>" or "bundled (assets/<file>)"
      - template_version: "v4" / "v5" / "v6" / "v8" (drives writer's targets.{v} block)

    Raises FileNotFoundError if neither uploaded file nor bundled file exists.
    """
    if uploaded_file is not None:
        b = uploaded_file.getvalue()
        name = getattr(uploaded_file, "name", "uw_template.xlsx")
        return b, f"uploaded: {name}", _detect_uw_template_version(b)
    if BUNDLED_UW_TEMPLATE_PATH.exists():
        b = BUNDLED_UW_TEMPLATE_PATH.read_bytes()
        return b, f"bundled (assets/{BUNDLED_UW_TEMPLATE_PATH.name})", BUNDLED_UW_TEMPLATE_VERSION
    raise FileNotFoundError(
        f"Bundled UW Template not found at {BUNDLED_UW_TEMPLATE_PATH}. "
        "Either restore the file in the repo's assets/ folder or upload a "
        "custom template via the Advanced expander in the intake panel."
    )


def _mf_t12_paste_csv(res) -> bytes:
    """Build a paste-ready CSV for the MF UW Model's T-12 Analysis Layer 1.

    Columns: Acct # | Account Name (raw) | <12 month labels> | T-12 Total | → MAPPING.
    Paste A->A, B->B, the 12 months across C-N, and the bucket into col P.
    """
    import csv
    import io

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Acct #", "Account Name (raw)"] + res.month_labels
               + ["T-12 Total", "→ MAPPING"])
    for ln in res.lines:
        w.writerow([ln.acct or "", ln.name]
                   + [round(x, 2) for x in ln.monthly]
                   + [round(ln.total, 2), ln.bucket])
    return buf.getvalue().encode("utf-8-sig")


def _render_mf_t12_detail(res) -> None:
    """Render the T-12 parse result: metrics, reconciliation, buckets, paste CSV."""
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Detected format", res.format_guess)
    c2.metric("GL lines", len(res.lines))
    c3.metric("Coverage", f"{res.coverage * 100:.0f}%")
    c4.metric("Period", res.period or "—")

    comp, rep = res.computed, res.reported
    recon = []
    for key, label, rk in [("income", "Total Income", "total_income"),
                           ("expense", "Total OpEx", "total_expense"),
                           ("noi", "NOI", "noi")]:
        r = rep.get(rk)
        recon.append({
            "Metric": label,
            "Computed": f"${comp[key]:,.0f}",
            "As-reported": f"${r:,.0f}" if r is not None else "—",
            "Δ": f"${comp[key] - r:,.0f}" if r is not None else "—",
        })
    st.markdown("**Reconciliation** (computed from mapped lines vs. the statement's own totals)")
    st.dataframe(pd.DataFrame(recon), hide_index=True, use_container_width=True)

    buckets: dict[str, float] = {}
    for ln in res.lines:
        buckets[ln.bucket] = buckets.get(ln.bucket, 0.0) + ln.total
    bdf = pd.DataFrame(
        [{"_StdCOA bucket": b, "T-12 Total": f"${v:,.0f}"}
         for b, v in sorted(buckets.items(), key=lambda kv: -abs(kv[1]))]
    )
    st.markdown("**Standardized buckets** (→ col P on the model)")
    st.dataframe(bdf, hide_index=True, use_container_width=True)
    for warn in res.warnings:
        st.warning(warn, icon="⚠️")
    st.download_button(
        "⬇️ T-12 paste-ready mapping (CSV)", data=_mf_t12_paste_csv(res),
        file_name="MF_T12_StdCOA_mapping.csv", mime="text/csv", key="mf_t12_csv",
    )


def _parse_currency(raw: str) -> int:
    """Parse '$18,000,000' / '18000000' / '18M' → int. Returns 0 if blank/invalid."""
    if not raw:
        return 0
    s = raw.replace("$", "").replace(",", "").strip().upper()
    mult = 1
    if s.endswith("M"):
        s, mult = s[:-1], 1_000_000
    elif s.endswith("K"):
        s, mult = s[:-1], 1_000
    try:
        return int(float(s) * mult)
    except ValueError:
        return 0


def _mf_file_token(f) -> str:
    """Stable per-upload token for cache keying. Streamlit's UploadedFile carries
    a `file_id` that persists across reruns (e.g. a download-button click) and
    changes only when a new file is chosen — far cheaper than hashing the bytes."""
    if f is None:
        return "none"
    fid = getattr(f, "file_id", None)
    if fid:
        return f"id:{fid}"
    return "h:" + hashlib.sha256(f.getvalue()).hexdigest()


def _mf_sig(rr_file, t12_file, ar_file, om_file, model_override, engine_ai, om_api_key) -> str | None:
    """Signature of all MF inputs. None when nothing is uploaded. Changing any
    file (or the OM engine/key) changes the signature → triggers a recompute."""
    tokens = [_mf_file_token(f) for f in
              (rr_file, t12_file, ar_file, model_override)]
    om_tok = _mf_file_token(om_file)
    if all(t == "none" for t in tokens) and om_tok == "none":
        return None
    extra = ""
    if om_file is not None:                       # OM result depends on engine/key
        extra = f"|{engine_ai}|{(om_api_key or '')[:8]}"
    return "|".join(tokens) + f"|{om_tok}{extra}"


def _compute_mf(rr_file, t12_file, ar_file, om_file, model_override,
                engine_ai, om_api_key, sig) -> dict:
    """Parse every uploaded MF doc + build the populated model, under the
    determinate progress overlay. Returns a result bundle (parsed objects,
    output bytes, report, per-step errors) so the caller can cache it and the
    renderer can redraw without re-parsing. This is the ONLY place the heavy
    work runs — a cache hit skips it entirely (no re-parse, no overlay)."""
    res = {"sig": sig, "rr": None, "rr_err": None, "t12": None, "t12_err": None,
           "ar": None, "ar_err": None, "ar_join": None, "om": None, "om_err": None,
           "out": None, "report": None, "build_err": None, "model_src": None,
           "prop_name": None}

    plan = []
    if rr_file is not None:
        plan.append((1.0, "Parsing rent roll…"))
    if t12_file is not None:
        plan.append((1.0, "Parsing T-12…"))
    if ar_file is not None:
        plan.append((1.0, "Parsing AR aging…"))
    if om_file is not None:
        plan.append((2.0 if engine_ai else 1.0, "Extracting OM…"))
    if rr_file is not None or t12_file is not None or om_file is not None:
        plan.append((3.0, "Building MF UW Model…"))
    pp = _PipelineProgress(plan)

    prop_name = None
    if rr_file is not None:
        try:
            with pp.stage():
                rr = parse_mf_rr(rr_file.getvalue())
            res["rr"] = rr
            prop_name = rr.property_hint or derive_property_name(rr_file.name)
        except Exception as exc:  # noqa: BLE001
            res["rr_err"] = str(exc)

    if t12_file is not None:
        try:
            with pp.stage():
                t12 = parse_mf_t12(t12_file.getvalue())
            res["t12"] = t12
            prop_name = prop_name or derive_property_name(t12_file.name)
        except Exception as exc:  # noqa: BLE001
            res["t12_err"] = str(exc)

    if ar_file is not None:
        try:
            with pp.stage():
                ar = parse_mf_ar(ar_file.getvalue())
            res["ar"] = ar
            if res["rr"] is not None:
                res["ar_join"] = join_ar_to_units(res["rr"].units, ar)
        except Exception as exc:  # noqa: BLE001
            res["ar_err"] = str(exc)

    if om_file is not None:
        engine = "llm" if engine_ai else "basic"
        with pp.stage():
            try:
                om = parse_mf_om(om_file.getvalue(), engine=engine,
                                 api_key=om_api_key or None)
                res["om"] = om
                prop_name = (prop_name or om.prop_info.property_name
                             or derive_property_name(om_file.name))
            except MFOMExtractorError as exc:
                res["om_err"] = f"OM extraction failed: {exc}"
            except Exception as exc:  # noqa: BLE001
                res["om_err"] = f"Could not read the OM: {exc}"

    res["prop_name"] = prop_name
    if res["rr"] is not None or res["t12"] is not None or res["om"] is not None:
        try:
            if model_override is not None:
                model_bytes = model_override.getvalue()
                res["model_src"] = "uploaded override"
            else:
                model_bytes = BUNDLED_MF_MODEL_PATH.read_bytes()
                res["model_src"] = "bundled MF_UW_Model_v15.xlsx"
            with pp.stage() as sub:
                out, report = populate_mf_model(
                    model_bytes, t12=res["t12"], rr=res["rr"], om=res["om"],
                    property_name=prop_name,
                    property_units=(res["rr"].unit_count if res["rr"] is not None else None),
                    progress=sub,
                )
            res["out"], res["report"] = out, report
        except Exception as exc:  # noqa: BLE001
            res["build_err"] = str(exc)
    return res


def _render_mf_result(res: dict) -> None:
    """Draw the MF results from a (possibly cached) result bundle. Cheap — no
    parsing — so it's safe to re-run on every Streamlit rerun (e.g. a download)."""
    # --- Rent Roll ---
    if res["rr_err"]:
        st.markdown("### \U0001F4C4 Rent Roll")
        st.error(f"Could not parse the rent roll: {res['rr_err']}")
    elif res["rr"] is not None:
        st.markdown("### \U0001F4C4 Rent Roll")
        rr = res["rr"]
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("Units", rr.unit_count)
        d2.metric("Occupied", rr.occupied)
        d3.metric("Vacant", rr.vacant)
        d4.metric("Legal / eviction", rr.legal_count)
        for w in rr.warnings:
            st.warning(w, icon="⚠️")

    # --- T-12 ---
    if res["t12_err"]:
        st.markdown("### \U0001F4C4 T-12 income statement")
        st.error(f"Could not parse the T-12: {res['t12_err']}")
    elif res["t12"] is not None:
        st.markdown("### \U0001F4C4 T-12 income statement")
        _render_mf_t12_detail(res["t12"])

    # --- AR aging ---
    if res["ar_err"]:
        st.markdown("### \U0001F4C4 AR aging")
        st.error(f"Could not parse the AR aging report: {res['ar_err']}")
    elif res["ar"] is not None:
        st.markdown("### \U0001F4C4 AR aging")
        ar = res["ar"]
        a1, a2, a3 = st.columns(3)
        a1.metric("AR rows", len(ar.rows))
        a2.metric("Total AR", f"${ar.total_ar:,.0f}")
        a3.metric("Period", ar.period_hint or "—")
        rep = res["ar_join"]
        if rep is not None:
            st.caption(f"Joined {rep.matched}/{len(ar.rows)} AR rows to units by Bldg-Unit.")
            for w in rep.warnings:
                st.warning(w, icon="⚠️")
        else:
            st.info("Upload the Rent Roll too — AR aging joins to units by Bldg-Unit.")

    # --- OM ---
    if res["om_err"]:
        st.markdown("### \U0001F4D5 Offering Memorandum")
        st.error(res["om_err"])
    elif res["om"] is not None:
        st.markdown("### \U0001F4D5 Offering Memorandum")
        om = res["om"]
        pi = om.prop_info
        o1, o2, o3, o4 = st.columns(4)
        o1.metric("Units (OM)", pi.units_total or "—")
        o2.metric("Year built", pi.year_built or "—")
        o3.metric("Rent comps", len(om.comps))
        o4.metric("Engine", om.engine.upper())
        with st.expander("OM extraction detail"):
            st.write({"property": pi.property_name, "address": pi.address,
                      "county": pi.county, "acres": pi.lot_acres,
                      "buildings": pi.num_buildings, "stories": pi.num_stories,
                      "class": pi.building_class, "unit_mix": len(pi.unit_mix),
                      "market": om.market.city_market})
            if om.comps:
                st.dataframe(pd.DataFrame(
                    [{"Comp": c.name, "Yr": c.year_built, "Units": c.units,
                      "Avg SF": c.avg_sf, "Asking": c.asking_rent,
                      "Occ": c.occupancy} for c in om.comps]),
                    hide_index=True, use_container_width=True)
        for w in om.warnings:
            st.caption(f"⚠️ {w}")

    # --- Populated MF UW Model ---
    if res["build_err"]:
        st.divider()
        st.markdown("### \U0001F9EE Populate the MF UW Model")
        st.error(f"Could not populate the MF UW Model: {res['build_err']}")
    elif res["out"] is not None:
        st.divider()
        st.markdown("### \U0001F9EE Populate the MF UW Model")
        report = res["report"]
        st.success(
            f"Populated **{report['rr_units']}** units + "
            f"**{report['t12_lines']}** T-12 lines + "
            f"**{report['om_prop_cells']}** Prop Info fields + "
            f"**{report['om_comps']}** rent comps into the {res['model_src']}."
        )
        for w in report["warnings"]:
            st.caption(f"⚠️ {w}")
        safe = (res["prop_name"] or "MF_Property").replace(" ", "_")
        st.download_button(
            "⬇️ Download populated MF UW Model (.xlsx)",
            data=res["out"],
            file_name=f"{safe}_MF_UW_Model_populated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="mf_model_dl",
        )
    else:
        st.info("Upload a Rent Roll and/or a T-12 to populate the MF UW Model.")


def _ck_chip(label: str, on: bool) -> str:
    cls = "ok" if on else "off"
    mark = "✓" if on else "—"
    return f'<span class="ck-chip {cls}">{label} {mark}</span>'


def _render_mf_intake() -> None:
    """MF (multifamily) mode — RR / T-12 / AR intake → populate the MF UW Model.

    Cockpit layout (2026-06-12): same chrome as ALF — `Intake` header, 3-col
    uploaders, Advanced expander, then the UW//DECK command bar with status
    chips between intake and results."""
    st.markdown("##### Intake")

    cu1, cu2, cu3 = st.columns(3)
    rr_file = cu1.file_uploader("Rent Roll (.xlsx/.xls)", type=["xlsx", "xlsm", "xls"], key="mf_rr_up")
    t12_file = cu2.file_uploader("T-12 (.xlsx)", type=["xlsx"], key="mf_t12_up")
    ar_file = cu3.file_uploader("AR aging (.xlsx)", type=["xlsx"], key="mf_ar_up")

    # --- OM (Offering Memorandum PDF) → Prop Info + Rental Comps ---
    om_file = st.file_uploader(
        "Offering Memorandum (.pdf) — property facts, market data & rent comps",
        type=["pdf"], key="mf_om_up")
    oc1, oc2 = st.columns([1, 2])
    om_engine = oc1.radio(
        "OM extraction engine", ["AI (Claude)", "Basic (no API)"],
        key="mf_om_engine", horizontal=False,
        help="AI reads the whole OM with Claude (needs an API key); Basic does a "
             "no-API labelled-facts scan (property details only, no comps/market).")
    _secret_key = st.secrets.get("ANTHROPIC_API_KEY", "") if hasattr(st, "secrets") else ""
    om_api_key = _secret_key or os.environ.get("ANTHROPIC_API_KEY", "")
    if om_engine.startswith("AI") and not om_api_key:
        om_api_key = oc2.text_input(
            "Anthropic API key", type="password", key="mf_om_key",
            help="Used only for this extraction; not stored. Or add ANTHROPIC_API_KEY "
                 "to Streamlit secrets. Switch to Basic to skip.")

    # ── Underwriting controls (mirrors the ALF row) ─────────────────────────
    mo1, _mo2, _mo3 = st.columns(3)
    with mo1:
        if "mf_pp_input" not in st.session_state:
            st.session_state["mf_pp_input"] = ""

        def _reformat_mf_pp() -> None:
            n = _parse_currency(st.session_state["mf_pp_input"])
            if n > 0:
                st.session_state["mf_pp_input"] = f"${n:,}"

        st.text_input(
            "Purchase price",
            key="mf_pp_input",
            on_change=_reformat_mf_pp,
            placeholder="$20,000,000",
            help=(
                "Drives the Going-in cap, Price/unit, Price/SF and GRM tiles "
                "on the MF Dashboard. Accepts `$20,000,000`, `20M`, `20000000`. "
                "Leave blank to skip — the dashboard still renders everything "
                "that doesn't need a price."
            ),
        )
    mf_purchase_price = _parse_currency(st.session_state.get("mf_pp_input", ""))

    with st.expander("Advanced — override MF UW Model template"):
        model_override = st.file_uploader("MF UW Model (.xlsx)", type=["xlsx"], key="mf_model_up")

    # Compute once per distinct set of uploads, then cache in session_state.
    # A Streamlit rerun (e.g. clicking the download button, or toggling any
    # unrelated widget) reuses the cached result instead of re-parsing — so the
    # overlay only appears on a genuine recompute. The signature changes only
    # when an uploaded file (or the OM engine/key) changes.
    _engine_ai = om_engine.startswith("AI")
    sig = _mf_sig(rr_file, t12_file, ar_file, om_file, model_override,
                  _engine_ai, om_api_key)
    if sig is None:
        res = None
    else:
        cached = st.session_state.get("mf_result")
        if cached is not None and cached.get("sig") == sig:
            res = cached                      # reuse — no re-parse, no overlay
        else:
            res = _compute_mf(rr_file, t12_file, ar_file, om_file, model_override,
                              _engine_ai, om_api_key, sig)
            st.session_state["mf_result"] = res

    # Cockpit command bar — mirrors the ALF bar so both modes read as the
    # same product: deal readout + intake status chips + version chrome.
    _deal = (res or {}).get("prop_name") or ""
    _deal_label = (
        f"{_deal.upper()}" if _deal
        else "NO DEAL LOADED — drop a rent roll or T-12 above"
    )
    st.markdown(
        f"""
        <div class="ck-bar">
            <span class="ck-brand">UW//DECK</span>
            <span class="ck-chip ok">MF</span>
            <span class="ck-deal">{_deal_label}</span>
            {_ck_chip("RR", rr_file is not None)}
            {_ck_chip("T12", t12_file is not None)}
            {_ck_chip("AR", ar_file is not None)}
            {_ck_chip("OM", om_file is not None)}
            <span class="ck-ver">MF UW MODEL v15</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Top-level tabs — mirror ALF (Dashboard first, Workspace = populate flow).
    mf_tab_dashboard, mf_tab_workspace = st.tabs(["Dashboard", "Workspace"])

    with mf_tab_workspace:
        if res is None:
            st.info("Upload a Rent Roll and/or a T-12 to populate the MF UW Model.")
        else:
            _render_mf_result(res)

        st.divider()
        st.caption(
            "Coming next: redIQ Sortable-RR ancillary-fee breakouts (cols W–AK)."
        )

    with mf_tab_dashboard:
        _rr_ok = res is not None and res.get("rr") is not None
        _t12_ok = res is not None and res.get("t12") is not None
        if not (_rr_ok or _t12_ok):
            st.info(
                "Upload a Rent Roll and/or a T-12 in the intake panel to "
                "populate the MF dashboard."
            )
        else:
            try:
                _mf_model = compute_mf_dashboard(
                    res.get("rr"), res.get("t12"),
                    purchase_price=mf_purchase_price or None,
                    property_name=res.get("prop_name") or "MF deal",
                    period_label=(
                        (res.get("t12").period if _t12_ok else "")
                        or (res.get("rr").period_hint if _rr_ok else "")
                    ),
                )
                render_mf_dashboard(_mf_model)
            except Exception as exc:  # noqa: BLE001
                st.error(f"MF dashboard failed to render: {exc}")


# ---------------------------------------------------------------------------
# Page setup
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Underwriting Intake",
    page_icon="🏢",
    layout="wide",
)

# Password gate runs first so it gets a clean slate for inject_landing_css().
# inject_brand_css() runs AFTER — the landing page never sees navy CSS because
# require_login() calls st.stop() before we reach this line when unauthenticated.
username = require_login()

# Now authenticated — apply brand styling, then layer the cockpit terminal
# theme on top (2026-06-12 UI redesign — graphite/teal, monospace chrome).
# The Light toggle (rendered in the top control row below) drives both the
# cockpit CSS variant and Streamlit's NATIVE theme — the latter so canvas-
# rendered surfaces (st.dataframe grids, Altair charts, widget internals)
# follow the flip, not just the custom chrome.


def _set_native_theme(light: bool) -> None:
    """Flip Streamlit's native theme to match the cockpit variant.

    Uses the semi-private ``streamlit.config`` set_option — wrapped so a
    future API change degrades gracefully to CSS-only theming (custom chrome
    still flips; native widgets keep the server default). Called from the
    Light toggle's on_change callback: callbacks run BEFORE the rerun's
    script executes, so the new theme rides the very next NewSession message
    — no explicit st.rerun() needed. (An early st.rerun() is also actively
    harmful here: rerunning before the toggle widget is re-instantiated
    makes Streamlit garbage-collect its pending state, snapping it back.)
    NOTE: config is process-wide, not per-session — fine for this app's
    single-operator usage; a second concurrent user would inherit the same
    base theme on their next rerun.
    """
    desired = "light" if light else "dark"
    tokens = {
        "light": {
            "backgroundColor": "#F2F4F6",
            "secondaryBackgroundColor": "#FFFFFF",
            "textColor": "#18202A",
            "primaryColor": "#0E8A63",
        },
        "dark": {
            "backgroundColor": "#101418",
            "secondaryBackgroundColor": "#1A2027",
            "textColor": "#E6EDF5",
            "primaryColor": "#5DCAA5",
        },
    }[desired]
    try:
        from streamlit import config as _st_config

        if _st_config.get_option("theme.base") != desired:
            _st_config.set_option("theme.base", desired)
            for _k, _v in tokens.items():
                _st_config.set_option(f"theme.{_k}", _v)
    except Exception:
        pass  # degrade to CSS-only theming


def _on_theme_toggle() -> None:
    _set_native_theme(bool(st.session_state.get("ck_light_mode", False)))


_ck_light = bool(st.session_state.get("ck_light_mode", False))
# Re-assert on every run too — covers a fresh session joining a process whose
# global theme was left in the other mode by a previous toggle.
_set_native_theme(_ck_light)
inject_brand_css()
inject_cockpit_css(light=_ck_light)

# ---------------------------------------------------------------------------
# Loading overlay slot (Track 5 v0.1.8 — shared by ALF + MF)
# ---------------------------------------------------------------------------
# Created here, ABOVE the mode dispatch and the top-level tabs, so the overlay's
# DOM is never inside an inactive/hidden container (Streamlit uses display:none
# on inactive tabs, which kills position:fixed descendants). The slot captures a
# DeltaGenerator bound to this DOM position; later _show_loading(...) calls from
# either mode or tab still render at this module-level position. CSS for
# `.t5-overlay` lives in `branding.inject_brand_css()` (injected just above).
import contextlib  # local import to keep the heavy imports at the top of the file

_overlay_slot = st.empty()


@contextlib.contextmanager
def _show_loading(label: str):
    """Full-page loading overlay context manager (visible across modes/tabs)."""
    _overlay_slot.markdown(
        f"""
        <div class="t5-overlay" role="status" aria-live="polite">
            <div class="t5-overlay-ring"></div>
            <div class="t5-overlay-label">{label}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    try:
        yield
    finally:
        _overlay_slot.empty()


def _render_overlay_pct(pct: float, label: str) -> None:
    """Determinate overlay: spinner + a gold % readout + a progress bar."""
    pct = max(1, min(100, int(round(pct))))
    _overlay_slot.markdown(
        f"""
        <div class="t5-overlay" role="status" aria-live="polite">
            <div class="t5-overlay-ring"></div>
            <div class="t5-overlay-pct">{pct}%</div>
            <div class="t5-overlay-bar"><div class="t5-overlay-bar-fill" style="width:{pct}%"></div></div>
            <div class="t5-overlay-label">{label}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


class _PipelineProgress:
    """Determinate progress across a sequence of pipeline stages.

    `stages` is an ordered list of (weight, label). The displayed % is the
    weighted fraction of stages completed plus any sub-progress within the
    current stage — so it genuinely reflects how much of the *whole job* is
    left, not a timer. Fast stages (file parses) just tick the bar forward as
    each completes; a slow stage (the model build) can report sub-progress via
    the callback yielded by `stage()`. The overlay clears after each stage so
    that stage's results render, then reappears (at a higher %) for the next.
    """

    def __init__(self, stages):
        self._stages = list(stages)
        self._total = sum(w for w, _ in self._stages) or 1
        self._before = 0.0   # weight completed before the current stage
        self._i = -1

    @contextlib.contextmanager
    def stage(self):
        self._i += 1
        _, label = self._stages[self._i]
        self._show(label, 0.0)
        try:
            yield self._sub          # pass to slow stages for sub-progress
        finally:
            self._before += self._stages[self._i][0]
            _overlay_slot.empty()

    def _sub(self, frac: float, label=None) -> None:
        _, lbl = self._stages[self._i]
        self._show(label or lbl, max(0.0, min(1.0, frac)))

    def _show(self, label: str, frac: float) -> None:
        w = self._stages[self._i][0]
        _render_overlay_pct((self._before + frac * w) / self._total * 100, label)

# No logo on the post-login app — logos live only on the white landing page.
# The mode selector is the first element inside.

# ---------------------------------------------------------------------------
# Mode selector — ALF (senior housing) vs MF (multifamily)
# ---------------------------------------------------------------------------
# Phase 0: every authenticated user may use both modes (see auth.allowed_modes).
# When exactly one mode is permitted, the selector is hidden and the app
# auto-routes — so the future per-user access-type gating needs no app.py
# change. The selected mode routes the whole pipeline; MF renders a placeholder
# and stops before the ALF pipeline runs.
_modes = allowed_modes(username)
# Top control row (shared by ALF + MF): mode selector on the left; light/dark
# toggle + username chip + Sign out clustered in the upper-right corner.
# The toggle's state is read ABOVE (before CSS injection) via session_state;
# flipping it reruns the script, which re-injects the matching theme.
_mode_col, _theme_col, _user_col = st.columns([4, 1, 2], vertical_alignment="center")
with _theme_col:
    st.toggle(
        "Light",
        key="ck_light_mode",
        on_change=_on_theme_toggle,
        help="Switch the cockpit between the dark terminal and light paper themes.",
    )
with _user_col:
    render_user_controls(username)
if len(_modes) == 1:
    app_mode = _modes[0]
else:
    with _mode_col:
        app_mode = st.radio(
            "Property type",
            options=_modes,
            format_func=lambda m: {
                "ALF": "ALF // senior housing",
                "MF": "MF // multifamily",
            }.get(m, m),
            horizontal=True,
            key="app_mode",
            label_visibility="collapsed",
            help=(
                "Switch between the senior-housing (ALF) normalizer and the "
                "multifamily (MF) intake pipeline."
            ),
        )

if app_mode == "MF":
    _render_mf_intake()
    st.stop()

# --- ALF mode (default): the existing pipeline runs below, unchanged. ---

# Cockpit command bar (replaces the old title row + version badge + caption).
# Rendered AFTER the intake panel below, where the upload widgets' state is
# known — the bar's status chips reflect which intake files are loaded.


# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Intake panel — main area, MF-style (2026-06-12 cockpit layout; the sidebar
# is gone — both modes render their uploaders in the page so ALF and MF read
# as the same product). Widget labels/keys unchanged from the sidebar era so
# session state carries across.
# ---------------------------------------------------------------------------
st.markdown("##### Intake")

iu1, iu2, iu3 = st.columns(3)

with iu1:
    rr_file = st.file_uploader(
        "Rent Roll (.xlsx / .xlsm / .xls) — required",
        type=["xlsx", "xlsm", "xls"],
        help=(
            "Any senior housing rent roll. Header doesn't need to be on row 1. "
            "Legacy .xls (binary Excel) supported via xlrd — Yardi exports "
            "frequently land in this format."
        ),
    )

    auto_detected_date = None
    if rr_file is not None:
        auto_detected_date = detect_period_date(getattr(rr_file, "name", ""))

    period_date_input = st.date_input(
        "RR Period Date",
        value=auto_detected_date or dt.date.today(),
        help=(
            "Written to column S of the Analyzer's Rent Roll Input sheet on "
            "every row. Auto-detected from the rent roll filename when "
            "possible. Override if needed."
        ),
    )
    if auto_detected_date:
        st.caption(f"Auto-detected: **{auto_detected_date.isoformat()}**")
    elif rr_file is not None:
        st.caption("Could not auto-detect — set manually.")

with iu2:
    raw_t12_file = st.file_uploader(
        "Raw T12 (.xlsx / .xlsm / .xls) — optional",
        type=["xlsx", "xlsm", "xls"],
        key="raw_t12_uploader",
        help=(
            "Optional. Upload a raw T12 export from Yardi (Income to Budget), "
            "MRI (R12MINCS), or a broker Financial Summary (`Historical "
            "Performance` header at A4). The app parses it, detects month "
            "labels, applies drop-rules, and writes the GL detail into the "
            "Analyzer's 'T12 Input' sheet. Mappings for any UNMATCHED "
            "descriptions can be filled in below before download."
        ),
    )

    # Cluster B (T12 v0.2.0): annualize toggle for partial-year T12 files. The
    # parser reads the value below; default OFF surfaces a warning instead of
    # silently scaling. Disabled in UI when no T12 is uploaded.
    annualize_partial_year = st.checkbox(
        "Annualize partial-year T12",
        value=False,
        key="annualize_partial_year",
        disabled=raw_t12_file is None,
        help=(
            "When the uploaded T12 covers fewer than 12 months, multiply every "
            "monthly value by 12/N (where N is months populated). Off by "
            "default — partial-year T12s surface as a warning so you decide "
            "explicitly. Use with caution: assumes flat seasonality."
        ),
    )

with iu3:
    ar_file = st.file_uploader(
        "AR Aging (.xlsx / .xlsm / .xls / .csv) — optional",
        type=["xlsx", "xlsm", "xls", "csv"],
        key="ar_aging_uploader",
        help=(
            "Optional. Upload an AR aging report from the operator. The app "
            "parses bucket totals (Current / 31-60 / 61-90 / 91-120 / 120+), "
            "payer mix, and roll-forward fields, then writes them into the "
            "Analyzer's 'AR & Collections' sheet (hidden by default; revealed "
            "when AR is uploaded). Adds Workbook Health P5 pre-export gate."
        ),
    )

    ar_as_of_override = None
    if ar_file is not None:
        ar_as_of_override = st.date_input(
            "AR as-of date (override)",
            value=period_date_input,
            key="ar_as_of_input",
            help=(
                "Defaults to the RR period date. Override if the operator's "
                "AR report carries a different as-of date — the P5 pre-export "
                "gate will flag if the AR date doesn't match RR."
            ),
        )

# ── Underwriting controls row ────────────────────────────────────────────────
io1, io2, io3 = st.columns(3)

with io1:
    # Track 4 / Phase 2.5 — UW Template scenario selector. The template file
    # itself is bundled by default (see _load_uw_template) and can be
    # overridden via the Advanced expander below. The scenario radio stays
    # always-visible because it's a per-deal underwriting choice, not an
    # operational config.
    uw_template_scenario = st.radio(
        "UW Template scenario",
        options=["normalized", "t12_actual"],
        index=0,
        horizontal=True,
        key="uw_template_scenario",
        help=(
            "Which UW Output column to write into the populated UW Template. "
            "**normalized** (col F) = analyst's stabilized underwriting "
            "assumption — the contract's underwriting figure. "
            "**t12_actual** (col E) = trailing-12 actuals, useful for a "
            "variance / sanity-check view. Defaults to normalized."
        ),
    )

with io2:
    # Auto-format-on-blur pattern: the on_change callback fires when the user
    # presses Enter or tabs away, parses whatever they typed, and writes the
    # formatted value back into session_state. Streamlit re-reads the widget's
    # value from session_state on the next rerun, so the field shows
    # "$18,000,000" after blur. Live per-keystroke formatting is not possible
    # with vanilla st.text_input (no per-key callback) — would require a
    # custom HTML/JS component.
    if "pp_input" not in st.session_state:
        st.session_state["pp_input"] = ""

    def _reformat_pp() -> None:
        n = _parse_currency(st.session_state["pp_input"])
        if n > 0:
            st.session_state["pp_input"] = f"${n:,}"
        # Invalid → leave raw intact so the user can correct it.

    st.text_input(
        "Purchase price",
        key="pp_input",
        on_change=_reformat_pp,
        placeholder="$18,000,000",
        help=(
            "Drives the Going-in cap rate, EBITDAR cap, and Price/bed "
            "tiles on the Dashboard. Accepts `$18,000,000`, `18000000`, "
            "`18M`, `500K`, etc. Auto-formats to `$#,###,###` on Enter "
            "or tab-out. Leave blank to skip — you can set it later in "
            "the downloaded Analyzer at T12 Analytics!E117."
        ),
    )
    purchase_price_input = _parse_currency(st.session_state.get("pp_input", ""))

with io3:
    care_type_default = st.selectbox(
        "Care Type default",
        options=["(none — flag missing)", "IL", "AL", "MC"],
        index=0,
        help=(
            "Applied when the rent roll source has no Care Type / Wing / "
            "Building column. For single-care-setting properties (e.g., a "
            "100% AL building) this fills in the Care Type for every bed. "
            "Source values always win — explicit Care Type columns in the "
            "rent roll override this default."
        ),
    )
    if care_type_default.startswith("("):
        care_type_default = ""

with st.expander("Advanced"):
    av1, av2 = st.columns(2)
    with av1:
        sheet_override = st.text_input(
            "RR sheet name (auto if blank)",
            value="",
            help="Defaults to 'Details' if present, otherwise the first sheet.",
        )
        mapping_file = st.file_uploader(
            "RR Mapping override (.xlsx)",
            type=["xlsx"],
            help=(
                "Override defaults for Apartment_Type_Rules, Bed_Status_Rules, "
                "Payer_Type_Rules, Care_Level_Rules, Care_Bucket_Rules. "
                "Any sheet you omit falls back to built-in defaults."
            ),
        )
    with av2:
        analyzer_override_file = st.file_uploader(
            "Analyzer template override (.xlsx)",
            type=["xlsx"],
            key="analyzer_override_uploader",
            help=(
                "By default the app uses the bundled Analyzer "
                "(`ALF_Financial_Analyzer_Only.xlsx`). Upload to override "
                "for this session only — uploads do not modify the bundled "
                "file."
            ),
        )
        uw_template_override_file = st.file_uploader(
            "UW Template override (.xlsx)",
            type=["xlsx"],
            key="uw_template_override_uploader",
            help=(
                "By default the app uses the bundled UW Template "
                "(`assets/ALF_UW_Template_v11.xlsx`). Upload to override "
                "for this session only — e.g. to populate against a "
                "legacy v6 / v5 / v4 template. Uploads do not modify the "
                "bundled file. The writer auto-detects the version from "
                "the file's structure."
            ),
        )
    st.caption(
        f"RR v{RR_VERSION} · T12 v{T12_VERSION} · AR v{AR_VERSION} "
        f"· T5 v{T5_VERSION} · UWT v{UWT_VERSION}"
    )

# ---------------------------------------------------------------------------
# Cockpit command bar — deal readout + intake status chips + version chrome
# ---------------------------------------------------------------------------
_deal_name = (
    derive_property_name(getattr(rr_file, "name", "")) if rr_file is not None else ""
)
_deal_label = (
    f"{_deal_name or 'DEAL'} · {period_date_input.strftime('%b %Y').upper()}"
    if rr_file is not None
    else "NO DEAL LOADED — drop a rent roll above"
)

# ALF mode chip mirrors the MF bar so the two pages read identically.
st.markdown(
    f"""
    <div class="ck-bar">
        <span class="ck-brand">UW//DECK</span>
        <span class="ck-chip ok">ALF</span>
        <span class="ck-deal">{_deal_label}</span>
        {_ck_chip("RR", rr_file is not None)}
        {_ck_chip("T12", raw_t12_file is not None)}
        {_ck_chip("AR", ar_file is not None)}
        <span class="ck-chip ok">{uw_template_scenario.replace("_", " ").upper()}</span>
        <span class="ck-ver">RR {RR_VERSION} · T12 {T12_VERSION} · SUB {ANALYZER_SUBSTRATE_VERSION} · UWT {UWT_VERSION}</span>
    </div>
    """,
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# Top-level switch tabs — Dashboard (clean slate) vs Workspace (everything else)
# ---------------------------------------------------------------------------
top_tab_dashboard, top_tab_workspace = st.tabs(["Dashboard", "Workspace"])

with top_tab_workspace:


    # ---------------------------------------------------------------------------
    # Resolve Analyzer source — bundled by default, override wins when present
    # ---------------------------------------------------------------------------
    try:
        analyzer_bytes_cached, analyzer_source_label, analyzer_substrate_ver = _load_analyzer(
            analyzer_override_file
        )
    except FileNotFoundError as e:
        st.error(str(e))
        st.stop()


    # ---------------------------------------------------------------------------
    # Main — empty state
    # ---------------------------------------------------------------------------
    if rr_file is None:
        st.info(f"Using Analyzer: **{analyzer_source_label}** (substrate {analyzer_substrate_ver}). Upload a rent roll to begin.")
        with st.expander("What the app does"):
            st.markdown(
                """
                The webapp runs **four** parallel data tracks that all
                converge on the bundled Analyzer workbook, then feeds the
                Analyzer to a downstream UW Template in one click.

                **Track 1 — Rent Roll Normalizer** *(RR v1.18.1)*

                - Detects the header row in the first ~20 rows.
                - Parses parent-apartment / child-bed layouts: apartment rows
                  establish context, child rows become normalized beds.
                - Auto-groups care charges by header prefix. Recognized
                  buckets (AL care, Med Mgmt, Pharmacy, per-fee ancillary
                  cols Meal/Scooter/HK/Laundry/Pet) get their own columns;
                  unrecognized care rolls into **Other LOC $** so revenue
                  never disappears.
                - Normalizes apt type, bed status (incl. Preleased), payer
                  type, and care level.
                - Preserves vacant beds.
                - Exports a 7-tab standalone normalized Excel for analyst
                  review, AND writes bed-level data to the Analyzer's
                  `Rent Roll Input` sheet (cols A–AJ).

                **Track 2 — T12 Normalizer** *(T12 v0.2.1)*

                - Detects T12 format (Yardi `Income to Budget`, MRI
                  `R12MINCS`, broker-financial-summary).
                - Reads month labels from the source and normalizes to
                  `MMM YYYY`.
                - Drops grand-total rows and explicit non-operating lines.
                - Writes GL detail to the Analyzer's `T12 Input` sheet.
                - Surfaces UNMATCHED descriptions for in-app mapping; new
                  mappings persist in your downloaded Analyzer.
                - Optional `Annualize partial-year T12` toggle for < 12-
                  month inputs.

                **Track 2 — AR & Collections module** *(AR v0.1.0,
                substrate v0.2.10+)*

                - Optional AR aging upload (.xlsx / .csv).
                - Parses bucket totals (Current / 31-60 / 61-90 / 91-120 /
                  120+), payer mix, and roll-forward fields.
                - Writes to the Analyzer's `AR & Collections` sheet
                  (hidden by default; revealed when AR is uploaded).
                - Adds a Workbook Health P5 pre-export gate that flags
                  AR-vs-RR period mismatches.

                **Track 4 — ALF UW Template integration** *(UWT v0.4.1,
                Phase 2.5)*

                - Optional UW Template upload (`ALF_UW_Template_v5.xlsx`
                  or v4).
                - Populates ~95 of 111 mapped concepts from the Analyzer
                  into the template in one click: T-12 Analysis Layer 3
                  (EGI, EBITDARM, EBITDA, full opex line-item map), Prop
                  Info (property name, licensed + occupied beds), Rent
                  Roll Analysis row 211+ (all 176-bed paste path with
                  position-shift handling for v5).
                - Scenario radio: **normalized** (col F, the underwriting
                  figure) vs **t12_actual** (col E, for variance views).
                - Surfaces a per-deal populated UW Template as a second
                  download alongside the populated Analyzer.
                - Drill-in `PopulateReport` expander surfaces warnings,
                  outcome counts, and per-error notes.
                - **⚠️ Cache caveat (READ THIS — known issue):**
                  openpyxl (Python's xlsx library) **doesn't compute
                  Excel formulas**. The Analyzer this app builds has
                  formula text but no cached values. The UW Template
                  writer reads cached values — so on first pass, the
                  populated UW Template's T-12 Analysis Layer 3 (EGI,
                  EBITDARM, opex line items) will be **BLANK**.
                  Workaround: (1) download Analyzer, (2) open in Excel,
                  let it compute, save, (3) upload as **"Analyzer
                  template override"** in the intake panel's Advanced expander,
                  (4) re-download UW Template — now fully populated.
                  An in-Python formula evaluator is on the roadmap to
                  eliminate this round-trip entirely.

                **Combined output:** When you upload RR + T12 (+ optional
                AR + optional UW Template), you get a single populated
                Analyzer with everything reconciled, plus any new T12
                mappings you supplied through the matcher form, plus a
                per-deal populated UW Template if you uploaded one.

                **Dashboard tab** *(Track 5, T5 v0.1.10)*

                Switch to the **Dashboard** tab above for a mobile-
                friendly view of the same headline KPIs the downloaded
                Analyzer surfaces in its `Dashboard` sheet — occupancy,
                EBITDARM margin, going-in cap, RevPOR, payer mix, care-
                type breakdown, 12-month EGI trend.

                **Analyzer source:** The app uses the bundled Analyzer
                (`ALF_Financial_Analyzer_Only.xlsx`, substrate v0.3.0)
                by default. To use a different Analyzer for one session
                — or to feed a pre-Excel-cached Analyzer back through the
                UW Template populate flow — expand
                **"Advanced — override Analyzer template"** in the
                intake panel.
                """
            )
        st.stop()


    # ---------------------------------------------------------------------------
    # Process — Rent Roll
    # ---------------------------------------------------------------------------
    try:
        with _show_loading("Parsing rent roll…"):
            mappings = load_mapping_workbook(mapping_file) if mapping_file else MappingSet()
            result = normalize_rent_roll(
                rr_file,
                sheet_name=sheet_override.strip() or None,
                mappings=mappings,
                property_care_type_default=care_type_default or None,
            )
    except Exception as e:
        st.error(f"Failed to process rent roll: {e}")
        st.stop()

    n = result.normalized
    c = result.condensed

    if n.empty:
        st.warning(
            "No bed rows detected. Check that the file has a parent-apartment / "
            "child-bed layout and that 'Bed' (or a similar column) identifies "
            "child rows."
        )
        st.stop()

    summary    = build_summary(n)
    by_type    = build_by_type(n)
    exceptions = build_exceptions(n, result.unmapped)


    # ---------------------------------------------------------------------------
    # Process — T12 (if uploaded)
    # ---------------------------------------------------------------------------
    # T12 parsing requires the Analyzer's Description_Map. Since the Analyzer is
    # always available now (bundled default + optional override), T12 parsing
    # proceeds whenever a raw T12 is uploaded — no Analyzer-upload prerequisite.
    t12_parse_result = None
    t12_parse_error = None
    descmap_labels_cached: list[str] = []

    if raw_t12_file is not None:
        try:
            analyzer_wb_for_descmap = openpyxl.load_workbook(
                pd.io.common.BytesIO(analyzer_bytes_cached), data_only=True
            )
            descmap = read_descmap_descriptions(analyzer_wb_for_descmap)
            descmap_labels_cached = _read_descmap_labels(analyzer_bytes_cached)
            with _show_loading("Parsing T12…"):
                t12_parse_result = parse_t12(
                    raw_t12_file.getvalue(),
                    descmap,
                    annualize_partial_year=annualize_partial_year,
                )
        except UnknownT12FormatError as e:
            t12_parse_error = (
                f"T12 format not recognized: {e}\n\n"
                "Currently supported: Yardi (Income to Budget), MRI (R12MINCS), "
                "Broker Financial Summary (`Historical Performance` header at A4). "
                "Adding a new format requires extending the format-registry in "
                "t12_normalizer.py — see SPEC-T12.md §\"Parser data flow\"."
            )
        except ValueError as e:
            t12_parse_error = f"T12 parse error: {e}"
        except Exception as e:
            t12_parse_error = f"Could not parse T12: {e}"


    # ---------------------------------------------------------------------------
    # UNMATCHED matcher form — session-state driven
    # ---------------------------------------------------------------------------
    if "t12_resolutions" not in st.session_state:
        st.session_state.t12_resolutions = {}

    unresolved_descriptions: list[str] = []
    if t12_parse_result is not None:
        unresolved_descriptions = [
            d for d in t12_parse_result.unmatched
            if d not in st.session_state.t12_resolutions
        ]


    # ---------------------------------------------------------------------------
    # Headline KPIs
    # ---------------------------------------------------------------------------
    colA, colB, colC, colD, colE = st.columns(5)
    total_beds = len(n)
    occ_beds = int((n["Status"] == "Occupied").sum())
    colA.metric("Total Beds", total_beds)
    colB.metric("Occupied", occ_beds)
    colC.metric(
        "Bed Occupancy",
        f"{100*occ_beds/total_beds:.1f}%" if total_beds else "0.0%",
    )
    colD.metric(
        "Avg Actual (occ)",
        f"${n.loc[n['Status']=='Occupied','Actual Rate'].mean():,.0f}" if occ_beds else "$0",
    )
    colE.metric("In-Place Monthly Rev", f"${n['Total Monthly Revenue'].sum():,.0f}")

    st.caption(
        f"Header detected on row {result.header_row_idx + 1} (1-indexed). "
        f"{len(result.care_groups)} care/ancillary column group(s) identified. "
        f"Analyzer: {analyzer_source_label} (substrate {analyzer_substrate_ver})."
    )

    if result.property_care_type_default:
        default_count = int((n["Care Type Source"] == "Property Default").sum())
        source_count = int((n["Care Type Source"] == "Source").sum())
        st.info(
            f"**Property Care Type default applied: {result.property_care_type_default}** — "
            f"used for {default_count} bed(s) where source had no Care Type. "
            f"{source_count} bed(s) used an explicit source value."
        )


    # ---------------------------------------------------------------------------
    # T12 status panel (only when relevant)
    # ---------------------------------------------------------------------------
    if raw_t12_file is not None:
        st.divider()
        st.subheader("T12 Normalizer")
        if t12_parse_error is not None:
            st.error(t12_parse_error)
        elif t12_parse_result is not None:
            # 5-column layout (was 4, with a duplicate-tc bug). Each metric in its
            # own column so all five display.
            ta, tb, tc, td, te = st.columns(5)
            ta.metric("Format", t12_parse_result.format_name)
            tb.metric("GL Rows Extracted", len(t12_parse_result.gl_rows))
            # Use the most-recent populated label as period. Partial-year files may
            # have leading "" labels (padded); skip those when picking display.
            labels = [lbl for lbl in t12_parse_result.month_labels if lbl]
            first_label = labels[0] if labels else "—"
            last_label = labels[-1] if labels else "—"
            tc.metric("Period (first month)", first_label)
            td.metric("Period (last month)", last_label)
            te.metric(
                "UNMATCHED",
                len(t12_parse_result.unmatched),
                help="Descriptions not found in the Analyzer's Description_Map.",
            )

            # Cluster B (B-2): partial-year detection. Surface as a warning when
            # < 12 months are populated. Annualization (if requested via the checkbox)
            # has already been applied by parse_t12; the warning text reflects that.
            if t12_parse_result.populated_months < 12:
                n = t12_parse_result.populated_months
                if t12_parse_result.was_annualized:
                    st.warning(
                        f"⚠ T12 is partial-year ({n} months populated). Values were "
                        f"scaled by 12/{n} per the annualize checkbox. Ratios assume "
                        f"flat seasonality — review against rent roll occupancy."
                    )
                else:
                    st.warning(
                        f"⚠ T12 is partial-year ({n} months populated). Ratios will "
                        f"be misleading without annualization. Toggle "
                        f"'Annualize partial-year T12' in the intake panel to scale "
                        f"values by 12/{n}, or proceed knowing downstream metrics "
                        f"reflect a {n}-month period."
                    )

            # Cluster B (B-1): sign-convention guards. Defensive — none of the
            # current verified fixtures trip these on standard signs.
            for warning in t12_parse_result.sign_warnings:
                st.warning(warning)

            if t12_parse_result.unmatched:
                n_resolved = len(t12_parse_result.unmatched) - len(unresolved_descriptions)
                if unresolved_descriptions:
                    st.warning(
                        f"⚠️ {len(unresolved_descriptions)} description(s) need mapping "
                        f"before the combined Analyzer download is enabled. "
                        f"({n_resolved} already resolved this session.)"
                    )

                    with st.form("unmatched_matcher", clear_on_submit=False):
                        st.markdown(
                            "**Map these descriptions before download.** Mappings "
                            "will be appended to your Analyzer's Description_Map "
                            "and persist for future uploads of the same operator."
                        )
                        new_resolutions: dict[str, dict] = {}

                        for i, desc in enumerate(unresolved_descriptions):
                            st.markdown(f"**{desc}**")
                            c1, c2, c3, c4 = st.columns([3, 2, 1, 2])
                            with c1:
                                label_options = ["(select…)"] + descmap_labels_cached
                                chosen_label = st.selectbox(
                                    "Label",
                                    options=label_options,
                                    key=f"label_{i}",
                                    label_visibility="collapsed",
                                )
                            with c2:
                                chosen_section = st.selectbox(
                                    "Section",
                                    options=["(select…)"] + DESCMAP_SECTIONS,
                                    key=f"section_{i}",
                                    label_visibility="collapsed",
                                )
                            with c3:
                                chosen_caretype = st.selectbox(
                                    "Care",
                                    options=DESCMAP_CARETYPES,
                                    index=0,
                                    key=f"caretype_{i}",
                                    label_visibility="collapsed",
                                )
                            with c4:
                                chosen_flag = st.selectbox(
                                    "Flag",
                                    options=DESCMAP_FLAGS,
                                    index=0,
                                    key=f"flag_{i}",
                                    label_visibility="collapsed",
                                )
                            new_resolutions[desc] = {
                                "description": desc,
                                "label": None if chosen_label == "(select…)" else chosen_label,
                                "section": None if chosen_section == "(select…)" else chosen_section,
                                "caretype": chosen_caretype,
                                "flag": chosen_flag or None,
                            }

                        submitted = st.form_submit_button(
                            "✓ Apply mappings & enable download",
                            use_container_width=True,
                        )
                        if submitted:
                            bad = [
                                d for d, m in new_resolutions.items()
                                if not m["label"] or not m["section"]
                            ]
                            if bad:
                                st.error(
                                    f"Each row needs a Label and Section. Missing: "
                                    f"{', '.join(bad[:3])}"
                                    f"{'…' if len(bad) > 3 else ''}"
                                )
                            else:
                                st.session_state.t12_resolutions.update(new_resolutions)
                                st.rerun()
                else:
                    st.success(
                        f"✓ All {len(t12_parse_result.unmatched)} UNMATCHED descriptions "
                        "resolved. Combined Analyzer download is enabled."
                    )
            else:
                st.success("✓ Zero UNMATCHED — every description in the T12 already "
                           "maps to a Label.")


    # ---------------------------------------------------------------------------
    # Tabs
    # ---------------------------------------------------------------------------
    st.divider()
    tab_condensed, tab_full, tab_summary, tab_bytype, tab_excep, tab_audit = st.tabs([
        "Condensed RR",
        "Normalized (full)",
        "Summary",
        "By Type",
        "Exceptions",
        "Mapping Audit",
    ])

    with tab_condensed:
        st.subheader("Condensed RR — underwriting view")
        st.caption(
            "Filter and sort columns before exporting. Use the three-dot menu on "
            "any column header to sort. Use the search box above the table to filter."
        )
        st.dataframe(
            c,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Market Rate":   st.column_config.NumberColumn(format="$%.0f"),
                "Actual Rate":   st.column_config.NumberColumn(format="$%.0f"),
                "Concession $":  st.column_config.NumberColumn(format="$%.0f"),
                "Care Level $":  st.column_config.NumberColumn(format="$%.0f"),
                "Med Mgmt $":    st.column_config.NumberColumn(format="$%.0f"),
                "Pharmacy $":    st.column_config.NumberColumn(format="$%.0f"),
                "Other LOC $":   st.column_config.NumberColumn(format="$%.0f"),
            },
        )

    with tab_full:
        st.subheader("Normalized_Beds — full detail")
        st.dataframe(n, use_container_width=True, hide_index=True)

    with tab_summary:
        st.subheader("RR_Summary — KPIs")
        st.dataframe(summary, use_container_width=True, hide_index=True)

    with tab_bytype:
        st.subheader("RR_By_Type — mix analysis")
        st.dataframe(by_type, use_container_width=True, hide_index=True)

    with tab_excep:
        st.subheader("RR_Exceptions — rows needing review")
        if exceptions.empty:
            st.success("No exceptions flagged.")
        else:
            st.warning(f"{len(exceptions)} issue(s) flagged.")
            st.dataframe(exceptions, use_container_width=True, hide_index=True)

    with tab_audit:
        st.subheader("Mapping_Reference — how source columns were classified")
        st.dataframe(result.mapping_audit, use_container_width=True, hide_index=True)
        with st.expander("Detected source headers"):
            st.write(result.source_headers)
        with st.expander("Unmapped values (add to your mapping workbook to clean up)"):
            st.json(result.unmapped)


    # ---------------------------------------------------------------------------
    # Export downloads (Track 1 standalone + Track 2/3 combined Analyzer)
    # ---------------------------------------------------------------------------
    st.divider()
    st.subheader("Export")

    run_meta = {
        "RR Version":          RR_VERSION,
        "RR Last Updated":     RR_LAST_UPDATED,
        "T12 Version":         T12_VERSION,
        "T12 Last Updated":    T12_LAST_UPDATED,
        "Run Timestamp":       dt.datetime.now().isoformat(timespec="seconds"),
        "Source File":         getattr(rr_file, "name", "uploaded"),
        "Mapping File":        getattr(mapping_file, "name", "(defaults only)"),
        "Analyzer Source":     analyzer_source_label,
        "Analyzer Substrate":  analyzer_substrate_ver,
        "Property Care Type Default": result.property_care_type_default or "(none)",
        "Header Row (1-idx)":  result.header_row_idx + 1,
        "Care Groups Detected": len(result.care_groups),
        "Total Beds":          len(n),
        "Occupied Beds":       occ_beds,
        "T12 File":            getattr(raw_t12_file, "name", "(not uploaded)"),
        "T12 Format Detected": t12_parse_result.format_name if t12_parse_result else "(n/a)",
        "T12 GL Rows":         len(t12_parse_result.gl_rows) if t12_parse_result else 0,
    }

    xlsx_bytes = write_output(
        condensed=c,
        normalized=n,
        mapping_audit=result.mapping_audit,
        summary=summary,
        by_type=by_type,
        exceptions=exceptions,
        run_metadata=run_meta,
    )

    out_name = _build_output_name(getattr(rr_file, "name", "rent_roll.xlsx"))

    dl_col1, dl_col2 = st.columns(2)

    # --- Download 1: Standalone Normalized Rent Roll (always available) ---
    with dl_col1:
        st.markdown("**Normalized Rent Roll**")
        st.caption("6-tab analyst workbook with formatting.")
        st.download_button(
            label=f"⬇️ Download {out_name}",
            data=xlsx_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl_rr",
        )

    # --- Download 2: Combined Analyzer (RR + optional T12) ---
    with dl_col2:
        st.markdown("**Analyzer with data**")

        # Gating: rent roll always required. T12 is optional. If T12 is uploaded,
        # all UNMATCHED descriptions must be resolved before download.
        has_t12 = raw_t12_file is not None
        t12_parsed_ok = t12_parse_result is not None
        t12_unmatched_remaining = (
            len([
                d for d in t12_parse_result.unmatched
                if d not in st.session_state.t12_resolutions
            ]) if t12_parsed_ok else 0
        )
        t12_blocking = has_t12 and (not t12_parsed_ok or t12_unmatched_remaining > 0)

        can_download = rr_file is not None and not t12_blocking

        if t12_blocking:
            if not t12_parsed_ok:
                st.caption("T12 parse failed — see error above.")
            else:
                st.caption(f"Resolve {t12_unmatched_remaining} UNMATCHED description(s) above to enable.")
        else:
            t12_caption = (
                f"T12 data → `T12 Input!A12+`. " if has_t12 else ""
            )
            ar_caption = (
                f"AR data → `AR & Collections` (revealed). "
                if ar_file is not None else ""
            )
            st.caption(
                f"RR data → `Rent Roll Input!A7+`. "
                f"{t12_caption}"
                f"{ar_caption}"
                f"Period {period_date_input.isoformat()} written to RR col S."
            )

        if can_download:
            try:
                with _show_loading("Building populated Analyzer…"):
                    # Step 1: Write RR data into the resolved Analyzer.
                    translated = translate_for_t12(c)
                    populated_after_rr = populate_rr_input(
                        analyzer_bytes_cached,
                        translated,
                        period_date_input,
                        source_filename=getattr(rr_file, "name", ""),
                    )

                    # Step 2: If T12 was uploaded, append session-state UNMATCHED
                    # resolutions and write GL detail on top of the RR-populated Analyzer.
                    if has_t12 and t12_parse_result is not None:
                        new_descmap_entries = list(st.session_state.t12_resolutions.values())
                        final_bytes = populate_t12_input(
                            populated_after_rr,
                            t12_parse_result,
                            new_descmap_entries=new_descmap_entries,
                            source_filename=getattr(raw_t12_file, "name", "raw_t12.xlsx"),
                            t12_version=T12_VERSION,
                            t12_last_updated=T12_LAST_UPDATED,
                        )
                    else:
                        final_bytes = populated_after_rr

                    # Step 3: If AR was uploaded, parse it and write to the
                    # AR & Collections sheet on top of the RR(+T12) result.
                    if ar_file is not None:
                        ar_result = parse_ar_file(ar_file)
                        as_of_str = (
                            ar_as_of_override.isoformat()
                            if ar_as_of_override is not None
                            else None
                        )
                        final_bytes = populate_ar_collections(
                            final_bytes,
                            ar_result,
                            as_of_date=as_of_str,
                            source_filename=getattr(ar_file, "name", "ar_aging.xlsx"),
                            ar_version=AR_VERSION,
                        )

                rr_stem = Path(getattr(rr_file, "name", "rent_roll.xlsx")).stem
                name_parts = [rr_stem]
                if has_t12:
                    name_parts.append(Path(getattr(raw_t12_file, "name", "raw_t12.xlsx")).stem)
                if ar_file is not None:
                    name_parts.append("AR")
                combined_out_name = (
                    f"Analyzer with {' + '.join(name_parts)} "
                    f"{period_date_input.isoformat()}.xlsx"
                )

                st.download_button(
                    label=f"⬇️ Download {combined_out_name[:60]}{'…' if len(combined_out_name) > 60 else ''}",
                    data=final_bytes,
                    file_name=combined_out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_combined",
                )

                # ─── Track 4 / Phase 2.5 — populated UW Template download ───
                # Mirrors the Analyzer pattern: bundled template loads by
                # default from `assets/ALF_UW_Template_v5.xlsx`; operator
                # can override via Advanced → "UW Template override". The
                # populate flow fires unconditionally on every successful
                # Analyzer build — no upload required.
                #
                # The writer reads cached formula values from the Analyzer,
                # so the bytes we just generated (via openpyxl, which DOES
                # NOT compute formulas) won't have UW Output values cached.
                # Workaround: warn the user — the populated UW Template will
                # be sparse unless the Analyzer is round-tripped through
                # Excel first. Future enhancement: invoke a formula engine
                # (pycel / formulas) to compute Analyzer values in-Python
                # before the writer reads them.
                st.markdown("---")
                st.markdown("##### 📋 Populate UW Template")
                try:
                    uw_template_bytes, uw_template_source, uw_template_version = (
                        _load_uw_template(uw_template_override_file)
                    )
                    st.caption(
                        f"Using UW Template: **{uw_template_source}** "
                        f"(`{uw_template_version}`)."
                    )

                    # Per-deal filename: <Property>_UW_Template_<period>_<scenario>.xlsx
                    property_name = (
                        derive_property_name(getattr(rr_file, "name", ""))
                        or "Property"
                    )

                    # In-Python UW Output evaluator (kills the cache caveat).
                    # The Analyzer the app just built via openpyxl has formula
                    # *text* but no cached values, so the writer's reads of
                    # `UW Output!{col}{row}` would all come back blank. We
                    # compute those values directly from the parsed RR + T12
                    # (mirroring T12 Analytics) and hand them to the writer as
                    # a fallback. Property name + period date are added here
                    # too (the app already has them; their Analyzer source
                    # cells are formula/blank on a fresh build).
                    uw_computed = compute_uw_output_values(
                        result,
                        t12_parse_result,
                        scenario=uw_template_scenario,
                    )
                    uw_computed.setdefault("property_name", property_name)
                    uw_computed.setdefault("rr_period_date", period_date_input)
                    # Monthly breakdown for the T-12 Analysis Layer-3 grid
                    # (cols B–M). Empty when no T12 — those rows stay blank.
                    uw_monthly = compute_uw_output_monthly(result, t12_parse_result)
                    # Summarized raw T-12 (by label) for Section I / Layer 1.
                    uw_raw_lines = compute_t12_raw_lines(t12_parse_result)

                    with _show_loading("Populating UW Template…"):
                        populated_uw, uw_report = populate_uw_template(
                            final_bytes,
                            uw_template_bytes,
                            scenario=uw_template_scenario,
                            template_version=uw_template_version,
                            computed_values=uw_computed,
                            computed_monthly=uw_monthly,
                            raw_t12_lines=uw_raw_lines,
                        )
                        # Sanitize for filename
                        safe_property = "".join(
                            c if c.isalnum() or c in " -_" else "_"
                            for c in property_name
                        ).strip().replace(" ", "_")
                        uw_out_name = (
                            f"{safe_property}_UW_Template_"
                            f"{period_date_input.isoformat()}_"
                            f"{uw_template_scenario}.xlsx"
                        )

                        # Inline summary
                        n_written = uw_report.summary.get("written", 0)
                        n_cells = uw_report.summary.get("cells_written", 0)
                        n_total = uw_report.summary.get("total_concepts", 0)
                        n_computed = uw_report.summary.get("computed_in_python", 0)
                        n_warn = len(uw_report.warnings)
                        st.caption(
                            f"Writer populated **{n_written} of {n_total}** "
                            f"concepts ({n_cells:,} cells). "
                            f"Scenario: `{uw_template_scenario}`. "
                            + (f"⚠️ {n_warn} warning(s)." if n_warn else "")
                        )

                        # Drill-in expander with the full PopulateReport
                        with st.expander(
                            "🔍 Populate report (details)",
                            expanded=(n_warn > 0),
                        ):
                            if uw_report.warnings:
                                st.markdown("**Warnings:**")
                                for w in uw_report.warnings:
                                    st.markdown(f"- {w}")

                            by_outcome = uw_report.by_outcome()
                            outcome_lines = []
                            for outcome in (
                                "written", "no_source", "skipped",
                                "no_target", "error",
                            ):
                                items = by_outcome.get(outcome, [])
                                if items:
                                    outcome_lines.append(
                                        f"- **{outcome}** — {len(items)} concept(s)"
                                    )
                            if outcome_lines:
                                st.markdown(
                                    "**Outcomes by category:**\n"
                                    + "\n".join(outcome_lines)
                                )

                            errors = by_outcome.get("error", [])
                            if errors:
                                st.markdown("**Errors:**")
                                for r in errors:
                                    st.markdown(
                                        f"- `{r.key}` → `{r.target_address}` — {r.notes}"
                                    )

                        st.download_button(
                            label=(
                                f"⬇️ Download {uw_out_name[:60]}"
                                f"{'…' if len(uw_out_name) > 60 else ''}"
                            ),
                            data=populated_uw,
                            file_name=uw_out_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="dl_uw_template",
                        )

                        # In-Python evaluator status. The cache caveat is now
                        # handled: the writer's UW-Output reads on a freshly
                        # built (openpyxl, no cached values) Analyzer fall back
                        # to values computed directly from the parsed RR + T12.
                        # No Excel round-trip needed. Surface a success line
                        # when the fallback did the work; only warn if T-12
                        # values still came through blank (e.g. no T12
                        # uploaded, or an analyst-override Analyzer with an
                        # unexpectedly empty UW Output).
                        n_monthly = uw_report.summary.get("monthly_cells_written", 0)
                        if n_computed > 0:
                            monthly_note = (
                                f" Plus **{n_monthly} monthly cells** across the "
                                f"T-12 Analysis Apr→Mar grid."
                                if n_monthly else ""
                            )
                            st.success(
                                f"✅ **{n_computed} UW Output value(s) computed "
                                f"in-Python** (EGI, EBITDARM, EBITDA, opex line "
                                f"items, bed counts, …) — the T-12 Analysis tab "
                                f"is populated directly from the parsed RR + "
                                f"T12. No Excel round-trip required." + monthly_note
                            )
                        t12_no_source = [
                            r for r in uw_report.results
                            if r.outcome == "no_source" and r.path == "t12"
                        ]
                        if t12_no_source and not has_t12:
                            st.info(
                                f"ℹ️ {len(t12_no_source)} T-12 Analysis value(s) "
                                f"are blank because no Raw T12 was uploaded. "
                                f"Upload a T12 to populate EGI / EBITDARM / opex "
                                f"line items."
                            )
                        elif t12_no_source:
                            st.info(
                                f"ℹ️ {len(t12_no_source)} T-12 value(s) came "
                                f"through blank. If you uploaded an Analyzer "
                                f"override, its `UW Output` cells may be empty — "
                                f"the in-Python evaluator only fills gaps when "
                                f"the parsed RR + T12 are available."
                            )

                except UWTemplateWriterError as e:
                    st.error(f"UW Template populate failed: {e}")
                except Exception as e:
                    st.error(f"Could not populate UW Template: {e}")
            except AnalyzerRRCapacityError as e:
                st.error(f"Rent Roll exceeds Analyzer capacity: {e}")
            except T12NormalizerCapacityError as e:
                st.error(f"T12 exceeds Analyzer capacity: {e}")
            except AROutputError as e:
                st.error(
                    f"Analyzer override is missing the 'AR & Collections' sheet "
                    f"(substrate v0.2.10+ required to use AR upload). {e}"
                )
            except ValueError as e:
                st.error(f"Analyzer / T12 / AR error: {e}")
            except Exception as e:
                st.error(f"Could not produce combined output: {e}")
        else:
            st.button(
                "⬇️ Combined download not yet available",
                disabled=True,
                use_container_width=True,
                key="dl_combined_disabled",
            )


# ---------------------------------------------------------------------------
# Top-level Dashboard tab — clean-slate render (Track 5)
# ---------------------------------------------------------------------------
with top_tab_dashboard:
    if rr_file is None:
        st.info("Upload a Rent Roll in the Workspace tab's intake panel to populate the dashboard.")
    elif t12_parse_result is None:
        st.info(
            "Rent Roll is parsed — upload a T12 in the intake panel "
            "to populate the financial metrics on the dashboard."
        )
    else:
        try:
            _period_lbl = (
                t12_parse_result.month_labels[-1]
                if t12_parse_result.month_labels and t12_parse_result.month_labels[-1]
                else period_date_input.isoformat()
            )
            _property_name = (
                derive_property_name(getattr(rr_file, "name", "")) or "Property"
            )
            _purchase_price = (
                float(purchase_price_input) if purchase_price_input and purchase_price_input > 0 else None
            )
            _model = compute_dashboard(
                rr_result=result,
                t12_result=t12_parse_result,
                ar_result=None,
                property_name=_property_name,
                period_label=_period_lbl,
                purchase_price=_purchase_price,
            )
            render_dashboard(_model)
        except Exception as e:  # noqa: BLE001
            st.error(f"Dashboard could not be rendered: {e}")
            st.caption(
                "This shouldn't happen — the downloaded Analyzer will still work. "
                "Please report the error above."
            )
