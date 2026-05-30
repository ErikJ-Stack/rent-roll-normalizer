"""
Fix a third v6-template repoint miss (same class as the two closed in UWT
v0.7.1) — surfaced 2026-05-30 on a populated Briar Glen output where the
T-12 Analysis Section D income-summary cells read $0.

  Section D (ECONOMIC vs. PHYSICAL OCCUPANCY RECONCILIATION) cells B22/B23/B24
  still reference the v5 income-block rows:
      B22  Gross Potential Rent (GPR)   =N58
      B23  Net Rent Revenue (billed)    =N63
      B24  Total Revenue / EGI          =N69
  In v5, N58=GPR, N63=Net Rent, N69=EGI — correct. But v6 rebuilt the income
  section (actual-T-12 build first, GPR waterfall demoted to a DIAGNOSTIC
  sub-block at N80-83, EGI moved up to N77), so in v6:
      N58 = "Base Rent — IL"      (→ should be GPR  at N80)
      N63 = "LOC / Care — AL"     (→ should be Net Rent at N83)
      N69 = "Meal Income"         (→ should be EGI at N77)
  all of which are $0 for a typical deal → the three Section-D headline cells
  (and B25 Economic Occupancy %, which divides B23/B22) read 0. The operator's
  openpyxl repointing pass repointed the EGI chain at B5/B9/B11 (N69→N77) but
  missed the Section-D chain — exactly the openpyxl-quirk-#4 partial-repoint
  that also left B56:M56 stale (fixed in v0.7.1).

  Fix: B22 =N80, B23 =N83, B24 =N77. B25 (=IFERROR(B23/B22,0)) then works.

Like v0.7.1, the openpyxl edit strips xl/metadata.xml, so we re-restore it +
the per-cell `cm` markers from assets/ALF_UW_Template_v5.xlsx via the proven
`_restore_dynamic_arrays` post-processor (verified-identical Section R/S
layout). NOT restored: xl/webextensions/ (Claude add-in — re-installable).

Idempotent: re-running is safe (cells already repointed → no change; metadata
already present → re-injected identically).
"""

from __future__ import annotations

import io
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
V6 = ROOT / "assets" / "ALF_UW_Template_v6.xlsx"
V5 = ROOT / "assets" / "ALF_UW_Template_v5.xlsx"

sys.path.insert(0, str(ROOT))
from uw_template_writer import _restore_dynamic_arrays  # noqa: E402

# Section-D income-summary repoints: cell -> (stale v5 ref, correct v6 ref, v6 label)
REPOINTS = {
    "B22": ("=N58", "=N80", "Gross Potential Rent (GPR)"),
    "B23": ("=N63", "=N83", "Net Rent (projected)"),
    "B24": ("=N69", "=N77", "EFFECTIVE GROSS INCOME (EGI)"),
}


def fix() -> int:
    wb = openpyxl.load_workbook(V6)
    t12 = wb["T-12 Analysis"]

    # --- Pre-flight: confirm the v6 target rows actually carry the labels we
    #     expect, so we never repoint blindly into a moved layout. ---
    expected_label = {
        "=N80": "Gross Potential Rent (GPR)",
        "=N83": "Net Rent (projected)",
        "=N77": "EFFECTIVE GROSS INCOME (EGI)",
    }
    preflight_ok = True
    for cell, (stale, correct, _lbl) in REPOINTS.items():
        target_row = int(correct.lstrip("=N"))
        actual = str(t12.cell(row=target_row, column=1).value or "").strip()
        want = expected_label[correct]
        ok = actual == want
        preflight_ok = preflight_ok and ok
        print(f"  preflight {correct}: A{target_row}={actual!r} "
              f"(want {want!r}) {'✓' if ok else '✗'}")
    if not preflight_ok:
        print("✗ Pre-flight failed — v6 layout differs from expected; aborting.")
        return 1

    # --- Step 1: repoint B22/B23/B24 ---
    fixed = []
    for cell, (stale, correct, _lbl) in REPOINTS.items():
        cur = t12[cell].value
        if cur != correct:
            t12[cell].value = correct
            fixed.append(f"{cell}: {cur!r} → {correct!r}")
    buf = io.BytesIO()
    wb.save(buf)
    fixed_bytes = buf.getvalue()
    print(f"\nStep 1 — Section D repoint: {len(fixed)} cell(s) changed")
    for f in fixed:
        print(f"  ✓ {f}")
    if not fixed:
        print("  (already repointed — no change)")

    # --- Step 2: restore xl/metadata.xml + cm markers from v5 ---
    v5_bytes = V5.read_bytes()
    restored = _restore_dynamic_arrays(fixed_bytes, v5_bytes)
    V6.write_bytes(restored)
    print(f"\nStep 2 — metadata.xml restore from v5: "
          f"{'applied' if restored is not fixed_bytes else 'NO-OP (v5 had no metadata?)'}")

    # --- Verify ---
    print("\nVerify:")
    checks = []
    with zipfile.ZipFile(V6) as z:
        parts = z.namelist()
        checks.append(("xl/metadata.xml present", "xl/metadata.xml" in parts))
        sheet8 = [p for p in parts if p.endswith("sheet8.xml")]
        cm_count = z.read(sheet8[0]).decode("utf-8").count('cm="') if sheet8 else 0
        checks.append((f"Section R/S cm markers present ({cm_count})", cm_count > 100))
    wb2 = openpyxl.load_workbook(V6)
    t = wb2["T-12 Analysis"]
    checks.append(("B22 == =N80 (GPR)", t["B22"].value == "=N80"))
    checks.append(("B23 == =N83 (Net Rent)", t["B23"].value == "=N83"))
    checks.append(("B24 == =N77 (EGI)", t["B24"].value == "=N77"))
    checks.append(("B25 econ-occ formula intact", t["B25"].value == "=IFERROR(B23/B22,0)"))
    # the v0.7.1 fixes must still be present (we didn't regress them)
    checks.append(("B56 == =C137 (v0.7.1)", t["B56"].value == "=C137"))
    checks.append(("M56 == =N137 (v0.7.1)", t["M56"].value == "=N137"))
    checks.append(("EGI N77 formula intact", str(t["N77"].value) == "=N61+N65+SUM(N66:N76)"))
    rra = wb2["Rent Roll Analysis"]
    checks.append(("Section R Z173 still ArrayFormula", hasattr(rra["Z173"].value, "text")))
    checks.append(("sheet count == 16", len(wb2.sheetnames) == 16))

    ok_all = True
    for label, ok in checks:
        ok_all = ok_all and ok
        print(f"  {'✓' if ok else '✗'} {label}")
    if not ok_all:
        return 1
    with zipfile.ZipFile(V6) as z:
        print(f"\nv6 now {len(z.namelist())} parts. "
              f"webextensions NOT restored (operator re-adds the Claude add-in if used).")
    print("\nAll checks pass.")
    return 0


if __name__ == "__main__":
    sys.exit(fix())
