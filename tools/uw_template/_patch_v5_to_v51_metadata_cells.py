"""
Patch assets/ALF_UW_Template_v5.xlsx with the two v5.1 metadata cells.

Surface-only edits above the data band:
  1. Cover!G1 = "Substrate:" (label) + Cover!H1 styled placeholder for the
     writer to populate from Analyzer's Cover!B8.  (E1/F1 unavailable —
     A1:F1 is the merged title band.)
  2. Rent Roll Analysis!B5 styled as a date cell (mm/dd/yyyy) for the writer
     to populate from RR_Period_Date.

Idempotent — re-running is a no-op on already-patched files.

Defensive: prints a pre/post fidelity report (sheet count, chart count,
spot-checks on critical formula cells from v0.4.3 patch).
"""

from __future__ import annotations

import sys
from pathlib import Path
from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "assets" / "ALF_UW_Template_v5.xlsx"
BACKUP = ROOT / "assets" / "ALF_UW_Template_v5.xlsx.bak-pre-v51"

LABEL_TEXT = "Substrate:"
LABEL_FONT = Font(
    name="Calibri",
    size=9,
    italic=True,
    color="FF595959",
)
VALUE_FONT = Font(
    name="Calibri",
    size=9,
    italic=True,
    color="FF595959",
    bold=True,
)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


def fidelity_snapshot(path: Path) -> dict:
    """Read fidelity-critical attributes of the template for before/after comparison."""
    wb = openpyxl.load_workbook(path, data_only=False)
    snap = {
        "sheet_count": len(wb.sheetnames),
        "sheets": list(wb.sheetnames),
    }
    chart_total = 0
    for ws in wb.worksheets:
        n = len(ws._charts) if hasattr(ws, "_charts") else 0
        chart_total += n
    snap["chart_total"] = chart_total

    # Spot-check critical cells from v0.4.3 patch (Section R/S formula fill-downs)
    rra = wb["Rent Roll Analysis"]
    snap["Z173"] = type(rra["Z173"].value).__name__  # should be ArrayFormula
    snap["C173"] = type(rra["C173"].value).__name__  # should be ArrayFormula
    snap["W211"] = rra["W211"].value
    snap["W610"] = rra["W610"].value
    snap["W611"] = rra["W611"].value  # must remain None — guard
    snap["AC211"] = rra["AC211"].value

    # T-12 Analysis monthly header formulas (v5 feature)
    t12 = wb["T-12 Analysis"]
    snap["B56"] = t12["B56"].value
    snap["M56"] = t12["M56"].value

    return snap


def patch(path: Path) -> dict:
    """Apply the two v5.1 cell additions. Returns operation report."""
    wb = openpyxl.load_workbook(path, data_only=False)
    report = {"operations": [], "skipped": []}

    # === Patch 1: Cover!G1 + H1 ===
    # E1/F1 unavailable — A1:F1 is the merged title band.
    cover = wb["Cover"]
    g1 = cover["G1"]
    h1 = cover["H1"]

    # Idempotency: if G1 already says "Substrate:", we've already patched
    if g1.value == LABEL_TEXT:
        report["skipped"].append("Cover!G1 already set — no-op")
    else:
        g1.value = LABEL_TEXT
        g1.font = LABEL_FONT
        g1.alignment = ALIGN_RIGHT
        report["operations"].append("Set Cover!G1 = 'Substrate:' (italic gray 9pt, right-aligned)")

    if h1.value is None:
        # Style the empty cell so the writer's eventual paste gets the right look
        h1.font = VALUE_FONT
        h1.alignment = ALIGN_LEFT
        report["operations"].append("Styled Cover!H1 (italic gray 9pt bold, left-aligned) — placeholder for writer-populated substrate version")
    else:
        report["skipped"].append(f"Cover!H1 already has value {h1.value!r} — left alone")

    # === Patch 2: Rent Roll Analysis!B5 ===
    rra = wb["Rent Roll Analysis"]
    b5 = rra["B5"]
    if b5.value is None:
        b5.number_format = "mm/dd/yyyy"
        # Match the visual style of D5 (=TODAY()) — Calibri default
        b5.font = Font(name="Calibri", size=11)
        b5.alignment = Alignment(horizontal="left", vertical="center")
        report["operations"].append("Styled Rent Roll Analysis!B5 (mm/dd/yyyy, Calibri 11) — placeholder for writer-populated RR period")
    else:
        report["skipped"].append(f"Rent Roll Analysis!B5 already has value {b5.value!r} — left alone")

    if report["operations"]:
        wb.save(path)
        report["saved"] = True
    else:
        report["saved"] = False

    return report


def main():
    print(f"Pre-patch fidelity snapshot of {TEMPLATE.name}:")
    pre = fidelity_snapshot(TEMPLATE)
    for k, v in pre.items():
        print(f"  {k:14s} = {v!r}")

    print(f"\nApplying patch...")
    report = patch(TEMPLATE)
    for op in report["operations"]:
        print(f"  ✓ {op}")
    for sk in report["skipped"]:
        print(f"  · {sk}")
    print(f"  Saved: {report['saved']}")

    print(f"\nPost-patch fidelity snapshot:")
    post = fidelity_snapshot(TEMPLATE)
    for k, v in post.items():
        print(f"  {k:14s} = {v!r}")

    print(f"\nDelta check:")
    drift = False
    for k in pre:
        if k in ("Z173", "C173", "B56", "M56", "W211", "W610", "AC211", "sheet_count", "chart_total", "sheets"):
            if pre[k] != post[k]:
                print(f"  ✗ {k}: {pre[k]!r} → {post[k]!r}")
                drift = True
    if not drift:
        print(f"  ✓ All critical attributes preserved (sheets, charts, ArrayFormula objects, v5 monthly headers, v0.4.3 W formulas).")

    # Confirm new cells landed
    print(f"\nNew cells:")
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    print(f"  Cover!G1                  = {wb['Cover']['G1'].value!r} | font={wb['Cover']['G1'].font.size}pt italic={wb['Cover']['G1'].font.italic} color={wb['Cover']['G1'].font.color.rgb if wb['Cover']['G1'].font.color else None}")
    print(f"  Cover!H1                  = {wb['Cover']['H1'].value!r} | number_format={wb['Cover']['H1'].number_format!r}")
    print(f"  Rent Roll Analysis!A5     = {wb['Rent Roll Analysis']['A5'].value!r} (unchanged)")
    print(f"  Rent Roll Analysis!B5     = {wb['Rent Roll Analysis']['B5'].value!r} | number_format={wb['Rent Roll Analysis']['B5'].number_format!r}")

    return 0 if not drift else 1


if __name__ == "__main__":
    sys.exit(main())
