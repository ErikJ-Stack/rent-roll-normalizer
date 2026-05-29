"""
Fix the two v6-template bugs caught during the UWT v0.7.0 absorption:

  1. Layer-3 monthly headers B56:M56 still reference =C122..=N122, but the raw
     month-header row moved to 137 (Section I shifted +15 in the income
     restructure). The operator's openpyxl repointing pass missed this chain
     (openpyxl quirk #4). Fix: B56=C137, C56=D137, ..., M56=N137.

  2. The v6 binary is the pre-Excel-resave openpyxl version (39 zip parts,
     missing xl/metadata.xml → Section R/S dynamic-array spills degrade in
     Excel). Restore xl/metadata.xml + the per-cell `cm` markers via the
     proven `_restore_dynamic_arrays` post-processor, sourcing from
     assets/ALF_UW_Template_v5.xlsx (= v5.1 content; identical Rent Roll
     Analysis / Section R-S layout — verified cell-by-cell before running).

Order matters: do the openpyxl B56 edit FIRST (it strips metadata anyway),
then restore metadata via zipfile.

NOT restored: xl/webextensions/ (the Claude-for-Excel taskpane add-in). It
carries v5's add-in fileId GUID and is re-installable by the operator opening
the add-in — not fabricated here.

Idempotent: re-running is safe (B56 already =C137 → no change; metadata already
present → _restore_dynamic_arrays still re-injects identically).
"""

from __future__ import annotations

import io
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
V6 = ROOT / "assets" / "ALF_UW_Template_v6.xlsx"
V5 = ROOT / "assets" / "ALF_UW_Template_v5.xlsx"

sys.path.insert(0, str(ROOT))
from uw_template_writer import _restore_dynamic_arrays  # noqa: E402


def fix() -> int:
    # --- Step 1: B56:M56 monthly-header repoint (row 122 → 137) ---
    wb = openpyxl.load_workbook(V6)
    t12 = wb["T-12 Analysis"]
    # B56..M56 = cols 2..13; each references col (n+1) at the raw-header row.
    from openpyxl.utils import get_column_letter
    fixed = []
    for col in range(2, 14):  # B(2)..M(13)
        cell = t12.cell(row=56, column=col)
        ref_col = get_column_letter(col + 1)  # B→C, ..., M→N
        want = f"={ref_col}137"
        cur = cell.value
        if cur != want:
            cell.value = want
            fixed.append(f"{get_column_letter(col)}56: {cur!r} → {want!r}")
    buf = io.BytesIO()
    wb.save(buf)
    fixed_bytes = buf.getvalue()
    print(f"Step 1 — B56:M56 header repoint: {len(fixed)} cells changed")
    for f in fixed:
        print(f"  ✓ {f}")

    # --- Step 2: restore xl/metadata.xml + cm markers from v5 ---
    v5_bytes = V5.read_bytes()
    restored = _restore_dynamic_arrays(fixed_bytes, v5_bytes)
    V6.write_bytes(restored)
    print(f"\nStep 2 — metadata.xml restore from v5: "
          f"{'applied' if restored is not fixed_bytes else 'NO-OP (v5 had no metadata?)'}")

    # --- Verify ---
    print("\nVerify:")
    checks = []
    with zipfile.ZipFile(V6) as z:
        parts = z.namelist()
        checks.append(("xl/metadata.xml present", "xl/metadata.xml" in parts))
        sheet8 = [p for p in parts if p.endswith("sheet8.xml")]
        cm_count = 0
        if sheet8:
            cm_count = z.read(sheet8[0]).decode("utf-8").count('cm="')
        checks.append((f"Section R/S cm markers present ({cm_count})", cm_count > 100))
    wb2 = openpyxl.load_workbook(V6)
    t = wb2["T-12 Analysis"]
    checks.append(("B56 == =C137", t["B56"].value == "=C137"))
    checks.append(("M56 == =N137", t["M56"].value == "=N137"))
    # formulas still intact (income restructure + EGI)
    checks.append(("EGI N77 formula intact", str(t["N77"].value) == "=N61+N65+SUM(N66:N76)"))
    checks.append(("Total Base N61 formula intact", str(t["N61"].value) == "=SUM(N58:N60)"))
    rra = wb2["Rent Roll Analysis"]
    z173 = rra["Z173"].value
    checks.append(("Section R Z173 still ArrayFormula", hasattr(z173, "text")))
    checks.append(("sheet count == 16", len(wb2.sheetnames) == 16))

    ok_all = True
    for label, ok in checks:
        ok_all = ok_all and ok
        print(f"  {'✓' if ok else '✗'} {label}")
    if not ok_all:
        return 1
    with zipfile.ZipFile(V6) as z:
        print(f"\nv6 now {len(z.namelist())} parts (was 39). "
              f"webextensions NOT restored (operator re-adds the Claude add-in if used).")
    print("\nAll checks pass.")
    return 0


if __name__ == "__main__":
    sys.exit(fix())
