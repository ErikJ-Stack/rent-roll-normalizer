"""
Fix the Section-D repoint miss in the operator's v6 *rev2* template (the
Other-Care revision absorbed 2026-06-03 as the new canonical asset).

Same bug class as the 2026-05-30 fix, but rev2's income restructure shifted the
correct target rows further AND the rev2 file's Section-D cells point at yet
different stale rows:

  Section D (ECONOMIC vs. PHYSICAL OCCUPANCY) — rev2 current → correct:
      B22  Gross Potential Rent (GPR)   =N58 (Base Rent IL)   → =N83  (GPR)
      B23  Net Rent Revenue (billed)    =N64 (LOC AL)         → =N86  (Net Rent projected)
      B24  Total Revenue / EGI          =N71 (Meal Income)    → =N80  (EGI)
  B25 (=IFERROR(B23/B22,0)) then resolves correctly.

Every other diagnostic ref in rev2 (B5/B9 EGI=N80, B13 Labor=N102, B14 Food=N104,
B15 Utilities=N113, B16 Overtime=N96, B12 Bad Debt=N124, D5 OpEx=N132, F5 EBITDAR=N135)
the operator repointed correctly — only Section D was missed.

openpyxl quirk #6: wb.save() strips xl/metadata.xml. We restore rev2's OWN metadata
(not v5's — rev2's Rent Roll Analysis layout differs, so v5's cm anchor cells would
not match). Faithful: we only edit B22/B23/B24, never a dynamic-array anchor cell.

Patches BOTH the committed canonical asset AND the operator's local ALF Templates copy.
Idempotent. Pre-flight aborts if the target rows don't carry the expected labels.
"""
from __future__ import annotations

import io
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
ASSET = ROOT / "assets" / "ALF_UW_Template_v6.xlsx"
DEALS = Path("/Users/erikjavellana/Library/CloudStorage/OneDrive-(na)/"
             "Deals/Acquisition/_Template/ALF Templates/ALF_UW_Template_v6.xlsx")

sys.path.insert(0, str(ROOT))
from uw_template_writer import _restore_dynamic_arrays  # noqa: E402

# cell -> (correct v6-rev2 ref, expected label at that row)
REPOINTS = {
    "B22": ("=N83", "Gross Potential Rent (GPR)"),
    "B23": ("=N86", "Net Rent (projected)"),
    "B24": ("=N80", "EFFECTIVE GROSS INCOME (EGI)"),
}


def fix_one(path: Path) -> int:
    original = path.read_bytes()
    wb = openpyxl.load_workbook(path)
    t12 = wb["T-12 Analysis"]

    # Pre-flight: confirm target rows carry the expected labels.
    ok = True
    for cell, (correct, want) in REPOINTS.items():
        row = int(correct.lstrip("=N"))
        actual = str(t12.cell(row=row, column=1).value or "").strip()
        hit = actual == want
        ok = ok and hit
        print(f"  preflight {correct}: A{row}={actual!r} (want {want!r}) {'✓' if hit else '✗'}")
    if not ok:
        print(f"✗ Pre-flight failed for {path.name}; aborting (no change).")
        return 1

    changed = []
    for cell, (correct, _want) in REPOINTS.items():
        if t12[cell].value != correct:
            changed.append(f"{cell}: {t12[cell].value!r} → {correct!r}")
            t12[cell].value = correct

    buf = io.BytesIO()
    wb.save(buf)
    restored = _restore_dynamic_arrays(buf.getvalue(), original)  # restore rev2's OWN metadata
    path.write_bytes(restored)

    print(f"  repointed {len(changed)} cell(s): " + ("; ".join(changed) or "(already correct)"))

    # Verify
    wb2 = openpyxl.load_workbook(path)
    t = wb2["T-12 Analysis"]
    checks = [
        ("B22 == =N83", t["B22"].value == "=N83"),
        ("B23 == =N86", t["B23"].value == "=N86"),
        ("B24 == =N80", t["B24"].value == "=N80"),
        ("B25 econ-occ intact", t["B25"].value == "=IFERROR(B23/B22,0)"),
        ("sheet count 16", len(wb2.sheetnames) == 16),
    ]
    with zipfile.ZipFile(path) as z:
        checks.append(("metadata.xml present", "xl/metadata.xml" in z.namelist()))
    all_ok = True
    for label, c in checks:
        all_ok = all_ok and c
        print(f"    {'✓' if c else '✗'} {label}")
    return 0 if all_ok else 1


def main() -> int:
    rc = 0
    for label, path in [("CANONICAL ASSET", ASSET), ("LOCAL ALF TEMPLATES", DEALS)]:
        print(f"\n=== {label}: {path} ===")
        if not path.exists():
            print(f"  ✗ missing: {path}")
            rc = 1
            continue
        rc |= fix_one(path)
    print("\n" + ("All patches verified." if rc == 0 else "✗ One or more patches failed."))
    return rc


if __name__ == "__main__":
    sys.exit(main())
