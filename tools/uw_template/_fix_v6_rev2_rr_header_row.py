"""
Restore the Rent Roll Analysis row-210 header row in the operator's v6 rev2
template. v5 carries a full 47-column header (Unit/Bed … Effective Conc $) at
row 210, but the operator's rev2 authoring dropped it entirely — so the writer
pastes data at 211+ with no column titles above it and emits the "A210 is blank"
warning. Rent Roll Analysis is unchanged v5→v6 (both 47 cols), so v5's header
row maps 1:1.

Copies value + style (font/fill/border/alignment/number_format) from
assets/ALF_UW_Template_v5.xlsx row 210 into rev2 row 210.

openpyxl quirk #6: wb.save() strips xl/metadata.xml — restore rev2's OWN metadata
after the edit (we touch only row 210, never a dynamic-array anchor cell).

Patches BOTH the canonical asset AND the operator's local ALF Templates copy.
Idempotent (re-run overwrites the same labels). Pre-flight aborts if the column
count differs between v5 and the target.
"""
from __future__ import annotations

import copy
import io
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
ASSET = ROOT / "assets" / "ALF_UW_Template_v6.xlsx"
V5 = ROOT / "assets" / "ALF_UW_Template_v5.xlsx"
DEALS = Path("/Users/erikjavellana/Library/CloudStorage/OneDrive-(na)/"
             "Deals/Acquisition/_Template/ALF Templates/ALF_UW_Template_v6.xlsx")
SHEET = "Rent Roll Analysis"
HEADER_ROW = 210


def fix_one(path: Path, v5_ws) -> int:
    original = path.read_bytes()
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET]

    if ws.max_column != v5_ws.max_column:
        print(f"  ✗ column count differs (v5 {v5_ws.max_column} vs target "
              f"{ws.max_column}); aborting {path.name}.")
        return 1

    copied = 0
    for c in range(1, v5_ws.max_column + 1):
        src = v5_ws.cell(HEADER_ROW, c)
        if src.value is None:
            continue
        dst = ws.cell(HEADER_ROW, c)
        dst.value = src.value
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format
        copied += 1

    buf = io.BytesIO()
    wb.save(buf)
    sys.path.insert(0, str(ROOT))
    from uw_template_writer import _restore_dynamic_arrays  # noqa: E402
    restored = _restore_dynamic_arrays(buf.getvalue(), original)
    path.write_bytes(restored)

    # verify
    wb2 = openpyxl.load_workbook(path)
    w = wb2[SHEET]
    a210 = w["A210"].value
    au210 = w["AU210"].value
    with zipfile.ZipFile(path) as z:
        has_md = "xl/metadata.xml" in z.namelist()
    ok = a210 == "Unit/Bed" and au210 == "Effective Conc $" and has_md and len(wb2.sheetnames) == 16
    print(f"  copied {copied} header cells | A210={a210!r} AU210={au210!r} "
          f"metadata={'✓' if has_md else '✗'} {'✓' if ok else '✗'}")
    return 0 if ok else 1


def main() -> int:
    v5_ws = openpyxl.load_workbook(V5)[SHEET]
    rc = 0
    for label, path in [("CANONICAL ASSET", ASSET), ("LOCAL ALF TEMPLATES", DEALS)]:
        print(f"\n=== {label}: {path} ===")
        if not path.exists():
            print(f"  ✗ missing: {path}")
            rc = 1
            continue
        rc |= fix_one(path, v5_ws)
    print("\n" + ("Header row restored." if rc == 0 else "✗ failed."))
    return rc


if __name__ == "__main__":
    sys.exit(main())
