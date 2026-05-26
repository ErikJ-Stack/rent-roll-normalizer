"""
Patch ALF_UW_Template_v5.xlsx — fill W/X/Y formula-column fill-downs
through row 610 to unblock Section R / Section S diagnostics.

THE BUG
  Operator-reported 2026-05-26: Section R (rows 170-181 "Unit Type Pricing
  by Care Level") on Rent Roll Analysis returns #CALC! everywhere because
  its dynamic-array driver at Z173:

      =SORT(UNIQUE(FILTER($X$211:$X$610,$X$211:$X$610<>"")))

  evaluates to #CALC! when X211:X610 is all empty (FILTER with no matches
  and no if_empty argument). That #CALC! poisons every spillover cell in
  C173:Q179. Same shape on Section S (rows 182-188) reading $AU$211:$AU$610.

  Root cause: the v5 template author left W/X/Y data-row formulas blank
  when cleaning the template. The Section R/S ArrayFormulas at row 173 are
  intact (verified — Z173 is an `openpyxl.worksheet.formula.ArrayFormula`)
  but they read from columns that have no fill-downs.

THE FIX (this script)
  Fill W/X/Y formulas for rows 211 through 610:

    W{r} = =AC{r}                         # Mirror AC (Apt Type), which the
                                          # writer already pastes from
                                          # Analyzer col F (normalized to
                                          # closed vocab: Studio /
                                          # 1 Bedroom / 2 Bedroom / 3 Bedroom
                                          # / Suite / Cottage). Required by
                                          # Section R's SqFt lookup at row
                                          # 171 (per template comment in
                                          # A171).
    X{r} = =IF(AND(D{r}="Occupied",       # Care|UnitType — OCCUPIED ONLY.
              C{r}<>"",W{r}<>""),          # Drives Section R's unique-key
              C{r}&"|"&W{r},"")            # SORT/UNIQUE/FILTER at Z173.
    Y{r} = =IF(AND(C{r}<>"",W{r}<>""),    # Care|Unit (all rows incl
              C{r}&"|"&W{r},"")            # vacants). Drives Section R
                                          # denominators (column population).

  Total: 400 rows × 3 cols = 1,200 formula cells.

WHAT THIS SCRIPT DOES NOT DO
  - Column AU (Conc Source) is left empty. The 2026-05-25 handoff contract
    §11 calls AU "Manual analyst entry," and the operator's diagnostic note
    flagged the auto-classifier sketch as "rough — confirm column meanings
    before using." Section S will continue to show 0 counts until analyst
    data is entered there. Auto-classification is a v5.1 template addition,
    not part of this patch.
  - Columns Z (_key), AA (Mkt-Actual $), AB (Mkt-Actual %) are also empty
    per contract §13's formula-derived list, but the operator hasn't
    reported them as blocking any section. Left alone.

IDEMPOTENCY
  Bails as a no-op if W211 already holds a formula or value. Re-runs are
  safe.

USAGE
  python tools/uw_template/_patch_v5_section_r_formulas.py
      [path/to/ALF_UW_Template_v5.xlsx]

  If no path is given, patches the committed asset at
  `assets/ALF_UW_Template_v5.xlsx`. The operator should also run this on
  their Deals-folder canonical copy if they author against it directly.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

DEFAULT_TARGET = Path(__file__).resolve().parent.parent.parent / "assets" / "ALF_UW_Template_v5.xlsx"

# v5 capacity: 400 unit-rows from 211 through 610 inclusive
START_ROW = 211
END_ROW = 610
SHEET_NAME = "Rent Roll Analysis"


def patch(template_path: Path) -> dict:
    """Idempotent: fill W/X/Y formulas in rows 211-610.

    Returns a dict with counts + a brief log.
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet {SHEET_NAME!r} not in template (got {wb.sheetnames})")
    ws = wb[SHEET_NAME]

    counts = {"w_filled": 0, "x_filled": 0, "y_filled": 0, "skipped": 0}
    log: list[str] = []

    # Idempotency gate
    if ws["W211"].value not in (None, ""):
        counts["skipped"] = 1
        log.append(
            f"SKIP: W211 already populated with {ws['W211'].value!r} — "
            f"no-op (re-run safe)"
        )
        return {"counts": counts, "log": log, "written": False}

    for r in range(START_ROW, END_ROW + 1):
        # W{r} = AC{r}  — mirror Apt Type, which is already writer-populated
        ws[f"W{r}"] = f"=AC{r}"
        counts["w_filled"] += 1

        # X{r}: Care|UnitType, occupied-only
        ws[f"X{r}"] = (
            f'=IF(AND(D{r}="Occupied",C{r}<>"",W{r}<>""),'
            f'C{r}&"|"&W{r},"")'
        )
        counts["x_filled"] += 1

        # Y{r}: Care|Unit, all care+type rows (incl vacants)
        ws[f"Y{r}"] = (
            f'=IF(AND(C{r}<>"",W{r}<>""),'
            f'C{r}&"|"&W{r},"")'
        )
        counts["y_filled"] += 1

    wb.save(template_path)

    log.append(
        f"PATCH: filled W/X/Y rows {START_ROW}-{END_ROW} "
        f"({counts['w_filled']} W + {counts['x_filled']} X + {counts['y_filled']} Y "
        f"= {counts['w_filled'] + counts['x_filled'] + counts['y_filled']} cells)"
    )
    log.append(f"SAVED: {template_path}")
    return {"counts": counts, "log": log, "written": True}


def verify(template_path: Path) -> dict:
    """Spot-check that the patch landed cleanly."""
    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb[SHEET_NAME]

    r = {}
    # Headers should be unchanged
    r["w210_header"] = ws["W210"].value
    r["x210_header"] = ws["X210"].value
    r["y210_header"] = ws["Y210"].value

    # Data rows — W/X/Y at start, mid, end of range
    for col in ("W", "X", "Y"):
        for row in (211, 400, 610):
            v = ws[f"{col}{row}"].value
            r[f"{col}{row}"] = v
            r[f"{col}{row}_is_formula"] = isinstance(v, str) and v.startswith("=")

    # Just-past-end should be empty (no over-write)
    r["W611"] = ws["W611"].value
    r["W611_empty"] = ws["W611"].value in (None, "")

    return r


def main(argv: list[str]) -> int:
    target = Path(argv[1]) if len(argv) > 1 else DEFAULT_TARGET
    print(f"Patching: {target}")

    result = patch(target)
    for line in result["log"]:
        print(f"  {line}")

    if not result["written"]:
        print("\n=== No-op (already patched) ===")
        return 0

    print("\n=== Verifying ===")
    v = verify(target)
    print(f"  Headers preserved: W210={v['w210_header']!r}, X210={v['x210_header']!r}, Y210={v['y210_header']!r}")
    checks = [
        ("W211 is formula", v["W211_is_formula"]),
        ("X211 is formula", v["X211_is_formula"]),
        ("Y211 is formula", v["Y211_is_formula"]),
        ("W400 is formula (mid-range)", v["W400_is_formula"]),
        ("W610 is formula (end of range)", v["W610_is_formula"]),
        ("W611 is empty (no over-write)", v["W611_empty"]),
    ]
    for desc, ok in checks:
        flag = "[OK]  " if ok else "[FAIL]"
        print(f"  {flag}  {desc}")
    all_ok = all(ok for _, ok in checks)

    print()
    print(f"  Sample formulas:")
    print(f"    W211 = {v['W211']!r}")
    print(f"    X211 = {v['X211']!r}")
    print(f"    Y211 = {v['Y211']!r}")

    print()
    print(f"=== {'[OK] Patch complete' if all_ok else '[FAIL] Patch verification failed'} ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
