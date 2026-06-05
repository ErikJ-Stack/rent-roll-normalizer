"""
Fix the B56:M56 monthly-header miss in the operator's v6 *rev2* template (the
Other-Care revision absorbed 2026-06-03 as the new canonical asset).

Same openpyxl-quirk-#4 partial-repoint class as the 2026-05-28 v0.7.1 fix
(=C122 → =C137). v0.7.1 repointed the Layer-3 monthly header row B56:M56 to the
rev1 raw header at row 137. The rev2 income restructure ("Other Care") shifted
the Layer-1 raw grid down to row 140 (raw header `C140:N140 = Apr-25 … Mar-26`),
and the rev2 repoint pass missed B56:M56 — leaving it pointing at row 125
("Permits, Licenses & Dues", an expense line = 0), so the Layer-3 header band
renders zeros instead of month labels.

  B56:M56 (12 cells) mirror C140:N140 (offset +1 column):
      B56 =C125 → =C140   ...   M56 =N125 → =N140
  N56 = "T-12 Total" (literal label) is already correct and is NOT touched.

Cosmetic only: no SUMIFS / total / diagnostic depends on B56:M56.

openpyxl quirk #6: wb.save() strips xl/metadata.xml. We restore rev2's OWN
metadata (faithful — we only edit B56:M56, never a dynamic-array anchor cell).

Patches the committed canonical asset only. The operator's Deals-folder copy is
covered by the 2026-06-05 handoff (durable Excel re-author). Idempotent.
Pre-flight aborts unless row 140 carries the raw month-header.
"""
from __future__ import annotations

import io
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
ASSET = ROOT / "assets" / "ALF_UW_Template_v6.xlsx"

sys.path.insert(0, str(ROOT))
from uw_template_writer import _restore_dynamic_arrays  # noqa: E402

RAW_HDR_ROW = 140          # rev2 Layer-1 raw month-header row (C140:N140)
HDR_ROW = 56               # Layer-3 monthly header row
# B56:M56 (cols 2..13) mirror C140:N140 (cols 3..14): target col = source col + 1
DEST_COLS = list(range(2, 14))   # B..M


def _col(n: int) -> str:
    return openpyxl.utils.get_column_letter(n)


def fix_one(path: Path) -> int:
    original = path.read_bytes()
    wb = openpyxl.load_workbook(path)
    t12 = wb["T-12 Analysis"]

    # Pre-flight: confirm row 140 is the raw month-header row.
    acct = str(t12.cell(row=RAW_HDR_ROW, column=1).value or "").strip()
    first_month = str(t12.cell(row=RAW_HDR_ROW, column=3).value or "").strip()
    pf_ok = acct == "Acct #" and "Apr" in first_month
    print(f"  preflight: A{RAW_HDR_ROW}={acct!r} C{RAW_HDR_ROW}={first_month!r} "
          f"{'✓' if pf_ok else '✗'}")
    if not pf_ok:
        print(f"✗ Pre-flight failed for {path.name}; aborting (no change).")
        return 1

    # Repoint B56:M56 -> =C140..=N140
    changed = []
    for c in DEST_COLS:
        cell = f"{_col(c)}{HDR_ROW}"
        correct = f"={_col(c + 1)}{RAW_HDR_ROW}"   # B56 -> =C140, etc.
        if t12[cell].value != correct:
            changed.append(f"{cell}: {t12[cell].value!r} → {correct!r}")
            t12[cell].value = correct

    buf = io.BytesIO()
    wb.save(buf)
    restored = _restore_dynamic_arrays(buf.getvalue(), original)  # rev2's OWN metadata
    path.write_bytes(restored)

    print(f"  repointed {len(changed)} cell(s): "
          + ("; ".join(changed) or "(already correct)"))

    # Verify
    wb2 = openpyxl.load_workbook(path)
    t = wb2["T-12 Analysis"]
    checks = [
        ("B56 == =C140", t["B56"].value == "=C140"),
        ("M56 == =N140", t["M56"].value == "=N140"),
        ('N56 == "T-12 Total" (untouched)', t["N56"].value == "T-12 Total"),
        ("sheet count 16", len(wb2.sheetnames) == 16),
    ]
    with zipfile.ZipFile(path) as z:
        checks.append(("metadata.xml present", "xl/metadata.xml" in z.namelist()))
    all_ok = True
    for label, c in checks:
        all_ok = all_ok and c
        print(f"    {'✓' if c else '✗'} {label}")
    return 0 if all_ok else 1


def main() -> int:
    print(f"=== CANONICAL ASSET: {ASSET} ===")
    if not ASSET.exists():
        print(f"  ✗ missing: {ASSET}")
        return 1
    rc = fix_one(ASSET)
    print("\n" + ("Patch verified." if rc == 0 else "✗ Patch failed."))
    return rc


if __name__ == "__main__":
    sys.exit(main())
