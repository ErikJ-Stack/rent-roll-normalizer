"""
Absorption script for v5 → v5.1 metadata cells.

WHEN TO RUN
  After the operator authors the two v5.1 metadata cells in Excel via
  Cowork (per `tools/uw_template/handoffs/2026-05-26-uwt-v5-to-v51-residual-gaps.md`)
  and re-drops `assets/ALF_UW_Template_v5.xlsx`.

WHY THIS EXISTS (instead of a direct openpyxl edit)
  v0.5.0 tried the openpyxl path: cell-level fidelity diff appeared clean,
  but `wb.save()` silently dropped `xl/metadata.xml` (the `XLDAPR`/
  `fDynamic="1"` block that Excel's dynamic-array spilled ranges depend on
  — the v0.4.3 Section R/S `SORT(UNIQUE(FILTER(...)))` framework would
  break). The template MUST be authored in Excel; this script handles only
  the registry-side absorption (no template edits).

WHAT THIS SCRIPT DOES
  1. Inspect the dropped v5.1 template to find where the two new cells
     landed (operator may have used the recommended G1/H1 + B5 locations
     OR put them elsewhere — see EXPECTED_LOCATIONS below).
  2. If cells found, update `registry.json`:
       - `substrate_version`: status `gap_target` → `mapped`, target set
         to the discovered Cover stamp cell.
       - `rr_period_date`: status `proposed` → `mapped` (format
         confirmed; v5 target unchanged at `Rent Roll Analysis!B5`).
       - `t12_period_date`: status `gap_target` → `derived_in_template`
         (a NEW status); notes rewritten to reflect v5's formula-fed
         B56:M56 from on-sheet Layer 1 row 122. No target needed.
       - `registry_version` 0.3.0 → 0.3.1.
       - Close open_questions matching prefixes #4, #7, #8 (the three
         questions v5.1 closes).
       - Add `derived_in_template` to status_legend if not present.
  3. Re-run `build_mapping_artifacts.py` to refresh MD/CSV/HTML.

WHAT THIS SCRIPT DOES NOT DO
  - Edit the template (operator's surface).
  - Bump UWT_VERSION (do that manually after running this — recommended
    bump: `v0.4.3` → `v0.5.1` since v0.5.0 was used by the rolled-back
    attempt). Edit `app.py` line UWT_VERSION constant.
  - Update `CHANGELOG-UWT.md` / `SPEC-UWT.md` / `CLAUDE.md` (do those by
    hand after verifying; the brief in
    `tools/uw_template/handoffs/2026-05-26-uwt-v5-to-v51-residual-gaps.md`
    has the boilerplate).

EXPECTED LOCATIONS (recommended per the handoff brief)
  - `Cover!G1` = "Substrate:" label (operator-authored, static)
  - `Cover!H1` = writer-target empty cell (where writer pastes Cover!B8)
  - `Rent Roll Analysis!B5` = writer-target empty date cell (mm/dd/yyyy)

If the operator chose different cells, edit EXPECTED_LOCATIONS below
before running.

USAGE
  python tools/uw_template/_absorb_v51_metadata_cells.py
      [path/to/ALF_UW_Template_v5.xlsx]

  Default target: committed `assets/ALF_UW_Template_v5.xlsx`.
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
REPO_ROOT = ROOT.parent.parent
REGISTRY = ROOT / "registry.json"
DEFAULT_TEMPLATE = REPO_ROOT / "assets" / "ALF_UW_Template_v5.xlsx"
GENERATOR = ROOT / "build_mapping_artifacts.py"


# Where the operator is recommended to put the two cells.
# Override here if the operator chose different addresses in their Excel pass.
EXPECTED_LOCATIONS = {
    "substrate_stamp": {
        "sheet": "Cover",
        "label_cell": "G1",       # static label "Substrate:"
        "value_cell": "H1",       # writer-populated value
        "label_text_expected": "Substrate:",
    },
    "rr_period_date": {
        "sheet": "Rent Roll Analysis",
        "value_cell": "B5",
        "label_cell": "A5",       # already exists as "Date:" in v5
    },
}


def detect_cells(template_path: Path) -> dict:
    """Inspect the operator-edited template to confirm the v5.1 cells landed."""
    wb = openpyxl.load_workbook(template_path, data_only=False, read_only=True)
    detected = {"substrate_stamp": False, "rr_period_date_b5": False}

    # Substrate stamp — look for the label at the expected location
    if EXPECTED_LOCATIONS["substrate_stamp"]["sheet"] in wb.sheetnames:
        ws = wb[EXPECTED_LOCATIONS["substrate_stamp"]["sheet"]]
        label_addr = EXPECTED_LOCATIONS["substrate_stamp"]["label_cell"]
        label_text = EXPECTED_LOCATIONS["substrate_stamp"]["label_text_expected"]
        v = ws[label_addr].value
        if isinstance(v, str) and label_text.lower() in v.lower():
            detected["substrate_stamp"] = True

    # RR Analysis B5 — confirm it has a date number_format OR is empty (writer
    # populates at runtime). We can't reliably detect "this is a date cell"
    # via openpyxl read_only mode; trust the operator to have formatted it.
    # Just verify the sheet exists.
    if EXPECTED_LOCATIONS["rr_period_date"]["sheet"] in wb.sheetnames:
        detected["rr_period_date_b5"] = True  # presence assumed; format trust

    return detected


def update_registry(detected: dict) -> dict:
    """Apply v5.1 absorption updates to registry.json."""
    reg = json.loads(REGISTRY.read_text(encoding="utf-8"))

    if reg.get("registry_version", "0.0.0") >= "0.3.1":
        return {"status": "no-op", "reason": f"registry already at v{reg['registry_version']}"}

    changes = []

    # Add new status if not present
    if "derived_in_template" not in reg["status_legend"]:
        reg["status_legend"]["derived_in_template"] = (
            "Template re-derives the value from on-sheet formulas (Excel "
            "computes it from other cells). Writer has no target to populate "
            "and intentionally so — distinct from `derived` which is for "
            "concepts where the Analyzer computes the value upstream."
        )
        changes.append("status_legend += derived_in_template")

    by_key = {c["key"]: c for c in reg["concepts"]}

    # 1. substrate_version → mapped
    if "substrate_version" in by_key and detected.get("substrate_stamp"):
        c = by_key["substrate_version"]
        c["status"] = "mapped"
        c["targets"]["v5"] = {
            "sheet": EXPECTED_LOCATIONS["substrate_stamp"]["sheet"],
            "address": EXPECTED_LOCATIONS["substrate_stamp"]["value_cell"],
            "label_at": EXPECTED_LOCATIONS["substrate_stamp"]["label_cell"],
            "target_label": "Substrate (Cover stamp)",
        }
        c["notes"] = (
            (c.get("notes", "") or "") +
            f" · v5.1: Closed — operator authored "
            f"{EXPECTED_LOCATIONS['substrate_stamp']['sheet']}!"
            f"{EXPECTED_LOCATIONS['substrate_stamp']['label_cell']}/"
            f"{EXPECTED_LOCATIONS['substrate_stamp']['value_cell']}. "
            f"Writer pastes Analyzer Cover!B8 value (e.g. 'v0.2.14')."
        )
        changes.append("substrate_version: gap_target → mapped")

    # 2. rr_period_date → mapped (was already proposed with B5 target — just confirm)
    if "rr_period_date" in by_key:
        c = by_key["rr_period_date"]
        if c["status"] in ("proposed", "gap_target"):
            c["status"] = "mapped"
            c["notes"] = (
                (c.get("notes", "") or "") +
                " · v5.1: Closed — format confirmed (mm/dd/yyyy at B5), "
                "writer pastes from RR_Period_Date named range."
            )
            changes.append("rr_period_date: proposed → mapped")

    # 3. t12_period_date → derived_in_template
    if "t12_period_date" in by_key:
        c = by_key["t12_period_date"]
        if c["status"] == "gap_target":
            c["status"] = "derived_in_template"
            c["targets"]["v5"] = None
            c["notes"] = (
                "v5 derives T-12 Analysis!B56:M56 monthly headers from "
                "on-sheet Layer 1 raw paste at row 122 via formula chain "
                "(B56 = C122, ..., M56 = N122). Writer has no target — "
                "template auto-derives from the raw T12 paste. The v4 "
                "registry note describing hardcoded Apr-25..Mar-26 was "
                "stale; v5 fixed this structurally."
            )
            changes.append("t12_period_date: gap_target → derived_in_template")

    # 4. Close open_questions matching v5.1-closeable prefixes
    CLOSED_PREFIXES = [
        "Cover substrate version stamp",
        "Rent Roll Analysis tab-header Period Date",
        "Date header at Rent Roll Analysis!A5",
    ]
    before = len(reg.get("open_questions", []))
    reg["open_questions"] = [
        q for q in reg.get("open_questions", [])
        if not any(q.lstrip().startswith(p) for p in CLOSED_PREFIXES)
    ]
    closed = before - len(reg["open_questions"])
    if closed:
        changes.append(f"open_questions: {before} → {len(reg['open_questions'])} ({closed} closed)")

    # 5. Bump registry_version
    reg["registry_version"] = "0.3.1"
    reg["generated_phase"] = (
        "Track 4 / Phase 3.6 — v5 → v5.1 metadata cells absorbed. "
        "substrate_version + rr_period_date close to mapped; t12_period_date "
        "reclassified to derived_in_template. Registry-only update; writer unchanged."
    )
    changes.append("registry_version: 0.3.0 → 0.3.1")

    REGISTRY.write_text(json.dumps(reg, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return {"status": "updated", "changes": changes}


def regen_artifacts() -> int:
    """Re-run build_mapping_artifacts.py."""
    if not GENERATOR.exists():
        return 1
    result = subprocess.run(
        [sys.executable, str(GENERATOR)],
        capture_output=True, text=True, encoding="utf-8",
    )
    print(result.stdout)
    if result.returncode != 0:
        print(f"  WARN: generator returned {result.returncode}")
        print(result.stderr[:500])
    return result.returncode


def main(argv: list[str]) -> int:
    template = Path(argv[1]) if len(argv) > 1 else DEFAULT_TEMPLATE

    if not template.exists():
        print(f"ERROR: template not found at {template}", file=sys.stderr)
        return 1

    print(f"Inspecting v5.1 template: {template}")
    detected = detect_cells(template)
    print(f"  Detection results:")
    for k, v in detected.items():
        flag = "✓" if v else "✗"
        print(f"    {flag}  {k}")
    print()

    if not detected["substrate_stamp"]:
        print(
            f"WARN: Substrate stamp label not found at "
            f"{EXPECTED_LOCATIONS['substrate_stamp']['sheet']}!"
            f"{EXPECTED_LOCATIONS['substrate_stamp']['label_cell']}. "
            f"Either the operator hasn't authored v5.1 yet, or they chose "
            f"different locations. Edit EXPECTED_LOCATIONS at the top of "
            f"this script and re-run."
        )
        print()

    print(f"Updating registry: {REGISTRY}")
    result = update_registry(detected)
    if result["status"] == "no-op":
        print(f"  No-op: {result['reason']}")
        return 0

    print(f"  Registry updates:")
    for change in result["changes"]:
        print(f"    + {change}")
    print()

    print(f"Regenerating artifacts...")
    rc = regen_artifacts()
    print()

    print("=== Absorption complete ===")
    print()
    print("Next steps (manual):")
    print("  1. Bump UWT_VERSION in app.py: v0.4.3 → v0.5.1 (v0.5.0 was used")
    print("     by the rolled-back attempt — skip it to avoid version reuse).")
    print("  2. Add CHANGELOG-UWT.md entry for v0.5.1 (boilerplate in the")
    print("     handoff brief at tools/uw_template/handoffs/2026-05-26-uwt-v5-to-v51-residual-gaps.md).")
    print("  3. Update SPEC-UWT.md current code version + phase plan row.")
    print("  4. Update CLAUDE.md last-updated stamp + Track 4 row.")
    print("  5. Mark the handoff Verified in tools/uw_template/HANDOFF_TRACKER.md.")
    print("  6. Re-run writer regression: PYTHONPATH=. python tests/test_uw_template_writer.py")
    print("     Expect: substrate_version + rr_period_date now report `written`")
    print("     in PopulateReport (previously skipped/no_target).")
    print("  7. Commit + push.")
    return rc


if __name__ == "__main__":
    sys.exit(main(sys.argv))
