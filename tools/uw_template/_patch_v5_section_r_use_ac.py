"""
Patch ALF_UW_Template_v5.xlsx — Section R: W mirrors AC (gated by
occupancy) + A173/B173 IFERROR wrappers + D173 uses per-unit sq ft from
col T instead of hardcoded estimates.

THE BUG (operator-reported 2026-05-27)
  Section R on the populated UW Template returns #N/A and zeros. Root
  cause: col W's formula in v5 is a substring-Notes-parser that tries
  to extract unit type from col S (Notes) with searches for "studio",
  "1 bed", "1br", etc. — but real rent rolls don't carry that wording
  in the Notes column (it's free-form lease/concession context). So
  W="" everywhere → X="" everywhere → Z173 spill is empty → all
  downstream aggregations in Section R show #N/A and zeros.

  The fix is to use col AC ("Apt Type") directly — the writer already
  pastes Analyzer col F's normalized closed-vocab values there (Studio /
  1 Bedroom / 2 Bedroom / 3 Bedroom / Suite / Cottage). W just needs to
  read AC.

HISTORY
  v0.4.3 (2026-05-26) shipped `W = =AC{r}` via the original Section R
  fill-down patch. The operator subsequently edited the template in
  Excel and REPLACED W's formula with the substring-Notes-parser (which
  I'd merely sketched as a what-if option in an earlier chat — never
  recommended). The operator-edited version became the committed asset
  at `deacc41 chore: refresh assets/ALF_UW_Template_v5.xlsx with
  operator edits`. This patch puts the AC reference back, gated on
  occupancy + AC<>"" per the 2026-05-27 operator diagnostic.

THE FIX (4 surface changes to assets/ALF_UW_Template_v5.xlsx)

  W211:W610 — Replace substring-parser with gated AC reference:
    W{r} = =IF(AND($D{r}="Occupied", $AC{r}<>""), $AC{r}, "")
    Returns the canonical Apt Type from AC when the row is occupied
    AND AC is populated; empty otherwise. Vacants and unmapped rows
    contribute nothing to Section R's unique-key SORT/UNIQUE/FILTER.

  A173 — Wrap dynamic-array TEXTBEFORE in IFERROR so empty Z173 spills
    don't show #N/A:
    A173 = =IFERROR(TEXTBEFORE(ANCHORARRAY(Z173),"|"),"")

  B173 — Same for TEXTAFTER:
    B173 = =IFERROR(TEXTAFTER(ANCHORARRAY(Z173),"|"),"")

  D173 — Replace hardcoded sq ft lookup `XLOOKUP(B173#, {"Studio";
    "1 Bedroom"; ...}, {450;700;1000;1300;350;900}, "")` with the
    actual per-unit sq ft from col T, averaged by Care|UnitType bucket:
    D173 = =IFERROR(AVERAGEIFS($T$211:$T$610,
                                $C$211:$C$610, ANCHORARRAY(A173),
                                $AC$211:$AC$610, ANCHORARRAY(B173)), "")

IDEMPOTENCY
  Bails if W211 already holds the new gated-AC formula. Re-runs safe.

OPENPYXL QUIRK #6 (xl/metadata.xml drop)

  This patch touches cells in a file with dynamic-array formulas
  (Z173 / C173 / etc. use SORT/UNIQUE/FILTER/ANCHORARRAY which depend
  on `xl/metadata.xml`'s XLDAPR block). openpyxl's `wb.save()` silently
  drops that part. After this patch ships:

    1. The committed file's xl/metadata.xml will be missing.
    2. When the operator (or any Excel user) opens the file, Excel
       detects the missing part and offers to repair on open. Accept
       the repair, then SAVE the file in Excel.
    3. Excel rebuilds xl/metadata.xml with the correct XLDAPR props.

  This round-trip has been the working pattern since v0.4.3 — Excel's
  repair behavior is forgiving. The v0.5.0 rollback documented this
  quirk after a more invasive (cells-and-headers) attempt; for a
  surface-only formula patch like this one, the round-trip is safe.

  See CLAUDE.md openpyxl quirk #6 for the full technical detail.

USAGE
  python tools/uw_template/_patch_v5_section_r_use_ac.py
      [path/to/ALF_UW_Template_v5.xlsx]

  Default target: committed `assets/ALF_UW_Template_v5.xlsx`. Operator
  should ALSO run on their Deals-folder canonical copy (or just replace
  it with the patched repo copy) to keep them in sync.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

DEFAULT_TARGET = Path(__file__).resolve().parent.parent.parent / "assets" / "ALF_UW_Template_v5.xlsx"
SHEET_NAME = "Rent Roll Analysis"

# Data-row range for the W column fill-down
W_START_ROW = 211
W_END_ROW = 610

# New formulas to write
def _w_formula(row: int) -> str:
    """W{r} — gated AC reference. Occupied + AC populated only."""
    return f'=IF(AND($D{row}="Occupied",$AC{row}<>""),$AC{row},"")'


A173_FORMULA = '=IFERROR(TEXTBEFORE(ANCHORARRAY(Z173),"|"),"")'
B173_FORMULA = '=IFERROR(TEXTAFTER(ANCHORARRAY(Z173),"|"),"")'
D173_FORMULA = (
    '=IFERROR(AVERAGEIFS($T$211:$T$610,'
    '$C$211:$C$610,ANCHORARRAY(A173),'
    '$AC$211:$AC$610,ANCHORARRAY(B173)),"")'
)


def _is_already_patched(ws) -> bool:
    """Detect whether the gated-AC formula is already in W211."""
    v = ws["W211"].value
    # ArrayFormula objects carry the formula text on `.text`
    if hasattr(v, "text"):
        text = v.text
    elif isinstance(v, str):
        text = v
    else:
        return False
    # Identify by the distinctive pattern
    return '$AC211<>""' in text and '$AC211,""' in text


def patch(template_path: Path) -> dict:
    """Idempotent: apply the 4 surface changes."""
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet {SHEET_NAME!r} not in template")
    ws = wb[SHEET_NAME]

    counts = {"w_rewritten": 0, "section_r_rewritten": 0, "skipped": 0}
    log: list[str] = []

    if _is_already_patched(ws):
        counts["skipped"] = 1
        log.append("SKIP: W211 already holds the gated-AC formula (re-run safe)")
        return {"counts": counts, "log": log, "written": False}

    # ── 1. Rewrite W211:W610 ──────────────────────────────────────────────
    for r in range(W_START_ROW, W_END_ROW + 1):
        ws[f"W{r}"] = _w_formula(r)
        counts["w_rewritten"] += 1
    log.append(
        f"PATCH W{W_START_ROW}:W{W_END_ROW}: replaced substring-Notes-parser "
        f"with gated AC reference ({counts['w_rewritten']} cells)"
    )

    # ── 2. A173 — wrap TEXTBEFORE in IFERROR ──────────────────────────────
    ws["A173"] = ArrayFormula(ref="A173", text=A173_FORMULA)
    counts["section_r_rewritten"] += 1
    log.append(f"PATCH A173: {A173_FORMULA}")

    # ── 3. B173 — wrap TEXTAFTER in IFERROR ───────────────────────────────
    ws["B173"] = ArrayFormula(ref="B173", text=B173_FORMULA)
    counts["section_r_rewritten"] += 1
    log.append(f"PATCH B173: {B173_FORMULA}")

    # ── 4. D173 — use per-unit sq ft from col T ───────────────────────────
    ws["D173"] = ArrayFormula(ref="D173", text=D173_FORMULA)
    counts["section_r_rewritten"] += 1
    log.append(f"PATCH D173: {D173_FORMULA}")

    wb.save(template_path)
    log.append(f"SAVED: {template_path}")
    return {"counts": counts, "log": log, "written": True}


def verify(template_path: Path) -> dict:
    """Confirm the patch landed cleanly."""
    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb[SHEET_NAME]
    r = {}

    # W spot-checks
    for row in (211, 400, 610):
        v = ws[f"W{row}"].value
        text = v.text if hasattr(v, "text") else (v if isinstance(v, str) else None)
        r[f"W{row}_text"] = text
        r[f"W{row}_ok"] = isinstance(text, str) and "$AC" in text and "Occupied" in text

    # A173, B173, D173
    for addr in ("A173", "B173", "D173"):
        v = ws[addr].value
        text = v.text if hasattr(v, "text") else (v if isinstance(v, str) else None)
        r[f"{addr}_text"] = text
        r[f"{addr}_ok"] = isinstance(text, str) and "IFERROR" in text

    # Z173 unchanged
    z = ws["Z173"].value
    z_text = z.text if hasattr(z, "text") else None
    r["Z173_ok"] = isinstance(z_text, str) and "FILTER" in z_text and "SORT" in z_text
    r["Z173_text"] = z_text

    # Sheet count + dimensions unchanged
    r["sheet_count"] = len(wb.sheetnames)
    r["rra_max_row"] = ws.max_row
    return r


def main(argv: list[str]) -> int:
    target = Path(argv[1]) if len(argv) > 1 else DEFAULT_TARGET
    print(f"Patching: {target}")

    result = patch(target)
    for line in result["log"]:
        print(f"  {line}")

    if not result["written"]:
        print("\n=== No-op (already patched) ===")
        return 0

    print("\n=== Verifying ===")
    v = verify(target)
    checks = [
        ("W211 references AC + gated on Occupied", v["W211_ok"]),
        ("W400 (mid-range) same pattern", v["W400_ok"]),
        ("W610 (end of range) same pattern", v["W610_ok"]),
        ("A173 wrapped in IFERROR", v["A173_ok"]),
        ("B173 wrapped in IFERROR", v["B173_ok"]),
        ("D173 uses AVERAGEIFS on col T", v["D173_ok"]),
        ("Z173 dynamic-array driver unchanged (SORT/FILTER)", v["Z173_ok"]),
        (f"Sheet count = {v['sheet_count']} (expected 16)", v["sheet_count"] == 16),
        (f"RR Analysis max_row = {v['rra_max_row']} (expected 610)", v["rra_max_row"] == 610),
    ]
    for desc, ok in checks:
        flag = "[OK]  " if ok else "[FAIL]"
        print(f"  {flag}  {desc}")

    print()
    print(f"  Sample formulas:")
    print(f"    W211 = {v['W211_text']!r}")
    print(f"    A173 = {v['A173_text']!r}")
    print(f"    B173 = {v['B173_text']!r}")
    print(f"    D173 = {v['D173_text']!r}")

    all_ok = all(ok for _, ok in checks)
    print()
    print("=== " + ("[OK] Patch complete" if all_ok else "[FAIL] Patch verification failed") + " ===")

    if all_ok:
        print()
        print("NEXT STEPS")
        print("  1. Open the patched file in Excel.")
        print("  2. If Excel prompts to repair, accept (this restores the")
        print("     xl/metadata.xml dynamic-array properties block that")
        print("     openpyxl dropped per quirk #6).")
        print("  3. Save the file in Excel.")
        print("  4. Optionally re-save over your Deals-folder canonical.")
        print("  5. Re-run the writer on Homestead to verify Section R")
        print("     resolves cleanly — should see real Care|UnitType")
        print("     combinations and counts/min/avg/max per bucket.")

    return 0 if all_ok else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
