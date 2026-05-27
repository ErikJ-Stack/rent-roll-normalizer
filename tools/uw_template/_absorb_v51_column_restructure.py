"""
Absorption script for v5 → v5.1 column restructure (Unit Type before Status).

WHEN TO RUN
  After the operator authors the v5.1 column restructure in Excel
  (per `tools/uw_template/handoffs/2026-05-27-uwt-v51-unit-type-restructure.md`)
  and overwrites `assets/ALF_UW_Template_v5.xlsx`.

THE OPERATOR'S CHANGES (template-side)
  - NEW column at Rent Roll Analysis!D = "Unit Type"
  - DROP column W "Unit Type (base)" (was a helper col mirroring AC)
  - DROP column AC "Apt Type" (was writer-paste target for rr_apt_type)
  - Everything from old D onward shifts right by 1
  - Everything from old AD onward shifts left by 1 (to close the AC hole)
  - All in-template formulas updated to reference new positions
  - Total col count: 48 → 47

THIS SCRIPT'S CHANGES (registry-side)
  Updates every rent_roll concept's v5 target address per the shift
  table. Re-targets rr_apt_type from AC to new D. Bumps registry
  0.3.0 → 0.4.0 (substantial structural restructure).

DOES NOT
  - Edit the template (operator owns that — see openpyxl quirk #6).
  - Bump UWT_VERSION (do that manually after running — recommended
    bump: v0.4.4 → v0.5.1, skipping v0.5.0 which was the rolled-back
    attempt).
  - Update CHANGELOG-UWT.md / SPEC-UWT.md / CLAUDE.md (manual after).

PRE-FLIGHT VERIFICATION
  Confirms the template has the new structure before applying registry
  changes. If the operator hasn't authored yet (or chose a different
  restructure shape), the script bails with a clear error.

USAGE
  python tools/uw_template/_absorb_v51_column_restructure.py
      [path/to/ALF_UW_Template_v5.xlsx]
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
REPO_ROOT = ROOT.parent.parent
REGISTRY = ROOT / "registry.json"
DEFAULT_TEMPLATE = REPO_ROOT / "assets" / "ALF_UW_Template_v5.xlsx"
GENERATOR = ROOT / "build_mapping_artifacts.py"

# ──────────────────────────────────────────────────────────────────────────────
# Shift table — v5 col → v5.1 col
#
# Encodes the column transformation:
#   1. Insert new D "Unit Type"  → old D-V shift right to E-W
#   2. Drop W "Unit Type (base)" (which was at OLD-W, NEW-X post-right-shift)
#      → wait, this is more subtle. Let me re-trace.
#
# Actually the net effect is simpler if we think of it as:
#   - Cols A, B, C unchanged
#   - NEW col D inserted (Unit Type)
#   - Cols old D..V shift right by 1 → become E..W
#   - Old W (Unit Type base) was deleted from the resulting position
#     → but since cols shifted right BEFORE the delete, the deleted col
#       is at the new X position (was old W) -- doesn't this re-shift?
#   - This gets complicated. The operator's net effect (per the handoff
#     brief recommendation): drop W AND AC, insert new D, end with 47 cols.
#
# CLEANER MODEL: just describe end-state mappings v5 → v5.1.
#
# Old col → New col (where applicable; "DROP" means concept's target removed
#                    because the col is gone in v5.1):
# ──────────────────────────────────────────────────────────────────────────────

SHIFT_TABLE: dict[str, str] = {
    # A, B, C unchanged
    "A": "A",
    "B": "B",
    "C": "C",
    # D-V shift right by 1 to E-W
    "D": "E",
    "E": "F",
    "F": "G",
    "G": "H",
    "H": "I",
    "I": "J",
    "J": "K",
    "K": "L",
    "L": "M",
    "M": "N",
    "N": "O",
    "O": "P",
    "P": "Q",
    "Q": "R",
    "R": "S",
    "S": "T",
    "T": "U",
    "U": "V",
    "V": "W",
    # W (Unit Type base) — DROPPED (no concept currently targeted W; was
    # template-internal helper col)
    "W": None,
    # X, Y, Z, AA, AB unchanged in position (formula cols — but their formulas
    # update on the template side to use new D not old W)
    "X": "X",
    "Y": "Y",
    "Z": "Z",
    "AA": "AA",
    "AB": "AB",
    # AC (Apt Type) — DROPPED. rr_apt_type retargets to new D (not via
    # SHIFT_TABLE — handled in SPECIAL_RETARGETS below)
    "AC": None,
    # AD-AV left-shift by 1 to AC-AU (closing the AC hole)
    "AD": "AC",
    "AE": "AD",
    "AF": "AE",
    "AG": "AF",
    "AH": "AG",
    "AI": "AH",
    "AJ": "AI",
    "AK": "AJ",
    "AL": "AK",
    "AM": "AL",
    "AN": "AM",
    "AO": "AN",
    "AP": "AO",
    "AQ": "AP",
    "AR": "AQ",
    "AS": "AR",
    "AT": "AS",
    "AU": "AT",
    "AV": "AU",
}

# Concepts that need special retarget (not just a column shift)
SPECIAL_RETARGETS: dict[str, str] = {
    # rr_apt_type was at AC (dropped); re-target to new D "Unit Type"
    "rr_apt_type": "D",
}


# ──────────────────────────────────────────────────────────────────────────────
# Pre-flight: verify the operator's v5.1 restructure landed
# ──────────────────────────────────────────────────────────────────────────────

def verify_template_restructured(template_path: Path) -> tuple[bool, list[str]]:
    """Check whether the template has been restructured to v5.1 shape.

    Returns (ok, issues_list).
    """
    issues = []
    wb = openpyxl.load_workbook(template_path, data_only=False, read_only=True)
    if "Rent Roll Analysis" not in wb.sheetnames:
        return False, ["Rent Roll Analysis sheet missing"]

    ws = wb["Rent Roll Analysis"]

    # D210 should now be "Unit Type"
    d210 = ws["D210"].value
    if not (isinstance(d210, str) and d210.strip().lower() in ("unit type", "unit type (base)")):
        issues.append(
            f"D210 expected 'Unit Type', got {d210!r}. Operator may not have "
            f"restructured yet, or chose a different col position."
        )

    # E210 should now be "Status" (was D in v5)
    e210 = ws["E210"].value
    if not (isinstance(e210, str) and e210.strip().lower() == "status"):
        issues.append(
            f"E210 expected 'Status' (was D in v5), got {e210!r}. Column "
            f"right-shift not detected."
        )

    # Max col should be 47 (was 48) — two cols dropped, one added
    if ws.max_column != 47:
        issues.append(
            f"Rent Roll Analysis max_column = {ws.max_column} (expected 47 after "
            f"dropping W + AC, inserting new D). Restructure incomplete."
        )

    # Confirm AC no longer says "Apt Type"
    ac210 = ws["AC210"].value
    if isinstance(ac210, str) and ac210.strip().lower() == "apt type":
        issues.append(
            f"AC210 still reads {ac210!r} — old AC should be dropped, with "
            f"Concession $ (old AD) now at AC. Restructure incomplete."
        )

    return len(issues) == 0, issues


# ──────────────────────────────────────────────────────────────────────────────
# Apply registry updates
# ──────────────────────────────────────────────────────────────────────────────

def shift_address(addr: str) -> str | None:
    """Shift a target address like 'D211+' or 'AC211' to its v5.1 equivalent.

    Returns the new address, or None if the column was dropped.
    """
    if not addr:
        return addr
    # Parse out col letters
    col = ""
    i = 0
    while i < len(addr) and addr[i].isalpha():
        col += addr[i]
        i += 1
    rest = addr[i:]
    if col not in SHIFT_TABLE:
        # Unknown col — leave alone (and warn)
        return addr
    new_col = SHIFT_TABLE[col]
    if new_col is None:
        return None
    return f"{new_col}{rest}"


def shift_target(target: dict | None) -> tuple[dict | None, str]:
    """Apply column shifts to a target dict. Returns (new_target, change_note).

    new_target is None if the target's column was dropped (concept becomes
    target-less).
    """
    if not target:
        return target, "no-target"

    old_addr = target.get("address", "")
    old_label_at = target.get("label_at", "")
    new_addr = shift_address(old_addr)

    if new_addr is None:
        return None, f"dropped (was {old_addr})"
    if new_addr == old_addr:
        return target, "unchanged"

    new_target = dict(target)
    new_target["address"] = new_addr
    if old_label_at:
        new_label_at = shift_address(old_label_at)
        if new_label_at is not None:
            new_target["label_at"] = new_label_at
    paste_anchor = target.get("paste_anchor")
    if paste_anchor and "!" in paste_anchor:
        sheet, addr = paste_anchor.split("!", 1)
        new_pa = shift_address(addr)
        if new_pa:
            new_target["paste_anchor"] = f"{sheet}!{new_pa}"

    return new_target, f"{old_addr} → {new_addr}"


def update_registry() -> dict:
    """Apply v5.1 column-restructure updates to registry.json.

    CRITICAL SCOPE NOTE: the column restructure ONLY affects the
    `Rent Roll Analysis` sheet. T-12 Analysis, Prop Info, Cover, etc.
    are UNTOUCHED. We filter on target.sheet to scope the shift
    correctly — pre-v0.5.1 bug-fix: previous version shifted ALL
    concepts, which incorrectly moved T-12 Analysis cells like
    `egi: N69 → O69` (T-12 Analysis cols didn't change).
    """
    reg = json.loads(REGISTRY.read_text(encoding="utf-8"))

    if reg.get("registry_version", "0.0.0") >= "0.4.0":
        return {"status": "no-op", "reason": f"registry at v{reg['registry_version']}; expected 0.3.x"}

    changes = []
    target_changes: list[tuple[str, str]] = []
    drop_changes: list[str] = []

    for concept in reg["concepts"]:
        key = concept.get("key", "")

        # Special retargets take precedence over shift table
        if key in SPECIAL_RETARGETS:
            new_col = SPECIAL_RETARGETS[key]
            targets = concept.setdefault("targets", {})
            old_v5 = targets.get("v5")
            if old_v5 and old_v5.get("sheet") == "Rent Roll Analysis":
                new_target = dict(old_v5)
                old_addr = old_v5.get("address", "")
                row_part = ""
                i = 0
                while i < len(old_addr) and old_addr[i].isalpha():
                    i += 1
                row_part = old_addr[i:]
                new_target["address"] = f"{new_col}{row_part}"
                if old_v5.get("label_at"):
                    la = old_v5["label_at"]
                    la_i = 0
                    while la_i < len(la) and la[la_i].isalpha():
                        la_i += 1
                    new_target["label_at"] = f"{new_col}{la[la_i:]}"
                targets["v5"] = new_target
                target_changes.append((key, f"SPECIAL: {old_addr} → {new_target['address']}"))
            continue

        # Standard shift via SHIFT_TABLE — SCOPED TO Rent Roll Analysis ONLY
        targets = concept.get("targets") or {}
        old_target = targets.get("v5")
        if not old_target:
            continue

        # ★ KEY FIX: only shift if the target is on Rent Roll Analysis.
        # T-12 Analysis, Prop Info, Cover, etc. are unaffected by the
        # column restructure.
        if old_target.get("sheet") != "Rent Roll Analysis":
            continue

        new_target, note = shift_target(old_target)
        if note in ("unchanged", "no-target"):
            continue

        if new_target is None:
            drop_changes.append(key)
            targets["v5"] = None
        else:
            targets["v5"] = new_target
            target_changes.append((key, note))

    # Update template metadata
    v5 = reg["templates"].setdefault("v5", {})
    v5["last_structural_update"] = "2026-05-27 (v5.1 column restructure: Unit Type inserted at new D; W + AC dropped)"
    # rent_roll col span unchanged (data still ends at row 610), but the new
    # last col is AU instead of AV
    v5["rent_roll_last_col"] = "AU"

    # Version bump
    reg["registry_version"] = "0.4.0"
    reg["generated_phase"] = (
        "Track 4 / Phase 4 — v5.1 column restructure absorbed. "
        "Rent Roll Analysis: Unit Type inserted at new col D (before Status); "
        "old W (Unit Type base helper) + old AC (Apt Type) dropped; D-V shifted "
        "right to E-W; AD-AV shifted left to AC-AU. Net col count 48 → 47. "
        "rr_apt_type re-targets AC → D. Writer no code change; registry-only."
    )
    changes.append("registry_version: 0.3.0 → 0.4.0")
    changes.append(f"target shifts applied to {len(target_changes)} concepts")
    if drop_changes:
        changes.append(f"targets dropped for {len(drop_changes)} concept(s): {drop_changes}")

    REGISTRY.write_text(json.dumps(reg, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return {
        "status": "updated",
        "changes": changes,
        "target_changes": target_changes,
        "drop_changes": drop_changes,
    }


def regen_artifacts() -> int:
    if not GENERATOR.exists():
        return 1
    result = subprocess.run(
        [sys.executable, str(GENERATOR)],
        capture_output=True, text=True, encoding="utf-8",
    )
    print(result.stdout)
    if result.returncode != 0:
        print(result.stderr[:500])
    return result.returncode


def main(argv: list[str]) -> int:
    target = Path(argv[1]) if len(argv) > 1 else DEFAULT_TEMPLATE

    if not target.exists():
        print(f"ERROR: template not found at {target}", file=sys.stderr)
        return 1

    print(f"=== Pre-flight: verify v5.1 restructure on {target} ===")
    ok, issues = verify_template_restructured(target)
    if not ok:
        print("PRE-FLIGHT FAILED — operator has not yet authored v5.1 restructure:")
        for issue in issues:
            print(f"  - {issue}")
        print()
        print("See `tools/uw_template/handoffs/2026-05-27-uwt-v51-unit-type-restructure.md`")
        print("for the spec the operator needs to follow in Excel.")
        return 1
    print("  ✓ D210 = 'Unit Type'")
    print("  ✓ E210 = 'Status' (right-shifted)")
    print("  ✓ Max column = 47 (was 48 — net -1)")
    print("  ✓ AC no longer 'Apt Type' (dropped)")
    print()

    print(f"=== Applying registry updates ===")
    result = update_registry()
    if result["status"] == "no-op":
        print(f"  No-op: {result['reason']}")
        return 0

    for change in result["changes"]:
        print(f"  + {change}")
    print()
    if result.get("target_changes"):
        print(f"  Per-concept target shifts ({len(result['target_changes'])}):")
        for key, note in result["target_changes"][:15]:
            print(f"    {key:35s}  {note}")
        if len(result["target_changes"]) > 15:
            print(f"    ... + {len(result['target_changes']) - 15} more")
    if result.get("drop_changes"):
        print(f"  Targets dropped: {result['drop_changes']}")
    print()

    print("=== Regenerating artifacts ===")
    regen_artifacts()
    print()

    print("=== Absorption complete ===")
    print()
    print("NEXT STEPS (manual)")
    print("  1. Bump UWT_VERSION in app.py: v0.4.4 → v0.5.1 (skip v0.5.0 — rolled-back).")
    print("  2. Add CHANGELOG-UWT.md v0.5.1 entry (boilerplate in the handoff brief).")
    print("  3. Update SPEC-UWT.md current code version + phase plan row.")
    print("  4. Update CLAUDE.md last-updated stamp + Track 4 row.")
    print("  5. Run writer regression: `PYTHONPATH=. python tests/test_uw_template_writer.py`")
    print("     Expect: cells_written near-identical to v5 (~3,232 on Homestead);")
    print("     rr_apt_type now writes to D211+ instead of AC211+; right-shifted")
    print("     and left-shifted concepts at their new positions.")
    print("  6. Optionally: rerun `populate_uw_template()` on Homestead, open in")
    print("     Excel, verify D211 = '1 Bedroom' (Janet Pierson's unit) and")
    print("     E211 = 'Occupied'.")
    print("  7. Mark handoff Verified in `tools/uw_template/HANDOFF_TRACKER.md`.")
    print("  8. Commit + push.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
