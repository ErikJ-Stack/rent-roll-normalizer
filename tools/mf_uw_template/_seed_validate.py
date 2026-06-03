"""
MF COA-seed validation harness (LOCAL tool, not the production parser).

Applies the COA -> _StdCOA seed dictionary (`coa_seed` below) to the real
operator T-12 samples and reports per-file coverage + unmapped lines. This is
the "first slice" prototype for the scoped MF parser build (SPEC-MF §2.8):
mf_t12_normalizer + the COA dictionary, validated against the hand-mapped
samples. Reads gitignored deal files via absolute paths — local-only, guarded
for existence; the production parser will take file inputs properly.

Run: python tools/mf_uw_template/_seed_validate.py
"""
from __future__ import annotations
import os, re, sys
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

# ---------- Yardi account-number (5-digit root) -> bucket ----------
ACCT = {
 41000:"Gross Potential Rent",41008:"Gross Potential Rent",41010:"Gain/Loss on Market Rent",
 41091:"Concessions",41100:"Vacancy Loss",41110:"Employee Units",41115:"Employee Units",
 41120:"Down Units Loss",41121:"Down Units Loss",41150:"Write-offs / Bad Debt",
 41153:"Write-offs / Bad Debt",41155:"Write-offs / Bad Debt",
 43010:"Admin Fees",43020:"Application Fees",43035:"Write-offs / Bad Debt",43055:"Misc Other Income",
 43056:"Misc Other Income",43063:"Admin Fees",43080:"Misc Other Income",43090:"Misc Other Income",
 43105:"Parking Income",43117:"Misc Other Income",43125:"Misc Other Income",43135:"Late Fees",
 43145:"Lease Break Fee",43150:"Misc Other Income",43160:"Misc Other Income",43170:"MTM Fees",
 43180:"Late Fees",43190:"Parking Income",43196:"Parking Income",43200:"Pet Fees",43201:"Pet Fees",
 43213:"Misc Other Income",43215:"Renter's Insurance",43225:"Reimbursement — Internet",
 43235:"Storage / Common Bins",43250:"Admin Fees",43258:"Utility Reimbursement",43261:"Utility Reimbursement",
 43262:"Utility Reimbursement",43264:"Utility Reimbursement",43265:"Misc Other Income",43267:"Misc Other Income",
 43277:"Admin Fees",43290:"Misc Other Income",
}
ACCT.update({n:"Payroll — On-Site" for n in
 [51010,51015,51017,51020,51024,51026,51030,51040,51045,51050,51055,51070,51090,51100,51110,51120,51140,51150]})
ACCT.update({n:"Repairs & Maintenance" for n in
 [52010,52020,52040,52050,52051,52055,52057,52060,52065,52070,52088,52100,52105,52110,52112,52125,52130,52140,52150,52155,52190,52200,52210,52230,52250]})
ACCT[52015]="Pest Control"; ACCT[52063]="Pool / Amenity Maintenance"
ACCT.update({n:"Make-Ready / Turnover" for n in [52605,52610,52620,52621,52640,52645,52650,52660,52670,52680,52685,52690,52700]})
ACCT.update({n:"Pool / Amenity Maintenance" for n in [52820,52830,52860,52880]})
ACCT.update({53030:"Utilities — Internet/Cable",53055:"Contract Services",53060:"Contract Services",
 53070:"Contract Services",53073:"Contract Services",53090:"Contract Services",53100:"Landscaping",
 53105:"Landscaping",53116:"Contract Services",53130:"Contract Services",53131:"Pool / Amenity Maintenance",
 53140:"Pest Control",53145:"Utility Reimbursement",53150:"Pool / Amenity Maintenance",53165:"Contract Services",
 53171:"Contract Services",53175:"Pest Control",53180:"Utilities — Trash",53182:"Utilities — Trash",
 53183:"Utility Reimbursement",53185:"Utility Reimbursement",53186:"Utilities — Trash"})
ACCT.update({n:"Leasing & Marketing" for n in
 [54005,54007,54010,54012,54013,54025,54028,54035,54038,54040,54042,54044,54055,54090,54100,54105,54110,54122,54126,54134]})
ACCT.update({n:"General & Administrative" for n in
 [58020,58025,58028,58030,58070,58080,58090,58100,58107,58115,58205,58210,58225,58238,58240,58245,58247,58250,58253,58280,58281,58283,58284,58290,58300,58320,58278,58110]})
ACCT.update({58260:"Legal / Eviction",58270:"Utilities — Internet/Cable",58275:"Legal / Eviction",58305:"Leasing & Marketing"})
ACCT.update({59010:"Utilities — Electric",59020:"Utilities — Electric",59030:"Utilities — Electric",59040:"Utilities — Electric",
 59050:"Utility Reimbursement",59100:"General & Administrative",59105:"Utility Reimbursement",
 59110:"Utilities — Water/Sewer",59115:"Utility Reimbursement"})
ACCT.update({61030:"Management Fee",62010:"Real Estate Taxes",62030:"Real Estate Taxes",
 63010:"Insurance",63020:"Insurance",63030:"Insurance",63035:"Insurance",63090:"Insurance"})


def acct_root(a):
    m = re.match(r"(\d{5})", str(a) if a is not None else "")
    return int(m.group(1)) if m else None


def acct_bucket(a):
    root = acct_root(a)
    if root is None:
        return None
    if 70000 <= root <= 89999:
        return "— EXCLUDED (non-OpEx) —"   # replacements / capital / non-operating / D&A
    return ACCT.get(root)


# ---------- name regex -> bucket (ordered specific -> generic) ----------
NAME = [
 (r"gain\s*/?\s*loss to lease", "Gain/Loss on Market Rent"),
 (r"vacancy", "Vacancy Loss"), (r"concession", "Concessions"),
 (r"employee unit|courtesy officer", "Employee Units"),
 (r"model|office unit|storage unit|down unit", "Down Units Loss"),
 (r"bad debt|write.?off", "Write-offs / Bad Debt"), (r"prepaid rent", "Prepaid Rent Change"),
 (r"delinquen", "Delinquency Change"),
 (r"market rent|base rent|subsidy|potential rent|rent - other|rental income", "Gross Potential Rent"),
 (r"application fee", "Application Fees"), (r"late fee|late charge|nsf", "Late Fees"),
 (r"month.?to.?month|\bmtm\b|m2m", "MTM Fees"),
 (r"lease (cancel|break|term)|early termination", "Lease Break Fee"),
 (r"\bpet\b", "Pet Fees"), (r"garage|parking|carport", "Parking Income"),
 (r"storage", "Storage / Common Bins"), (r"package|locker", "Package Service / Lockers"), (r"valet", "Valet Trash"),
 (r"renter.?s insurance|insurance fee income", "Renter's Insurance"), (r"insurance pass", "Insurance Pass-Thru"),
 (r"satellite|internet income|cable.*income|cable tv comm", "Reimbursement — Internet"),
 (r"reimburse|rebill|rubs", "Utility Reimbursement"),
 (r"admin|community fee|amenity fee|transfer fee|set up|risk fee", "Admin Fees"),
 (r"keys|access card|vending|stockwell|fines|forfeit|damages|consulting|interest income|vendor rebate|misc|other income|recovery|dwp|credit builder|utilities corporate|laundry", "Misc Other Income"),
 # expenses
 (r"make.?ready|redecorat|tub (glaz|refinish)|carpet (clean|repair)|unit clean|drywall|paint|resurfac|housekeeper|blinds", "Make-Ready / Turnover"),
 (r"exterminat|pest", "Pest Control"), (r"landscap", "Landscaping"),
 (r"pool|amenity|club room|exercise|gym|recreation", "Pool / Amenity Maintenance"),
 (r"trash|garbage|rubbish|recycl", "Utilities — Trash"),
 (r"water|sewer", "Utilities — Water/Sewer"), (r"propane|\bgas\b", "Utilities — Gas"),
 (r"vacant unit", "Utilities — Electric"),
 (r"electric|lights|lighting", "Utilities — Electric"),
 (r"internet|cable", "Utilities — Internet/Cable"),
 (r"utility processing|util.{0,4}process", "General & Administrative"),
 (r"management fee", "Management Fee"), (r"real estate tax|property tax|ad valorem", "Real Estate Taxes"),
 (r"insurance|workers.{0,3}comp", "Insurance"),
 (r"eviction|legal", "Legal / Eviction"),
 (r"advertis|marketing|leasing|brochure|promotion|\bils\b|seo|website|referral|prospect|reputation|online ad|paid email|tour experience|outreach|social media", "Leasing & Marketing"),
 (r"payroll|salar|wages|bonus|401k|burden|group insurance|benefit|simple ira|futa|suta|\bsui\b|medicare|soc sec|staffing|\bhealth\b", "Payroll — On-Site"),
 (r"repair|maintenance|supplies|hvac|plumb|electrical|appliance|lock|window|fire|alarm|elevator|snow|janitor|cleaning|carpentry|tile|tools", "Repairs & Maintenance"),
 (r"contract", "Contract Services"),
 (r"dues|office|bank|software|accounting|background|computer|data|uniform|printer|copier|license|permit|postage|mileage|telephone|\bcell|training|seminar|answering|technical support|verification|recruit|professional fee|travel|meeting|compliance|membership|association", "General & Administrative"),
]


def name_bucket(nm):
    s = (nm or "").lower()
    for pat, b in NAME:
        if re.search(pat, s):
            return b
    return None


def classify(acct, name):
    return acct_bucket(acct) or name_bucket(name)


# ---------- per-format leaf extractors ----------
def leaves_avana(ws):
    out = []
    for r in range(4, ws.max_row + 1):
        a, b, tot = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 15).value
        if isinstance(tot, (int, float)) and b and str(b).startswith(str(a) + " "):
            namepart = str(b).split(" - ", 1)[1] if " - " in str(b) else str(b)
            if namepart.startswith("Total"):   # roll-up account, not a leaf
                continue
            out.append((a, str(b), tot))
    return out


def leaves_ascend(ws):
    out = []
    for r in range(6, ws.max_row + 1):
        a, b, tot = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 15).value
        if isinstance(tot, (int, float)) and b not in (None, ""):
            if len(str(b)) - len(str(b).lstrip()) == 5:   # leaf indent level
                out.append((a, str(b).strip(), tot))
    return out


COPELAND_HDRS = {"income", "expenses", "rental income", "other income", "pass through reimbursement",
                 "operating expenses", "marketing and leasing", "utilities taxes and insurance",
                 "office and administrative", "payroll expense", "net income",
                 "net operating income", "total income", "total expenses"}


def leaves_copeland(ws):
    out = []
    for r in range(7, ws.max_row + 1):
        a, tot = ws.cell(r, 1).value, ws.cell(r, 14).value
        if isinstance(tot, (int, float)) and a not in (None, ""):
            s = str(a).strip()
            if not s.lower().startswith("total") and s.lower() not in COPELAND_HDRS:
                out.append((None, s, tot))
    return out


SPECS = [
 ("AVANA (Yardi)",
  r"C:/Users/erikj/Dropbox/Erik Javellana - Deal Review/Deals under review/MF_VA_Woodbridge_AvanaStoneyRidge/Operating Statement/Avana Stoney Ridge_April_T-12.xlsx",
  leaves_avana, 5346349.92, 3538883.61),
 ("ASCEND (Yardi/YSI)",
  r"C:/Users/erikj/Dropbox/Erik Javellana - Deal Review/Deals under review/MF_NC_Leland_AscendBrunswickVillage/Income Statements/Ascend Brunswick Village - T12 (2026.04).xlsx",
  leaves_ascend, 3572817.47, 2052515.08),
 ("COPELAND (Tzadik)",
  r"C:/Users/erikj/Dropbox/Erik Javellana - Deal Review/Deals under review/MF_FL_Tampa_CopelandVillage/Broker Docs/Financials/Copeland Village - P&L - 2026-04 T12.xlsx",
  leaves_copeland, 5016396.68, 3438832.68),
]


def emit_seed_csv():
    import csv
    out = os.path.join(os.path.dirname(__file__), "coa_seed.csv")
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["tier", "key", "std_bucket"])
        for num in sorted(ACCT):
            w.writerow(["acct_root", num, ACCT[num]])
        w.writerow(["acct_range", "70000-89999", "— EXCLUDED (non-OpEx) —"])
        for pat, b in NAME:
            w.writerow(["name_regex", pat, b])
    print(f"\nWROTE {out}  ({len(ACCT)} acct rules + {len(NAME)} name rules)")


def main():
    for name, F, extract, total_income, noi in SPECS:
        if not os.path.exists(F):
            print(f"\n##### {name}: FILE NOT FOUND (gitignored deal file) — skipped")
            continue
        wb = openpyxl.load_workbook(F, data_only=True, read_only=True)
        leaves = extract(wb.worksheets[0])
        unmapped = []
        mapped = 0
        bucket_tot = {}
        for a, nm, tot in leaves:
            b = classify(a, nm)
            if b:
                mapped += 1
                bucket_tot[b] = bucket_tot.get(b, 0.0) + tot
            else:
                unmapped.append((a, nm, tot))
        unmapped_amt = sum(u[2] for u in unmapped)
        pct = mapped / len(leaves) * 100 if leaves else 0
        print(f"\n##### {name}: {len(leaves)} leaves | mapped {mapped} ({pct:.0f}%) | "
              f"unmapped {len(unmapped)} lines / ${unmapped_amt:,.0f}")
        for u in unmapped:
            print(f"    UNMAPPED  {u[0] or ''}  {u[1]!r}  {u[2]:,.2f}")
        wb.close()
    emit_seed_csv()


if __name__ == "__main__":
    main()
