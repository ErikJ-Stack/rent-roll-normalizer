"""
migrate_to_v016.py — Substrate template v0.1.5 → v0.1.6

Branches 1 + 4 of the optimization effort tracked in OPTIMIZATION-DECISIONS.md.
This is the workbook-side migration; Cluster B (sign guards, partial-year T12)
is code-side and ships separately per D-12.

What this script does (in order):

  Cluster A — Correctness fixes
    A-1  UW Output R29 (Bonus wages)         — sibling-pattern fill
    A-2  UW Output R57 (Bad debt expense)    — sibling-pattern fill
    A-3  UW Output R61 (Lease / ground lease)— sibling-pattern fill + indent fix
    A-4  Rent Roll Recon H20                 — chunk over-255-char string literals

  Cluster C — Workbook Health (new sheet)
    C-1  Add 'Workbook Health' sheet at last position, hidden
    C-2  Section 1: Workbook Map (formula-driven from per-sheet AZ1:AZ5)
    C-3  Section 2: Validation (live $ checks, ±$1 tolerance per D-08)
    C-4  Section 3: Diagnostics (formula-error counts, capacity, version pills)

  Cluster D — Cover sheet + supporting work
    D-1  Add 'Cover' sheet at position 0
    D-2  Populate AZ1:AZ5 anchor cells on all 13 sheets
    D-3  Add 5 named ranges (D-14)
    D-4  Set T12 Analytics!B2 = =Property_Name
    D-5  Add light-coverage cell comments (D-13)

Idempotent: re-running on a v0.1.6 file is a no-op (skips already-applied changes).

Usage:
    python migrate_to_v016.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

SUBSTRATE_FROM = "v0.1.5"
SUBSTRATE_TO = "v0.1.6"

# Navy section header style — matches existing UW Output!R65 / R69 convention
NAVY_FILL = PatternFill(patternType="solid", fgColor="FF2F5597")
WHITE_BOLD = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")

# Per-sheet anchor metadata (D-10 + D-14 anchor population table)
ANCHOR_DATA = {
    "Cover": ("Workbook landing — versions, links, orientation", "reference", "visible", SUBSTRATE_TO, ""),
    "T12 Analytics": ("Per-Label T12 aggregation; main feed for UW Output", "aggregator", "visible", SUBSTRATE_TO, ""),
    "T12 Input": ("T12 raw paste area (MRI / Yardi / normalizer output)", "input", "visible", SUBSTRATE_TO, ""),
    "T12 Raw Data": ("Description→Label rollup with monthly trending", "aggregator", "visible", SUBSTRATE_TO, ""),
    "Rent Roll Input": ("RR normalized output paste area", "input", "visible", SUBSTRATE_TO, ""),
    "Rent Roll Recon": ("RR ↔ T12 reconciliation diagnostic", "aggregator", "visible", SUBSTRATE_TO, ""),
    "Monthly Trending": ("Per-Label monthly summary by Group", "aggregator", "visible", SUBSTRATE_TO, ""),
    "UW Output": ("Final UW-ready summary; copy to downstream sheet", "output", "visible", SUBSTRATE_TO, ""),
    "Mapping Review": ("Description_Map review for UNMATCHED descriptions", "reference", "visible", SUBSTRATE_TO, ""),
    "Description_Map": ("Canonical Description→Label vocabulary", "reference", "visible", SUBSTRATE_TO, ""),
    "RR_Calc": ("RR helper calculations", "reference", "hidden", SUBSTRATE_TO, ""),
    "T12_Calc": ("T12 helper calculations", "reference", "hidden", SUBSTRATE_TO, ""),
    "Workbook Health": ("Map / Validation / Diagnostics", "health", "hidden", SUBSTRATE_TO, ""),
}


# ============================================================================
# Helpers
# ============================================================================

def is_already_v016(wb) -> bool:
    """Idempotency guard: a v0.1.6 workbook has Cover at position 0 with B8 == v0.1.6."""
    if wb.sheetnames[0] == "Cover":
        cover = wb["Cover"]
        if cover["B8"].value == SUBSTRATE_TO:
            return True
    return False


def style_section_header(cell) -> None:
    cell.fill = NAVY_FILL
    cell.font = WHITE_BOLD
    cell.alignment = HEADER_ALIGN


def populate_anchors(wb) -> None:
    """D-2: populate AZ1:AZ5 on all 13 sheets per ANCHOR_DATA."""
    for sheet_name, (purpose, category, visibility, version, notes) in ANCHOR_DATA.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws["AZ1"] = purpose
        ws["AZ2"] = category
        ws["AZ3"] = visibility
        ws["AZ4"] = version
        # Leave AZ5 truly empty if no notes (avoids the "" → 0 render issue)
        if notes:
            ws["AZ5"] = notes


# ============================================================================
# Cluster A — Correctness fixes
# ============================================================================

def apply_cluster_a(wb) -> None:
    ws = wb["UW Output"]

    # A-1: R29 Bonus wages
    ws["B29"] = "-"
    ws["C29"] = "-"
    ws["D29"] = "-"
    ws["E29"] = "='T12 Analytics'!E64"
    ws["F29"] = "='T12 Analytics'!F64"
    # G29 already has =F29-E29

    # A-2: R57 Bad debt expense
    ws["B57"] = "-"
    ws["C57"] = "-"
    ws["D57"] = "-"
    ws["E57"] = "='T12 Analytics'!E98"
    ws["F57"] = "='T12 Analytics'!F98"
    # G57 already has =F57-E57

    # A-3: R61 Lease / ground lease (paper over per D-04)
    ws["B61"] = "-"
    ws["C61"] = "-"
    ws["D61"] = "-"
    ws["E61"] = "='T12 Analytics'!E102"
    ws["F61"] = "='T12 Analytics'!F102"
    ws["G61"] = "=F61-E61"
    # Indent fix per F-6
    a61 = ws["A61"]
    a61.alignment = Alignment(
        horizontal=a61.alignment.horizontal,
        vertical=a61.alignment.vertical,
        indent=1,
    )

    # A-4: Rent Roll Recon H20 — chunked literals (D-09)
    rr = wb["Rent Roll Recon"]
    rr["H20"] = (
        '=IF(E20=0,"Gap = $0 — RR and T12 are perfectly aligned.",'
        'IF(ABS(E20/E19)<=0.02,"Gap = "&TEXT(E20,"$#,##0")&" ("&TEXT(E20/E19,"0.0%")'
        '&") — within 2%, normal timing variance: partial-month move-ins/outs or rounding.",'
        'IF(E20<0,'
        '"⚠ Gap = "&TEXT(E20,"$#,##0")&" ("&TEXT(E20/E19,"0.0%")'
        '&"). T12 collected MORE than RR projects. Investigate: '
        '(1) Occ was higher earlier in T12 — property trending down; '
        '(2) Rates were higher in prior months — compression or new concessions; '
        '(3) Active concessions are newer than T12 average — not in full T12; ("'
        '&"4) Partial-month collections in T12 for move-ins/outs; '
        '(5) One-time adjustments or reversals in T12 income statement.",'
        '"⚠ Gap = "&TEXT(E20,"$#,##0")&" ("&TEXT(E20/E19,"0.0%")'
        '&"). RR projects MORE than T12 collected. Investigate: '
        '(1) Occupancy has improved during T12 — positive trend; '
        '(2) Rates were raised mid-T12 — RR reflects new higher rates; '
        '(3) Bad debt or uncollected rent in T12 not visible in RR; '
        '(4) Notice residents bill"'
        '&"ed but not yet collected; '
        '(5) One-time credits or refunds in T12 reduced collected revenue."'
        ')))'
    )


# ============================================================================
# Cluster D — Cover sheet
# ============================================================================

def add_cover_sheet(wb) -> None:
    """D-1: Add Cover sheet at position 0 with Property / Versions / Links / About."""
    if "Cover" in wb.sheetnames:
        return  # idempotency
    cover = wb.create_sheet("Cover", 0)

    cover["A1"] = "ALF Financial Analyzer"
    cover["A1"].font = Font(name="Arial", size=18, bold=True)
    cover["A2"] = (
        "Senior-housing underwriting workbook — RR + T12 reconciliation, "
        "UW-ready output"
    )
    cover["A2"].font = Font(name="Arial", size=10, italic=True)

    # Property block
    cover["A4"] = "Property"
    style_section_header(cover["A4"])
    cover["A5"] = "Property name"
    cover["B5"] = ""  # canonical home of Property_Name named range

    # Versions block
    cover["A7"] = "Versions"
    style_section_header(cover["A7"])
    cover["A8"] = "Substrate template"; cover["B8"] = SUBSTRATE_TO
    cover["A9"] = "Rent Roll Normalizer (app)"; cover["B9"] = "v1.12.0"
    cover["A10"] = "T12 Normalizer"; cover["B10"] = "v0.1.0"

    # Links block
    cover["A12"] = "Links"
    style_section_header(cover["A12"])
    cover["A13"] = "GitHub"
    cover["B13"] = "https://github.com/<your-repo>/rent-roll-normalizer"  # user fills final URL
    cover["A14"] = "App URL"
    cover["B14"] = "https://rrnormalizer.streamlit.app/"

    # About block
    cover["A16"] = "About"
    style_section_header(cover["A16"])
    cover["A17"] = (
        "This workbook is the underwriting destination for the Rent Roll Normalizer "
        "and T12 Normalizer pipeline. Paste normalized rent roll data into 'Rent Roll "
        "Input', T12 data into 'T12 Input', and the analytical sheets recalculate "
        "automatically."
    )
    cover["A18"] = (
        "Visible tabs flow left to right: T12 inputs → RR inputs → reconciliation → "
        "UW Output. The 'Workbook Health' sheet is hidden by default — right-click any "
        "tab and choose 'Unhide' to access it for diagnostics, validation checks, and "
        "version pills."
    )
    cover["A19"] = (
        "Property name entered at B5 above propagates to T12 Analytics via the "
        "Property_Name named range."
    )
    for r in (17, 18, 19):
        cover[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    cover.column_dimensions["A"].width = 28
    cover.column_dimensions["B"].width = 60


# ============================================================================
# Cluster C — Workbook Health sheet
# ============================================================================

def add_workbook_health_sheet(wb) -> None:
    """C-1 through C-4: hidden sheet at last position, three sections."""
    if "Workbook Health" in wb.sheetnames:
        return
    wh = wb.create_sheet("Workbook Health")
    wh.sheet_state = "hidden"

    wh["A1"] = "Workbook Health"
    wh["A1"].font = Font(name="Arial", size=14, bold=True)
    wh["A2"] = "Hidden by default. Un-hide via right-click → Unhide."
    wh["A2"].font = Font(italic=True, size=9)

    # ----- Section 1: Workbook Map -----
    wh["A4"] = "1 · WORKBOOK MAP"
    style_section_header(wh["A4"])
    headers = ["Sheet", "Purpose", "Category", "Visibility", "Version", "Notes"]
    for i, h in enumerate(headers):
        c = wh.cell(5, i + 1, h)
        c.font = Font(bold=True)

    # Order matches workbook left-to-right after migration
    map_order = [
        "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
        "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
        "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
        "Workbook Health",
    ]
    for i, sheet_name in enumerate(map_order):
        row = 6 + i
        wh.cell(row, 1, sheet_name)
        # Formula refs use single quotes around sheet names with spaces
        ref = f"'{sheet_name}'!" if " " in sheet_name else f"{sheet_name}!"
        wh.cell(row, 2, f"={ref}AZ1")
        wh.cell(row, 3, f"={ref}AZ2")
        wh.cell(row, 4, f"={ref}AZ3")
        wh.cell(row, 5, f"={ref}AZ4")
        # AZ5 may be empty; wrap in IF to render blank instead of 0
        wh.cell(row, 6, f'=IF({ref}AZ5="","",{ref}AZ5)')

    # ----- Section 2: Validation -----
    val_start = 6 + len(map_order) + 2
    wh.cell(val_start, 1, "2 · VALIDATION")
    style_section_header(wh.cell(val_start, 1))
    val_headers = ["Check", "Result", "Status"]
    for i, h in enumerate(val_headers):
        c = wh.cell(val_start + 1, i + 1, h)
        c.font = Font(bold=True)

    # V1: Source $ → Operating $ leakage (±$1 per D-08)
    # Compares T12 Raw Data F-column total against UW Output operating revenue
    # Defensive: any non-numeric ref returns "-"
    v1_row = val_start + 2
    wh.cell(v1_row, 1, "V1 · Source $ → Operating $ leakage (±$1)")
    wh.cell(v1_row, 2, "=IFERROR(ROUND(SUM('T12 Raw Data'!F6:F60)-'T12 Analytics'!E52,2),\"-\")")
    wh.cell(v1_row, 3, '=IF(ISNUMBER(B' + str(v1_row) + '),IF(ABS(B' + str(v1_row) + ')<=1,"✓","⚠"),"-")')

    # V2: UNMATCHED count in Description_Map
    v2_row = val_start + 3
    wh.cell(v2_row, 1, "V2 · Description_Map UNMATCHED count")
    wh.cell(v2_row, 2, '=COUNTIF(Description_Map!B:B,"UNMATCHED")')
    wh.cell(v2_row, 3, f'=IF(B{v2_row}=0,"✓","⚠")')

    # V3: RR period date populated
    v3_row = val_start + 4
    wh.cell(v3_row, 1, "V3 · RR period date selected")
    wh.cell(v3_row, 2, "=IF(ISNUMBER(RR_Period_Date),TEXT(RR_Period_Date,\"yyyy-mm-dd\"),\"missing\")")
    wh.cell(v3_row, 3, '=IF(ISNUMBER(RR_Period_Date),"✓","⚠")')

    # V4: T12 period date populated
    v4_row = val_start + 5
    wh.cell(v4_row, 1, "V4 · T12 period date populated")
    wh.cell(v4_row, 2, "=IF(ISNUMBER(T12_Period_Date),TEXT(T12_Period_Date,\"yyyy-mm-dd\"),\"missing\")")
    wh.cell(v4_row, 3, '=IF(ISNUMBER(T12_Period_Date),"✓","⚠")')

    # V5: Property name populated
    v5_row = val_start + 6
    wh.cell(v5_row, 1, "V5 · Property name set")
    wh.cell(v5_row, 2, '=IF(LEN(TRIM(Property_Name))>0,Property_Name,"missing")')
    wh.cell(v5_row, 3, '=IF(LEN(TRIM(Property_Name))>0,"✓","⚠")')

    # V6: RR data row count
    v6_row = val_start + 7
    wh.cell(v6_row, 1, "V6 · RR Input rows populated")
    wh.cell(v6_row, 2, "=COUNTA('Rent Roll Input'!A7:A606)")
    wh.cell(v6_row, 3, f'=IF(B{v6_row}>0,"✓","⚠")')

    # V7: T12 data row count
    v7_row = val_start + 8
    wh.cell(v7_row, 1, "V7 · T12 Input rows populated")
    wh.cell(v7_row, 2, "=COUNTA('T12 Input'!B12:B511)")
    wh.cell(v7_row, 3, f'=IF(B{v7_row}>0,"✓","⚠")')

    # ----- Section 3: Diagnostics -----
    diag_start = val_start + 10
    wh.cell(diag_start, 1, "3 · DIAGNOSTICS")
    style_section_header(wh.cell(diag_start, 1))
    diag_headers = ["Metric", "Value"]
    for i, h in enumerate(diag_headers):
        c = wh.cell(diag_start + 1, i + 1, h)
        c.font = Font(bold=True)

    # G1-G3: Capacity utilization (reads from existing UW Output capacity inputs at R70-R73)
    cap_row = diag_start + 2
    wh.cell(cap_row, 1, "G1 · Total licensed beds (from UW Output)")
    wh.cell(cap_row, 2, "='UW Output'!E73")
    wh.cell(cap_row + 1, 1, "G2 · Total available beds (from UW Output)")
    wh.cell(cap_row + 1, 2, "='UW Output'!E74")
    wh.cell(cap_row + 2, 1, "G3 · Avg occupied beds (T12)")
    wh.cell(cap_row + 2, 2, "='T12 Analytics'!E7")
    wh.cell(cap_row + 3, 1, "G4 · Capacity utilization (T12 occ / available)")
    wh.cell(cap_row + 3, 2, f"=IFERROR(B{cap_row + 2}/B{cap_row + 1},\"-\")")

    # G5-G7: Version pills (read from Cover)
    ver_row = cap_row + 5
    wh.cell(ver_row, 1, "G5 · Substrate version")
    wh.cell(ver_row, 2, "=Cover!B8")
    wh.cell(ver_row + 1, 1, "G6 · RR Normalizer version")
    wh.cell(ver_row + 1, 2, "=Cover!B9")
    wh.cell(ver_row + 2, 1, "G7 · T12 Normalizer version")
    wh.cell(ver_row + 2, 2, "=Cover!B10")

    # G8: Last-open timestamp
    ts_row = ver_row + 4
    wh.cell(ts_row, 1, "G8 · Last opened (volatile)")
    wh.cell(ts_row, 2, '=TEXT(NOW(),"yyyy-mm-dd hh:mm")')

    # Column widths
    wh.column_dimensions["A"].width = 48
    wh.column_dimensions["B"].width = 32
    wh.column_dimensions["C"].width = 12
    wh.column_dimensions["D"].width = 12
    wh.column_dimensions["E"].width = 12
    wh.column_dimensions["F"].width = 24


# ============================================================================
# Cluster D — Named ranges + Property_Name wiring
# ============================================================================

def add_named_ranges(wb) -> None:
    """D-3: 5 names per D-14."""
    targets = {
        "RR_Period_Date": "'Rent Roll Recon'!$B$2",
        "T12_Period_Date": "'T12 Analytics'!$E$2",
        "RR_Input_Data": "'Rent Roll Input'!$A$7:$S$606",
        "T12_Input_Data": "'T12 Input'!$A$12:$O$511",
        "Property_Name": "Cover!$B$5",
    }
    for name, ref in targets.items():
        # Idempotency: skip if already defined
        if name in wb.defined_names:
            continue
        dn = DefinedName(name=name, attr_text=ref)
        wb.defined_names[name] = dn


def wire_property_name(wb) -> None:
    """D-4: T12 Analytics!B2 = =Property_Name (additive — currently empty cell)."""
    ws = wb["T12 Analytics"]
    if ws["B2"].value is None or ws["B2"].value == "":
        ws["B2"] = "=Property_Name"


# ============================================================================
# Cluster D — Cell comments (light scope per D-13)
# ============================================================================

def add_cell_comments(wb) -> None:
    """D-5: comments on the 4-5 hardest-to-decode formulas."""
    AUTHOR = "Substrate v0.1.6"

    # 1. Monthly Trending B6 — INDEX/MATCH lookup pattern (representative of T12 Raw Data rollup)
    mt = wb["Monthly Trending"]
    if mt["B5"].comment is None:
        mt["B5"].comment = Comment(
            "T12 rollup pattern: INDEX/MATCH locates the Label row in T12 Raw Data "
            "and pulls the F-column total. IFERROR returns 0 if the Label is missing "
            "(e.g. a Description hasn't been mapped yet). Same pattern repeats for "
            "every Label row on this sheet.",
            AUTHOR,
        )

    # 2. T12 Analytics R37 — GPR formula (gross potential rent; key definition)
    ta = wb["T12 Analytics"]
    if ta["E37"].comment is None:
        ta["E37"].comment = Comment(
            "Gross Potential Rent (GPR): T12 actual base rent grossed up for "
            "vacancy. Pulled from T12 Raw Data 'Gross Rent Revenue' Label. Returns "
            'blank if zero so the per-care-type columns sum cleanly.',
            AUTHOR,
        )

    # 3. T12 Analytics R52 — EGI calc
    if ta["E52"].comment is None:
        ta["E52"].comment = Comment(
            "EGI = base rent (E16) + ancillary income (E23) + other line items "
            "(E47:E50). Definition matches Monthly Trending R21 and feeds UW Output "
            "as the top of the operating P&L.",
            AUTHOR,
        )

    # 4. T12 Analytics R110 — EBITDAR (post-mgmt-fee)
    if ta["E110"].comment is None:
        ta["E110"].comment = Comment(
            "EBITDAR (after mgmt fee) = EBITDARM (E108) − management fee (E106). "
            "This is the lender/DSCR convention NOI. EBITDARM at R108 is the "
            "pre-mgmt-fee version used in IRR/cap rate convention.",
            AUTHOR,
        )

    # 5. Rent Roll Recon H20 — diagnostic chain
    rr = wb["Rent Roll Recon"]
    if rr["H20"].comment is None:
        rr["H20"].comment = Comment(
            "Diagnostic for the RR-vs-T12 base rent gap (E20). Four cases: aligned "
            "(=$0), within tolerance (≤2%), T12 higher (gap<0), or RR higher "
            "(gap>0). The two ⚠ cases include a 5-item investigation list. "
            "Literals are chunked sub-255-char per Excel's per-literal cap "
            "(see OPTIMIZATION-DECISIONS.md A-4 / D-09).",
            AUTHOR,
        )


# ============================================================================
# Verification
# ============================================================================

def verify_migration(wb) -> dict:
    """Sanity checks. Returns a results dict for the caller to print."""
    results = {}

    # Cover sheet at position 0
    results["cover_at_position_0"] = wb.sheetnames[0] == "Cover"
    results["cover_substrate_version"] = wb["Cover"]["B8"].value if "Cover" in wb.sheetnames else None

    # Workbook Health hidden
    if "Workbook Health" in wb.sheetnames:
        results["wh_hidden"] = wb["Workbook Health"].sheet_state == "hidden"
    else:
        results["wh_hidden"] = False

    # Named ranges exist
    expected_names = {"RR_Period_Date", "T12_Period_Date", "RR_Input_Data",
                      "T12_Input_Data", "Property_Name"}
    actual_names = set(wb.defined_names)
    results["all_named_ranges_present"] = expected_names.issubset(actual_names)
    results["missing_named_ranges"] = list(expected_names - actual_names)

    # Anchor cells populated on all sheets in ANCHOR_DATA
    anchor_status = {}
    for s in ANCHOR_DATA:
        if s in wb.sheetnames:
            anchor_status[s] = wb[s]["AZ1"].value is not None
    results["anchor_population"] = anchor_status
    results["all_anchors_populated"] = all(anchor_status.values())

    # UW Output bug fills
    uwo = wb["UW Output"]
    results["uwo_R29_filled"] = uwo["E29"].value == "='T12 Analytics'!E64"
    results["uwo_R57_filled"] = uwo["E57"].value == "='T12 Analytics'!E98"
    results["uwo_R61_filled"] = uwo["E61"].value == "='T12 Analytics'!E102"

    # H20 chunked (no _xlfn._LONGTEXT)
    h20 = wb["Rent Roll Recon"]["H20"].value or ""
    results["h20_no_longtext"] = "_LONGTEXT" not in h20
    results["h20_starts_with_if"] = h20.startswith("=IF(")

    # T12 Analytics B2 wired to Property_Name
    results["t12_b2_wired"] = wb["T12 Analytics"]["B2"].value == "=Property_Name"

    return results


# ============================================================================
# Main
# ============================================================================

def main(input_path: str, output_path: str) -> None:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v016(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op.")
        wb.save(dst)
        return

    print(f"Migrating {SUBSTRATE_FROM} → {SUBSTRATE_TO}...")

    # Cluster D first — Cover must exist before anchors and named ranges reference it
    add_cover_sheet(wb)

    # Cluster A — single-cell formula edits
    apply_cluster_a(wb)

    # Cluster C — Workbook Health depends on AZ anchors (added below) but uses formulas,
    # so cell values resolve at workbook open. Order: create the sheet first, populate anchors
    # second (so all sheets including the new ones get anchors).
    add_workbook_health_sheet(wb)

    # D-2: anchors on all 13 sheets (Cover + 11 existing + Workbook Health)
    populate_anchors(wb)

    # D-3 + D-4: named ranges and Property_Name wiring
    add_named_ranges(wb)
    wire_property_name(wb)

    # D-5: cell comments
    add_cell_comments(wb)

    # Save
    print(f"Saving to {dst}...")
    wb.save(dst)

    # Reload and verify
    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    results = verify_migration(wb2)
    print()
    print("=== Verification results ===")
    for k, v in results.items():
        marker = "✓" if (v is True or (isinstance(v, str) and v == SUBSTRATE_TO)
                         or (isinstance(v, list) and not v)
                         or (isinstance(v, dict) and all(v.values()))) else (
                "✗" if v is False or (isinstance(v, list) and v) else "·")
        print(f"  {marker} {k}: {v}")

    all_ok = all([
        results["cover_at_position_0"],
        results["cover_substrate_version"] == SUBSTRATE_TO,
        results["wh_hidden"],
        results["all_named_ranges_present"],
        results["all_anchors_populated"],
        results["uwo_R29_filled"],
        results["uwo_R57_filled"],
        results["uwo_R61_filled"],
        results["h20_no_longtext"],
        results["h20_starts_with_if"],
        results["t12_b2_wired"],
    ])
    print()
    print(f"=== {'✓ Migration complete' if all_ok else '✗ Migration incomplete — see above'} ===")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v016.py input.xlsx output.xlsx")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
