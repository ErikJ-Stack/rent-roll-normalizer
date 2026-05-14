"""
migrate_to_v022.py — Substrate template v0.2.1 → v0.2.2

User-feedback round (Homestead populated workbook review). Three coordinated
fixes to Rent Roll Input that close out the visual + structural gaps left by
the v0.1.10 / v0.1.13 column extensions.

THREE STEPS:

  A. **Format consistency on cols V–AG.** v0.1.10 added cols V–AB
     (`2nd Person Rent $`, Move-out Date, Balance, Notes, Market PSF,
     Actual PSF, ACH) and v0.1.13 added cols AC–AG (5 per-fee ancillary
     cols). Both extensions used a different default style than the
     pre-existing cols A–U:
       - Header font sz=10 (vs 8 for A–U)
       - Data cells: no number format (General — `100` instead of
         `$100.00`)
       - Data cells: no row fill (transparent — vs the `FFFFFFC7`
         pale-yellow data band on A–U)
       - No column widths
     Step A applies the matching styling: header sz=8, $-currency on
     all monetary cols, mm/dd/yyyy on W (Move-out Date), pale-yellow
     fill on rows 7–606, sensible column widths.

  B. **Split T (Total LOC $) + add new col AH `Total Ancillary $`.**
     Currently T = `=L+M+N+O+IFERROR(AC,0)+...+IFERROR(AG,0)` —
     a mixed-semantics rollup that combines true Level-of-Care charges
     (L–O) with the 5 ancillary fees (AC–AG). The label "Total LOC $"
     no longer reflects what's in the cell. v0.2.2 splits this:
       - T (Total LOC $) reverts to pure LOC: `=L+M+N+O`
       - **NEW col AH (Total Ancillary $)** = V + AC + AD + AE + AF + AG
         (the 2nd Person Rent + 5 per-fee ancillary cols)

  C. **Rewrite U (Total Monthly Rev) for transparency.**
     Currently U = `=H + IFERROR(I,0) + T + IFERROR(V,0)` — implicit
     dependency on T's mixed semantics. After step B, U becomes
     `=H + IFERROR(I,0) + T + AH` — every contributor is visible at one
     hop, and the math is structurally identical (V is now inside AH).

Companion app.py change (RR v1.17.4):
  Adds a v0.2.2+ sentinel check to `_detect_substrate_version()`
  (Rent Roll Input!AH4 == "Total\\nAncillary $").

Companion normalizer.py change (RR v1.17.4):
  Adds `_reroute_recurring_concessions()` post-process pass that detects
  Notes-buried concession amounts (currently routed to `Other LOC $` as
  negative) and moves them to `Concession $` when the Notes pattern
  identifies them as recurring (e.g. `$XXX/mo concession`,
  `$XXX concession ending DATE`, `$XXX concession remaining`). One-time
  concessions (`(half off $XXX concession)`) are left alone.

Idempotency: gate checks BOTH `Cover!B8 == "v0.2.2"` AND
`Rent Roll Input!AH4 == "Total\\nAncillary $"`. Re-run is a no-op.

Usage:
    python tools/migration/migrate_to_v022.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from copy import copy
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SUBSTRATE_FROM = "v0.2.1"
SUBSTRATE_TO = "v0.2.2"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

RRI_SHEET = "Rent Roll Input"
DATA_FIRST_ROW = 7
DATA_LAST_ROW = 606
HEADER_ROW = 4

# Column indices (1-based)
COL_T = 20   # Total LOC $
COL_U = 21   # Total Monthly Rev
COL_V = 22   # 2nd Person Rent $
COL_W = 23   # Move-out Date
COL_X = 24   # Balance
COL_Y = 25   # Notes
COL_Z = 26   # Market PSF
COL_AA = 27  # Actual PSF
COL_AB = 28  # ACH
COL_AC = 29  # Meal Plan $
COL_AD = 30  # Scooter Fee $
COL_AE = 31  # Housekeeping $
COL_AF = 32  # Laundry $
COL_AG = 33  # Pet $
COL_AH = 34  # NEW: Total Ancillary $

NEW_AH_HEADER = "Total\nAncillary $"

# Number formats
FMT_DOLLAR = '$#,##0.00;"($"#,##0.00);-'
FMT_DOLLAR_PSF = '$#,##0.00'
FMT_DATE = "mm/dd/yyyy"
FMT_GENERAL = "General"

# Column metadata: (col_idx, number_format, width)
COL_FORMAT_SPECS = [
    (COL_V,  FMT_DOLLAR,     11),  # 2nd Person Rent $
    (COL_W,  FMT_DATE,       12),  # Move-out Date
    (COL_X,  FMT_DOLLAR,     11),  # Balance
    (COL_Y,  FMT_GENERAL,    30),  # Notes (wider for text)
    (COL_Z,  FMT_DOLLAR_PSF, 10),  # Market PSF
    (COL_AA, FMT_DOLLAR_PSF, 10),  # Actual PSF
    (COL_AB, FMT_GENERAL,     8),  # ACH (X indicator)
    (COL_AC, FMT_DOLLAR,     11),  # Meal Plan $
    (COL_AD, FMT_DOLLAR,     11),  # Scooter Fee $
    (COL_AE, FMT_DOLLAR,     11),  # Housekeeping $
    (COL_AF, FMT_DOLLAR,     11),  # Laundry $
    (COL_AG, FMT_DOLLAR,     11),  # Pet $
    (COL_AH, FMT_DOLLAR,     13),  # NEW Total Ancillary $
]

# Reference cells for style copying — H is the canonical $-currency col
# in the pre-existing layout (Actual Rate). We mirror its header + data
# styles to the V–AH cols.
REF_HEADER_COL = 8   # H4 — pre-existing header style (sz=8, navy fill, white font, bold)
REF_DATA_COL = 8     # H7 — pre-existing data fill (FFFFFFC7 pale yellow)


# ---------------------------------------------------------------------------
# Idempotency gate
# ---------------------------------------------------------------------------

def is_already_v022(wb: openpyxl.Workbook) -> bool:
    cover_b8_ok = False
    try:
        cover_b8 = wb["Cover"]["B8"].value
        cover_b8_ok = (cover_b8 == SUBSTRATE_TO)
    except Exception:
        pass

    sentinel_ok = False
    try:
        v = wb[RRI_SHEET].cell(HEADER_ROW, COL_AH).value
        sentinel_ok = (v == NEW_AH_HEADER)
    except Exception:
        pass

    return cover_b8_ok and sentinel_ok


# ---------------------------------------------------------------------------
# A. Format consistency on cols V–AH
# ---------------------------------------------------------------------------

def step_format_consistency(wb: openpyxl.Workbook) -> int:
    """Apply matching header + data styles + column widths to V–AH.

    Mutates style objects defensively — openpyxl Cell.font / .fill /
    .alignment / .border are all read-only on assignment from another
    cell; we copy() the StyleProxy attribute objects.

    Returns the number of cells modified.
    """
    ws = wb[RRI_SHEET]
    ref_h = ws.cell(HEADER_ROW, REF_HEADER_COL)
    ref_d = ws.cell(DATA_FIRST_ROW, REF_DATA_COL)

    cells_modified = 0

    for col_idx, fmt, width in COL_FORMAT_SPECS:
        col_letter = get_column_letter(col_idx)

        # 1) Header row 4 — copy font, alignment, border from H4. Keep the
        # NAVY header fill that already exists on V–AG headers (verified
        # FF1F3864). Only fix the font size / weight to match.
        h_cell = ws.cell(HEADER_ROW, col_idx)
        # Build a Font object that matches H4's (sz=8, bold, white, Calibri)
        # but preserves any existing color
        h_cell.font = Font(
            name=ref_h.font.name or "Calibri",
            size=ref_h.font.size,
            bold=True,
            color=ref_h.font.color,
        )
        h_cell.alignment = copy(ref_h.alignment)
        if ref_h.border:
            h_cell.border = copy(ref_h.border)
        cells_modified += 1

        # 2) Data rows 7..606 — fill (pale yellow) + number_format
        for row in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
            d_cell = ws.cell(row, col_idx)
            # Apply the same fill as H7 (FFFFFFC7)
            if ref_d.fill:
                d_cell.fill = copy(ref_d.fill)
            # Apply the per-col number format
            d_cell.number_format = fmt
            # Apply the same alignment + border as H7
            if ref_d.alignment:
                d_cell.alignment = copy(ref_d.alignment)
            if ref_d.border:
                d_cell.border = copy(ref_d.border)
            cells_modified += 1

        # 3) Column width
        ws.column_dimensions[col_letter].width = width

    return cells_modified


# ---------------------------------------------------------------------------
# B. Split T + add new col AH (Total Ancillary $)
# ---------------------------------------------------------------------------

def step_split_t_add_ancillary(wb: openpyxl.Workbook) -> int:
    """Rewrite T (Total LOC $) to pure LOC, add new AH (Total Ancillary $).

    Returns count of cells modified (T rows rewritten + AH header + AH rows).
    """
    ws = wb[RRI_SHEET]

    # 1) Rewrite T7:T606 — drop the IFERROR-wrapped AC..AG terms
    cells = 0
    for row in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        cell = ws.cell(row, COL_T)
        cell.value = f"=IFERROR(L{row}+M{row}+N{row}+O{row},0)"
        cells += 1

    # 2) Set new AH header — preserves the FFFFFFC7-yellow fill in step A's
    # data zone but the header itself was just stamped with the pre-existing
    # navy header fill (already there from v0.1.10) + sz=8 from step A.
    # Set the value here.
    ws.cell(HEADER_ROW, COL_AH).value = NEW_AH_HEADER
    cells += 1

    # 3) AH7:AH606 — sum V + AC + AD + AE + AF + AG (the 6 ancillary cols)
    for row in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        cell = ws.cell(row, COL_AH)
        cell.value = (
            f"=IFERROR("
            f"IFERROR(V{row},0)"
            f"+IFERROR(AC{row},0)"
            f"+IFERROR(AD{row},0)"
            f"+IFERROR(AE{row},0)"
            f"+IFERROR(AF{row},0)"
            f"+IFERROR(AG{row},0)"
            f",0)"
        )
        cells += 1

    return cells


# ---------------------------------------------------------------------------
# C. Rewrite U (Total Monthly Rev)
# ---------------------------------------------------------------------------

def step_rewrite_total_monthly_rev(wb: openpyxl.Workbook) -> int:
    """U = H + IFERROR(I,0) + T + AH. Was: H + IFERROR(I,0) + T + IFERROR(V,0).
    V is now included in AH, so we drop +V and add +AH.
    """
    ws = wb[RRI_SHEET]
    cells = 0
    for row in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        cell = ws.cell(row, COL_U)
        cell.value = f"=IFERROR(H{row}+IFERROR(I{row},0)+T{row}+IFERROR(AH{row},0),0)"
        cells += 1
    return cells


# ---------------------------------------------------------------------------
# Stamping
# ---------------------------------------------------------------------------

def stamp_versions(wb: openpyxl.Workbook) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

def verify_migration(wb: openpyxl.Workbook) -> dict:
    r: dict = {}

    # 1. Cover!B8 stamp
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    # 2. All 14 AZ4 anchors
    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    ws = wb[RRI_SHEET]

    # 3. Sentinel: AH4 header = "Total\nAncillary $"
    r["ah4"] = ws.cell(HEADER_ROW, COL_AH).value
    r["ah4_ok"] = r["ah4"] == NEW_AH_HEADER

    # 4. T7 rewritten — pure LOC (no AC reference)
    t7 = ws.cell(DATA_FIRST_ROW, COL_T).value
    r["t7"] = t7
    r["t7_ok"] = (
        isinstance(t7, str) and "L7+M7+N7+O7" in t7 and "AC7" not in t7
    )

    # 5. AH7 formula — sum of V + AC..AG
    ah7 = ws.cell(DATA_FIRST_ROW, COL_AH).value
    r["ah7"] = ah7
    r["ah7_ok"] = (
        isinstance(ah7, str)
        and "V7" in ah7 and "AC7" in ah7 and "AD7" in ah7
        and "AE7" in ah7 and "AF7" in ah7 and "AG7" in ah7
    )

    # 6. U7 rewritten — references AH (no longer +V directly)
    u7 = ws.cell(DATA_FIRST_ROW, COL_U).value
    r["u7"] = u7
    r["u7_ok"] = (
        isinstance(u7, str)
        and "H7" in u7 and "I7" in u7 and "T7" in u7 and "AH7" in u7
    )

    # 7. Header font size on V4 should be 8 (was 10)
    v4_size = ws.cell(HEADER_ROW, COL_V).font.size
    r["v4_font_size"] = v4_size
    r["v4_font_size_ok"] = v4_size == 8

    # 8. AC7 should now have $-currency number format
    ac7_fmt = ws.cell(DATA_FIRST_ROW, COL_AC).number_format
    r["ac7_fmt"] = ac7_fmt
    r["ac7_fmt_ok"] = "$" in ac7_fmt

    # 9. AC7 should have a non-transparent fill now
    ac7_fill = ws.cell(DATA_FIRST_ROW, COL_AC).fill
    fill_rgb = ac7_fill.fgColor.rgb if ac7_fill and ac7_fill.fgColor else None
    r["ac7_fill"] = fill_rgb
    r["ac7_fill_ok"] = fill_rgb == "FFFFFFC7"

    # 10. W7 (Move-out Date) should have date format
    w7_fmt = ws.cell(DATA_FIRST_ROW, COL_W).number_format
    r["w7_fmt"] = w7_fmt
    r["w7_fmt_ok"] = "mm/dd/yyyy" in w7_fmt

    # 11. Column widths set on V–AH
    widths_ok = True
    for col_idx, _, expected_w in COL_FORMAT_SPECS:
        cd = ws.column_dimensions.get(get_column_letter(col_idx))
        if not cd or cd.width != expected_w:
            widths_ok = False
            break
    r["widths_ok"] = widths_ok

    # 12. T7 last row (T606) rewritten too
    t606 = ws.cell(DATA_LAST_ROW, COL_T).value
    r["t606_ok"] = (
        isinstance(t606, str) and f"L{DATA_LAST_ROW}" in t606 and "AC" not in t606
    )

    return r


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v022(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...\n")

    print("Step A — format consistency on Rent Roll Input cols V-AH:")
    nA = step_format_consistency(wb)
    print(f"  modified {nA} cells (header + data + widths)")

    print("\nStep B — split T into pure LOC + add new col AH (Total Ancillary $):")
    nB = step_split_t_add_ancillary(wb)
    print(f"  modified {nB} cells (T rewrite + AH header + AH formula)")

    print("\nStep C — rewrite U (Total Monthly Rev) to reference AH instead of V:")
    nC = step_rewrite_total_monthly_rev(wb)
    print(f"  modified {nC} cells")

    stamp_versions(wb)
    print(f"\nStep D — stamped substrate version -> {SUBSTRATE_TO} (14 anchors)")

    print(f"\nSaving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"   1. Cover!B8 = {r['cover_b8']!r}                  : {r['cover_b8_ok']}")
    print(f"   2. All 14 AZ4 = {SUBSTRATE_TO}                                 : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"   3. AH4 sentinel = 'Total\\nAncillary $'                        : {r['ah4_ok']}")
    print(f"   4. T7 = pure LOC (L+M+N+O, no AC..AG)                          : {r['t7_ok']}")
    print(f"   5. AH7 = V+AC+AD+AE+AF+AG (Total Ancillary)                    : {r['ah7_ok']}")
    print(f"   6. U7 = H+I+T+AH (rewritten, no +V)                            : {r['u7_ok']}")
    print(f"   7. V4 header font size = 8 (was 10)                            : {r['v4_font_size_ok']}")
    print(f"   8. AC7 number_format includes $ sign                           : {r['ac7_fmt_ok']}")
    print(f"   9. AC7 fill = FFFFFFC7 (pale yellow)                           : {r['ac7_fill_ok']}")
    print(f"  10. W7 number_format = mm/dd/yyyy (Move-out Date)               : {r['w7_fmt_ok']}")
    print(f"  11. Column widths set on V-AH per spec                          : {r['widths_ok']}")
    print(f"  12. T{DATA_LAST_ROW} (last row) also rewritten (sample check)              : {r['t606_ok']}")

    all_ok = all([
        r["cover_b8_ok"], r["az4_all"], r["ah4_ok"],
        r["t7_ok"], r["ah7_ok"], r["u7_ok"],
        r["v4_font_size_ok"], r["ac7_fmt_ok"], r["ac7_fill_ok"],
        r["w7_fmt_ok"], r["widths_ok"], r["t606_ok"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v022.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
