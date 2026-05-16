"""
migrate_to_v024.py — Substrate template v0.2.3 → v0.2.4

Two cosmetic-but-visible fixes against the v0.2.3 substrate, both surfaced
by the user during inspection of the populated Homestead Analyzer on
2026-05-16:

  BL-0016  Rent Roll Input!AH4 — missing header fill
           When the AH "Total Ancillary $" column was added in
           substrate v0.2.2, the header cell received the correct
           white+bold font but no background fill (fill=00000000,
           transparent). White-on-default renders as a blank cell in
           Excel — the column header is invisible. By the substrate's
           own header palette, AH is a *computed* column (formula
           =IFERROR(V+AC+AD+AE+AF+AG,0)) and should match the green
           fill FF1F6B52 already on T4 ("Total LOC $") and U4
           ("Total Monthly Rev"), not the navy FF1F3864 used for input
           columns A4..AG4.

  BL-0017  Consistent "intentionally blank" treatment, workbook-wide
           144 cells across the substrate hold the 3-character literal
           text `"-"` (with quote chars in the payload) as an
           "intentionally blank" placeholder marker. In Excel, the
           quote chars render visibly — so what should signal "no data
           by design" instead reads as a typo. v0.2.4 replaces the
           per-cell payload with a coherent visual treatment:

             - value  : em-dash `—` (plain text, no quotes/formula)
             - fill   : solid FFF2F2F2 (very light gray)
             - font   : preserve size/bold/italic, override color → FFA0A0A0
             - align  : horizontal=center (preserve vertical/wrap/indent)

           Distribution of the 144 targets (all sheets):
             - T12 Analytics  : E36, G36   (2 cells — flagged in user
                                            diagnosis; H36's design note
                                            says "col E blank by design")
             - Rent Roll Recon: D109       (1 cell  — MC "Other / unmapped"
                                            Avg Care Level has no avg-rate
                                            concept)
             - UW Output      : 141 cells  — cols B/C/D × rows
                                            {8-12, 22-28, 30-36, 38-56,
                                            58-60, 62-64, 66-68}; per-care-type
                                            columns adjacent to total-only
                                            formulas (cols E/F)

           NOT touched: formula-conditional blanks like T12 Analytics
           E37/G37/H38 that return "" only when source data is missing.
           Those are "blank when data isn't here" not "blank by design";
           styling them permanently would mislead when they populate.

OPERATIONS:

  A. Set Rent Roll Input!AH4 fill = PatternFill(FF1F6B52, solid)
     to match T4/U4 (computed-column header palette).
  B. Apply intentional-blank treatment to all 144 placeholder cells:
     value=`—`, fill=light gray solid, font=medium gray, center align.
  C. Stamp Cover!B8 + 14 AZ4 anchors to v0.2.4.
  D. Verification block (12 checks).

Idempotent: gate checks (i) version stamp, (ii) AH4 fill is the green
FF1F6B52, (iii) UW Output!B8 sentinel has been re-styled (value=`—` +
gray fill). Re-running on a partial-state file safely re-applies any
unfinished step.

Usage:
    python tools/migration/migrate_to_v024.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SUBSTRATE_FROM = "v0.2.3"
SUBSTRATE_TO = "v0.2.4"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# Step A — AH4 green fill (matches T4/U4 computed-column header palette).
GREEN_FILL_RGB = "FF1F6B52"
GREEN_FILL = PatternFill(
    start_color=GREEN_FILL_RGB, end_color=GREEN_FILL_RGB, fill_type="solid"
)

# Step B — intentional-blank treatment.
EMDASH = "—"  # em-dash
BLANK_FILL_RGB = "FFF2F2F2"
BLANK_FONT_RGB = "FFA0A0A0"
BLANK_FILL = PatternFill(
    start_color=BLANK_FILL_RGB, end_color=BLANK_FILL_RGB, fill_type="solid"
)
DASH_ARTIFACT = '"-"'  # literal-text artifact being replaced

# UW Output row groups — cols B/C/D × these row ranges hold the placeholder.
UW_OUTPUT_ROW_GROUPS = [
    (8, 12),    # Other-revenue lines + EGI
    (22, 28),   # OpEx section 1
    (30, 36),   # OpEx section 2
    (38, 56),   # OpEx detail
    (58, 60),   # Total OpEx + NOI lines
    (62, 64),   # Capex
    (66, 68),   # Below-the-line
]


def build_blank_targets() -> list[tuple[str, str]]:
    """Enumerate all 144 intentional-blank placeholder coords."""
    targets: list[tuple[str, str]] = []
    targets.append(("T12 Analytics", "E36"))
    targets.append(("T12 Analytics", "G36"))
    targets.append(("Rent Roll Recon", "D109"))
    for start_row, end_row in UW_OUTPUT_ROW_GROUPS:
        for r in range(start_row, end_row + 1):
            for col in ("B", "C", "D"):
                targets.append(("UW Output", f"{col}{r}"))
    return targets


BLANK_TARGETS = build_blank_targets()
assert len(BLANK_TARGETS) == 144, f"Expected 144 targets, built {len(BLANK_TARGETS)}"


def _is_already_styled(cell) -> bool:
    """Cell has the intentional-blank treatment fully applied."""
    if cell.value != EMDASH:
        return False
    fg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    return fg == BLANK_FILL_RGB


def is_already_v024(wb) -> bool:
    """Gate: version stamp + AH4 green fill + UW Output sentinel styled."""
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    ah4 = wb["Rent Roll Input"]["AH4"]
    fg = ah4.fill.fgColor.rgb if ah4.fill and ah4.fill.fgColor else None
    if fg != GREEN_FILL_RGB:
        return False
    sentinel = wb["UW Output"]["B8"]
    return _is_already_styled(sentinel)


def fix_ah4_fill(wb) -> int:
    """A: apply green PatternFill to Rent Roll Input!AH4."""
    ws = wb["Rent Roll Input"]
    ws["AH4"].fill = GREEN_FILL
    return 1


def apply_blank_styling(wb) -> tuple[int, int]:
    """B: apply intentional-blank treatment to the 144 placeholder cells.

    Idempotent — skips cells that already have the styling. Accepts any
    of three pre-states per cell:
      1. literal `"-"` text       (the v0.2.3 substrate's stored value)
      2. None / blank             (e.g., E36/G36 from an aborted earlier
                                   v0.2.4 attempt that cleared but didn't
                                   restyle)
      3. already-styled em-dash   (this run is a no-op for that cell)

    Returns (styled_count, skipped_count).
    """
    styled = 0
    skipped = 0
    for sheet_name, coord in BLANK_TARGETS:
        cell = wb[sheet_name][coord]

        if _is_already_styled(cell):
            skipped += 1
            continue

        # value: em-dash, plain text (not formula, not quoted).
        cell.value = EMDASH

        # fill: solid light gray.
        cell.fill = BLANK_FILL

        # font: preserve size/bold/italic/name, override color to medium gray.
        old = cell.font
        cell.font = Font(
            name=old.name,
            size=old.size,
            bold=old.bold,
            italic=old.italic,
            underline=old.underline,
            strike=old.strike,
            color=BLANK_FONT_RGB,
            family=old.family,
            scheme=old.scheme,
        )

        # alignment: center horizontal; preserve vertical/wrap/indent.
        oa = cell.alignment
        cell.alignment = Alignment(
            horizontal="center",
            vertical=oa.vertical,
            indent=oa.indent,
            wrap_text=oa.wrap_text,
            shrink_to_fit=oa.shrink_to_fit,
            text_rotation=oa.text_rotation,
        )

        styled += 1
    return styled, skipped


def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


def verify_migration(wb) -> dict:
    r = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # AH4 fill: must be the green FF1F6B52 matching T4/U4
    ws_rr = wb["Rent Roll Input"]
    ah4 = ws_rr["AH4"]
    ah4_fg = ah4.fill.fgColor.rgb if ah4.fill and ah4.fill.fgColor else None
    r["ah4_fill"] = ah4_fg
    r["ah4_fill_green"] = ah4_fg == GREEN_FILL_RGB
    r["ah4_text_ok"] = ah4.value == "Total\nAncillary $"
    r["ah4_font_ok"] = bool(ah4.font and ah4.font.bold)
    t4_fg = ws_rr["T4"].fill.fgColor.rgb if ws_rr["T4"].fill else None
    u4_fg = ws_rr["U4"].fill.fgColor.rgb if ws_rr["U4"].fill else None
    r["t4_u4_intact"] = t4_fg == GREEN_FILL_RGB and u4_fg == GREEN_FILL_RGB
    ag4_fg = ws_rr["AG4"].fill.fgColor.rgb if ws_rr["AG4"].fill else None
    r["ag4_navy_intact"] = ag4_fg == "FF1F3864"

    # All 144 targets must be fully styled.
    unstyled = []
    for sheet_name, coord in BLANK_TARGETS:
        if not _is_already_styled(wb[sheet_name][coord]):
            unstyled.append(f"{sheet_name}!{coord}")
    r["all_styled"] = len(unstyled) == 0
    r["unstyled_examples"] = unstyled[:5]
    r["target_count"] = len(BLANK_TARGETS)

    # Sample-check a few targets for full treatment (font color + alignment).
    sentinel = wb["UW Output"]["B8"]
    r["sentinel_font_color"] = (
        sentinel.font.color.rgb if sentinel.font and sentinel.font.color else None
    ) == BLANK_FONT_RGB
    r["sentinel_align_center"] = sentinel.alignment.horizontal == "center"

    # No remaining literal `"-"` text anywhere in the workbook.
    remaining_dash = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "s" and cell.value == DASH_ARTIFACT:
                    remaining_dash.append(f"{sheet_name}!{cell.coordinate}")
    r["dash_remaining"] = remaining_dash
    r["dash_clean"] = len(remaining_dash) == 0

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v024(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    n_a = fix_ah4_fill(wb)
    print(f"  A: AH4 fill -> {GREEN_FILL_RGB} ({n_a} cell)")

    styled, skipped = apply_blank_styling(wb)
    print(
        f"  B: intentional-blank treatment applied to {styled} cell(s)"
        f" ({skipped} already styled, skipped)"
    )

    stamp_versions(wb)
    print(f"  C: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:24s}    : {r['cover_b8_ok']}")
    print(f"  All 14 AZ4 = {SUBSTRATE_TO}              : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  AH4 fill = {r['ah4_fill']!r:14s} (target {GREEN_FILL_RGB!r}): {r['ah4_fill_green']}")
    print(f"  AH4 text 'Total\\nAncillary $' intact   : {r['ah4_text_ok']}")
    print(f"  AH4 font bold intact                   : {r['ah4_font_ok']}")
    print(f"  T4 + U4 reference fills still green    : {r['t4_u4_intact']}")
    print(f"  AG4 (last input col) still navy        : {r['ag4_navy_intact']}")
    print(f"  All {r['target_count']} placeholder cells styled       : {r['all_styled']}")
    if not r["all_styled"]:
        for x in r["unstyled_examples"]:
            print(f"      unstyled: {x}")
    print(f"  Sentinel font color = {BLANK_FONT_RGB}        : {r['sentinel_font_color']}")
    print(f"  Sentinel alignment center              : {r['sentinel_align_center']}")
    print(f"  Literal `\"-\"` text remaining: {len(r['dash_remaining'])}     : {r['dash_clean']}")

    all_ok = (
        r["cover_b8_ok"] and r["az4_all"]
        and r["ah4_fill_green"]
        and r["ah4_text_ok"] and r["ah4_font_ok"]
        and r["t4_u4_intact"] and r["ag4_navy_intact"]
        and r["all_styled"]
        and r["sentinel_font_color"] and r["sentinel_align_center"]
        and r["dash_clean"]
    )
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v024.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
