"""
migrate_to_v0214.py — Substrate template v0.2.13 → v0.2.14

Closes the AI-column conflict surfaced by Track 4 / UWT v0.2.0 (2026-05-25):

  - 2026-05-25 ALF UW Template handoff contract reserves
    `Rent Roll Input!AI` for **Deposit** (UW Template Rent Roll Analysis
    col M).
  - RR v1.18.0 + substrate v0.2.13 (also 2026-05-25, earlier in the day)
    put **Preleased Date** at AI.

Per user decision: Deposit gets AI; Preleased Date relocates to AJ.

THE FIX (4 surface changes):

  1. Relocate `Rent Roll Input!AI4` header ("Preleased\\nDate") → AJ4.
     Copy style + column width along with the header.
  2. If any rows in AI7:AI606 hold Preleased Date data (populated user
     workbooks coming forward from v0.2.13), relocate values + number
     formats to AJ7:AJ606. Bundled-file case is a no-op.
  3. Write new AI4 header = "Deposit". Style + column width copied from
     X4 (Balance — existing $-typed input column). Restyle AI7:AI606
     number_format to match X7 ($-typed data cells).
  4. Widen `RR_Input_Data` named range:
       'Rent Roll Input'!$A$7:$S$606  →  'Rent Roll Input'!$A$7:$AJ$606
     The legacy A:S scope predates the v1.16.0+ ancillary columns; the
     handoff contract's Rent Roll path needs the full row.

Section N on Rent Roll Recon (v0.2.13) uses COUNTIFS on Status="Preleased"
in col E — NOT on the Preleased Date column directly — so relocating the
date column from AI to AJ has zero formula impact on the rest of the
substrate. Verified by grep on the Rent Roll Recon sheet pre-flight.

Idempotency:
  - Gate: Cover!B8 == "v0.2.14" AND AI4 == "Deposit".
  - AI4 → AJ4 header relocation: only runs if AI4 currently holds the
    Preleased header text.
  - AI7:AI606 data relocation: per-cell guard (only if AI[r] is populated).
  - Named range widen: only writes if current range != target.

BUNDLED FILE STATUS:
  Bundled `ALF_Financial_Analyzer_Only.xlsx` (currently v0.2.13) is updated
  in place by running this migration on it. Empty data block (AI7:AI606
  all None) — the relocation is header-only on the bundled file.

CROSS-TRACK:
  Companion change in `analyzer_rr_writer.py` (RR v1.18.1): COL_AI_INDEX
  renamed COL_AJ_INDEX = 36 for Preleased Date routing. New (unused)
  COL_AI_INDEX = 35 added for Deposit slot — clear-only, no parser support
  yet. Deposit field parsing (mappings.py / normalizer.py / Condensed_RR)
  deferred until a source fixture with Deposit data exists.

Usage:
    python tools/migration/migrate_to_v0214.py input.xlsx output.xlsx
"""
from __future__ import annotations

import copy
import sys
from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

SUBSTRATE_FROM = "v0.2.13"
SUBSTRATE_TO = "v0.2.14"

# 16-sheet anchor list — unchanged since v0.2.10.
ANCHOR_SHEETS = (
    "Cover", "Dashboard",
    "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending",
    "AR & Collections",
    "UW Output", "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

RR_INPUT_SHEET = "Rent Roll Input"
DATA_START_ROW = 7
DATA_END_ROW = 606

# Header text constants
PRELEASED_HEADER_TEXT = "Preleased\nDate"
DEPOSIT_HEADER_TEXT = "Deposit"

# Style sources
PRELEASED_STYLE_SOURCE_HEADER = "Q4"   # date-typed input header
PRELEASED_STYLE_SOURCE_DATA = "Q7"
DEPOSIT_STYLE_SOURCE_HEADER = "X4"     # Balance — $-typed input header
DEPOSIT_STYLE_SOURCE_DATA = "X7"

# Named range
NAMED_RANGE = "RR_Input_Data"
NAMED_RANGE_OLD_VALUE = "'Rent Roll Input'!$A$7:$S$606"
NAMED_RANGE_NEW_VALUE = "'Rent Roll Input'!$A$7:$AJ$606"


def is_already_v0214(wb) -> bool:
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    ws = wb[RR_INPUT_SHEET]
    return ws["AI4"].value == DEPOSIT_HEADER_TEXT


def _copy_style(target_cell, source_cell) -> None:
    if not source_cell.has_style:
        return
    target_cell.font = copy.copy(source_cell.font)
    target_cell.fill = copy.copy(source_cell.fill)
    target_cell.alignment = copy.copy(source_cell.alignment)
    target_cell.border = copy.copy(source_cell.border)
    target_cell.number_format = source_cell.number_format


# ----------------------------------------------------------------------
# Step 1: Relocate AI4 header → AJ4
# ----------------------------------------------------------------------
def relocate_preleased_header(wb) -> dict:
    ws = wb[RR_INPUT_SHEET]
    counts = {"patched": 0, "skipped": 0}
    log = []

    ai_cell = ws["AI4"]
    aj_cell = ws["AJ4"]

    if ai_cell.value != PRELEASED_HEADER_TEXT:
        # Either already migrated, or in an unexpected state. If AJ already
        # holds the Preleased header, treat as already-migrated.
        if aj_cell.value == PRELEASED_HEADER_TEXT:
            counts["skipped"] += 1
            log.append("  SKIP relocate header: AJ4 already holds Preleased header")
            return {"counts": counts, "log": log}
        counts["skipped"] += 1
        log.append(
            f"  SKIP relocate header: AI4 = {ai_cell.value!r}, expected "
            f"{PRELEASED_HEADER_TEXT!r}"
        )
        return {"counts": counts, "log": log}

    # Copy header text + style to AJ
    aj_cell.value = PRELEASED_HEADER_TEXT
    _copy_style(aj_cell, ai_cell)

    # Copy column width AI → AJ
    if "AI" in ws.column_dimensions:
        ws.column_dimensions["AJ"].width = ws.column_dimensions["AI"].width

    # Clear AI4 (will be overwritten by Step 3 with Deposit)
    ai_cell.value = None
    # Don't clear AI4 style — Step 3 will reapply.

    counts["patched"] += 1
    log.append(f"  PATCH AI4 → AJ4: relocated Preleased header text + style + width")
    return {"counts": counts, "log": log}


# ----------------------------------------------------------------------
# Step 2: Relocate AI7:AI606 data → AJ7:AJ606 (if any)
# ----------------------------------------------------------------------
def relocate_preleased_data(wb) -> dict:
    ws = wb[RR_INPUT_SHEET]
    counts = {"rows_moved": 0, "rows_empty": 0}
    log = []

    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        ai_cell = ws.cell(row=r, column=35)  # AI
        aj_cell = ws.cell(row=r, column=36)  # AJ

        if ai_cell.value is None:
            counts["rows_empty"] += 1
            continue

        # Move value
        aj_cell.value = ai_cell.value
        # Preserve number format if it was set (date format)
        if ai_cell.number_format and ai_cell.number_format != "General":
            aj_cell.number_format = ai_cell.number_format
        # Clear AI cell value (style reset comes in Step 3)
        ai_cell.value = None

        counts["rows_moved"] += 1

    if counts["rows_moved"]:
        log.append(
            f"  PATCH relocate data: {counts['rows_moved']} populated rows "
            f"moved AI → AJ"
        )
    else:
        log.append("  PATCH relocate data: no populated rows (bundled-file case)")
    return {"counts": counts, "log": log}


# ----------------------------------------------------------------------
# Step 3: Write Deposit header at AI4 + restyle AI column
# ----------------------------------------------------------------------
def write_deposit_header(wb) -> dict:
    ws = wb[RR_INPUT_SHEET]
    counts = {"patched": 0, "skipped": 0}
    log = []

    ai_cell = ws["AI4"]
    if ai_cell.value == DEPOSIT_HEADER_TEXT:
        counts["skipped"] += 1
        log.append("  SKIP Deposit header: AI4 already holds Deposit")
    else:
        ai_cell.value = DEPOSIT_HEADER_TEXT
        src = ws[DEPOSIT_STYLE_SOURCE_HEADER]
        _copy_style(ai_cell, src)
        if "X" in ws.column_dimensions:
            ws.column_dimensions["AI"].width = ws.column_dimensions["X"].width
        counts["patched"] += 1
        log.append(
            f"  PATCH AI4: wrote {DEPOSIT_HEADER_TEXT!r} with "
            f"{DEPOSIT_STYLE_SOURCE_HEADER} style + col width"
        )

    # Apply $-typed number format to AI7:AI606 data cells.
    # Pull number_format from X7. We only restyle empty cells to avoid
    # disturbing any analyst-typed Deposit values (idempotency).
    src_data = ws[DEPOSIT_STYLE_SOURCE_DATA]
    src_fmt = src_data.number_format
    if src_fmt and src_fmt != "General":
        restyled = 0
        for r in range(DATA_START_ROW, DATA_END_ROW + 1):
            cell = ws.cell(row=r, column=35)  # AI
            if cell.value is None and cell.number_format != src_fmt:
                cell.number_format = src_fmt
                restyled += 1
        log.append(
            f"  PATCH AI{DATA_START_ROW}:AI{DATA_END_ROW}: "
            f"applied number_format={src_fmt!r} to {restyled} empty cells"
        )

    return {"counts": counts, "log": log}


# ----------------------------------------------------------------------
# Step 4: Widen RR_Input_Data named range
# ----------------------------------------------------------------------
def widen_named_range(wb) -> dict:
    counts = {"patched": 0, "skipped": 0}
    log = []

    if NAMED_RANGE not in wb.defined_names:
        log.append(f"  WARN: named range {NAMED_RANGE!r} not found")
        return {"counts": counts, "log": log}

    cur = wb.defined_names[NAMED_RANGE].value
    if cur == NAMED_RANGE_NEW_VALUE:
        counts["skipped"] += 1
        log.append(f"  SKIP named range: already at {cur!r}")
        return {"counts": counts, "log": log}

    # openpyxl v3.x: assignment via DefinedName replacement is the canonical
    # mutation pattern (defined_names.append removed).
    wb.defined_names[NAMED_RANGE] = DefinedName(NAMED_RANGE, attr_text=NAMED_RANGE_NEW_VALUE)
    counts["patched"] += 1
    log.append(
        f"  PATCH named range {NAMED_RANGE}: {cur!r} → {NAMED_RANGE_NEW_VALUE!r}"
    )
    return {"counts": counts, "log": log}


# ----------------------------------------------------------------------
# Step 5: Version stamps
# ----------------------------------------------------------------------
def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ----------------------------------------------------------------------
# Verify
# ----------------------------------------------------------------------
def verify_migration(wb) -> dict:
    r = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    ws = wb[RR_INPUT_SHEET]
    r["ai4"] = ws["AI4"].value
    r["ai4_ok"] = r["ai4"] == DEPOSIT_HEADER_TEXT
    r["aj4"] = ws["AJ4"].value
    r["aj4_ok"] = r["aj4"] == PRELEASED_HEADER_TEXT

    # Named range
    nr = wb.defined_names.get(NAMED_RANGE)
    r["named_range_value"] = nr.value if nr else None
    r["named_range_ok"] = r["named_range_value"] == NAMED_RANGE_NEW_VALUE

    # Anchors
    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_count"] = len(az4)
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())

    # Sheet count unchanged at 16
    r["sheet_count"] = len(wb.sheetnames)
    r["sheet_count_ok"] = r["sheet_count"] == 16

    # AI column width copied (sanity: AI width should match X width if X has one)
    ai_w = ws.column_dimensions.get("AI")
    x_w = ws.column_dimensions.get("X")
    r["ai_width"] = ai_w.width if ai_w else None
    r["x_width"] = x_w.width if x_w else None

    # AJ column width copied (sanity: AJ width should match what AI had before — Q width)
    aj_w = ws.column_dimensions.get("AJ")
    q_w = ws.column_dimensions.get("Q")
    r["aj_width"] = aj_w.width if aj_w else None
    r["q_width"] = q_w.width if q_w else None

    return r


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0214(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    cur_version = wb["Cover"]["B8"].value
    if cur_version != SUBSTRATE_FROM:
        print(
            f"  WARN: Cover!B8 = {cur_version!r}, expected {SUBSTRATE_FROM!r}. "
            f"Proceeding anyway."
        )

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    hdr_result = relocate_preleased_header(wb)
    for line in hdr_result["log"]:
        print(line)

    data_result = relocate_preleased_data(wb)
    for line in data_result["log"]:
        print(line)

    dep_result = write_deposit_header(wb)
    for line in dep_result["log"]:
        print(line)

    nr_result = widen_named_range(wb)
    for line in nr_result["log"]:
        print(line)

    stamp_versions(wb)
    print(
        f"  Stamped substrate version -> {SUBSTRATE_TO} on "
        f"Cover!B8 + {len(ANCHOR_SHEETS)} AZ4 anchors"
    )

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    checks = [
        ("Cover!B8 = " + repr(r["cover_b8"]), r["cover_b8_ok"]),
        ("AI4 = " + repr(r["ai4"]), r["ai4_ok"]),
        ("AJ4 = " + repr(r["aj4"]), r["aj4_ok"]),
        (
            f"RR_Input_Data range = {r['named_range_value']!r}",
            r["named_range_ok"],
        ),
        (f"Sheet count = {r['sheet_count']} (expected 16)", r["sheet_count_ok"]),
        (f"All {r['az4_count']} AZ4 anchors = {SUBSTRATE_TO}", r["az4_all"]),
        (
            f"AI col width = {r['ai_width']!r} (vs X width {r['x_width']!r})",
            r["ai_width"] is not None
            and r["x_width"] is not None
            and abs((r["ai_width"] or 0) - (r["x_width"] or 0)) < 0.01,
        ),
        (
            f"AJ col width = {r['aj_width']!r} (vs Q width {r['q_width']!r})",
            r["aj_width"] is not None,  # AJ width should be set; equality vs Q optional
        ),
    ]
    for desc, ok in checks:
        flag = "[OK]  " if ok else "[FAIL]"
        print(f"  {flag}  {desc}")

    all_ok = all(ok for _, ok in checks)

    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0214.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
