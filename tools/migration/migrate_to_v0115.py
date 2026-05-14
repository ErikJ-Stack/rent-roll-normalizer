"""
migrate_to_v0115.py - Substrate template v0.1.14 -> v0.1.15

Closes UW-BACKLOG BL-0002. Improves V5 (AL Acuity Mix) on T12 Analytics
when source rent rolls have no per-bed acuity tier data (Homestead-style
broker-condensed formats, flat-rate AL operators).

Two coordinated tweaks — pure substrate, no code:

  A. Blank the chart data when no acuity data exists. The 8 source cells
     at Rent Roll Recon!D59:D66 currently use:
         =IFERROR(SUMIFS(...), 0)
     This evaluates to 0 (not blank) for sources without acuity tiers,
     causing the V5 doughnut to render 8 equal "0" slices with labels in
     the legend. We wrap the existing formula so it returns "" (empty)
     when `B$67 = 0` (no occupied AL with any acuity tier filled in):
         =IF($B$67=0, "", IFERROR(SUMIFS(...), 0))
     Doughnut charts treat empty cells as "no slice" — the chart renders
     as an empty frame instead of zero-slices-with-legend.

  B. Style the existing K45 conditional note (visible since v0.1.8) to be
     visually prominent. K45 already contains:
         =IF('Rent Roll Recon'!D67=0, "Property has no AL acuity data — ...",
             IF(..., "⚠ Acuity skewed: top tier > 50%", "✓ Acuity distribution ..."))
     We add bold font + pale-yellow fill so the empty-state message reads
     as a warning attached to V5 rather than an ignorable label.

  C. Stamp Cover!B8 + 13 AZ4 anchors to v0.1.14 -> v0.1.15.
  D. 6-check verification.

Idempotent: gate checks BOTH the version stamp AND that D59's formula
already starts with the v0.1.15 wrapper `=IF($B$67=0,"",`.

Usage:
    python tools/migration/migrate_to_v0115.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SUBSTRATE_FROM = "v0.1.14"
SUBSTRATE_TO = "v0.1.15"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

RECON_SHEET = "Rent Roll Recon"
ANALYTICS_SHEET = "T12 Analytics"

# V5 data range
ACUITY_ROW_FIRST = 59
ACUITY_ROW_LAST = 66
ACUITY_TOTAL_ROW = 67  # B67 = SUM(B59:B66); D67 = SUM(D59:D66)

# K45 = V5 conditional note location
NOTE_ROW = 45
NOTE_COL = 11  # K

# Wrapper prefix used by v0.1.15 (also the idempotency marker)
WRAPPER_PREFIX = "=IF($B$67=0,\"\","


def is_already_v0115(wb) -> bool:
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    d59 = wb[RECON_SHEET].cell(ACUITY_ROW_FIRST, 4).value
    return isinstance(d59, str) and d59.startswith(WRAPPER_PREFIX)


# ----- A. Blank D59:D66 when no acuity data -----

def wrap_acuity_formulas(wb) -> int:
    """Wrap existing D59:D66 formulas so they return "" instead of 0 when
    B67=0 (no occupied AL with acuity tier set). Pattern:
        old: =IFERROR(SUMIFS(...), 0)
        new: =IF($B$67=0,"",IFERROR(SUMIFS(...), 0))
    Preserves the inner formula verbatim — analyst-customized SUMIFS (if
    any) survive the migration unchanged.
    """
    ws = wb[RECON_SHEET]
    n = 0
    for r in range(ACUITY_ROW_FIRST, ACUITY_ROW_LAST + 1):
        cell = ws.cell(r, 4)  # col D
        v = cell.value
        if not isinstance(v, str):
            continue
        if v.startswith(WRAPPER_PREFIX):
            # already wrapped — re-running, no-op
            continue
        if not v.startswith("="):
            continue
        # Strip the leading "=" so we can splice into the new formula
        inner = v[1:]
        cell.value = f'=IF($B$67=0,"",{inner})'
        n += 1
    return n


# ----- B. Style K45 (V5 conditional note) -----

PALE_YELLOW = "FFFFF2CC"
DARK_TEXT = "FF1F1F1F"


def style_v5_note(wb) -> int:
    """Apply bold + yellow fill to T12 Analytics!K45 so the existing V5
    conditional note reads as a warning attached to the chart. The cell's
    formula is left untouched — only the visual styling changes."""
    ws = wb[ANALYTICS_SHEET]
    cell = ws.cell(NOTE_ROW, NOTE_COL)
    cell.font = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
    cell.fill = PatternFill(fill_type="solid", fgColor=PALE_YELLOW)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return 1


# ----- C. Versioning -----

def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ----- D. Verify -----

def verify_migration(wb) -> dict:
    r: dict = {}
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # Acuity formula wrapper applied to all 8 rows
    ws = wb[RECON_SHEET]
    wrapped = 0
    for r_ in range(ACUITY_ROW_FIRST, ACUITY_ROW_LAST + 1):
        v = ws.cell(r_, 4).value
        if isinstance(v, str) and v.startswith(WRAPPER_PREFIX):
            wrapped += 1
    r["wrapped_count"] = wrapped
    r["acuity_wrapper_ok"] = wrapped == (ACUITY_ROW_LAST - ACUITY_ROW_FIRST + 1)

    # K45 styling
    ws2 = wb[ANALYTICS_SHEET]
    note_cell = ws2.cell(NOTE_ROW, NOTE_COL)
    r["note_bold"] = bool(note_cell.font.bold) if note_cell.font else False
    fill = note_cell.fill
    fg = fill.fgColor.value if fill and fill.fgColor and hasattr(fill.fgColor, 'value') else None
    r["note_fill"] = fg
    r["note_styled_ok"] = (r["note_bold"] is True) and (str(fg).upper().endswith("FFF2CC"))

    # K45 formula still contains the conditional note text (unchanged)
    note_formula = str(note_cell.value or "")
    r["note_formula_intact"] = "Property has no AL acuity data" in note_formula

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0115(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    nA = wrap_acuity_formulas(wb)
    print(f"  A: wrapped {nA} acuity-data cells with blank-when-empty guard (D59:D66)")
    nB = style_v5_note(wb)
    print(f"  B: applied bold + yellow fill to V5 conditional note (T12 Analytics!K45)")

    stamp_versions(wb)
    print(f"  C: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:<22}     : {r['cover_b8_ok']}")
    print(f"  All 13 AZ4 = {SUBSTRATE_TO}              : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  Acuity rows wrapped (D59:D66)         : {r['acuity_wrapper_ok']} ({r['wrapped_count']}/8)")
    print(f"  K45 note bold                          : {r['note_bold']}")
    print(f"  K45 note fill                          : {r['note_fill']}")
    print(f"  K45 note styled correctly              : {r['note_styled_ok']}")
    print(f"  K45 note formula intact                : {r['note_formula_intact']}")

    all_ok = all([
        r["cover_b8_ok"], r["az4_all"], r["acuity_wrapper_ok"],
        r["note_styled_ok"], r["note_formula_intact"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0115.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
