"""
migrate_to_v0213.py — Substrate template v0.2.12 → v0.2.13

Opens AND closes UW-BACKLOG **BL-0025** (move-out & preleased exposure
surface).

THE GAP:
  Move-out Date is captured per-resident at Rent Roll Input!W (substrate
  v0.1.10 / RR v1.16.0) but no Analyzer formula reads it. "Vacant w/
  Prelease" units (Homestead-style) were collapsing into plain `Vacant`,
  inflating exposure on the populated Analyzer's Rent Roll Recon
  section A counts. No section surfaces an underwriting "exposure"
  view (gross / net of preleased, forward NTV departure timeline).

THE FIX (3 surface changes + parser additions in RR v1.18.0):

  1. Extend Rent Roll Input!E7:E606 data validation list:
       "Occupied,Vacant,Notice,Eviction"
       → "Occupied,Vacant,Notice,Eviction,Preleased"

  2. Add Rent Roll Input!AI column header at AI4 = "Preleased Date".
     AH=Total Ancillary $ formula (substrate v0.1.13) is preserved; the
     new column sits at the next free col. analyzer_rr_writer
     (v1.18.0) routes Condensed_RR's `Preleased Date` field here.

  3. Append a new Section N to Rent Roll Recon at rows 178-198. Pure
     append (no insert_rows) — avoids the BL-0001 qualified-range-
     endpoint shift trap on cross-sheet ranges. Two sub-sections:

       N1 — Point-in-time exposure  (as of selected period)
         r182 Occupied               (mirrors r8)
         r183 On notice              (mirrors r10)
         r184 Vacant                 (mirrors r9 — now plain Vacant only)
         r185 Preleased              NEW: COUNTIFS Status="Preleased"
         r186 Total beds             (mirrors r7)
         r187 Gross exposure         =Notice + Vacant
         r188 Net exposure           =Gross − Preleased
         r189 Net exposure %         =Net / Total beds

       N2 — Forward NTV departures by move-out date
         r193 ≤30 days               COUNTIFS Status=Notice, W in [B2, B2+30]
         r194 31-60 days             COUNTIFS Status=Notice, W in (B2+30, B2+60]
         r195 61-90 days             COUNTIFS Status=Notice, W in (B2+60, B2+90]
         r196 91+ days               COUNTIFS Status=Notice, W > B2+90
         r197 No date / past         =Total Notice − sum(B193:B196)
         r198 Total Notice           =SUM(B193:B197); sanity = B183

  Status="Preleased" is the new bed status added in RR v1.18.0
  (mappings.py: `(r"\\bprelease", "Preleased")` ordered before
  `\\bvacant\\b` so "Vacant w/ Prelease" hits Preleased first).

Idempotency:
  - Gate: Cover!B8 == "v0.2.13" AND Rent Roll Recon!A178 starts with "N".
  - DV extension: only rewrites if "Preleased" not already in formula1.
  - AI4 header: only writes if currently blank.
  - Section N append: only writes if A178 is blank (no row inserts ever).

BUNDLED FILE STATUS:
  Bundled `ALF_Financial_Analyzer_Only.xlsx` (currently v0.2.12) is
  updated in place by running this migration on it.

Usage:
    python tools/migration/migrate_to_v0213.py input.xlsx output.xlsx
"""
from __future__ import annotations

import copy
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

SUBSTRATE_FROM = "v0.2.12"
SUBSTRATE_TO = "v0.2.13"

# 16-sheet anchor list — unchanged since v0.2.10 (no sheet adds/removes).
ANCHOR_SHEETS = (
    "Cover", "Dashboard",
    "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending",
    "AR & Collections",
    "UW Output", "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# -- DV extension on Rent Roll Input!E -----------------------------------
RR_INPUT_SHEET = "Rent Roll Input"
STATUS_DV_RANGE = "E7:E606"
STATUS_DV_OLD_LIST = "Occupied,Vacant,Notice,Eviction"
STATUS_DV_NEW_LIST = "Occupied,Vacant,Notice,Eviction,Preleased"

# -- New AI4 header on Rent Roll Input -----------------------------------
PRELEASED_HEADER_CELL = "AI4"
PRELEASED_HEADER_TEXT = "Preleased\nDate"
PRELEASED_HEADER_STYLE_SOURCE = "Q4"  # Move-in Date — same shape (date col)

# -- Rent Roll Recon Section N -------------------------------------------
RR_RECON_SHEET = "Rent Roll Recon"
SECTION_HEADER_ROW = 178
N1_SUBHEAD_ROW = 180
N1_COL_HEADER_ROW = 181
N1_OCCUPIED_ROW = 182
N1_NOTICE_ROW = 183
N1_VACANT_ROW = 184
N1_PRELEASED_ROW = 185
N1_TOTAL_BEDS_ROW = 186
N1_GROSS_EXPOSURE_ROW = 187
N1_NET_EXPOSURE_ROW = 188
N1_NET_EXPOSURE_PCT_ROW = 189
N2_SUBHEAD_ROW = 191
N2_COL_HEADER_ROW = 192
N2_30D_ROW = 193
N2_60D_ROW = 194
N2_90D_ROW = 195
N2_91PLUS_ROW = 196
N2_NO_DATE_ROW = 197
N2_TOTAL_NOTICE_ROW = 198

# Cross-sheet ref fragments — qualified so the BL-0001 endpoint shift trap
# doesn't apply (no insert_rows here anyway, but keeps formulas portable).
RR_S = "'Rent Roll Input'!$S$7:$S$606"   # Period Date column
RR_D = "'Rent Roll Input'!$D$7:$D$606"   # Care Type
RR_E = "'Rent Roll Input'!$E$7:$E$606"   # Status
RR_A = "'Rent Roll Input'!$A$7:$A$606"   # Unit # (used for "has-a-unit" filter)
RR_W = "'Rent Roll Input'!$W$7:$W$606"   # Move-out Date
PERIOD = "'Rent Roll Recon'!$B$2"        # Selected-period dropdown


def _care_filter(care: str) -> str:
    """Return the COUNTIFS arg-pair for filtering on a care type."""
    return f'{RR_D},"{care}"'


def _occupied_formula(care: str) -> str:
    return (
        f'=COUNTIFS({RR_S},{PERIOD},{_care_filter(care)},'
        f'{RR_E},"Occupied")'
    )


def _notice_formula(care: str) -> str:
    return (
        f'=COUNTIFS({RR_S},{PERIOD},{_care_filter(care)},'
        f'{RR_E},"Notice")'
    )


def _vacant_formula(care: str) -> str:
    return (
        f'=COUNTIFS({RR_S},{PERIOD},{_care_filter(care)},'
        f'{RR_E},"Vacant")'
    )


def _preleased_formula(care: str) -> str:
    return (
        f'=COUNTIFS({RR_S},{PERIOD},{_care_filter(care)},'
        f'{RR_E},"Preleased")'
    )


def _total_beds_formula(care: str) -> str:
    # Mirrors Rent Roll Recon!B7: count units with a populated Unit # for
    # the selected period + care type.
    return (
        f'=COUNTIFS({RR_S},{PERIOD},{RR_A},"<>",'
        f'{_care_filter(care)})'
    )


def _ntv_window_formula(care: str, lo_op: str, lo_off: int,
                       hi_op: str | None = None, hi_off: int | None = None) -> str:
    """COUNTIFS NTV rows with W in (or beyond) a window relative to period."""
    parts = [
        f'COUNTIFS({RR_S},{PERIOD},{_care_filter(care)},'
        f'{RR_E},"Notice",'
        f'{RR_W},"{lo_op}"&({PERIOD}+{lo_off})'
    ]
    if hi_op is not None:
        parts.append(f',{RR_W},"{hi_op}"&({PERIOD}+{hi_off})')
    return "=" + "".join(parts) + ")"


CARES = ("IL", "AL", "MC")
CARE_COLS = {"IL": "B", "AL": "C", "MC": "D"}  # B/C/D as in existing sections


def is_already_v0213(wb) -> bool:
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    ws = wb[RR_RECON_SHEET]
    a178 = ws.cell(SECTION_HEADER_ROW, 1).value
    if not isinstance(a178, str) or not a178.startswith("N"):
        return False
    return True


# ----------------------------------------------------------------------
# Step 1: Extend Status DV on Rent Roll Input!E7:E606
# ----------------------------------------------------------------------
def extend_status_dv(wb) -> dict:
    ws = wb[RR_INPUT_SHEET]
    counts = {"patched": 0, "skipped_already": 0, "missing": 0}
    log = []

    # Find the DV that covers E7:E606
    target_dv = None
    for dv in ws.data_validations.dataValidation:
        ranges = [str(r) for r in dv.sqref.ranges] if dv.sqref else []
        if STATUS_DV_RANGE in ranges and dv.type == "list":
            target_dv = dv
            break

    if target_dv is None:
        counts["missing"] = 1
        log.append(f"  MISSING: no list DV on {STATUS_DV_RANGE}")
        return {"counts": counts, "log": log}

    cur = target_dv.formula1 or ""
    # formula1 in openpyxl includes the outer quotes for inline lists, e.g.
    #   '"Occupied,Vacant,Notice,Eviction"'
    if "Preleased" in cur:
        counts["skipped_already"] += 1
        log.append(f"  SKIP DV: 'Preleased' already in formula1 = {cur!r}")
        return {"counts": counts, "log": log}

    # Build new list — same structure with the outer quotes preserved.
    target_dv.formula1 = f'"{STATUS_DV_NEW_LIST}"'
    counts["patched"] += 1
    log.append(f"  PATCH DV {STATUS_DV_RANGE}: {cur!r} -> {target_dv.formula1!r}")
    return {"counts": counts, "log": log}


# ----------------------------------------------------------------------
# Step 2: Add Rent Roll Input!AI4 header
# ----------------------------------------------------------------------
def add_preleased_header(wb) -> dict:
    ws = wb[RR_INPUT_SHEET]
    counts = {"patched": 0, "skipped": 0}
    log = []

    cell = ws[PRELEASED_HEADER_CELL]
    if cell.value is not None:
        counts["skipped"] += 1
        log.append(f"  SKIP {PRELEASED_HEADER_CELL}: already populated ({cell.value!r})")
        return {"counts": counts, "log": log}

    cell.value = PRELEASED_HEADER_TEXT

    # Copy style from Q4 (Move-in Date — same role: date-typed input column).
    src = ws[PRELEASED_HEADER_STYLE_SOURCE]
    if src.has_style:
        cell.font = copy.copy(src.font)
        cell.fill = copy.copy(src.fill)
        cell.alignment = copy.copy(src.alignment)
        cell.border = copy.copy(src.border)
        cell.number_format = src.number_format
    # Also copy column width from Q
    src_col_letter = PRELEASED_HEADER_STYLE_SOURCE[0]  # 'Q'
    if src_col_letter in ws.column_dimensions:
        from openpyxl.utils import get_column_letter
        ai_letter = "".join(c for c in PRELEASED_HEADER_CELL if c.isalpha())
        ws.column_dimensions[ai_letter].width = ws.column_dimensions[src_col_letter].width

    counts["patched"] += 1
    log.append(f"  PATCH {PRELEASED_HEADER_CELL}: wrote {PRELEASED_HEADER_TEXT!r} with Q4 style")
    return {"counts": counts, "log": log}


# ----------------------------------------------------------------------
# Step 3: Append Section N to Rent Roll Recon
# ----------------------------------------------------------------------
def _apply_style_from(target_cell, source_cell):
    if not source_cell.has_style:
        return
    target_cell.font = copy.copy(source_cell.font)
    target_cell.fill = copy.copy(source_cell.fill)
    target_cell.alignment = copy.copy(source_cell.alignment)
    target_cell.border = copy.copy(source_cell.border)
    target_cell.number_format = source_cell.number_format


def append_section_n(wb) -> dict:
    ws = wb[RR_RECON_SHEET]
    counts = {"section_n_written": 0, "skipped_already": 0}
    log = []

    # Idempotency gate — if A178 already has content, skip.
    if ws.cell(SECTION_HEADER_ROW, 1).value:
        counts["skipped_already"] += 1
        log.append(f"  SKIP Section N: A{SECTION_HEADER_ROW} already populated")
        return {"counts": counts, "log": log}

    # Style sources — borrow from existing peers.
    # A6 / A15 = section header (large dark fill). A6 is merged A6:H6.
    style_section_header = ws["A6"]
    # A121 / A133 etc are subheads ("M1 — ...").
    style_subhead = ws["A121"]
    # Row 4 = column headers (Metric / IL / AL / MC / Total / Flag).
    style_col_header_label = ws["A4"]
    style_col_header_care = ws["B4"]
    style_col_header_total = ws["E4"]
    style_col_header_flag = ws["H4"]
    # Row 7-10 = regular metric rows (label, IL/AL/MC formulas, total, flag).
    style_metric_label = ws["A7"]
    style_metric_num = ws["B7"]
    style_metric_total = ws["E7"]
    style_metric_flag = ws["H7"]
    # Row 12 has a % formula — use as % style template.
    style_metric_pct = ws["B12"]

    # ----- Section N header (row 178) — merged A:H -----------------------
    sh = ws.cell(SECTION_HEADER_ROW, 1)
    sh.value = "N  ·  EXPOSURE  (forward-looking move-out risk)"
    _apply_style_from(sh, style_section_header)
    ws.merge_cells(start_row=SECTION_HEADER_ROW, start_column=1,
                   end_row=SECTION_HEADER_ROW, end_column=8)

    # ----- N1 subhead (row 180) — merged A:H -----------------------------
    sub1 = ws.cell(N1_SUBHEAD_ROW, 1)
    sub1.value = "N1  —  Point-in-time exposure  (as of selected period)"
    _apply_style_from(sub1, style_subhead)
    ws.merge_cells(start_row=N1_SUBHEAD_ROW, start_column=1,
                   end_row=N1_SUBHEAD_ROW, end_column=8)

    # ----- N1 column header row (181) ------------------------------------
    headers = [
        (1, "Metric", style_col_header_label),
        (2, "IL", style_col_header_care),
        (3, "AL", style_col_header_care),
        (4, "MC", style_col_header_care),
        (5, "Total", style_col_header_total),
        (8, "Flag / Note", style_col_header_flag),
    ]
    for col, val, src in headers:
        c = ws.cell(N1_COL_HEADER_ROW, col)
        c.value = val
        _apply_style_from(c, src)

    # ----- N1 data rows --------------------------------------------------
    # Schema for each row: (row, label, formula_factory, flag, total_formula_override)
    # formula_factory takes a care-type code ('IL'/'AL'/'MC') and returns
    # the formula text. Total col E uses =SUM(B:D) unless overridden.
    n1_rows = [
        (N1_OCCUPIED_ROW,    "Occupied",
         _occupied_formula, "", None),
        (N1_NOTICE_ROW,      "On notice",
         _notice_formula,   "", None),
        (N1_VACANT_ROW,      "Vacant",
         _vacant_formula,   "Excl. Vacant w/ Prelease (counted on row " +
                            str(N1_PRELEASED_ROW) + ")", None),
        (N1_PRELEASED_ROW,   "Preleased",
         _preleased_formula, "Vacant w/ signed prelease — offsets gross exposure", None),
        (N1_TOTAL_BEDS_ROW,  "Total beds",
         _total_beds_formula, "", None),
        (N1_GROSS_EXPOSURE_ROW, "Gross exposure (Notice + Vacant)",
         None, "Includes only currently-empty + leaving units", None),
        (N1_NET_EXPOSURE_ROW, "Net exposure (Notice + Vacant − Preleased)",
         None, "Subtracts preleased units already lined up to fill", None),
        (N1_NET_EXPOSURE_PCT_ROW, "Net exposure %",
         None, "Net exposure ÷ Total beds", None),
    ]

    for r, label, factory, flag, total_override in n1_rows:
        # Col A label
        a = ws.cell(r, 1)
        a.value = label
        _apply_style_from(a, style_metric_label)

        # Cols B-D
        if factory is not None:
            for care in CARES:
                col = ord(CARE_COLS[care]) - ord("A") + 1
                c = ws.cell(r, col)
                c.value = factory(care)
                _apply_style_from(c, style_metric_num)
            # Col E — Total = SUM(B:D)
            e = ws.cell(r, 5)
            e.value = f"=SUM(B{r}:D{r})"
            _apply_style_from(e, style_metric_total)

        # Col H — flag note
        if flag:
            h = ws.cell(r, 8)
            h.value = flag
            _apply_style_from(h, style_metric_flag)

    # ----- N1 derived rows: gross, net, pct ------------------------------
    for care in CARES:
        col = ord(CARE_COLS[care]) - ord("A") + 1
        col_letter = CARE_COLS[care]
        # Gross exposure = Notice + Vacant
        c = ws.cell(N1_GROSS_EXPOSURE_ROW, col)
        c.value = (
            f"={col_letter}{N1_NOTICE_ROW}+{col_letter}{N1_VACANT_ROW}"
        )
        _apply_style_from(c, style_metric_num)

        # Net exposure = Gross − Preleased
        c = ws.cell(N1_NET_EXPOSURE_ROW, col)
        c.value = (
            f"={col_letter}{N1_GROSS_EXPOSURE_ROW}-{col_letter}{N1_PRELEASED_ROW}"
        )
        _apply_style_from(c, style_metric_num)

        # Net exposure %
        c = ws.cell(N1_NET_EXPOSURE_PCT_ROW, col)
        c.value = (
            f'=IFERROR({col_letter}{N1_NET_EXPOSURE_ROW}/'
            f'{col_letter}{N1_TOTAL_BEDS_ROW},"-")'
        )
        _apply_style_from(c, style_metric_pct)
        c.number_format = "0.0%"

    # Col E totals for derived rows
    e_gross = ws.cell(N1_GROSS_EXPOSURE_ROW, 5)
    e_gross.value = f"=SUM(B{N1_GROSS_EXPOSURE_ROW}:D{N1_GROSS_EXPOSURE_ROW})"
    _apply_style_from(e_gross, style_metric_total)

    e_net = ws.cell(N1_NET_EXPOSURE_ROW, 5)
    e_net.value = f"=SUM(B{N1_NET_EXPOSURE_ROW}:D{N1_NET_EXPOSURE_ROW})"
    _apply_style_from(e_net, style_metric_total)

    e_pct = ws.cell(N1_NET_EXPOSURE_PCT_ROW, 5)
    e_pct.value = (
        f'=IFERROR(E{N1_NET_EXPOSURE_ROW}/E{N1_TOTAL_BEDS_ROW},"-")'
    )
    _apply_style_from(e_pct, style_metric_total)
    e_pct.number_format = "0.0%"

    # ----- N2 subhead (row 191) — merged A:H -----------------------------
    sub2 = ws.cell(N2_SUBHEAD_ROW, 1)
    sub2.value = "N2  —  Forward NTV departures by move-out date  (Notice rows only, relative to period)"
    _apply_style_from(sub2, style_subhead)
    ws.merge_cells(start_row=N2_SUBHEAD_ROW, start_column=1,
                   end_row=N2_SUBHEAD_ROW, end_column=8)

    # ----- N2 column header (row 192) ------------------------------------
    n2_headers = [
        (1, "Window", style_col_header_label),
        (2, "IL", style_col_header_care),
        (3, "AL", style_col_header_care),
        (4, "MC", style_col_header_care),
        (5, "Total", style_col_header_total),
        (8, "Flag / Note", style_col_header_flag),
    ]
    for col, val, src in n2_headers:
        c = ws.cell(N2_COL_HEADER_ROW, col)
        c.value = val
        _apply_style_from(c, src)

    # ----- N2 bucket rows ------------------------------------------------
    # (row, label, lo_op, lo_off, hi_op, hi_off, flag)
    # lo_op = comparison vs (B2 + lo_off). hi_op may be None for open-ended top.
    n2_bucket_rows = [
        (N2_30D_ROW,   "≤ 30 days",   ">=", 0,  "<=", 30,  ""),
        (N2_60D_ROW,   "31–60 days",  ">",  30, "<=", 60,  ""),
        (N2_90D_ROW,   "61–90 days",  ">",  60, "<=", 90,  ""),
        (N2_91PLUS_ROW, "91+ days",   ">",  90, None, None, "Beyond 90-day window"),
    ]

    for r, label, lo_op, lo_off, hi_op, hi_off, flag in n2_bucket_rows:
        a = ws.cell(r, 1)
        a.value = label
        _apply_style_from(a, style_metric_label)

        for care in CARES:
            col = ord(CARE_COLS[care]) - ord("A") + 1
            c = ws.cell(r, col)
            c.value = _ntv_window_formula(care, lo_op, lo_off, hi_op, hi_off)
            _apply_style_from(c, style_metric_num)

        e = ws.cell(r, 5)
        e.value = f"=SUM(B{r}:D{r})"
        _apply_style_from(e, style_metric_total)

        if flag:
            h = ws.cell(r, 8)
            h.value = flag
            _apply_style_from(h, style_metric_flag)

    # ----- N2 "No move-out date set" row (residual) ----------------------
    a = ws.cell(N2_NO_DATE_ROW, 1)
    a.value = "No move-out date set / past"
    _apply_style_from(a, style_metric_label)
    for care in CARES:
        col_letter = CARE_COLS[care]
        col_idx = ord(col_letter) - ord("A") + 1
        c = ws.cell(N2_NO_DATE_ROW, col_idx)
        c.value = (
            f"={col_letter}{N1_NOTICE_ROW}-"
            f"({col_letter}{N2_30D_ROW}+{col_letter}{N2_60D_ROW}+"
            f"{col_letter}{N2_90D_ROW}+{col_letter}{N2_91PLUS_ROW})"
        )
        _apply_style_from(c, style_metric_num)
    e = ws.cell(N2_NO_DATE_ROW, 5)
    e.value = f"=SUM(B{N2_NO_DATE_ROW}:D{N2_NO_DATE_ROW})"
    _apply_style_from(e, style_metric_total)
    h = ws.cell(N2_NO_DATE_ROW, 8)
    h.value = "NTV without a scheduled move-out date — unknown timing"
    _apply_style_from(h, style_metric_flag)

    # ----- N2 Total Notice row (sanity) ----------------------------------
    a = ws.cell(N2_TOTAL_NOTICE_ROW, 1)
    a.value = "Total Notice  (sanity = N1 On notice)"
    _apply_style_from(a, style_metric_label)
    for care in CARES:
        col_letter = CARE_COLS[care]
        col_idx = ord(col_letter) - ord("A") + 1
        c = ws.cell(N2_TOTAL_NOTICE_ROW, col_idx)
        c.value = f"=SUM({col_letter}{N2_30D_ROW}:{col_letter}{N2_NO_DATE_ROW})"
        _apply_style_from(c, style_metric_num)
    e = ws.cell(N2_TOTAL_NOTICE_ROW, 5)
    e.value = f"=SUM(B{N2_TOTAL_NOTICE_ROW}:D{N2_TOTAL_NOTICE_ROW})"
    _apply_style_from(e, style_metric_total)

    counts["section_n_written"] = 1
    log.append(f"  PATCH Section N: rows {SECTION_HEADER_ROW}-{N2_TOTAL_NOTICE_ROW} written")
    return {"counts": counts, "log": log}


# ----------------------------------------------------------------------
# Step 4: Version stamps
# ----------------------------------------------------------------------
def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ----------------------------------------------------------------------
# Verify
# ----------------------------------------------------------------------
def verify_migration(wb) -> dict:
    r = {}

    # Cover stamp
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    # DV extension
    ws_in = wb[RR_INPUT_SHEET]
    dv_ok = False
    dv_formula = None
    for dv in ws_in.data_validations.dataValidation:
        ranges = [str(rng) for rng in dv.sqref.ranges] if dv.sqref else []
        if STATUS_DV_RANGE in ranges and dv.type == "list":
            dv_formula = dv.formula1
            if "Preleased" in (dv_formula or ""):
                dv_ok = True
            break
    r["dv_formula"] = dv_formula
    r["dv_ok"] = dv_ok

    # AI4 header
    r["ai4"] = ws_in[PRELEASED_HEADER_CELL].value
    r["ai4_ok"] = r["ai4"] == PRELEASED_HEADER_TEXT

    # Section N — spot-check 6 cells
    ws_re = wb[RR_RECON_SHEET]
    spot_checks = [
        (SECTION_HEADER_ROW, 1, "N  ·  EXPOSURE", "section header"),
        (N1_SUBHEAD_ROW, 1, "N1", "N1 subhead"),
        (N1_PRELEASED_ROW, 1, "Preleased", "N1 Preleased label"),
        (N1_NET_EXPOSURE_ROW, 1, "Net exposure", "N1 Net exposure label"),
        (N2_SUBHEAD_ROW, 1, "N2", "N2 subhead"),
        (N2_30D_ROW, 1, "30 days", "N2 ≤30d row"),
    ]
    spot_results = []
    spot_all_ok = True
    for row, col, must_contain, label in spot_checks:
        v = ws_re.cell(row, col).value
        ok = isinstance(v, str) and must_contain in v
        spot_results.append((row, col, label, ok, v))
        spot_all_ok = spot_all_ok and ok
    r["section_n_spot_checks"] = spot_results
    r["section_n_spot_ok"] = spot_all_ok

    # Section N — formula presence on row 185 (Preleased) and row 188 (Net)
    r["preleased_il_formula"] = ws_re.cell(N1_PRELEASED_ROW, 2).value
    r["preleased_formula_ok"] = (
        isinstance(r["preleased_il_formula"], str)
        and "\"Preleased\"" in r["preleased_il_formula"]
    )
    r["net_il_formula"] = ws_re.cell(N1_NET_EXPOSURE_ROW, 2).value
    r["net_formula_ok"] = (
        isinstance(r["net_il_formula"], str)
        and "-B" + str(N1_PRELEASED_ROW) in r["net_il_formula"]
    )

    # Anchors
    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # Sheet count unchanged at 16
    r["sheet_count"] = len(wb.sheetnames)
    r["sheet_count_ok"] = r["sheet_count"] == 16

    return r


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0213(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    cur_version = wb["Cover"]["B8"].value
    if cur_version != SUBSTRATE_FROM:
        print(
            f"  WARN: Cover!B8 = {cur_version!r}, expected {SUBSTRATE_FROM!r}. "
            f"Proceeding anyway."
        )

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    dv_result = extend_status_dv(wb)
    for line in dv_result["log"]:
        print(line)

    hdr_result = add_preleased_header(wb)
    for line in hdr_result["log"]:
        print(line)

    sec_result = append_section_n(wb)
    for line in sec_result["log"]:
        print(line)

    stamp_versions(wb)
    print(
        f"  Stamped substrate version -> {SUBSTRATE_TO} on "
        f"Cover!B8 + {len(ANCHOR_SHEETS)} AZ4 anchors"
    )

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    checks = [
        ("Cover!B8 = " + repr(r["cover_b8"]), r["cover_b8_ok"]),
        ("DV E7:E606 contains 'Preleased' (formula1=" + repr(r["dv_formula"]) + ")", r["dv_ok"]),
        ("AI4 = " + repr(r["ai4"]), r["ai4_ok"]),
        (f"Section N spot-checks ({len(r['section_n_spot_checks'])} cells)", r["section_n_spot_ok"]),
        ("N1!B185 formula references Status=\"Preleased\"", r["preleased_formula_ok"]),
        ("N1!B188 net formula references Preleased row", r["net_formula_ok"]),
        (f"Sheet count = {r['sheet_count']} (expected 16)", r["sheet_count_ok"]),
        (f"All {r['az4_count']} AZ4 anchors = {SUBSTRATE_TO}", r["az4_all"]),
    ]
    for desc, ok in checks:
        flag = "[OK]  " if ok else "[FAIL]"
        print(f"  {flag}  {desc}")

    if not r["section_n_spot_ok"]:
        print()
        print("  Section N spot-check detail:")
        for row, col, label, ok, val in r["section_n_spot_checks"]:
            mark = "OK" if ok else "FAIL"
            print(f"    {mark}  r{row}c{col} {label}: {val!r}")

    all_ok = all(ok for _, ok in checks)

    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0213.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
