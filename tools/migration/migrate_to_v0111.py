"""
migrate_to_v0111.py - Substrate template v0.1.10 -> v0.1.11

Patch fix for the chart category-axis positioning bug introduced when the
five T12 Analytics charts were added at substrate v0.1.8 (Branch 3
analytical coverage). All three axis-bearing charts on T12 Analytics
have `<catAx><axPos val="l"/></catAx>` (category axis claiming the LEFT
position) when they should have `val="b"` (BOTTOM) for vertical column
and line charts.

The bug visibly manifests on the LineChart (V4 - T12 Revenue Trend):
when openpyxl writes out the chart XML, Excel can't reconcile two axes
both claiming the left position, falls back to rendering the categories
(month labels) as legend entries with no plot area. V1 and V2 (BarCharts
with barDir="col") are also technically buggy but Excel tolerates them
visually -- this fix brings them to spec-compliance and avoids stricter
future Excel versions silently breaking them too.

Operations:

  A. For each chart on T12 Analytics with a category axis at "l", flip
     it to "b". V3 and V5 are doughnut charts with no axes - skipped.
     Expected: 3 charts fixed (V1 BarChart, V2 BarChart, V4 LineChart).
     Value axes are NOT touched - they're correctly at "l" already.
  B. Stamp Cover!B8 + 13 AZ4 anchors to v0.1.11
  C. 6-check verification (3 charts confirmed + Cover stamp + AZ4 stamps
     + spot-check that valAx didn't move)

Idempotent: gate checks BOTH version stamp AND that V4's catAx already
reads "b", so re-running on a partial-state file safely re-applies.

Usage:
    python tools/migration/migrate_to_v0111.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

SUBSTRATE_FROM = "v0.1.10"
SUBSTRATE_TO = "v0.1.11"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

ANALYTICS_SHEET = "T12 Analytics"

# Chart titles that should have their category axis flipped to "b".
# V3 (Payer Mix doughnut) and V5 (AL Acuity Mix doughnut) are excluded -
# doughnut charts have no x/y axes.
TARGET_CHART_TITLE_FRAGMENTS = (
    "Occupancy by Care Type",        # V1 BarChart, barDir=col
    "Rate Dispersion by Care Type",  # V2 BarChart, barDir=col
    "T12 Revenue Trend",             # V4 LineChart
)


def _chart_title(chart) -> str:
    """Best-effort extract of a chart's title text."""
    try:
        if chart.title is None or chart.title.tx is None or chart.title.tx.rich is None:
            return ""
        out = ""
        for p in chart.title.tx.rich.paragraphs:
            for r in p.r or []:
                out += (r.t or "")
        return out
    except Exception:
        return ""


def _is_target_chart(chart) -> bool:
    title = _chart_title(chart)
    return any(frag in title for frag in TARGET_CHART_TITLE_FRAGMENTS)


def is_already_v0111(wb) -> bool:
    """Gate: version stamp AND V4 category axis already at 'b'."""
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    if ANALYTICS_SHEET not in wb.sheetnames:
        return False
    ws = wb[ANALYTICS_SHEET]
    for ch in ws._charts:
        if "T12 Revenue Trend" in _chart_title(ch):
            x_axis = getattr(ch, "x_axis", None)
            return x_axis is not None and getattr(x_axis, "axPos", None) == "b"
    return False


def fix_chart_axes(wb) -> dict:
    """A. Flip catAx axPos 'l' -> 'b' on V1/V2/V4. Leave valAx alone.

    Returns a dict mapping chart title -> action ('fixed' | 'already_ok' |
    'skipped:no_x_axis' | 'unexpected_pos:<value>').
    """
    ws = wb[ANALYTICS_SHEET]
    actions: dict[str, str] = {}
    for ch in ws._charts:
        title = _chart_title(ch)
        if not _is_target_chart(ch):
            continue
        x_axis = getattr(ch, "x_axis", None)
        if x_axis is None:
            actions[title] = "skipped:no_x_axis"
            continue
        cur = getattr(x_axis, "axPos", None)
        if cur == "b":
            actions[title] = "already_ok"
        elif cur == "l":
            x_axis.axPos = "b"
            actions[title] = "fixed"
        else:
            actions[title] = f"unexpected_pos:{cur!r}"
    return actions


def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


def verify_migration(wb) -> dict:
    r: dict = {}

    # 1. Cover!B8 stamp
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    # 2. AZ4 stamps
    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # 3. Three target charts at 'b'; valAx unmoved at 'l'
    ws = wb[ANALYTICS_SHEET]
    fixed_count = 0
    valax_ok = True
    for ch in ws._charts:
        if not _is_target_chart(ch):
            continue
        x_axis = getattr(ch, "x_axis", None)
        y_axis = getattr(ch, "y_axis", None)
        if x_axis is not None and getattr(x_axis, "axPos", None) == "b":
            fixed_count += 1
        if y_axis is not None and getattr(y_axis, "axPos", None) != "l":
            valax_ok = False
    r["fixed_count"] = fixed_count
    r["fixed_all_three"] = fixed_count == 3
    r["valax_intact"] = valax_ok

    # 4. Doughnut charts (V3, V5) still axis-less - sanity
    doughnut_titles_seen = []
    for ch in ws._charts:
        title = _chart_title(ch)
        if "Payer Mix" in title or "AL Acuity Mix" in title:
            doughnut_titles_seen.append(title)
    r["doughnuts_seen"] = len(doughnut_titles_seen)

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0111(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    actions = fix_chart_axes(wb)
    print(f"  A: chart axis fixes:")
    for title, action in actions.items():
        print(f"     - {title!r}: {action}")

    stamp_versions(wb)
    print(f"  B: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:<24s}    : {r['cover_b8_ok']}")
    print(f"  All 13 AZ4 = {SUBSTRATE_TO}              : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  3 target charts fixed (catAx='b')     : {r['fixed_all_three']} ({r['fixed_count']}/3)")
    print(f"  valAx unmoved at 'l'                  : {r['valax_intact']}")
    print(f"  Doughnut charts still present         : {r['doughnuts_seen']}/2")

    all_ok = (
        r["cover_b8_ok"] and r["az4_all"]
        and r["fixed_all_three"]
        and r["valax_intact"]
        and r["doughnuts_seen"] == 2
    )
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0111.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
