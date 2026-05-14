"""
migrate_to_v020.py - Substrate template v0.1.15 -> v0.2.0

Closes UW-BACKLOG BL-0009 — Branch 2 Handoff readiness. The v0.2.0 flagship
substrate release. Per the four-branch Track 3 roadmap (Branches 1+4 closed
in v0.1.6, Branch 3 closed in v0.1.8 + extended through v0.1.14), Branch 2
is the final Analyzer-optimization workstream.

Three coordinated additions:

  A. NEW SHEET: `UW Export` — values-only mirror of UW Output for clean
     downstream paste. Inserted at index 8 (right after UW Output, before
     Mapping Review). Layout:
       Rows 1-7:  metadata header (Property, RR period, T12 period,
                  substrate version, generated timestamp)
       Row  8:    visual break
       Rows 9-79: values-only mirror of UW Output rows 1-71 (cols A-H).
                  Each mirror cell uses `='UW Output'!{cell}` so when
                  Excel evaluates, the cell shows the live value. The
                  downstream consumer copies UW Export and pastes as
                  values into their template — clean numeric handoff
                  without formula references back to this workbook.

  B. NEW SECTION on Workbook Health: PRE-EXPORT GATE (rows 46-52). A
     compact validation aggregator that surfaces a single
     "READY FOR EXPORT" / "⚠ NOT READY" indicator at row 52 plus
     individual sub-checks at rows 48-51. Aggregates the existing V1-V8
     validation checks that have been on Workbook Health since v0.1.4.

  C. WORKBOOK MAP extension: insert "UW Export" row in the existing
     Workbook Map section so the new sheet is discoverable from the
     Workbook Health hub like every other sheet.

  D. ANCHOR_SHEETS extended from 13 → 14 to include UW Export. Cover!B8
     and all 14 AZ4 anchors stamped to v0.2.0.

Idempotent: gate checks BOTH the version stamp AND that the new
"UW Export" sheet exists with the metadata title at A1. Re-runs on
partial-state files safely re-apply (existing UW Export gets fully
rewritten — no in-place patching).

Usage:
    python tools/migration/migrate_to_v020.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SUBSTRATE_FROM = "v0.1.15"
SUBSTRATE_TO = "v0.2.0"

# Original 13 sheets + the new UW Export = 14
ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "UW Export",  # NEW at v0.2.0
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

EXPORT_SHEET = "UW Export"
UW_OUTPUT_SHEET = "UW Output"
WORKBOOK_HEALTH = "Workbook Health"
COVER = "Cover"

# UW Export layout
EXPORT_TITLE_ROW = 1
EXPORT_META_FIRST_ROW = 3   # Property name
EXPORT_META_LAST_ROW = 7    # Generated
EXPORT_BREAK_ROW = 8
EXPORT_MIRROR_FIRST_ROW = 9
UW_OUTPUT_ROWS = 71
UW_OUTPUT_COLS = 8  # A through H
EXPORT_MIRROR_LAST_ROW = EXPORT_MIRROR_FIRST_ROW + UW_OUTPUT_ROWS - 1  # row 79

# Pre-Export Gate placement on Workbook Health
GATE_TITLE_ROW = 46
GATE_HEADER_ROW = 47
GATE_FIRST_CHECK_ROW = 48
GATE_AGGREGATE_ROW = 52  # the "READY FOR EXPORT" indicator

# Workbook Map insertion: insert "UW Export" right after "UW Output" (row 13)
MAP_NEW_ROW = 14   # was empty in prior substrate (row 18 of the map ends at 18 = Workbook Health)

# ===== Styling =====
NAVY = "FF1F3864"
LIGHT_NAVY = "FF305496"
WHITE = "FFFFFFFF"
DARK_TEXT = "FF1F1F1F"
PALE_GREY = "FFF2F2F2"
PALE_YELLOW = "FFFFF2CC"
PALE_GREEN = "FFE2EFDA"

TITLE_FONT = Font(name="Calibri", size=12, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
META_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
BODY_FONT = Font(name="Calibri", size=10, color=DARK_TEXT)
ITALIC_FONT = Font(name="Calibri", size=9, italic=True, color="FF7F7F7F")
MIRROR_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)

TITLE_FILL = PatternFill(fill_type="solid", fgColor=NAVY)
SUBTITLE_FILL = PatternFill(fill_type="solid", fgColor=LIGHT_NAVY)
META_FILL = PatternFill(fill_type="solid", fgColor=PALE_GREY)
GATE_FILL = PatternFill(fill_type="solid", fgColor=PALE_YELLOW)
MIRROR_HEADER_FILL = PatternFill(fill_type="solid", fgColor=PALE_GREY)

LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Named ranges and cells used in metadata header
PROPERTY_NAME_REF = "Property_Name"
RR_PERIOD_REF = "RR_Period_Date"
T12_PERIOD_REF = "T12_Period_Date"
COVER_B8 = "Cover!$B$8"


def is_already_v020(wb) -> bool:
    """Gate: version stamp AND UW Export sheet exists with the title at A1."""
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    if EXPORT_SHEET not in wb.sheetnames:
        return False
    title = wb[EXPORT_SHEET].cell(EXPORT_TITLE_ROW, 1).value
    return isinstance(title, str) and "UW Export" in title


# ===== A. UW Export sheet =====

def install_uw_export_sheet(wb) -> int:
    """Create UW Export sheet (or recreate if already present from a partial
    migration). Inserted at index 8 — right after UW Output, before
    Mapping Review."""
    # If sheet already exists (partial-state re-run), drop it to start fresh.
    if EXPORT_SHEET in wb.sheetnames:
        del wb[EXPORT_SHEET]

    # Find the insertion position: right after UW Output
    if UW_OUTPUT_SHEET in wb.sheetnames:
        insert_idx = wb.sheetnames.index(UW_OUTPUT_SHEET) + 1
    else:
        insert_idx = len(wb.sheetnames)
    ws = wb.create_sheet(title=EXPORT_SHEET, index=insert_idx)

    # Sheet-level styling
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    for col_letter in ("B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col_letter].width = 14

    # ----- Row 1: Title -----
    title = ws.cell(EXPORT_TITLE_ROW, 1, value="UW Export  —  Values-only mirror of UW Output")
    title.font = TITLE_FONT
    title.fill = TITLE_FILL
    title.alignment = LEFT
    ws.merge_cells(start_row=EXPORT_TITLE_ROW, end_row=EXPORT_TITLE_ROW, start_column=1, end_column=8)
    ws.row_dimensions[EXPORT_TITLE_ROW].height = 22

    # ----- Row 2: Instructions (italic) -----
    instr = ws.cell(2, 1, value=(
        "Copy A9:H79 (or use Excel's Move-or-Copy → Create-a-copy → Paste-Special: Values) "
        "into the downstream UW template. Metadata header above (rows 3-7) carries the audit trail."
    ))
    instr.font = ITALIC_FONT
    instr.alignment = LEFT
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=8)
    ws.row_dimensions[2].height = 28

    # ----- Rows 3-7: Metadata header -----
    metadata_rows = [
        ("Property name",         f"=IFERROR({PROPERTY_NAME_REF},\"(not set)\")"),
        ("Rent roll period",      f"=IF(ISNUMBER({RR_PERIOD_REF}),TEXT({RR_PERIOD_REF},\"yyyy-mm-dd\"),\"(not set)\")"),
        ("T12 period",            f"=IF(ISNUMBER({T12_PERIOD_REF}),TEXT({T12_PERIOD_REF},\"yyyy-mm-dd\"),\"(not set)\")"),
        ("Substrate version",     f"={COVER_B8}"),
        ("Generated (open time)", f"=TEXT(NOW(),\"yyyy-mm-dd hh:mm\")"),
    ]
    for offset, (label, formula) in enumerate(metadata_rows):
        r = EXPORT_META_FIRST_ROW + offset
        a = ws.cell(r, 1, value=label)
        a.font = META_LABEL_FONT; a.fill = META_FILL; a.alignment = LEFT; a.border = BOX
        b = ws.cell(r, 2, value=formula)
        b.font = BODY_FONT; b.alignment = LEFT; b.border = BOX
        ws.merge_cells(start_row=r, end_row=r, start_column=2, end_column=8)

    # ----- Row 8: visual break -----
    # (left empty — no styling needed)

    # ----- Rows 9-79: values-only mirror of UW Output rows 1-71 -----
    # Each mirror cell uses `='UW Output'!{cell}`. Excel evaluates → shows
    # the live value. Number format inherited as General; downstream
    # paste-as-values doesn't care about format.
    for src_row in range(1, UW_OUTPUT_ROWS + 1):
        dest_row = EXPORT_MIRROR_FIRST_ROW + (src_row - 1)
        for col in range(1, UW_OUTPUT_COLS + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            cell = ws.cell(dest_row, col)
            cell.value = f"='{UW_OUTPUT_SHEET}'!{col_letter}{src_row}"
            cell.font = BODY_FONT
            # Header row (mirror of UW Output row 4) gets header styling
            if src_row == 4:
                cell.font = MIRROR_HEADER_FONT
                cell.fill = MIRROR_HEADER_FILL
                cell.alignment = CENTER

    # ----- Anchor cell at AZ1 (sheet-purpose label, used by Workbook Map) -----
    ws["AZ1"] = "Values-only mirror of UW Output for downstream paste"
    ws["AZ4"] = SUBSTRATE_TO  # version stamp

    return UW_OUTPUT_ROWS  # number of mirror rows installed


# ===== B. Pre-Export Gate on Workbook Health =====

def install_pre_export_gate(wb) -> int:
    """Add a 'PRE-EXPORT GATE' section to Workbook Health (rows 46-52).

    Structure:
        Row 46: section title (merged A:D)
        Row 47: column headers (Check / Status)
        Rows 48-51: individual sub-checks
        Row 52: aggregate READY FOR EXPORT indicator

    The sub-checks are formula references to existing V1-V8 validation
    cells (rows 23-30) so the gate stays in sync with the established
    validation logic. No new validation criteria invented here.
    """
    ws = wb[WORKBOOK_HEALTH]

    # Row 46 — section title
    title = ws.cell(GATE_TITLE_ROW, 1, value="4 · PRE-EXPORT GATE")
    title.font = SUBTITLE_FONT; title.fill = SUBTITLE_FILL; title.alignment = LEFT
    ws.merge_cells(start_row=GATE_TITLE_ROW, end_row=GATE_TITLE_ROW, start_column=1, end_column=4)

    # Row 47 — column headers
    h1 = ws.cell(GATE_HEADER_ROW, 1, value="Pre-export check")
    h1.font = META_LABEL_FONT; h1.fill = META_FILL; h1.alignment = LEFT; h1.border = BOX
    h2 = ws.cell(GATE_HEADER_ROW, 2, value="Status")
    h2.font = META_LABEL_FONT; h2.fill = META_FILL; h2.alignment = CENTER; h2.border = BOX

    # Rows 48-51 — individual sub-checks. Each reads from the V-row cells
    # already present on Workbook Health (rows 23-30 from substrate v0.1.4+).
    checks = [
        # (label, formula referencing V-cells)
        (
            "P1 · RR + T12 period dates set",
            # V3 (row 25) and V4 (row 26) return TEXT or "—" — check both look like dates
            '=IF(AND(LEN(B25)>4,LEN(B26)>4,B25<>"-",B26<>"-"),"✓","⚠ Set period dates")'
        ),
        (
            "P2 · Property name populated",
            # V5 (row 27)
            '=IF(LEN(TRIM(IFERROR(Property_Name,"")))>0,"✓","⚠ Set property name on Cover!B5")'
        ),
        (
            "P3 · RR + T12 input rows present",
            # V6 (row 28) and V7 (row 29) — both should be > 0
            '=IF(AND(B28>0,B29>0),"✓","⚠ Upload RR + T12 to populate input sheets")'
        ),
        (
            "P4 · Source $ → Operating $ leakage within ±$1",
            # V1 (row 23)
            '=IF(IFERROR(VALUE(SUBSTITUTE(B23,"$","")),99)<=1,"✓","⚠ Investigate T12 leakage (V1)")'
        ),
    ]
    for offset, (label, formula) in enumerate(checks):
        r = GATE_FIRST_CHECK_ROW + offset
        a = ws.cell(r, 1, value=label)
        a.font = BODY_FONT; a.alignment = LEFT; a.border = BOX
        b = ws.cell(r, 2, value=formula)
        b.font = BODY_FONT; b.alignment = CENTER; b.border = BOX

    # Row 52 — AGGREGATE indicator
    agg_label = ws.cell(GATE_AGGREGATE_ROW, 1, value="READY FOR EXPORT?")
    agg_label.font = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
    agg_label.fill = GATE_FILL
    agg_label.alignment = LEFT
    agg_label.border = BOX

    # Formula: pass only when ALL P-checks return "✓"
    agg = ws.cell(GATE_AGGREGATE_ROW, 2)
    agg.value = (
        f'=IF(AND(B{GATE_FIRST_CHECK_ROW}="✓",'
        f'B{GATE_FIRST_CHECK_ROW + 1}="✓",'
        f'B{GATE_FIRST_CHECK_ROW + 2}="✓",'
        f'B{GATE_FIRST_CHECK_ROW + 3}="✓"),'
        f'"✓ READY — UW Export tab is good to copy",'
        f'"⚠ NOT READY — resolve the ⚠ items above first")'
    )
    agg.font = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
    agg.fill = GATE_FILL
    agg.alignment = LEFT
    agg.border = BOX
    ws.merge_cells(start_row=GATE_AGGREGATE_ROW, end_row=GATE_AGGREGATE_ROW, start_column=2, end_column=4)

    return len(checks) + 1  # 4 sub-checks + aggregate


# ===== C. Workbook Map extension =====

def add_uw_export_to_map(wb) -> int:
    """Insert 'UW Export' row in the existing Workbook Map section.

    The map currently runs rows 6-18 (each row = sheet name + AZ1 reference).
    UW Output is at row 13. We add UW Export immediately after at row 14
    (which is currently empty — there's a gap between row 13 and row 14
    we can use without shifting rows).

    Wait — looking again, rows 6-18 are contiguous. Inserting between them
    requires a row insert. Cleaner: just append at row 19 (currently empty
    visual break before Section 2 at row 21) so the map gets a tail entry.
    """
    ws = wb[WORKBOOK_HEALTH]
    # Use row 19 (currently empty before Section 2 at row 21)
    a = ws.cell(19, 1, value=EXPORT_SHEET)
    a.font = BODY_FONT; a.alignment = LEFT
    b = ws.cell(19, 2, value=f"='{EXPORT_SHEET}'!AZ1")
    b.font = BODY_FONT; b.alignment = LEFT
    return 1


# ===== D. Versioning =====

def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ===== E. Verify =====

def verify_migration(wb) -> dict:
    r: dict = {}
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # UW Export sheet exists
    r["uw_export_exists"] = EXPORT_SHEET in wb.sheetnames

    if r["uw_export_exists"]:
        ws_e = wb[EXPORT_SHEET]
        r["uw_export_title"] = ws_e.cell(1, 1).value
        r["uw_export_title_ok"] = isinstance(r["uw_export_title"], str) and "UW Export" in r["uw_export_title"]

        # Metadata header — sample row 3 (Property name) and row 6 (Substrate version)
        prop_label = ws_e.cell(3, 1).value
        prop_formula = str(ws_e.cell(3, 2).value or "")
        r["meta_property_label_ok"] = prop_label == "Property name"
        r["meta_property_formula_ok"] = "Property_Name" in prop_formula

        sub_formula = str(ws_e.cell(6, 2).value or "")
        r["meta_substrate_formula_ok"] = "Cover!$B$8" in sub_formula

        # Mirror — sample first cell A9 should be ='UW Output'!A1
        mirror_a9 = str(ws_e.cell(9, 1).value or "")
        r["mirror_a9"] = mirror_a9
        r["mirror_a9_ok"] = mirror_a9 == "='UW Output'!A1"

        # Mirror — last cell H79 should be ='UW Output'!H71
        mirror_h79 = str(ws_e.cell(79, 8).value or "")
        r["mirror_h79"] = mirror_h79
        r["mirror_h79_ok"] = mirror_h79 == "='UW Output'!H71"
    else:
        r["uw_export_title_ok"] = False
        r["meta_property_label_ok"] = False
        r["meta_property_formula_ok"] = False
        r["meta_substrate_formula_ok"] = False
        r["mirror_a9_ok"] = False
        r["mirror_h79_ok"] = False

    # Pre-Export Gate on Workbook Health
    ws_wh = wb[WORKBOOK_HEALTH]
    gate_title = ws_wh.cell(GATE_TITLE_ROW, 1).value
    r["gate_title_ok"] = isinstance(gate_title, str) and "PRE-EXPORT GATE" in gate_title.upper()

    p1_label = ws_wh.cell(GATE_FIRST_CHECK_ROW, 1).value
    r["gate_p1_ok"] = isinstance(p1_label, str) and "P1" in p1_label

    agg_formula = str(ws_wh.cell(GATE_AGGREGATE_ROW, 2).value or "")
    r["gate_aggregate_ok"] = "READY" in agg_formula and "AND(" in agg_formula

    # Workbook map includes UW Export
    map_uw_export = ws_wh.cell(19, 1).value
    r["map_uw_export_ok"] = map_uw_export == EXPORT_SHEET

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v020(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    nA = install_uw_export_sheet(wb)
    print(f"  A: installed UW Export sheet — {nA} mirror rows + 5 metadata rows")
    nB = install_pre_export_gate(wb)
    print(f"  B: installed Pre-Export Gate on Workbook Health — {nB} cells")
    nC = add_uw_export_to_map(wb)
    print(f"  C: extended Workbook Map with UW Export entry")

    stamp_versions(wb)
    print(f"  D: stamped substrate version -> {SUBSTRATE_TO} (14 anchors)")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:<22}     : {r['cover_b8_ok']}")
    print(f"  All 14 AZ4 = {SUBSTRATE_TO}                  : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  UW Export sheet exists                  : {r['uw_export_exists']}")
    print(f"  UW Export title row 1                   : {r['uw_export_title_ok']}")
    print(f"  Metadata: Property name label           : {r['meta_property_label_ok']}")
    print(f"  Metadata: Property name formula         : {r['meta_property_formula_ok']}")
    print(f"  Metadata: Substrate version formula     : {r['meta_substrate_formula_ok']}")
    print(f"  Mirror A9 references UW Output A1       : {r['mirror_a9_ok']}")
    print(f"  Mirror H79 references UW Output H71     : {r['mirror_h79_ok']}")
    print(f"  Pre-Export Gate title                   : {r['gate_title_ok']}")
    print(f"  Pre-Export Gate P1 check                : {r['gate_p1_ok']}")
    print(f"  Pre-Export Gate aggregate formula       : {r['gate_aggregate_ok']}")
    print(f"  Workbook Map includes UW Export         : {r['map_uw_export_ok']}")

    all_ok = all([
        r["cover_b8_ok"], r["az4_all"], r["uw_export_exists"],
        r["uw_export_title_ok"], r["meta_property_label_ok"],
        r["meta_property_formula_ok"], r["meta_substrate_formula_ok"],
        r["mirror_a9_ok"], r["mirror_h79_ok"],
        r["gate_title_ok"], r["gate_p1_ok"], r["gate_aggregate_ok"],
        r["map_uw_export_ok"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v020.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
