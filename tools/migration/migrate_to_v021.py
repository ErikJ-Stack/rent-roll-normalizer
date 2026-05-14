"""
migrate_to_v021.py — Substrate template v0.2.0 → v0.2.1

Closes UW-BACKLOG BL-0001 — finer ancillary Labels in `Description_Map`.

Surfaced by substrate v0.1.12 Section M (Operator Fee Schedule &
Ancillary Reconciliation). M2 / M3 / M4 currently report 5 of the 7
default Homestead IL fees against the shared catchall Label
"Other community revenue", with a "(shared — see row N)" detection
that prevents double-reporting but also prevents per-fee T12
attribution. This release adds dedicated Labels so each of those 5
fees can be attributed at the GL level.

NEW LABELS (5):
    Meal Income           — formerly mapped to "Other community revenue"
    Housekeeping Income   — formerly mapped to "Other community revenue"
    Laundry Income        — formerly mapped to "Other community revenue"
    Scooter Fee Revenue   — formerly mapped to "Other community revenue"
    Transfer Fee Revenue  — formerly mapped to "Other community revenue"

Closed vocabulary grows 55 → 60 Labels.

THREE COORDINATED CHANGES:

  A. T12 Raw Data — insert 5 rows at R16 (between "2nd Person Revenue"
     at R15 and "Other community revenue" at R16). Each new row mirrors
     the R15 (2nd Person Revenue) template: col A="Revenue", col B=new
     Label, cols F-Q=SUMIF formulas against T12_Calc with the Label
     literal swapped, col R=SUM(F:Q) annual total. After insert, sweep
     all workbook formulas referencing T12 Raw Data row ≥ 16 by +5.

  B. Monthly Trending — insert 5 rows at R20 (between "2nd Person
     Revenue" at R19 and "Other community revenue" at R20). Each new
     row mirrors the R19 template: col A=new Label, cols B-M=
     INDEX/MATCH against T12 Raw Data, col N=SUM(B:M). After insert,
     sweep all workbook formulas referencing Monthly Trending row ≥ 20
     by +5. Patch the EGI formula at the new R26 (was R21) to include
     the 5 new rows (R20-R24).

  C. Description_Map — append 14 new Description→Label rows starting
     after the last populated row, mapping the typical operator-side
     descriptions for each new Label (e.g. "Meal Plan Revenue" →
     "Meal Income"; "Motorized Scooter Fee" → "Scooter Fee Revenue").
     No row insertion needed — Description_Map uses dynamic
     defined-name ranges (`DescMap_Description`, `DescMap_Label`) that
     auto-extend based on COUNTA.

  D. Rent Roll Recon Section M — re-point 5 of the M1 D-column "T12
     Label" cells from "Other community revenue" to the new specific
     Labels:
         D124: → "Transfer Fee Revenue"  (Elective Transfer Fee)
         D125: → "Meal Income"           (Meal Delivery)
         D126: → "Scooter Fee Revenue"   (Motorized Scooter Fee)
         D128: → "Housekeeping Income"   (Housekeeping)
         D129: → "Laundry Income"        (Laundry)
     M3's "(shared — see row N)" detection resolves automatically — the
     COUNTIF check finds no duplicates so per-fee VLOOKUPs flow through.
     M5's "Other community revenue" residual subtraction (B172) still
     works: it sums only fees whose D-column STILL says
     "Other community revenue", which after migration is just the
     Community Fee row (D123, which always pointed at
     "Community / move-in fees" anyway, so B172 = 0 → B173 = B170 with
     the per-fee fees now living in their own labeled rows).

Idempotency: gate checks BOTH the substrate version stamp AND that
T12 Raw Data!B16 says "Meal Income". Re-running is a no-op.

Mechanics on openpyxl row insertion (replicates the v0.1.5 pattern):
  insert_rows() shifts cell positions and formatting, but does NOT
  rewrite formula text. shift_all_formulas() does the full-workbook
  regex sweep after each insert.

Usage:
    python tools/migration/migrate_to_v021.py input.xlsx output.xlsx
"""
from __future__ import annotations

import re
import sys
from copy import copy
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.utils import get_column_letter


SUBSTRATE_FROM = "v0.2.0"
SUBSTRATE_TO = "v0.2.1"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# 5 new Labels — order is the order of T12 Raw Data row insertion + the
# Monthly Trending row insertion. M1's D-column re-pointing matches by
# *position* in this list to a specific M1 fee row.
NEW_LABELS: Tuple[str, ...] = (
    "Meal Income",
    "Housekeeping Income",
    "Laundry Income",
    "Scooter Fee Revenue",
    "Transfer Fee Revenue",
)

# T12 Raw Data layout
RAW_INSERT_AT_ROW = 16     # before existing "Other community revenue"
RAW_TEMPLATE_ROW = 15      # "2nd Person Revenue" — formulas to mirror
RAW_MONTH_FIRST_COL = 6    # F
RAW_T12_COL = 18           # R (annual = SUM(F:Q))

# Monthly Trending layout
MT_INSERT_AT_ROW = 20      # before existing "Other community revenue"
MT_TEMPLATE_ROW = 19       # "2nd Person Revenue"
MT_FIRST_DATA_COL = 2      # B
MT_LAST_MONTH_COL = 13     # M
MT_TOTAL_COL = 14          # N (=SUM(B:M))
MT_EGI_ORIGINAL_ROW = 21   # post-shift will be at MT_EGI_ORIGINAL_ROW + 5 = 26

# Description_Map appends
DESC_MAP_DESC_COL = 1
DESC_MAP_LABEL_COL = 2

# Section M1 D-column re-points (Rent Roll Recon)
SECTION_M_REPOINTS: List[Tuple[int, str]] = [
    (124, "Transfer Fee Revenue"),    # Elective Transfer Fee
    (125, "Meal Income"),             # Meal Delivery
    (126, "Scooter Fee Revenue"),     # Motorized Scooter Fee
    (128, "Housekeeping Income"),     # Housekeeping
    (129, "Laundry Income"),          # Laundry
]


# ---------------------------------------------------------------------------
# Description_Map appends — typical operator-side descriptions
# (Description, Label) pairs. Conservatively chosen: 2-4 per Label,
# covering the common chart-of-accounts spellings.
# ---------------------------------------------------------------------------
DESC_MAP_APPENDS: List[Tuple[str, str]] = [
    # Meal Income (4)
    ("Meal Income",          "Meal Income"),
    ("Meal Plan Revenue",    "Meal Income"),
    ("Meal Plan Income",     "Meal Income"),
    ("Dining Revenue",       "Meal Income"),
    # Housekeeping Income (3)
    ("Housekeeping Income",   "Housekeeping Income"),
    ("Housekeeping Revenue",  "Housekeeping Income"),
    ("H/K Income",            "Housekeeping Income"),
    # Laundry Income (2)
    ("Laundry Income",   "Laundry Income"),
    ("Laundry Revenue",  "Laundry Income"),
    # Scooter Fee Revenue (3)
    ("Scooter Fee",            "Scooter Fee Revenue"),
    ("Motorized Scooter Fee",  "Scooter Fee Revenue"),
    ("Mobility Fee",           "Scooter Fee Revenue"),
    # Transfer Fee Revenue (2)
    ("Transfer Fee",            "Transfer Fee Revenue"),
    ("Elective Transfer Fee",   "Transfer Fee Revenue"),
]


# ---------------------------------------------------------------------------
# Formula row-shift utility (copied verbatim from migrate_to_v015.py)
# ---------------------------------------------------------------------------

def shift_row_refs_in_formula(
    formula: str,
    threshold: int,
    delta: int,
    target_sheet: str,
    same_sheet: bool,
) -> str:
    """Increment every row reference in `formula` by `delta` if the row >=
    `threshold` AND the reference points at `target_sheet`."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    norm_target = target_sheet.lower()
    out = formula

    qualified_pattern = re.compile(
        r"('([^']+)'|([A-Za-z_][A-Za-z0-9_ ]*))!(\$?)([A-Z]+)(\$?)(\d+)"
    )

    def replace_qualified(m: re.Match) -> str:
        sheet_quoted = m.group(2)
        sheet_unquoted = m.group(3)
        sheet = sheet_quoted if sheet_quoted is not None else sheet_unquoted
        col_dollar = m.group(4)
        col = m.group(5)
        row_dollar = m.group(6)
        row_num = int(m.group(7))

        if sheet.lower() != norm_target:
            return m.group(0)
        if row_num < threshold:
            return m.group(0)

        new_row = row_num + delta
        sheet_qual = f"'{sheet}'" if sheet_quoted is not None else sheet
        return f"{sheet_qual}!{col_dollar}{col}{row_dollar}{new_row}"

    out = qualified_pattern.sub(replace_qualified, out)

    if same_sheet:
        unqualified_pattern = re.compile(
            r"(?<![A-Za-z_!])(\$?)([A-Z]+)(\$?)(\d+)\b"
        )

        def replace_unqualified(m: re.Match) -> str:
            col_dollar = m.group(1)
            col = m.group(2)
            row_dollar = m.group(3)
            row_num = int(m.group(4))
            if row_num < threshold:
                return m.group(0)
            return f"{col_dollar}{col}{row_dollar}{row_num + delta}"

        out = unqualified_pattern.sub(replace_unqualified, out)

    return out


def shift_merged_cells(ws, threshold: int, delta: int) -> int:
    """Shift merged-cell range definitions to keep up with insert_rows.
    Critical: do NOT use unmerge_cells (wipes displaced content).
    """
    shifted = 0
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= threshold:
            mr.shift(col_shift=0, row_shift=delta)
            shifted += 1
    return shifted


def shift_all_formulas(
    wb: openpyxl.Workbook,
    target_sheet: str,
    threshold: int,
    delta: int,
) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        modified = 0
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                same_sheet = (sheet_name == target_sheet)
                new_v = shift_row_refs_in_formula(
                    v, threshold, delta, target_sheet, same_sheet,
                )
                if new_v != v:
                    ws.cell(cell.row, cell.column, new_v)
                    modified += 1
        if modified > 0:
            counts[sheet_name] = modified
    return counts


def copy_row_formatting(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(src_row, col)
        dst_cell = ws.cell(dst_row, col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.border = copy(src_cell.border)


# ---------------------------------------------------------------------------
# Idempotency gate
# ---------------------------------------------------------------------------

def is_already_v021(wb: openpyxl.Workbook) -> bool:
    """Idempotent: both the version stamp AND the sentinel must agree."""
    cover_b8_ok = False
    try:
        cover_b8 = wb["Cover"]["B8"].value
        cover_b8_ok = (cover_b8 == SUBSTRATE_TO)
    except Exception:
        pass

    sentinel_ok = False
    try:
        v = wb["T12 Raw Data"].cell(RAW_INSERT_AT_ROW, 2).value
        sentinel_ok = (v == NEW_LABELS[0])  # "Meal Income"
    except Exception:
        pass

    return cover_b8_ok and sentinel_ok


# ---------------------------------------------------------------------------
# A. T12 Raw Data — insert 5 rows + shift + populate
# ---------------------------------------------------------------------------

def step_t12_raw_data(wb: openpyxl.Workbook) -> int:
    ws = wb["T12 Raw Data"]
    delta = len(NEW_LABELS)
    target = RAW_INSERT_AT_ROW
    template_row = RAW_TEMPLATE_ROW

    # We read template formulas AFTER the shift sweep, not before. Reading
    # pre-shift would cache the OLD `$N$500` endpoints; the shift pass would
    # then bump R15's pre-existing endpoints to `$N$505` while the new rows
    # we populate stay at `$N$500`. Reading post-shift keeps every row's
    # SUMIF range endpoints consistent. (The drift is a known openpyxl quirk
    # — the unqualified-ref regex catches the colon-suffix endpoint of a
    # qualified `T12_Calc!$N$1:$N$NNN` range; it's the same artifact that
    # produced the v0.1.7 "SUMIFS N501 vs N500 cosmetic" sweep.)
    template_label = ws.cell(template_row, 2).value
    print(f"  T12 Raw Data: template = R{template_row} ({template_label!r})")
    print(f"  T12 Raw Data: inserting {delta} rows at R{target}")
    ws.insert_rows(target, amount=delta)

    n_merges = shift_merged_cells(ws, target, +delta)
    print(f"  T12 Raw Data: shifted {n_merges} merged-cell range(s)")

    counts = shift_all_formulas(wb, "T12 Raw Data", target, +delta)
    total = sum(counts.values())
    print(f"  T12 Raw Data: shifted formula refs in {total} cells: {counts}")

    # NOW capture template formulas — post-shift, so any range-endpoint
    # drift in R15 is already baked in.
    template_formulas: Dict[int, str] = {}
    for c in range(RAW_MONTH_FIRST_COL, RAW_T12_COL + 1):
        f = ws.cell(template_row, c).value
        if isinstance(f, str) and f.startswith("="):
            template_formulas[c] = f

    # Populate each new row by mirroring the template
    for i, new_label in enumerate(NEW_LABELS):
        row = target + i
        ws.cell(row, 1, "Revenue")
        ws.cell(row, 2, new_label)
        copy_row_formatting(ws, template_row, row)

        for c, template in template_formulas.items():
            new_formula = template.replace(f'"{template_label}"', f'"{new_label}"')
            # Update bare row refs (template_row → row)
            new_formula = re.sub(
                rf"\b([A-Z]+){template_row}\b",
                lambda m, _r=row: f"{m.group(1)}{_r}",
                new_formula,
            )
            ws.cell(row, c, new_formula)
        print(f"    R{row}: {new_label!r} ({len(template_formulas)} formulas)")

    return delta


# ---------------------------------------------------------------------------
# B. Monthly Trending — insert 5 rows + shift + populate + patch EGI
# ---------------------------------------------------------------------------

def step_monthly_trending(wb: openpyxl.Workbook) -> int:
    ws = wb["Monthly Trending"]
    delta = len(NEW_LABELS)
    target = MT_INSERT_AT_ROW
    template_row = MT_TEMPLATE_ROW

    # Same post-shift template read as step_t12_raw_data — uniformity over
    # any latent cross-sheet range drift.
    template_label = ws.cell(template_row, 1).value
    print(f"  Monthly Trending: template = R{template_row} ({template_label!r})")
    print(f"  Monthly Trending: inserting {delta} rows at R{target}")
    ws.insert_rows(target, amount=delta)

    n_merges = shift_merged_cells(ws, target, +delta)
    print(f"  Monthly Trending: shifted {n_merges} merged-cell range(s)")

    counts = shift_all_formulas(wb, "Monthly Trending", target, +delta)
    total = sum(counts.values())
    print(f"  Monthly Trending: shifted formula refs in {total} cells: {counts}")

    template_formulas: Dict[int, str] = {}
    for c in range(MT_FIRST_DATA_COL, MT_TOTAL_COL + 1):
        f = ws.cell(template_row, c).value
        if isinstance(f, str) and f.startswith("="):
            template_formulas[c] = f

    # Populate each new row
    for i, new_label in enumerate(NEW_LABELS):
        row = target + i
        ws.cell(row, 1, new_label)
        copy_row_formatting(ws, template_row, row)

        for c, template in template_formulas.items():
            new_formula = template.replace(
                f'MATCH("{template_label}"',
                f'MATCH("{new_label}"',
            )
            # Update bare row refs (template_row → row) for col N self-sum
            new_formula = re.sub(
                rf"\b([A-Z]+){template_row}\b",
                lambda m, _r=row: f"{m.group(1)}{_r}",
                new_formula,
            )
            ws.cell(row, c, new_formula)
        print(f"    R{row}: {new_label!r} ({len(template_formulas)} formulas)")

    # --- Patch EGI to include the new rows ---
    # Pre-migration EGI at MT_EGI_ORIGINAL_ROW (21) was
    #   =B8+B10+B11+B15+B16+B17+B18+B19+B20
    # After insert+shift, EGI sits at MT_EGI_ORIGINAL_ROW + delta (26) with
    # the B20 ref shifted to B25 (Other community revenue's new row).
    # We add the 5 new rows R20-R24.
    egi_row = MT_EGI_ORIGINAL_ROW + delta  # 26
    new_rows = list(range(target, target + delta))  # [20,21,22,23,24]
    for c in range(MT_FIRST_DATA_COL, MT_TOTAL_COL + 1):
        existing = ws.cell(egi_row, c).value
        if not isinstance(existing, str) or not existing.startswith("="):
            continue
        col_letter = get_column_letter(c)
        # Pre-existing rows in EGI: 8, 10, 11, 15, 16, 17, 18, 19, 25 (post-shift)
        # New rows: 20, 21, 22, 23, 24
        pre_existing = [8, 10, 11, 15, 16, 17, 18, 19, target + delta]  # 25
        all_rows = sorted(pre_existing + new_rows)
        new_egi = "=" + "+".join(f"{col_letter}{r}" for r in all_rows)
        ws.cell(egi_row, c, new_egi)
    print(f"  Monthly Trending: EGI at R{egi_row} rewritten to include R{target}-R{target+delta-1}")

    return delta


# ---------------------------------------------------------------------------
# C. Description_Map appends
# ---------------------------------------------------------------------------

def step_description_map(wb: openpyxl.Workbook) -> int:
    """Append the 14 new (Description, Label) rows to Description_Map.

    Uses ws.max_row to find the bottom — but ws.max_row can include
    formatted-but-empty trailing rows on some workbooks. Walk back from
    max_row to find the last row with an actual value in col A.
    """
    ws = wb["Description_Map"]
    last_used = 4  # header at row 4; data starts at row 5
    for r in range(ws.max_row, 4, -1):
        if ws.cell(r, DESC_MAP_DESC_COL).value or ws.cell(r, DESC_MAP_LABEL_COL).value:
            last_used = r
            break

    print(f"  Description_Map: last populated row = R{last_used}; appending {len(DESC_MAP_APPENDS)} new rows")

    for i, (desc, label) in enumerate(DESC_MAP_APPENDS):
        row = last_used + 1 + i
        ws.cell(row, DESC_MAP_DESC_COL, desc)
        ws.cell(row, DESC_MAP_LABEL_COL, label)
        # Mirror formatting from a reference data row (R5)
        copy_row_formatting(ws, 5, row)

    return len(DESC_MAP_APPENDS)


# ---------------------------------------------------------------------------
# D. Section M D-column re-points (Rent Roll Recon)
# ---------------------------------------------------------------------------

def step_section_m_repoint(wb: openpyxl.Workbook) -> int:
    """Re-point M1's "T12 Label" column from the catchall to the new specific
    Labels for 5 of the 7 default fees. M2 / M3 / M4 read D124-D131 via
    relative references, so the change propagates automatically without
    further edits.
    """
    ws = wb["Rent Roll Recon"]
    for row, new_label in SECTION_M_REPOINTS:
        prev = ws.cell(row, 4).value  # col D
        ws.cell(row, 4, new_label)
        print(f"    D{row}: {prev!r} → {new_label!r}")
    return len(SECTION_M_REPOINTS)


# ---------------------------------------------------------------------------
# Stamping
# ---------------------------------------------------------------------------

def stamp_versions(wb: openpyxl.Workbook) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

def verify_migration(wb: openpyxl.Workbook) -> dict:
    r: dict = {}

    # 1. Cover!B8 stamp
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    # 2. All 14 AZ4 anchors
    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # 3. T12 Raw Data rows 16-20 have the 5 new Labels
    ws = wb["T12 Raw Data"]
    actual_labels = [ws.cell(RAW_INSERT_AT_ROW + i, 2).value for i in range(5)]
    r["raw_labels"] = actual_labels
    r["raw_labels_ok"] = actual_labels == list(NEW_LABELS)

    # 4. T12 Raw Data R21 now says "Other community revenue" (shifted from R16)
    r["raw_other_at_21"] = ws.cell(21, 2).value
    r["raw_other_at_21_ok"] = r["raw_other_at_21"] == "Other community revenue"

    # 5. T12 Raw Data new rows have SUMIF formulas referencing the new Label
    f16 = ws.cell(16, 6).value  # F16 = SUMIF for Meal Income
    r["raw_f16"] = f16
    r["raw_f16_ok"] = isinstance(f16, str) and '"Meal Income"' in f16

    # 6. Monthly Trending rows 20-24 have the 5 new Labels
    ws = wb["Monthly Trending"]
    actual_mt = [ws.cell(MT_INSERT_AT_ROW + i, 1).value for i in range(5)]
    r["mt_labels"] = actual_mt
    r["mt_labels_ok"] = actual_mt == list(NEW_LABELS)

    # 7. Monthly Trending R25 is "Other community revenue", R26 is EGI
    r["mt_other_at_25"] = ws.cell(25, 1).value
    r["mt_other_at_25_ok"] = r["mt_other_at_25"] == "Other community revenue"
    r["mt_egi_at_26"] = ws.cell(26, 1).value
    r["mt_egi_at_26_ok"] = isinstance(r["mt_egi_at_26"], str) and "EFFECTIVE GROSS INCOME" in r["mt_egi_at_26"]

    # 8. EGI formula at R26 includes the 5 new rows
    egi_b26 = ws.cell(26, 2).value
    r["mt_egi_b26"] = egi_b26
    r["mt_egi_b26_ok"] = (
        isinstance(egi_b26, str)
        and "B20" in egi_b26 and "B21" in egi_b26
        and "B22" in egi_b26 and "B23" in egi_b26 and "B24" in egi_b26
        and "B25" in egi_b26  # original Other community revenue, shifted from B20
    )

    # 9. INDEX/MATCH on Monthly Trending row 20 references "Meal Income"
    mt_b20 = ws.cell(20, 2).value
    r["mt_b20"] = mt_b20
    r["mt_b20_ok"] = isinstance(mt_b20, str) and '"Meal Income"' in mt_b20

    # 10. Section M re-points — D124-D129
    ws = wb["Rent Roll Recon"]
    repoints_ok = True
    for row, expected in SECTION_M_REPOINTS:
        if ws.cell(row, 4).value != expected:
            repoints_ok = False
            break
    r["section_m_repoints_ok"] = repoints_ok

    # 11. Description_Map appends — find the 14 new rows
    ws = wb["Description_Map"]
    descs = set()
    for row in range(5, ws.max_row + 1):
        v = ws.cell(row, DESC_MAP_DESC_COL).value
        if v:
            descs.add(str(v).strip())
    expected_descs = {d for d, _ in DESC_MAP_APPENDS}
    r["desc_map_appended_count"] = len(expected_descs & descs)
    r["desc_map_appends_ok"] = r["desc_map_appended_count"] == len(DESC_MAP_APPENDS)

    # 12. EGI absolute reference in Section M5 (Rent Roll Recon B174) shifted
    ws = wb["Rent Roll Recon"]
    b174 = ws.cell(174, 2).value
    r["rr_b174"] = b174
    r["rr_b174_ok"] = isinstance(b174, str) and "'Monthly Trending'!$N$26" in b174

    # 13. Sentinel check (also tested in is_already_v021)
    r["sentinel_ok"] = wb["T12 Raw Data"].cell(RAW_INSERT_AT_ROW, 2).value == NEW_LABELS[0]

    return r


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v021(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...\n")

    print("Step A — T12 Raw Data: insert 5 rows + shift refs + populate new Labels:")
    step_t12_raw_data(wb)
    print()

    print("Step B — Monthly Trending: insert 5 rows + shift refs + populate + patch EGI:")
    step_monthly_trending(wb)
    print()

    print("Step C — Description_Map: append typical Description->Label rows for the 5 new Labels:")
    step_description_map(wb)
    print()

    print("Step D — Rent Roll Recon Section M: re-point 5 D-column entries from 'Other community revenue':")
    step_section_m_repoint(wb)
    print()

    stamp_versions(wb)
    print(f"Step E — stamped substrate version -> {SUBSTRATE_TO} (14 anchors)")

    print(f"\nSaving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"   1. Cover!B8 = {r['cover_b8']!r:<10}                                         : {r['cover_b8_ok']}")
    print(f"   2. All 14 AZ4 = {SUBSTRATE_TO}                                          : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"   3. T12 Raw Data R16-R20 = {NEW_LABELS}")
    print(f"      actual: {r['raw_labels']}")
    print(f"      ok                                                          : {r['raw_labels_ok']}")
    print(f"   4. T12 Raw Data R21 = 'Other community revenue' (shifted)             : {r['raw_other_at_21_ok']}")
    print(f"   5. T12 Raw Data F16 SUMIF refers to 'Meal Income'                     : {r['raw_f16_ok']}")
    print(f"   6. Monthly Trending R20-R24 = 5 new labels                            : {r['mt_labels_ok']}")
    print(f"   7. Monthly Trending R25 = 'Other community revenue' & R26 = EGI       : {r['mt_other_at_25_ok'] and r['mt_egi_at_26_ok']}")
    print(f"   8. Monthly Trending B26 EGI includes new rows B20-B24 + B25           : {r['mt_egi_b26_ok']}")
    print(f"   9. Monthly Trending B20 INDEX/MATCH refers to 'Meal Income'           : {r['mt_b20_ok']}")
    print(f"  10. Rent Roll Recon Section M D124-D129 re-pointed to new Labels       : {r['section_m_repoints_ok']}")
    print(f"  11. Description_Map appended {r['desc_map_appended_count']}/{len(DESC_MAP_APPENDS)} new descriptions                : {r['desc_map_appends_ok']}")
    print(f"  12. Rent Roll Recon B174 EGI reference shifted to Monthly Trending N26 : {r['rr_b174_ok']}")
    print(f"  13. Sentinel: T12 Raw Data!B16 = {NEW_LABELS[0]!r}                       : {r['sentinel_ok']}")

    all_ok = all([
        r["cover_b8_ok"], r["az4_all"], r["raw_labels_ok"],
        r["raw_other_at_21_ok"], r["raw_f16_ok"], r["mt_labels_ok"],
        r["mt_other_at_25_ok"], r["mt_egi_at_26_ok"], r["mt_egi_b26_ok"],
        r["mt_b20_ok"], r["section_m_repoints_ok"], r["desc_map_appends_ok"],
        r["rr_b174_ok"], r["sentinel_ok"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v021.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
