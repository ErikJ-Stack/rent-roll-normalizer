"""
Migrate ALF_Financial_Analyzer_Only.xlsx from substrate v0.1.4 to v0.1.5.

v0.1.5 ADDS one new revenue Label: "2nd Person Revenue".

The Label tracks second-occupant revenue separately from base rent so that
per-bed base-rate calculations (Base rent ÷ ADC) stay clean. First surfaced
on the Homestead Pensacola broker file (2026-05) where second-person
revenue was reported as its own line item by care type, but the substrate
had no Label to receive it without inflating Base rent.

Mechanics: openpyxl's `insert_rows()` shifts cell values down but does NOT
update formula text — so every formula in the workbook that references a
shifted row is now off by 1 until we explicitly rewrite it. This script
does the full sweep across all sheets after each insert.

CHANGES (idempotent — script detects v0.1.5 and exits if already applied):

  1. T12 Raw Data: insert one row at R15 (before "Other community
     revenue") with col A="Revenue", col B="2nd Person Revenue", and
     SUMIF/SUM formulas built from the R14 (Respite care) template with
     row refs and the matched-label string updated. Then shift all
     workbook formulas referencing T12 Raw Data row >= 15 by +1.

  2. Monthly Trending: insert one row at R19 (before "Other community
     revenue"). Shift all workbook formulas referencing Monthly Trending
     row >= 19 by +1. Update EGI formula at (post-insert) R21 to include
     the new R19 (2nd Person Revenue) — this addition is NOT covered by
     the row-shift sweep because EGI's original formula never referenced
     the "2nd Person" position (the label didn't exist before this
     migration).

Description_Map gets new entries at runtime via the parser/writer flow,
not as part of this substrate migration. The new Label becomes available
in the substrate's MATCH/SUMIF logic the moment it appears in Monthly
Trending and T12 Raw Data.

After this migration runs, recalc the file with LibreOffice (scripts/
recalc.py) before consuming it as a destination workbook — openpyxl
writes formulas but does not evaluate them, and downstream code reading
data_only=True needs the recalculated cache.
"""

from __future__ import annotations

import re
import sys
from copy import copy
from pathlib import Path
from typing import Dict

import openpyxl


# ---------------------------------------------------------------------------
# Constants — match the v0.1.4 substrate layout
# ---------------------------------------------------------------------------

NEW_LABEL = "2nd Person Revenue"

MT_INSERT_BEFORE_ROW = 19
RAW_INSERT_BEFORE_ROW = 15

# T12 Raw Data layout
RAW_MONTH_FIRST_COL = 6   # F
RAW_MONTH_LAST_COL = 17   # Q
RAW_T12_COL = 18          # R


# ---------------------------------------------------------------------------
# Formula row-shift utility
# ---------------------------------------------------------------------------

def shift_row_refs_in_formula(
    formula: str,
    threshold: int,
    delta: int,
    target_sheet: str,
    same_sheet: bool,
) -> str:
    """Increment every row reference in `formula` by `delta` if the row >=
    `threshold` AND the reference points at `target_sheet`.

    Two passes:
      Pass 1 — qualified refs ('Sheet'!A19 or Sheet!A19): shift only when
        the qualifier matches `target_sheet` (case-insensitive).
      Pass 2 — unqualified refs (B19, $A$5): only shift when the formula
        lives IN `target_sheet` itself (`same_sheet=True`).

    Edge cases handled:
      - Absolute refs ($A$19, A$19, $A19) — preserved with `$` markers.
      - Range refs (A19:B30) — both endpoints get evaluated and shifted.
      - String literals like "Some text 19" — the lookbehind on the
        unqualified pattern excludes letter-then-digit inside strings
        because they're preceded by alphabetic text.
      - Function names like SUM, MATCH — excluded by lookbehind from
        unqualified-ref matching.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    norm_target = target_sheet.lower()
    out = formula

    # ---------- Pass 1: qualified refs ----------
    qualified_pattern = re.compile(
        r"('([^']+)'|([A-Za-z_][A-Za-z0-9_ ]*))!(\$?)([A-Z]+)(\$?)(\d+)"
    )

    def replace_qualified(m: re.Match) -> str:
        sheet_quoted = m.group(2)
        sheet_unquoted = m.group(3)
        sheet = sheet_quoted if sheet_quoted is not None else sheet_unquoted
        col_dollar = m.group(4)
        col = m.group(5)
        row_dollar = m.group(6)
        row_num = int(m.group(7))

        if sheet.lower() != norm_target:
            return m.group(0)
        if row_num < threshold:
            return m.group(0)

        new_row = row_num + delta
        sheet_qual = f"'{sheet}'" if sheet_quoted is not None else sheet
        return f"{sheet_qual}!{col_dollar}{col}{row_dollar}{new_row}"

    out = qualified_pattern.sub(replace_qualified, out)

    # ---------- Pass 2: unqualified refs ----------
    if same_sheet:
        # Match a cell ref NOT preceded by '!' (which means it's qualified
        # to another sheet) or by another letter/underscore (function name
        # like SUM or string-literal context). Colon IS allowed: the second
        # endpoint of a range like A1:B19 must also be matched and shifted.
        unqualified_pattern = re.compile(
            r"(?<![A-Za-z_!])(\$?)([A-Z]+)(\$?)(\d+)\b"
        )

        def replace_unqualified(m: re.Match) -> str:
            col_dollar = m.group(1)
            col = m.group(2)
            row_dollar = m.group(3)
            row_num = int(m.group(4))
            if row_num < threshold:
                return m.group(0)
            return f"{col_dollar}{col}{row_dollar}{row_num + delta}"

        out = unqualified_pattern.sub(replace_unqualified, out)

    return out


def shift_merged_cells(
    ws,
    threshold: int,
    delta: int,
) -> int:
    """Shift any merged-cell range whose top row >= threshold by `delta` rows.

    openpyxl's `insert_rows()` shifts cell positions but NOT the merged-cell
    range definitions. This function patches that omission. Without this,
    rows that get displaced by the insert end up still inside the OLD merge
    range, which causes Excel to silently drop their values on save (because
    merged cells only retain the top-left cell's value).

    Critically: we must NOT use `ws.unmerge_cells()` to do this. unmerge
    clears all non-top-left cells of the previously merged range — but
    after insert_rows, the cells at those positions now contain DIFFERENT
    content (formulas/values that got displaced into them by the insert).
    Calling unmerge would silently wipe that content. Instead, we use the
    merge range's `shift()` method, which mutates the bounds directly.

    Returns the number of merge ranges shifted.
    """
    shifted = 0
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= threshold:
            mr.shift(col_shift=0, row_shift=delta)
            shifted += 1
    return shifted


def shift_all_formulas(
    wb: openpyxl.Workbook,
    target_sheet: str,
    threshold: int,
    delta: int,
) -> Dict[str, int]:
    """Walk every cell in every sheet and shift formula refs to `target_sheet`
    rows >= threshold by `delta`. Returns {sheet_name: cells_modified}.
    """
    counts: Dict[str, int] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        modified = 0
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                same_sheet = (sheet_name == target_sheet)
                new_v = shift_row_refs_in_formula(
                    v, threshold, delta, target_sheet, same_sheet,
                )
                if new_v != v:
                    ws.cell(cell.row, cell.column, new_v)
                    modified += 1
        if modified > 0:
            counts[sheet_name] = modified
    return counts


# ---------------------------------------------------------------------------
# Formatting copy
# ---------------------------------------------------------------------------

def copy_row_formatting(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(src_row, col)
        dst_cell = ws.cell(dst_row, col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.border = copy(src_cell.border)


# ---------------------------------------------------------------------------
# Idempotency check
# ---------------------------------------------------------------------------

def detect_already_v015(wb: openpyxl.Workbook) -> bool:
    if "T12 Raw Data" not in wb.sheetnames:
        return False
    ws = wb["T12 Raw Data"]
    for r in range(4, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v and str(v).strip() == NEW_LABEL:
            return True
    return False


# ---------------------------------------------------------------------------
# Migration steps
# ---------------------------------------------------------------------------

def step_t12_raw_data(wb: openpyxl.Workbook) -> None:
    """Insert new row at R15 + shift all dependent formula refs + populate."""
    ws = wb["T12 Raw Data"]
    target = RAW_INSERT_BEFORE_ROW

    template_row = target - 1
    template_label = ws.cell(template_row, 2).value
    template_formulas: Dict[int, str] = {}
    for c in range(RAW_MONTH_FIRST_COL, RAW_T12_COL + 1):
        f = ws.cell(template_row, c).value
        if isinstance(f, str) and f.startswith("="):
            template_formulas[c] = f

    print(f"  T12 Raw Data: template = R{template_row} ({template_label!r})")

    ws.insert_rows(target)
    print(f"  T12 Raw Data: row inserted at R{target}")

    n_merges = shift_merged_cells(ws, target, +1)
    print(f"  T12 Raw Data: shifted {n_merges} merged-cell range(s)")

    counts = shift_all_formulas(wb, "T12 Raw Data", target, +1)
    total = sum(counts.values())
    print(f"  T12 Raw Data: shifted formula refs in {total} cells: {counts}")

    # Populate new row
    ws.cell(target, 1, "Revenue")
    ws.cell(target, 2, NEW_LABEL)
    copy_row_formatting(ws, target + 1, target)  # template now sits at target+1

    for c, template in template_formulas.items():
        # (a) Replace SUMIF label literal
        new_formula = template.replace(f'"{template_label}"', f'"{NEW_LABEL}"')
        # (b) Update bare row refs (template_row → target)
        new_formula = re.sub(
            rf"\b([A-Z]+){template_row}\b",
            lambda m: f"{m.group(1)}{target}",
            new_formula,
        )
        ws.cell(target, c, new_formula)
    print(f"  T12 Raw Data: populated R{target} with {len(template_formulas)} formulas")


def step_monthly_trending(wb: openpyxl.Workbook) -> None:
    """Insert new row at R19 + shift all dependent formula refs + populate
    + patch EGI to add the new R19 to the sum."""
    ws = wb["Monthly Trending"]
    target = MT_INSERT_BEFORE_ROW

    template_row = target - 1
    template_label = ws.cell(template_row, 1).value
    template_formulas: Dict[int, str] = {}
    for c in range(2, 15):
        f = ws.cell(template_row, c).value
        if isinstance(f, str) and f.startswith("="):
            template_formulas[c] = f

    print(f"  Monthly Trending: template = R{template_row} ({template_label!r})")

    ws.insert_rows(target)
    print(f"  Monthly Trending: row inserted at R{target}")

    n_merges = shift_merged_cells(ws, target, +1)
    print(f"  Monthly Trending: shifted {n_merges} merged-cell range(s)")

    counts = shift_all_formulas(wb, "Monthly Trending", target, +1)
    total = sum(counts.values())
    print(f"  Monthly Trending: shifted formula refs in {total} cells: {counts}")

    # Populate new row
    ws.cell(target, 1, NEW_LABEL)
    copy_row_formatting(ws, target + 1, target)

    for c, template in template_formulas.items():
        # Replace MATCH("...") label literal
        new_formula = template.replace(
            f'MATCH("{template_label}"',
            f'MATCH("{NEW_LABEL}"',
        )
        # Update bare row refs (template_row → target) for the col N
        # =SUM(B##:M##) self-reference
        new_formula = re.sub(
            rf"\b([A-Z]+){template_row}\b",
            lambda m: f"{m.group(1)}{target}",
            new_formula,
        )
        ws.cell(target, c, new_formula)
    print(f"  Monthly Trending: populated R{target} with {len(template_formulas)} formulas")

    # --- Patch EGI to include the new R19 ---
    # Pre-migration EGI at R20 was =B8+B10+B11+B15+B16+B17+B18+B19. After
    # insert+shift, EGI sits at R21 with formula =B8+B10+B11+B15+B16+B17+B18+B20
    # (B19 → B20 because old "Other community revenue" shifted from R19 → R20).
    # We need to ALSO add the new R19 (2nd Person Revenue) to that sum.
    egi_row = target + 2  # 21
    for c in range(2, 15):
        existing = ws.cell(egi_row, c).value
        if not isinstance(existing, str) or not existing.startswith("="):
            continue
        col_letter = ws.cell(egi_row, c).column_letter
        # Rewrite explicitly to avoid regex fragility
        new_egi = (
            f"={col_letter}8+{col_letter}10+{col_letter}11+"
            f"{col_letter}15+{col_letter}16+{col_letter}17+"
            f"{col_letter}18+{col_letter}19+{col_letter}20"
        )
        ws.cell(egi_row, c, new_egi)
    print(f"  Monthly Trending: EGI at R{egi_row} rewritten to include R19 (2nd Person)")


# ---------------------------------------------------------------------------
# Top-level migration
# ---------------------------------------------------------------------------

def migrate(input_path: Path, output_path: Path) -> None:
    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path, data_only=False)

    if detect_already_v015(wb):
        print(f"  Detected '{NEW_LABEL}' already in T12 Raw Data — workbook "
              f"is already v0.1.5 or later. Exiting (idempotent).")
        return

    print(f"\nApplying v0.1.4 → v0.1.5 migration...\n")
    print("Step 1 — T12 Raw Data:")
    step_t12_raw_data(wb)
    print("\nStep 2 — Monthly Trending:")
    step_monthly_trending(wb)

    print(f"\nSaving → {output_path}")
    wb.save(output_path)
    print("Done.\n")
    print("NOTE: openpyxl writes formulas but does not evaluate them. Recalc "
          "the file with scripts/recalc.py before downstream consumption.")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: migrate_to_v015.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    migrate(Path(sys.argv[1]), Path(sys.argv[2]))
