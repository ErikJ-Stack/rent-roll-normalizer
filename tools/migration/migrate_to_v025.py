"""
migrate_to_v025.py — Substrate template v0.2.4 → v0.2.5

Closes UW-BACKLOG BL-0012 — Section M Misc/Diabetes credit reconciliation
against T12 `Concessions & specials`.

Surfaced by RR v1.17.0 (BL-0003) "Side observation worth tracking" in
CHANGELOG-RR.md: Homestead's residual `Other LOC $` post-split is
**-$12,146.75** (Diabetes + Misc, both partially negative — net credit).
The negative residual likely reflects discount/credit postings that operators
sometimes route through Other LOC instead of the formal `Concessions` GL.
Section M5 currently treats negative residuals the same as positive, so it
surfaces a misleading "✓ Misc. income share within band" note when the bucket
is actually a net credit.

ONE STEP:

  A. Add Section M6 below the existing M5 (rows 178-184) on Rent Roll Recon.
     M6 only fires on NEGATIVE residuals (positive ones are M5's domain).
     Layout:

       R178: Section header — "M6  —  Negative residual check
              (Misc. credits vs T12 Concessions & specials)"
       R179: "Residual from M5 (annual)" | =B173
       R180: "T12 'Concessions & specials' — annual"
              | =IFERROR(VLOOKUP("Concessions & specials",
                'T12 Raw Data'!$B:$R, 17, 0), 0)
       R181: "Residual / T12 Concessions (abs)"
              | =IFERROR(ABS(B179)/ABS(B180), 0)
       R183: (merged A:I) conditional note:
              - If B179 >= 0: "" (positive residuals are M5's domain)
              - elif B180 == 0: ⚠ "Negative residual but T12 has no
                Concessions line — verify GL routing"
              - elif ABS(B179)/ABS(B180) > 10%: ⚠ "Likely misposted
                concessions — review GL"
              - else: ✓ "Within reconciliation tolerance (≤10% of T12
                Concessions)"

  B. Stamp Cover!B8 + 15 AZ4 anchors to v0.2.5.

Idempotency: gate checks BOTH `Cover!B8 == "v0.2.5"` AND
`Rent Roll Recon!A178` starts with "M6". Re-run is a no-op.

Usage:
    python tools/migration/migrate_to_v025.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SUBSTRATE_FROM = "v0.2.4"
SUBSTRATE_TO = "v0.2.5"

# 15-sheet anchor list (unchanged from v0.2.4)
ANCHOR_SHEETS = (
    "Cover", "Investment Dashboard",
    "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

RR_RECON = "Rent Roll Recon"

# Section M6 layout (rows 178-184; R177 left blank for spacing)
M6_HEADER_ROW = 178           # "M6 — Negative residual check ..."
M6_RESIDUAL_ROW = 179         # Residual from M5
M6_T12_CONC_ROW = 180         # T12 Concessions & specials annual
M6_RATIO_ROW = 181            # |Residual| / |T12 Concessions|
M6_NOTE_ROW = 183             # Conditional note (merged A:I)

# Style references — mirror the existing M5 block (R169 header, R170 data row,
# R176 conditional note)
STYLE_REF_HEADER_ROW = 169
STYLE_REF_DATA_ROW = 170
STYLE_REF_NOTE_ROW = 176

# Threshold for the ⚠ warning. 10% of T12 Concessions absolute value.
RESIDUAL_THRESHOLD = 0.10


# ---------------------------------------------------------------------------
# Idempotency gate
# ---------------------------------------------------------------------------

def is_already_v025(wb: openpyxl.Workbook) -> bool:
    cover_b8_ok = False
    try:
        cover_b8 = wb["Cover"]["B8"].value
        cover_b8_ok = (cover_b8 == SUBSTRATE_TO)
    except Exception:
        pass

    sentinel_ok = False
    try:
        v = wb[RR_RECON].cell(M6_HEADER_ROW, 1).value
        sentinel_ok = isinstance(v, str) and v.strip().startswith("M6")
    except Exception:
        pass

    return cover_b8_ok and sentinel_ok


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _copy_cell_style(src_cell, dst_cell) -> None:
    """Copy font/fill/alignment/border/number_format from src to dst, defensively."""
    if src_cell.font:
        dst_cell.font = copy(src_cell.font)
    if src_cell.fill:
        dst_cell.fill = copy(src_cell.fill)
    if src_cell.alignment:
        dst_cell.alignment = copy(src_cell.alignment)
    if src_cell.border:
        dst_cell.border = copy(src_cell.border)
    dst_cell.number_format = src_cell.number_format


# ---------------------------------------------------------------------------
# A. Install Section M6
# ---------------------------------------------------------------------------

def install_section_m6(wb: openpyxl.Workbook) -> int:
    ws = wb[RR_RECON]

    cells_set = 0

    # --- Header row (R178) — merge A:G to match M5 header style ---
    hdr = ws.cell(M6_HEADER_ROW, 1)
    hdr.value = "M6  —  Negative residual check  (Misc. credits vs T12 Concessions & specials)"
    _copy_cell_style(ws.cell(STYLE_REF_HEADER_ROW, 1), hdr)
    # Match the M5 header's merge if present — A:G covers the header span
    # Don't merge if a merge already exists there from a prior partial run
    try:
        ws.merge_cells(start_row=M6_HEADER_ROW, end_row=M6_HEADER_ROW,
                       start_column=1, end_column=7)
    except Exception:
        pass
    cells_set += 1

    # --- Data rows (R179-R181) — mirror M5 data styling ---
    label_style = ws.cell(STYLE_REF_DATA_ROW, 1)
    value_style = ws.cell(STYLE_REF_DATA_ROW, 2)

    # R179 — Residual from M5
    a179 = ws.cell(M6_RESIDUAL_ROW, 1)
    a179.value = "Residual from M5 (annual)"
    _copy_cell_style(label_style, a179)
    b179 = ws.cell(M6_RESIDUAL_ROW, 2)
    b179.value = "=B173"
    _copy_cell_style(value_style, b179)
    cells_set += 2

    # R180 — T12 Concessions & specials annual
    a180 = ws.cell(M6_T12_CONC_ROW, 1)
    a180.value = "T12 'Concessions & specials' — annual"
    _copy_cell_style(label_style, a180)
    b180 = ws.cell(M6_T12_CONC_ROW, 2)
    b180.value = "=IFERROR(VLOOKUP(\"Concessions & specials\", 'T12 Raw Data'!$B:$R, 17, 0), 0)"
    _copy_cell_style(value_style, b180)
    cells_set += 2

    # R181 — Ratio
    a181 = ws.cell(M6_RATIO_ROW, 1)
    a181.value = "Residual / T12 Concessions (abs)"
    _copy_cell_style(label_style, a181)
    b181 = ws.cell(M6_RATIO_ROW, 2)
    b181.value = "=IFERROR(ABS(B179)/ABS(B180), 0)"
    _copy_cell_style(value_style, b181)
    b181.number_format = "0.0%"
    cells_set += 2

    # --- Conditional note (R183) — merged A:I, mirror M5 note style ---
    note_cell = ws.cell(M6_NOTE_ROW, 1)
    note_cell.value = (
        '=IF(B179>=0,'
        '"",'                                            # positive residuals are M5's
        'IF(B180=0,'
        '"⚠ Negative residual = "&TEXT(B179,"$#,##0")&" in Other LOC bucket, '
        'but T12 has no Concessions & specials line for reconciliation. '
        'Verify GL routing.",'
        'IF(ABS(B179)/ABS(B180)>' + f'{RESIDUAL_THRESHOLD},'
        '"⚠ Negative residual = "&TEXT(B179,"$#,##0")&" is "&TEXT(ABS(B179)/ABS(B180),"0.0%")&'
        '" of T12 Concessions ("&TEXT(ABS(B180),"$#,##0")&"). Likely misposted '
        'concessions — review GL routing for Other LOC credits.",'
        '"✓ Negative residual = "&TEXT(B179,"$#,##0")&" within reconciliation '
        'tolerance (≤"&TEXT(' + f'{RESIDUAL_THRESHOLD},"0%")' + '&" of T12 Concessions "&'
        'TEXT(ABS(B180),"$#,##0")&").")))'
    )
    _copy_cell_style(ws.cell(STYLE_REF_NOTE_ROW, 1), note_cell)
    try:
        ws.merge_cells(start_row=M6_NOTE_ROW, end_row=M6_NOTE_ROW,
                       start_column=1, end_column=9)
    except Exception:
        pass
    cells_set += 1

    return cells_set


# ---------------------------------------------------------------------------
# Stamping
# ---------------------------------------------------------------------------

def stamp_versions(wb: openpyxl.Workbook) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

def verify_migration(wb: openpyxl.Workbook) -> dict:
    r: dict = {}
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    ws = wb[RR_RECON]

    # Section header
    a178 = ws.cell(M6_HEADER_ROW, 1).value
    r["m6_header"] = a178
    r["m6_header_ok"] = isinstance(a178, str) and a178.strip().startswith("M6")

    # Residual row reads from B173
    b179 = ws.cell(M6_RESIDUAL_ROW, 2).value
    r["b179_ok"] = isinstance(b179, str) and "B173" in b179

    # T12 Concessions VLOOKUP
    b180 = ws.cell(M6_T12_CONC_ROW, 2).value
    r["b180_ok"] = (
        isinstance(b180, str)
        and "VLOOKUP" in b180
        and "Concessions & specials" in b180
        and "T12 Raw Data" in b180
    )

    # Ratio formula
    b181 = ws.cell(M6_RATIO_ROW, 2).value
    r["b181_ok"] = isinstance(b181, str) and "ABS(B179)" in b181 and "ABS(B180)" in b181

    # Conditional note structure
    a183 = ws.cell(M6_NOTE_ROW, 1).value
    r["a183_ok"] = (
        isinstance(a183, str)
        and "B179>=0" in a183
        and "Likely misposted" in a183
        and "reconciliation" in a183
    )

    # Merged cells on note row
    note_merged = False
    for mr in ws.merged_cells.ranges:
        if mr.min_row == M6_NOTE_ROW and mr.max_row == M6_NOTE_ROW \
                and mr.min_col == 1 and mr.max_col == 9:
            note_merged = True
            break
    r["note_merged_ok"] = note_merged

    # Header merge
    hdr_merged = False
    for mr in ws.merged_cells.ranges:
        if mr.min_row == M6_HEADER_ROW and mr.max_row == M6_HEADER_ROW \
                and mr.min_col == 1 and mr.max_col == 7:
            hdr_merged = True
            break
    r["hdr_merged_ok"] = hdr_merged

    return r


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v025(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...\n")

    print("Step A — install Section M6 on Rent Roll Recon (rows 178-183):")
    nA = install_section_m6(wb)
    print(f"  set {nA} cells (header + 3 data rows + conditional note + 2 merges)")

    stamp_versions(wb)
    print(f"\nStep B — stamped substrate version -> {SUBSTRATE_TO} ({len(ANCHOR_SHEETS)} anchors)")

    print(f"\nSaving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  1. Cover!B8 = {r['cover_b8']!r}                          : {r['cover_b8_ok']}")
    print(f"  2. All 15 AZ4 = {SUBSTRATE_TO}                            : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  3. M6 header at R{M6_HEADER_ROW} starts with 'M6'                : {r['m6_header_ok']}")
    print(f"  4. R{M6_RESIDUAL_ROW} B references B173 (M5 residual)              : {r['b179_ok']}")
    print(f"  5. R{M6_T12_CONC_ROW} B = VLOOKUP('Concessions & specials', ...)   : {r['b180_ok']}")
    print(f"  6. R{M6_RATIO_ROW} B = ABS(B179)/ABS(B180) ratio                : {r['b181_ok']}")
    print(f"  7. R{M6_NOTE_ROW} conditional note (4 branches)               : {r['a183_ok']}")
    print(f"  8. R{M6_NOTE_ROW} merged A:I                                  : {r['note_merged_ok']}")
    print(f"  9. R{M6_HEADER_ROW} merged A:G                                  : {r['hdr_merged_ok']}")

    all_ok = all([
        r["cover_b8_ok"], r["az4_all"],
        r["m6_header_ok"], r["b179_ok"], r["b180_ok"], r["b181_ok"],
        r["a183_ok"], r["note_merged_ok"], r["hdr_merged_ok"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v025.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
