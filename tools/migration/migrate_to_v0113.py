"""
migrate_to_v0113.py - Substrate template v0.1.12 -> v0.1.13

Companion to RR v1.17.0 (UW-BACKLOG BL-0003 — RR Input expansion +
Section M2/M4 per-fee capture rewrite). Two coordinated changes:

  A. Rent Roll Input new columns at AC-AG (5 cols)
     AC: Meal Plan $        (was: lumped into Other LOC $)
     AD: Scooter Fee $      (was: lumped into Other LOC $)
     AE: Housekeeping $     (was: lumped into Other LOC $)
     AF: Laundry $          (was: lumped into Other LOC $)
     AG: Pet $              (was: lumped into Other LOC $)

     Other LOC $ (col O) remains as the catchall for unmatched care
     headers (Diabetes, Misc, etc.).

  B. Total LOC $ formula at T7:T606 extended to include AC-AG:
       Old: =IFERROR(L7+M7+N7+O7,0)
       New: =IFERROR(L7+M7+N7+O7+IFERROR(AC7,0)+IFERROR(AD7,0)+
                     IFERROR(AE7,0)+IFERROR(AF7,0)+IFERROR(AG7,0),0)
     (Total Monthly Rev formula at U7:U606 already references T7 so it
     auto-picks up the extension; no change needed.)

  C. Section M (Rent Roll Recon rows 121-167) — extend M1 with a 5th
     column "RR Input Col" (col E), pre-populate it for the 5 default
     fees that now have direct RR matches, rewrite M2 and M4 to use
     INDIRECT off the new col so per-fee capture rates & implied rates
     are computed for Meal Delivery / Motorized Scooter / Second Person
     Fee / Housekeeping / Laundry. M5 SUMPRODUCT picks up the new
     attributions automatically (no formula change needed there — the
     v0.1.12 formula was forward-compatible).

     Default RR Input Col mappings (per Homestead schedule + RR v1.17.0
     col layout):
       Community Fee      → ""    (one-time event; no per-fee RR col)
       Elective Transfer  → ""    (rare event; no per-fee RR col)
       Meal Delivery      → "AC"
       Motorized Scooter  → "AD"
       Second Person Fee  → "V"   (existing v0.1.10 column)
       Housekeeping       → "AE"
       Laundry            → "AF"

     M2 eligibility denominator unified: ALL occupied beds in the
     selected period (was: occupied IL only for SP). Per user spec
     2026-05-13 — couples can occur in any care type.

  D. Stamp Cover!B8 + 13 AZ4 anchors to v0.1.12 -> v0.1.13.
  E. 12-check verification block.

Idempotent: gate checks BOTH version stamp AND that Rent Roll Input!AC4
already reads "Meal Plan $". Re-runs on a partial-state file safely
re-apply.

Usage:
    python tools/migration/migrate_to_v0113.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

SUBSTRATE_FROM = "v0.1.12"
SUBSTRATE_TO = "v0.1.13"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

RRI_SHEET = "Rent Roll Input"
RECON_SHEET = "Rent Roll Recon"

# RRI column constants
RRI_HEADER_ROW = 4
RRI_DATA_START = 7
RRI_DATA_END = 606
# New columns at v0.1.13
NEW_COLS = [
    (29, "AC", "Meal Plan\n$"),
    (30, "AD", "Scooter Fee\n$"),
    (31, "AE", "Housekeeping\n$"),
    (32, "AF", "Laundry\n$"),
    (33, "AG", "Pet\n$"),
]

# Section M (mirrors layout from migrate_to_v0112.py)
M1_FIRST_DATA = 123
M1_LAST_DATA = 131
M1_HEADER_ROW = 122
M2_FIRST_DATA = 135
M2_LAST_DATA = 143
M2_HEADER_ROW = 134
M3_FIRST_DATA = 147
M3_LAST_DATA = 155
M4_FIRST_DATA = 159
M4_LAST_DATA = 167
M4_HEADER_ROW = 158

# Default 7 fees — same order as v0.1.12. Maps fee name → RR Input Col.
# Empty string = no per-fee RR col (event-based fees: Community / Transfer).
DEFAULT_FEE_RR_COLS = {
    "Community Fee":         "",
    "Elective Transfer Fee": "",
    "Meal Delivery":         "AC",
    "Motorized Scooter Fee": "AD",
    "Second Person Fee":     "V",
    "Housekeeping":          "AE",
    "Laundry":               "AF",
}

# Total LOC $ formula extension
def total_loc_formula(r: int) -> str:
    return (
        f"=IFERROR(L{r}+M{r}+N{r}+O{r}"
        f"+IFERROR(AC{r},0)+IFERROR(AD{r},0)+IFERROR(AE{r},0)+IFERROR(AF{r},0)+IFERROR(AG{r},0),0)"
    )

OLD_TOTAL_LOC_PATTERN_PREFIX = "=IFERROR(L"

# ----- Styling -----
NAVY = "FF1F3864"
WHITE = "FFFFFFFF"
DARK_TEXT = "FF1F1F1F"
PALE_GREY = "FFF2F2F2"
PALE_YELLOW = "FFFFF2CC"
PALE_GREEN = "FFE2EFDA"

HEADER_FONT_NAVY = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_FILL_NAVY = PatternFill(fill_type="solid", fgColor=NAVY)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADER_FONT_GREY = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
HEADER_FILL_GREY = PatternFill(fill_type="solid", fgColor=PALE_GREY)

BODY_FONT = Font(name="Calibri", size=10, color=DARK_TEXT)
PASTE_FILL = PatternFill(fill_type="solid", fgColor=PALE_YELLOW)
AUTO_FILL = PatternFill(fill_type="solid", fgColor=PALE_GREEN)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# RRI cross-row references reused in Section M formulas
RRI_OCC_RANGE   = "'Rent Roll Input'!$E$7:$E$606"
RRI_PERIOD_RNG  = "'Rent Roll Input'!$S$7:$S$606"
PERIOD_SELECTOR = "$B$2"


def is_already_v0113(wb) -> bool:
    """Gate: version stamp AND Rent Roll Input!AC4 reads the new header."""
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    rri = wb[RRI_SHEET]
    ac4 = rri.cell(RRI_HEADER_ROW, 29).value
    return isinstance(ac4, str) and "Meal Plan" in ac4


# ----- A. RRI new column headers -----

def install_new_rri_headers(wb) -> int:
    rri = wb[RRI_SHEET]
    for col_idx, _letter, label in NEW_COLS:
        c = rri.cell(row=RRI_HEADER_ROW, column=col_idx, value=label)
        c.font = HEADER_FONT_NAVY
        c.fill = HEADER_FILL_NAVY
        c.alignment = HEADER_ALIGN
    return len(NEW_COLS)


# ----- B. Total LOC $ formula extension -----

def extend_total_loc_formula(wb) -> int:
    """Rewrite T7:T606 from the 4-col sum to the 9-col sum (adds AC-AG)."""
    rri = wb[RRI_SHEET]
    n = 0
    for r in range(RRI_DATA_START, RRI_DATA_END + 1):
        cell = rri.cell(row=r, column=20)  # col T
        v = cell.value
        if not isinstance(v, str) or not v.startswith(OLD_TOTAL_LOC_PATTERN_PREFIX):
            continue
        # Match expected old shape exactly to avoid clobbering customized formulas
        old = f"=IFERROR(L{r}+M{r}+N{r}+O{r},0)"
        if v == old:
            cell.value = total_loc_formula(r)
            n += 1
        elif "IFERROR(AC" in v:
            # already migrated — skip
            continue
    return n


# ----- C. Section M rewrite -----

def install_m1_rr_input_col(wb) -> int:
    """Add 5th column 'RR Input Col' to Section M1 (col E) with default per-fee mappings."""
    ws = wb[RECON_SHEET]

    # Header at E122
    c = ws.cell(M1_HEADER_ROW, 5, value="RR Input Col")
    c.font = HEADER_FONT_GREY
    c.fill = HEADER_FILL_GREY
    c.alignment = CENTER_ALIGN
    c.border = BOX

    # Pre-populate E123:E129 for the 7 default fees
    n = 0
    for offset in range(7):
        m1_row = M1_FIRST_DATA + offset
        fee_name = ws.cell(m1_row, 1).value
        if fee_name in DEFAULT_FEE_RR_COLS:
            default_col = DEFAULT_FEE_RR_COLS[fee_name]
            cell = ws.cell(m1_row, 5, value=default_col)
            cell.fill = PASTE_FILL
            cell.font = BODY_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = BOX
            n += 1
    # Style blank rows (130-131) too
    for offset in (7, 8):
        m1_row = M1_FIRST_DATA + offset
        cell = ws.cell(m1_row, 5)
        cell.fill = PASTE_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BOX
    return n


def rewrite_m2_capture(wb) -> int:
    """Universal M2 formulas: capture # via INDIRECT off M1 col E. Eligible
    # = all occupied beds in selected period (no longer IL-only for SP)."""
    ws = wb[RECON_SHEET]
    n = 0
    for offset in range(M1_LAST_DATA - M1_FIRST_DATA + 1):
        m1_row = M1_FIRST_DATA + offset
        m2_row = M2_FIRST_DATA + offset

        # A: fee name (already references M1; preserve)
        a = ws.cell(m2_row, 1)
        a.value = f"=A{m1_row}"

        # B: Eligible # — all occupied beds in selected period
        b = ws.cell(m2_row, 2)
        b.value = (
            f"=COUNTIFS({RRI_OCC_RANGE},\"Occupied\","
            f"{RRI_PERIOD_RNG},{PERIOD_SELECTOR})"
        )
        b.fill = AUTO_FILL
        b.font = BODY_FONT
        b.alignment = LEFT_ALIGN
        b.border = BOX
        b.number_format = "0"

        # C: Capturing # — INDIRECT off E{m1_row}; show "—" if unset
        c = ws.cell(m2_row, 3)
        c.value = (
            f"=IF(TRIM(E{m1_row})=\"\",\"—\","
            f"COUNTIF(INDIRECT(\"'Rent Roll Input'!\"&E{m1_row}&\"7:\"&E{m1_row}&\"606\"),\">0\"))"
        )
        c.fill = AUTO_FILL
        c.font = BODY_FONT
        c.alignment = LEFT_ALIGN
        c.border = BOX
        c.number_format = "0"

        # D: Capture %
        d = ws.cell(m2_row, 4)
        d.value = f"=IFERROR(C{m2_row}/B{m2_row},\"\")"
        d.fill = AUTO_FILL
        d.font = BODY_FONT
        d.alignment = LEFT_ALIGN
        d.border = BOX
        d.number_format = "0.0%"

        # E: Note
        e = ws.cell(m2_row, 5)
        e.value = (
            f"=IF(TRIM(E{m1_row})=\"\","
            f"\"No per-fee RR column (event-based / not captured per-resident)\","
            f"\"✓ Direct RR match (col \"&TRIM(E{m1_row})&\")\")"
        )
        e.fill = AUTO_FILL
        e.font = BODY_FONT
        e.alignment = LEFT_ALIGN
        e.border = BOX

        n += 1
    return n


def rewrite_m4_implied(wb) -> int:
    """Universal M4 formulas: implied per-resident rate (T12 monthly ÷ RR
    capture #) vs. schedule. Variance % + conditional note for any row
    with an RR Input Col set."""
    ws = wb[RECON_SHEET]
    n = 0
    for offset in range(M1_LAST_DATA - M1_FIRST_DATA + 1):
        m1_row = M1_FIRST_DATA + offset
        m2_row = M2_FIRST_DATA + offset
        m3_row = M3_FIRST_DATA + offset
        m4_row = M4_FIRST_DATA + offset

        # A: ref M1 fee name
        a = ws.cell(m4_row, 1)
        a.value = f"=A{m1_row}"

        # B: T12 $/mo (from M3 col C — same as v0.1.12)
        b = ws.cell(m4_row, 2)
        b.value = f"=C{m3_row}"
        b.fill = AUTO_FILL
        b.font = BODY_FONT
        b.alignment = LEFT_ALIGN
        b.border = BOX
        b.number_format = "$#,##0;($#,##0);\"\""

        # C: RR # capturing — pull from M2; show "—" if no RR col set
        c = ws.cell(m4_row, 3)
        c.value = (
            f"=IF(TRIM(E{m1_row})=\"\",\"—\",C{m2_row})"
        )
        c.fill = AUTO_FILL
        c.font = BODY_FONT
        c.alignment = LEFT_ALIGN
        c.border = BOX
        c.number_format = "0"

        # D: Implied $/resident — only when RR col set AND counts numeric
        d = ws.cell(m4_row, 4)
        d.value = (
            f"=IF(AND(TRIM(E{m1_row})<>\"\",ISNUMBER(B{m4_row}),ISNUMBER(C{m4_row}),C{m4_row}>0),"
            f"B{m4_row}/C{m4_row},\"\")"
        )
        d.fill = AUTO_FILL
        d.font = BODY_FONT
        d.alignment = LEFT_ALIGN
        d.border = BOX
        d.number_format = "$#,##0;($#,##0);\"\""

        # E: Schedule $ (from M1 col B)
        e = ws.cell(m4_row, 5)
        e.value = f"=B{m1_row}"
        e.fill = AUTO_FILL
        e.font = BODY_FONT
        e.alignment = LEFT_ALIGN
        e.border = BOX
        e.number_format = "$#,##0;($#,##0);\"\""

        # F: Variance %
        f = ws.cell(m4_row, 6)
        f.value = (
            f"=IF(AND(ISNUMBER(D{m4_row}),ISNUMBER(E{m4_row}),E{m4_row}<>0),"
            f"(D{m4_row}-E{m4_row})/E{m4_row},\"\")"
        )
        f.fill = AUTO_FILL
        f.font = BODY_FONT
        f.alignment = LEFT_ALIGN
        f.border = BOX
        f.number_format = "0.0%;(0.0%);\"\""

        # G: Conditional note
        g = ws.cell(m4_row, 7)
        g.value = (
            f"=IF(TRIM(E{m1_row})=\"\","
            f"\"Falls into M5 Misc. (event-based; reconcile via M3 ÷ annual turnover)\","
            f"IF(NOT(ISNUMBER(F{m4_row})),\"\","
            f"IF(ABS(F{m4_row})>0.05,"
            f"\"⚠ Implied rate differs from schedule by \"&TEXT(F{m4_row},\"0.0%\")&"
            f" \" — legacy in-place residents or schedule out of date\","
            f"\"✓ Implied rate within 5% of schedule\")))"
        )
        g.fill = AUTO_FILL
        g.font = BODY_FONT
        g.alignment = LEFT_ALIGN
        g.border = BOX

        n += 1
    return n


# ----- D. Stamp -----

def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ----- E. Verification -----

def verify_migration(wb) -> dict:
    r: dict = {}
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    rri = wb[RRI_SHEET]
    found_headers = []
    for col_idx, letter, expected in NEW_COLS:
        v = rri.cell(RRI_HEADER_ROW, col_idx).value
        if isinstance(v, str) and expected.split("\n")[0] in v:
            found_headers.append(letter)
    r["new_headers"] = found_headers
    r["new_headers_ok"] = len(found_headers) == len(NEW_COLS)

    # Total LOC $ formula extension — sample 3 rows
    extended = 0
    sampled = 0
    for sample_row in (RRI_DATA_START, RRI_DATA_START + 100, RRI_DATA_END - 1):
        v = str(rri.cell(sample_row, 20).value or "")
        if v:
            sampled += 1
            if "IFERROR(AC" in v:
                extended += 1
    r["tlc_extended"] = f"{extended}/{sampled}"
    r["tlc_extended_ok"] = extended == sampled and sampled > 0

    # Section M
    ws = wb[RECON_SHEET]
    r["m1_e_header"] = ws.cell(M1_HEADER_ROW, 5).value
    r["m1_e_header_ok"] = r["m1_e_header"] == "RR Input Col"

    # Default RR col values populated
    e_values = {
        ws.cell(M1_FIRST_DATA + i, 1).value: ws.cell(M1_FIRST_DATA + i, 5).value
        for i in range(7)
    }
    r["m1_e_defaults"] = e_values
    expected_with_col = {"Meal Delivery", "Motorized Scooter Fee", "Second Person Fee",
                         "Housekeeping", "Laundry"}
    populated = {k for k, v in e_values.items()
                 if isinstance(v, str) and v.strip() and v != ""}
    r["m1_e_5_populated"] = expected_with_col.issubset(populated)

    # M2 generic INDIRECT formula
    sp_offset = 4  # Second Person Fee is 5th in default list
    sp_m2_c = ws.cell(M2_FIRST_DATA + sp_offset, 3).value
    r["m2_c_formula"] = sp_m2_c
    r["m2_indirect_ok"] = isinstance(sp_m2_c, str) and "INDIRECT" in sp_m2_c

    # M2 eligibility unified to all-occupied
    sp_m2_b = str(ws.cell(M2_FIRST_DATA + sp_offset, 2).value or "")
    r["m2_b_unified_ok"] = (
        "Occupied" in sp_m2_b
        and '\\"IL\\"' not in sp_m2_b  # no IL filter
        and 'D$7:$D' not in sp_m2_b    # no Care Type filter
    )

    # M4 generic implied formula
    sp_m4_d = ws.cell(M4_FIRST_DATA + sp_offset, 4).value
    r["m4_d_formula"] = sp_m4_d
    r["m4_implied_ok"] = isinstance(sp_m4_d, str) and "B" in sp_m4_d and "C" in sp_m4_d

    # Sections K and L untouched
    r["section_k_intact"] = "IL " in (ws.cell(86, 1).value or "")
    r["section_l_intact"] = "MC CARE STRUCTURE" in (ws.cell(102, 1).value or "").upper()

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0113(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    n_h = install_new_rri_headers(wb)
    print(f"  A: installed {n_h} new RRI headers (AC-AG)")

    n_t = extend_total_loc_formula(wb)
    print(f"  B: extended Total LOC $ formula in {n_t} rows of col T")

    n_m1 = install_m1_rr_input_col(wb)
    print(f"  C1: installed M1 'RR Input Col' header + populated {n_m1} defaults")

    n_m2 = rewrite_m2_capture(wb)
    print(f"  C2: rewrote {n_m2} M2 rows with universal INDIRECT formulas")

    n_m4 = rewrite_m4_implied(wb)
    print(f"  C3: rewrote {n_m4} M4 rows with universal implied-rate formulas")

    stamp_versions(wb)
    print(f"  D: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:<22}     : {r['cover_b8_ok']}")
    print(f"  All 13 AZ4 = {SUBSTRATE_TO}              : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  5 new RRI headers (AC-AG)             : {r['new_headers_ok']} ({r['new_headers']})")
    print(f"  Total LOC $ formula extended (sample) : {r['tlc_extended']} ({r['tlc_extended_ok']})")
    print(f"  M1 col E header 'RR Input Col'        : {r['m1_e_header_ok']}")
    print(f"  M1 col E 5 default mappings populated : {r['m1_e_5_populated']}")
    print(f"  M2 capture formula uses INDIRECT      : {r['m2_indirect_ok']}")
    print(f"  M2 eligibility unified all-occupied   : {r['m2_b_unified_ok']}")
    print(f"  M4 implied-rate formula generic       : {r['m4_implied_ok']}")
    print(f"  Section K (IL deep-dive) intact       : {r['section_k_intact']}")
    print(f"  Section L (MC structure) intact       : {r['section_l_intact']}")

    all_ok = all([
        r["cover_b8_ok"], r["az4_all"], r["new_headers_ok"],
        r["tlc_extended_ok"], r["m1_e_header_ok"], r["m1_e_5_populated"],
        r["m2_indirect_ok"], r["m2_b_unified_ok"], r["m4_implied_ok"],
        r["section_k_intact"], r["section_l_intact"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0113.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
