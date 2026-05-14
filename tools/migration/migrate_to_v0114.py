"""
migrate_to_v0114.py - Substrate template v0.1.13 -> v0.1.14

Three small Track 3 patches surfaced in v0.1.10 carry-forwards. Bundled
into a single substrate increment per UW-BACKLOG.md prioritization:

  A. BL-0004 — T12 Analytics: 2P revenue reconciliation row.
     Compares `SUM('Rent Roll Input'!V) × 12` (RR-projected annual 2P
     revenue from per-bed entries) against `T12 Raw Data!R15` (T12
     actual annual 2P revenue). Slots into the empty rows 42-44 between
     the GPR Revenue Waterfall (ends row 41) and the Other Revenue
     Normalization Bridge (starts row 45). Conditional note fires when
     |variance| > 10%.

  B. BL-0005 — Workbook Health: total outstanding AR aggregation.
     Sums `Rent Roll Input!X` (Balance, captured at v1.16.0) into a new
     diagnostic row. Adds AR ÷ monthly EGI % as a collection-velocity
     indicator. Conditional note fires when AR > 5% of monthly EGI.
     Slots into rows 43-45 after the existing G8 'Last opened' volatile
     timestamp.

  C. BL-0006 — Rent Roll Recon Section K: PSF dispersion.
     Adds col I 'Avg Actual PSF' to the IL unit-type mix table (rows
     87-93). Sources from `Rent Roll Input!AA` (Actual PSF, captured at
     v1.16.0). Mirrors the existing AVERAGEIFS pattern from col D
     (Avg Rate).

  D. Stamp Cover!B8 + 13 AZ4 anchors to v0.1.13 -> v0.1.14.
  E. 10-check verification block.

Idempotent: gate checks BOTH the version stamp AND that the three
sentinel cells are present (T12 Analytics!A42 starts with "RR ↔ T12",
Workbook Health!A43 starts with "G9", Rent Roll Recon!I87 reads
"Avg Actual PSF"). Re-runs on a partial-state file safely re-apply.

Usage:
    python tools/migration/migrate_to_v0114.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SUBSTRATE_FROM = "v0.1.13"
SUBSTRATE_TO = "v0.1.14"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

ANALYTICS = "T12 Analytics"
RECON = "Rent Roll Recon"
HEALTH = "Workbook Health"

# Cross-sheet references used in formulas
T12RD_R15_2P = "'T12 Raw Data'!$R$15"   # 2nd Person Revenue annual total
EGI_ANNUAL   = "'Monthly Trending'!$N$21"
RRI_V_RANGE  = "'Rent Roll Input'!$V$7:$V$606"   # 2P Rent per-bed
RRI_X_RANGE  = "'Rent Roll Input'!$X$7:$X$606"   # Balance per-bed
RRI_AA_RANGE = "'Rent Roll Input'!$AA$7:$AA$606" # Actual PSF per-bed
RRI_PERIOD   = "'Rent Roll Input'!$S$7:$S$606"
RRI_STATUS   = "'Rent Roll Input'!$E$7:$E$606"
RRI_CARETYPE = "'Rent Roll Input'!$D$7:$D$606"
RRI_APTTYPE  = "'Rent Roll Input'!$F$7:$F$606"
RECON_B2     = "'Rent Roll Recon'!$B$2"

# Variance thresholds
VARIANCE_2P_THRESHOLD = 0.10      # 10% variance fires note on BL-0004
AR_PCT_EGI_THRESHOLD  = 0.05      # 5% of monthly EGI fires note on BL-0005

# ===== Styling =====
NAVY = "FF1F3864"
LIGHT_NAVY = "FF305496"
WHITE = "FFFFFFFF"
DARK_TEXT = "FF1F1F1F"
PALE_GREY = "FFF2F2F2"
PALE_GREEN = "FFE2EFDA"
HEADER_FILL = PatternFill(fill_type="solid", fgColor=PALE_GREY)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
SUBTITLE_FILL = PatternFill(fill_type="solid", fgColor=LIGHT_NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
AUTO_FILL = PatternFill(fill_type="solid", fgColor=PALE_GREEN)
BODY_FONT = Font(name="Calibri", size=10, color=DARK_TEXT)
ITALIC_FONT = Font(name="Calibri", size=9, italic=True, color="FF7F7F7F")

LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def is_already_v0114(wb) -> bool:
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    # BL-0004 sentinel: T12 Analytics!A168 has the recon section title
    ta_a168 = wb[ANALYTICS].cell(168, 1).value
    if not (isinstance(ta_a168, str) and "Reconciliation" in ta_a168):
        return False
    # BL-0005 sentinel: Workbook Health!A43 starts with "G9"
    wh_a43 = wb[HEALTH].cell(43, 1).value
    if not (isinstance(wh_a43, str) and "G9" in wh_a43):
        return False
    # BL-0006 sentinel: Rent Roll Recon!I87 reads "Avg Actual\nPSF"
    rr_i87 = wb[RECON].cell(87, 9).value
    return isinstance(rr_i87, str) and "Actual" in rr_i87 and "PSF" in rr_i87


# ===== BL-0004: T12 Analytics 2P recon =====

def install_bl0004(wb) -> int:
    """BL-0004: 2P revenue reconciliation at rows 168-171 (after the existing
    KPI Dashboard + COLOR KEY at row 166). The sections within rows 30-50
    have horizontal merges (A43:H43, A45:H45) we don't want to disrupt;
    rows 168+ are clean working space."""
    ws = wb[ANALYTICS]

    # Row 168 — subsection title (merged A:G)
    title = ws.cell(168, 1)
    title.value = "RR ↔ T12 — 2nd Person Revenue Reconciliation  (BL-0004)"
    title.fill = SUBTITLE_FILL
    title.font = SUBTITLE_FONT
    title.alignment = LEFT
    ws.merge_cells(start_row=168, end_row=168, start_column=1, end_column=7)

    # Row 169 — column headers
    headers = [
        (1, "Metric"),
        (2, "RR projection\n(annual)"),
        (3, "T12 actual\n(annual)"),
        (4, "Variance %"),
        (5, "Note"),
    ]
    for col, label in headers:
        c = ws.cell(169, col, value=label)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BOX
    # Merge note column E:G
    ws.merge_cells(start_row=169, end_row=169, start_column=5, end_column=7)

    # Row 170 — data
    a170 = ws.cell(170, 1, value="2P revenue\n(Σ Rent Roll Input!V × 12  vs.  T12 Raw Data!R15)")
    a170.fill = AUTO_FILL; a170.font = BODY_FONT; a170.alignment = LEFT; a170.border = BOX

    b170 = ws.cell(170, 2, value=f"=SUM({RRI_V_RANGE})*12")
    b170.fill = AUTO_FILL; b170.font = BODY_FONT; b170.alignment = RIGHT; b170.border = BOX
    b170.number_format = "$#,##0;($#,##0);\"\""

    c170 = ws.cell(170, 3, value=f"=IFERROR({T12RD_R15_2P},0)")
    c170.fill = AUTO_FILL; c170.font = BODY_FONT; c170.alignment = RIGHT; c170.border = BOX
    c170.number_format = "$#,##0;($#,##0);\"\""

    d170 = ws.cell(170, 4, value=f"=IFERROR((B170-C170)/C170,\"\")")
    d170.fill = AUTO_FILL; d170.font = BODY_FONT; d170.alignment = RIGHT; d170.border = BOX
    d170.number_format = "0.0%;(0.0%);\"\""

    # Note (E170, spans E:G)
    e170 = ws.cell(170, 5, value=(
        f"=IF(NOT(ISNUMBER(D170)),\"\","
        f"IF(ABS(D170)>{VARIANCE_2P_THRESHOLD},"
        f"\"⚠ 2P revenue gap of \"&TEXT(D170,\"0.0%\")&"
        f"\" — verify SP column matches operator's published 2P rate\","
        f"\"✓ 2P revenue reconciles within \"&TEXT({VARIANCE_2P_THRESHOLD},\"0%\")&\" of T12\"))"
    ))
    e170.fill = AUTO_FILL; e170.font = ITALIC_FONT; e170.alignment = LEFT; e170.border = BOX
    ws.merge_cells(start_row=170, end_row=170, start_column=5, end_column=7)

    return 3  # 1 title row + 1 header row + 1 data row


# ===== BL-0005: Workbook Health AR aggregation =====

def install_bl0005(wb) -> int:
    ws = wb[HEALTH]

    # Row 43 — G9 Total outstanding AR
    a43 = ws.cell(43, 1, value="G9 · Total outstanding AR  (Σ Rent Roll Input!X)")
    a43.font = BODY_FONT; a43.alignment = LEFT
    b43 = ws.cell(43, 2, value=f"=SUM({RRI_X_RANGE})")
    b43.font = BODY_FONT; b43.alignment = RIGHT
    b43.number_format = "$#,##0;($#,##0);\"\""

    # Row 44 — G10 AR / monthly EGI %
    a44 = ws.cell(44, 1, value="G10 · AR ÷ monthly EGI  (collection-velocity indicator)")
    a44.font = BODY_FONT; a44.alignment = LEFT
    b44 = ws.cell(44, 2, value=f"=IFERROR(B43/({EGI_ANNUAL}/12),0)")
    b44.font = BODY_FONT; b44.alignment = RIGHT
    b44.number_format = "0.0%;(0.0%);\"\""

    # Row 45 — conditional note
    note = ws.cell(45, 1, value=(
        f"=IF(B43=0,\"\","
        f"IF(B44>{AR_PCT_EGI_THRESHOLD},"
        f"\"⚠ AR is \"&TEXT(B44,\"0.0%\")&\" of monthly EGI — collection risk; review aging\","
        f"\"✓ AR within \"&TEXT({AR_PCT_EGI_THRESHOLD},\"0%\")&\" of monthly EGI\"))"
    ))
    note.font = ITALIC_FONT; note.alignment = LEFT
    ws.merge_cells(start_row=45, end_row=45, start_column=1, end_column=4)

    return 3  # G9, G10, note


# ===== BL-0006: Section K Avg Actual PSF =====

def _psf_per_unit_type_formula(unit_type: str) -> str:
    return (
        f"=IFERROR(AVERAGEIFS({RRI_AA_RANGE},{RRI_PERIOD},{RECON_B2},"
        f"{RRI_STATUS},\"<>Vacant\",{RRI_STATUS},\"<>Eviction\","
        f"{RRI_CARETYPE},\"IL\",{RRI_APTTYPE},\"{unit_type}\"),\"-\")"
    )


def _psf_total_il_formula() -> str:
    return (
        f"=IFERROR(AVERAGEIFS({RRI_AA_RANGE},{RRI_PERIOD},{RECON_B2},"
        f"{RRI_STATUS},\"<>Vacant\",{RRI_STATUS},\"<>Eviction\","
        f"{RRI_CARETYPE},\"IL\"),\"-\")"
    )


def install_bl0006(wb) -> int:
    ws = wb[RECON]

    # I87 — header
    h = ws.cell(87, 9, value="Avg Actual\nPSF")
    h.fill = HEADER_FILL; h.font = HEADER_FONT; h.alignment = CENTER; h.border = BOX

    # I88-I92 — per unit type
    UNIT_TYPES = ["Studio", "1 Bedroom", "2 Bedroom", "Cottage / Villa", "Other"]
    for offset, ut in enumerate(UNIT_TYPES):
        r = 88 + offset
        c = ws.cell(r, 9, value=_psf_per_unit_type_formula(ut))
        c.font = BODY_FONT; c.alignment = RIGHT; c.border = BOX
        c.number_format = "$#,##0.00;($#,##0.00);\"-\""

    # I93 — Total IL
    t = ws.cell(93, 9, value=_psf_total_il_formula())
    t.font = BODY_FONT; t.alignment = RIGHT; t.border = BOX
    t.number_format = "$#,##0.00;($#,##0.00);\"-\""

    return 7  # 1 header + 5 unit types + 1 total


# ===== Versioning =====

def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ===== Verify =====

def verify_migration(wb) -> dict:
    r: dict = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # BL-0004 — rows 168 (title) / 169 (headers) / 170 (data)
    ws_ta = wb[ANALYTICS]
    r["bl4_title"] = ws_ta.cell(168, 1).value
    r["bl4_title_ok"] = isinstance(r["bl4_title"], str) and "Reconciliation" in r["bl4_title"]
    rr_proj = str(ws_ta.cell(170, 2).value or "")
    r["bl4_rr_proj_ok"] = "SUM('Rent Roll Input'!$V$7:$V$606)*12" in rr_proj
    t12_actual = str(ws_ta.cell(170, 3).value or "")
    r["bl4_t12_actual_ok"] = "T12 Raw Data" in t12_actual and "$R$15" in t12_actual

    # BL-0005
    ws_wh = wb[HEALTH]
    r["bl5_g9_label"] = ws_wh.cell(43, 1).value
    r["bl5_g9_ok"] = isinstance(r["bl5_g9_label"], str) and "G9" in r["bl5_g9_label"]
    ar_formula = str(ws_wh.cell(43, 2).value or "")
    r["bl5_ar_formula_ok"] = "SUM('Rent Roll Input'!$X$7:$X$606)" in ar_formula
    g10_label = ws_wh.cell(44, 1).value
    r["bl5_g10_ok"] = isinstance(g10_label, str) and "G10" in g10_label

    # BL-0006
    ws_rr = wb[RECON]
    i87 = ws_rr.cell(87, 9).value
    r["bl6_i87"] = i87
    # Header is "Avg Actual\nPSF" — check for both tokens to allow the embedded newline
    r["bl6_i87_ok"] = isinstance(i87, str) and "Actual" in i87 and "PSF" in i87
    i88 = str(ws_rr.cell(88, 9).value or "")
    r["bl6_i88_ok"] = "AVERAGEIFS" in i88 and "$AA" in i88 and '"Studio"' in i88
    i93 = str(ws_rr.cell(93, 9).value or "")
    r["bl6_i93_ok"] = "AVERAGEIFS" in i93 and "$AA" in i93 and '"IL"' in i93

    # Sanity: prior substrate sections still intact
    r["section_k_unit_table_intact"] = ws_rr.cell(87, 1).value == "Unit Type"
    r["section_l_intact"] = "MC CARE STRUCTURE" in (ws_rr.cell(102, 1).value or "").upper()

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0114(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    n4 = install_bl0004(wb)
    print(f"  A: BL-0004 T12 Analytics 2P recon — {n4} cells")
    n5 = install_bl0005(wb)
    print(f"  B: BL-0005 Workbook Health AR aggregation — {n5} cells")
    n6 = install_bl0006(wb)
    print(f"  C: BL-0006 Section K Avg Actual PSF column — {n6} cells")

    stamp_versions(wb)
    print(f"  D: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:<22}     : {r['cover_b8_ok']}")
    print(f"  All 13 AZ4 = {SUBSTRATE_TO}              : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  BL-0004 title row 42                   : {r['bl4_title_ok']}")
    print(f"  BL-0004 RR projection formula          : {r['bl4_rr_proj_ok']}")
    print(f"  BL-0004 T12 actual formula             : {r['bl4_t12_actual_ok']}")
    print(f"  BL-0005 G9 label                       : {r['bl5_g9_ok']}")
    print(f"  BL-0005 AR SUM formula                 : {r['bl5_ar_formula_ok']}")
    print(f"  BL-0005 G10 label                      : {r['bl5_g10_ok']}")
    print(f"  BL-0006 I87 header 'Avg Actual PSF'    : {r['bl6_i87_ok']}")
    print(f"  BL-0006 I88 Studio AVERAGEIFS          : {r['bl6_i88_ok']}")
    print(f"  BL-0006 I93 Total IL AVERAGEIFS        : {r['bl6_i93_ok']}")
    print(f"  Section K unit-type table intact       : {r['section_k_unit_table_intact']}")
    print(f"  Section L (MC) intact                  : {r['section_l_intact']}")

    all_ok = all([
        r["cover_b8_ok"], r["az4_all"],
        r["bl4_title_ok"], r["bl4_rr_proj_ok"], r["bl4_t12_actual_ok"],
        r["bl5_g9_ok"], r["bl5_ar_formula_ok"], r["bl5_g10_ok"],
        r["bl6_i87_ok"], r["bl6_i88_ok"], r["bl6_i93_ok"],
        r["section_k_unit_table_intact"], r["section_l_intact"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0114.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
