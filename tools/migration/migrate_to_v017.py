"""
migrate_to_v017.py — Substrate template v0.1.6 → v0.1.7

Companion to T12 code v0.2.0 (BrokerFinancialSummaryFormat + Cluster B). All
edits are additive or single-cell formula replacements; no row inserts (so no
openpyxl formula-shift dance). Idempotent — re-running on a v0.1.7 file is a
no-op.

What this script does:

  A. Deferred bug fixes (logged in v0.1.6 changelog as "Out of this round")
    A-1  T12 Analytics!E102 (`Lease / ground lease`) — replace `=0` with
         INDEX/MATCH against T12 Raw Data Label column. Fills R61 of UW Output
         with real lease data when source has it.
    A-2  T12 Analytics!F102 — set to `=E102` matching the sibling pattern.
    A-3  T12 Raw Data SUMIFS — sweep all 636 cells in F:Q replacing
         `T12_Calc!$X$1:$X$501` with `$X$1:$X$500`. Cosmetic; T12_Calc has
         500 data rows so 501 reads empty either way.

  B. Workbook Health additions (Cluster B carry-forward surfaced in workbook)
    B-1  Add V8 row "T12 month coverage" to the Workbook Health Validation
         section. Counts populated months in T12 Input C11:N11; ✓ on 12, ⚠
         otherwise.

  C. Description_Map additions (Homestead vocabulary derived in Phase 4)
    C-1  Append 99 prefixed entries derived from the populated_analyzer's
         per-banner mappings. Resolves UNMATCHED on Homestead and March_2026
         broker fixtures end-to-end.

  D. Version stamps
    D-1  Cover!B8 → v0.1.7
    D-2  All 13 sheets' AZ4 → v0.1.7

  E. Verification block: 13 boolean checks; reports pass/fail per check.

Usage:
    python tools/migration/migrate_to_v017.py input.xlsx output.xlsx
"""
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Tuple

import openpyxl
from openpyxl.styles import Font

SUBSTRATE_FROM = "v0.1.6"
SUBSTRATE_TO = "v0.1.7"

# ----------------------------------------------------------------------------
# Sheets that need version-stamp bumps via AZ4
# ----------------------------------------------------------------------------
ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)


# ----------------------------------------------------------------------------
# C-1: Homestead vocabulary additions (derived 2026-05-08; see SPEC-T12.md
# §"Template substrate" v0.1.7 entry). 99 prefixed entries; suffix-Label
# mappings inherited from the populated_analyzer's v0.1.5 Option-C work.
# ----------------------------------------------------------------------------
HOMESTEAD_DESCMAP_ADDITIONS: Tuple[Tuple[str, str], ...] = (
    ('Activities | Contract - Activities', 'Contract / agency labor'),
    ('Activities | Other', 'Other / miscellaneous'),
    ('Activities | Payroll - Holiday', 'Overtime wages'),
    ('Activities | Payroll - Overtime', 'Overtime wages'),
    ('Activities | Payroll - PTO', 'PTO wages'),
    ('Activities | Payroll - Wages', 'Activities labor'),
    ('Activities | Supplies - Activities', 'Recreation & activity suppl.'),
    ('Care Level Revenue | Assisted Living', 'LOC revenue — AL'),
    ('Dietary | Other', 'Other / miscellaneous'),
    ('Dietary | Payroll - Holiday', 'Overtime wages'),
    ('Dietary | Payroll - Overtime', 'Overtime wages'),
    ('Dietary | Payroll - PTO', 'PTO wages'),
    ('Dietary | Payroll - Wages', 'Dining / food service labor'),
    ('Dietary | Raw Food', 'Food cost'),
    ('Dietary | Supplies', 'Dining & kitchen supplies'),
    ('Direct Care | Equipment - Rentals', 'R&M variable'),
    ('Direct Care | Nursing Supplies - General', 'Nursing & care supplies'),
    ('Direct Care | Other', 'Other / miscellaneous'),
    ('Direct Care | Payroll - Holiday', 'Overtime wages'),
    ('Direct Care | Payroll - Overtime', 'Overtime wages'),
    ('Direct Care | Payroll - PTO', 'PTO wages'),
    ('Direct Care | Payroll - Wages', 'Care staff labor'),
    ('Employee Benefits | Bonus Program (403b replace)', 'Employee 401(k)'),
    ('Employee Benefits | Health Insurance', 'Employee benefits'),
    ('Employee Benefits | Other', 'Other / miscellaneous'),
    ('Employee Benefits | Payroll Taxes', 'Payroll taxes'),
    ('Employee Benefits | Training & Orientation', 'Office, admin & G&A'),
    ('Employee Benefits | Workers Compensation', "Workers' comp insurance"),
    ('General & Administrative | Bank Fees', 'Office, admin & G&A'),
    ('General & Administrative | Data Processing', 'Office, admin & G&A'),
    ('General & Administrative | Dues & Subscriptions', 'Permits, licenses & dues'),
    ('General & Administrative | Fines & Penalties', 'Office, admin & G&A'),
    ('General & Administrative | IT & Computers', 'Telephone / IT'),
    ('General & Administrative | Licenses & Permits', 'Permits, licenses & dues'),
    ('General & Administrative | Motor Vehicles', 'Auto Expense'),
    ('General & Administrative | Office Equipment', 'Office, admin & G&A'),
    ('General & Administrative | Office Supplies', 'Office, admin & G&A'),
    ('General & Administrative | Office Telecom', 'Telephone / IT'),
    ('General & Administrative | Other', 'Other / miscellaneous'),
    ('General & Administrative | Payroll - Bonus', 'Bonus wages'),
    ('General & Administrative | Payroll - Overtime', 'Overtime wages'),
    ('General & Administrative | Payroll - PTO', 'PTO wages'),
    ('General & Administrative | Payroll - Wages', 'Administrative labor'),
    ('General & Administrative | Postage', 'Office, admin & G&A'),
    ('General & Administrative | Professional Fees - Legal', 'Legal expenses'),
    ('General & Administrative | Recruiting/Hiring Costs', 'Office, admin & G&A'),
    ('General & Administrative | Travel & Meals', 'Office, admin & G&A'),
    ('Housekeeping & Laundry | Payroll - Holiday', 'Overtime wages'),
    ('Housekeeping & Laundry | Payroll - Overtime', 'Overtime wages'),
    ('Housekeeping & Laundry | Payroll - PTO', 'PTO wages'),
    ('Housekeeping & Laundry | Payroll - Wages', 'Maint. & housekeeping labor'),
    ('Housekeeping & Laundry | Supplies', 'HK & laundry supplies'),
    ('Maintenance | Equipment', 'R&M variable'),
    ('Maintenance | Extermination', 'Pest elimination'),
    ('Maintenance | Landscaping', 'R&M variable'),
    ('Maintenance | Payroll - Holiday', 'Overtime wages'),
    ('Maintenance | Payroll - Overtime', 'Overtime wages'),
    ('Maintenance | Payroll - PTO', 'PTO wages'),
    ('Maintenance | Payroll - Wages', 'Maint. & housekeeping labor'),
    ('Maintenance | R&M - Contract Service', 'R&M fixed'),
    ('Maintenance | R&M - General/Other', 'R&M variable'),
    ('Maintenance | R&M - Supplies', 'R&M variable'),
    ('Maintenance | Trash Removal', 'Utilities'),
    ('Management Fee & Bad Debt | Bad Debt', 'Bad debt expense'),
    ('Management Fee & Bad Debt | Management Fee', 'Management fee'),
    ('Marketing | Advertising', 'Sales, adv. & marketing'),
    ('Marketing | Executive Spend', 'Sales, adv. & marketing'),
    ('Marketing | Marketing Events', 'Sales, adv. & marketing'),
    ('Marketing | Marketing Supplies', 'Sales, adv. & marketing'),
    ('Marketing | Other', 'Other / miscellaneous'),
    ('Marketing | Payroll - Bonus', 'Bonus wages'),
    ('Marketing | Payroll - Wages', 'Administrative labor'),
    ('Marketing | Referral Fees', 'Referral fees'),
    ('Nursing Administration | Payroll - Bonus', 'Bonus wages'),
    ('Nursing Administration | Payroll - Holiday', 'Overtime wages'),
    ('Nursing Administration | Payroll - Overtime', 'Overtime wages'),
    ('Nursing Administration | Payroll - PTO', 'PTO wages'),
    ('Nursing Administration | Payroll - Wages', 'Wellness / care coordinators'),
    ('Other Income', 'Other community revenue'),
    ('Pharmacy | Contract Service - Pharmacy', 'Nursing & care supplies'),
    ('Respite Revenue', 'Respite care'),
    ('Room & Board Revenue | Assisted Living', 'Base rent — AL'),
    ('Room & Board Revenue | Independent Living', 'Base rent — IL'),
    ('Room & Board Revenue | Memory Care', 'Base rent — MC'),
    ('Second Persons Revenue | Assisted Living', 'Base rent — AL'),
    ('Second Persons Revenue | Independent Living', 'Base rent — IL'),
    ('Second Persons Revenue | Memory Care', 'Base rent — MC'),
    ('Taxes & Insurance | Insurance - GLPL', 'P&C insurance (bundled)'),
    ('Taxes & Insurance | Insurance - Other', 'P&C insurance (bundled)'),
    ('Taxes & Insurance | Insurance - Property', 'P&C insurance (bundled)'),
    ('Taxes & Insurance | Tax - Personal Property', 'Personal property taxes'),
    ('Taxes & Insurance | Tax - Real Property', 'Real estate taxes'),
    ('Taxes & Insurance | Tax - Sales & Use', 'Office, admin & G&A'),
    ('Utilities | Cable TV', 'Telephone / IT'),
    ('Utilities | Electric', 'Utilities'),
    ('Utilities | Gas', 'Utilities'),
    ('Utilities | Internet', 'Telephone / IT'),
    ('Utilities | Telephone', 'Telephone / IT'),
    ('Utilities | Water & Sewer', 'Utilities'),
)


# ============================================================================
# Idempotency guard
# ============================================================================

def is_already_v017(wb) -> bool:
    if "Cover" in wb.sheetnames and wb["Cover"]["B8"].value == SUBSTRATE_TO:
        return True
    return False


# ============================================================================
# A — T12 Analytics R102 + T12 Raw Data N501 sweep
# ============================================================================

LEASE_FORMULA = (
    '=IFERROR(INDEX(\'T12 Raw Data\'!R:R,'
    'MATCH("Lease / ground lease",\'T12 Raw Data\'!B:B,0)),0)'
)


def apply_lease_formula(wb) -> None:
    """A-1, A-2: T12 Analytics row 102 lease formula + sibling F102."""
    ta = wb["T12 Analytics"]
    # Idempotent: only replace if currently `=0`. Don't clobber custom formulas.
    if ta["E102"].value in (0, "=0", None):
        ta["E102"] = LEASE_FORMULA
    if ta["F102"].value in (0, "=0", None):
        ta["F102"] = "=E102"


N501_PATTERN = re.compile(r"\$1:\$([A-Z]+)\$501")


def sweep_n501_to_n500(wb) -> int:
    """A-3: walk all formula cells in T12 Raw Data and replace any
    `T12_Calc!$X$1:$X$501` reference with `$X$1:$X$500`. Returns count."""
    trd = wb["T12 Raw Data"]
    count = 0
    for r in range(1, trd.max_row + 1):
        for c in range(1, trd.max_column + 1):
            cell = trd.cell(r, c)
            if not isinstance(cell.value, str):
                continue
            if "$501" not in cell.value:
                continue
            new_val = N501_PATTERN.sub(r"$1:$\1$500", cell.value)
            if new_val != cell.value:
                cell.value = new_val
                count += 1
    return count


# ============================================================================
# B — Workbook Health: V8 partial-year T12 row
# ============================================================================

def add_partial_year_validation(wb) -> None:
    """B-1: add V8 'T12 month coverage' to Validation section.

    Validation block layout from migrate_to_v016:
      row 21: section header '2 · VALIDATION'
      row 22: column header 'Check / Result / Status'
      rows 23-29: V1-V7
      row 30: blank gutter (re-purposed here for V8)
      row 31: section header '3 · DIAGNOSTICS'
    """
    wh = wb["Workbook Health"]
    # Idempotency: skip if V8 already there
    if wh.cell(30, 1).value and "T12 month coverage" in str(wh.cell(30, 1).value):
        return
    wh.cell(30, 1, "V8 · T12 month coverage")
    # Count populated month-label cells in T12 Input C11:N11. Returns 12 when
    # full year, lower for partial-year T12s (broker files, mid-year deals).
    wh.cell(30, 2, '=COUNTA(\'T12 Input\'!C11:N11)')
    wh.cell(30, 3, '=IF(B30=12,"✓","⚠")')


# ============================================================================
# C — Description_Map appends
# ============================================================================

def append_descmap_entries(wb) -> int:
    """C-1: append 99 prefixed Homestead vocabulary entries. Idempotent —
    skips entries whose key (col A) already exists."""
    ws = wb["Description_Map"]
    existing_keys = set()
    last_row = 0
    for r in range(5, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None:
            s = str(v).strip()
            existing_keys.add(s)
            last_row = r
    if last_row == 0:
        last_row = 4  # write to row 5 first

    appended = 0
    next_row = last_row + 1
    for desc, label in HOMESTEAD_DESCMAP_ADDITIONS:
        if desc in existing_keys:
            continue
        ws.cell(next_row, 1, desc)
        ws.cell(next_row, 2, label)
        existing_keys.add(desc)
        next_row += 1
        appended += 1
    return appended


# ============================================================================
# D — Version stamps
# ============================================================================

def stamp_versions(wb) -> None:
    """D-1, D-2: bump Cover!B8 and all 13 sheets' AZ4 to SUBSTRATE_TO."""
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ============================================================================
# Verification
# ============================================================================

def verify_migration(wb) -> dict:
    results = {}

    # Cover B8 stamped
    results["cover_b8"] = wb["Cover"]["B8"].value if "Cover" in wb.sheetnames else None

    # All 13 anchor AZ4 cells == v0.1.7
    az4 = {}
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            az4[s] = wb[s]["AZ4"].value
    results["az4_per_sheet"] = az4
    results["all_az4_v017"] = all(v == SUBSTRATE_TO for v in az4.values())

    # T12 Analytics E102 holds the lease formula (not =0)
    e102 = wb["T12 Analytics"]["E102"].value or ""
    results["e102_resolved"] = "INDEX" in e102 and "Lease / ground lease" in e102
    results["f102_sibling"] = wb["T12 Analytics"]["F102"].value == "=E102"

    # T12 Raw Data: no remaining $501 refs
    trd = wb["T12 Raw Data"]
    n501_remaining = 0
    for r in range(1, trd.max_row + 1):
        for c in range(1, trd.max_column + 1):
            v = trd.cell(r, c).value
            if isinstance(v, str) and "$501" in v:
                n501_remaining += 1
    results["n501_remaining"] = n501_remaining
    results["n501_swept_clean"] = n501_remaining == 0

    # Workbook Health V8 row exists and is the partial-year coverage check
    wh = wb["Workbook Health"]
    v8_label = wh.cell(30, 1).value
    results["v8_present"] = (
        v8_label is not None and "T12 month coverage" in str(v8_label)
    )

    # Description_Map: count of HOMESTEAD_DESCMAP_ADDITIONS keys present
    dm = wb["Description_Map"]
    keys = set()
    for r in range(5, dm.max_row + 1):
        v = dm.cell(r, 1).value
        if v is not None:
            keys.add(str(v).strip())
    target_descs = {desc for desc, _ in HOMESTEAD_DESCMAP_ADDITIONS}
    present = len(target_descs & keys)
    results["homestead_entries_present"] = present
    results["all_homestead_entries"] = present == len(target_descs)

    return results


# ============================================================================
# Main
# ============================================================================

def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v017(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    # A — formula fixes
    apply_lease_formula(wb)
    n501_swept = sweep_n501_to_n500(wb)
    print(f"  A-3: swept {n501_swept} $501 -> $500 references in T12 Raw Data")

    # B — Workbook Health
    add_partial_year_validation(wb)
    print("  B-1: added V8 partial-year validation row")

    # C — Description_Map appends
    appended = append_descmap_entries(wb)
    print(f"  C-1: appended {appended} Description_Map entries (Homestead vocab)")

    # D — version stamps
    stamp_versions(wb)
    print(f"  D: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    # Reload and verify
    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    results = verify_migration(wb2)
    print()
    print("=== Verification ===")
    print(f"  Cover!B8                  : {results['cover_b8']}")
    print(f"  All 13 AZ4 == {SUBSTRATE_TO}        : {results['all_az4_v017']}")
    print(f"  E102 lease formula        : {results['e102_resolved']}")
    print(f"  F102 = =E102              : {results['f102_sibling']}")
    print(f"  T12 Raw Data $501 remain  : {results['n501_remaining']} (expect 0)")
    print(f"  V8 partial-year row       : {results['v8_present']}")
    print(
        f"  Homestead descmap entries : "
        f"{results['homestead_entries_present']}/{len(HOMESTEAD_DESCMAP_ADDITIONS)}"
    )

    all_ok = (
        results["cover_b8"] == SUBSTRATE_TO
        and results["all_az4_v017"]
        and results["e102_resolved"]
        and results["f102_sibling"]
        and results["n501_swept_clean"]
        and results["v8_present"]
        and results["all_homestead_entries"]
    )
    print()
    print(
        "=== "
        + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete")
        + " ==="
    )
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v017.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
