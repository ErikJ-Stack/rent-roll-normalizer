"""
migrate_to_v028.py — Substrate template v0.2.7 → v0.2.8

Closes UW-BACKLOG **BL-0022** — `Cover!B5` doesn't auto-resolve the property name.

Background:
  RR v1.15.0 + T12 v0.2.1 writers auto-stamp the property name into
  `Rent Roll Input!A3` and `T12 Input!A10`. `T12 Analytics!B2` resolves
  via a 3-priority chain (RR → T12 → `Property_Name = Cover!B5`). But
  `Cover!B5` itself was left as the bottom-priority manual-entry cell —
  so when path 1 (RR) wins on T12 Analytics, Cover itself stays blank.
  Consumers of `Cover!B5` directly (`Dashboard!B2` title formula,
  `UW Export!B3`, `Workbook Health!B27`, Pre-Export Gate `B49`) all
  render "missing" / "(not set)" / empty even though the writer-stamped
  inputs are present and T12 Analytics resolves correctly.

  v0.2.8 rewrites `Cover!B5` to a 2-priority resolver formula:
    =IFERROR(IF(LEN(TRIM('Rent Roll Input'!A3))>0,'Rent Roll Input'!A3,
            IF(LEN(TRIM('T12 Input'!A10))>0,'T12 Input'!A10,"")),"")
  — no fallback to `Property_Name` since that would be circular
  (`Property_Name → Cover!B5`).

  Defensive skip: if `Cover!B5` already contains static (non-formula)
  text at migration time, the rewrite is skipped and the user's typed
  value is preserved. This handles the case where someone has manually
  set the property name and doesn't want it auto-overwritten.

  `Cover!A19` docstring also updated — old text described B5 as a
  manual-entry cell ("Property name entered at B5 above..."), which is
  now inaccurate.

OPERATIONS:

  A. Rewrite Cover!B5 = 2-priority property-name resolver (skip if user text)
  B. Rewrite Cover!A19 docstring (preserve user customization if present)
  C. Stamp Cover!B8 = v0.2.8
  D. Stamp 15 AZ4 anchors = v0.2.8

Idempotent: gate checks `Cover!B8 == "v0.2.8"`. Re-running is a no-op.

Usage:
    python tools/migration/migrate_to_v028.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

SUBSTRATE_FROM = "v0.2.7"
SUBSTRATE_TO = "v0.2.8"

# 15-sheet anchor list — same as v0.2.7 (no sheet adds/removes in v0.2.8).
ANCHOR_SHEETS = (
    "Cover", "Dashboard",
    "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# --- A: Cover!B5 property-name resolver ------------------------------------

NEW_B5_FORMULA = (
    '=IFERROR('
    'IF(LEN(TRIM(\'Rent Roll Input\'!A3))>0,\'Rent Roll Input\'!A3,'
    'IF(LEN(TRIM(\'T12 Input\'!A10))>0,\'T12 Input\'!A10,'
    '""'
    '))'
    ',"")'
)

# --- B: Cover!A19 docstring update -----------------------------------------

# Variants observed across substrate versions; treat any as safe to overwrite.
OLD_A19_TEXTS = (
    "Property name entered at B5 above propagates to T12 Analytics via the Property_Name named range.",
)
NEW_A19_TEXT = (
    "Property name at B5 auto-resolves from Rent Roll Input!A3 → T12 Input!A10 "
    "(writer-stamped). Type into B5 to manually override. Propagates to all "
    "consumers via the Property_Name named range."
)


def is_already_v028(wb) -> bool:
    return wb["Cover"]["B8"].value == SUBSTRATE_TO


def looks_like_user_text(value) -> bool:
    """True if Cover!B5 contains a static string (not a formula and not empty).
    Respect user-typed values; don't overwrite with our formula."""
    if value is None:
        return False
    if isinstance(value, str):
        return not value.lstrip().startswith("=")
    # Number / date / other types — treat as user data
    return True


def rewrite_cover_b5(wb) -> str:
    """A: Cover!B5 resolver. Returns 'wrote' or 'skipped_user_text'."""
    b5 = wb["Cover"]["B5"]
    if looks_like_user_text(b5.value):
        return "skipped_user_text"
    b5.value = NEW_B5_FORMULA
    return "wrote"


def rewrite_cover_a19(wb) -> str:
    """B: Cover!A19 docstring. Returns 'wrote' or 'skipped_user_text'."""
    a19 = wb["Cover"]["A19"]
    v = a19.value
    if v is None or (isinstance(v, str) and (v in OLD_A19_TEXTS or v == NEW_A19_TEXT)):
        a19.value = NEW_A19_TEXT
        return "wrote"
    return "skipped_user_text"


def stamp_versions(wb) -> int:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    stamped = 0
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO
            stamped += 1
    return stamped


def verify_migration(wb) -> dict:
    r = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # A: Cover!B5 either has the new formula OR is user-typed static text
    b5 = wb["Cover"]["B5"].value
    r["b5_formula"] = b5 == NEW_B5_FORMULA
    r["b5_user_text"] = (
        isinstance(b5, str) and not b5.lstrip().startswith("=") and b5.strip() != ""
    ) or isinstance(b5, (int, float))
    r["b5_valid_post_state"] = r["b5_formula"] or r["b5_user_text"]

    # B: Cover!A19 either has new docstring or user-customized non-stale text
    a19 = wb["Cover"]["A19"].value
    r["a19_updated"] = a19 == NEW_A19_TEXT or (
        isinstance(a19, str) and a19 not in OLD_A19_TEXTS and a19 is not None
    )

    # Sanity: existing structures preserved
    # M5 + M6 on Rent Roll Recon (BL-0012 from v0.2.5) should be intact
    rrr = wb["Rent Roll Recon"]
    r["m5_intact"] = str(rrr["A169"].value or "").lstrip().startswith("M5")
    r["m6_intact"] = str(rrr["A178"].value or "").lstrip().startswith("M6")
    # T12 Analytics!B2 3-priority resolver (BL untouched here)
    b2 = str(wb["T12 Analytics"]["B2"].value or "")
    r["t12analytics_b2_intact"] = (
        "Rent Roll Input" in b2 and "T12 Input" in b2 and "Property_Name" in b2
    )
    # Dashboard sheet at index 1 (BL-0018 from v0.2.7)
    r["dashboard_at_index_1"] = (
        len(wb.sheetnames) > 1 and wb.sheetnames[1] == "Dashboard"
    )

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v028(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    b5_status = rewrite_cover_b5(wb)
    print(f"  A: Cover!B5 resolver — {b5_status}")

    a19_status = rewrite_cover_a19(wb)
    print(f"  B: Cover!A19 docstring — {a19_status}")

    n_stamp = stamp_versions(wb)
    print(f"  C+D: Cover!B8 + {n_stamp} AZ4 anchors -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:24s}    : {r['cover_b8_ok']}")
    print(f"  All AZ4 anchors = {SUBSTRATE_TO}            : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  Cover!B5 valid (formula or user text) : {r['b5_valid_post_state']}")
    print(f"    - formula injected                  : {r['b5_formula']}")
    print(f"    - user-typed text preserved         : {r['b5_user_text']}")
    print(f"  Cover!A19 docstring updated/preserved : {r['a19_updated']}")
    print(f"  M5 (R169) intact                      : {r['m5_intact']}")
    print(f"  M6 (R178) intact                      : {r['m6_intact']}")
    print(f"  T12 Analytics!B2 3-priority intact    : {r['t12analytics_b2_intact']}")
    print(f"  Dashboard at sheetnames index 1       : {r['dashboard_at_index_1']}")

    all_ok = (
        r["cover_b8_ok"] and r["az4_all"]
        and r["b5_valid_post_state"] and r["a19_updated"]
        and r["m5_intact"] and r["m6_intact"]
        and r["t12analytics_b2_intact"] and r["dashboard_at_index_1"]
    )
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v028.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
