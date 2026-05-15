"""
migrate_to_v023.py — Substrate template v0.2.2 → v0.2.3

Realigns Rent Roll Recon row 16 — "RR gross contracted base rent / mo" —
with the intent already documented in column H ("Gross contracted rates
before concessions"). Old formula at B16/C16/D16 summed Actual Rate
(`'Rent Roll Input'!$H`) over occupied units only — producing
"current contracted at actual rate" rather than the Gross Potential Rent
at 100% occupancy that the row's role as the underwriting anchor demands.

  Old B16/C16/D16:  SUMIFS($H, period, E<>Vacant, E<>Eviction, D=<care>)
                    → Actual Rate × occupied (Homestead: $565,140)

  New B16/C16/D16:  SUMIFS($G, period, D=<care>)
                    → Market Rate × all units, 100% occupancy
                       (Homestead: $809,567)

Row 17 (effective net after concessions) is unchanged — its formula
`SUMIFS($H, occupied) + SUMIFS($I, occupied)` is already correct because
concessions are negative-signed (per SPEC-RR.md L184). Rows 18-20 chain
off row 17 and need no change.

This was previously implemented as substrate v0.1.11 in PR #12 (closed
unmerged after main moved 23 commits to v0.2.2). The fix is identical;
this script targets the v0.2.2 → v0.2.3 transition with the current
14-sheet anchor list (UW Export was added in v0.2.0).

Closes UW-BACKLOG BL-0015.

OPERATIONS:

  A. Rewrite B16, C16, D16 with new SUMIFS($G, ...) formulas
  B. Rewrite A16 label to "RR Gross Potential Rent / mo (Market × all units)"
  C. Rewrite H16 note to reflect GPR semantics
  D. Stamp Cover!B8 + 14 AZ4 anchors to v0.2.3
  E. 9-check verification

Idempotent: gate checks BOTH version stamp AND that B16 references col G
(not col H), so re-running on a partial-state file safely re-applies.

Usage:
    python tools/migration/migrate_to_v023.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

SUBSTRATE_FROM = "v0.2.2"
SUBSTRATE_TO = "v0.2.3"

# v0.2.0 added "UW Export" sheet (BL-0009 / Branch 2 Handoff readiness).
ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# Row 16 — new formulas (GPR = Market Rate × all units in period, by care type)
NEW_B16 = (
    '=IFERROR(SUMIFS(\'Rent Roll Input\'!$G$7:$G$606,'
    '\'Rent Roll Input\'!$S$7:$S$606,\'Rent Roll Recon\'!$B$2,'
    '\'Rent Roll Input\'!$D$7:$D$606,"IL"),0)'
)
NEW_C16 = (
    '=IFERROR(SUMIFS(\'Rent Roll Input\'!$G$7:$G$606,'
    '\'Rent Roll Input\'!$S$7:$S$606,\'Rent Roll Recon\'!$B$2,'
    '\'Rent Roll Input\'!$D$7:$D$606,"AL"),0)'
)
NEW_D16 = (
    '=IFERROR(SUMIFS(\'Rent Roll Input\'!$G$7:$G$606,'
    '\'Rent Roll Input\'!$S$7:$S$606,\'Rent Roll Recon\'!$B$2,'
    '\'Rent Roll Input\'!$D$7:$D$606,"MC"),0)'
)

NEW_A16 = "RR Gross Potential Rent / mo  (Market × all units)"
NEW_H16 = (
    "Gross Potential Rent — Market Rate × all units at 100% occupancy. "
    "Excludes concessions & vacancy loss. Row 16 − Row 17 = vacancy + "
    "market-vs-actual gap."
)


def is_already_v023(wb) -> bool:
    """Gate: version stamp AND B16 formula references col G (not col H)."""
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    rrr = wb["Rent Roll Recon"]
    b16 = str(rrr["B16"].value or "")
    return "$G$7:$G$606" in b16


def update_row_16(wb) -> dict:
    """A + B + C: rewrite row 16 formulas, label, and note."""
    rrr = wb["Rent Roll Recon"]
    n = {"formulas": 0, "label": 0, "note": 0}

    rrr["B16"].value = NEW_B16
    rrr["C16"].value = NEW_C16
    rrr["D16"].value = NEW_D16
    n["formulas"] = 3

    # E16 stays =SUM(B16:D16) — already correct, no rewrite needed.

    rrr["A16"].value = NEW_A16
    n["label"] = 1

    rrr["H16"].value = NEW_H16
    n["note"] = 1

    return n


def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


def verify_migration(wb) -> dict:
    r = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    rrr = wb["Rent Roll Recon"]

    # Formula checks — all three must sum $G (not $H) and key off the care code
    b16 = str(rrr["B16"].value or "")
    c16 = str(rrr["C16"].value or "")
    d16 = str(rrr["D16"].value or "")

    r["b16_market"] = "$G$7:$G$606" in b16 and '"IL"' in b16
    r["c16_market"] = "$G$7:$G$606" in c16 and '"AL"' in c16
    r["d16_market"] = "$G$7:$G$606" in d16 and '"MC"' in d16
    r["row16_no_status_filter"] = (
        "Vacant" not in b16 and "Eviction" not in b16
        and "Vacant" not in c16 and "Eviction" not in c16
        and "Vacant" not in d16 and "Eviction" not in d16
    )

    # E16 sum still wires up the three care subtotals
    e16 = str(rrr["E16"].value or "")
    r["e16_sum_ok"] = e16 == "=SUM(B16:D16)"

    # Label + note updated
    a16 = str(rrr["A16"].value or "")
    h16 = str(rrr["H16"].value or "")
    r["a16_label_ok"] = "Gross Potential Rent" in a16
    r["h16_note_ok"] = "Gross Potential Rent" in h16 and "100%" in h16

    # Sanity: row 17 was NOT touched (still references col H and col I, still
    # has the status filter). If this trips, the migration accidentally
    # rewrote the wrong row.
    b17 = str(rrr["B17"].value or "")
    r["row17_intact"] = (
        "$H$7:$H$606" in b17
        and "$I$7:$I$606" in b17
        and "Vacant" in b17
    )

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v023(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    n = update_row_16(wb)
    print(f"  A+B+C: row 16 — {n['formulas']} formulas, "
          f"{n['label']} label, {n['note']} note updated")

    stamp_versions(wb)
    print(f"  D: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:24s}    : {r['cover_b8_ok']}")
    print(f"  All 14 AZ4 = {SUBSTRATE_TO}              : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  B16 sums $G + care=IL                 : {r['b16_market']}")
    print(f"  C16 sums $G + care=AL                 : {r['c16_market']}")
    print(f"  D16 sums $G + care=MC                 : {r['d16_market']}")
    print(f"  Row 16 has no Vacant/Eviction filter  : {r['row16_no_status_filter']}")
    print(f"  E16 = SUM(B16:D16) (unchanged)        : {r['e16_sum_ok']}")
    print(f"  A16 label updated                     : {r['a16_label_ok']}")
    print(f"  H16 note updated                      : {r['h16_note_ok']}")
    print(f"  Row 17 untouched (still $H + $I + filter): {r['row17_intact']}")

    all_ok = (
        r["cover_b8_ok"] and r["az4_all"]
        and r["b16_market"] and r["c16_market"] and r["d16_market"]
        and r["row16_no_status_filter"]
        and r["e16_sum_ok"]
        and r["a16_label_ok"] and r["h16_note_ok"]
        and r["row17_intact"]
    )
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v023.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
