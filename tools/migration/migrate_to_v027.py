"""
migrate_to_v027.py — Substrate template v0.2.6 → v0.2.7

Closes UW-BACKLOG **BL-0018** (Dashboard sheet redesign).

Replaces the v0.2.4 "Investment Dashboard" with a redesigned "Dashboard"
sheet authored by the user in Excel. The new Dashboard is denser (437
cells vs 340), visually richer (6 native Excel charts vs 0), and uses a
17-column layout instead of the 52-column Investment Dashboard layout.

Like Investment Dashboard, the new Dashboard is **purely referential**:
every value cell is either a label, a static text string, or a formula
referencing T12 Analytics / Rent Roll Recon / Monthly Trending / Cover.
96 unique cross-sheet refs total; 95 resolve to populated cells on a
v0.2.6 baseline; the 1 outlier is `Cover!B5` (Property Name) which is
user-populated at runtime via the `Property_Name` named range.

Why a template asset rather than programmatic construction:
  - Same reason as v0.2.4: 437 styled cells + 6 Excel charts +
    72 merged ranges + custom tab color. Encoding all of that as Python
    style objects would balloon this script. The Dashboard is captured
    once into `v027_assets/dashboard_template.xlsx` and lifted into the
    real workbook at migration time.

Why deepcopy charts instead of cell-by-cell:
  - The v0.2.4 _copy_cell helper preserves font/fill/border/alignment/
    number_format/protection but NOT charts. openpyxl Chart objects can
    be deep-copied and re-attached to a different worksheet without the
    series references breaking, since series formulas like
    `Dashboard!$O$2:$O$5` are string-typed and travel with the chart.

OPERATIONS:

  A. Remove the existing "Investment Dashboard" sheet (the v0.2.4
     predecessor — fully superseded by the new Dashboard).
  B. Insert "Dashboard" worksheet at index 1, copying cells + styles +
     column widths + row heights + merged cells + 6 charts + tab color
     from `v027_assets/dashboard_template.xlsx`.
  C. Stamp AZ1:AZ5 anchor metadata on the new sheet.
  D. Stamp Cover!B8 → v0.2.7 + AZ4 on all 15 sheets (anchor list
     re-rolls: "Investment Dashboard" → "Dashboard").

Idempotency: gate checks BOTH `Cover!B8 == "v0.2.7"` AND `"Dashboard"`
at index 1 AND `"Investment Dashboard"` absent. Re-running on an
already-migrated workbook is a no-op (just re-saves). Partial pre-state
(e.g. Dashboard inserted but Investment Dashboard not yet removed) is
handled by the per-step has_X / has_Y guards.

Usage:
    python tools/migration/migrate_to_v027.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from copy import copy, deepcopy
from pathlib import Path

import openpyxl

SUBSTRATE_FROM = "v0.2.6"
SUBSTRATE_TO = "v0.2.7"

NEW_SHEET = "Dashboard"
OLD_SHEET = "Investment Dashboard"
NEW_SHEET_INDEX = 1  # immediately after Cover

# 15-sheet anchor list — same shape as v0.2.6 but Investment Dashboard
# is renamed to Dashboard.
ANCHOR_SHEETS = (
    "Cover", "Dashboard",
    "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

ANCHOR_PURPOSE = "Underwriting at-a-glance KPI dashboard with embedded charts"
ANCHOR_CATEGORY = "Analytical (handoff)"
ANCHOR_VISIBILITY = "visible"
ANCHOR_NOTES = (
    "All value cells are formula references into T12 Analytics, Rent Roll Recon, "
    "Monthly Trending, and Cover. 6 native Excel charts embedded. No source-of-"
    "truth data lives here. Supersedes the v0.2.4-v0.2.6 Investment Dashboard."
)

TEMPLATE_PATH = (
    Path(__file__).parent / "v027_assets" / "dashboard_template.xlsx"
)


def is_already_v027(wb) -> bool:
    """Gate: version stamp + new sheet at expected position + old sheet gone."""
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    if NEW_SHEET not in wb.sheetnames:
        return False
    if wb.sheetnames.index(NEW_SHEET) != NEW_SHEET_INDEX:
        return False
    if OLD_SHEET in wb.sheetnames:
        return False
    return True


def _copy_cell(src_cell, dst_cell) -> None:
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def remove_old_dashboard(wb) -> bool:
    if OLD_SHEET in wb.sheetnames:
        del wb[OLD_SHEET]
        return True
    return False


def insert_dashboard(wb) -> dict:
    """Insert the Dashboard sheet from the template at index 1."""
    n = {"cells": 0, "col_widths": 0, "row_heights": 0, "merges": 0, "charts": 0}

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Template asset missing: {TEMPLATE_PATH}. "
            "This migration cannot proceed without the bundled dashboard template."
        )

    tmpl_wb = openpyxl.load_workbook(TEMPLATE_PATH)
    src_ws = tmpl_wb[NEW_SHEET]

    dst_ws = wb.create_sheet(NEW_SHEET)

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column)
            _copy_cell(cell, new_cell)
            if cell.value is not None:
                n["cells"] += 1

    for letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[letter].width = dim.width
            n["col_widths"] += 1
        if dim.hidden:
            dst_ws.column_dimensions[letter].hidden = dim.hidden

    for r, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[r].height = dim.height
            n["row_heights"] += 1

    for mr in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(mr))
        n["merges"] += 1

    if src_ws.sheet_view.showGridLines is False:
        dst_ws.sheet_view.showGridLines = False

    # Tab color — Dashboard sheet has a navy tab color FF1F4E79.
    if src_ws.sheet_properties.tabColor is not None:
        dst_ws.sheet_properties.tabColor = copy(src_ws.sheet_properties.tabColor)

    # Charts — deepcopy each. Anchor info travels with the chart object;
    # series references are string formulas (e.g. "Dashboard!$O$2:$O$5") so
    # they stay valid as long as the sheet retains its name.
    for chart in src_ws._charts:
        new_chart = deepcopy(chart)
        dst_ws._charts.append(new_chart)
        n["charts"] += 1

    # Move to index 1 (immediately after Cover).
    current_index = wb.sheetnames.index(NEW_SHEET)
    offset = NEW_SHEET_INDEX - current_index
    if offset != 0:
        wb.move_sheet(NEW_SHEET, offset=offset)

    return n


def stamp_anchor_cells(wb) -> None:
    ws = wb[NEW_SHEET]
    ws["AZ1"] = ANCHOR_PURPOSE
    ws["AZ2"] = ANCHOR_CATEGORY
    ws["AZ3"] = ANCHOR_VISIBILITY
    ws["AZ4"] = SUBSTRATE_TO
    ws["AZ5"] = ANCHOR_NOTES


def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


def verify_migration(wb) -> dict:
    r = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    r["new_sheet_exists"] = NEW_SHEET in wb.sheetnames
    r["new_sheet_position_ok"] = (
        r["new_sheet_exists"]
        and wb.sheetnames.index(NEW_SHEET) == NEW_SHEET_INDEX
    )
    r["old_sheet_removed"] = OLD_SHEET not in wb.sheetnames
    r["sheet_count"] = len(wb.sheetnames)
    r["sheet_count_ok"] = r["sheet_count"] == 15

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    if r["new_sheet_exists"]:
        ws = wb[NEW_SHEET]
        r["sheet_dimensions"] = ws.dimensions
        r["sheet_dimensions_ok"] = ws.max_row >= 90 and ws.max_column >= 15
        r["chart_count"] = len(ws._charts)
        r["chart_count_ok"] = r["chart_count"] == 6

        # Spot-check key cells
        b2 = str(ws["B2"].value or "")
        b6 = str(ws["B6"].value or "")
        b5 = str(ws["B5"].value or "")
        r["b2_title_ok"] = "UNDERWRITING DASHBOARD" in b2.upper()
        r["b6_kpi_formula_ok"] = b6.startswith("=") and "T12 Analytics" in b6
        r["b5_label_ok"] = "OCCUPANCY" in b5.upper()

        # AZ1:AZ5 anchor metadata
        r["az1_purpose_ok"] = ws["AZ1"].value == ANCHOR_PURPOSE
        r["az3_visibility_ok"] = ws["AZ3"].value == ANCHOR_VISIBILITY
        r["az4_self_stamp_ok"] = ws["AZ4"].value == SUBSTRATE_TO
    else:
        r["sheet_dimensions"] = None
        r["sheet_dimensions_ok"] = False
        r["chart_count"] = 0
        r["chart_count_ok"] = False
        r["b2_title_ok"] = False
        r["b6_kpi_formula_ok"] = False
        r["b5_label_ok"] = False
        r["az1_purpose_ok"] = False
        r["az3_visibility_ok"] = False
        r["az4_self_stamp_ok"] = False

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v027(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    removed = remove_old_dashboard(wb)
    if removed:
        print(f"  A: removed '{OLD_SHEET}' sheet")
    else:
        print(f"  A: '{OLD_SHEET}' already absent")

    if NEW_SHEET not in wb.sheetnames:
        n = insert_dashboard(wb)
        print(f"  B: inserted '{NEW_SHEET}' at index {NEW_SHEET_INDEX} — "
              f"{n['cells']} cells, {n['col_widths']} col widths, "
              f"{n['row_heights']} row heights, {n['merges']} merges, "
              f"{n['charts']} charts")
    else:
        print(f"  B: '{NEW_SHEET}' already present, skipping sheet copy")
        if wb.sheetnames.index(NEW_SHEET) != NEW_SHEET_INDEX:
            current_index = wb.sheetnames.index(NEW_SHEET)
            wb.move_sheet(NEW_SHEET, offset=NEW_SHEET_INDEX - current_index)
            print(f"     repositioned to index {NEW_SHEET_INDEX}")

    stamp_anchor_cells(wb)
    print(f"  C: stamped AZ1:AZ5 on '{NEW_SHEET}'")

    stamp_versions(wb)
    print(f"  D: stamped substrate version -> {SUBSTRATE_TO} on "
          f"Cover!B8 + {len(ANCHOR_SHEETS)} AZ4 anchors")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:24s}    : {r['cover_b8_ok']}")
    print(f"  '{NEW_SHEET}' sheet exists                 : {r['new_sheet_exists']}")
    print(f"  Sheet at position {NEW_SHEET_INDEX} (after Cover)        : {r['new_sheet_position_ok']}")
    print(f"  '{OLD_SHEET}' removed       : {r['old_sheet_removed']}")
    print(f"  Sheet count = {r['sheet_count']} (expected 15)            : {r['sheet_count_ok']}")
    print(f"  Sheet dimensions = {r['sheet_dimensions']!s:10s}      : {r['sheet_dimensions_ok']}")
    print(f"  Chart count = {r['chart_count']} (expected 6)            : {r['chart_count_ok']}")
    print(f"  B2 title 'UNDERWRITING DASHBOARD'          : {r['b2_title_ok']}")
    print(f"  B5 'OCCUPANCY' label                       : {r['b5_label_ok']}")
    print(f"  B6 KPI cell formula into T12 Analytics     : {r['b6_kpi_formula_ok']}")
    print(f"  AZ1 purpose stamped                        : {r['az1_purpose_ok']}")
    print(f"  AZ3 visibility stamped                     : {r['az3_visibility_ok']}")
    print(f"  AZ4 self-stamp = {SUBSTRATE_TO}                  : {r['az4_self_stamp_ok']}")
    print(f"  All 15 AZ4 = {SUBSTRATE_TO}                     : {r['az4_all']} ({r['az4_count']} sheets)")

    all_ok = (
        r["cover_b8_ok"]
        and r["new_sheet_exists"] and r["new_sheet_position_ok"]
        and r["old_sheet_removed"]
        and r["sheet_count_ok"]
        and r["sheet_dimensions_ok"] and r["chart_count_ok"]
        and r["b2_title_ok"] and r["b5_label_ok"] and r["b6_kpi_formula_ok"]
        and r["az1_purpose_ok"] and r["az3_visibility_ok"]
        and r["az4_self_stamp_ok"] and r["az4_all"]
    )
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v027.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
