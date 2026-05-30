"""
migrate_to_v0216.py — Substrate template v0.2.15 → v0.2.16

Closes UW-BACKLOG BL-0028 — T12 Analytics missing an "Auto Expense"
non-labor row.

THE BUG
-------
`T12 Raw Data!B63` carries the GL label "Auto Expense" (Description_Map
routes "Auto Expenses" / "Auto and Mileage Expense" / "Bus/Shuttle Service"
/ "General & Administrative | Motor Vehicles" → "Auto Expense"), so the
GL dollars aggregate into the raw sheet. But `T12 Analytics` Section 3's
non-labor block (A79:A102, summed at E103=`=SUM(E79:E102)`) has NO Auto
Expense line — only "Auto insurance" at A91. So the Auto Expense dollars
fall out of the non-labor total, and EBITDARM/EBITDAR/EBITDA (and the
UW Output / UW Export layers that mirror them) OVERSTATE NOI by exactly
the Auto Expense amount.

On Homestead that gap is $6,061.32 — the precise standardized-vs-
as-reported NOI delta ($1,417,385 standardized vs $1,411,324 as-reported).
The Python engine (`dashboard_model._LABELS_NON_LABOR`, UWT v0.7.0)
already folds Auto Expense in; this migration brings the Excel substrate
into agreement. After the fix, standardized NOI ties to as-reported.

Monthly Trending ALREADY carries an "Auto Expense" row (A59, right after
"Auto insurance" A58) — confirming the intended design includes it; only
T12 Analytics missed it. No Monthly Trending / UW Output change is needed:
UW Output's "Total non-labor" (E62) is a *pull* from `'T12 Analytics'!E103`
(not a sum of its own lines), so the shift sweep re-points it to the
corrected total automatically.

THE FIX
-------
Insert ONE row in T12 Analytics at row 92 (immediately after "Auto
insurance" at A91), mirroring the A91 row:
    A92 = "Auto Expense"
    E92 = =IFERROR(INDEX('T12 Raw Data'!R:R,MATCH("Auto Expense",'T12 Raw Data'!B:B,0)),0)
    F92 = =E92
    G92 = =F92-E92
Then run the full-workbook row-shift sweep for every formula referencing
T12 Analytics rows >= 92, +1. This:
  - extends the non-labor SUM endpoint E103=`=SUM(E79:E102)` →
    (now at E104) `=SUM(E79:E103)`, which captures BOTH the new Auto
    Expense row (92) and all the shifted real rows (93-103); the new row
    sits INSIDE the original range, so the endpoint auto-extend is exact
    (sidesteps the BL-0001 endpoint-trap — see note below);
  - bumps the EBITDA chain (E105→E106 `=E76+E104`, E108→E109 `=E52-E106`,
    E110→E111, E113→E114, …) and every Section 4/5 ratio that reads them;
  - re-points the 187 cross-sheet single-cell refs into T12 Analytics
    (Dashboard ×152, UW Output ×34, AR & Collections ×1).

BL-0001 endpoint-trap note: a verified pre-flight scan confirmed there
are NO cross-sheet *range* refs into T12 Analytics (all 187 external refs
are single cells) and NO chart series reference T12 Analytics. The only
within-sheet range crossing row 92 is the non-labor SUM we WANT to extend.
So the classic qualified-range-endpoint drift does not apply here.

Idempotency: gate checks BOTH the version stamp (Cover!B8 == v0.2.16)
AND the sentinel (T12 Analytics!A92 == "Auto Expense"). Re-running is
a no-op.

Usage:
    python tools/migration/migrate_to_v0216.py input.xlsx output.xlsx
"""
from __future__ import annotations

import re
import sys
from copy import copy
from pathlib import Path
from typing import Dict

import openpyxl


SUBSTRATE_FROM = "v0.2.15"
SUBSTRATE_TO = "v0.2.16"

# Full 16-sheet anchor list (stamped at AZ4 on every sheet).
ANCHOR_SHEETS = (
    "Cover", "Dashboard", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending",
    "AR & Collections", "UW Output", "UW Export", "Mapping Review",
    "Description_Map", "RR_Calc", "T12_Calc", "Workbook Health",
)

SHEET = "T12 Analytics"
INSERT_AT_ROW = 92          # new "Auto Expense" lands here
TEMPLATE_ROW = 91           # "Auto insurance" — the row we mirror
NEW_LABEL = "Auto Expense"
NEW_MATCH_LABEL = "Auto Expense"   # the T12 Raw Data!B label to MATCH
TEMPLATE_MATCH_LABEL = "Auto insurance"
COL_LABEL = 1               # A
COL_FIRST = 5               # E (INDEX/MATCH)
COL_LAST = 7                # G (=F-E)


# ---------------------------------------------------------------------------
# Formula row-shift utility (copied verbatim from migrate_to_v021.py)
# ---------------------------------------------------------------------------

def shift_row_refs_in_formula(
    formula: str,
    threshold: int,
    delta: int,
    target_sheet: str,
    same_sheet: bool,
) -> str:
    """Increment every row reference in `formula` by `delta` if the row >=
    `threshold` AND the reference points at `target_sheet`."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    norm_target = target_sheet.lower()
    out = formula

    qualified_pattern = re.compile(
        r"('([^']+)'|([A-Za-z_][A-Za-z0-9_ ]*))!(\$?)([A-Z]+)(\$?)(\d+)"
    )

    def replace_qualified(m: re.Match) -> str:
        sheet_quoted = m.group(2)
        sheet_unquoted = m.group(3)
        sheet = sheet_quoted if sheet_quoted is not None else sheet_unquoted
        col_dollar = m.group(4)
        col = m.group(5)
        row_dollar = m.group(6)
        row_num = int(m.group(7))

        if sheet.lower() != norm_target:
            return m.group(0)
        if row_num < threshold:
            return m.group(0)

        new_row = row_num + delta
        sheet_qual = f"'{sheet}'" if sheet_quoted is not None else sheet
        return f"{sheet_qual}!{col_dollar}{col}{row_dollar}{new_row}"

    out = qualified_pattern.sub(replace_qualified, out)

    if same_sheet:
        unqualified_pattern = re.compile(
            r"(?<![A-Za-z_!])(\$?)([A-Z]+)(\$?)(\d+)\b"
        )

        def replace_unqualified(m: re.Match) -> str:
            col_dollar = m.group(1)
            col = m.group(2)
            row_dollar = m.group(3)
            row_num = int(m.group(4))
            if row_num < threshold:
                return m.group(0)
            return f"{col_dollar}{col}{row_dollar}{row_num + delta}"

        out = unqualified_pattern.sub(replace_unqualified, out)

    return out


def shift_merged_cells(ws, threshold: int, delta: int) -> int:
    """Shift merged-cell range definitions to keep up with insert_rows.
    Critical: do NOT use unmerge_cells (wipes displaced content)."""
    shifted = 0
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= threshold:
            mr.shift(col_shift=0, row_shift=delta)
            shifted += 1
    return shifted


def shift_all_formulas(
    wb: openpyxl.Workbook,
    target_sheet: str,
    threshold: int,
    delta: int,
) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        modified = 0
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                same_sheet = (sheet_name == target_sheet)
                new_v = shift_row_refs_in_formula(
                    v, threshold, delta, target_sheet, same_sheet,
                )
                if new_v != v:
                    ws.cell(cell.row, cell.column, new_v)
                    modified += 1
        if modified > 0:
            counts[sheet_name] = modified
    return counts


def copy_row_formatting(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(src_row, col)
        dst_cell = ws.cell(dst_row, col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.border = copy(src_cell.border)


# ---------------------------------------------------------------------------
# Idempotency gate
# ---------------------------------------------------------------------------

def is_already_v0216(wb: openpyxl.Workbook) -> bool:
    """Idempotent: both the version stamp AND the sentinel must agree."""
    cover_ok = False
    try:
        cover_ok = (wb["Cover"]["B8"].value == SUBSTRATE_TO)
    except Exception:
        pass

    sentinel_ok = False
    try:
        v = wb[SHEET].cell(INSERT_AT_ROW, COL_LABEL).value
        sentinel_ok = (v == NEW_LABEL)
    except Exception:
        pass

    return cover_ok and sentinel_ok


# ---------------------------------------------------------------------------
# Step — insert the Auto Expense row + shift + populate
# ---------------------------------------------------------------------------

def step_insert_auto_expense(wb: openpyxl.Workbook) -> None:
    ws = wb[SHEET]

    template_label = ws.cell(TEMPLATE_ROW, COL_LABEL).value
    print(f"  {SHEET}: template = row {TEMPLATE_ROW} ({template_label!r})")
    print(f"  {SHEET}: inserting 1 row at row {INSERT_AT_ROW}")
    ws.insert_rows(INSERT_AT_ROW, amount=1)

    n_merges = shift_merged_cells(ws, INSERT_AT_ROW, +1)
    print(f"  {SHEET}: shifted {n_merges} merged-cell range(s)")

    counts = shift_all_formulas(wb, SHEET, INSERT_AT_ROW, +1)
    total = sum(counts.values())
    print(f"  {SHEET}: shifted formula refs in {total} cells: {counts}")

    # Capture the template (Auto insurance, row 91) formulas POST-shift.
    # Row 91 < threshold so it is untouched by the sweep — but we read it
    # after the shift on principle (BL-0001 discipline).
    template_formulas: Dict[int, str] = {}
    for c in range(COL_FIRST, COL_LAST + 1):
        f = ws.cell(TEMPLATE_ROW, c).value
        if isinstance(f, str) and f.startswith("="):
            template_formulas[c] = f

    # Populate the new row by mirroring the template row.
    ws.cell(INSERT_AT_ROW, COL_LABEL, NEW_LABEL)
    copy_row_formatting(ws, TEMPLATE_ROW, INSERT_AT_ROW)

    for c, template in template_formulas.items():
        new_formula = template.replace(
            f'"{TEMPLATE_MATCH_LABEL}"', f'"{NEW_MATCH_LABEL}"'
        )
        # Bump bare self-refs from the template row to the new row
        # (F91 -> F92, etc.) — only matters for cols F/G.
        new_formula = re.sub(
            rf"(?<![A-Za-z_!])([A-Z]+){TEMPLATE_ROW}\b",
            lambda m: f"{m.group(1)}{INSERT_AT_ROW}",
            new_formula,
        )
        ws.cell(INSERT_AT_ROW, c, new_formula)
    print(f"    row {INSERT_AT_ROW}: {NEW_LABEL!r} ({len(template_formulas)} formulas)")


# ---------------------------------------------------------------------------
# Stamping
# ---------------------------------------------------------------------------

def stamp_versions(wb: openpyxl.Workbook) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

def verify_migration(wb: openpyxl.Workbook) -> dict:
    r: dict = {}
    ws = wb[SHEET]

    # 1. Cover!B8 stamp
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    # 2. All 16 AZ4 anchors
    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # 3. New row label at A92 = "Auto Expense"
    r["a92"] = ws.cell(92, 1).value
    r["a92_ok"] = r["a92"] == NEW_LABEL

    # 4. Row above (A91) still "Auto insurance"
    r["a91"] = ws.cell(91, 1).value
    r["a91_ok"] = r["a91"] == "Auto insurance"

    # 5. Row below shifted: A93 = "Fire / security monitoring"
    r["a93"] = ws.cell(93, 1).value
    r["a93_ok"] = r["a93"] == "Fire / security monitoring"

    # 6. E92 INDEX/MATCH references "Auto Expense"
    e92 = ws.cell(92, 5).value
    r["e92"] = e92
    r["e92_ok"] = (
        isinstance(e92, str)
        and '"Auto Expense"' in e92
        and "INDEX" in e92 and "MATCH" in e92
    )

    # 7. F92 = "=E92", G92 = "=F92-E92"
    r["f92"] = ws.cell(92, 6).value
    r["g92"] = ws.cell(92, 7).value
    r["f92_ok"] = r["f92"] == "=E92"
    r["g92_ok"] = r["g92"] == "=F92-E92"

    # 8. Non-labor total shifted to A104 and SUM endpoint extended to E103.
    r["a104"] = ws.cell(104, 1).value
    r["a104_ok"] = (str(r["a104"]).strip() == "Total non-labor opex")
    r["e104"] = ws.cell(104, 5).value
    r["e104_ok"] = r["e104"] == "=SUM(E79:E103)"
    r["f104"] = ws.cell(104, 6).value
    r["f104_ok"] = r["f104"] == "=SUM(F79:F103)"

    # 9. EBITDA chain shifted + re-pointed.
    r["e106"] = ws.cell(106, 5).value          # Total Op Ex (excl mgmt)
    r["e106_ok"] = r["e106"] == "=E76+E104"
    r["e109"] = ws.cell(109, 5).value          # EBITDARM
    r["e109_ok"] = r["e109"] == "=E52-E106"
    r["a109"] = ws.cell(109, 1).value
    r["a109_ok"] = r["a109"] == "EBITDARM"
    r["e111"] = ws.cell(111, 5).value          # EBITDAR
    r["e111_ok"] = r["e111"] == "=E109-E107"
    r["e114"] = ws.cell(114, 5).value          # EBITDA
    r["e114_ok"] = r["e114"] == "=E111-E113"
    r["a114"] = ws.cell(114, 1).value
    r["a114_ok"] = r["a114"] == "EBITDA"

    # 10. UW Output "Total non-labor" pull re-pointed E103 -> E104.
    uo = wb["UW Output"]
    r["uw_e62"] = uo.cell(62, 5).value
    r["uw_e62_ok"] = r["uw_e62"] == "='T12 Analytics'!E104"
    # And the EBITDARM mirror re-pointed E108 -> E109.
    r["uw_e66"] = uo.cell(66, 5).value
    r["uw_e66_ok"] = r["uw_e66"] == "='T12 Analytics'!E109"

    # 11. Dashboard headline EBITDARM-margin ref re-pointed (was F162).
    #     E162 = =IFERROR(E108/E52,...) shifted to E163 referencing E109.
    ta_e163 = ws.cell(163, 5).value
    r["ta_e163"] = ta_e163
    r["ta_e163_ok"] = isinstance(ta_e163, str) and "E109" in ta_e163 and "E52" in ta_e163

    # 12. Sheet count unchanged at 16.
    r["sheet_count"] = len(wb.sheetnames)
    r["sheet_count_ok"] = r["sheet_count"] == 16

    return r


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0216(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...\n")

    print("Step A — T12 Analytics: insert Auto Expense row + shift refs + populate:")
    step_insert_auto_expense(wb)
    print()

    stamp_versions(wb)
    print(f"Step B — stamped substrate version -> {SUBSTRATE_TO} ({len(ANCHOR_SHEETS)} anchors)")

    print(f"\nSaving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"   1. Cover!B8 = {r['cover_b8']!r:<10}                                  : {r['cover_b8_ok']}")
    print(f"   2. All {r['az4_count']} AZ4 = {SUBSTRATE_TO}                               : {r['az4_all']}")
    print(f"   3. A92 = 'Auto Expense'                                       : {r['a92_ok']} ({r['a92']!r})")
    print(f"   4. A91 = 'Auto insurance' (unchanged above)                   : {r['a91_ok']}")
    print(f"   5. A93 = 'Fire / security monitoring' (shifted)               : {r['a93_ok']}")
    print(f"   6. E92 INDEX/MATCH refers to 'Auto Expense'                   : {r['e92_ok']}")
    print(f"   7. F92 = '=E92' and G92 = '=F92-E92'                          : {r['f92_ok'] and r['g92_ok']}")
    print(f"   8. Non-labor total at A104, E104 = '=SUM(E79:E103)'           : {r['a104_ok'] and r['e104_ok'] and r['f104_ok']}")
    print(f"        E104 = {r['e104']!r}")
    print(f"   9. EBITDA chain: E106='=E76+E104', E109(EBITDARM)='=E52-E106' : {r['e106_ok'] and r['e109_ok'] and r['a109_ok']}")
    print(f"        E111(EBITDAR)='=E109-E107', E114(EBITDA)='=E111-E113'    : {r['e111_ok'] and r['e114_ok'] and r['a114_ok']}")
    print(f"  10. UW Output E62='='T12 Analytics'!E104', E66->E109          : {r['uw_e62_ok'] and r['uw_e66_ok']}")
    print(f"        UW Output E62 = {r['uw_e62']!r}")
    print(f"  11. Dashboard EBITDARM-margin source (T12A E163) -> E109/E52   : {r['ta_e163_ok']}")
    print(f"  12. Sheet count unchanged at 16                               : {r['sheet_count_ok']} ({r['sheet_count']})")

    all_ok = all([
        r["cover_b8_ok"], r["az4_all"], r["a92_ok"], r["a91_ok"], r["a93_ok"],
        r["e92_ok"], r["f92_ok"], r["g92_ok"],
        r["a104_ok"], r["e104_ok"], r["f104_ok"],
        r["e106_ok"], r["e109_ok"], r["a109_ok"],
        r["e111_ok"], r["e114_ok"], r["a114_ok"],
        r["uw_e62_ok"], r["uw_e66_ok"], r["ta_e163_ok"], r["sheet_count_ok"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0216.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
