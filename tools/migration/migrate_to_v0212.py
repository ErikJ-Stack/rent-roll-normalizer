"""
migrate_to_v0212.py — Substrate template v0.2.11 → v0.2.12

Closes UW-BACKLOG **BL-0024** (Dashboard blended-vs-segment formula
mis-references).

THE BUG:
  Twelve Dashboard cells pull from `'T12 Analytics'!F134`, `F140`, or
  `F143` — but those source cells are segment-specific (F134 = AL,
  F140 = MC, F143 = MC) while the Dashboard cells are labeled as
  blended / community-wide.

  Concrete examples on the Homestead fixture:
    - Dashboard!B6 (OCCUPANCY tile, labeled "Normalized community
      occupancy") shows 64.5% via F134=C11/C6 (AL-only) instead of
      the actual blended 72.7% (128 occupied / 176 licensed).
    - Dashboard!F20 (labeled "Blended ADR / day") shows MC ADR via
      F140=D20/(D11*12) instead of blended.
    - Dashboard!K6 (REVPOR tile, labeled "Normalized RevPOR per
      resident") shows MC RevPOR via F143=(D20+D27)/(D11*12).

  The bug was authored on Dashboard's side (T12 Analytics col A
  labels even say "— AL" / "— MC" for those F-cells; the Dashboard
  copy-paste pointed at the wrong column when wiring blended tiles).

THE FIX (inline approach — no T12 Analytics edits):
  Rewrite all 12 Dashboard formulas to compute blended directly from
  T12 Analytics column E primitives:
    E6  = total licensed beds  (=SUM(B6:D6))
    E11 = total stabilized occupied beds  (=SUM(B11:D11), where
                                            B11:D11 = B6*B10 etc.)
    E20 = blended stabilized annual base rent  (=SUM(B20:D20))
    E27 = blended stabilized LOC revenue       (=SUM(B27:D27))

  Twelve cells patched on Dashboard:

  ┌──────┬────────────────────────────────────┬────────────────────┐
  │ Cell │ Label / role                        │ Old → new ref      │
  ├──────┼────────────────────────────────────┼────────────────────┤
  │ B6   │ OCCUPANCY headline tile (TEXT %)    │ F134 → E11/E6      │
  │ B8   │ Status text under OCCUPANCY tile    │ F134 → E11/E6      │
  │ C21  │ "Community occupancy %" detail row  │ F134 → E11/E6      │
  │ D35  │ "Community occupancy" side card     │ F134 → E11/E6      │
  │ E55  │ "Occupancy gap to market" delta     │ F134 → E11/E6      │
  │ G55  │ Risk-row flag emoji                 │ F134 → E11/E6      │
  │ H55  │ Risk-row flag text                  │ F134 → E11/E6      │
  │ P5   │ Upper-right "Blended" anchor pin    │ F134 → E11/E6      │
  │ F20  │ "Blended ADR / day" tile            │ F140 → E20/(E11*12)│
  │ K6   │ REVPOR headline tile (TEXT $)       │ F143 → (E20+E27)/  │
  │      │                                     │   (E11*12)         │
  │ K8   │ Status text under REVPOR tile       │ F143 → (E20+E27)/  │
  │      │                                     │   (E11*12)         │
  │ F21  │ "RevPOR (blended)" detail row       │ F143 → (E20+E27)/  │
  │      │                                     │   (E11*12)         │
  └──────┴────────────────────────────────────┴────────────────────┘

  All rewrites wrap the inline division in IFERROR(...,"—") so the
  div-by-zero path (uninstalled workbook with E6=0 or E11=0) renders
  the same em-dash placeholder the old formulas produced. Threshold-
  comparison shapes (B8, G55, H55) preserve the existing ✓/⚠/✗
  branches and "— Source not populated" fallback exactly.

  No T12 Analytics edits — pure Dashboard surface.

Idempotency:
  - Gate: `Cover!B8 == "v0.2.12"` AND `Dashboard!B6` formula does NOT
    contain "F134" (i.e. already rewritten).
  - Each cell patch is self-idempotent: write-once via "current
    formula contains F134/F140/F143 substring" check. Already-fixed
    cells are skipped.

BUNDLED FILE STATUS:
  Bundled `ALF_Financial_Analyzer_Only.xlsx` (currently v0.2.11) is
  updated in place by running this migration on it.

Usage:
    python tools/migration/migrate_to_v0212.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

SUBSTRATE_FROM = "v0.2.11"
SUBSTRATE_TO = "v0.2.12"

DASHBOARD_SHEET = "Dashboard"

# 16-sheet anchor list — unchanged from v0.2.10 / v0.2.11 (no sheet adds/removes).
ANCHOR_SHEETS = (
    "Cover", "Dashboard",
    "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending",
    "AR & Collections",
    "UW Output", "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# Blended-ratio expression building blocks (Excel formula text fragments).
# Wrapping each in T12 Analytics ref to keep the cross-sheet sheet-name
# qualifier on every cell ref.
_E6 = "'T12 Analytics'!E6"
_E11 = "'T12 Analytics'!E11"
_E20 = "'T12 Analytics'!E20"
_E27 = "'T12 Analytics'!E27"

OCC_EXPR = f"{_E11}/{_E6}"                       # blended occupancy
ADR_EXPR = f"{_E20}/({_E11}*12)"                  # blended ADR / day
REVPOR_EXPR = f"({_E20}+{_E27})/({_E11}*12)"      # blended RevPOR

# ----------------------------------------------------------------------
# Target formulas — one row per patched cell.
# ----------------------------------------------------------------------
# Each entry: (cell_ref, buggy_substring, new_formula, label_for_logging)
PATCHES = [
    # --- OCCUPANCY (F134 → blended E11/E6) ---
    (
        "B6",
        "F134",
        f'=IFERROR(TEXT({OCC_EXPR},"0.0%"),"—")',
        "OCCUPANCY tile",
    ),
    (
        "B8",
        "F134",
        (
            f'=IFERROR(IF({OCC_EXPR}>=0.9,'
            f'"✓ At/above 90% stabilized hurdle",'
            f'IF({OCC_EXPR}>=0.85,'
            f'"⚠ Below stabilized, lease-up risk",'
            f'"✗ Distressed occupancy")),'
            f'"— Source not populated")'
        ),
        "OCCUPANCY status text",
    ),
    (
        "C21",
        "F134",
        f'=IFERROR({OCC_EXPR},"—")',
        "Community occupancy % row",
    ),
    (
        "D35",
        "F134",
        f'=IFERROR({OCC_EXPR},"—")',
        "Community occupancy card",
    ),
    (
        "E55",
        "F134",
        f'=IFERROR({OCC_EXPR}-0.895,"—")',
        "Occupancy gap-to-market delta",
    ),
    (
        "G55",
        "F134",
        (
            f'=IFERROR(IF({OCC_EXPR}>=0.895,"✅",'
            f'IF({OCC_EXPR}>=0.85,"⚠️","❌")),"⚪")'
        ),
        "Occupancy risk flag emoji",
    ),
    (
        "H55",
        "F134",
        (
            f'=IFERROR(IF({OCC_EXPR}>=0.895,"At or above NIC MAP",'
            f'IF({OCC_EXPR}>=0.85,"Below market — lease-up risk",'
            f'"Significantly below market")),"No data")'
        ),
        "Occupancy risk flag text",
    ),
    (
        "P5",
        "F134",
        f'=IFERROR({OCC_EXPR},NA())',
        "Upper-right Blended anchor",
    ),
    # --- ADR (F140 → blended E20/(E11*12)) ---
    (
        "F20",
        "F140",
        f'=IFERROR({ADR_EXPR},"—")',
        "Blended ADR tile",
    ),
    # --- REVPOR (F143 → blended (E20+E27)/(E11*12)) ---
    (
        "K6",
        "F143",
        f'=IFERROR(TEXT({REVPOR_EXPR},"$#,##0"),"—")',
        "REVPOR tile",
    ),
    (
        "K8",
        "F143",
        (
            f'=IFERROR(IF({REVPOR_EXPR}>0,'
            f'"vs blended market $5,500-$6,500",'
            f'"— Source not populated"),'
            f'"— Source not populated")'
        ),
        "REVPOR status text",
    ),
    (
        "F21",
        "F143",
        f'=IFERROR({REVPOR_EXPR},"—")',
        "RevPOR (blended) detail row",
    ),
]


def is_already_v0212(wb) -> bool:
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    ws = wb[DASHBOARD_SHEET]
    b6 = ws["B6"].value
    if not isinstance(b6, str) or "F134" in b6:
        return False
    return True


def patch_dashboard_formulas(wb) -> dict:
    """Rewrite the 12 buggy Dashboard cells. Each cell is self-idempotent."""
    ws = wb[DASHBOARD_SHEET]
    counts = {"patched": 0, "skipped_already_fixed": 0, "skipped_unexpected": 0}
    log = []

    for cell_ref, buggy_substr, new_formula, label in PATCHES:
        current = ws[cell_ref].value
        if not isinstance(current, str):
            counts["skipped_unexpected"] += 1
            log.append(f"  SKIP {cell_ref} ({label}): not a string ({current!r})")
            continue
        if buggy_substr not in current:
            # Already fixed (or hand-edited away from the buggy ref).
            counts["skipped_already_fixed"] += 1
            log.append(f"  SKIP {cell_ref} ({label}): no '{buggy_substr}' in formula")
            continue
        ws[cell_ref] = new_formula
        counts["patched"] += 1
        log.append(f"  PATCH {cell_ref} ({label})")

    return {"counts": counts, "log": log}


def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


def verify_migration(wb) -> dict:
    r = {}
    ws = wb[DASHBOARD_SHEET]

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    # Each cell should no longer contain its buggy substring,
    # and SHOULD now contain at least one E-column primitive ref.
    all_clean = True
    cell_results = []
    for cell_ref, buggy_substr, _, label in PATCHES:
        v = ws[cell_ref].value
        clean = isinstance(v, str) and buggy_substr not in v and (
            f"!E11" in v or f"!E20" in v or f"!E6" in v
        )
        all_clean = all_clean and clean
        cell_results.append((cell_ref, label, clean))
    r["all_cells_clean"] = all_clean
    r["cell_results"] = cell_results

    # No F134/F140/F143 anywhere on Dashboard anymore (broader sweep).
    leftover = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and (
                "F134" in v or "F140" in v or "F143" in v
            ):
                leftover.append((cell.coordinate, v))
    r["leftover_buggy_refs"] = leftover
    r["no_leftover_ok"] = len(leftover) == 0

    # Anchor stamps
    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    r["sheet_count"] = len(wb.sheetnames)
    r["sheet_count_ok"] = r["sheet_count"] == 16

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0212(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    cur_version = wb["Cover"]["B8"].value
    if cur_version != SUBSTRATE_FROM:
        print(
            f"  WARN: Cover!B8 = {cur_version!r}, expected {SUBSTRATE_FROM!r}. "
            f"Proceeding anyway."
        )

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    result = patch_dashboard_formulas(wb)
    for line in result["log"]:
        print(line)
    c = result["counts"]
    print(
        f"  Dashboard patches: {c['patched']} rewritten, "
        f"{c['skipped_already_fixed']} already-fixed, "
        f"{c['skipped_unexpected']} unexpected-shape skipped"
    )

    stamp_versions(wb)
    print(
        f"  Stamped substrate version -> {SUBSTRATE_TO} on "
        f"Cover!B8 + {len(ANCHOR_SHEETS)} AZ4 anchors"
    )

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:14s}                       : {r['cover_b8_ok']}")
    print(f"  All 12 cells clean (no F134/F140/F143, has E-ref)  : {r['all_cells_clean']}")
    for cell_ref, label, ok in r["cell_results"]:
        print(f"    {cell_ref:4s} {label:38s}                  : {ok}")
    print(f"  No leftover F134/F140/F143 anywhere on Dashboard  : {r['no_leftover_ok']}")
    if r["leftover_buggy_refs"]:
        for coord, val in r["leftover_buggy_refs"]:
            print(f"    LEFTOVER {coord}: {val!r}")
    print(f"  Sheet count = {r['sheet_count']} (expected 16)                   : {r['sheet_count_ok']}")
    print(f"  All 16 AZ4 = {SUBSTRATE_TO}                          : {r['az4_all']} ({r['az4_count']} sheets)")

    all_ok = all([
        r["cover_b8_ok"], r["all_cells_clean"], r["no_leftover_ok"],
        r["sheet_count_ok"], r["az4_all"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0212.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
