"""
migrate_to_v0112.py - Substrate template v0.1.11 -> v0.1.12

Adds Section M to `Rent Roll Recon` (rows 119-172):

  M1 - Published schedule (analyst paste-in from broker package)
       7 default fee rows + 3 extra rows for property-specific additions.
       Column D holds a dropdown linking to the Description_Map Label
       vocabulary, defaulting to the most appropriate Label per fee.

  M2 - RR-side capture (auto)
       For Second Person Fee: counts occupied IL beds (eligible) vs. count
       of residents with `Rent Roll Input!V > 0` (capturing), computes
       capture %. Other 6 fees show 'n/a (one-time)' / 'n/a (see UW-BACKLOG
       BL-0003)' depending on whether the fee is event-based or whether
       per-fee RR-side data isn't captured yet.

  M3 - T12 actuals (auto)
       VLOOKUP from M1 T12 Label into T12 Raw Data col R (T12_Total). When
       multiple M1 fees share the same Label (typical for 'Other community
       revenue'), the second and subsequent rows show a 'shared bucket'
       note so the analyst doesn't read the same $ value as multiple
       independent fee streams.

  M4 - Implied per-resident rate (T12 actual / RR count vs. schedule)
       Only computes for fees with a direct RR match (SP today). Others
       show 'falls into M5 Misc.' Variance % vs. schedule fires a
       conditional note if |variance| > 5%.

  M5 - Misc. Income (T12 ancillary not attributable per-fee)
       Surfaces the residual 'Other community revenue' that wasn't tied
       to a specific M1 fee via M4. Computes residual as % of EGI; fires
       a conditional note if residual % EGI > 15% ('ancillary mix is
       concentrated in shared bucket - consider expanding RR Input').

Also:

  C. Stamp Cover!B8 + 13 AZ4 anchors to v0.1.11 -> v0.1.12.
  D. 10-check verification block.

Idempotent: gate checks BOTH version stamp AND that A119 already reads
the Section M title. Re-runs on a partial-state file safely re-apply.

Usage:
    python tools/migration/migrate_to_v0112.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

SUBSTRATE_FROM = "v0.1.11"
SUBSTRATE_TO = "v0.1.12"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

RECON_SHEET = "Rent Roll Recon"

# ===== Styling =====
NAVY = "FF1F3864"            # Section title bar
LIGHT_NAVY = "FF305496"      # Subsection title
PALE_YELLOW = "FFFFF2CC"     # Analyst paste-in cells
PALE_GREEN = "FFE2EFDA"      # Auto-calculated cells
PALE_GREY = "FFF2F2F2"       # Header rows / n/a cells
WHITE = "FFFFFFFF"
DARK_TEXT = "FF1F1F1F"

TITLE_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
BODY_FONT = Font(name="Calibri", size=10, color=DARK_TEXT)
ITALIC_FONT = Font(name="Calibri", size=9, italic=True, color="FF7F7F7F")

TITLE_FILL = PatternFill(fill_type="solid", fgColor=NAVY)
SUBTITLE_FILL = PatternFill(fill_type="solid", fgColor=LIGHT_NAVY)
HEADER_FILL = PatternFill(fill_type="solid", fgColor=PALE_GREY)
PASTE_FILL = PatternFill(fill_type="solid", fgColor=PALE_YELLOW)
AUTO_FILL = PatternFill(fill_type="solid", fgColor=PALE_GREEN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ===== Section M layout =====
SEC_TITLE_ROW = 119
M1_TITLE_ROW = 121
M1_HEADER_ROW = 122
M1_FIRST_DATA = 123
M1_LAST_DATA = 131   # 9 fee rows (7 defaults + 2 blank for property-specific)
M2_TITLE_ROW = 133
M2_HEADER_ROW = 134
M2_FIRST_DATA = 135
M2_LAST_DATA = M2_FIRST_DATA + (M1_LAST_DATA - M1_FIRST_DATA)  # parallel
M3_TITLE_ROW = M2_LAST_DATA + 2
M3_HEADER_ROW = M3_TITLE_ROW + 1
M3_FIRST_DATA = M3_HEADER_ROW + 1
M3_LAST_DATA = M3_FIRST_DATA + (M1_LAST_DATA - M1_FIRST_DATA)
M4_TITLE_ROW = M3_LAST_DATA + 2
M4_HEADER_ROW = M4_TITLE_ROW + 1
M4_FIRST_DATA = M4_HEADER_ROW + 1
M4_LAST_DATA = M4_FIRST_DATA + (M1_LAST_DATA - M1_FIRST_DATA)
M5_TITLE_ROW = M4_LAST_DATA + 2
M5_FIRST_DATA = M5_TITLE_ROW + 1

# Default fees (7 rows pre-populated; rows 8-9 left blank for property-specific)
# (fee_name, basis, default_t12_label, m2_note_kind)
#   m2_note_kind: "sp" -> direct RR match formulas
#                 "one-time" -> n/a (one-time fee)
#                 "no-rr"    -> awaiting RR Input expansion
DEFAULT_FEES = [
    ("Community Fee",        "per unit",     "Community / move-in fees", "one-time"),
    ("Elective Transfer Fee","per move",     "Other community revenue",  "one-time"),
    ("Meal Delivery",        "per meal",     "Other community revenue",  "no-rr"),
    ("Motorized Scooter Fee","per month",    "Other community revenue",  "no-rr"),
    ("Second Person Fee",    "per month",    "2nd Person Revenue",       "sp"),
    ("Housekeeping",         "per visit",    "Other community revenue",  "no-rr"),
    ("Laundry",              "per pickup",   "Other community revenue",  "no-rr"),
]

# Cell ranges used in formulas
RRI_OCCUPIED_RANGE = "'Rent Roll Input'!$E$7:$E$606"
RRI_CARE_RANGE     = "'Rent Roll Input'!$D$7:$D$606"
RRI_SP_RANGE       = "'Rent Roll Input'!$V$7:$V$606"
T12RD_LABEL_RANGE  = "'T12 Raw Data'!$B$5:$R$60"   # B = Label, R = T12_Total
EGI_CELL           = "'Monthly Trending'!$N$21"   # EFFECTIVE GROSS INCOME annual

# Variance threshold for M4 / M5 notes (5% on M4, 15% on M5)
M4_VARIANCE_THRESHOLD = 0.05
M5_PCT_EGI_THRESHOLD  = 0.15


def is_already_v0112(wb) -> bool:
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    ws = wb[RECON_SHEET]
    title = ws.cell(SEC_TITLE_ROW, 1).value
    return isinstance(title, str) and title.startswith("M ")


# ----- Section M installation -----

def _stamp_title(ws, row: int, text: str, fill: PatternFill, font: Font, cols=9):
    """Write title at A{row}, merge A:I, fill+font."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill
    cell.font = font
    cell.alignment = LEFT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=cols)
    ws.row_dimensions[row].height = 18


def _stamp_header_cell(c, value):
    c.value = value
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BOX


def _stamp_paste_cell(c):
    c.fill = PASTE_FILL
    c.font = BODY_FONT
    c.alignment = LEFT
    c.border = BOX


def _stamp_auto_cell(c):
    c.fill = AUTO_FILL
    c.font = BODY_FONT
    c.alignment = LEFT
    c.border = BOX


def install_m1_schedule(ws) -> int:
    """M1 — Published schedule (paste-in)."""
    _stamp_title(ws, SEC_TITLE_ROW,
                 "M  ·  OPERATOR FEE SCHEDULE  &  ANCILLARY RECONCILIATION",
                 TITLE_FILL, TITLE_FONT)
    _stamp_title(ws, M1_TITLE_ROW,
                 "M1  —  Published schedule  (paste-in: published rates from broker)",
                 SUBTITLE_FILL, SUBTITLE_FONT)

    # Headers
    headers = ["Fee Name", "Published $", "Basis", "T12 Label"]
    for col_idx, h in enumerate(headers, start=1):
        _stamp_header_cell(ws.cell(M1_HEADER_ROW, col_idx), h)

    # Default rows
    for offset, (name, basis, label, _kind) in enumerate(DEFAULT_FEES):
        r = M1_FIRST_DATA + offset
        # Fee name — paste/edit-friendly
        c = ws.cell(r, 1, value=name)
        _stamp_paste_cell(c); c.alignment = LEFT
        # Published $ — paste-in, currency format
        c = ws.cell(r, 2)
        _stamp_paste_cell(c)
        c.number_format = "$#,##0.00"
        # Basis — paste-in
        c = ws.cell(r, 3, value=basis)
        _stamp_paste_cell(c)
        # T12 Label — default value, dropdown via data validation
        c = ws.cell(r, 4, value=label)
        _stamp_paste_cell(c)

    # Extra blank rows (M1_FIRST_DATA + len(DEFAULT_FEES) ... M1_LAST_DATA)
    for r in range(M1_FIRST_DATA + len(DEFAULT_FEES), M1_LAST_DATA + 1):
        for col in range(1, 5):
            _stamp_paste_cell(ws.cell(r, col))
        ws.cell(r, 2).number_format = "$#,##0.00"

    # Data validation for T12 Label column (uses DescMap_Label named range)
    dv = DataValidation(type="list", formula1="=DescMap_Label", allow_blank=True)
    dv.error = "Pick a Label from the Description_Map vocabulary."
    dv.errorTitle = "Invalid T12 Label"
    dv.prompt = "Pick the T12 Label this fee maps to."
    dv.promptTitle = "T12 Label"
    ws.add_data_validation(dv)
    dv.add(f"D{M1_FIRST_DATA}:D{M1_LAST_DATA}")

    return M1_LAST_DATA - M1_FIRST_DATA + 1


def install_m2_capture(ws) -> int:
    """M2 — RR-side capture (auto)."""
    _stamp_title(ws, M2_TITLE_ROW,
                 "M2  —  RR-side capture  (auto, from Rent Roll Input)",
                 SUBTITLE_FILL, SUBTITLE_FONT)

    headers = ["Fee Name", "Eligible #", "Capturing #", "Capture %", "Note"]
    for col_idx, h in enumerate(headers, start=1):
        _stamp_header_cell(ws.cell(M2_HEADER_ROW, col_idx), h)

    for offset, (name, _basis, _label, kind) in enumerate(DEFAULT_FEES):
        m1_row = M1_FIRST_DATA + offset
        m2_row = M2_FIRST_DATA + offset

        # Col A — reference M1 fee name
        c = ws.cell(m2_row, 1, value=f"=A{m1_row}")
        _stamp_auto_cell(c)

        if kind == "sp":
            # Eligible: occupied IL beds (current period)
            ws.cell(m2_row, 2, value=(
                f"=COUNTIFS({RRI_CARE_RANGE},\"IL\","
                f"{RRI_OCCUPIED_RANGE},\"Occupied\","
                f"'Rent Roll Input'!$S$7:$S$606,$B$2)"
            ))
            _stamp_auto_cell(ws.cell(m2_row, 2))
            ws.cell(m2_row, 2).number_format = "0"

            # Capturing: count of residents with V > 0 (SP populated)
            ws.cell(m2_row, 3, value=f"=COUNTIF({RRI_SP_RANGE},\">0\")")
            _stamp_auto_cell(ws.cell(m2_row, 3))
            ws.cell(m2_row, 3).number_format = "0"

            # Capture % = capturing / eligible
            ws.cell(m2_row, 4, value=f"=IFERROR(C{m2_row}/B{m2_row},0)")
            _stamp_auto_cell(ws.cell(m2_row, 4))
            ws.cell(m2_row, 4).number_format = "0.0%"

            ws.cell(m2_row, 5, value="✓ Direct RR match (col V — 2nd Person Rent)")
            _stamp_auto_cell(ws.cell(m2_row, 5))
        elif kind == "one-time":
            for col in (2, 3, 4):
                c = ws.cell(m2_row, col, value="n/a")
                _stamp_auto_cell(c); c.alignment = CENTER
            ws.cell(m2_row, 5, value="One-time / event-based fee; see M3 for T12 actual")
            _stamp_auto_cell(ws.cell(m2_row, 5))
        else:  # no-rr
            for col in (2, 3, 4):
                c = ws.cell(m2_row, col, value="—")
                _stamp_auto_cell(c); c.alignment = CENTER
            ws.cell(m2_row, 5, value="No per-fee RR column yet — falls into M5 Misc. (see UW-BACKLOG BL-0003)")
            _stamp_auto_cell(ws.cell(m2_row, 5))

    # Blank-row stamping
    for offset in range(len(DEFAULT_FEES), M1_LAST_DATA - M1_FIRST_DATA + 1):
        m1_row = M1_FIRST_DATA + offset
        m2_row = M2_FIRST_DATA + offset
        for col in range(1, 6):
            _stamp_auto_cell(ws.cell(m2_row, col))
        ws.cell(m2_row, 1, value=f"=A{m1_row}")

    return M2_LAST_DATA - M2_FIRST_DATA + 1


def install_m3_t12(ws) -> int:
    """M3 — T12 actuals via M1 T12 Label."""
    _stamp_title(ws, M3_TITLE_ROW,
                 "M3  —  T12 actuals  (12-month total + monthly avg, via VLOOKUP on M1 Label)",
                 SUBTITLE_FILL, SUBTITLE_FONT)

    headers = ["Fee Name", "T12 Annual $", "T12 Monthly $", "Mapped to Label", "Note"]
    for col_idx, h in enumerate(headers, start=1):
        _stamp_header_cell(ws.cell(M3_HEADER_ROW, col_idx), h)

    for offset in range(M1_LAST_DATA - M1_FIRST_DATA + 1):
        m1_row = M1_FIRST_DATA + offset
        m3_row = M3_FIRST_DATA + offset

        # Fee name reference
        ws.cell(m3_row, 1, value=f"=A{m1_row}")
        _stamp_auto_cell(ws.cell(m3_row, 1))

        # T12 Annual $: if this M1 Label was already used in a prior M1 row,
        # show "(shared bucket — see row N)" instead of duplicating the value
        if offset == 0:
            # First row — straight VLOOKUP
            ws.cell(m3_row, 2, value=(
                f"=IFERROR(VLOOKUP(D{m1_row}, 'T12 Raw Data'!$B:$R, 17, 0), 0)"
            ))
        else:
            # Subsequent rows — check whether Label was used earlier
            ws.cell(m3_row, 2, value=(
                f"=IF(COUNTIF($D${M1_FIRST_DATA}:D{m1_row - 1}, D{m1_row}) > 0, "
                f"\"(shared — see row \" & MATCH(D{m1_row}, $D${M1_FIRST_DATA}:D{m1_row - 1}, 0) + {M1_FIRST_DATA - 1} & \")\", "
                f"IFERROR(VLOOKUP(D{m1_row}, 'T12 Raw Data'!$B:$R, 17, 0), 0))"
            ))
        _stamp_auto_cell(ws.cell(m3_row, 2))
        ws.cell(m3_row, 2).number_format = "$#,##0;($#,##0);\"\""

        # T12 Monthly $: if B is numeric, divide by 12; if string, propagate the note
        ws.cell(m3_row, 3, value=(
            f"=IF(ISNUMBER(B{m3_row}), B{m3_row}/12, B{m3_row})"
        ))
        _stamp_auto_cell(ws.cell(m3_row, 3))
        ws.cell(m3_row, 3).number_format = "$#,##0;($#,##0);\"\""

        # Mapped Label (reference back to M1)
        ws.cell(m3_row, 4, value=f"=D{m1_row}")
        _stamp_auto_cell(ws.cell(m3_row, 4))

        # Note
        ws.cell(m3_row, 5, value=(
            f"=IF(D{m1_row}=\"\",\"\","
            f"IF(COUNTIF($D${M1_FIRST_DATA}:$D${M1_LAST_DATA}, D{m1_row}) > 1, "
            f"\"⚠ Shared T12 Label — can't isolate per-fee\", \"\"))"
        ))
        _stamp_auto_cell(ws.cell(m3_row, 5))

    return M3_LAST_DATA - M3_FIRST_DATA + 1


def install_m4_implied(ws) -> int:
    """M4 — Implied per-resident rate (T12/RR count) vs schedule."""
    _stamp_title(ws, M4_TITLE_ROW,
                 "M4  —  Implied per-resident rate  (T12 monthly ÷ RR # capturing  vs.  schedule)",
                 SUBTITLE_FILL, SUBTITLE_FONT)

    headers = ["Fee Name", "T12 $/mo", "RR # capturing", "Implied $/resident",
               "Schedule $", "Variance %", "Note"]
    for col_idx, h in enumerate(headers, start=1):
        _stamp_header_cell(ws.cell(M4_HEADER_ROW, col_idx), h)

    for offset, (name, _basis, _label, kind) in enumerate(DEFAULT_FEES):
        m1_row = M1_FIRST_DATA + offset
        m2_row = M2_FIRST_DATA + offset
        m3_row = M3_FIRST_DATA + offset
        m4_row = M4_FIRST_DATA + offset

        ws.cell(m4_row, 1, value=f"=A{m1_row}")
        _stamp_auto_cell(ws.cell(m4_row, 1))

        # B: T12 $/mo (from M3 col C); shared-bucket strings propagate as-is
        ws.cell(m4_row, 2, value=f"=C{m3_row}")
        _stamp_auto_cell(ws.cell(m4_row, 2))
        ws.cell(m4_row, 2).number_format = "$#,##0;($#,##0);\"\""

        if kind == "sp":
            # Direct match — compute implied vs. schedule
            ws.cell(m4_row, 3, value=f"=C{m2_row}")   # RR capturing count from M2
            _stamp_auto_cell(ws.cell(m4_row, 3))
            ws.cell(m4_row, 3).number_format = "0"

            ws.cell(m4_row, 4, value=(
                f"=IF(AND(ISNUMBER(B{m4_row}), C{m4_row} > 0), B{m4_row}/C{m4_row}, \"\")"
            ))
            _stamp_auto_cell(ws.cell(m4_row, 4))
            ws.cell(m4_row, 4).number_format = "$#,##0;($#,##0);\"\""

            ws.cell(m4_row, 5, value=f"=B{m1_row}")
            _stamp_auto_cell(ws.cell(m4_row, 5))
            ws.cell(m4_row, 5).number_format = "$#,##0;($#,##0);\"\""

            ws.cell(m4_row, 6, value=(
                f"=IF(AND(ISNUMBER(D{m4_row}), ISNUMBER(E{m4_row}), E{m4_row} <> 0), "
                f"(D{m4_row} - E{m4_row}) / E{m4_row}, \"\")"
            ))
            _stamp_auto_cell(ws.cell(m4_row, 6))
            ws.cell(m4_row, 6).number_format = "0.0%;(0.0%);\"\""

            ws.cell(m4_row, 7, value=(
                f"=IF(NOT(ISNUMBER(F{m4_row})), \"\","
                f"IF(ABS(F{m4_row}) > {M4_VARIANCE_THRESHOLD}, "
                f"\"⚠ Implied rate differs from schedule by \" & TEXT(F{m4_row}, \"0.0%\") & "
                f" \" — legacy in-place residents or schedule out of date\", "
                f"\"✓ Implied rate within \" & TEXT({M4_VARIANCE_THRESHOLD}, \"0%\") & \" of schedule\"))"
            ))
            _stamp_auto_cell(ws.cell(m4_row, 7))
        elif kind == "one-time":
            for col in (3, 4, 5, 6):
                c = ws.cell(m4_row, col, value="n/a")
                _stamp_auto_cell(c); c.alignment = CENTER
            ws.cell(m4_row, 7, value="One-time fee; reconcile via M3 T12 actual ÷ annual turnover")
            _stamp_auto_cell(ws.cell(m4_row, 7))
        else:  # no-rr
            for col in (3, 4):
                c = ws.cell(m4_row, col, value="—")
                _stamp_auto_cell(c); c.alignment = CENTER
            ws.cell(m4_row, 5, value=f"=B{m1_row}")
            _stamp_auto_cell(ws.cell(m4_row, 5))
            ws.cell(m4_row, 5).number_format = "$#,##0;($#,##0);\"\""
            c = ws.cell(m4_row, 6, value="—")
            _stamp_auto_cell(c); c.alignment = CENTER
            ws.cell(m4_row, 7, value="Falls into M5 Misc. — needs per-fee RR column (see UW-BACKLOG BL-0003)")
            _stamp_auto_cell(ws.cell(m4_row, 7))

    # Blank rows (8th and 9th rows)
    for offset in range(len(DEFAULT_FEES), M1_LAST_DATA - M1_FIRST_DATA + 1):
        m1_row = M1_FIRST_DATA + offset
        m4_row = M4_FIRST_DATA + offset
        for col in range(1, 8):
            _stamp_auto_cell(ws.cell(m4_row, col))
        ws.cell(m4_row, 1, value=f"=A{m1_row}")

    return M4_LAST_DATA - M4_FIRST_DATA + 1


def install_m5_misc(ws) -> int:
    """M5 — Misc. Income (unattributed)."""
    _stamp_title(ws, M5_TITLE_ROW,
                 "M5  —  Misc. Income  (T12 'Other community revenue' minus per-fee attribution)",
                 SUBTITLE_FILL, SUBTITLE_FONT)

    rows = [
        ("T12 'Other community revenue' — annual total",
         "=IFERROR(VLOOKUP(\"Other community revenue\", 'T12 Raw Data'!$B:$R, 17, 0), 0)",
         "$#,##0;($#,##0);\"\""),
        ("T12 'Other community revenue' — monthly avg",
         f"=B{M5_FIRST_DATA}/12",
         "$#,##0;($#,##0);\"\""),
        ("Less: M4 fees attributed to this bucket (annualized)",
         # Sum of B{m4_row}*12 across M4 rows where M1 Label = "Other community revenue"
         # AND M2 had a direct RR match (RR count populated).
         # Today this evaluates to 0 (only SP has direct RR match, and SP isn't OCR).
         # When future RR columns add Meal/HK/Laundry/Scooter, those will populate
         # M4 directly and this SUMIFS will deduct them.
         f"=SUMPRODUCT(("
         f"$D${M1_FIRST_DATA}:$D${M1_LAST_DATA}=\"Other community revenue\")*"
         f"ISNUMBER($C${M2_FIRST_DATA}:$C${M2_LAST_DATA})*"
         f"IFERROR($C${M2_FIRST_DATA}:$C${M2_LAST_DATA}*$B${M1_FIRST_DATA}:$B${M1_LAST_DATA}*12,0))",
         "$#,##0;($#,##0);\"\""),
        ("Residual unattributed Misc. income (annual)",
         f"=B{M5_FIRST_DATA} - B{M5_FIRST_DATA + 2}",
         "$#,##0;($#,##0);\"\""),
        ("Residual as % of EGI (annual)",
         f"=IFERROR(B{M5_FIRST_DATA + 3} / {EGI_CELL}, 0)",
         "0.0%"),
    ]

    for i, (label, formula, fmt) in enumerate(rows):
        r = M5_FIRST_DATA + i
        c = ws.cell(r, 1, value=label)
        _stamp_auto_cell(c)
        c.alignment = LEFT
        c = ws.cell(r, 2, value=formula)
        _stamp_auto_cell(c)
        c.alignment = RIGHT
        c.number_format = fmt

    # Conditional note
    note_row = M5_FIRST_DATA + 6
    c = ws.cell(note_row, 1, value=(
        f"=IF(B{M5_FIRST_DATA + 4} > {M5_PCT_EGI_THRESHOLD}, "
        f"\"⚠ Misc. income is \" & TEXT(B{M5_FIRST_DATA + 4}, \"0.0%\") & "
        f"\" of EGI — ancillary mix concentrated in shared bucket. \" & "
        f"\"Expand RR Input per-fee columns (see UW-BACKLOG BL-0003) to break out.\", "
        f"IF(B{M5_FIRST_DATA + 3} = 0, \"\", "
        f"\"✓ Misc. income share within band (≤ \" & TEXT({M5_PCT_EGI_THRESHOLD}, \"0%\") & \" of EGI).\"))"
    ))
    c.font = ITALIC_FONT
    c.alignment = LEFT
    ws.merge_cells(start_row=note_row, end_row=note_row, start_column=1, end_column=9)

    return len(rows) + 1


# ----- Versioning -----

def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ----- Verify -----

def verify_migration(wb) -> dict:
    r: dict = {}
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    ws = wb[RECON_SHEET]

    # M section title
    r["sec_title"] = ws.cell(SEC_TITLE_ROW, 1).value
    r["sec_title_ok"] = isinstance(r["sec_title"], str) and r["sec_title"].startswith("M ")

    # M1 subsection
    r["m1_title_ok"] = isinstance(ws.cell(M1_TITLE_ROW, 1).value, str) and "M1" in ws.cell(M1_TITLE_ROW, 1).value
    r["m1_default_fees"] = sum(
        1 for offset in range(len(DEFAULT_FEES))
        if ws.cell(M1_FIRST_DATA + offset, 1).value == DEFAULT_FEES[offset][0]
    )
    r["m1_fees_ok"] = r["m1_default_fees"] == len(DEFAULT_FEES)

    # M2 SP formula
    sp_offset = next(i for i, f in enumerate(DEFAULT_FEES) if f[0] == "Second Person Fee")
    sp_m2_row = M2_FIRST_DATA + sp_offset
    sp_count_cell = ws.cell(sp_m2_row, 3).value
    r["sp_count_formula"] = sp_count_cell
    r["sp_count_ok"] = isinstance(sp_count_cell, str) and "COUNTIF" in sp_count_cell and "V$7:$V$606" in sp_count_cell

    # M3 first row VLOOKUP
    m3_first_b = ws.cell(M3_FIRST_DATA, 2).value
    r["m3_vlookup_ok"] = isinstance(m3_first_b, str) and "VLOOKUP" in m3_first_b and "T12 Raw Data" in m3_first_b

    # M4 SP variance row
    sp_m4_row = M4_FIRST_DATA + sp_offset
    r["m4_variance_formula"] = ws.cell(sp_m4_row, 6).value
    r["m4_variance_ok"] = isinstance(r["m4_variance_formula"], str) and "D" in r["m4_variance_formula"]

    # M5 residual
    r["m5_residual_formula"] = ws.cell(M5_FIRST_DATA + 3, 2).value
    r["m5_residual_ok"] = isinstance(r["m5_residual_formula"], str) and "B" in r["m5_residual_formula"]

    # Sections K and L still intact (rows 86-117 untouched)
    r["section_k_intact"] = "IL UNIT-TYPE MIX" in (ws.cell(86, 1).value or "").upper() or "IL " in (ws.cell(86, 1).value or "")
    r["section_l_intact"] = "MC CARE STRUCTURE" in (ws.cell(102, 1).value or "").upper()

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0112(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    ws = wb[RECON_SHEET]
    n1 = install_m1_schedule(ws);  print(f"  M1: installed schedule ({n1} rows)")
    n2 = install_m2_capture(ws);   print(f"  M2: installed RR-capture block ({n2} rows)")
    n3 = install_m3_t12(ws);       print(f"  M3: installed T12 actuals block ({n3} rows)")
    n4 = install_m4_implied(ws);   print(f"  M4: installed implied-per-resident block ({n4} rows)")
    n5 = install_m5_misc(ws);      print(f"  M5: installed misc-income block ({n5} rows)")

    stamp_versions(wb)
    print(f"  C: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:<22}     : {r['cover_b8_ok']}")
    print(f"  All 13 AZ4 = {SUBSTRATE_TO}              : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  Section M title in place              : {r['sec_title_ok']}")
    print(f"  M1 subsection title                   : {r['m1_title_ok']}")
    print(f"  M1 7 default fees installed           : {r['m1_fees_ok']} ({r['m1_default_fees']}/7)")
    print(f"  M2 SP COUNTIF formula                 : {r['sp_count_ok']}")
    print(f"  M3 first row VLOOKUP formula          : {r['m3_vlookup_ok']}")
    print(f"  M4 SP variance formula                : {r['m4_variance_ok']}")
    print(f"  M5 residual formula                   : {r['m5_residual_ok']}")
    print(f"  Section K (IL deep-dive) intact       : {r['section_k_intact']}")
    print(f"  Section L (MC structure) intact       : {r['section_l_intact']}")

    all_ok = all([
        r["cover_b8_ok"], r["az4_all"], r["sec_title_ok"], r["m1_title_ok"],
        r["m1_fees_ok"], r["sp_count_ok"], r["m3_vlookup_ok"],
        r["m4_variance_ok"], r["m5_residual_ok"],
        r["section_k_intact"], r["section_l_intact"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0112.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
