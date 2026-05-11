"""
migrate_to_v0110.py — Substrate template v0.1.9 -> v0.1.10

Companion to RR v1.16.0 (Tier 1.2 + Tier 2 + Tier 3 data-capture expansion).
Adds 7 new data columns to the Analyzer's `Rent Roll Input` sheet, after the
existing T/U formula columns:

  V  — 2nd Person Rent $   (Tier 1.2; aligns with T12 substrate's `2nd Person Revenue` Label)
  W  — Move-out Date       (Tier 2.1; vacate forecasting)
  X  — Balance             (Tier 2.2; bad-debt indicator)
  Y  — Notes               (Tier 2.3; free-form context)
  Z  — Market PSF          (Tier 3.1; rate per sqft)
  AA — Actual PSF          (Tier 3.1; rate per sqft)
  AB — ACH                 (Tier 3.2; collection-velocity flag)

Updates the Total Monthly Rev formula at U7:U606 to include 2nd Person Rent,
since 2P is incremental housing revenue (not LOC) and would otherwise be
silently excluded from per-resident TMR.

Operations:

  A. Write column headers at row 4 cols V-AB on Rent Roll Input (styled to
     match existing navy header).
  B. Replace U7:U606 formula from:
         =IFERROR(H7+IFERROR(I7,0)+T7,0)
     to:
         =IFERROR(H7+IFERROR(I7,0)+T7+IFERROR(V7,0),0)
  C. Stamp Cover!B8 + 13 AZ4 anchors to v0.1.10
  D. 8-check verification

Idempotent: gate checks BOTH version stamp AND that row-4 V header is
present, so re-running on a partial-state file safely re-applies.

Usage:
    python tools/migration/migrate_to_v0110.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SUBSTRATE_FROM = "v0.1.9"
SUBSTRATE_TO = "v0.1.10"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# New columns at V-AB on Rent Roll Input
NEW_HEADERS = [
    (22, "V",  "2nd Person\nRent $"),
    (23, "W",  "Move-out\nDate"),
    (24, "X",  "Balance"),
    (25, "Y",  "Notes"),
    (26, "Z",  "Market\nPSF"),
    (27, "AA", "Actual\nPSF"),
    (28, "AB", "ACH"),
]

HEADER_ROW = 4
DATA_START = 7
DATA_END = 606

# Header style — match the existing navy header row 4 convention
NAVY = "FF1F3864"
WHITE = "FFFFFFFF"
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill(fill_type="solid", fgColor=NAVY)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def is_already_v0110(wb) -> bool:
    """Gate: version stamp AND new V4 header present."""
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    rri = wb["Rent Roll Input"]
    v4 = rri.cell(row=HEADER_ROW, column=22).value
    return isinstance(v4, str) and "2nd Person" in v4


def install_new_headers(wb) -> int:
    """A. Write headers at row 4 cols V-AB on Rent Roll Input."""
    rri = wb["Rent Roll Input"]
    n = 0
    for col_idx, _letter, label in NEW_HEADERS:
        c = rri.cell(row=HEADER_ROW, column=col_idx, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        n += 1
    return n


def update_tmr_formula(wb) -> int:
    """B. Extend Total Monthly Rev (U7:U606) to include +V (2nd Person Rent).

    Pattern: =IFERROR(H{r}+IFERROR(I{r},0)+T{r},0)
        ->   =IFERROR(H{r}+IFERROR(I{r},0)+T{r}+IFERROR(V{r},0),0)

    Only rewrites cells whose current value matches the old pattern, so
    customized formulas (if any) are left intact.
    """
    rri = wb["Rent Roll Input"]
    n = 0
    for r in range(DATA_START, DATA_END + 1):
        u = rri.cell(row=r, column=21)  # col U
        v = u.value
        if not isinstance(v, str):
            continue
        old_pattern = f"=IFERROR(H{r}+IFERROR(I{r},0)+T{r},0)"
        new_formula = f"=IFERROR(H{r}+IFERROR(I{r},0)+T{r}+IFERROR(V{r},0),0)"
        if v == old_pattern:
            u.value = new_formula
            n += 1
    return n


def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


def verify_migration(wb) -> dict:
    r = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # New headers V-AB on Rent Roll Input
    rri = wb["Rent Roll Input"]
    found_headers = []
    for col_idx, letter, expected in NEW_HEADERS:
        v = rri.cell(row=HEADER_ROW, column=col_idx).value
        if not isinstance(v, str):
            continue
        # Compare on the leading distinguishing token
        key = expected.split("\n")[0].split()[0].lower()
        if key in v.lower():
            found_headers.append(letter)
    r["new_headers"] = found_headers
    r["all_new_headers"] = len(found_headers) == len(NEW_HEADERS)

    # TMR formula extension — sample 3 rows
    extended = 0
    sampled = 0
    for sample_row in (DATA_START, DATA_START + 100, DATA_END - 1):
        u = str(rri.cell(row=sample_row, column=21).value or "")
        if u:
            sampled += 1
            if f"+IFERROR(V{sample_row},0)" in u:
                extended += 1
    r["tmr_extended_samples"] = f"{extended}/{sampled}"
    r["tmr_extended_ok"] = extended == sampled and sampled > 0

    # Sanity: existing A-R headers still in place
    a4 = rri.cell(row=HEADER_ROW, column=1).value
    r["a4_unit"] = a4
    r["existing_a_r_intact"] = a4 == "Unit #"

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0110(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    n_headers = install_new_headers(wb)
    print(f"  A: installed {n_headers} new headers at Rent Roll Input row 4 cols V-AB")

    n_tmr = update_tmr_formula(wb)
    print(f"  B: extended Total Monthly Rev formula in {n_tmr} cells of col U")

    stamp_versions(wb)
    print(f"  C: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:24s}    : {r['cover_b8_ok']}")
    print(f"  All 13 AZ4 = {SUBSTRATE_TO}              : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  New headers V-AB present              : {r['all_new_headers']} ({r['new_headers']})")
    print(f"  TMR formula extended (sample rows)    : {r['tmr_extended_samples']} ({r['tmr_extended_ok']})")
    print(f"  Existing A4 = 'Unit #' (intact)       : {r['existing_a_r_intact']}")

    all_ok = (
        r["cover_b8_ok"] and r["az4_all"]
        and r["all_new_headers"]
        and r["tmr_extended_ok"]
        and r["existing_a_r_intact"]
    )
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0110.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
