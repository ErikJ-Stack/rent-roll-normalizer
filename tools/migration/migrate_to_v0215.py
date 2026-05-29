"""
Substrate migration v0.2.14 → v0.2.15 — 2nd Person Revenue Description_Map re-map.

Part of the UWT v6 absorption (T-12 Analysis income restructure). The v6
template breaks 2nd Person Revenue out as its own income line (N66), distinct
from Base Rent. For that line to populate, the Analyzer's Description_Map must
route the 2nd-person GL descriptions to a dedicated "2nd Person Revenue" label
instead of folding them into "Base rent — IL/AL/MC".

Four Description_Map rows re-pointed (col B label):
  r127  Second Person Fee                          Base rent — IL → 2nd Person Revenue
  r400  Second Persons Revenue | Assisted Living    Base rent — AL → 2nd Person Revenue
  r401  Second Persons Revenue | Independent Living Base rent — IL → 2nd Person Revenue
  r402  Second Persons Revenue | Memory Care        Base rent — MC → 2nd Person Revenue

(r127 included per operator decision 2026-05-28 — treat the move-in "Second
Person Fee" as 2nd-person revenue too.)

Impact: Base Rent IL/AL/MC totals drop by the 2nd-person amount; "2nd Person
Revenue" picks it up.

**Companion EGI-formula fix (critical).** The Analyzer's EGI
(`T12 Analytics!E52 = E16+E23+E47+E48+E49+E50`, and the normalized
`F52 = E20+E27+F47+F48+F49+F50`) does NOT sum the "2nd Person Revenue" label.
Re-mapping alone would move the 2nd-person dollars into a label EGI ignores →
EGI would DROP by that amount (a regression the upstream handoff's "EGI
unchanged" note missed — caught by verifying the EGI formula chain 2026-05-28).
This migration therefore ALSO amends E52 + F52 to add a direct INDEX/MATCH
term for "2nd Person Revenue" (mirroring rows 47-50). With both changes the
re-map is EGI-neutral. On Homestead the move is real: r127 "Second Person Fee"
carries ~$32,220 — broken out of base rent and added back into EGI via the new
term, so EGI stays $7,001,957.

Idempotent. Gate: Cover!B8 == "v0.2.15" (already migrated → no-op).
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
ANALYZER = ROOT / "ALF_Financial_Analyzer_Only.xlsx"

OLD_VERSION = "v0.2.14"
NEW_VERSION = "v0.2.15"
NEW_LABEL = "2nd Person Revenue"

# (row, expected current label) — re-pointed to NEW_LABEL
REMAP_ROWS = {
    127: "Base rent — IL",
    400: "Base rent — AL",
    401: "Base rent — IL",
    402: "Base rent — MC",
}


def migrate(path: Path = ANALYZER) -> None:
    wb = openpyxl.load_workbook(path)

    # Idempotency gate
    if wb["Cover"]["B8"].value == NEW_VERSION:
        print(f"No-op — already at {NEW_VERSION}.")
        return

    dm = wb["Description_Map"]
    ops = []

    # --- Step 1: re-map the 4 label cells ---
    for row, expected in REMAP_ROWS.items():
        cur = dm.cell(row, 2).value
        if cur == NEW_LABEL:
            continue  # already re-mapped
        if cur != expected:
            print(f"  ⚠ r{row} label is {cur!r}, expected {expected!r} — re-mapping anyway "
                  f"(verify the description in col A is a 2nd-person line).")
        dm.cell(row, 2).value = NEW_LABEL
        ops.append(f"B{row}: {cur!r} → {NEW_LABEL!r}  (desc: {dm.cell(row,1).value!r})")

    # --- Step 1b: amend T12 Analytics EGI to include "2nd Person Revenue" ---
    # Without this the re-mapped 2nd-person dollars drop out of EGI (E52/F52
    # don't sum that label). Add an INDEX/MATCH term mirroring rows 47-50.
    ta = wb["T12 Analytics"]
    _2P_TERM = ("+IFERROR(INDEX('T12 Raw Data'!R:R,"
                "MATCH(\"2nd Person Revenue\",'T12 Raw Data'!B:B,0)),0)")
    for cell, base in (("E52", "=E16+E23+E47+E48+E49+E50"),
                       ("F52", "=E20+E27+F47+F48+F49+F50")):
        cur = ta[cell].value
        if cur and "2nd Person Revenue" in str(cur):
            continue  # already amended
        ta[cell].value = base + _2P_TERM
        ops.append(f"{cell}: appended 2nd-Person-Revenue term to EGI")

    # --- Step 2: bump version stamp on Cover!B8 + every sheet's AZ4 anchor ---
    wb["Cover"]["B8"].value = NEW_VERSION
    anchors = 0
    for ws in wb.worksheets:
        if ws["AZ4"].value == OLD_VERSION:
            ws["AZ4"].value = NEW_VERSION
            anchors += 1
    ops.append(f"version stamp {OLD_VERSION} → {NEW_VERSION} (Cover!B8 + {anchors} AZ4 anchors)")

    wb.save(path)

    print("Applied:")
    for op in ops:
        print(f"  ✓ {op}")

    # --- Verify ---
    print("\nVerify:")
    chk = openpyxl.load_workbook(path)
    dm2 = chk["Description_Map"]
    checks = []
    checks.append(("Cover!B8 == v0.2.15", chk["Cover"]["B8"].value == NEW_VERSION))
    for row in REMAP_ROWS:
        checks.append((f"B{row} == '{NEW_LABEL}'", dm2.cell(row, 2).value == NEW_LABEL))
    ta2 = chk["T12 Analytics"]
    checks.append(("E52 EGI includes 2nd Person Revenue",
                   "2nd Person Revenue" in str(ta2["E52"].value)))
    checks.append(("F52 EGI includes 2nd Person Revenue",
                   "2nd Person Revenue" in str(ta2["F52"].value)))
    n_anchor_new = sum(1 for ws in chk.worksheets if ws["AZ4"].value == NEW_VERSION)
    checks.append((f"all 16 AZ4 anchors == v0.2.15", n_anchor_new == 16))
    checks.append(("sheet count == 16", len(chk.sheetnames) == 16))
    for label, ok in checks:
        print(f"  {'✓' if ok else '✗'} {label}")
    if not all(ok for _, ok in checks):
        sys.exit("VERIFY FAILED")
    print("\nAll checks pass.")


if __name__ == "__main__":
    migrate()
