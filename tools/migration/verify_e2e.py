"""
verify_e2e.py — paste Salem and Briar Glen GL detail into the migrated
Analyzer, force LibreOffice recalc, and check EGI / EBITDARM.

Targets per HANDOFF:
  Salem (Yardi):      73 GL rows, EGI = $2,201,865, EBITDARM = $329,550
  Briar Glen (MRI):   91 GL rows, EGI = $3,763,229, EBITDARM = -$595,387
"""

import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

WORK = Path("/home/claude/work")

# ---------------------------------------------------------------------------
# Format-specific extractors (mini parsers)
# ---------------------------------------------------------------------------

DROP_PREFIXES = ("TOTAL ", "Total ", "NET ", "Net ", "Total - ")
DROP_EXACT = {"NET INCOME", "NET OPERATING INCOME", "Net Income",
              "Net Operating Income", "TOTAL REVENUE", "OPERATING EXPENSES",
              "REVENUE"}
DROP_LIST = {"Other Non Operating Revenue & Expense"}
DROP_KEYWORDS = ("EBITDA", "EBITDAR", "EBITDARM")


def is_grand_total(desc):
    if not desc:
        return False
    d = str(desc).strip()
    if d in DROP_EXACT or d in DROP_LIST:
        return True
    for p in DROP_PREFIXES:
        if d.startswith(p):
            return True
    for kw in DROP_KEYWORDS:
        if kw in d.upper():
            return True
    return False


def extract_yardi_salem(path):
    """Yardi 'Income to Budget' format. Account # in col A, description in col B,
    12 monthly amounts in cols C-N, total in col O. Month-end dates in row 9 cols C-N."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Month labels from row 9 cols C-N
    month_labels = []
    for c in range(3, 15):  # C..N
        v = ws.cell(row=9, column=c).value
        if isinstance(v, datetime):
            month_labels.append(v.strftime("%b %Y"))
        elif v is not None:
            # Try to parse as date string
            try:
                d = datetime.strptime(str(v).strip(), "%m/%d/%Y")
                month_labels.append(d.strftime("%b %Y"))
            except Exception:
                month_labels.append(str(v))
        else:
            month_labels.append("")

    rows = []
    for r in range(11, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value  # account # (with indent)
        b = ws.cell(row=r, column=2).value  # description
        if b is None:
            continue
        desc = str(b).strip()
        if not desc:
            continue
        # Only lines with an account number are GL detail in Yardi
        if a is None:
            continue
        a_str = str(a).strip()
        if not a_str.isdigit():
            continue  # section header or subtotal
        if is_grand_total(desc):
            continue

        monthly = []
        for c in range(3, 15):
            v = ws.cell(row=r, column=c).value
            try:
                monthly.append(float(v) if v not in (None, "", " ") else 0.0)
            except (TypeError, ValueError):
                monthly.append(0.0)

        # Drop rule: no $ value (all zero AND total zero)
        total_v = ws.cell(row=r, column=15).value
        try:
            total = float(total_v) if total_v not in (None, "", " ") else 0.0
        except (TypeError, ValueError):
            total = 0.0
        if all(v == 0 for v in monthly) and total == 0:
            continue

        rows.append({
            "account": a_str,
            "description": desc,
            "monthly": monthly,
            "total": total,
        })

    return rows, month_labels


def extract_mri_briar(path):
    """MRI R12MINCS format. No account #. Description in col A, 12 monthly
    amounts in cols B-M, total in col N. Month labels in row 11 cols B-M."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Month labels from row 11 cols B-M ('01/25' → Jan 2025)
    month_labels = []
    for c in range(2, 14):
        v = ws.cell(row=11, column=c).value
        if v is None:
            month_labels.append("")
            continue
        s = str(v).strip()
        try:
            d = datetime.strptime(s, "%m/%y")
            month_labels.append(d.strftime("%b %Y"))
        except Exception:
            month_labels.append(s)

    rows = []
    for r in range(12, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value  # description
        if a is None:
            continue
        desc = str(a).strip()
        if not desc:
            continue
        if is_grand_total(desc):
            continue

        # Read monthly + total
        monthly = []
        for c in range(2, 14):
            v = ws.cell(row=r, column=c).value
            try:
                monthly.append(float(v) if v not in (None, "", " ") else 0.0)
            except (TypeError, ValueError):
                monthly.append(0.0)
        total_v = ws.cell(row=r, column=14).value
        try:
            total = float(total_v) if total_v not in (None, "", " ") else 0.0
        except (TypeError, ValueError):
            total = 0.0

        # Drop rule: must have at least one non-zero $ value
        if all(v == 0 for v in monthly) and total == 0:
            continue

        rows.append({
            "account": "",
            "description": desc,
            "monthly": monthly,
            "total": total,
        })

    return rows, month_labels


# ---------------------------------------------------------------------------
# Write GL detail to T12 Input + recalc
# ---------------------------------------------------------------------------

def write_to_analyzer(template_path, output_path, gl_rows, month_labels):
    """Write GL detail to T12 Input!A12+, month labels to C11:N11."""
    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["T12 Input"]

    # Month labels at row 11 C-N
    for i, label in enumerate(month_labels):
        ws.cell(row=11, column=3 + i).value = label

    # GL detail starting at row 12
    for i, row in enumerate(gl_rows):
        r = 12 + i
        ws.cell(row=r, column=1).value = row["account"] or None
        ws.cell(row=r, column=2).value = row["description"]
        for j, v in enumerate(row["monthly"]):
            ws.cell(row=r, column=3 + j).value = v
        ws.cell(row=r, column=15).value = row["total"]

    wb.save(output_path)


def recalc_with_libreoffice(xlsx_path):
    """Force formula recalculation by round-tripping through LibreOffice."""
    out_dir = Path(xlsx_path).parent
    # LibreOffice headless conversion forces recalc on open
    cmd = [
        "libreoffice", "--headless", "--calc",
        "--convert-to", "xlsx",
        "--outdir", str(out_dir / "recalc"),
        str(xlsx_path),
    ]
    (out_dir / "recalc").mkdir(exist_ok=True)
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        print(f"LibreOffice stderr: {result.stderr}")
        raise RuntimeError("LibreOffice conversion failed")
    recalc_path = out_dir / "recalc" / Path(xlsx_path).name
    return recalc_path


def read_results(recalc_path):
    wb = openpyxl.load_workbook(recalc_path, data_only=True)
    ws_mt = wb["Monthly Trending"]
    egi = ws_mt.cell(row=20, column=14).value      # N20 = T12 EGI
    ebitdarm = ws_mt.cell(row=68, column=14).value # N68 = T12 EBITDARM

    # Coverage check on T12 Input
    ws_in = wb["T12 Input"]
    unmatched = 0
    matched = 0
    for r in range(12, 512):
        v = ws_in.cell(row=r, column=16).value
        if v == "UNMATCHED":
            unmatched += 1
            desc = ws_in.cell(row=r, column=2).value
            print(f"    UNMATCHED at row {r}: {desc!r}")
        elif v and v != "":
            matched += 1

    return {
        "egi": egi,
        "ebitdarm": ebitdarm,
        "matched": matched,
        "unmatched": unmatched,
    }


# ---------------------------------------------------------------------------
# Verification harness
# ---------------------------------------------------------------------------

def verify_case(name, source_path, extractor, expected, template_path):
    print(f"\n{'='*60}")
    print(f"  {name}")
    print(f"{'='*60}")

    rows, labels = extractor(source_path)
    print(f"  GL rows extracted: {len(rows)} (expected {expected['gl_rows']})")
    print(f"  Month labels: {labels}")

    out_path = WORK / f"verify_{name.lower().replace(' ', '_')}.xlsx"
    write_to_analyzer(template_path, out_path, rows, labels)
    print(f"  Wrote populated workbook → {out_path.name}")

    recalc_path = recalc_with_libreoffice(out_path)
    print(f"  Recalculated via LibreOffice")

    results = read_results(recalc_path)

    egi_match = abs((results["egi"] or 0) - expected["egi"]) < 1.0
    ebitdarm_match = abs((results["ebitdarm"] or 0) - expected["ebitdarm"]) < 1.0
    rows_match = len(rows) == expected["gl_rows"]
    no_unmatched = results["unmatched"] == 0

    print(f"\n  Results:")
    print(f"    GL rows:    {len(rows)}/{expected['gl_rows']}  "
          f"{'✓' if rows_match else '✗'}")
    print(f"    Matched:    {results['matched']}")
    print(f"    UNMATCHED:  {results['unmatched']}  "
          f"{'✓' if no_unmatched else '✗'}")
    print(f"    EGI:        ${results['egi']:>14,.2f}  "
          f"(expected ${expected['egi']:,.2f})  "
          f"{'✓' if egi_match else '✗'}")
    print(f"    EBITDARM:   ${results['ebitdarm']:>14,.2f}  "
          f"(expected ${expected['ebitdarm']:,.2f})  "
          f"{'✓' if ebitdarm_match else '✗'}")

    return all([rows_match, no_unmatched, egi_match, ebitdarm_match])


def main():
    template = WORK / "master_v014.xlsx"
    if not template.exists():
        print(f"ERROR: {template} not found — run migrate_analyzer.py first")
        sys.exit(1)

    salem_ok = verify_case(
        name="Salem",
        source_path=WORK / "salem.xlsx",
        extractor=extract_yardi_salem,
        expected={"gl_rows": 73, "egi": 2201864.71, "ebitdarm": 329549.93},
        template_path=template,
    )

    briar_ok = verify_case(
        name="Briar Glen",
        source_path=WORK / "briar.xlsx",
        extractor=extract_mri_briar,
        expected={"gl_rows": 91, "egi": 3763228.77, "ebitdarm": -595387.41},
        template_path=template,
    )

    print(f"\n{'='*60}")
    print(f"  OVERALL: {'PASSED' if salem_ok and briar_ok else 'FAILED'}")
    print(f"{'='*60}")
    sys.exit(0 if (salem_ok and briar_ok) else 1)


if __name__ == "__main__":
    main()
