"""
migrate_to_v0210.py — Substrate template v0.2.9 → v0.2.10

Opens UW-BACKLOG **BL-0023** (AR & Collections module). Inserts a new
"AR & Collections" analytical sheet at index 8 (between Monthly Trending
and UW Output) and extends Workbook Health with conditional AR-source
logic + an optional P5 pre-export gate.

The AR sheet is created HIDDEN by default. ar_writer.py (Track 1) flips
sheet_state to 'visible' when an AR aging file is uploaded for the
analysis cycle. When no AR file is uploaded, the sheet stays hidden and
every downstream integration falls back to its pre-AR behavior — so this
migration is fully additive (no regression for non-AR analyses).

KEY ARCHITECTURAL DECISIONS (locked with user 2026-05-23):

  1. Sheet position: index 8 (data → reconcile → trend → AR review → export).
  2. Sheet hidden by default; visibility flipped by ar_writer.py per cycle.
  3. AR-presence indicator: 'AR & Collections'!Z1 holds an integer flag
     (0 = no AR data, 1 = AR populated). ar_writer.py sets Z1=1 when it
     writes data; the migration creates Z1=0.
  4. Workbook Health B43 (Total outstanding AR) wraps the existing
     RR-derived SUM in an IF guard: when Z1=1, reads from AR sheet's
     Total AR (C15); otherwise keeps the SUM('Rent Roll Input'!X) formula.
  5. P5 gate added at Workbook Health row 52, READY-FOR-EXPORT summary
     shifts from row 52 → row 53. Safe because no external references
     to Workbook Health!B52 exist (verified pre-migration).
  6. P5 evaluates: when Z1=0 → "✓" (no AR required); when Z1=1 → compares
     AR sheet's as-of date (C3) to RR_Period_Date named range.
  7. Payer buckets on §3: 7 rows reflecting all mappings.py normalization
     targets (Private Pay, Medicaid, Medicare, Managed Care, VA Benefit,
     LTC Insurance, Self-Pay + Other). AR ingest constructs MappingSet
     with payer_fallback="Self-Pay + Other".
  8. Cross-sheet pins:
       Monthly Trending!N26  -> annualized EGI (DSO + AR÷EGI denominators)
       T12 Analytics!E7      -> avg occupied beds (per-bed metric)
       T12 Analytics!E98     -> bad debt expense (variance cross-check)
       RR_Period_Date        -> named range for P5 period comparison

OPERATIONS:

  A. Insert 'AR & Collections' sheet at index 8 (hidden), programmatically
     populated with all 5 sections from the spec + Z1 presence flag +
     AZ1:AZ5 anchor block.
  B. Wrap Workbook Health!B43 in IF guard (AR-presence pivot).
  C. Move Workbook Health row 52 (summary) → row 53; insert P5 at row 52.
     Rewrite the now-row-53 summary formula to AND in B52.
  D. Stamp Cover!B8 → v0.2.10 + AZ4 on all 16 sheets (anchor list grows
     15 → 16 because AR & Collections joins).

Idempotency:
  - Gate: Cover!B8 == "v0.2.10" AND "AR & Collections" sheet at index 8.
  - All step-internal mutations are guarded so partial-migration state
    is recoverable. Re-running on a migrated workbook is a no-op
    (re-save only).

Why no template asset (unlike v0.2.4 / v0.2.7):
  - AR sheet has no charts, only formula references and labels. ~110
    populated cells vs Dashboard's 437 + 6 charts. Programmatic
    construction is reasonable and keeps the sheet source-controlled
    as code rather than as an opaque .xlsx blob.

BUNDLED FILE STATUS:
  - As of 2026-05-23 the bundled ALF_Financial_Analyzer_Only.xlsx is at
    Cover!B8 = "v0.2.4" (user-managed per BL-0021). This migration does
    NOT update the bundled file. Bundled-file AR support requires either
    (a) forward-rolling the bundled through v0.2.5 → v0.2.10, or
    (b) a separate one-off bundled patch script — TBD with user.

Usage:
    python tools/migration/migrate_to_v0210.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SUBSTRATE_FROM = "v0.2.9"
SUBSTRATE_TO = "v0.2.10"

NEW_SHEET = "AR & Collections"
NEW_SHEET_INDEX = 8  # between Monthly Trending (7) and UW Output

# 16-sheet anchor list — AR & Collections inserted at index 8.
ANCHOR_SHEETS = (
    "Cover", "Dashboard",
    "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending",
    "AR & Collections",                                              # NEW
    "UW Output", "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# AR sheet anchor metadata
ANCHOR_PURPOSE = "AR aging, collections KPIs, roll-forward & bad-debt reconciliation"
ANCHOR_CATEGORY = "Analytical (diligence)"
ANCHOR_VISIBILITY = "hidden (visible when AR uploaded)"
ANCHOR_NOTES = (
    "Sheet populated by ar_writer.py per analysis cycle. Z1 = AR presence "
    "flag (0=no AR data, 1=populated). Workbook Health B43 + P5 gate "
    "read Z1 to decide RR-fallback vs AR-source behavior. Hidden by "
    "default so non-AR analyses see no surface change."
)

# Payer order — must match mappings.py normalization targets to render
# correctly when ar_writer.py SUMIFs by payer.
PAYER_ROWS = (
    "Private Pay",
    "Medicaid",
    "Medicare",
    "Managed Care",
    "VA Benefit",
    "LTC Insurance",
    "Self-Pay + Other",
)

# Workbook Health row mutations
WH_AR_BALANCE_ROW = 43          # B43 wrap in IF guard
WH_P5_ROW = 52                  # NEW gate row
WH_SUMMARY_OLD_ROW = 52         # READY FOR EXPORT old position
WH_SUMMARY_NEW_ROW = 53         # READY FOR EXPORT new position

# Original B43 formula (we wrap, not replace — keep RR fallback exact)
RR_AR_SUM_FORMULA = "=SUM('Rent Roll Input'!$X$7:$X$606)"
RR_AR_SUM_FORMULA_INNER = "SUM('Rent Roll Input'!$X$7:$X$606)"

# B43 conditional: AR present → use AR Total AR; else → RR-derived SUM
WH_B43_NEW_FORMULA = (
    f"=IF('{NEW_SHEET}'!Z1=1,"
    f"'{NEW_SHEET}'!C15,"
    f"{RR_AR_SUM_FORMULA_INNER})"
)

# P5 formula: AR absent → ✓ (gate inert); AR present → compare as-of
# date (C3) to RR_Period_Date named range.
WH_P5_FORMULA = (
    f"=IF('{NEW_SHEET}'!Z1=0,\"✓\","
    f"IF('{NEW_SHEET}'!C3=RR_Period_Date,\"✓\","
    f"\"⚠ AR period ≠ RR period\"))"
)

# Original summary at row 52 (we move to row 53 and extend the AND)
WH_SUMMARY_NEW_FORMULA = (
    "=IF(AND(B48=\"✓\",B49=\"✓\",B50=\"✓\",B51=\"✓\",B52=\"✓\"),"
    "\"✓ READY — UW Export tab is good to copy\","
    "\"⚠ NOT READY — resolve the ⚠ items above first\")"
)


# ---------------------------------------------------------------------------
# Style helpers — minimal, taste-consistent with other analytical sheets.
# ---------------------------------------------------------------------------

NAVY = "FF1F4E79"
LIGHT_GRAY = "FFE7E6E6"
WHITE = "FFFFFFFF"
DARK_GRAY = "FF595959"
AMBER = "FFD97706"
BORDER_GRAY = "FFA6A6A6"

thin_bottom = Border(bottom=Side(style="thin", color=BORDER_GRAY))
thin_top = Border(top=Side(style="thin", color=BORDER_GRAY))
thin_top_bottom = Border(
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)

title_font = Font(name="Calibri", size=14, bold=True, color=WHITE)
title_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
section_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
subsection_font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
header_font = Font(name="Calibri", size=10, bold=True)
header_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
label_font = Font(name="Calibri", size=10)
note_font = Font(name="Calibri", size=9, italic=True, color=DARK_GRAY)

left_indent = Alignment(horizontal="left", vertical="center", indent=1)
left_align = Alignment(horizontal="left", vertical="center")
right_align = Alignment(horizontal="right", vertical="center")
center_align = Alignment(horizontal="center", vertical="center")


def _set(ws, ref, value, font=None, fill=None, align=None, border=None, number_format=None):
    cell = ws[ref]
    cell.value = value
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def _section_header(ws, row, text):
    ref_b = f"B{row}"
    _set(ws, ref_b, text, font=section_font, fill=section_fill, align=left_indent)
    # extend fill across B:F
    for col in ("C", "D", "E", "F"):
        cell = ws[f"{col}{row}"]
        cell.fill = section_fill
    ws.merge_cells(f"B{row}:F{row}")


def _subsection_header(ws, row, text):
    _set(ws, f"B{row}", text, font=subsection_font, align=left_align, border=thin_bottom)
    for col in ("C", "D", "E", "F"):
        ws[f"{col}{row}"].border = thin_bottom


# ---------------------------------------------------------------------------
def is_already_v0210(wb) -> bool:
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    if NEW_SHEET not in wb.sheetnames:
        return False
    if wb.sheetnames.index(NEW_SHEET) != NEW_SHEET_INDEX:
        return False
    return True


# ---------------------------------------------------------------------------
# Step A — build the AR & Collections sheet
# ---------------------------------------------------------------------------

def build_ar_sheet(wb) -> dict:
    """Construct the AR & Collections sheet from scratch.

    Returns counts dict for verification reporting.
    """
    n = {"cells_populated": 0, "sections": 0}

    ws = wb.create_sheet(NEW_SHEET)

    # Tab color — amber for diligence sheet, distinguishes from analytical navy.
    from openpyxl.styles.colors import Color
    ws.sheet_properties.tabColor = Color(rgb=AMBER)

    # Hide by default — ar_writer.py flips visible when AR uploaded.
    ws.sheet_state = "hidden"

    # Hide gridlines (matches Dashboard convention).
    ws.sheet_view.showGridLines = False

    # ---- Column widths ----
    widths = {"A": 3, "B": 48, "C": 16, "D": 14, "E": 14, "F": 22, "G": 3, "Z": 6}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---- Z1: AR presence flag ----
    # 0 by default; ar_writer.py sets to 1 when AR data is written.
    # Kept in column Z to be out of the way of section content but
    # discoverable by anyone inspecting the sheet.
    _set(ws, "Z1", 0, font=note_font, align=center_align,
         number_format="0")
    n["cells_populated"] += 1

    # ---- Title (row 1, merged B:F) ----
    _set(ws, "B1", "AR & COLLECTIONS",
         font=title_font, fill=title_fill, align=left_indent)
    for col in ("C", "D", "E", "F"):
        ws[f"{col}1"].fill = title_fill
    ws.merge_cells("B1:F1")
    ws.row_dimensions[1].height = 26
    n["cells_populated"] += 1

    # ---- Settings band (rows 3-5) ----
    # Row 3: AR as-of date
    _set(ws, "B3", "AR as-of date:", font=label_font, align=right_align)
    _set(ws, "C3", "=IFERROR(RR_Period_Date,\"\")",
         font=label_font, align=center_align, number_format="yyyy-mm-dd")
    _set(ws, "D3", "(analyst override OK)", font=note_font, align=left_align)
    n["cells_populated"] += 3

    # Row 4: Period basis
    _set(ws, "B4", "Period basis:", font=label_font, align=right_align)
    _set(ws, "C4", "Monthly", font=label_font, align=center_align)
    _set(ws, "D4", "Monthly / Quarterly / Annual", font=note_font, align=left_align)
    n["cells_populated"] += 3

    # Row 5: Variance tolerance
    _set(ws, "B5", "Variance tolerance:", font=label_font, align=right_align)
    _set(ws, "C5", 0.20, font=label_font, align=center_align, number_format="0.0%")
    _set(ws, "D5", "(default 20%, analyst-editable)", font=note_font, align=left_align)
    n["cells_populated"] += 3

    # ---- Section 1: Aging Summary (rows 7-18) ----
    _section_header(ws, 7, "1 · AGING SUMMARY")
    n["sections"] += 1

    _set(ws, "B8", "Bucket", font=header_font, fill=header_fill, align=left_indent, border=thin_top_bottom)
    _set(ws, "C8", "$ outstanding", font=header_font, fill=header_fill, align=right_align, border=thin_top_bottom)
    _set(ws, "D8", "% of total", font=header_font, fill=header_fill, align=right_align, border=thin_top_bottom)
    n["cells_populated"] += 3

    aging_rows = [
        (9, "Current (0-30 days)"),
        (10, "31-60 days"),
        (11, "61-90 days"),
        (12, "91-120 days"),
        (13, "Over 120 days"),
    ]
    for r, label in aging_rows:
        _set(ws, f"B{r}", label, font=label_font, align=left_indent)
        # C: $ amount — ar_writer.py populates; default 0
        _set(ws, f"C{r}", 0, font=label_font, align=right_align, number_format="#,##0")
        # D: % of total — formula
        _set(ws, f"D{r}", f"=IFERROR(C{r}/$C$15,0)",
             font=label_font, align=right_align, number_format="0.0%")
        n["cells_populated"] += 3

    # Row 15: TOTAL AR
    _set(ws, "B15", "TOTAL AR", font=header_font, align=left_indent, border=thin_top_bottom)
    _set(ws, "C15", "=SUM(C9:C13)", font=header_font, align=right_align,
         number_format="#,##0", border=thin_top_bottom)
    _set(ws, "D15", "=IF(C15=0,0,1)", font=header_font, align=right_align,
         number_format="0.0%", border=thin_top_bottom)
    n["cells_populated"] += 3

    # Row 17: 90+ subtotal
    _set(ws, "B17", "90+ subtotal", font=label_font, align=left_indent)
    _set(ws, "C17", "=C12+C13", font=label_font, align=right_align, number_format="#,##0")
    n["cells_populated"] += 2

    # Row 18: % of total in 90+
    _set(ws, "B18", "% of total in 90+", font=label_font, align=left_indent)
    _set(ws, "C18", "=IFERROR(C17/C15,0)", font=label_font, align=right_align, number_format="0.0%")
    n["cells_populated"] += 2

    # ---- Section 2: Key Ratios / KPIs (rows 20-26) ----
    _section_header(ws, 20, "2 · KEY RATIOS / KPIs")
    n["sections"] += 1

    _set(ws, "B21", "KPI", font=header_font, fill=header_fill, align=left_indent, border=thin_top_bottom)
    _set(ws, "C21", "Value", font=header_font, fill=header_fill, align=right_align, border=thin_top_bottom)
    _set(ws, "D21", "Threshold flag", font=header_font, fill=header_fill, align=center_align, border=thin_top_bottom)
    _set(ws, "E21", "Threshold", font=header_font, fill=header_fill, align=left_indent, border=thin_top_bottom)
    n["cells_populated"] += 4

    # Row 22: DSO
    _set(ws, "B22", "DSO (days sales outstanding)", font=label_font, align=left_indent)
    _set(ws, "C22", "=IFERROR(C15/('Monthly Trending'!$N$26/365),0)",
         font=label_font, align=right_align, number_format="0.0")
    _set(ws, "D22",
         '=IF(C22=0,"—",IF(C22<30,"✓",IF(C22<=45,"⚠ "&TEXT(C22,"0")&" days","✗ "&TEXT(C22,"0")&" days")))',
         font=label_font, align=center_align)
    _set(ws, "E22", "< 30 ✓ · 30-45 ⚠ · > 45 ✗", font=note_font, align=left_indent)
    n["cells_populated"] += 4

    # Row 23: AR ÷ monthly EGI
    _set(ws, "B23", "AR ÷ monthly EGI", font=label_font, align=left_indent)
    _set(ws, "C23", "=IFERROR(C15/('Monthly Trending'!$N$26/12),0)",
         font=label_font, align=right_align, number_format="0.0%")
    _set(ws, "D23",
         '=IF(C23=0,"—",IF(C23<0.05,"✓","⚠ "&TEXT(C23,"0.0%")))',
         font=label_font, align=center_align)
    _set(ws, "E23", "< 5% ✓ (matches Workbook Health G10)", font=note_font, align=left_indent)
    n["cells_populated"] += 4

    # Row 24: % aged 90+
    _set(ws, "B24", "% aged 90+", font=label_font, align=left_indent)
    _set(ws, "C24", "=C18", font=label_font, align=right_align, number_format="0.0%")
    _set(ws, "D24",
         '=IF(C24=0,"—",IF(C24<0.05,"✓",IF(C24<=0.10,"⚠ "&TEXT(C24,"0.0%"),"✗ "&TEXT(C24,"0.0%"))))',
         font=label_font, align=center_align)
    _set(ws, "E24", "< 5% ✓ · 5-10% ⚠ · > 10% ✗", font=note_font, align=left_indent)
    n["cells_populated"] += 4

    # Row 25: Collection effectiveness — populated by ar_writer.py when present
    _set(ws, "B25", "Collection effectiveness", font=label_font, align=left_indent)
    _set(ws, "C25", "", font=label_font, align=right_align, number_format="0.0%")
    _set(ws, "D25",
         '=IF(C25="","—",IF(C25>=0.97,"✓","⚠ "&TEXT(C25,"0.0%")))',
         font=label_font, align=center_align)
    _set(ws, "E25", "> 97% target  (— if charges absent)", font=note_font, align=left_indent)
    n["cells_populated"] += 4

    # Row 26: Avg balance per occupied bed
    _set(ws, "B26", "Avg balance per occupied bed", font=label_font, align=left_indent)
    _set(ws, "C26", "=IFERROR(C15/'T12 Analytics'!$E$7,0)",
         font=label_font, align=right_align, number_format="#,##0")
    _set(ws, "D26", "", font=label_font, align=center_align)
    _set(ws, "E26", "(contextual benchmark)", font=note_font, align=left_indent)
    n["cells_populated"] += 4

    # ---- Section 3: By-Payer Mix (rows 28-37) ----
    _section_header(ws, 28, "3 · BY-PAYER MIX")
    n["sections"] += 1

    _set(ws, "B29", "Payer", font=header_font, fill=header_fill, align=left_indent, border=thin_top_bottom)
    _set(ws, "C29", "$ outstanding", font=header_font, fill=header_fill, align=right_align, border=thin_top_bottom)
    _set(ws, "D29", "% of total", font=header_font, fill=header_fill, align=right_align, border=thin_top_bottom)
    _set(ws, "E29", "% aged 90+", font=header_font, fill=header_fill, align=right_align, border=thin_top_bottom)
    _set(ws, "F29", "Concentration flag", font=header_font, fill=header_fill, align=center_align, border=thin_top_bottom)
    n["cells_populated"] += 5

    for i, payer in enumerate(PAYER_ROWS):
        r = 30 + i
        _set(ws, f"B{r}", payer, font=label_font, align=left_indent)
        _set(ws, f"C{r}", 0, font=label_font, align=right_align, number_format="#,##0")
        _set(ws, f"D{r}", f"=IFERROR(C{r}/$C$15,0)",
             font=label_font, align=right_align, number_format="0.0%")
        _set(ws, f"E{r}", 0, font=label_font, align=right_align, number_format="0.0%")
        # Flag when payer's 90+ rate > 2× facility-wide 90+ rate
        _set(ws, f"F{r}",
             f'=IF(C{r}=0,"",IF(E{r}>2*$C$18,"⚠ 2× facility 90+",""))',
             font=label_font, align=center_align)
        n["cells_populated"] += 5

    # ---- Section 4: Roll-Forward & Bad-Debt Reconciliation (rows 40-58) ----
    _section_header(ws, 40, "4 · ROLL-FORWARD & BAD-DEBT RECONCILIATION")
    n["sections"] += 1

    # Roll-forward block (rows 42-49)
    _subsection_header(ws, 41, "Roll-forward (period-over-period)")
    n["cells_populated"] += 1

    rollforward_rows = [
        (42, "Prior period AR", "", "(manual entry or prior-period file)"),
        (43, "+ Charges (period)", 0, "(optional — from charges_period col)"),
        (44, "− Collections (period)", 0, "(optional — from collections_period col)"),
        (45, "− Write-offs (period)", 0, "(optional — from writeoffs_period col)"),
        (46, "+/− Adjustments", 0, "(optional — from adjustments_period col)"),
    ]
    for r, label, value, note in rollforward_rows:
        _set(ws, f"B{r}", label, font=label_font, align=left_indent)
        _set(ws, f"C{r}", value, font=label_font, align=right_align, number_format="#,##0")
        _set(ws, f"E{r}", note, font=note_font, align=left_indent)
        n["cells_populated"] += 3

    # Row 47: Implied closing AR = prior + charges - collections - writeoffs + adjustments
    _set(ws, "B47", "= Implied closing AR", font=header_font, align=left_indent, border=thin_top)
    _set(ws, "C47", "=IFERROR(C42+C43-C44-C45+C46,0)",
         font=header_font, align=right_align, number_format="#,##0", border=thin_top)
    n["cells_populated"] += 2

    # Row 48: Reported closing AR (from §1)
    _set(ws, "B48", "Reported closing AR", font=label_font, align=left_indent)
    _set(ws, "C48", "=C15", font=label_font, align=right_align, number_format="#,##0")
    n["cells_populated"] += 2

    # Row 49: Reconciliation gap
    _set(ws, "B49", "RECONCILIATION GAP", font=header_font, align=left_indent, border=thin_top)
    _set(ws, "C49", "=C47-C48",
         font=header_font, align=right_align, number_format="#,##0;[Red]-#,##0", border=thin_top)
    _set(ws, "D49",
         '=IF(C15=0,"",IF(ABS(C49)>0.01*C15,"⚠ |gap| > 1% of AR","✓"))',
         font=header_font, align=center_align, border=thin_top)
    n["cells_populated"] += 3

    # Bad-debt cross-check block (rows 51-57)
    _subsection_header(ws, 51, "Bad-debt cross-check")
    n["cells_populated"] += 1

    _set(ws, "B52", "Variance tolerance", font=label_font, align=left_indent)
    _set(ws, "C52", "=C5", font=label_font, align=right_align, number_format="0.0%")
    _set(ws, "E52", "(refs settings band C5)", font=note_font, align=left_indent)
    n["cells_populated"] += 3

    _set(ws, "B53", "T-12 bad debt expense (annual)", font=label_font, align=left_indent)
    _set(ws, "C53", "=IFERROR('T12 Analytics'!$E$98,0)",
         font=label_font, align=right_align, number_format="#,##0")
    _set(ws, "E53", "(as reported — not modified)", font=note_font, align=left_indent)
    n["cells_populated"] += 3

    _set(ws, "B54", "Period write-offs (annualized)", font=label_font, align=left_indent)
    # Annualization: monthly × 12. Period basis from C4 (future: switch on monthly/quarterly/annual).
    _set(ws, "C54", "=C45*12", font=label_font, align=right_align, number_format="#,##0")
    _set(ws, "E54", "(period basis × 12 — monthly default)", font=note_font, align=left_indent)
    n["cells_populated"] += 3

    _set(ws, "B55", "Variance (T12 − annualized write-offs)", font=label_font, align=left_indent)
    _set(ws, "C55", "=C53-C54",
         font=label_font, align=right_align, number_format="#,##0;[Red]-#,##0")
    n["cells_populated"] += 2

    _set(ws, "B56", "Variance flag", font=header_font, align=left_indent, border=thin_top)
    _set(ws, "C56",
         '=IF(C53=0,"⚪ no T12 bad debt",'
         'IF(ABS(C55)<=C52*C53,"✓ within "&TEXT(C52,"0%")&" tolerance",'
         '"⚠ DIVERGENCE — investigate"))',
         font=header_font, align=left_indent, border=thin_top)
    n["cells_populated"] += 2

    _set(ws, "B57", "Implied reserve change", font=label_font, align=left_indent)
    _set(ws, "C57", "=-C55",
         font=label_font, align=right_align, number_format="#,##0;[Red]-#,##0")
    _set(ws, "E57", "(sign-reversed variance — context only)", font=note_font, align=left_indent)
    n["cells_populated"] += 3

    # ---- Section 5: Flags & Exceptions (rows 60-66) ----
    _section_header(ws, 60, "5 · FLAGS & EXCEPTIONS")
    n["sections"] += 1

    _set(ws, "B61", "Flag", font=header_font, fill=header_fill, align=left_indent, border=thin_top_bottom)
    _set(ws, "C61", "Count / status", font=header_font, fill=header_fill, align=right_align, border=thin_top_bottom)
    _set(ws, "E61", "Trigger", font=header_font, fill=header_fill, align=left_indent, border=thin_top_bottom)
    n["cells_populated"] += 3

    flag_rows = [
        (62, "Resident in 90+ with active concession", 0,
         "RR concession & AR 90+ > 0"),
        (63, "Vacant bed with non-zero AR", 0,
         "RR status=Vacant & AR > 0"),
        (64, "Payer-type concentration", 0,
         "Single payer > 60% of 90+ balance"),
        (65, "Sum-check mismatch (rows)", 0,
         "Row total_balance ≠ Σ buckets (ingest)"),
        (66, "Period-date mismatch", 0,
         "AR as-of date ≠ RR period date"),
    ]
    for r, label, value, trigger in flag_rows:
        _set(ws, f"B{r}", label, font=label_font, align=left_indent)
        _set(ws, f"C{r}", value, font=label_font, align=right_align, number_format="0")
        _set(ws, f"E{r}", trigger, font=note_font, align=left_indent)
        n["cells_populated"] += 3

    # ---- AZ1:AZ5 anchor block ----
    _set(ws, "AZ1", ANCHOR_PURPOSE)
    _set(ws, "AZ2", ANCHOR_CATEGORY)
    _set(ws, "AZ3", ANCHOR_VISIBILITY)
    _set(ws, "AZ4", SUBSTRATE_TO)
    _set(ws, "AZ5", ANCHOR_NOTES)
    n["cells_populated"] += 5

    # ---- Move to index 8 ----
    current_index = wb.sheetnames.index(NEW_SHEET)
    offset = NEW_SHEET_INDEX - current_index
    if offset != 0:
        wb.move_sheet(NEW_SHEET, offset=offset)

    return n


# ---------------------------------------------------------------------------
# Step B — wrap Workbook Health!B43 in AR-presence IF guard
# ---------------------------------------------------------------------------

def wrap_workbook_health_ar_balance(wb) -> bool:
    ws = wb["Workbook Health"]
    cur = ws[f"B{WH_AR_BALANCE_ROW}"].value
    # Idempotency: if already wrapped, no-op.
    if isinstance(cur, str) and f"'{NEW_SHEET}'!Z1=1" in cur:
        return False
    if cur != RR_AR_SUM_FORMULA:
        # Defensive: don't blindly overwrite if the formula has been
        # hand-edited to something else. Surface and skip.
        print(f"  WARN: Workbook Health!B{WH_AR_BALANCE_ROW} not the expected "
              f"RR-derived formula — skipping wrap. Got: {cur!r}")
        return False
    ws[f"B{WH_AR_BALANCE_ROW}"].value = WH_B43_NEW_FORMULA
    return True


# ---------------------------------------------------------------------------
# Step C — add P5 gate at row 52; move READY-FOR-EXPORT summary to row 53
# ---------------------------------------------------------------------------

def add_p5_gate(wb) -> dict:
    ws = wb["Workbook Health"]
    res = {"summary_moved": False, "p5_added": False, "summary_rewritten": False}

    # Idempotency: if A52 already says "P5", we've migrated.
    a52 = ws["A52"].value
    if isinstance(a52, str) and a52.startswith("P5"):
        return res

    # 1. Move row 52 (summary) → row 53.
    src_a = ws["A52"]
    src_b = ws["B52"]
    dst_a = ws["A53"]
    dst_b = ws["B53"]

    dst_a.value = src_a.value
    if src_a.has_style:
        dst_a.font = copy(src_a.font)
        dst_a.fill = copy(src_a.fill)
        dst_a.border = copy(src_a.border)
        dst_a.alignment = copy(src_a.alignment)
        dst_a.number_format = src_a.number_format
        dst_a.protection = copy(src_a.protection)

    # Rewrite the summary formula to AND in B52 (P5).
    dst_b.value = WH_SUMMARY_NEW_FORMULA
    if src_b.has_style:
        dst_b.font = copy(src_b.font)
        dst_b.fill = copy(src_b.fill)
        dst_b.border = copy(src_b.border)
        dst_b.alignment = copy(src_b.alignment)
        dst_b.number_format = src_b.number_format
        dst_b.protection = copy(src_b.protection)
    res["summary_moved"] = True
    res["summary_rewritten"] = True

    # 2. Clear original row 52 cells and stamp P5 in their place.
    ws["A52"].value = "P5 · AR period matches RR period (inert if no AR)"
    # Inherit A48's style for consistency with the other gates.
    a48 = ws["A48"]
    if a48.has_style:
        ws["A52"].font = copy(a48.font)
        ws["A52"].fill = copy(a48.fill)
        ws["A52"].border = copy(a48.border)
        ws["A52"].alignment = copy(a48.alignment)
        ws["A52"].number_format = a48.number_format

    ws["B52"].value = WH_P5_FORMULA
    b48 = ws["B48"]
    if b48.has_style:
        ws["B52"].font = copy(b48.font)
        ws["B52"].fill = copy(b48.fill)
        ws["B52"].border = copy(b48.border)
        ws["B52"].alignment = copy(b48.alignment)
        ws["B52"].number_format = b48.number_format
    res["p5_added"] = True

    return res


# ---------------------------------------------------------------------------
# Step D — stamp substrate version on Cover!B8 + every anchor sheet AZ4
# ---------------------------------------------------------------------------

def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ---------------------------------------------------------------------------
def verify_migration(wb) -> dict:
    r = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    r["new_sheet_exists"] = NEW_SHEET in wb.sheetnames
    r["new_sheet_position_ok"] = (
        r["new_sheet_exists"]
        and wb.sheetnames.index(NEW_SHEET) == NEW_SHEET_INDEX
    )
    r["sheet_count"] = len(wb.sheetnames)
    r["sheet_count_ok"] = r["sheet_count"] == 16

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    if r["new_sheet_exists"]:
        ws = wb[NEW_SHEET]
        r["ar_hidden_default"] = ws.sheet_state == "hidden"
        r["ar_z1_flag"] = ws["Z1"].value
        r["ar_z1_default_ok"] = r["ar_z1_flag"] == 0
        # Spot-check titles
        r["ar_title_ok"] = "AR & COLLECTIONS" in str(ws["B1"].value or "")
        r["ar_total_formula_ok"] = ws["C15"].value == "=SUM(C9:C13)"
        # Spot-check formula-references into the wider Analyzer
        r["ar_egi_ref_ok"] = "'Monthly Trending'!$N$26" in str(ws["C23"].value or "")
        r["ar_beds_ref_ok"] = "'T12 Analytics'!$E$7" in str(ws["C26"].value or "")
        r["ar_baddebt_ref_ok"] = "'T12 Analytics'!$E$98" in str(ws["C53"].value or "")
        # AZ anchor
        r["ar_az4_self"] = ws["AZ4"].value == SUBSTRATE_TO
    else:
        for k in ("ar_hidden_default", "ar_z1_default_ok", "ar_title_ok",
                  "ar_total_formula_ok", "ar_egi_ref_ok", "ar_beds_ref_ok",
                  "ar_baddebt_ref_ok", "ar_az4_self"):
            r[k] = False
        r["ar_z1_flag"] = None

    # Workbook Health B43 conditional wrap
    wh = wb["Workbook Health"]
    b43 = wh[f"B{WH_AR_BALANCE_ROW}"].value
    r["b43_wrapped"] = isinstance(b43, str) and f"'{NEW_SHEET}'!Z1=1" in b43
    r["b43_fallback_preserved"] = isinstance(b43, str) and RR_AR_SUM_FORMULA_INNER in b43

    # P5 at row 52
    a52 = wh["A52"].value
    b52 = wh["B52"].value
    r["p5_label_ok"] = isinstance(a52, str) and a52.startswith("P5")
    r["p5_formula_ok"] = isinstance(b52, str) and f"'{NEW_SHEET}'!Z1=0" in b52

    # Summary at row 53 includes B52
    b53 = wh["B53"].value
    r["summary_moved_ok"] = (
        isinstance(b53, str)
        and 'B52="✓"' in b53
        and "READY" in b53
    )
    # Original row 52 summary should no longer be the summary formula
    r["summary_not_at_52"] = not (isinstance(b52, str) and "READY" in b52)

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0210(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    cur_version = wb["Cover"]["B8"].value
    if cur_version != SUBSTRATE_FROM:
        print(f"  WARN: Cover!B8 = {cur_version!r}, expected {SUBSTRATE_FROM!r}. "
              f"Proceeding anyway — but consider forward-rolling first.")

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    # Step A
    if NEW_SHEET not in wb.sheetnames:
        n = build_ar_sheet(wb)
        print(f"  A: built '{NEW_SHEET}' at index {NEW_SHEET_INDEX} — "
              f"{n['cells_populated']} cells across {n['sections']} sections "
              f"(hidden by default)")
    else:
        print(f"  A: '{NEW_SHEET}' already exists, skipping build")
        if wb.sheetnames.index(NEW_SHEET) != NEW_SHEET_INDEX:
            cur = wb.sheetnames.index(NEW_SHEET)
            wb.move_sheet(NEW_SHEET, offset=NEW_SHEET_INDEX - cur)
            print(f"     repositioned to index {NEW_SHEET_INDEX}")

    # Step B
    wrapped = wrap_workbook_health_ar_balance(wb)
    print(f"  B: Workbook Health!B{WH_AR_BALANCE_ROW} AR-presence wrap: "
          f"{'applied' if wrapped else 'skipped (already wrapped or non-default formula)'}")

    # Step C
    res = add_p5_gate(wb)
    if res["p5_added"]:
        print(f"  C: added P5 gate at row {WH_P5_ROW}, "
              f"moved READY-FOR-EXPORT summary to row {WH_SUMMARY_NEW_ROW}")
    else:
        print(f"  C: P5 gate already present — skipped")

    # Step D
    stamp_versions(wb)
    print(f"  D: stamped substrate version -> {SUBSTRATE_TO} on "
          f"Cover!B8 + {len(ANCHOR_SHEETS)} AZ4 anchors")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:14s}                    : {r['cover_b8_ok']}")
    print(f"  '{NEW_SHEET}' sheet exists                : {r['new_sheet_exists']}")
    print(f"  Sheet at position {NEW_SHEET_INDEX}                            : {r['new_sheet_position_ok']}")
    print(f"  Sheet count = {r['sheet_count']} (expected 16)              : {r['sheet_count_ok']}")
    print(f"  AR sheet hidden by default                  : {r['ar_hidden_default']}")
    print(f"  AR!Z1 presence flag = 0                     : {r['ar_z1_default_ok']} (got {r['ar_z1_flag']!r})")
    print(f"  AR!B1 title 'AR & COLLECTIONS'              : {r['ar_title_ok']}")
    print(f"  AR!C15 = SUM(C9:C13)                        : {r['ar_total_formula_ok']}")
    print(f"  AR!C23 refs Monthly Trending!N26 (EGI)      : {r['ar_egi_ref_ok']}")
    print(f"  AR!C26 refs T12 Analytics!E7  (beds)        : {r['ar_beds_ref_ok']}")
    print(f"  AR!C53 refs T12 Analytics!E98 (bad debt)    : {r['ar_baddebt_ref_ok']}")
    print(f"  AR sheet AZ4 self-stamp = {SUBSTRATE_TO}            : {r['ar_az4_self']}")
    print(f"  WH!B43 wrapped in AR-presence IF            : {r['b43_wrapped']}")
    print(f"  WH!B43 RR fallback preserved                : {r['b43_fallback_preserved']}")
    print(f"  WH!A52 = 'P5 ...'                           : {r['p5_label_ok']}")
    print(f"  WH!B52 P5 formula references AR!Z1          : {r['p5_formula_ok']}")
    print(f"  WH!B53 = summary, includes B52              : {r['summary_moved_ok']}")
    print(f"  Original summary cleared from row 52        : {r['summary_not_at_52']}")
    print(f"  All 16 AZ4 = {SUBSTRATE_TO}                       : {r['az4_all']} ({r['az4_count']} sheets)")

    all_ok = all([
        r["cover_b8_ok"],
        r["new_sheet_exists"], r["new_sheet_position_ok"], r["sheet_count_ok"],
        r["ar_hidden_default"], r["ar_z1_default_ok"], r["ar_title_ok"],
        r["ar_total_formula_ok"], r["ar_egi_ref_ok"], r["ar_beds_ref_ok"],
        r["ar_baddebt_ref_ok"], r["ar_az4_self"],
        r["b43_wrapped"], r["b43_fallback_preserved"],
        r["p5_label_ok"], r["p5_formula_ok"],
        r["summary_moved_ok"], r["summary_not_at_52"],
        r["az4_all"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0210.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
