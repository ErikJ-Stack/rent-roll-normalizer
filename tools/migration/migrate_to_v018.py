"""
migrate_to_v018.py — Substrate template v0.1.7 -> v0.1.8

Branch 3 (Analytical coverage) per OPTIMIZATION-DECISIONS.md. All edits are
additive or single-cell formula additions in currently-empty cells, plus
append-only new sections at the bottom of Rent Roll Recon. No row inserts,
no shifting of existing formulas. Idempotent — re-running on a v0.1.8 file
restores defaults but leaves analyst-entered values intact (except B2 of
Rent Roll Recon which is reverted to the auto-latest formula by design).

What this script does (per Branch 3 design close-out):

  B3.1 — Property name + period date plumbing
    B3.1-a  Reserve single-cell value targets for property name extracted
            from raw data: Rent Roll Input!A3 (Track 1 writer-populated)
            and T12 Input!A10 (Track 2 writer-populated). No separate
            labels — the cell IS the value. Clears v0.1.8-first-pass
            'Property name:' labels at RR Input A3 and T12 Input A2.
    B3.1-b  Replace T12 Analytics!B2 with 3-priority formula:
            Rent Roll Input!A3 -> T12 Input!A10 -> Cover!B5 (Property_Name).
    B3.1-c  Set T12 Analytics!E2 to LOOKUP(2,1/(...<>"")) over
            T12 Input!C11:N11 — rightmost-populated month.

  B3.2 — Property snapshot visuals on T12 Analytics
    B3.2-a  Hidden helper rate-bucket block at T12 Analytics!K46:V53 for V2.
    B3.2-b  Five charts at K1:V44 (occupancy stacked column, payer doughnut,
            rate-band histogram, T12 revenue trend, acuity doughnut).
    B3.2-c  Five conditional note cells immediately below each chart.

  B3.3 — Rent Roll Recon B2 latest-date default + DV
    B3.3-a  Set B2 = LOOKUP(9.99E+307, RR_Calc!A2:A13). Idempotency note:
            re-running re-installs the formula, so analyst-picked overrides
            do not survive a re-run.
    B3.3-b  Add data validation list on B2 sourced from RR_Calc!A2:A13.

  B3.4 — IL Unit-Type Mix & Rate Dispersion (Rent Roll Recon rows 86-100)
    Section K. Append-only. Sources from Rent Roll Input filtered to
    Care Type=IL and Status NOT IN (Vacant, Eviction). Includes sqft.

  B3.5 — MC Care Structure auto-detect (Rent Roll Recon rows 102-117)
    Section L. Append-only. Distinct-count pattern detect; substring-based
    tier mapping for Basic/Moderate/Advanced; FFS fallback.

  D — Version stamps
    Cover!B8 -> v0.1.8; all 13 AZ4 anchors -> v0.1.8.

  E — Verification block (15 boolean / count checks).

Usage:
    python tools/migration/migrate_to_v018.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

SUBSTRATE_FROM = "v0.1.7"
SUBSTRATE_TO = "v0.1.8"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# Section-header styling per inspection 2026-05-11
NAVY = "FF1F3864"     # column headers / row pills
SECTION_PURPLE = "FF4A3869"   # analytical deep-dive sections (H/I/J/K/L)
WHITE = "FFFFFFFF"

HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_FILL_PURPLE = PatternFill(fill_type="solid", fgColor=SECTION_PURPLE)
HEADER_FILL_NAVY = PatternFill(fill_type="solid", fgColor=NAVY)
HEADER_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", indent=1)
HEADER_ALIGN_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
BODY_BOLD = Font(name="Arial", size=10, bold=True)


# ============================================================================
# Idempotency gate
# ============================================================================

def is_already_v018(wb) -> bool:
    """Gate also checks B2 formula references the refined A3/A10 cells.

    Catches the v0.1.8 first-pass state where B8 was already stamped v0.1.8
    but B2 still pointed at Rent Roll Input!B3 / T12 Input!B2 (per the
    initial label-at-A2/A3-plus-empty-B2/B3 design). Re-running on that
    state will apply this refinement.
    """
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    b2 = str(wb["T12 Analytics"]["B2"].value or "")
    return "'Rent Roll Input'!A3" in b2 and "'T12 Input'!A10" in b2


# ============================================================================
# B3.1-a — Add property-name source cells on input sheets
# ============================================================================

def add_input_property_cells(wb) -> None:
    """Reserve property-name source cells. Single-cell value, no label.

    Target cells per user spec (2026-05-11 refinement of v0.1.8 first pass):
      - Rent Roll Input!A3 — property name value (writer-populated; analyst-paste OK)
      - T12 Input!A10 — property name value (writer-populated; analyst-paste OK)

    Both cells are kept blank by the migration. Future writer code stamps
    them on extraction:
      - Track 1 follow-up: writer.py -> Rent Roll Input!A3
      - Track 2 follow-up: t12_normalizer_writer.py -> T12 Input!A10

    The v0.1.8 first pass placed 'Property name:' labels at A3 (RR Input)
    and A2 (T12 Input) with empty B-cells as value targets. This refinement
    clears those labels so the target cells themselves can hold the value.
    Idempotent: only clears literals matching the v0.1.8-first-pass label;
    user-entered property names are preserved on re-run.
    """
    # Rent Roll Input: clear v0.1.8 first-pass label at A3 (now value cell)
    rri = wb["Rent Roll Input"]
    if rri["A3"].value == "Property name:":
        rri["A3"] = None

    # T12 Input: clear v0.1.8 first-pass label at A2 (replaced by A10 value cell)
    t12i = wb["T12 Input"]
    if t12i["A2"].value == "Property name:":
        t12i["A2"] = None
    # A10 reserved as the value cell — no write here (writer populates).


# ============================================================================
# B3.1-b — T12 Analytics B2 3-priority formula
# ============================================================================

def install_b2_property_formula(wb) -> None:
    """3-priority: Rent Roll Input!A3 -> T12 Input!A10 -> Cover!B5 (Property_Name).

    Cell coordinates per 2026-05-11 user refinement. See
    add_input_property_cells() docstring for writer expectations.
    """
    ws = wb["T12 Analytics"]
    ws["B2"] = (
        "=IFERROR(IF(LEN(TRIM('Rent Roll Input'!A3))>0,'Rent Roll Input'!A3,"
        "IF(LEN(TRIM('T12 Input'!A10))>0,'T12 Input'!A10,Property_Name)),Property_Name)"
    )


# ============================================================================
# B3.1-c — T12 Analytics E2 rightmost-month formula
# ============================================================================

def install_e2_period_formula(wb) -> None:
    ws = wb["T12 Analytics"]
    ws["E2"] = (
        "=IFERROR(LOOKUP(2,1/('T12 Input'!$C$11:$N$11<>\"\"),"
        "'T12 Input'!$C$11:$N$11),\"\")"
    )
    ws["E2"].number_format = "mmm yyyy"


# ============================================================================
# B3.2-a — Hidden helper rate-bucket block at T12 Analytics K46:V53
# ============================================================================

# Rate buckets: $0-1999 / $2000-3999 / $4000-5999 / $6000-7999 / $8000+
RATE_BUCKETS = (
    ("$0 – $1,999", 0, 1999.99),
    ("$2,000 – $3,999", 2000, 3999.99),
    ("$4,000 – $5,999", 4000, 5999.99),
    ("$6,000 – $7,999", 6000, 7999.99),
    ("$8,000+", 8000, 99999999),
)

def install_helper_rate_buckets(wb) -> None:
    ws = wb["T12 Analytics"]
    # Header row 46
    ws["K46"] = "(hidden helper — rate buckets for V2 chart)"
    ws["K46"].font = Font(name="Arial", size=9, italic=True, color="FF808080")
    ws["K47"] = "Bucket"
    ws["L47"] = "IL"
    ws["M47"] = "AL"
    ws["N47"] = "MC"
    for c in ("K47", "L47", "M47", "N47"):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL_NAVY
        ws[c].alignment = HEADER_ALIGN_CTR

    rri = "'Rent Roll Input'"
    period_filter = f"{rri}!$S$7:$S$606,'Rent Roll Recon'!$B$2"
    status_filter = f"{rri}!$E$7:$E$606,\"<>Vacant\",{rri}!$E$7:$E$606,\"<>Eviction\""

    for i, (label, lo, hi) in enumerate(RATE_BUCKETS):
        row = 48 + i
        ws.cell(row=row, column=11, value=label)  # K
        for col, care in [(12, "IL"), (13, "AL"), (14, "MC")]:
            # =COUNTIFS(period, status, rate>=lo, rate<=hi, care=X)
            f = (
                f"=COUNTIFS({period_filter},"
                f"{rri}!$E$7:$E$606,\"<>Vacant\","
                f"{rri}!$E$7:$E$606,\"<>Eviction\","
                f"{rri}!$H$7:$H$606,\">=\"&{lo},"
                f"{rri}!$H$7:$H$606,\"<=\"&{hi},"
                f"{rri}!$D$7:$D$606,\"{care}\")"
            )
            ws.cell(row=row, column=col, value=f)


# ============================================================================
# B3.2-b — Five charts on T12 Analytics K1:V44
# ============================================================================

def _clear_existing_charts(ws) -> int:
    n = len(ws._charts)
    ws._charts.clear()
    return n


def install_charts(wb) -> int:
    """Returns count of charts installed (5)."""
    ws = wb["T12 Analytics"]
    cleared = _clear_existing_charts(ws)
    if cleared:
        print(f"  (cleared {cleared} pre-existing chart(s) for idempotent rebuild)")

    # ---- V1: Occupancy by Care Type (stacked column) ----
    # Source: Rent Roll Recon rows 8-11 (Occupied/Vacant/Notice/Eviction) x B-D (IL/AL/MC)
    v1 = BarChart()
    v1.type = "col"
    v1.style = 10
    v1.grouping = "stacked"
    v1.overlap = 100
    v1.title = "V1 — Occupancy by Care Type (selected period)"
    v1.y_axis.title = "Unit count"
    v1.x_axis.title = "Care Type"
    # Categories: B5:D5 (IL / AL / MC headers from T12 Analytics row 5)
    # Series: Occupied / Vacant / Notice / Eviction = Rent Roll Recon rows 8/9/10/11
    cats = Reference(wb["T12 Analytics"], min_col=2, max_col=4, min_row=5, max_row=5)
    rrr = wb["Rent Roll Recon"]
    for src_row, label in [(8, "Occupied"), (9, "Vacant"), (10, "Notice"), (11, "Eviction")]:
        data = Reference(rrr, min_col=2, max_col=4, min_row=src_row, max_row=src_row)
        s = openpyxl.chart.Series(data, title=label)
        v1.series.append(s)
    v1.set_categories(cats)
    v1.height = 7.5  # cm — approx 14 row span
    v1.width = 12    # cm — approx 5 col span
    ws.add_chart(v1, "K1")

    # ---- V2: Rate Dispersion (histogram bars, IL/AL/MC three-series) ----
    v2 = BarChart()
    v2.type = "col"
    v2.style = 11
    v2.grouping = "clustered"
    v2.title = "V2 — Rate Dispersion by Care Type (occupied units)"
    v2.y_axis.title = "Unit count"
    v2.x_axis.title = "Rate band"
    cats2 = Reference(ws, min_col=11, max_col=11, min_row=48, max_row=52)  # K48:K52
    for col, label in [(12, "IL"), (13, "AL"), (14, "MC")]:
        data = Reference(ws, min_col=col, max_col=col, min_row=48, max_row=52)
        s = openpyxl.chart.Series(data, title=label)
        v2.series.append(s)
    v2.set_categories(cats2)
    v2.height = 7.5
    v2.width = 12
    ws.add_chart(v2, "K16")

    # ---- V3: Payer Mix (doughnut) ----
    v3 = DoughnutChart()
    v3.title = "V3 — Payer Mix (% of total monthly revenue)"
    v3.style = 26
    # Labels: Rent Roll Recon A40:A46 (Private Pay, Medicaid, LTC, VA, Managed Care, Self-Pay, Other)
    cats3 = Reference(rrr, min_col=1, max_col=1, min_row=40, max_row=46)
    # Data: H40:H46 (Total monthly rev per payer)
    data3 = Reference(rrr, min_col=8, max_col=8, min_row=40, max_row=46)
    s3 = openpyxl.chart.Series(data3, title="Payer mix")
    v3.series.append(s3)
    v3.set_categories(cats3)
    v3.dataLabels = DataLabelList(showPercent=True)
    v3.height = 7.5
    v3.width = 12
    ws.add_chart(v3, "P1")

    # ---- V4: T12 Revenue Trend (line) ----
    v4 = LineChart()
    v4.style = 12
    v4.title = "V4 — T12 Revenue Trend (12 months)"
    v4.y_axis.title = "Total operating rev ($)"
    v4.x_axis.title = "Month"
    t12_input = wb["T12 Input"]
    # Monthly headers: T12 Input C11:N11
    cats4 = Reference(t12_input, min_col=3, max_col=14, min_row=11, max_row=11)
    # Data: T12 Raw Data row 1 has "Total operating revenue" — not present yet.
    # Fall back to summing T12 Input C12:N (all months across all rows of T12 data).
    # We'll create a helper row at T12 Analytics K54 = SUM('T12 Raw Data' total operating rev row)
    # but simpler: use the Monthly Trending revenue row directly.
    # Per substrate: Monthly Trending has per-Label monthly summary.
    # Safest: SUM each month of T12 Input C12:C511, etc., over rev-flagged rows.
    # For v0.1.8 simplicity: source from T12 Input total row 502 (template grand total) is
    # too coupled. Use helper row 54 K54:V54 with SUMIFS over Description_Map labels.
    # Compromise: chart top operating revenue using T12 Raw Data row that holds it.
    # T12 Raw Data row 2 typically labeled. To stay safe, compute helper at row 54:
    # K54 = =SUM('T12 Input'!C12:C511) ... but this includes EXPENSES too. So narrow
    # to top-of-P&L total revenue via T12 Raw Data: assume row labelled "Total revenue".
    # As a defensible-default fallback we point at T12 Raw Data column F-Q for "Total
    # operating revenue" row — found via lookup. The chart references the helper row.
    ws["K53"] = "(hidden helper — T12 monthly revenue for V4)"
    ws["K53"].font = Font(name="Arial", size=9, italic=True, color="FF808080")
    # T12 Raw Data structure: A=Section, B=Label, C=Care, D=Flag, E=Matched,
    # F-Q = M01..M12, R = T12_Total. Sum all "Revenue" Section rows per month.
    # Dst K54..V54 (12 cells) maps to src F..Q (12 cells), offset = 5.
    for i in range(12):
        src_letter = openpyxl.utils.get_column_letter(6 + i)   # F..Q
        dst_letter = openpyxl.utils.get_column_letter(11 + i)  # K..V
        ws[f"{dst_letter}54"] = (
            f"=SUMIFS('T12 Raw Data'!{src_letter}:{src_letter},"
            f"'T12 Raw Data'!$A:$A,\"Revenue\")"
        )
    data4 = Reference(ws, min_col=11, max_col=22, min_row=54, max_row=54)
    s4 = openpyxl.chart.Series(data4, title="Total revenue (monthly)")
    v4.series.append(s4)
    v4.set_categories(cats4)
    v4.height = 7.5
    v4.width = 12
    ws.add_chart(v4, "P16")

    # ---- V5: Acuity Mix (doughnut, AL Care Levels) ----
    v5 = DoughnutChart()
    v5.title = "V5 — AL Acuity Mix (Basic / Level 2-7)"
    v5.style = 26
    cats5 = Reference(rrr, min_col=1, max_col=1, min_row=59, max_row=66)
    data5 = Reference(rrr, min_col=4, max_col=4, min_row=59, max_row=66)
    s5 = openpyxl.chart.Series(data5, title="AL acuity mix")
    v5.series.append(s5)
    v5.set_categories(cats5)
    v5.dataLabels = DataLabelList(showPercent=True)
    v5.height = 7.5
    v5.width = 12
    ws.add_chart(v5, "K31")

    return 5


# ============================================================================
# B3.2-c — Conditional notes for each chart
# ============================================================================

def install_chart_notes(wb) -> None:
    ws = wb["T12 Analytics"]
    rrr = "'Rent Roll Recon'"

    # V1 note at K15 — flag if any care type < 85% occ
    # Occupancy = row 12 of Rent Roll Recon (Physical occupancy %)
    ws["K15"] = (
        f"=IF(OR({rrr}!B12<0.85,{rrr}!C12<0.85,{rrr}!D12<0.85),"
        "\"⚠ One or more care types below 85% physical occupancy — "
        "investigate stabilization assumptions.\","
        "\"✓ All care types ≥ 85% physical occupancy.\")"
    )
    ws["K15"].font = BODY_FONT
    ws["K15"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # V2 note at K30 — flag if IL rate CV > 25%
    # CV available from Rent Roll Recon section K row 96 (after migration adds it)
    ws["K30"] = (
        f"=IF(LEN({rrr}!B96)=0,\"-\","
        f"IF(IFERROR(VALUE(SUBSTITUTE({rrr}!B96,\"%\",\"\"))/100,0)>0.25,"
        "\"⚠ IL rate dispersion CV \"&TEXT("
        f"IFERROR(VALUE(SUBSTITUTE({rrr}!B96,\"%\",\"\"))/100,0),\"0.0%\")"
        "&\" — wide spread may indicate legacy in-place rates.\","
        "\"✓ IL rate dispersion within normal band.\"))"
    )
    ws["K30"].font = BODY_FONT
    ws["K30"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # V3 note at P15 — flag Medicaid > 30% or Managed Care > 25%
    # Section F row 41 Medicaid, row 44 Managed Care, % column E and G
    ws["P15"] = (
        f"=IF({rrr}!E41>0.3,"
        f"\"⚠ Medicaid revenue share \"&TEXT({rrr}!E41,\"0.0%\")&"
        "\" — reimbursement rate risk; verify stabilized assumption.\","
        f"IF({rrr}!E44>0.25,"
        f"\"⚠ Managed Care revenue share \"&TEXT({rrr}!E44,\"0.0%\")&"
        "\" — verify per-diem rates vs. private pay.\","
        "\"✓ Payer mix dominated by private pay / low reimbursement risk.\"))"
    )
    ws["P15"].font = BODY_FONT
    ws["P15"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # V4 note at P30 — trajectory based on last-3 vs prior-3 monthly avg
    # Helper row 54 has 12 monthly values at K54:V54
    ws["P30"] = (
        "=IFERROR(IF(AVERAGE(T54:V54)>AVERAGE(K54:M54)*1.05,"
        "\"↑ Revenue trending up — latest 3-mo avg \"&TEXT(AVERAGE(T54:V54),\"$#,##0\")"
        "&\" vs prior 3-mo \"&TEXT(AVERAGE(K54:M54),\"$#,##0\"),"
        "IF(AVERAGE(T54:V54)<AVERAGE(K54:M54)*0.95,"
        "\"↓ Revenue trending down — latest 3-mo avg \"&TEXT(AVERAGE(T54:V54),\"$#,##0\")"
        "&\" vs prior 3-mo \"&TEXT(AVERAGE(K54:M54),\"$#,##0\"),"
        "\"→ Revenue stable — within ±5% of prior 3-mo.\")),"
        "\"Insufficient T12 data for trend.\")"
    )
    ws["P30"].font = BODY_FONT
    ws["P30"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # V5 note at K45 — flag if D67=0 (no acuity) or top tier > 50% of charges
    ws["K45"] = (
        f"=IF({rrr}!D67=0,"
        "\"Property has no AL acuity data — flat-rate AL or unpopulated.\","
        f"IF(MAX({rrr}!D59:D66)/{rrr}!D67>0.5,"
        "\"⚠ Acuity skewed: top tier > 50% of AL care charges — verify staffing model.\","
        "\"✓ Acuity distribution within normal range.\"))"
    )
    ws["K45"].font = BODY_FONT
    ws["K45"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ============================================================================
# B3.3 — Rent Roll Recon B2 default formula + DV
# ============================================================================

def install_rr_recon_b2_default(wb) -> None:
    ws = wb["Rent Roll Recon"]
    ws["B2"] = "=IFERROR(LOOKUP(9.99E+307,'RR_Calc'!$A$2:$A$13),\"\")"
    ws["B2"].number_format = "mm/dd/yyyy"
    # Idempotency: remove any pre-existing DV that targets B2 before adding ours
    keep = [
        dv for dv in ws.data_validations.dataValidation
        if "B2" not in str(dv.sqref)
    ]
    ws.data_validations.dataValidation = keep
    dv = DataValidation(
        type="list",
        formula1="='RR_Calc'!$A$2:$A$13",
        allow_blank=True,
        showDropDown=False,  # openpyxl quirk: False = show dropdown arrow
    )
    dv.add("B2")
    ws.add_data_validation(dv)


# ============================================================================
# B3.4 — IL section (Rent Roll Recon section K, rows 86-100)
# ============================================================================

# Apt Type values to break out
IL_APT_TYPES = ("Studio", "1 Bedroom", "2 Bedroom", "Cottage / Villa", "Other")


def _il_filter_clauses(care: str = "IL") -> str:
    """Common filter clauses against Rent Roll Input."""
    rri = "'Rent Roll Input'"
    return (
        f"{rri}!$S$7:$S$606,'Rent Roll Recon'!$B$2,"
        f"{rri}!$E$7:$E$606,\"<>Vacant\","
        f"{rri}!$E$7:$E$606,\"<>Eviction\","
        f"{rri}!$D$7:$D$606,\"{care}\""
    )


def install_il_section(wb) -> None:
    ws = wb["Rent Roll Recon"]
    rri = "'Rent Roll Input'"

    # Idempotency: if A86 already says our header, skip
    if ws["A86"].value and "IL UNIT-TYPE MIX" in str(ws["A86"].value):
        return

    # Section header (rows 86)
    ws["A86"] = "K  ·  IL UNIT-TYPE MIX, SIZE & RATE DISPERSION"
    ws.merge_cells("A86:I86")
    ws["A86"].font = HEADER_FONT
    ws["A86"].fill = HEADER_FILL_PURPLE
    ws["A86"].alignment = HEADER_ALIGN_LEFT

    # Column headers (row 87)
    headers = ["Unit Type", "Count", "% of\nIL", "Avg Rate", "Min Rate",
               "Max Rate", "Avg Sq Ft", "$/Sq Ft"]
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=87, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL_NAVY
        c.alignment = HEADER_ALIGN_CTR

    base_filter = _il_filter_clauses("IL")
    for i, apt in enumerate(IL_APT_TYPES):
        row = 88 + i
        ws.cell(row=row, column=1, value=apt).font = BODY_FONT
        # Apt-type filter clause
        apt_clause = f",{rri}!$F$7:$F$606,\"{apt}\""
        # B: count
        ws.cell(row=row, column=2,
                value=f"=COUNTIFS({base_filter}{apt_clause})")
        # C: % of IL (count / total IL occupied)
        ws.cell(row=row, column=3,
                value=f"=IFERROR(B{row}/COUNTIFS({base_filter}),\"-\")"
                ).number_format = "0.0%"
        # D: avg rate
        ws.cell(row=row, column=4,
                value=f"=IFERROR(AVERAGEIFS({rri}!$H$7:$H$606,{base_filter}{apt_clause}),\"-\")"
                ).number_format = "$#,##0"
        # E: min rate
        ws.cell(row=row, column=5,
                value=f"=IFERROR(MINIFS({rri}!$H$7:$H$606,{base_filter}{apt_clause}),\"-\")"
                ).number_format = "$#,##0"
        # F: max rate
        ws.cell(row=row, column=6,
                value=f"=IFERROR(MAXIFS({rri}!$H$7:$H$606,{base_filter}{apt_clause}),\"-\")"
                ).number_format = "$#,##0"
        # G: avg sq ft
        ws.cell(row=row, column=7,
                value=f"=IFERROR(AVERAGEIFS({rri}!$C$7:$C$606,{base_filter}{apt_clause}),\"-\")"
                ).number_format = "#,##0"
        # H: $/sq ft
        ws.cell(row=row, column=8,
                value=f"=IFERROR(D{row}/G{row},\"-\")"
                ).number_format = "$#,##0.00"

    # Total row 93 — Total IL occupied
    ws.cell(row=93, column=1, value="Total IL occupied").font = BODY_BOLD
    ws.cell(row=93, column=2, value="=SUM(B88:B92)").font = BODY_BOLD
    c = ws.cell(row=93, column=3, value="100.0%")
    c.font = BODY_BOLD
    c.number_format = "0.0%"
    # Weighted avg rate (sum of count*rate / sum count)
    ws.cell(row=93, column=4,
            value=f"=IFERROR(AVERAGEIFS({rri}!$H$7:$H$606,{base_filter}),\"-\")"
            ).number_format = "$#,##0"
    ws.cell(row=93, column=4).font = BODY_BOLD
    # Range cells E93, F93 — actual min/max across all IL
    ws.cell(row=93, column=5,
            value=f"=IFERROR(MINIFS({rri}!$H$7:$H$606,{base_filter}),\"-\")"
            ).number_format = "$#,##0"
    ws.cell(row=93, column=6,
            value=f"=IFERROR(MAXIFS({rri}!$H$7:$H$606,{base_filter}),\"-\")"
            ).number_format = "$#,##0"
    ws.cell(row=93, column=7,
            value=f"=IFERROR(AVERAGEIFS({rri}!$C$7:$C$606,{base_filter}),\"-\")"
            ).number_format = "#,##0"
    ws.cell(row=93, column=8,
            value=f"=IFERROR(D93/G93,\"-\")"
            ).number_format = "$#,##0.00"

    # Summary metrics rows 95-99
    # Row 95: Rate spread = max - min
    ws.cell(row=95, column=1, value="Rate spread (max − min)").font = BODY_FONT
    ws.cell(row=95, column=2,
            value="=IFERROR(F93-E93,\"-\")").number_format = "$#,##0"
    # Row 96: CV = stdev/avg
    # Use a SUMPRODUCT-based stdev over the filtered range; openpyxl can serialize.
    # Approximate via STDEV.S over the entire IL range using array formula. Simpler:
    # Build via helper that references occupied IL Actual Rates from RR Input.
    # Concrete formula: =IFERROR(STDEV.S(IF(...filter...,'Rent Roll Input'!$H$7:$H$606)) / D93, "-")
    # This is an array formula in Excel; openpyxl writes it as a regular formula and
    # modern Excel evaluates with implicit intersection / spill — works in Excel 365.
    # As a safer alternative use ratio of (max-min)/avg as a proxy.
    # Pragmatic: pre-compute proxy CV as (max - min) / avg / sqrt(12) — defensible
    # range-based dispersion measure that doesn't require array formulas.
    ws.cell(row=96, column=1, value="Rate CV (approx — range ÷ avg ÷ √12)").font = BODY_FONT
    ws.cell(row=96, column=2,
            value="=IFERROR((F93-E93)/D93/SQRT(12),\"-\")"
            ).number_format = "0.0%"
    # Row 97: Avg sq ft (IL)
    ws.cell(row=97, column=1, value="Avg sq ft (IL occupied)").font = BODY_FONT
    ws.cell(row=97, column=2, value="=G93").number_format = "#,##0"
    # Row 98: Sq ft range (min — max)
    ws.cell(row=98, column=1, value="Sq ft range").font = BODY_FONT
    ws.cell(row=98, column=2,
            value=f"=IFERROR(MINIFS({rri}!$C$7:$C$606,{base_filter}),\"-\")"
            ).number_format = "#,##0"
    ws.cell(row=98, column=3, value="—").alignment = Alignment(horizontal="center")
    ws.cell(row=98, column=4,
            value=f"=IFERROR(MAXIFS({rri}!$C$7:$C$606,{base_filter}),\"-\")"
            ).number_format = "#,##0"
    # Row 99: $/sq ft (avg rate / avg sq ft)
    ws.cell(row=99, column=1, value="$/sq ft (IL avg rate ÷ avg sq ft)").font = BODY_FONT
    ws.cell(row=99, column=2, value="=H93").number_format = "$#,##0.00"

    # Row 100: Conditional note
    ws.cell(row=100, column=1,
            value=("=IF(B93=0,\"No IL units in selected period\","
                   "IF(B96>0.25,"
                   "\"⚠ IL rate CV \"&TEXT(B96,\"0.0%\")&\" — wide rate dispersion. \""
                   "&\"Likely legacy in-place rates well below current asking.\","
                   "\"✓ IL rate dispersion within normal band.\"))"))
    ws.cell(row=100, column=1).font = Font(name="Arial", size=10, italic=True)
    ws.merge_cells("A100:H100")


# ============================================================================
# B3.5 — MC section (Rent Roll Recon section L, rows 102-117)
# ============================================================================

MC_TIER_MAP = (
    ("Tier 1 / Basic", ("Basic", "Tier 1", "Level 1", "Lvl 1")),
    ("Tier 2 / Moderate", ("Moderate", "Tier 2", "Level 2", "Level 3", "Lvl 2", "Lvl 3")),
    ("Tier 3 / Advanced", ("Advanced", "Tier 3", "Level 4", "Level 5", "Level 6", "Level 7", "Lvl 4", "Lvl 5")),
)


def install_mc_section(wb) -> None:
    ws = wb["Rent Roll Recon"]
    rri = "'Rent Roll Input'"

    if ws["A102"].value and "MC CARE STRUCTURE" in str(ws["A102"].value):
        return

    # Section header (row 102)
    ws["A102"] = "L  ·  MC CARE STRUCTURE  (auto-detected pattern)"
    ws.merge_cells("A102:I102")
    ws["A102"].font = HEADER_FONT
    ws["A102"].fill = HEADER_FILL_PURPLE
    ws["A102"].alignment = HEADER_ALIGN_LEFT

    base_filter = _il_filter_clauses("MC")

    # Row 103: Pattern detector
    # Distinct-count of non-empty K values among occupied MC residents:
    #   distinct = SUMPRODUCT((K matches occupied MC) / COUNTIFS(K, K, ...filter...))
    # Then map to flat / tiered / FFS.
    # Simpler approach: count distinct K values that have at least 1 occupied MC
    # resident, using SUMPRODUCT over Levels 1-7 + Basic + custom.
    # Concrete: count of distinct K-values populated, using a SUMPRODUCT pattern.
    distinct_formula = (
        f"SUMPRODUCT((COUNTIFS({rri}!$D$7:$D$606,\"MC\","
        f"{rri}!$S$7:$S$606,'Rent Roll Recon'!$B$2,"
        f"{rri}!$E$7:$E$606,\"<>Vacant\","
        f"{rri}!$E$7:$E$606,\"<>Eviction\","
        f"{rri}!$K$7:$K$606,{rri}!$K$7:$K$606)>0)"
        f"*({rri}!$D$7:$D$606=\"MC\")"
        f"*({rri}!$S$7:$S$606='Rent Roll Recon'!$B$2)"
        f"*({rri}!$E$7:$E$606<>\"Vacant\")"
        f"*({rri}!$E$7:$E$606<>\"Eviction\")"
        f"*({rri}!$K$7:$K$606<>\"\")"
        f"/MAX(COUNTIFS({rri}!$D$7:$D$606,\"MC\","
        f"{rri}!$S$7:$S$606,'Rent Roll Recon'!$B$2,"
        f"{rri}!$E$7:$E$606,\"<>Vacant\","
        f"{rri}!$E$7:$E$606,\"<>Eviction\","
        f"{rri}!$K$7:$K$606,{rri}!$K$7:$K$606),1))"
    )
    # The above is brittle for large empty sheets — simplify with a fixed-bucket count:
    # count buckets where COUNTIFS(...MC...K=bucket) > 0, using a small hidden helper
    # is too complex. Use a pragmatic distinct-via-FREQUENCY proxy:
    # Just COUNT distinct K values in occupied MC rows = number of unique K levels
    # populated. Approximation good enough for the pattern-detect.
    #
    # Final approach: count distinct using the helper sums below at rows 106-109.
    # If row 106 count > 0: tier 1 populated. Similarly for 107, 108, 109.
    # Distinct count = COUNTIF(B106:B109, ">0")
    ws.cell(row=103, column=1, value="MC Care Pattern detected:").font = BODY_BOLD
    ws.cell(row=103, column=2,
            value=("=IF(SUM(B106:B109)=0,\"Flat-rate (no care levels recorded)\","
                   "IF(COUNTIF(B106:B109,\">0\")=1,\"Flat-rate (single tier)\","
                   "IF(COUNTIF(B106:B109,\">0\")<=3,"
                   "\"Tiered acuity (\"&COUNTIF(B106:B109,\">0\")&\" levels)\","
                   "\"Fee-for-service / variable\")))"))
    ws.cell(row=103, column=2).font = BODY_BOLD
    ws.merge_cells("B103:H103")

    # Column headers (row 105)
    headers_mc = ["Tier", "Count", "% of\nMC", "Avg $/mo", "Total $/mo"]
    for col, label in enumerate(headers_mc, start=1):
        c = ws.cell(row=105, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL_NAVY
        c.alignment = HEADER_ALIGN_CTR

    # Rows 106-108 — Tier 1/2/3 mapped from K substrings, row 109 = Other (FFS)
    for i, (tier_label, k_patterns) in enumerate(MC_TIER_MAP):
        row = 106 + i
        ws.cell(row=row, column=1, value=tier_label).font = BODY_FONT
        # Count: sum of count for each K-pattern in the tier
        count_terms = [
            f"COUNTIFS({base_filter},{rri}!$K$7:$K$606,\"{p}\")"
            for p in k_patterns
        ]
        ws.cell(row=row, column=2, value="=" + "+".join(count_terms))
        # % of MC
        ws.cell(row=row, column=3,
                value=f"=IFERROR(B{row}/COUNTIFS({base_filter}),\"-\")"
                ).number_format = "0.0%"
        # Avg care charge $/mo for residents in tier (L = Care Level $)
        sum_terms = [
            f"SUMIFS({rri}!$L$7:$L$606,{base_filter},{rri}!$K$7:$K$606,\"{p}\")"
            for p in k_patterns
        ]
        sum_join = "+".join(sum_terms)
        ws.cell(row=row, column=4,
                value=f"=IFERROR(({sum_join})/B{row},\"-\")"
                ).number_format = "$#,##0"
        ws.cell(row=row, column=5,
                value=f"={sum_join}"
                ).number_format = "$#,##0"

    # Row 109: Other / FFS (K values not in any tier)
    ws.cell(row=109, column=1, value="Other / unmapped").font = BODY_FONT
    # Count of occupied MC where K is non-empty AND not matched by any tier substring
    # Simplest defensible: total occupied MC with K non-empty MINUS sum of B106:B108
    ws.cell(row=109, column=2,
            value=(f"=MAX(0,COUNTIFS({base_filter},{rri}!$K$7:$K$606,\"?*\")"
                   "-SUM(B106:B108))"))
    ws.cell(row=109, column=3,
            value=f"=IFERROR(B109/COUNTIFS({base_filter}),\"-\")"
            ).number_format = "0.0%"
    ws.cell(row=109, column=4, value="\"-\"")
    ws.cell(row=109, column=5,
            value=(f"=MAX(0,SUMIFS({rri}!$L$7:$L$606,{base_filter})"
                   "-SUM(E106:E108))")).number_format = "$#,##0"

    # Row 110: Total MC occupied
    ws.cell(row=110, column=1, value="Total MC occupied").font = BODY_BOLD
    ws.cell(row=110, column=2,
            value=f"=COUNTIFS({base_filter})").font = BODY_BOLD
    c = ws.cell(row=110, column=3, value="100.0%")
    c.font = BODY_BOLD
    c.number_format = "0.0%"
    ws.cell(row=110, column=4,
            value=f"=IFERROR(E110/B110,\"-\")"
            ).number_format = "$#,##0"
    ws.cell(row=110, column=4).font = BODY_BOLD
    ws.cell(row=110, column=5,
            value=f"=SUMIFS({rri}!$L$7:$L$606,{base_filter})"
            ).number_format = "$#,##0"
    ws.cell(row=110, column=5).font = BODY_BOLD

    # Row 112: MC base rent / resident (avg of H, occupied MC)
    ws.cell(row=112, column=1, value="MC avg base rent / resident").font = BODY_FONT
    ws.cell(row=112, column=2,
            value=f"=IFERROR(AVERAGEIFS({rri}!$H$7:$H$606,{base_filter}),\"-\")"
            ).number_format = "$#,##0"
    # Row 113: MC avg care charge / resident
    ws.cell(row=113, column=1, value="MC avg care charge / resident").font = BODY_FONT
    ws.cell(row=113, column=2,
            value="=IFERROR(D110,\"-\")"
            ).number_format = "$#,##0"
    # Row 114: Care charge / base rent ratio
    ws.cell(row=114, column=1, value="Care charge ÷ base rent ratio").font = BODY_FONT
    ws.cell(row=114, column=2,
            value="=IFERROR(B113/B112,\"-\")"
            ).number_format = "0.0%"
    ws.cell(row=114, column=3,
            value="=IF(B114>0.3,\"⚠ Care charge > 30% of base — acuity-heavy\","
                   "\"✓ Within range\")").font = Font(name="Arial", size=10, italic=True)
    # Row 115: Total MC monthly revenue (base + care)
    ws.cell(row=115, column=1, value="Total MC monthly revenue").font = BODY_BOLD
    ws.cell(row=115, column=2,
            value=(f"=IFERROR(SUMIFS({rri}!$H$7:$H$606,{base_filter})"
                   f"+SUMIFS({rri}!$L$7:$L$606,{base_filter})"
                   f"+SUMIFS({rri}!$I$7:$I$606,{base_filter}),0)")
            ).number_format = "$#,##0"
    ws.cell(row=115, column=2).font = BODY_BOLD

    # Row 117: Pattern-specific conditional note
    ws.cell(row=117, column=1, value=(
        "=IF(B110=0,\"No MC units in selected period.\","
        "IF(LEFT(B103,9)=\"Flat-rate\","
        "\"Flat-rate MC detected. Tier analysis not applicable; \""
        "&\"see base rent + care charge totals only.\","
        "IF(LEFT(B103,7)=\"Tiered \","
        "\"Tiered MC detected. Verify per-tier staffing model supports the implied acuity mix.\","
        "\"Fee-for-service MC detected. Charges vary per resident — \""
        "&\"review individual care plans for sustainability.\")))"
    ))
    ws.cell(row=117, column=1).font = Font(name="Arial", size=10, italic=True)
    ws.merge_cells("A117:H117")


# ============================================================================
# D — Version stamps
# ============================================================================

def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ============================================================================
# Verification
# ============================================================================

def verify_migration(wb) -> dict:
    r = {}

    # 1. Cover B8 stamped
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    # 2. All 13 anchor AZ4 cells
    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all_v018"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # 3. T12 Analytics B2 — 3-priority formula present (refined to A3/A10)
    b2 = wb["T12 Analytics"]["B2"].value or ""
    r["t12a_b2_priority"] = (
        "'Rent Roll Input'!A3" in b2 and "'T12 Input'!A10" in b2 and "Property_Name" in b2
    )

    # 4. T12 Analytics E2 — rightmost-month formula present
    e2 = wb["T12 Analytics"]["E2"].value or ""
    r["t12a_e2_lookup"] = "LOOKUP" in e2 and "C$11:$N$11" in e2

    # 5. Charts on T12 Analytics
    r["t12a_chart_count"] = len(wb["T12 Analytics"]._charts)
    r["t12a_charts_ok"] = r["t12a_chart_count"] == 5

    # 6. Helper rate-bucket block present at K47:N52
    r["helper_buckets_ok"] = (
        wb["T12 Analytics"]["K47"].value == "Bucket"
        and wb["T12 Analytics"]["K48"].value == "$0 – $1,999"
        and wb["T12 Analytics"]["K52"].value == "$8,000+"
    )

    # 7. Helper monthly revenue row 54 present (SUMIFS over Revenue Section)
    k54 = str(wb["T12 Analytics"]["K54"].value or "")
    r["helper_v4_ok"] = (
        "SUMIFS" in k54 and "T12 Raw Data" in k54 and "\"Revenue\"" in k54
    )

    # 8. Chart notes present at K15, K30, P15, P30, K45
    notes_present = all(
        wb["T12 Analytics"][c].value is not None
        for c in ("K15", "K30", "P15", "P30", "K45")
    )
    r["chart_notes_ok"] = notes_present

    # 9. Rent Roll Recon B2 formula
    b2_rr = wb["Rent Roll Recon"]["B2"].value or ""
    r["rr_recon_b2_ok"] = "LOOKUP(9.99E+307" in b2_rr

    # 10. Rent Roll Recon B2 data validation present
    rr_recon = wb["Rent Roll Recon"]
    has_b2_dv = any(
        "B2" in str(dv.sqref) and dv.type == "list"
        for dv in rr_recon.data_validations.dataValidation
    )
    r["rr_recon_b2_dv_ok"] = has_b2_dv

    # 11. IL section K header at A86
    r["il_section_header_ok"] = (
        "IL UNIT-TYPE MIX" in str(rr_recon["A86"].value or "")
    )
    # IL total row 93 formula
    r["il_total_row_ok"] = rr_recon["B93"].value == "=SUM(B88:B92)"

    # 12. MC section L header at A102
    r["mc_section_header_ok"] = (
        "MC CARE STRUCTURE" in str(rr_recon["A102"].value or "")
    )
    # MC pattern detector at B103
    r["mc_pattern_detector_ok"] = (
        "Flat-rate" in str(rr_recon["B103"].value or "")
    )

    # 13. Input sheets — leftover v0.1.8-first-pass labels cleared from
    #     RR Input A3 and T12 Input A2 (now reserved as value cells / passed-through)
    r["rri_a3_clear"] = wb["Rent Roll Input"]["A3"].value != "Property name:"
    r["t12i_a2_clear"] = wb["T12 Input"]["A2"].value != "Property name:"

    # 14. Named ranges intact
    names = {n for n in wb.defined_names}
    expected = {
        "RR_Period_Date", "T12_Period_Date", "RR_Input_Data",
        "T12_Input_Data", "Property_Name", "DescMap_Description", "DescMap_Label",
    }
    r["named_ranges_ok"] = expected.issubset(names)

    return r


# ============================================================================
# Main
# ============================================================================

def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v018(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    # B3.1
    add_input_property_cells(wb)
    print("  B3.1-a: reserved property-name value cells (RR Input A3, T12 Input A10)")
    install_b2_property_formula(wb)
    print("  B3.1-b: installed T12 Analytics B2 3-priority property-name formula")
    install_e2_period_formula(wb)
    print("  B3.1-c: installed T12 Analytics E2 rightmost-month formula")

    # B3.2
    install_helper_rate_buckets(wb)
    print("  B3.2-a: installed hidden helper rate-bucket block (K46:V53)")
    n_charts = install_charts(wb)
    print(f"  B3.2-b: installed {n_charts} chart objects")
    install_chart_notes(wb)
    print("  B3.2-c: installed 5 conditional note cells")

    # B3.3
    install_rr_recon_b2_default(wb)
    print("  B3.3: installed Rent Roll Recon B2 latest-date default formula + DV")

    # B3.4
    install_il_section(wb)
    print("  B3.4: installed Rent Roll Recon section K (IL deep-dive)")

    # B3.5
    install_mc_section(wb)
    print("  B3.5: installed Rent Roll Recon section L (MC deep-dive)")

    # D
    stamp_versions(wb)
    print(f"  D: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    # Reload and verify
    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']:20s}  : {r['cover_b8_ok']}")
    print(f"  All 13 AZ4 = {SUBSTRATE_TO}                : {r['az4_all_v018']} ({r['az4_count']} sheets)")
    print(f"  T12 Analytics B2 3-priority         : {r['t12a_b2_priority']}")
    print(f"  T12 Analytics E2 rightmost-month    : {r['t12a_e2_lookup']}")
    print(f"  T12 Analytics chart count = 5       : {r['t12a_charts_ok']} (found {r['t12a_chart_count']})")
    print(f"  Helper rate-bucket block            : {r['helper_buckets_ok']}")
    print(f"  Helper V4 monthly revenue row       : {r['helper_v4_ok']}")
    print(f"  Chart notes (K15/K30/P15/P30/K45)   : {r['chart_notes_ok']}")
    print(f"  Rent Roll Recon B2 default formula  : {r['rr_recon_b2_ok']}")
    print(f"  Rent Roll Recon B2 data validation  : {r['rr_recon_b2_dv_ok']}")
    print(f"  IL section K header (A86)           : {r['il_section_header_ok']}")
    print(f"  IL total row (B93)                  : {r['il_total_row_ok']}")
    print(f"  MC section L header (A102)          : {r['mc_section_header_ok']}")
    print(f"  MC pattern detector (B103)          : {r['mc_pattern_detector_ok']}")
    print(f"  Rent Roll Input A3 reserved (cleared): {r['rri_a3_clear']}")
    print(f"  T12 Input A2 cleared (A10 is target) : {r['t12i_a2_clear']}")
    print(f"  Named ranges intact                 : {r['named_ranges_ok']}")

    all_ok = (
        r["cover_b8_ok"] and r["az4_all_v018"]
        and r["t12a_b2_priority"] and r["t12a_e2_lookup"]
        and r["t12a_charts_ok"] and r["helper_buckets_ok"] and r["helper_v4_ok"]
        and r["chart_notes_ok"]
        and r["rr_recon_b2_ok"] and r["rr_recon_b2_dv_ok"]
        and r["il_section_header_ok"] and r["il_total_row_ok"]
        and r["mc_section_header_ok"] and r["mc_pattern_detector_ok"]
        and r["rri_a3_clear"] and r["t12i_a2_clear"]
        and r["named_ranges_ok"]
    )
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v018.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
