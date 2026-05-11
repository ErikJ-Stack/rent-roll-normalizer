"""
migrate_to_v019.py — Substrate template v0.1.8 -> v0.1.9

Bug fix for the period-dropdown / latest-period-default not populating in
Excel (reported 2026-05-11 against substrate v0.1.8). Two root causes:

  1. RR_Calc!A2:A13 used `_xludf.minifs(...)` — Google Sheets / LibreOffice
     UDF prefix that Excel does NOT recognize. In Excel each cell resolves
     to #NAME?, the IFERROR returns "", the dropdown source is empty, and
     the LOOKUP-based default formula on Rent Roll Recon!B2 finds nothing.
     Fix: drop the `_xludf.` prefix and use Excel-native MINIFS.

  2. Rent Roll Recon!B2 default formula (installed at v0.1.8) was
     `=LOOKUP(9.99E+307, 'RR_Calc'!$A$2:$A$13)` — a transitive dependency
     on RR_Calc working. Belt-and-suspenders fix: B2 now reads directly
     from Rent Roll Input!$S$7:$S$606 via MAX, so it works even if RR_Calc
     ever drifts again.

This is a Cluster A-style correctness fix, analogous to the v0.1.6 H20
`_xlfn._LONGTEXT` cleanup. Architectural constraint of "additive only"
is preserved — these are formula-text repairs to an existing aggregator,
not a rewrite of its logic.

Operations:

  A. Drop `_xludf.` prefix from 12 cells in RR_Calc!A2:A13
  B. Replace Rent Roll Recon!B2 with direct-from-Input MAX formula
  C. Stamp Cover!B8 and all 13 AZ4 anchors to v0.1.9
  D. Verification block (6 checks)

Usage:
    python tools/migration/migrate_to_v019.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

SUBSTRATE_FROM = "v0.1.8"
SUBSTRATE_TO = "v0.1.9"

ANCHOR_SHEETS = (
    "Cover", "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)


def is_already_v019(wb) -> bool:
    """Gate verifies the workbook is fully at v0.1.9.

    Checks both the version stamp AND that no _xludf prefix remains in
    RR_Calc — defends against partial-migration states.
    """
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    ws = wb["RR_Calc"]
    for r in range(2, 14):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "_xludf." in v:
            return False
    return True


def fix_rr_calc_xludf(wb) -> int:
    """Drop `_xludf.` prefix from RR_Calc!A2:A13. Returns count fixed."""
    ws = wb["RR_Calc"]
    n = 0
    for r in range(2, 14):
        cell = ws.cell(row=r, column=1)
        v = cell.value
        if isinstance(v, str) and "_xludf." in v:
            cell.value = v.replace("_xludf.", "")
            n += 1
    return n


def install_b2_direct_max(wb) -> None:
    """Replace Rent Roll Recon!B2 with MAX-of-Input!S formula.

    Removes the transitive dependency on RR_Calc. Returns "" when the
    input range is empty (MAX of empties = 0; IF guards against it).
    Data validation on B2 (dropdown sourced from RR_Calc) is left in
    place — the v0.1.9 RR_Calc fix makes it work properly too.
    """
    ws = wb["Rent Roll Recon"]
    ws["B2"] = (
        "=IF(MAX('Rent Roll Input'!$S$7:$S$606)>0,"
        "MAX('Rent Roll Input'!$S$7:$S$606),\"\")"
    )
    ws["B2"].number_format = "mm/dd/yyyy"


def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


def verify_migration(wb) -> dict:
    r = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all_v019"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # No _xludf left anywhere in the workbook
    xludf_remaining = 0
    for s in wb.sheetnames:
        ws = wb[s]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "_xludf." in cell.value:
                    xludf_remaining += 1
    r["xludf_remaining"] = xludf_remaining
    r["xludf_swept_clean"] = xludf_remaining == 0

    # RR_Calc!A2 uses native MINIFS
    rrc_a2 = str(wb["RR_Calc"]["A2"].value or "")
    r["rr_calc_native_minifs"] = "_xludf." not in rrc_a2 and "MINIFS(" in rrc_a2.upper()

    # Rent Roll Recon!B2 reads directly from Input!S via MAX
    b2 = str(wb["Rent Roll Recon"]["B2"].value or "")
    r["b2_direct_max"] = "MAX('Rent Roll Input'!$S$7:$S$606)" in b2

    # B2 data validation still in place
    dvs = list(wb["Rent Roll Recon"].data_validations.dataValidation)
    r["b2_dv_intact"] = any(
        "B2" in str(dv.sqref) and dv.type == "list" for dv in dvs
    )

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v019(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    n_fixed = fix_rr_calc_xludf(wb)
    print(f"  A: dropped _xludf prefix from {n_fixed} cells in RR_Calc")

    install_b2_direct_max(wb)
    print("  B: replaced Rent Roll Recon!B2 with direct MAX on Rent Roll Input!S")

    stamp_versions(wb)
    print(f"  C: stamped substrate version -> {SUBSTRATE_TO}")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:20s}      : {r['cover_b8_ok']}")
    print(f"  All 13 AZ4 = {SUBSTRATE_TO}                : {r['az4_all_v019']} ({r['az4_count']} sheets)")
    print(f"  _xludf remaining in workbook        : {r['xludf_remaining']} (expect 0)")
    print(f"  RR_Calc!A2 uses native MINIFS       : {r['rr_calc_native_minifs']}")
    print(f"  Rent Roll Recon!B2 direct MAX       : {r['b2_direct_max']}")
    print(f"  Rent Roll Recon!B2 data validation  : {r['b2_dv_intact']}")

    all_ok = (
        r["cover_b8_ok"]
        and r["az4_all_v019"]
        and r["xludf_swept_clean"]
        and r["rr_calc_native_minifs"]
        and r["b2_direct_max"]
        and r["b2_dv_intact"]
    )
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v019.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
