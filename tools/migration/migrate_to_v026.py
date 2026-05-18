"""
migrate_to_v026.py — Substrate template v0.2.5 → v0.2.6

Closes UW-BACKLOG **BL-0016** + **BL-0017**. Both were originally deferred to
manual Excel handling per the BL entries' "do not re-attempt as substrate
migration without re-confirming" notes — user re-confirmed on 2026-05-18 to
ship via substrate migration after all. This file ports the abandoned
implementation from commit `fac129d` on branch `claude/serene-panini-3ad41d`
(originally targeted v0.2.4 — re-numbered here to v0.2.6 to chain after
v0.2.5 / BL-0012).

TWO STEPS:

  A. **BL-0016 — Rent Roll Input!AH4 fill.** The AH "Total Ancillary $"
     header was added in v0.2.2 with white-bold font but transparent fill —
     so it rendered white-on-default (invisible). v0.2.6 applies the green
     `FF1F6B52` PatternFill matching `T4` / `U4` (computed-column header
     palette: navy `FF1F3864` is for input columns, green for formula
     columns).

  B. **BL-0017 — Workbook-wide "intentionally blank" visual convention.**
     144 cells across 3 sheets currently store the literal 3-character
     string `"-"` (a double-quote, dash, double-quote payload — renders
     in Excel with visible quote marks). All 144 share the same
     "intentionally blank, not just missing data" design intent. v0.2.6
     applies the user-approved gray-with-em-dash treatment:
       - `value="—"` (em-dash plain text, no quote marks)
       - `fill=PatternFill(start_color="FFF2F2F2", solid)` (light gray)
       - `font.color="FFA0A0A0"` (medium gray, preserving size/bold/italic)
       - `alignment.horizontal="center"` (preserving vertical/wrap/indent)

     Target cell inventory (144 total):
       - T12 Analytics: E36, G36 (2)
       - Rent Roll Recon: D109 (1)
       - UW Output: cols B/C/D × rows {8-12, 22-28, 30-36, 38-56, 58-60,
         62-64, 66-68} (141)

     New user-facing rule established: **gray + em-dash = "blank by
     design"; truly empty = "data not yet populated"**.

     Out of scope (deliberately): formula-conditional blanks like
     `T12 Analytics!E37/G37/H38` that return `""` only when source data
     is missing. Those are "blank when data isn't here" — permanent
     styling would mislead when they populate. Defer to a future v0.2.7+
     if a clean Excel-conditional-formatting approach surfaces.

Idempotency: gate checks BOTH `Cover!B8 == "v0.2.6"` AND `Rent Roll
Input!AH4` fill is the green `FF1F6B52` AND `UW Output!B8` already has
the styled-blank treatment. Re-run on already-migrated workbook is a
no-op; partial pre-state (e.g. some cells styled, some not) is
gracefully re-applied via the per-cell skip-if-styled check.

Usage:
    python tools/migration/migrate_to_v026.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


SUBSTRATE_FROM = "v0.2.5"
SUBSTRATE_TO = "v0.2.6"

# 15-sheet anchor list (post-v0.2.4 with Investment Dashboard)
ANCHOR_SHEETS = (
    "Cover", "Investment Dashboard",
    "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# Step A — AH4 green fill (matches T4/U4 computed-column header palette).
GREEN_FILL_RGB = "FF1F6B52"
GREEN_FILL = PatternFill(
    start_color=GREEN_FILL_RGB, end_color=GREEN_FILL_RGB, fill_type="solid"
)

# Step B — intentional-blank treatment.
EMDASH = "—"  # em-dash
BLANK_FILL_RGB = "FFF2F2F2"
BLANK_FONT_RGB = "FFA0A0A0"
BLANK_FILL = PatternFill(
    start_color=BLANK_FILL_RGB, end_color=BLANK_FILL_RGB, fill_type="solid"
)
DASH_ARTIFACT = '"-"'  # literal-text artifact being replaced

# UW Output row groups — cols B/C/D × these row ranges hold the placeholder.
UW_OUTPUT_ROW_GROUPS = [
    (8, 12),    # Other-revenue lines + EGI
    (22, 28),   # OpEx section 1
    (30, 36),   # OpEx section 2
    (38, 56),   # OpEx detail
    (58, 60),   # Total OpEx + NOI lines
    (62, 64),   # Capex
    (66, 68),   # Below-the-line
]


def build_blank_targets() -> list[tuple[str, str]]:
    """Enumerate all 144 intentional-blank placeholder coords."""
    targets: list[tuple[str, str]] = []
    targets.append(("T12 Analytics", "E36"))
    targets.append(("T12 Analytics", "G36"))
    targets.append(("Rent Roll Recon", "D109"))
    for start_row, end_row in UW_OUTPUT_ROW_GROUPS:
        for r in range(start_row, end_row + 1):
            for col in ("B", "C", "D"):
                targets.append(("UW Output", f"{col}{r}"))
    return targets


BLANK_TARGETS = build_blank_targets()
assert len(BLANK_TARGETS) == 144, f"Expected 144 targets, built {len(BLANK_TARGETS)}"


def _is_already_styled(cell) -> bool:
    """Cell has the intentional-blank treatment fully applied."""
    if cell.value != EMDASH:
        return False
    fg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    return fg == BLANK_FILL_RGB


def is_already_v026(wb) -> bool:
    """Gate: version stamp + AH4 green fill + UW Output sentinel styled."""
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    ah4 = wb["Rent Roll Input"]["AH4"]
    fg = ah4.fill.fgColor.rgb if ah4.fill and ah4.fill.fgColor else None
    if fg != GREEN_FILL_RGB:
        return False
    sentinel = wb["UW Output"]["B8"]
    return _is_already_styled(sentinel)


def fix_ah4_fill(wb) -> int:
    """A: apply green PatternFill to Rent Roll Input!AH4."""
    ws = wb["Rent Roll Input"]
    ws["AH4"].fill = GREEN_FILL
    return 1


def apply_blank_styling(wb) -> tuple[int, int]:
    """B: apply intentional-blank treatment to the 144 placeholder cells.

    Idempotent — skips cells that already have the styling. Accepts any
    of three pre-states per cell:
      1. literal `"-"` text       (the v0.2.3 substrate's stored value)
      2. None / blank             (e.g., E36/G36 from an aborted earlier
                                   attempt that cleared but didn't restyle)
      3. already-styled em-dash   (this run is a no-op for that cell)

    Returns (styled_count, skipped_count).
    """
    styled = 0
    skipped = 0
    for sheet_name, coord in BLANK_TARGETS:
        cell = wb[sheet_name][coord]

        if _is_already_styled(cell):
            skipped += 1
            continue

        # value: em-dash, plain text (not formula, not quoted).
        cell.value = EMDASH

        # fill: solid light gray.
        cell.fill = BLANK_FILL

        # font: preserve size/bold/italic/name, override color to medium gray.
        old = cell.font
        cell.font = Font(
            name=old.name,
            size=old.size,
            bold=old.bold,
            italic=old.italic,
            underline=old.underline,
            strike=old.strike,
            color=BLANK_FONT_RGB,
            family=old.family,
            scheme=old.scheme,
        )

        # alignment: center horizontal; preserve vertical/wrap/indent.
        oa = cell.alignment
        cell.alignment = Alignment(
            horizontal="center",
            vertical=oa.vertical,
            indent=oa.indent,
            wrap_text=oa.wrap_text,
            shrink_to_fit=oa.shrink_to_fit,
            text_rotation=oa.text_rotation,
        )

        styled += 1
    return styled, skipped


def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


def verify_migration(wb) -> dict:
    r: dict = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # AH4 fill: must be the green FF1F6B52 matching T4/U4
    ws_rr = wb["Rent Roll Input"]
    ah4 = ws_rr["AH4"]
    fg = ah4.fill.fgColor.rgb if ah4.fill and ah4.fill.fgColor else None
    r["ah4_fill"] = fg
    r["ah4_fill_green"] = fg == GREEN_FILL_RGB

    # AH4 text + font preserved
    r["ah4_text_ok"] = ah4.value == "Total\nAncillary $"
    r["ah4_font_ok"] = bool(ah4.font.bold)

    # All 144 blank targets styled correctly
    fully_styled = 0
    sample_failures: list[str] = []
    for sheet_name, coord in BLANK_TARGETS:
        cell = wb[sheet_name][coord]
        if _is_already_styled(cell):
            fully_styled += 1
        else:
            sample_failures.append(f"{sheet_name}!{coord}")
    r["styled_count"] = fully_styled
    r["target_count"] = len(BLANK_TARGETS)
    r["blanks_all_styled"] = fully_styled == len(BLANK_TARGETS)
    r["blank_failure_samples"] = sample_failures[:5]

    # Sample 1: T12 Analytics E36 details
    e36 = wb["T12 Analytics"]["E36"]
    r["e36_value"] = e36.value
    r["e36_fill"] = e36.fill.fgColor.rgb if e36.fill and e36.fill.fgColor else None
    r["e36_font_color"] = e36.font.color.rgb if e36.font.color else None
    r["e36_align"] = e36.alignment.horizontal

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v026(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...\n")

    print("Step A — BL-0016: fix Rent Roll Input!AH4 header fill (green FF1F6B52):")
    n_a = fix_ah4_fill(wb)
    print(f"  set fill on {n_a} cell")

    print("\nStep B — BL-0017: apply 'intentionally blank' styling to 144 placeholders:")
    n_styled, n_skipped = apply_blank_styling(wb)
    print(f"  styled {n_styled} cells, skipped {n_skipped} already-styled "
          f"(total target: {len(BLANK_TARGETS)})")

    stamp_versions(wb)
    print(f"\nStep C — stamped substrate version -> {SUBSTRATE_TO} ({len(ANCHOR_SHEETS)} anchors)")

    print(f"\nSaving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  1. Cover!B8 = {r['cover_b8']!r}                          : {r['cover_b8_ok']}")
    print(f"  2. All 15 AZ4 = {SUBSTRATE_TO}                            : {r['az4_all']} ({r['az4_count']} sheets)")
    print(f"  3. AH4 fill = {r['ah4_fill']!r} (target {GREEN_FILL_RGB!r})   : {r['ah4_fill_green']}")
    print(f"  4. AH4 text 'Total\\nAncillary $' preserved                  : {r['ah4_text_ok']}")
    print(f"  5. AH4 font bold preserved                                  : {r['ah4_font_ok']}")
    print(f"  6. All 144 blank-targets styled ({r['styled_count']}/{r['target_count']})           : {r['blanks_all_styled']}")
    if r["blank_failure_samples"]:
        print(f"       first failures: {r['blank_failure_samples']}")
    print(f"  7. Sample (T12 Analytics E36):")
    print(f"       value='{r['e36_value']}' (target '—')")
    print(f"       fill={r['e36_fill']!r} (target {BLANK_FILL_RGB!r})")
    print(f"       font color={r['e36_font_color']!r} (target {BLANK_FONT_RGB!r})")
    print(f"       align horizontal={r['e36_align']!r} (target 'center')")

    all_ok = all([
        r["cover_b8_ok"], r["az4_all"],
        r["ah4_fill_green"], r["ah4_text_ok"], r["ah4_font_ok"],
        r["blanks_all_styled"],
        r["e36_value"] == EMDASH,
        r["e36_fill"] == BLANK_FILL_RGB,
        r["e36_font_color"] == BLANK_FONT_RGB,
        r["e36_align"] == "center",
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v026.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
