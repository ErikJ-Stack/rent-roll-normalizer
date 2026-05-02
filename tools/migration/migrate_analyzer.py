"""
migrate_analyzer.py — apply template-substrate v0.1.4 edits to a master Analyzer.

Usage:
    python3 migrate_analyzer.py <master_input.xlsx> <v014_reference.xlsx> <output.xlsx>

Reads the master Analyzer at the pre-v0.1.0 state, applies the 5 edit batches
described in HANDOFF-Analyzer-Migration.md, writes a migrated copy.

Fail-loud philosophy: any deviation from expected pre-state is logged as a
WARNING. Caller should investigate WARNINGs before treating the output as
ship-ready. End-to-end Salem/Briar Glen verification is a separate step.
"""

import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

T12_RAW_DATA_MONTH_COLS_MT = "FGHIJKLMNOPQ"  # T12 Raw Data F..Q ↔ Monthly Trending B..M

DESC_NAMED_RANGE = (
    "Description_Map!$A$5:INDEX(Description_Map!$A:$A, "
    "MAX(5, COUNTA(Description_Map!$A:$A)+4))"
)
LABEL_NAMED_RANGE = (
    "Description_Map!$B$5:INDEX(Description_Map!$B:$B, "
    "MAX(5, COUNTA(Description_Map!$A:$A)+4))"
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def warn(msg):
    print(f"  WARNING: {msg}")


def info(msg):
    print(f"  {msg}")


def section(title):
    print(f"\n=== {title} ===")


def read_row(ws, r, n_cols):
    """Capture full cell state (value + style) for one row, for safe shifting."""
    out = []
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=r, column=c)
        out.append({
            "value": cell.value,
            "has_style": cell.has_style,
            "font": copy(cell.font) if cell.has_style else None,
            "fill": copy(cell.fill) if cell.has_style else None,
            "border": copy(cell.border) if cell.has_style else None,
            "alignment": copy(cell.alignment) if cell.has_style else None,
            "number_format": cell.number_format,
        })
    return out


def write_row(ws, r, cells):
    for c, cd in enumerate(cells, start=1):
        target = ws.cell(row=r, column=c)
        target.value = cd["value"]
        if cd["has_style"]:
            target.font = cd["font"]
            target.fill = cd["fill"]
            target.border = cd["border"]
            target.alignment = cd["alignment"]
            target.number_format = cd["number_format"]


def mirror_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ---------------------------------------------------------------------------
# Batch 1 — Description_Map cleanup + 82 new entries
# ---------------------------------------------------------------------------

def batch1_description_map(wb, wb_ref):
    section("Batch 1 — Description_Map")
    ws = wb["Description_Map"]
    ws_ref = wb_ref["Description_Map"]

    # 1a. Find and delete duplicate "Auto Expenses" row (the one mapping to
    #     "Office, admin & G&A" — keep the one mapping to "Auto Expense").
    dup_rows = []
    for r in range(5, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Auto Expenses":
            dup_rows.append(r)
    info(f"Found {len(dup_rows)} 'Auto Expenses' row(s): {dup_rows}")

    if len(dup_rows) == 2:
        # Identify which one maps to "Auto Expense" (keep) vs other (delete)
        labels = [ws.cell(row=r, column=2).value for r in dup_rows]
        if labels[0] == "Auto Expense" and labels[1] != "Auto Expense":
            to_delete = dup_rows[1]
        elif labels[1] == "Auto Expense" and labels[0] != "Auto Expense":
            to_delete = dup_rows[0]
        else:
            warn(f"Both 'Auto Expenses' rows have same label {labels} — "
                 f"deleting the lower one defensively")
            to_delete = max(dup_rows)
        ws.delete_rows(to_delete, 1)
        info(f"Deleted row {to_delete}")
    elif len(dup_rows) == 1:
        warn("Only one 'Auto Expenses' row found — duplicate may already be removed")
    else:
        warn(f"Unexpected 'Auto Expenses' count: {len(dup_rows)}")

    # 1b. Build set of existing descriptions (post-delete) to enable auto-skip
    existing = set()
    for r in range(5, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None:
            existing.add(str(v).strip())
    info(f"Existing Description_Map entries after delete: {len(existing)}")

    # Find first blank row in master (where to start appending)
    append_start = 5
    for r in range(5, ws.max_row + 2):
        if ws.cell(row=r, column=1).value is None:
            append_start = r
            break
    info(f"Will append starting at row {append_start}")

    # 1c. Pull v0.1.4 reference rows R235:R316 (the 82 new MRI entries)
    new_entries = []
    for r in range(235, 317):
        desc = ws_ref.cell(row=r, column=1).value
        if desc is None:
            continue
        new_entries.append({
            "description": desc,
            "label":      ws_ref.cell(row=r, column=2).value,
            "section":    ws_ref.cell(row=r, column=3).value,
            "caretype":   ws_ref.cell(row=r, column=4).value,
            "flag":       ws_ref.cell(row=r, column=5).value,
        })
    info(f"v0.1.4 reference has {len(new_entries)} new MRI entries (R235:R316)")

    # 1d. Append, skipping any already in master
    ref_style_cell = lambda c: ws.cell(row=5, column=c)  # mirror R5 formatting
    appended = 0
    skipped = 0
    write_row_cursor = append_start
    for entry in new_entries:
        if str(entry["description"]).strip() in existing:
            skipped += 1
            continue
        for c, key in enumerate(["description", "label", "section", "caretype", "flag"], start=1):
            cell = ws.cell(row=write_row_cursor, column=c)
            cell.value = entry[key]
            mirror_style(ref_style_cell(c), cell)
        existing.add(str(entry["description"]).strip())
        write_row_cursor += 1
        appended += 1

    info(f"Appended {appended}, auto-skipped {skipped} (already present)")

    # Verification
    final_count = sum(1 for r in range(5, ws.max_row + 1)
                      if ws.cell(row=r, column=1).value is not None)
    info(f"Final Description_Map data row count: {final_count}")


# ---------------------------------------------------------------------------
# Batch 2 — Named ranges
# ---------------------------------------------------------------------------

def batch2_named_ranges(wb):
    section("Batch 2 — Named ranges")

    # Defensive: remove if exists, then add fresh
    for name in ("DescMap_Description", "DescMap_Label"):
        if name in wb.defined_names:
            del wb.defined_names[name]
            info(f"Removed pre-existing '{name}'")

    wb.defined_names["DescMap_Description"] = DefinedName(
        "DescMap_Description", attr_text=DESC_NAMED_RANGE
    )
    wb.defined_names["DescMap_Label"] = DefinedName(
        "DescMap_Label", attr_text=LABEL_NAMED_RANGE
    )
    info("Defined: DescMap_Description, DescMap_Label")


# ---------------------------------------------------------------------------
# Batch 3 — T12 Input formula + structure changes
# ---------------------------------------------------------------------------

def batch3_t12_input(wb):
    section("Batch 3 — T12 Input")
    ws = wb["T12 Input"]

    # 3a. Unmerge A11:O11
    merged_to_drop = []
    for mr in list(ws.merged_cells.ranges):
        if str(mr) == "A11:O11":
            merged_to_drop.append(str(mr))
    for mr in merged_to_drop:
        ws.unmerge_cells(mr)
        info(f"Unmerged {mr}")
    if not merged_to_drop:
        warn("A11:O11 was not merged in master — already unmerged?")

    # 3b. Rewrite 500 col P formulas
    rewritten = 0
    skipped = 0
    for r in range(12, 512):
        cell = ws.cell(row=r, column=16)
        expected_old = (
            f'=IF(ISNUMBER(VALUE(TRIM(A{r}))),'
            f'IFERROR(INDEX(Description_Map!$B$5:$B$284,'
            f'MATCH(TRIM(B{r}),Description_Map!$A$5:$A$284,0)),"UNMATCHED"),"")'
        )
        if cell.value != expected_old:
            warn(f"P{r} not in expected pre-state: {cell.value!r}")
            skipped += 1
            continue
        cell.value = (
            f'=IF(TRIM(B{r})<>"",'
            f'IFERROR(INDEX(DescMap_Label,MATCH(TRIM(B{r}),DescMap_Description,0)),'
            f'"UNMATCHED"),"")'
        )
        rewritten += 1
    info(f"Rewrote {rewritten}/500 col P formulas (skipped {skipped})")

    # 3c. Row 11 column headers — A, B, O (P11 already 'Coverage Check' in pre-state? verify)
    ref = ws.cell(row=11, column=16)  # P11
    p11_text = ref.value
    if p11_text != "Coverage Check":
        warn(f"P11 was {p11_text!r} (expected 'Coverage Check') — overwriting")
        ref.value = "Coverage Check"

    def set_header(col, text):
        cell = ws.cell(row=11, column=col)
        cell.value = text
        cell.font = Font(
            name=ref.font.name or "Calibri",
            size=ref.font.size or 11,
            bold=True,
            color=ref.font.color,
        )
        if ref.fill.fill_type:
            cell.fill = PatternFill(
                fill_type=ref.fill.fill_type,
                fgColor=ref.fill.fgColor,
                bgColor=ref.fill.bgColor,
            )
        cell.alignment = Alignment(
            horizontal=ref.alignment.horizontal or "center",
            vertical=ref.alignment.vertical or "center",
            wrap_text=ref.alignment.wrap_text or False,
        )

    set_header(1, "Account #")
    set_header(2, "Description")
    set_header(15, "T12 Total")
    info("Set row 11 headers: A=Account #, B=Description, O=T12 Total, P=Coverage Check")

    # 3d. Rewrite rows 4-7 instructions (still merged A4:O4 etc — value lives in col A)
    ws.cell(row=4, column=1).value = "1.  Upload this Analyzer + your raw T12 to the Streamlit app."
    ws.cell(row=5, column=1).value = "2.  Map any UNMATCHED descriptions in-app before downloading."
    ws.cell(row=6, column=1).value = "3.  Open the populated Analyzer — Coverage Check (col P) confirms all rows mapped."
    ws.cell(row=7, column=1).value = "4.  Mappings persist in Description_Map for next upload."
    info("Rewrote rows 4-7 with app-driven workflow instructions")

    # 3e. Row 9 layout note
    ws.cell(row=9, column=1).value = (
        "Layout: Account # (col A, optional, populated for Yardi-style T12s only), "
        "Description (col B, required), 12 months (cols C-N), T12 Total (col O). "
        "Headers in row 11 reflect the source T12 period."
    )
    info("Updated row 9 layout note")


# ---------------------------------------------------------------------------
# Batch 4 — T12_Calc helper col + T12 Raw Data SUMIF rewrite
# ---------------------------------------------------------------------------

def batch4_t12_calc_and_raw(wb):
    section("Batch 4 — T12_Calc + T12 Raw Data")

    # 4a. Helper col N on T12_Calc
    ws_calc = wb["T12_Calc"]
    for r in range(1, 501):
        ws_calc.cell(row=r, column=14).value = (
            f"=IFERROR(INDEX(DescMap_Label,MATCH(A{r},DescMap_Description,0)),\"\")"
        )
    info("Wrote helper formulas to T12_Calc!N1:N500")

    # 4b. Rewrite SUMIF formulas in T12 Raw Data
    ws_raw = wb["T12 Raw Data"]
    RAW_MONTH_COLS = list(range(6, 18))   # F..Q on T12 Raw Data
    CALC_MONTH_COLS = list(range(2, 14))  # B..M on T12_Calc

    rewrite_count = 0
    for r in range(5, ws_raw.max_row + 1):
        label = ws_raw.cell(row=r, column=2).value
        if not label or not str(label).strip():
            continue
        label_str = str(label).strip()
        for raw_col, calc_col in zip(RAW_MONTH_COLS, CALC_MONTH_COLS):
            cell = ws_raw.cell(row=r, column=raw_col)
            existing = cell.value
            if existing is None:
                continue
            if not (isinstance(existing, str)
                    and existing.startswith("=")
                    and "SUMIF" in existing):
                continue
            calc_letter = chr(64 + calc_col)
            cell.value = (
                f'=SUMIF(T12_Calc!$N$1:$N$500,"{label_str}",'
                f'T12_Calc!${calc_letter}$1:${calc_letter}$500)'
            )
            rewrite_count += 1

    info(f"Rewrote {rewrite_count} SUMIF formulas (expected ~612)")
    if rewrite_count != 612:
        warn(f"SUMIF rewrite count {rewrite_count} != 612 — check master pre-state")

    # 4c. Add new label rows R57 (Auto Expense), R58 (Lease / ground lease)
    def add_label_row(target_row, sec, label, care_type, flag, matched_descs):
        ws_raw.cell(row=target_row, column=1).value = sec
        ws_raw.cell(row=target_row, column=2).value = label
        ws_raw.cell(row=target_row, column=3).value = care_type
        ws_raw.cell(row=target_row, column=4).value = flag
        ws_raw.cell(row=target_row, column=5).value = matched_descs

        for raw_col in range(6, 18):
            calc_col = raw_col - 4
            calc_letter = chr(64 + calc_col)
            ws_raw.cell(row=target_row, column=raw_col).value = (
                f'=SUMIF(T12_Calc!$N$1:$N$500,"{label}",'
                f'T12_Calc!${calc_letter}$1:${calc_letter}$500)'
            )
        ws_raw.cell(row=target_row, column=18).value = f"=SUM(F{target_row}:Q{target_row})"

        # Mirror styling from R42 (Auto insurance — clean Non-Labor row)
        for c in range(1, 19):
            mirror_style(ws_raw.cell(row=42, column=c), ws_raw.cell(row=target_row, column=c))

    add_label_row(
        target_row=57,
        sec="Non-Labor",
        label="Auto Expense",
        care_type="-",
        flag=None,
        matched_descs="Auto Expenses | Auto and Mileage Expense | Bus/Shuttle Service",
    )
    info("Added T12 Raw Data R57 = Auto Expense")

    add_label_row(
        target_row=58,
        sec="Non-Labor",
        label="Lease / ground lease",
        care_type="-",
        flag=None,
        matched_descs=("Lease Expense | Ground Lease | Land Lease | "
                       "Facility Lease | Building Lease | Rent Expense"),
    )
    info("Added T12 Raw Data R58 = Lease / ground lease")


# ---------------------------------------------------------------------------
# Batch 5 — Monthly Trending fixes
# ---------------------------------------------------------------------------

def batch5_monthly_trending(wb):
    section("Batch 5 — Monthly Trending")
    ws_mt = wb["Monthly Trending"]

    # 5a. R10/R11 — drop ABS, return 0 when source missing
    for trending_letter, raw_letter in zip("BCDEFGHIJKLM", T12_RAW_DATA_MONTH_COLS_MT):
        col = ord(trending_letter) - ord("A") + 1
        ws_mt.cell(row=10, column=col).value = (
            f"=IFERROR(INDEX('T12 Raw Data'!{raw_letter}:{raw_letter},"
            f'MATCH("Physical Vacancy",\'T12 Raw Data\'!B:B,0)),0)'
        )
        ws_mt.cell(row=11, column=col).value = (
            f"=IFERROR(INDEX('T12 Raw Data'!{raw_letter}:{raw_letter},"
            f'MATCH("Loss to Lease",\'T12 Raw Data\'!B:B,0)),0)'
        )
    ws_mt.cell(row=10, column=14).value = "=SUM(B10:M10)"
    ws_mt.cell(row=11, column=14).value = "=SUM(B11:M11)"
    info("Fixed R10 (Physical Vacancy) and R11 (Loss to Lease)")

    # 5b. R20 EGI = include Vacancy (10) + L2L (11)
    for trending_letter in "BCDEFGHIJKLM":
        col = ord(trending_letter) - ord("A") + 1
        L = trending_letter
        ws_mt.cell(row=20, column=col).value = (
            f"={L}8+{L}10+{L}11+{L}15+{L}16+{L}17+{L}18+{L}19"
        )
    ws_mt.cell(row=20, column=14).value = "=N8+N10+N11+N15+N16+N17+N18+N19"
    info("Extended R20 EGI formula to include Vacancy + L2L")

    # 5c. Manual row insert at R53 — Auto Expense
    # Capture R52..R68 (17 rows: Auto insurance through EBITDAR)
    captured = {r: read_row(ws_mt, r, n_cols=14) for r in range(52, 69)}

    # Shift R53..R68 down to R54..R69 (R52 stays put)
    for old_r in range(53, 69):
        write_row(ws_mt, old_r + 1, captured[old_r])

    # Build new R53 = Auto Expense, mirror R52 (Auto insurance) formatting
    ref_row_data = captured[52]
    auto_expense_cells = []
    for c in range(1, 15):
        cd = {
            "value": None,
            "has_style": ref_row_data[c-1]["has_style"],
            "font": ref_row_data[c-1]["font"],
            "fill": ref_row_data[c-1]["fill"],
            "border": ref_row_data[c-1]["border"],
            "alignment": ref_row_data[c-1]["alignment"],
            "number_format": ref_row_data[c-1]["number_format"],
        }
        auto_expense_cells.append(cd)

    auto_expense_cells[0]["value"] = "Auto Expense"  # col A
    for trending_letter, raw_letter in zip("BCDEFGHIJKLM", T12_RAW_DATA_MONTH_COLS_MT):
        idx = ord(trending_letter) - ord("A")
        auto_expense_cells[idx]["value"] = (
            f"=IFERROR(INDEX('T12 Raw Data'!{raw_letter}:{raw_letter},"
            f'MATCH("Auto Expense",\'T12 Raw Data\'!B:B,0)),0)'
        )
    auto_expense_cells[13]["value"] = "=SUM(B53:M53)"  # col N

    write_row(ws_mt, 53, auto_expense_cells)
    info("Inserted R53 = Auto Expense; shifted old R53..R68 → R54..R69")

    # 5d. R64 Lease — replace =0 with proper INDEX/MATCH
    for trending_letter, raw_letter in zip("BCDEFGHIJKLM", T12_RAW_DATA_MONTH_COLS_MT):
        col = ord(trending_letter) - ord("A") + 1
        ws_mt.cell(row=64, column=col).value = (
            f"=IFERROR(INDEX('T12 Raw Data'!{raw_letter}:{raw_letter},"
            f'MATCH("Lease / ground lease",\'T12 Raw Data\'!B:B,0)),0)'
        )
    ws_mt.cell(row=64, column=14).value = "=SUM(B64:M64)"
    info("Replaced R64 Lease =0 placeholder with INDEX/MATCH")

    # 5e. R65 Total non-labor opex — rebuild full sum range B40:B64
    for trending_letter in "BCDEFGHIJKLM":
        col = ord(trending_letter) - ord("A") + 1
        parts = [f"{trending_letter}{r}" for r in range(40, 65)]
        ws_mt.cell(row=65, column=col).value = "=" + "+".join(parts)
    n_parts = [f"N{r}" for r in range(40, 65)]
    ws_mt.cell(row=65, column=14).value = "=" + "+".join(n_parts)
    info("Rebuilt R65 Total non-labor opex sum range to B40:B64 (25 rows)")

    # 5f. R66 / R68 / R69 — references shifted due to insert
    for trending_letter in "BCDEFGHIJKLM":
        col = ord(trending_letter) - ord("A") + 1
        L = trending_letter
        ws_mt.cell(row=66, column=col).value = f"={L}38+{L}65"
        ws_mt.cell(row=68, column=col).value = f"={L}20-{L}66"
        ws_mt.cell(row=69, column=col).value = f"={L}68-{L}67"
    ws_mt.cell(row=66, column=14).value = "=N38+N65"
    ws_mt.cell(row=68, column=14).value = "=N20-N66"
    ws_mt.cell(row=69, column=14).value = "=N68-N67"
    info("Fixed R66 (TOTAL OPEX), R68 (EBITDARM), R69 (EBITDAR) references")

    # 5g. N-column self-references on rows 54-63 and 67 — broken by row shift
    for r in [54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 67]:
        ws_mt.cell(row=r, column=14).value = f"=SUM(B{r}:M{r})"
    info("Fixed N-column self-references on rows 54-63 and 67")


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

def verify(wb):
    section("Post-migration verification")
    ok = True

    # 1. Named ranges present
    for nm in ("DescMap_Description", "DescMap_Label"):
        if nm not in wb.defined_names:
            warn(f"Missing named range: {nm}")
            ok = False
    info("Named ranges present: ✓")

    # 2. Spot-check T12 Input col P
    ws = wb["T12 Input"]
    for r in (12, 100, 511):
        f = ws.cell(row=r, column=16).value
        if not (f and "DescMap_Label" in f and "DescMap_Description" in f and 'TRIM(B{0})<>""'.format(r) in f):
            warn(f"T12 Input!P{r} formula not as expected: {f!r}")
            ok = False
    info("T12 Input col P formulas: ✓")

    # 3. Row 11 headers
    headers = {1: "Account #", 2: "Description", 15: "T12 Total", 16: "Coverage Check"}
    for c, expected in headers.items():
        actual = ws.cell(row=11, column=c).value
        if actual != expected:
            warn(f"T12 Input!{chr(64+c)}11 = {actual!r}, expected {expected!r}")
            ok = False
    info("T12 Input row 11 headers: ✓")

    # 4. T12_Calc col N spot-check
    ws_calc = wb["T12_Calc"]
    for r in (1, 100, 500):
        f = ws_calc.cell(row=r, column=14).value
        if not f or "DescMap_Label" not in f:
            warn(f"T12_Calc!N{r} formula not as expected: {f!r}")
            ok = False
    info("T12_Calc helper col: ✓")

    # 5. T12 Raw Data new rows
    ws_raw = wb["T12 Raw Data"]
    if ws_raw.cell(row=57, column=2).value != "Auto Expense":
        warn(f"T12 Raw Data!B57 = {ws_raw.cell(row=57, column=2).value!r}, expected 'Auto Expense'")
        ok = False
    if ws_raw.cell(row=58, column=2).value != "Lease / ground lease":
        warn(f"T12 Raw Data!B58 = {ws_raw.cell(row=58, column=2).value!r}, expected 'Lease / ground lease'")
        ok = False
    info("T12 Raw Data new label rows: ✓")

    # 6. Monthly Trending key formulas
    ws_mt = wb["Monthly Trending"]
    if ws_mt.cell(row=53, column=1).value != "Auto Expense":
        warn(f"Monthly Trending!A53 = {ws_mt.cell(row=53, column=1).value!r}, expected 'Auto Expense'")
        ok = False
    if ws_mt.cell(row=20, column=2).value != "=B8+B10+B11+B15+B16+B17+B18+B19":
        warn(f"Monthly Trending!B20 EGI not as expected: {ws_mt.cell(row=20, column=2).value!r}")
        ok = False
    if ws_mt.cell(row=68, column=2).value != "=B20-B66":
        warn(f"Monthly Trending!B68 EBITDARM not as expected: {ws_mt.cell(row=68, column=2).value!r}")
        ok = False
    info("Monthly Trending structural formulas: ✓")

    # 7. RR-side sheets untouched (presence check)
    rr_sheets = ["Rent Roll Input", "Rent Roll Recon", "T12 Analytics", "UW Output"]
    for s in rr_sheets:
        if s not in wb.sheetnames:
            warn(f"Expected RR-side sheet missing: {s}")
            ok = False
    info(f"RR-side sheets present (untouched): {rr_sheets}")

    print(f"\n{'='*40}")
    print(f"Verification: {'PASSED' if ok else 'FAILED — check WARNINGS above'}")
    print(f"{'='*40}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)

    in_path = Path(sys.argv[1])
    ref_path = Path(sys.argv[2])
    out_path = Path(sys.argv[3])

    print(f"Master input:    {in_path}")
    print(f"v0.1.4 reference: {ref_path}")
    print(f"Output:          {out_path}")

    wb = openpyxl.load_workbook(in_path)
    wb_ref = openpyxl.load_workbook(ref_path, read_only=False)

    # Sanity: master should have all 11 sheets
    expected_sheets = {
        "Rent Roll Input", "Rent Roll Recon", "T12 Input", "T12 Raw Data",
        "T12 Analytics", "Monthly Trending", "Mapping Review", "Description_Map",
        "RR_Calc", "T12_Calc", "UW Output",
    }
    actual_sheets = set(wb.sheetnames)
    missing = expected_sheets - actual_sheets
    extra = actual_sheets - expected_sheets
    if missing:
        warn(f"Master is missing expected sheets: {missing}")
    if extra:
        info(f"Master has extra sheets (will leave alone): {extra}")

    # Apply the 5 batches in order
    batch1_description_map(wb, wb_ref)
    batch2_named_ranges(wb)
    batch3_t12_input(wb)
    batch4_t12_calc_and_raw(wb)
    batch5_monthly_trending(wb)

    # Verify
    verify(wb)

    # Save
    wb.save(out_path)
    print(f"\nSaved migrated Analyzer → {out_path}")


if __name__ == "__main__":
    main()
