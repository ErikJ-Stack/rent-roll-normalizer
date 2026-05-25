"""
migrate_to_v0211.py — Substrate template v0.2.10 → v0.2.11

Closes the Track-3-shaped portion of BL-0023 (AR & Collections module).
Adds two presentation-layer touches that surface AR signal in the
workbook's top-level views:

  1. Dashboard variance tile at K10:L13 — surfaces the bad-debt
     variance flag from `'AR & Collections'!C56`. Shows "— upload AR
     to populate" when Z1=0 (no AR uploaded), otherwise the live
     ⚪/✓/⚠ flag string. K10:L13 was confirmed empty in the v0.2.10
     bundled Analyzer (verified pre-migration).

  2. Cover row 11 — "AR Module" version line at A11/B11. Sits in the
     existing empty row between T12 Normalizer (R10) and the Links
     section header (R12), so no row inserts needed.

OPERATIONS:

  A. Dashboard tile build at K10:L13:
       K10:L10 (merged) — title "BAD DEBT VARIANCE"
                          (Calibri 9pt bold white on blue FF5B9BD5)
       K11:L12 (merged) — flag value formula, wrapped, centered
                          (Calibri 12pt bold navy on white)
       K13:L13 (merged) — footnote "T12 BD − annualized AR write-offs"
                          (Calibri 9pt italic gray on light-gray FFF2F2F2)
                          NOTE: footnote text intentionally does NOT
                          start with "=" — openpyxl auto-classifies any
                          string beginning with "=" as a formula and
                          writes it into the <f> element, which Excel
                          then tries to parse and strips during repair
                          ("Removed Records: Formula from sheet2.xml").
                          The tile-above relationship is structurally
                          obvious; the "=" prefix is implied.
       Tile is dormant when Z1=0 (shows "— upload AR to populate"),
       live when Z1=1.

  B. Cover A11/B11 — "AR Module" / "v0.1.0" stamp. Borrows label-cell
     style from A10 (T12 Normalizer label) and value-cell style from
     B10 (T12 version value) for consistency.

  C. Stamp Cover!B8 → v0.2.11 + AZ4 on all 16 sheets.

Idempotency:
  - Gate: Cover!B8 == "v0.2.11" AND K10 populated AND A11 == "AR Module"
  - Each op guarded so partial-state migrations recover cleanly.
  - K10 build skipped if a value is already there (regardless of which
    state — protects user hand-edits).

BUNDLED FILE STATUS:
  - After this migration runs on the bundled v0.2.10 Analyzer, both the
    chain output and the bundled file will be at v0.2.11 with identical
    Dashboard + Cover state. (Same shipping pattern as v0.2.10.)

Usage:
    python tools/migration/migrate_to_v0211.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SUBSTRATE_FROM = "v0.2.10"
SUBSTRATE_TO = "v0.2.11"

# 16-sheet anchor list — same as v0.2.10 (no sheet adds/removes).
ANCHOR_SHEETS = (
    "Cover", "Dashboard",
    "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending",
    "AR & Collections",
    "UW Output", "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# Dashboard tile cell map (K10:L13)
TILE_TITLE_RANGE = "K10:L10"
TILE_TITLE_CELL = "K10"
TILE_VALUE_RANGE = "K11:L12"
TILE_VALUE_CELL = "K11"
TILE_FOOTNOTE_RANGE = "K13:L13"
TILE_FOOTNOTE_CELL = "K13"

TILE_TITLE_TEXT = "BAD DEBT VARIANCE"
TILE_VALUE_FORMULA = (
    "=IF('AR & Collections'!Z1=0,"
    "\"— upload AR to populate\","
    "'AR & Collections'!C56)"
)
# Footnote text intentionally does NOT start with "=" — see top-of-file
# note. openpyxl's value setter classifies leading-"=" strings as formulas
# and Excel rejects them on open (sheet2.xml repair). The tile-above
# relationship makes the "= ..." prefix implicit.
TILE_FOOTNOTE_TEXT = "T12 bad debt − annualized AR write-offs"

# Cover AR module line
COVER_AR_LABEL_CELL = "A11"
COVER_AR_VALUE_CELL = "B11"
COVER_AR_LABEL = "AR Module"
COVER_AR_VERSION = "v0.1.0"

# Style references (match REVPOR tile at K5-K8)
NAVY = "FF1F4E79"
TILE_TITLE_FILL = "FF5B9BD5"
TILE_FOOTNOTE_FILL = "FFF2F2F2"
WHITE = "FFFFFFFF"
DARK_GRAY = "FF595959"


def is_already_v0211(wb) -> bool:
    if wb["Cover"]["B8"].value != SUBSTRATE_TO:
        return False
    ws = wb["Dashboard"]
    if ws[TILE_TITLE_CELL].value != TILE_TITLE_TEXT:
        return False
    if wb["Cover"][COVER_AR_LABEL_CELL].value != COVER_AR_LABEL:
        return False
    return True


# ---------------------------------------------------------------------------
# Step A — build Dashboard variance tile at K10:L13
# ---------------------------------------------------------------------------

def build_dashboard_tile(wb) -> dict:
    """Construct the AR variance tile at Dashboard!K10:L13.

    Returns counts dict.
    """
    ws = wb["Dashboard"]
    n = {"cells_written": 0, "merges": 0}

    # Idempotency: skip if title cell already populated (regardless of
    # source — protects against double-stamping AND against user hand-
    # edits in that range).
    if ws[TILE_TITLE_CELL].value is not None and ws[TILE_TITLE_CELL].value != "":
        return n

    # Borrow style from REVPOR tile (K5:K8) for consistency
    src_title = ws["K5"]
    src_value = ws["K6"]
    src_footnote = ws["K8"]

    # --- K10: title ---
    title = ws[TILE_TITLE_CELL]
    title.value = TILE_TITLE_TEXT
    title.font = Font(
        name=src_title.font.name or "Calibri",
        size=src_title.font.size or 9,
        bold=True,
        color=WHITE,
    )
    title.fill = PatternFill(
        start_color=TILE_TITLE_FILL,
        end_color=TILE_TITLE_FILL,
        fill_type="solid",
    )
    title.alignment = Alignment(horizontal="center", vertical="center")
    # Extend fill to L10 before merging
    ws["L10"].fill = copy(title.fill)
    ws.merge_cells(TILE_TITLE_RANGE)
    n["cells_written"] += 1
    n["merges"] += 1

    # --- K11:L12: value (merged, formula) ---
    value = ws[TILE_VALUE_CELL]
    value.value = TILE_VALUE_FORMULA
    value.font = Font(
        name="Calibri",
        size=12,            # smaller than headline tile (28pt) — string can be long
        bold=True,
        color=NAVY,
    )
    value.fill = PatternFill(
        start_color=WHITE,
        end_color=WHITE,
        fill_type="solid",
    )
    value.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    # Match fill across the merge range
    for cell_ref in ("L11", "K12", "L12"):
        ws[cell_ref].fill = copy(value.fill)
    ws.merge_cells(TILE_VALUE_RANGE)
    n["cells_written"] += 1
    n["merges"] += 1

    # --- K13: footnote ---
    foot = ws[TILE_FOOTNOTE_CELL]
    foot.value = TILE_FOOTNOTE_TEXT
    foot.font = Font(
        name="Calibri",
        size=9,
        italic=True,
        color=DARK_GRAY,
    )
    foot.fill = PatternFill(
        start_color=TILE_FOOTNOTE_FILL,
        end_color=TILE_FOOTNOTE_FILL,
        fill_type="solid",
    )
    foot.alignment = Alignment(horizontal="center", vertical="center")
    ws["L13"].fill = copy(foot.fill)
    ws.merge_cells(TILE_FOOTNOTE_RANGE)
    n["cells_written"] += 1
    n["merges"] += 1

    return n


# ---------------------------------------------------------------------------
# Step B — Cover AR module version line at A11/B11
# ---------------------------------------------------------------------------

def add_cover_ar_line(wb) -> bool:
    ws = wb["Cover"]

    # Idempotency
    if ws[COVER_AR_LABEL_CELL].value == COVER_AR_LABEL:
        return False

    # Guard: only write if A11 currently empty (don't overwrite hand-edits)
    if ws[COVER_AR_LABEL_CELL].value not in (None, ""):
        print(f"  WARN: Cover!{COVER_AR_LABEL_CELL} non-empty "
              f"({ws[COVER_AR_LABEL_CELL].value!r}); skipping AR line.")
        return False

    # Borrow styles from T12 Normalizer line (A10/B10) for consistency
    src_label = ws["A10"]
    src_value = ws["B10"]

    label = ws[COVER_AR_LABEL_CELL]
    label.value = COVER_AR_LABEL
    if src_label.has_style:
        label.font = copy(src_label.font)
        label.fill = copy(src_label.fill)
        label.border = copy(src_label.border)
        label.alignment = copy(src_label.alignment)
        label.number_format = src_label.number_format

    value = ws[COVER_AR_VALUE_CELL]
    value.value = COVER_AR_VERSION
    if src_value.has_style:
        value.font = copy(src_value.font)
        value.fill = copy(src_value.fill)
        value.border = copy(src_value.border)
        value.alignment = copy(src_value.alignment)
        value.number_format = src_value.number_format

    return True


# ---------------------------------------------------------------------------
# Step C — version stamping
# ---------------------------------------------------------------------------

def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


# ---------------------------------------------------------------------------
def verify_migration(wb) -> dict:
    r = {}

    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    # Cover AR line
    r["cover_ar_label"] = wb["Cover"][COVER_AR_LABEL_CELL].value
    r["cover_ar_label_ok"] = r["cover_ar_label"] == COVER_AR_LABEL
    r["cover_ar_value"] = wb["Cover"][COVER_AR_VALUE_CELL].value
    r["cover_ar_value_ok"] = r["cover_ar_value"] == COVER_AR_VERSION

    # Dashboard tile
    ws = wb["Dashboard"]
    r["tile_title"] = ws[TILE_TITLE_CELL].value
    r["tile_title_ok"] = r["tile_title"] == TILE_TITLE_TEXT
    r["tile_value_formula"] = ws[TILE_VALUE_CELL].value
    r["tile_value_refs_ar"] = (
        isinstance(r["tile_value_formula"], str)
        and "'AR & Collections'!Z1=0" in r["tile_value_formula"]
        and "'AR & Collections'!C56" in r["tile_value_formula"]
    )
    r["tile_footnote"] = ws[TILE_FOOTNOTE_CELL].value
    r["tile_footnote_ok"] = r["tile_footnote"] == TILE_FOOTNOTE_TEXT

    # Verify the three merges exist
    merge_strs = {str(mr) for mr in ws.merged_cells.ranges}
    r["title_merge_ok"] = TILE_TITLE_RANGE in merge_strs
    r["value_merge_ok"] = TILE_VALUE_RANGE in merge_strs
    r["footnote_merge_ok"] = TILE_FOOTNOTE_RANGE in merge_strs

    # Anchor stamps
    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    # Sheet count should stay at 16 (no adds/removes)
    r["sheet_count"] = len(wb.sheetnames)
    r["sheet_count_ok"] = r["sheet_count"] == 16

    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v0211(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    cur_version = wb["Cover"]["B8"].value
    if cur_version != SUBSTRATE_FROM:
        print(f"  WARN: Cover!B8 = {cur_version!r}, expected {SUBSTRATE_FROM!r}. "
              f"Proceeding anyway.")

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")

    # Step A
    n = build_dashboard_tile(wb)
    if n["cells_written"]:
        print(f"  A: built variance tile at Dashboard!K10:L13 — "
              f"{n['cells_written']} cells, {n['merges']} merges")
    else:
        print(f"  A: Dashboard tile already present — skipped")

    # Step B
    added = add_cover_ar_line(wb)
    if added:
        print(f"  B: added Cover!{COVER_AR_LABEL_CELL}/{COVER_AR_VALUE_CELL} "
              f"AR module line — {COVER_AR_LABEL} / {COVER_AR_VERSION}")
    else:
        print(f"  B: Cover AR module line already present or A11 non-empty")

    # Step C
    stamp_versions(wb)
    print(f"  C: stamped substrate version -> {SUBSTRATE_TO} on "
          f"Cover!B8 + {len(ANCHOR_SHEETS)} AZ4 anchors")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:14s}                       : {r['cover_b8_ok']}")
    print(f"  Cover!A11 = 'AR Module'                          : {r['cover_ar_label_ok']}")
    print(f"  Cover!B11 = 'v0.1.0'                             : {r['cover_ar_value_ok']}")
    print(f"  Dashboard!K10 title = 'BAD DEBT VARIANCE'        : {r['tile_title_ok']}")
    print(f"  Dashboard!K11 formula refs AR sheet              : {r['tile_value_refs_ar']}")
    print(f"  Dashboard!K13 footnote present                   : {r['tile_footnote_ok']}")
    print(f"  K10:L10 merge present                            : {r['title_merge_ok']}")
    print(f"  K11:L12 merge present                            : {r['value_merge_ok']}")
    print(f"  K13:L13 merge present                            : {r['footnote_merge_ok']}")
    print(f"  Sheet count = {r['sheet_count']} (expected 16)                   : {r['sheet_count_ok']}")
    print(f"  All 16 AZ4 = {SUBSTRATE_TO}                          : {r['az4_all']} ({r['az4_count']} sheets)")

    all_ok = all([
        r["cover_b8_ok"], r["cover_ar_label_ok"], r["cover_ar_value_ok"],
        r["tile_title_ok"], r["tile_value_refs_ar"], r["tile_footnote_ok"],
        r["title_merge_ok"], r["value_merge_ok"], r["footnote_merge_ok"],
        r["sheet_count_ok"], r["az4_all"],
    ])
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v0211.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
