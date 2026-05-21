"""
migrate_to_v029.py — Substrate template v0.2.8 → v0.2.9

Closes UW-BACKLOG **BL-0020** (Dashboard chart-data-link bug fixes) — as a
proper forward-rolling migration on the chain.

History / why this exists:
  BL-0020's three chart-data-link fixes were originally implemented as a
  `migrate_to_v028.py` on branch `claude/bl-0020-dashboard-data-link-fixes`
  (PR #34), but that PR was closed unmerged when the user opted to
  wholesale-replace the bundled Analyzer with their hand-edited copy
  (BL-0021). The fixes ended up present ONLY in the bundled file, not in
  any migration on `main`. Meanwhile the `v0.2.8` number was re-used on
  `main` for BL-0022 (`Cover!B5` resolver). So:

    - main's `migrate_to_v028.py` = BL-0022 (Cover!B5)
    - the v027 Dashboard asset still ships the THREE chart bugs

  Net effect: anyone forward-rolling a workbook through the chain gets a
  Dashboard with broken charts (v0.2.7 inserts the buggy Dashboard from
  `v027_assets/dashboard_template.xlsx`; nothing downstream fixes it).
  This migration restores reproducibility: after v0.2.9 the chain produces
  a Dashboard with correct charts. It also resolves the v0.2.8 collision —
  BL-0020 is now unambiguously v0.2.9, BL-0022 keeps v0.2.8.

THE THREE BUGS (as they exist in the v0.2.7-produced Dashboard):

  1. "Monthly EGI Trend" line chart (data at Dashboard!C97:C108) points at
     `Monthly Trending!B21:M21` — row 21 = Housekeeping Income since
     v0.1.7 / BL-0001 moved EGI to row 26. Fix: row 21 → 26.

  2. "Payer Mix — Revenue Share" pie (data at Dashboard!F90:F93) points at
     `Rent Roll Recon!B40:B43` — column B = COUNTIFS unit counts. The
     chart title says Revenue Share, so the source should be col I
     (`H/H47` revenue ratios). Fix: col B → I.

  3. Doughnut chart [1] series range `Dashboard!$O$8:$O$19` (12 rows) holds
     payer labels at O8 + O14:O19 with a 5-row gap (O9:O13 empty) — Excel
     renders 5 empty slices. Fix: move the O14:O19 payer rows up to
     O9:O14 (contiguous with O8), clear O15:O19, and shrink the chart
     series range to `Dashboard!$O$8:$O$14` (cat) / `$P$8:$P$14` (val).

This is a surgical cell + chart patch — NO template asset, NO sheet
add/remove, NO row inserts. Charts [3] (EGI) and [4] (Payer Mix) auto-pick
up corrected values because only the underlying C/F cell formulas change,
not those charts' own ranges. Only chart [1] (doughnut) needs a range
mutation because its SOURCE RANGE changes.

Idempotency:
  - Gate: `Cover!B8 == "v0.2.9"`.
  - The C97:C108 / F90:F93 rewrites are self-idempotent (rebuilding to the
    target ref is a no-op if already there).
  - The doughnut data-move is the ONLY non-idempotent op, so it is GUARDED:
    it only fires when the Dashboard is in the buggy state (`O9` empty AND
    `O14 == "Medicaid"`). If applied to an already-fixed Dashboard (e.g.
    the user's bundled copy), the move is skipped — preventing corruption.

Usage:
    python tools/migration/migrate_to_v029.py input.xlsx output.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.chart.data_source import NumData, NumVal, StrData, StrVal

SUBSTRATE_FROM = "v0.2.8"
SUBSTRATE_TO = "v0.2.9"

DASHBOARD_SHEET = "Dashboard"

# 15-sheet anchor list — unchanged from v0.2.7 / v0.2.8 (no sheet adds/removes).
ANCHOR_SHEETS = (
    "Cover", "Dashboard",
    "T12 Analytics", "T12 Input", "T12 Raw Data",
    "Rent Roll Input", "Rent Roll Recon", "Monthly Trending", "UW Output",
    "UW Export",
    "Mapping Review", "Description_Map", "RR_Calc", "T12_Calc",
    "Workbook Health",
)

# Fix 1: EGI series. C97 -> col B, C98 -> col C, ..., C108 -> col M.
EGI_FIRST_ROW = 97
EGI_LAST_ROW = 108

# Fix 2: Payer Mix pie. F90 -> Rent Roll Recon row 40, ..., F93 -> row 43.
PAYER_FIRST_ROW = 90
PAYER_LAST_ROW = 93

# Fix 3: doughnut payer rows. Buggy data sits at O8 + O14:O19; fixed layout
# is contiguous O8:O14. The move shifts O14:O19 -> O9:O14.
DOUGHNUT_SRC_ROWS = range(14, 20)   # O14:O19 / P14:P19 (the misplaced rows)
DOUGHNUT_DST_START = 9              # first destination row (O9 / P9)


def is_already_v029(wb) -> bool:
    return wb["Cover"]["B8"].value == SUBSTRATE_TO


def fix_egi_series(ws) -> int:
    """Rewrite C97:C108 from Monthly Trending row 21 -> row 26."""
    n = 0
    for i, r in enumerate(range(EGI_FIRST_ROW, EGI_LAST_ROW + 1)):
        col = chr(ord("B") + i)  # B..M
        target = f"=IFERROR('Monthly Trending'!{col}26,0)"
        if ws[f"C{r}"].value != target:
            ws[f"C{r}"].value = target
            n += 1
    return n


def fix_payer_pie(ws) -> int:
    """Rewrite F90:F93 from Rent Roll Recon col B -> col I (both refs in the cell)."""
    n = 0
    for r in range(PAYER_FIRST_ROW, PAYER_LAST_ROW + 1):
        recon_row = 40 + (r - PAYER_FIRST_ROW)
        cur = ws[f"F{r}"].value
        if isinstance(cur, str) and f"!B{recon_row}" in cur:
            ws[f"F{r}"].value = cur.replace(f"!B{recon_row}", f"!I{recon_row}")
            n += 1
    return n


def fix_doughnut_layout(ws) -> bool:
    """Move misplaced payer rows O14:O19 -> O9:O14 (guarded; only if buggy)."""
    buggy = ws["O9"].value is None and ws["O14"].value == "Medicaid"
    if not buggy:
        return False

    from copy import copy as _copy

    # Move each source row up to its destination, copying value + style.
    for offset, src_row in enumerate(DOUGHNUT_SRC_ROWS):
        dst_row = DOUGHNUT_DST_START + offset
        for col in ("O", "P"):
            src = ws[f"{col}{src_row}"]
            dst = ws[f"{col}{dst_row}"]
            dst.value = src.value
            if src.has_style:
                dst.font = _copy(src.font)
                dst.fill = _copy(src.fill)
                dst.border = _copy(src.border)
                dst.alignment = _copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = _copy(src.protection)

    # Clear the now-vacated tail rows (everything past the new contiguous block).
    new_last = DOUGHNUT_DST_START + len(DOUGHNUT_SRC_ROWS) - 1  # O14
    for r in range(new_last + 1, 20):  # O15:O19 / P15:P19
        for col in ("O", "P"):
            ws[f"{col}{r}"].value = None
    return True


def fix_doughnut_chart(ws) -> bool:
    """Shrink doughnut chart [1] series range O8:O19 -> O8:O14 and rebuild caches."""
    if len(ws._charts) < 2:
        return False
    ch = ws._charts[1]
    changed = False
    for s in ch.series:
        if s.cat and s.cat.strRef and s.cat.strRef.f and "$O$8:$O$19" in s.cat.strRef.f:
            s.cat.strRef.f = "Dashboard!$O$8:$O$14"
            labels = [ws[f"O{r}"].value for r in range(8, 15)]
            s.cat.strRef.strCache = StrData(
                ptCount=len(labels),
                pt=[StrVal(idx=i, v=("" if v is None else str(v))) for i, v in enumerate(labels)],
            )
            changed = True
        if s.val and s.val.numRef and s.val.numRef.f and "$P$8:$P$19" in s.val.numRef.f:
            s.val.numRef.f = "Dashboard!$P$8:$P$14"
            s.val.numRef.numCache = NumData(
                formatCode="General",
                ptCount=7,
                pt=[NumVal(idx=i, v=0.0) for i in range(7)],
            )
            changed = True
    return changed


def stamp_versions(wb) -> None:
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B8"] = SUBSTRATE_TO
    for s in ANCHOR_SHEETS:
        if s in wb.sheetnames:
            wb[s]["AZ4"] = SUBSTRATE_TO


def verify_migration(wb) -> dict:
    r = {}
    r["cover_b8"] = wb["Cover"]["B8"].value
    r["cover_b8_ok"] = r["cover_b8"] == SUBSTRATE_TO

    az4 = {s: wb[s]["AZ4"].value for s in ANCHOR_SHEETS if s in wb.sheetnames}
    r["az4_all"] = all(v == SUBSTRATE_TO for v in az4.values())
    r["az4_count"] = len(az4)

    ws = wb[DASHBOARD_SHEET] if DASHBOARD_SHEET in wb.sheetnames else None
    if ws is None:
        for k in ("egi_ok", "payer_ok", "doughnut_data_ok", "doughnut_chart_ok"):
            r[k] = False
        return r

    # Fix 1: every C97:C108 references row 26.
    r["egi_ok"] = all(
        isinstance(ws[f"C{r_}"].value, str) and "26,0)" in ws[f"C{r_}"].value
        for r_ in range(EGI_FIRST_ROW, EGI_LAST_ROW + 1)
    )
    # Fix 2: every F90:F93 references col I (no col B).
    r["payer_ok"] = all(
        isinstance(ws[f"F{r_}"].value, str)
        and f"!I{40 + (r_ - PAYER_FIRST_ROW)}" in ws[f"F{r_}"].value
        and f"!B{40 + (r_ - PAYER_FIRST_ROW)}" not in ws[f"F{r_}"].value
        for r_ in range(PAYER_FIRST_ROW, PAYER_LAST_ROW + 1)
    )
    # Fix 3a: payer rows contiguous O8:O14, tail O15:O19 empty.
    contiguous = all(ws[f"O{r_}"].value is not None for r_ in range(8, 15))
    tail_empty = all(ws[f"O{r_}"].value is None for r_ in range(15, 20))
    r["doughnut_data_ok"] = contiguous and tail_empty
    # Fix 3b: chart [1] series range shrunk.
    chart_ok = False
    if len(ws._charts) >= 2:
        for s in ws._charts[1].series:
            cat = s.cat.strRef.f if s.cat and s.cat.strRef else ""
            val = s.val.numRef.f if s.val and s.val.numRef else ""
            if "$O$8:$O$14" in (cat or "") and "$P$8:$P$14" in (val or ""):
                chart_ok = True
    r["doughnut_chart_ok"] = chart_ok
    return r


def main(input_path: str, output_path: str) -> int:
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src)

    if is_already_v029(wb):
        print(f"Workbook is already at {SUBSTRATE_TO}. No-op (will re-save).")
        wb.save(dst)
        return 0

    if DASHBOARD_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"No '{DASHBOARD_SHEET}' sheet — run migrate_to_v027.py first "
            "(this migration patches the Dashboard the v0.2.7 step inserts)."
        )

    print(f"Migrating {SUBSTRATE_FROM} -> {SUBSTRATE_TO}...")
    ws = wb[DASHBOARD_SHEET]

    n_egi = fix_egi_series(ws)
    print(f"  A: EGI series — rewrote {n_egi} cells (C97:C108 -> Monthly Trending row 26)")

    n_payer = fix_payer_pie(ws)
    print(f"  B: Payer Mix pie — rewrote {n_payer} cells (F90:F93 -> Rent Roll Recon col I)")

    moved = fix_doughnut_layout(ws)
    print(f"  C: doughnut data move (O14:O19 -> O9:O14): {'applied' if moved else 'skipped (not in buggy state)'}")

    chart_fixed = fix_doughnut_chart(ws)
    print(f"  D: doughnut chart range O8:O19 -> O8:O14: {'applied' if chart_fixed else 'skipped (already shrunk)'}")

    stamp_versions(wb)
    print(f"  E: stamped substrate version -> {SUBSTRATE_TO} on Cover!B8 + {len(ANCHOR_SHEETS)} AZ4 anchors")

    print(f"Saving to {dst}...")
    wb.save(dst)

    print(f"Verifying {dst}...")
    wb2 = openpyxl.load_workbook(dst)
    r = verify_migration(wb2)

    print()
    print("=== Verification ===")
    print(f"  Cover!B8 = {r['cover_b8']!r:24s}     : {r['cover_b8_ok']}")
    print(f"  Fix 1 — C97:C108 all ref row 26          : {r['egi_ok']}")
    print(f"  Fix 2 — F90:F93 all ref col I            : {r['payer_ok']}")
    print(f"  Fix 3a — payer rows contiguous O8:O14    : {r['doughnut_data_ok']}")
    print(f"  Fix 3b — doughnut chart range O8:O14     : {r['doughnut_chart_ok']}")
    print(f"  All 15 AZ4 = {SUBSTRATE_TO}                    : {r['az4_all']} ({r['az4_count']} sheets)")

    all_ok = (
        r["cover_b8_ok"]
        and r["egi_ok"] and r["payer_ok"]
        and r["doughnut_data_ok"] and r["doughnut_chart_ok"]
        and r["az4_all"]
    )
    print()
    print("=== " + ("[OK] Migration complete" if all_ok else "[FAIL] Migration incomplete") + " ===")
    return 0 if all_ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_to_v029.py input.xlsx output.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
