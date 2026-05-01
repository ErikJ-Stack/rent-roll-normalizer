"""
T12 Writer
==========

Loads a user-supplied T12 intake workbook, writes the translated rent roll
into the "Rent Roll Input" sheet starting at row 7 (cols A through S), and
returns the modified workbook as bytes.

Critical preservation guarantees:
  - All other tabs are untouched
  - Cols T and U (formulas) are untouched
  - Data validations on cols D, E, F, K, P remain in place
  - Header/formatting on rows 1-6 unchanged
  - Old data in A7:S606 from previous runs is cleared before writing
    (so the file is idempotent — re-running on a different rent roll
    won't leave ghost rows from the prior run)

Limit: max 600 bed rows per run (matches the formula extent in cols T-U).
"""

from __future__ import annotations

import datetime as dt
import io
from typing import Optional

import openpyxl
import pandas as pd


# Layout constants — match the T12 we inspected
SHEET_NAME = "Rent Roll Input"
DATA_START_ROW = 7
DATA_END_ROW = 606  # also the formula extent in cols T, U
COL_A_TO_R_COUNT = 18  # 18 columns from Condensed_RR
COL_S_INDEX = 19       # Period Date column

# The 18 source columns in the order the T12 expects them.
# These names must match the Condensed_RR column names exactly.
SOURCE_COLUMNS_A_TO_R = [
    "Unit #",          # A
    "Room #",          # B
    "Sq Ft",           # C
    "Care Type",       # D
    "Status",          # E
    "Apt Type",        # F
    "Market Rate",     # G
    "Actual Rate",     # H
    "Concession $",    # I
    "Concession End Date",  # J
    "Care Level",      # K
    "Care Level $", # L
    "Med Mgmt $",      # M
    "Pharmacy $",      # N
    "Other LOC $",     # O
    "Payer Type",      # P
    "Move-in Date",    # Q
    "Resident Name",   # R
]


class T12CapacityError(Exception):
    """Raised when the rent roll has more rows than the T12 can hold."""


def _coerce_value(v):
    """Convert pandas/numpy scalars to native Python types openpyxl can write.

    Handles: NaN/NaT → None, numpy ints/floats → int/float, pandas Timestamps
    → datetime.date, anything else → unchanged.
    """
    if v is None:
        return None
    # Pandas null sentinels
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        # pd.isna on some array-like objects raises; skip for those
        pass

    # Pandas/numpy datetime-like → date
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime().date()
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v

    # Numpy scalars
    if hasattr(v, "item"):
        try:
            return v.item()
        except (ValueError, AttributeError):
            pass

    return v


def populate_t12(
    t12_bytes: bytes,
    translated_df: pd.DataFrame,
    period_date: Optional[dt.date],
) -> bytes:
    """Populate a T12 workbook with translated rent roll data and return as bytes.

    Args:
        t12_bytes: Raw bytes of the user-uploaded T12 .xlsx file
        translated_df: DataFrame from t12_translator.translate_for_t12()
            — must have the 18 source columns in SOURCE_COLUMNS_A_TO_R
        period_date: Date written to col S on every populated row. Required.

    Raises:
        T12CapacityError: if the rent roll exceeds DATA_END_ROW - DATA_START_ROW + 1 rows
        ValueError: if the T12 doesn't contain the expected sheet
    """
    if period_date is None:
        raise ValueError("period_date is required to populate the T12.")

    n_rows = len(translated_df)
    max_rows = DATA_END_ROW - DATA_START_ROW + 1
    if n_rows > max_rows:
        raise T12CapacityError(
            f"Rent roll has {n_rows} bed rows, but the T12 'Rent Roll Input' "
            f"sheet's formulas only extend to row {DATA_END_ROW} "
            f"(max {max_rows} rows). Either trim the rent roll or extend "
            f"the T12 formulas to additional rows."
        )

    # Load the user's T12 (preserves formulas, formatting, validations, other tabs)
    wb = openpyxl.load_workbook(io.BytesIO(t12_bytes), data_only=False)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"T12 workbook is missing the required '{SHEET_NAME}' sheet. "
            f"Found sheets: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]

    # --- Step 1: Clear any pre-existing data in A7:S606 -------------------
    # This makes the operation idempotent. We only clear cols A-S; cols T-U
    # contain formulas which we leave alone.
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        for c in range(1, COL_S_INDEX + 1):  # cols 1-19 = A-S
            ws.cell(row=r, column=c).value = None

    # --- Step 2: Write the translated rent roll ---------------------------
    # Match by column NAME (not position) — defensive against future
    # Condensed_RR column reordering.
    missing = [c for c in SOURCE_COLUMNS_A_TO_R if c not in translated_df.columns]
    if missing:
        raise ValueError(
            f"Translated DataFrame is missing required columns: {missing}"
        )

    for i, (_, row) in enumerate(translated_df.iterrows()):
        excel_row = DATA_START_ROW + i
        # Cols A-R (1-18) ← the 18 source columns in fixed order
        for col_idx, src_col in enumerate(SOURCE_COLUMNS_A_TO_R, start=1):
            value = _coerce_value(row[src_col])
            ws.cell(row=excel_row, column=col_idx).value = value
        # Col S (19) ← period date on every populated row, formatted as date
        s_cell = ws.cell(row=excel_row, column=COL_S_INDEX)
        s_cell.value = period_date
        s_cell.number_format = "mm/dd/yyyy"

    # --- Step 3: Save and return ------------------------------------------
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
