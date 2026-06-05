"""
MF rent-roll normalizer — parse an operator Rent Roll into per-unit records
mapped to the MF UW Model's `Rent Roll Analysis` grid (cols A–AK, anchor A273).

Handles three row shapes (auto-detected): (a) one-row-per-unit (Yardi-CIM "Rent
Roll - Cim", Hidden Lakes — 143 units); (b) the itemized "Rent Roll
(Operations) - Rent Related Charge Codes" format (Avana — 263 units) where each
unit's identity is on a header row and its charges (Amenity/Base/etc.) are
itemized across continuation rows with a blank Bldg-Unit, summed into the unit;
and (c) the RealPage **OneSite "RENT ROLL DETAIL"** export (Ascend Brunswick —
334 units) where a unit repeats across multiple lease rows (current resident +
a future Applicant / Pending-renewal row) and charges are spread *horizontally*
across per-code columns (RENT / INTERNET / TRASH / …). OneSite lease rows are
deduped to one record per physical unit.
Header-driven column mapping (needle-priority) so it tolerates column reordering. AR aging (cols Q–T)
is filled later by `mf_ar_parser` via a Bldg-Unit join; the W–AK ancillary
breakouts are best-effort from the redIQ Sortable-RR (decision §2.7.2) and are
left empty here (the OneSite path fills them from its per-code columns).

File formats: .xlsx / .xlsm (openpyxl) and legacy .xls (xlrd — OneSite/RealPage
and older Yardi exports), auto-detected by OLE2 magic bytes.

Public API:
    parse_mf_rr(source) -> MFRRResult
        source: path str | pathlib.Path | bytes | file-like (BytesIO w/ .name)
"""
from __future__ import annotations

import datetime as _dt
import io
import re
from dataclasses import dataclass, field

import openpyxl

from mf_mappings import UNMAPPED_STATUS, classify_charge_code, normalize_status

# Header label -> canonical field. Matched case-insensitively on a normalized
# (lowercased, punctuation-stripped) header string; first containing match wins.
# Ordered by PRIORITY (not column order): the first needle that finds a column
# claims the field. So "unit type" wins over "floor plan", "unit status" over a
# bare "occupancy", etc. Matched against the normalized (lowercased,
# punctuation-stripped) header text.
_HEADER_MAP = [
    ("bldg-unit", "bldg_unit"), ("bldg", "bldg_unit"), ("unit no", "bldg_unit"),
    ("unit id", "bldg_unit"),
    ("unit type", "unit_type"), ("floor plan", "unit_type"),
    ("sqft", "sqft"), ("sq ft", "sqft"), ("net sf", "sqft"), ("square f", "sqft"),
    ("unit status", "status"), ("occupancy type", "status"), ("status", "status"),
    ("resident", "resident"), ("tenant", "resident"),
    ("charge code", "charge_code"),
    ("move in", "move_in"), ("move-in", "move_in"),
    ("lease start", "lease_start"), ("lease sign", "lease_start"),
    ("lease end", "lease_end"), ("lease exp", "lease_end"), ("expiration", "lease_end"),
    ("expected move", "exp_move_out"), ("move out", "exp_move_out"), ("move-out", "exp_move_out"),
    ("market rent", "market_rent"), ("gpr market", "market_rent"),
    ("actual charge", "actual_charges"), ("recurring", "actual_charges"),
    ("scheduled charge", "scheduled_charges"),
    ("balance", "balance"),
    ("deposit", "deposit"),
]


def unit_key(s: str) -> str:
    """Normalized Bldg-Unit join key (uppercase, alnum only). 'A1' / 'A-1' -> 'A1'."""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


@dataclass
class MFUnit:
    bldg_unit: str = ""       # raw source Bldg-Unit (join key for AR)
    bldg: str = ""
    unit: str = ""
    unit_type: str = ""
    sqft: float | None = None
    status: str = ""
    resident: str = ""
    legal: bool = False
    move_in: _dt.datetime | None = None
    lease_start: _dt.datetime | None = None
    lease_end: _dt.datetime | None = None
    exp_move_out: _dt.datetime | None = None
    market_rent: float = 0.0
    actual_charges: float = 0.0
    scheduled_charges: float = 0.0
    balance: float = 0.0
    deposit: float = 0.0
    # filled later by mf_ar_parser (Q–T) + redIQ Source Data (W–AK)
    ar_0_30: float = 0.0
    ar_31_60: float = 0.0
    ar_61_90: float = 0.0
    ar_90_plus: float = 0.0
    notes: str = ""
    ancillary: dict = field(default_factory=dict)


@dataclass
class MFRRResult:
    units: list[MFUnit]
    period_hint: str = ""
    property_hint: str = ""
    warnings: list[str] = field(default_factory=list)

    @property
    def unit_count(self) -> int:
        return len(self.units)

    @property
    def occupied(self) -> int:
        return sum(1 for u in self.units if u.status.startswith("Occupied"))

    @property
    def vacant(self) -> int:
        return sum(1 for u in self.units if u.status.startswith("Vacant"))

    @property
    def legal_count(self) -> int:
        return sum(1 for u in self.units if u.legal)


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9 ]", "", str(s).strip().lower()) if s is not None else ""


def _num(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("$", "").strip()
        neg = s.startswith("(") and s.endswith(")")
        s = s.strip("()")
        try:
            return -float(s) if neg else float(s)
        except ValueError:
            return 0.0
    return 0.0


def _date(v):
    return v if isinstance(v, (_dt.datetime, _dt.date)) else None


def _split_bldg_unit(raw: str) -> tuple[str, str]:
    s = str(raw).strip()
    if "-" in s:
        b, u = s.split("-", 1)
        return b.strip(), u.strip()
    m = re.match(r"^([A-Za-z]+)\s*(\d.*)$", s)
    if m:
        return m.group(1), m.group(2)
    return "", s


_OLE2_MAGIC = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"  # legacy .xls (OLE2 compound doc)


def _to_bytes_and_name(source) -> tuple[bytes, str]:
    """Normalize any accepted source to (raw bytes, name hint)."""
    name = str(getattr(source, "name", "") or "")
    if isinstance(source, (bytes, bytearray)):
        return bytes(source), name
    if hasattr(source, "read"):                # file-like (e.g. Streamlit upload)
        data = source.read()
        try:
            source.seek(0)
        except Exception:
            pass
        return data, name
    name = str(source)                         # path-like
    with open(source, "rb") as fh:
        return fh.read(), name


def _read_grid(source) -> list[tuple]:
    """Read the first worksheet into a list of value-tuples. Supports .xlsx/.xlsm
    (openpyxl) and legacy .xls (xlrd), detected by OLE2 magic bytes so a
    mislabeled extension still routes correctly. .xls date cells are converted
    back to datetimes (xlrd hands them out as serials)."""
    data, name = _to_bytes_and_name(source)
    is_xls = data[:8] == _OLE2_MAGIC or (
        name.lower().endswith(".xls")
        and not name.lower().endswith((".xlsx", ".xlsm")))
    if is_xls:
        import xlrd  # legacy .xls only; pinned in requirements.txt
        book = xlrd.open_workbook(file_contents=data)
        sh = book.sheet_by_index(0)
        dm = book.datemode
        rows: list[tuple] = []
        for r in range(sh.nrows):
            row = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                v = cell.value
                if cell.ctype == 3:            # XL_CELL_DATE -> datetime
                    try:
                        v = _dt.datetime(*xlrd.xldate_as_tuple(v, dm))
                    except Exception:
                        pass
                elif cell.ctype == 0 or v == "":   # empty / blank
                    v = None
                row.append(v)
            rows.append(tuple(row))
        return rows
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        return list(wb.worksheets[0].iter_rows(values_only=True))
    finally:
        wb.close()


def _header_band_property(rows, header_row) -> str:
    """Property name from the header band (col A above the grid header) — the
    operator file carries a clean name (e.g. "Avana Stoney Ridge"); far more
    reliable than parsing the filename."""
    for i in range(header_row):
        v = rows[i][0] if rows[i] else None
        if isinstance(v, str):
            s = v.strip()
            low = s.lower()
            if (s and not re.match(r"\d", s)
                    and not any(k in low for k in ("rent roll", "report", "operations",
                                                   "charge code", "unit details",
                                                   "onesite", "parameters", "as of"))):
                return s
    return ""


def parse_mf_rr(source) -> MFRRResult:
    rows = _read_grid(source)
    maxc = max((len(r) for r in rows), default=0)

    # RealPage OneSite "RENT ROLL DETAIL" is a distinct shape (units repeat
    # across lease rows, charges spread horizontally) — route to its own parser.
    if _is_onesite(rows):
        return _parse_onesite(rows)

    # 1) header row = first row containing a Bldg-Unit-style label + a rent label
    header_row = None
    for i, row in enumerate(rows[:20]):
        norm = [_norm(c) for c in row]
        joined = " | ".join(norm)
        if ("bldg" in joined or "unit no" in joined or "unit id" in joined) and "rent" in joined:
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not locate the rent-roll header row.")

    property_hint = _header_band_property(rows, header_row)

    # 2) map columns — needle PRIORITY (map order), first column that matches wins
    hdr = [_norm(v) for v in rows[header_row]]
    col: dict[str, int] = {}
    for needle, fname in _HEADER_MAP:
        if fname in col:
            continue
        for c, h in enumerate(hdr):
            if h and needle in h:
                col[fname] = c
                break
    if "bldg_unit" not in col or "status" not in col:
        raise ValueError(f"Rent-roll header missing Bldg-Unit/Status (found {sorted(col)}).")

    def g(row, fname):
        idx = col.get(fname)
        return row[idx] if idx is not None and idx < len(row) else None

    # 3) walk data rows into per-unit blocks. A header row carries a Bldg-Unit;
    #    in the itemized "charge codes" format each unit is followed by
    #    continuation rows (blank Bldg-Unit) holding additional charge lines.
    #    Scheduled/Actual charges are summed across the block's charge-code rows
    #    (avoiding the L-blank per-unit total row). Stop at the trailing
    #    summary / future-resident blocks.
    itemized = "charge_code" in col
    units: list[MFUnit] = []
    warnings: list[str] = []
    unknown_status = 0
    cur: MFUnit | None = None

    def _add_charges(u, row):
        sched = _num(g(row, "scheduled_charges"))
        u.scheduled_charges += sched
        u.actual_charges += _num(g(row, "actual_charges"))
        bucket = classify_charge_code(g(row, "charge_code"))   # break out non-rent codes
        if bucket and sched:
            u.ancillary[bucket] = u.ancillary.get(bucket, 0.0) + sched

    for row in rows[header_row + 1:]:
        a = g(row, "bldg_unit")
        if a in (None, ""):
            # continuation charge-code line for the current unit
            if cur is not None and itemized and g(row, "charge_code") not in (None, ""):
                _add_charges(cur, row)
            continue
        a_norm = _norm(a)
        if a_norm.startswith(("future resident", "charge code", "total ", "resident total",
                              "ledger", "selected report", "grand total")):
            break  # end of the unit grid
        status = normalize_status(g(row, "status"))
        if not status or status == UNMAPPED_STATUS:
            cur = None
            if g(row, "status") not in (None, ""):
                unknown_status += 1
            continue

        bldg, unit = _split_bldg_unit(a)
        resident = str(g(row, "resident") or "").strip()
        legal = resident.startswith("**")
        if legal:
            resident = resident.lstrip("*").strip()
        if "vacant" in resident.lower():
            resident = ""

        cur = MFUnit(
            bldg_unit=str(a).strip(), bldg=bldg, unit=unit,
            unit_type=str(g(row, "unit_type") or "").strip(),
            sqft=(_num(g(row, "sqft")) or None),
            status=status, resident=resident, legal=legal,
            move_in=_date(g(row, "move_in")),
            lease_start=_date(g(row, "lease_start")),
            lease_end=_date(g(row, "lease_end")),
            exp_move_out=_date(g(row, "exp_move_out")),
            market_rent=_num(g(row, "market_rent")),
            balance=_num(g(row, "balance")),
            deposit=_num(g(row, "deposit")),
        )
        units.append(cur)
        if itemized:
            # the header row itself may carry the first charge-code line
            if g(row, "charge_code") not in (None, ""):
                _add_charges(cur, row)
        else:
            cur.scheduled_charges = _num(g(row, "scheduled_charges"))
            cur.actual_charges = _num(g(row, "actual_charges"))

    if not units:
        warnings.append("No unit rows parsed — check the rent-roll layout.")
    if unknown_status:
        warnings.append(f"{unknown_status} row(s) had an unrecognized status and were skipped "
                        "(likely charge-code-summary lines).")
    return MFRRResult(units=units, property_hint=property_hint, warnings=warnings)


# ---------------------------------------------------------------------------
# RealPage OneSite "RENT ROLL DETAIL" parser
# ---------------------------------------------------------------------------
# OneSite column header -> field. Bare "unit" (no "no"/"id") is the Bldg-Unit
# identifier here. Each field claims at most one column (used-set), so "lease
# rent" takes col "Lease Rent" before bare "rent" can — order matters below.
_ONESITE_COLS = [
    ("unitlease status", "status"), ("status", "status"),
    ("floorplan", "unit_type"), ("floor plan", "unit_type"),
    ("unit designation", "_skip"),          # claim before bare "unit" needle
    ("unit", "bldg_unit"),
    ("sqft", "sqft"), ("sq ft", "sqft"),
    ("name", "resident"),
    ("movein", "move_in"), ("move in", "move_in"),
    ("moveout", "exp_move_out"), ("move out", "exp_move_out"),
    ("lease start", "lease_start"),
    ("lease end", "lease_end"),
    ("market", "market_rent"),              # "Market + Addl."
    ("dep on hand", "deposit"),
    ("balance", "balance"),
    ("lease rent", "scheduled_charges"),    # base contracted rent (Scheduled GPR)
    ("rent", "actual_charges"),             # actual base rent billed (col "RENT")
    ("total billing", "_total"),
]

# OneSite recurring-charge column header -> W–AK ancillary bucket. Columns not
# listed (RENT = base rent; CONC/UP, CONC/CO, EMPDISC = concessions/discounts)
# are NOT bucketed as ancillary income.
_ONESITE_CHARGE_BUCKETS = {
    "internet": "utility_reimb", "cable": "utility_reimb", "satellite": "utility_reimb",
    "trash": "utility_reimb", "pest": "utility_reimb", "water": "utility_reimb",
    "sewer": "utility_reimb", "package": "package", "petrent": "pet", "pet rent": "pet",
    "garage": "parking", "carport": "parking", "parking": "parking",
    "storage": "storage", "commfee": "admin", "amenity": "amenity", "valet": "valet",
}


def _is_onesite(rows) -> bool:
    """RealPage OneSite signature: a 'Total Billing' or 'Unit/Lease Status' header."""
    for row in rows[:15]:
        joined = " | ".join(_norm(c) for c in row)
        if "total billing" in joined or "unitlease status" in joined:
            return True
    return False


def _is_secondary(status) -> bool:
    """A future-lease row (Applicant / Pending …) — not the unit's primary state."""
    s = str(status or "").lower()
    return "applicant" in s or "pending" in s


def _is_committed_secondary(status) -> bool:
    """An Applicant / Pending-resident row that carries a committed (pre-leased)
    rent — distinct from 'Pending renewal' (a renewal of the current lease)."""
    s = str(status or "").lower()
    return "applicant" in s or "pending resident" in s


def _onesite_period(rows, header_row) -> str:
    """Pull the 'As of Date: mm/dd/yyyy' stamp from the header band, if present."""
    for i in range(header_row):
        v = rows[i][0] if rows[i] else None
        if isinstance(v, str):
            m = re.search(r"as of[^0-9]*(\d{1,2}/\d{1,2}/\d{2,4})", v, re.IGNORECASE)
            if m:
                return m.group(1)
    return ""


def _parse_onesite(rows) -> MFRRResult:
    header_row = None
    for i, row in enumerate(rows[:15]):
        joined = " | ".join(_norm(c) for c in row)
        if "unitlease status" in joined or ("unit" in joined and "total billing" in joined):
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not locate the OneSite rent-roll header row.")

    hdr = [_norm(v) for v in rows[header_row]]
    col: dict[str, int] = {}
    used: set[int] = set()
    for needle, fname in _ONESITE_COLS:
        if fname in col:
            continue
        for c, h in enumerate(hdr):
            if c in used or not h:
                continue
            if needle in h:
                col[fname] = c
                used.add(c)
                break
    if "bldg_unit" not in col or "status" not in col:
        raise ValueError(f"OneSite header missing Unit/Status (found {sorted(col)}).")

    charge_cols: list[tuple[int, str]] = []
    for c, h in enumerate(hdr):
        for key, bucket in _ONESITE_CHARGE_BUCKETS.items():
            if key in h:
                charge_cols.append((c, bucket))
                break

    property_hint = _header_band_property(rows, header_row)
    period_hint = _onesite_period(rows, header_row)

    def g(row, fname):
        idx = col.get(fname)
        return row[idx] if idx is not None and idx < len(row) else None

    # group the lease rows by physical unit, preserving first-seen order
    groups: dict[str, list] = {}
    order: list[str] = []
    for row in rows[header_row + 1:]:
        a = g(row, "bldg_unit")
        if a in (None, ""):
            continue
        a_norm = _norm(a)
        if a_norm.startswith(("future resident", "total ", "grand total", "summary",
                              "selected report", "report ", "ledger")):
            break  # end of the unit grid
        key = str(a).strip()
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(row)

    units: list[MFUnit] = []
    warnings: list[str] = []
    unknown = 0
    for raw_unit in order:
        lease_rows = groups[raw_unit]
        # primary = first unit-state row (not an Applicant/Pending future lease)
        prim = next((r for r in lease_rows if not _is_secondary(g(r, "status"))),
                    lease_rows[0])
        sec = next((r for r in lease_rows if _is_committed_secondary(g(r, "status"))),
                   None)
        status = normalize_status(g(prim, "status"))
        if not status or status == UNMAPPED_STATUS:
            if g(prim, "status") not in (None, ""):
                unknown += 1
            continue

        bldg, unit = _split_bldg_unit(raw_unit)
        resident = str(g(prim, "resident") or "").strip()
        legal = resident.startswith("**")
        if legal:
            resident = resident.lstrip("*").strip()
        if "vacant" in resident.lower():
            resident = ""

        occupied = status.startswith("Occupied")
        # Charge source: occupied -> the current lease (primary row). Vacant but
        # pre-leased -> the committed Applicant row (secondary): scheduled = its
        # lease rent, actual = 0 (not billing yet). Plain vacant/down -> nothing.
        if occupied:
            crow = prim
            actual = _num(g(prim, "actual_charges"))
            sched = _num(g(prim, "scheduled_charges"))
        elif sec is not None:
            crow = sec
            actual = 0.0
            sched = _num(g(sec, "scheduled_charges"))
        else:
            crow = None
            actual = sched = 0.0

        u = MFUnit(
            bldg_unit=raw_unit, bldg=bldg, unit=unit,
            unit_type=str(g(prim, "unit_type") or "").strip(),
            sqft=(_num(g(prim, "sqft")) or None),
            status=status, resident=resident, legal=legal,
            move_in=_date(g(prim, "move_in")),
            lease_start=_date(g(prim, "lease_start")),
            lease_end=_date(g(prim, "lease_end")),
            exp_move_out=_date(g(prim, "exp_move_out")),
            market_rent=_num(g(prim, "market_rent")),
            actual_charges=actual, scheduled_charges=sched,
            balance=_num(g(prim, "balance")),
            deposit=_num(g(prim, "deposit")),
        )
        # Ancillary income only for occupied units (currently realized); a vacant
        # pre-leased unit isn't generating fee income yet.
        if occupied and crow is not None:
            for c, bucket in charge_cols:
                amt = _num(crow[c]) if c < len(crow) else 0.0
                if amt > 0:
                    u.ancillary[bucket] = u.ancillary.get(bucket, 0.0) + amt
        units.append(u)

    if not units:
        warnings.append("No OneSite unit rows parsed — check the rent-roll layout.")
    if unknown:
        warnings.append(f"{unknown} OneSite row(s) had an unrecognized status and were skipped.")
    return MFRRResult(units=units, period_hint=period_hint,
                      property_hint=property_hint, warnings=warnings)


if __name__ == "__main__":
    import sys
    r = parse_mf_rr(sys.argv[1])
    print(f"units={r.unit_count} | occupied={r.occupied} | vacant={r.vacant} | legal={r.legal_count}")
    for w in r.warnings:
        print("  WARN:", w)
    for u in r.units[:3]:
        print("  ", u.bldg, u.unit, u.unit_type, u.status, repr(u.resident),
              u.market_rent, u.scheduled_charges, u.balance)
