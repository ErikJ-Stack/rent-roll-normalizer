"""
MF T-12 normalizer — parse any of the known operator T-12 formats into
standardized GL lines bucketed to `_StdCOA`.

General by design (not 5 brittle per-format branches): it auto-detects the
month-header row, the monthly column set (contiguous OR odd-spaced), the total
column, and whether account numbers are present — then extracts leaf GL lines,
strips subtotals/headers, and classifies each via `mf_mappings`. Validated
against all 5 catalogued formats (PSI flat, QuickBooks nested, Yardi numbered,
Yardi/YSI, Tzadik name-only) — see `tools/mf_uw_template/COA-SEED.md`.

Public API:
    parse_mf_t12(source) -> MFT12Result
        source: path str | pathlib.Path | bytes | file-like (BytesIO w/ .name)
"""
from __future__ import annotations

import datetime as _dt
import io
import re
from dataclasses import dataclass, field

import openpyxl

from mf_mappings import (CONTROL_BUCKETS, EXCLUDED, bucket_side,
                         classify_t12_account)

_MONTH = r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)"
_MONTH_RE = re.compile(rf"^\s*{_MONTH}[a-z]*[\s\-/.,]*\d{{2,4}}\s*$", re.IGNORECASE)
_ACCT_RE = re.compile(r"^\s*\d{4,5}(-\d{1,3})?\b")
_TOTAL_RE = re.compile(r"^\s*total\b", re.IGNORECASE)
# Section/group HEADER words — skipped only when the row carries no value (a true
# header). A valued row with one of these names (e.g. Copeland's "Insurance"
# $231,828 leaf) is a real GL line and must be kept.
_SECTION_RE = re.compile(
    r"^\s*(income|expenses?|operating expenses|other income|rental income|"
    r"ordinary income/expense|pass.?through reimbursement|"
    r"general (maintenance|& administrative)|make.?ready|"
    r"contract services|recreational amenities|office expenses|"
    r"other general & administrative|marketing and leasing|payroll( expense| & benefits)|"
    r"non.?operating|capital|routine replacement|condominium|"
    r"partnership|depreciation|utilities taxes and insurance|repairs & maintenance)\s*$",
    re.IGNORECASE,
)
# Grand-total / P&L summary rows — always skipped even when valued.
_SUMMARY_RE = re.compile(
    r"^\s*(gross profit|operating income|net operating income|net ordinary income|"
    r"net income( before allocations)?|noi after replacements)\s*$",
    re.IGNORECASE,
)


@dataclass
class MFT12Line:
    acct: str | None
    name: str
    monthly: list[float]      # 12 values aligned to month_labels
    total: float
    bucket: str
    side: str                 # bucket side: income | expense | excluded | control
    section: str = "income"   # source section on the statement: income | expense | excluded


@dataclass
class MFT12Result:
    format_guess: str
    month_labels: list[str]
    period: str
    lines: list[MFT12Line]
    unmapped: list[MFT12Line] = field(default_factory=list)
    reported: dict = field(default_factory=dict)     # as-reported Total Income / OpEx / NOI
    computed: dict = field(default_factory=dict)      # income/expense/NOI from mapped leaves
    warnings: list[str] = field(default_factory=list)

    @property
    def coverage(self) -> float:
        n = len(self.lines)
        if not n:
            return 0.0
        mapped = sum(1 for ln in self.lines if ln.bucket not in CONTROL_BUCKETS or ln.bucket == EXCLUDED)
        return mapped / n


# ---------------------------------------------------------------------------
def _is_month(v) -> bool:
    if isinstance(v, (_dt.datetime, _dt.date)):
        return True
    return bool(v) and isinstance(v, str) and bool(_MONTH_RE.match(v))


def _month_label(v) -> str:
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%b %Y")
    return str(v).strip()


def _num(v) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("$", "").strip()
        neg = s.startswith("(") and s.endswith(")")
        s = s.strip("()")
        try:
            return -float(s) if neg else float(s)
        except ValueError:
            return None
    return None


def _load_ws(source):
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    # Prefer a sheet that looks like a statement; default to the first.
    ws = wb.worksheets[0]
    return wb, ws


def parse_mf_t12(source) -> MFT12Result:
    wb, ws = _load_ws(source)
    rows = list(ws.iter_rows(values_only=False))
    maxc = ws.max_column

    # 1) month-header row = row with the most month-like cells (>= 6)
    best_row = best_cols = None
    for r_idx in range(min(len(rows), 20)):
        cols = [c + 1 for c in range(maxc) if _is_month(ws.cell(r_idx + 1, c + 1).value)]
        if len(cols) >= 6 and (best_cols is None or len(cols) > len(best_cols)):
            best_row, best_cols = r_idx + 1, cols
    if not best_cols:
        wb.close()
        raise ValueError("Could not locate a monthly header row (need >= 6 month columns).")
    month_cols = sorted(best_cols)[:12]
    month_labels = [_month_label(ws.cell(best_row, c).value) for c in month_cols]
    first_month_col = month_cols[0]

    # 2) total column: a 'Total' header to the right of months, else last month + 1
    total_col = None
    for c in range(month_cols[-1] + 1, maxc + 1):
        hv = ws.cell(best_row, c).value
        if isinstance(hv, str) and hv.strip().lower() == "total":
            total_col = c
            break
    if total_col is None:
        total_col = month_cols[-1] + 1

    # 3) account-number presence (sampled)
    acct_seen = 0
    for r_idx in range(best_row, min(best_row + 60, len(rows))):
        a = ws.cell(r_idx + 1, 1).value
        if a is not None and _ACCT_RE.match(str(a)):
            acct_seen += 1
    has_acct = acct_seen >= 5

    # 4) extract leaves
    #   Subtotal/rollup detection (cross-format):
    #     - Yardi: account suffix in {098,099,199,090,999} = rollup (e.g. 41999-099).
    #     - any: de-prefixed name starts with "Total" or is a bare section header.
    #     - acct-bearing charts (PSI/Yardi): a leaf must carry an account number
    #       (PSI subtotals like "Net Rent Revenue" have none).
    ROLLUP_SUFFIX = {"098", "099", "199", "090", "999"}
    lines: list[MFT12Line] = []
    reported: dict = {}
    in_expense = False   # for no-account charts: flipped by the EXPENSE header
    for r_idx in range(best_row, len(rows)):
        rr = r_idx + 1
        acct = None
        name = None
        for c in range(1, first_month_col):
            v = ws.cell(rr, c).value
            if v in (None, ""):
                continue
            sv = str(v).strip()
            if _ACCT_RE.match(sv) and (acct is None) and not re.search(r"[a-zA-Z]{3}", sv):
                acct = sv
            else:
                name = sv  # last text cell wins (handles indented + B-col names)
        if name is None and acct is not None:
            name = acct
        if not name:
            continue
        tot = _num(ws.cell(rr, total_col).value)
        monthly = [(_num(ws.cell(rr, c).value) or 0.0) for c in month_cols]
        if tot is None:
            tot = sum(monthly)

        clean = re.sub(r"^\s*\d{4,5}(-\d{1,3})?\s*-\s*", "", name).strip()
        low = clean.lower()
        has_value = (tot not in (None, 0)) or any(monthly)
        is_total = bool(_TOTAL_RE.match(clean))
        is_summary = bool(_SUMMARY_RE.match(clean))
        is_header = bool(_SECTION_RE.match(clean)) and not has_value

        # capture as-reported grand totals (largest match wins) before skipping
        if tot:
            if re.fullmatch(r"total income", low):
                reported["total_income"] = max(reported.get("total_income", 0), tot)
            elif re.search(r"total operating / non|total operating expense|^total expenses?$", low):
                reported["total_expense"] = max(reported.get("total_expense", 0), tot)
            elif re.fullmatch(r"net operating income", low):
                reported["noi"] = tot
            elif low in ("net income", "net ordinary income") and "noi" not in reported:
                reported["noi"] = tot

        # track the income -> expense boundary for no-account charts
        if re.fullmatch(r"expenses?", low):
            in_expense = True

        # --- subtotal / summary / header / empty filters ---
        if is_total or is_summary or is_header:
            continue
        suffix_m = re.match(r"\d{5}-(\d{3})", str(acct) if acct else "")
        if suffix_m and suffix_m.group(1) in ROLLUP_SUFFIX:
            continue  # Yardi rollup account
        if has_acct and acct is None:
            continue  # acct-bearing chart: a leaf must have an account number
        if tot == 0 and not any(monthly):
            continue  # empty placeholder

        bucket = classify_t12_account(acct, clean)
        # source section: by Yardi/PSI account leading digit (4=income, 5/6=expense,
        # 7/8/9=below-NOI); for no-account charts, by the in_expense flag.
        root_m = re.match(r"(\d)", str(acct) if acct else "")
        if root_m:
            lead = root_m.group(1)
            section = "income" if lead == "4" else "expense" if lead in "56" else "excluded"
        else:
            section = "expense" if in_expense else "income"
        if bucket == EXCLUDED:
            section = "excluded"
        lines.append(MFT12Line(acct, clean, monthly, tot, bucket, bucket_side(bucket), section))

    wb.close()

    # 5) format guess
    if has_acct:
        prefixed = sum(1 for ln in lines if ln.acct and ln.name != ln.acct
                       and re.match(rf"^{re.escape(ln.acct)}", str(ln.acct)))
        fmt = "yardi_numbered" if first_month_col >= 3 and acct_seen and any(
            re.match(r"\d{5}-\d{3}", str(l.acct or "")) for l in lines) else "psi_flat"
    elif month_cols != list(range(first_month_col, first_month_col + len(month_cols))):
        fmt = "quickbooks_nested"   # odd-spaced month columns
    elif first_month_col <= 2:
        fmt = "tzadik_nameonly"
    else:
        fmt = "unknown"

    # 6) computed reconciliation — by SOURCE SECTION (so utility-rebill contras in
    #    the expense section reduce opex rather than being mis-summed as income).
    inc = sum(ln.total for ln in lines if ln.section == "income")
    exp = sum(ln.total for ln in lines if ln.section == "expense")
    computed = {"income": inc, "expense": exp, "noi": inc - exp,
                "excluded": sum(ln.total for ln in lines if ln.section == "excluded")}

    unmapped = [ln for ln in lines if ln.bucket == "— UNMAPPED —"]
    warnings = []
    if unmapped:
        warnings.append(f"{len(unmapped)} GL line(s) unmapped — analyst must classify.")
    if reported.get("noi") is not None and abs(reported["noi"] - computed["noi"]) > 1.0:
        warnings.append(
            f"NOI reconciliation off by ${reported['noi'] - computed['noi']:,.2f} "
            f"(reported {reported['noi']:,.0f} vs computed {computed['noi']:,.0f}) — "
            "check utility-rebill section handling / unmapped lines.")

    return MFT12Result(
        format_guess=fmt,
        month_labels=month_labels,
        period=f"{month_labels[0]} – {month_labels[-1]}" if month_labels else "",
        lines=lines, unmapped=unmapped, reported=reported, computed=computed,
        warnings=warnings,
    )


if __name__ == "__main__":
    import sys
    res = parse_mf_t12(sys.argv[1])
    print(f"format={res.format_guess} | period={res.period} | lines={len(res.lines)} "
          f"| coverage={res.coverage*100:.0f}%")
    print(f"computed income={res.computed['income']:,.0f} expense={res.computed['expense']:,.0f} "
          f"NOI={res.computed['noi']:,.0f} | reported={res.reported}")
    for w in res.warnings:
        print("  WARN:", w)
    for ln in res.unmapped:
        print("  UNMAPPED:", ln.acct, repr(ln.name), f"{ln.total:,.2f}")
