"""
Excel output writer — six-tab workbook with full professional formatting.

Theme: charcoal + white headers, modern/minimal aesthetic.
Includes: alternating row banding, color-coded status/care levels, currency
formatting, date formatting, auto-filters, frozen panes, print setup,
section dividers in summary, severity-coded exceptions.
"""

from __future__ import annotations

import io
from typing import Dict, List, Optional

import pandas as pd
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ============================================================================
# Theme — Charcoal + white, modern minimal
# ============================================================================
FONT_NAME = "Calibri"

# Header (charcoal)
HEADER_FILL = PatternFill("solid", fgColor="2B2B2B")
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Body
BODY_FONT = Font(name=FONT_NAME, size=10, color="222222")
BODY_FONT_BOLD = Font(name=FONT_NAME, size=10, color="222222", bold=True)

# Row banding (subtle gray every other row)
BAND_FILL = PatternFill("solid", fgColor="F5F5F5")

# Borders
THIN_GRAY = Side(border_style="thin", color="D4D4D4")
MEDIUM_DARK = Side(border_style="medium", color="2B2B2B")
ALL_BORDERS = Border(top=THIN_GRAY, left=THIN_GRAY, right=THIN_GRAY, bottom=THIN_GRAY)

# Status colors
STATUS_FILLS = {
    "Occupied": PatternFill("solid", fgColor="D4EDDA"),
    "Vacant":   PatternFill("solid", fgColor="F8D7DA"),
    "Hold":     PatternFill("solid", fgColor="FFF3CD"),
    "Notice":   PatternFill("solid", fgColor="FFF3CD"),
    "Model":    PatternFill("solid", fgColor="FFF3CD"),
    "Down":     PatternFill("solid", fgColor="E2E3E5"),
}

# Care Level colors (pale → dark by intensity)
CARE_LEVEL_FILLS = {
    "Level 1":  PatternFill("solid", fgColor="E8EEF7"),
    "Level 2":  PatternFill("solid", fgColor="C5D5EC"),
    "Level 3":  PatternFill("solid", fgColor="9CB7DC"),
    "Level 4":  PatternFill("solid", fgColor="6E94C9"),
    "Level 5":  PatternFill("solid", fgColor="3D6BB1"),
    "Level 6+": PatternFill("solid", fgColor="1E4480"),
}
CARE_LEVEL_FONTS_DARK = {
    "Level 4":  Font(name=FONT_NAME, size=10, color="FFFFFF"),
    "Level 5":  Font(name=FONT_NAME, size=10, color="FFFFFF", bold=True),
    "Level 6+": Font(name=FONT_NAME, size=10, color="FFFFFF", bold=True),
}

# Care Type colors
CARE_TYPE_FILLS = {
    "IL": PatternFill("solid", fgColor="E1F5E7"),
    "AL": PatternFill("solid", fgColor="E3EDF7"),
    "MC": PatternFill("solid", fgColor="F4E4F2"),
}

# Number formats
FMT_CURRENCY = '$#,##0;($#,##0);"-"'
FMT_INT = '#,##0;(#,##0);"-"'
FMT_DATE = "mm/dd/yyyy"

CURRENCY_COLS = {
    "Market Rate", "Actual Rate", "Concession $", "Care Level $",
    "Med Mgmt $", "Pharmacy $", "Other LOC $", "Total LOC $",
    "Rate Gap", "Total Monthly Revenue", "Avg Actual Rate", "Total Actual $",
}
DATE_COLS = {
    "Move-in Date", "Move-out Date", "Concession End Date",
}


# ============================================================================
# Helpers
# ============================================================================
def _calc_column_widths(df: pd.DataFrame, sample_rows: int = 200) -> List[float]:
    widths = []
    for col in df.columns:
        header_len = len(str(col))
        sample = df[col].head(sample_rows).astype(str)
        max_data_len = sample.str.len().max() if not sample.empty else 0
        widths.append(min(max(header_len + 2, max_data_len + 2, 10), 36))
    return widths


def _apply_table_style(ws, df: pd.DataFrame, *, banding: bool = True,
                       autofilter: bool = True, freeze_first_col: bool = False):
    n_rows = len(df) + 1
    n_cols = len(df.columns)
    if n_cols == 0:
        return

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = Border(top=MEDIUM_DARK, bottom=MEDIUM_DARK, left=THIN_GRAY, right=THIN_GRAY)
    ws.row_dimensions[1].height = 32

    for r in range(2, n_rows + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = ALL_BORDERS
            if banding and (r % 2 == 0):
                cell.fill = BAND_FILL

    for i, w in enumerate(_calc_column_widths(df), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "B2" if freeze_first_col else "A2"
    if autofilter and n_rows > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.gridLines = False
    ws.print_title_rows = "1:1"
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4


def _apply_number_formats(ws, df: pd.DataFrame):
    for i, col in enumerate(df.columns, start=1):
        fmt = None
        if col in CURRENCY_COLS:
            fmt = FMT_CURRENCY
        elif col in DATE_COLS:
            fmt = FMT_DATE
        elif col == "Sq Ft":
            fmt = FMT_INT
        if fmt is None:
            continue
        for r in range(2, len(df) + 2):
            cell = ws.cell(row=r, column=i)
            cell.number_format = fmt
            if col in CURRENCY_COLS or col == "Sq Ft":
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_status_colors(ws, df: pd.DataFrame):
    if "Status" not in df.columns:
        return
    idx = list(df.columns).index("Status") + 1
    for r in range(2, len(df) + 2):
        cell = ws.cell(row=r, column=idx)
        v = str(cell.value or "").strip()
        if v in STATUS_FILLS:
            cell.fill = STATUS_FILLS[v]
            cell.font = BODY_FONT_BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_care_level_colors(ws, df: pd.DataFrame):
    if "Care Level" not in df.columns:
        return
    idx = list(df.columns).index("Care Level") + 1
    for r in range(2, len(df) + 2):
        cell = ws.cell(row=r, column=idx)
        v = str(cell.value or "").strip()
        if v in CARE_LEVEL_FILLS:
            cell.fill = CARE_LEVEL_FILLS[v]
            cell.font = CARE_LEVEL_FONTS_DARK.get(v, BODY_FONT_BOLD)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_care_type_colors(ws, df: pd.DataFrame):
    if "Care Type" not in df.columns:
        return
    idx = list(df.columns).index("Care Type") + 1
    for r in range(2, len(df) + 2):
        cell = ws.cell(row=r, column=idx)
        v = str(cell.value or "").strip()
        if v in CARE_TYPE_FILLS:
            cell.fill = CARE_TYPE_FILLS[v]
            cell.font = BODY_FONT_BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center")


# ============================================================================
# Tab-specific writers
# ============================================================================
def _style_data_tab(ws, df: pd.DataFrame, freeze_first_col=True):
    _apply_table_style(ws, df, banding=True, autofilter=True, freeze_first_col=freeze_first_col)
    _apply_number_formats(ws, df)
    _apply_status_colors(ws, df)
    _apply_care_level_colors(ws, df)
    _apply_care_type_colors(ws, df)


def _style_summary(ws, df: pd.DataFrame):
    if df.empty:
        _apply_table_style(ws, df)
        return

    n_rows = len(df) + 1

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = Border(top=MEDIUM_DARK, bottom=MEDIUM_DARK, left=THIN_GRAY, right=THIN_GRAY)
    ws.row_dimensions[1].height = 32

    section_starts = {
        "Total Units":                       "Inventory",
        "Bed Occupancy %":                   "Occupancy",
        "Avg Market Rate (all beds)":        "Pricing",
        "Total Market Rate $ (all beds)":    "Revenue (Monthly)",
    }

    for r in range(2, n_rows + 1):
        label_cell = ws.cell(row=r, column=1)
        value_cell = ws.cell(row=r, column=2)

        label_cell.font = BODY_FONT_BOLD
        label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        value_cell.font = BODY_FONT
        value_cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        if r % 2 == 0:
            label_cell.fill = BAND_FILL
            value_cell.fill = BAND_FILL

        v = value_cell.value
        if isinstance(v, (int, float)):
            label = str(label_cell.value or "")
            if "$" in label:
                value_cell.number_format = FMT_CURRENCY
            elif "%" in label or "Occupancy" in label:
                pass
            elif "Avg" in label or "Rate" in label:
                value_cell.number_format = FMT_CURRENCY
            else:
                value_cell.number_format = FMT_INT

        label_text = str(label_cell.value or "").strip()
        if label_text in section_starts:
            top_border = Side(border_style="medium", color="2B2B2B")
            label_cell.border = Border(top=top_border, left=THIN_GRAY, right=THIN_GRAY, bottom=THIN_GRAY)
            value_cell.border = Border(top=top_border, left=THIN_GRAY, right=THIN_GRAY, bottom=THIN_GRAY)
        else:
            label_cell.border = ALL_BORDERS
            value_cell.border = ALL_BORDERS

        ws.row_dimensions[r].height = 20

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = "A2"

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.gridLines = False
    ws.print_title_rows = "1:1"


def _style_by_type(ws, df: pd.DataFrame):
    _apply_table_style(ws, df, banding=True, autofilter=True)
    _apply_number_formats(ws, df)

    if "Category" in df.columns:
        cat_idx = list(df.columns).index("Category") + 1
        category_fill = PatternFill("solid", fgColor="DCE6F2")
        for r in range(2, len(df) + 2):
            cell = ws.cell(row=r, column=cat_idx)
            cell.fill = category_fill
            cell.font = BODY_FONT_BOLD
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _style_exceptions(ws, df: pd.DataFrame):
    _apply_table_style(ws, df, banding=True, autofilter=True)

    if "Issue" not in df.columns:
        return
    idx = list(df.columns).index("Issue") + 1

    severe = Font(name=FONT_NAME, size=10, color="C00000", bold=True)
    moderate = Font(name=FONT_NAME, size=10, color="C65911")
    info = Font(name=FONT_NAME, size=10, color="9C6500")

    for r in range(2, len(df) + 2):
        cell = ws.cell(row=r, column=idx)
        text = str(cell.value or "").lower()
        if any(kw in text for kw in [
            "missing resident", "zero/blank actual", "missing move-in",
            "vacant bed has resident", "missing care type",
        ]):
            cell.font = severe
        elif "large market-to-actual gap" in text:
            cell.font = moderate
        elif "unmapped" in text or "no care charge" in text:
            cell.font = info
        else:
            cell.font = BODY_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 22


def _style_mapping_audit(ws, df: pd.DataFrame):
    _apply_table_style(ws, df, banding=True, autofilter=True)


# ============================================================================
# Main entry point
# ============================================================================
def write_output(
    condensed: pd.DataFrame,
    normalized: pd.DataFrame,
    mapping_audit: pd.DataFrame,
    summary: pd.DataFrame,
    by_type: pd.DataFrame,
    exceptions: pd.DataFrame,
    run_metadata: Optional[Dict] = None,
) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        condensed.to_excel(writer, sheet_name="Condensed_RR", index=False)
        normalized.to_excel(writer, sheet_name="Normalized_Beds", index=False)
        summary.to_excel(writer, sheet_name="RR_Summary", index=False)
        by_type.to_excel(writer, sheet_name="RR_By_Type", index=False)
        exceptions.to_excel(writer, sheet_name="RR_Exceptions", index=False)
        mapping_audit.to_excel(writer, sheet_name="Mapping_Reference", index=False)

        wb = writer.book

        _style_data_tab(wb["Condensed_RR"], condensed)
        _style_data_tab(wb["Normalized_Beds"], normalized)
        _style_summary(wb["RR_Summary"], summary)
        _style_by_type(wb["RR_By_Type"], by_type)
        _style_exceptions(wb["RR_Exceptions"], exceptions)
        _style_mapping_audit(wb["Mapping_Reference"], mapping_audit)

        if run_metadata:
            run_df = pd.DataFrame({
                "Key": list(run_metadata.keys()),
                "Value": [str(v) for v in run_metadata.values()],
            })
            run_df.to_excel(writer, sheet_name="Run_Info", index=False)
            _style_summary(wb["Run_Info"], run_df)

        # Hide gridlines and color tabs across the workbook
        for sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_view.showGridLines = False
            wb[sheet_name].sheet_properties.tabColor = "2B2B2B"

        # Open on the analyst view by default
        wb.active = wb.sheetnames.index("Condensed_RR")

    return buf.getvalue()
