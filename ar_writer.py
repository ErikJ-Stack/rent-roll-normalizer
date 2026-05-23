"""
ar_writer.py — Write parsed AR data into the Analyzer's AR & Collections sheet.

Mirrors the shape of analyzer_rr_writer.populate_rr_input() and
t12_normalizer_writer.populate_t12_input(): takes Analyzer bytes (output
of the prior writer step), mutates the AR & Collections sheet, flips it
from hidden to visible, and returns updated bytes.

The cell map below MUST stay in sync with tools/migration/migrate_to_v0210.py
which constructs the sheet structurally. If either file changes, update
both.

Z1 = 1 is the AR-presence pivot — it enables the Workbook Health B43
conditional read and activates the P5 pre-export gate. Z1 = 0 (the
default after migration) keeps everything inert so non-AR runs are
fully regression-clean.
"""
from __future__ import annotations

import io
from typing import Optional

import openpyxl

from ar_normalizer import AROutput


AR_SHEET = "AR & Collections"

# Aging-bucket rows — keyed by the AROutput attribute *suffix*
# (the full attr is "total_" + suffix).
AGING_ROW_MAP = {
    "current_0_30": 9,
    "days_31_60": 10,
    "days_61_90": 11,
    "days_91_120": 12,
    "over_120": 13,
}

# Payer rows in §3 By-Payer Mix. Order matches mappings.py normalization
# targets (incl. the v0.2.10 Managed Care addition) and the substrate
# sheet layout. Any payer in AROutput.payer_outstanding NOT in this map
# is silently skipped — should never happen unless mappings.py grows
# new buckets without updating this writer + the substrate sheet.
PAYER_ROW_MAP = {
    "Private Pay": 30,
    "Medicaid": 31,
    "Medicare": 32,
    "Managed Care": 33,
    "VA Benefit": 34,
    "LTC Insurance": 35,
    "Self-Pay + Other": 36,
}

# §4 Roll-Forward rows — (substrate_row, AROutput attribute)
ROLLFORWARD_MAP = (
    (42, "prior_period_balance"),
    (43, "charges_period"),
    (44, "collections_period"),
    (45, "writeoffs_period"),
    (46, "adjustments_period"),
)

# §5 Flag rows
FLAG_ROW_90PLUS_CONCESSION = 62   # RR-join required; stubbed to 0 for now
FLAG_ROW_VACANT_WITH_AR = 63      # RR-join required; stubbed to 0 for now
FLAG_ROW_PAYER_CONCENTRATION = 64
FLAG_ROW_SUM_CHECK = 65
FLAG_ROW_PERIOD_MISMATCH = 66

# Settings band
CELL_AS_OF_DATE = "C3"
CELL_AR_PRESENCE = "Z1"


class AROutputError(Exception):
    """Raised when the Analyzer workbook is incompatible with the AR writer.

    Most common cause: workbook hasn't had substrate v0.2.10 applied, so
    the AR & Collections sheet doesn't exist.
    """


def populate_ar_collections(
    analyzer_input,
    ar_output: AROutput,
    as_of_date: Optional[str] = None,
    source_filename: Optional[str] = None,
    ar_version: str = "0.1.0",
) -> bytes:
    """Write parsed AR into the Analyzer's AR & Collections sheet.

    Args:
        analyzer_input: bytes (typical — from RR/T12 writer output) or
            a path-like / file-like accepted by openpyxl.load_workbook.
        ar_output: result from ar_normalizer.parse_ar_file()
        as_of_date: ISO date string ("YYYY-MM-DD") to write to C3,
            overriding the default formula =IFERROR(RR_Period_Date,"").
            Pass None to keep the default.
        source_filename: reserved for future traceability stamping.
            Currently unused.
        ar_version: AR module version, reserved for future Cover stamp.

    Returns:
        bytes of the modified workbook.

    Raises:
        AROutputError: if AR & Collections sheet is absent (substrate
            v0.2.10 not applied to this Analyzer).
    """
    if isinstance(analyzer_input, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(analyzer_input))
    else:
        wb = openpyxl.load_workbook(analyzer_input)

    if AR_SHEET not in wb.sheetnames:
        raise AROutputError(
            f"Analyzer is missing '{AR_SHEET}' sheet — substrate v0.2.10 "
            f"not applied. Run tools/migration/migrate_to_v0210.py on this "
            f"workbook (or replace it with the bundled Analyzer)."
        )

    ws = wb[AR_SHEET]

    # Step 1: presence flag — enables Workbook Health B43 pivot + P5 gate.
    ws[CELL_AR_PRESENCE] = 1

    # Step 2: AR as-of override (analyst's date wins over the
    # default =IFERROR(RR_Period_Date,"") formula).
    if as_of_date:
        ws[CELL_AS_OF_DATE] = as_of_date

    # Step 3: §1 Aging Summary buckets
    for suffix, row in AGING_ROW_MAP.items():
        attr = f"total_{suffix}"
        ws[f"C{row}"] = float(getattr(ar_output, attr, 0.0))

    # Step 4: §3 By-Payer Mix
    for payer, row in PAYER_ROW_MAP.items():
        outstanding = float(ar_output.payer_outstanding.get(payer, 0.0))
        ninety_plus = float(ar_output.payer_90_plus.get(payer, 0.0))
        ws[f"C{row}"] = outstanding
        pct_90 = (ninety_plus / outstanding) if outstanding > 0 else 0.0
        ws[f"E{row}"] = pct_90

    # Step 5: §4 Roll-Forward (only write present fields — None preserves
    # the substrate's 0 / blank defaults).
    for row, attr in ROLLFORWARD_MAP:
        value = getattr(ar_output, attr, None)
        if value is not None:
            ws[f"C{row}"] = float(value)

    # Step 6: §5 Flags
    ws[f"C{FLAG_ROW_SUM_CHECK}"] = int(ar_output.sum_check_mismatch_count)

    # Payer concentration: 1 if any single payer holds > 60% of 90+ balance
    concentration_flag = 0
    if ar_output.total_90_plus > 0 and ar_output.payer_90_plus:
        max_90 = max(ar_output.payer_90_plus.values())
        if max_90 / ar_output.total_90_plus > 0.60:
            concentration_flag = 1
    ws[f"C{FLAG_ROW_PAYER_CONCENTRATION}"] = concentration_flag

    # RR-join flags — TBD when AR↔RR join logic ships. Stub to 0 so the
    # cells don't display stale data. ar_writer extension can populate
    # these by reading 'Rent Roll Input' from the same workbook.
    ws[f"C{FLAG_ROW_90PLUS_CONCESSION}"] = 0
    ws[f"C{FLAG_ROW_VACANT_WITH_AR}"] = 0

    # Period-date mismatch: only fires when analyst overrode C3 and the
    # override differs from the RR period. If as_of_date wasn't passed,
    # the default formula keeps C3 == RR_Period_Date and no mismatch.
    # We don't try to resolve the RR period here from Python — the cell
    # comparison happens at Excel formula time elsewhere. Set 0 stub.
    ws[f"C{FLAG_ROW_PERIOD_MISMATCH}"] = 0

    # Step 7: Flip visibility — analyst now sees the tab.
    ws.sheet_state = "visible"

    # Save to bytes (pipeline pattern, matches RR/T12 writers)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
