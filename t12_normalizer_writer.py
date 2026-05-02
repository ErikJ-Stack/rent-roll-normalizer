"""
T12 Normalizer Writer — Track 2 destination writer
====================================================

Loads the user's analysis workbook (`ALF_Financial_Analyzer_Only.xlsx` or the
standalone `ALF_T12-_Normalizer.xlsx` at template v0.1.4), writes the parsed
T12 GL detail into the `T12 Input` sheet, and returns the modified workbook
as bytes.

Naming history (see SPEC-T12.md §"Module naming history"):
  - `t12_writer.py` — Track 1 module that writes RR data into `Rent Roll Input`.
  - `t12_normalizer_writer.py` — THIS module, writes T12 data into `T12 Input`.
  Both write into a T12-shaped destination workbook but handle different
  inputs and different sheets.

Preservation guarantees (per SPEC-T12.md §"T12 file expected structure"):
  - Rows 1-10 of `T12 Input` (title, instructions, layout note, blank): untouched
  - Row 11 col A, B, O, P (column headers): untouched
  - Row 11 cols C-N (month labels): cleared and rewritten per upload
  - Rows 12-511 cols A-O (data area): cleared and rewritten per upload
  - Rows 12-511 col P (Coverage Check formula): UNTOUCHED (left in place)
  - All other sheets (T12_Calc helper col N, T12 Raw Data, Monthly Trending,
    Description_Map, Mapping Review, T12 Analytics, Rent Roll Input,
    Rent Roll Recon, RR_Calc, UW Output, Cover): untouched
  - Workbook-scoped named ranges (`DescMap_Description`, `DescMap_Label`):
    untouched

Idempotent re-run: clearing happens before writing so re-uploading a different
T12 doesn't leave ghost rows.

Capacity: 500 GL rows max. Salem produces 72, Briar Glen 91. Plenty of
headroom. Raises `T12NormalizerCapacityError` if exceeded — workbook would
need col P formula extension first.
"""

from __future__ import annotations

import datetime as dt
import io
from typing import Any, Dict, List, Optional, Sequence

import openpyxl

from t12_normalizer import T12ParseResult


# ---------------------------------------------------------------------------
# Layout constants — must match the v0.1.4 substrate
# ---------------------------------------------------------------------------

T12_INPUT_SHEET = "T12 Input"

DATA_START_ROW = 12
DATA_END_ROW = 511                  # also the col P formula extent
MAX_GL_ROWS = DATA_END_ROW - DATA_START_ROW + 1  # 500

# Data area column indices (1-based)
COL_ACCOUNT = 1                     # A
COL_DESCRIPTION = 2                 # B
COL_MONTH_START = 3                 # C
COL_MONTH_END = 14                  # N
COL_TOTAL = 15                      # O
COL_COVERAGE_CHECK = 16             # P (preserve)

LABEL_ROW = 11                      # row that carries month labels (C11:N11)
LABEL_COL_START = 3                 # C
LABEL_COL_END = 14                  # N

# Description_Map layout
DESCMAP_SHEET = "Description_Map"
DESCMAP_DATA_START_ROW = 5
DESCMAP_COL_DESC = 1                # A
DESCMAP_COL_LABEL = 2               # B
DESCMAP_COL_SECTION = 3             # C
DESCMAP_COL_CARETYPE = 4            # D
DESCMAP_COL_FLAG = 5                # E

# Run_Info tab — created if absent, otherwise appended to
RUN_INFO_SHEET = "Run_Info"


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class T12NormalizerCapacityError(Exception):
    """Raised when the parsed T12 has more GL rows than the workbook can hold."""


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def populate_t12_input(
    analyzer_bytes: bytes,
    parse_result: T12ParseResult,
    *,
    new_descmap_entries: Optional[Sequence[Dict[str, Any]]] = None,
    source_filename: str = "",
    t12_version: str = "",
    t12_last_updated: str = "",
) -> bytes:
    """Write T12 GL detail into the user's Analyzer / Normalizer workbook.

    Args:
        analyzer_bytes: bytes of the uploaded destination workbook
        parse_result: T12ParseResult from `t12_normalizer.parse_t12()`
        new_descmap_entries: optional list of dicts to append to Description_Map.
            Each dict has keys: description, label, section, caretype, flag.
            Used when the user resolves UNMATCHED via the in-app matcher.
        source_filename: original T12 filename for the Run_Info tab
        t12_version: T12 module version string for Run_Info
        t12_last_updated: T12 module last-updated date for Run_Info

    Returns:
        bytes of the modified .xlsx workbook

    Raises:
        T12NormalizerCapacityError: if more than 500 GL rows
        ValueError: if the workbook is missing required sheets
    """
    n_rows = len(parse_result.gl_rows)
    if n_rows > MAX_GL_ROWS:
        raise T12NormalizerCapacityError(
            f"T12 has {n_rows} GL detail rows, but the destination workbook's "
            f"`{T12_INPUT_SHEET}` sheet only has capacity for {MAX_GL_ROWS} rows "
            f"(rows {DATA_START_ROW}-{DATA_END_ROW}). The col P Coverage Check "
            f"formula would need to be extended in the workbook before this "
            f"T12 can be loaded."
        )

    if len(parse_result.month_labels) != 12:
        raise ValueError(
            f"Expected 12 month labels, got {len(parse_result.month_labels)}: "
            f"{parse_result.month_labels}"
        )

    wb = openpyxl.load_workbook(io.BytesIO(analyzer_bytes), data_only=False)

    if T12_INPUT_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Destination workbook is missing required sheet '{T12_INPUT_SHEET}'. "
            f"Found: {wb.sheetnames}. Is this a v0.1.4 substrate Analyzer / "
            f"Normalizer template?"
        )

    ws = wb[T12_INPUT_SHEET]

    # --- Step 1: Clear prior data in A12:O511 (col P preserved) -----------
    # This is the idempotency guarantee: re-uploading a different T12 must
    # not leave ghost rows from a prior run.
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        for c in range(COL_ACCOUNT, COL_TOTAL + 1):  # cols A-O = 1-15
            ws.cell(row=r, column=c).value = None

    # Also clear C11:N11 (the month-label row) before writing the new labels.
    # A11/B11/O11/P11 (column headers) are left in place.
    for c in range(LABEL_COL_START, LABEL_COL_END + 1):
        ws.cell(row=LABEL_ROW, column=c).value = None

    # --- Step 2: Write the 12 month labels to C11:N11 ---------------------
    for i, label in enumerate(parse_result.month_labels):
        cell = ws.cell(row=LABEL_ROW, column=LABEL_COL_START + i)
        cell.value = label
        # The substrate's C11:N11 already had whatever date format it had;
        # explicit text format avoids Excel auto-converting "Feb 2025" to a
        # date serial when the user re-saves.
        cell.number_format = "@"

    # --- Step 3: Write GL detail rows to A12:O... -------------------------
    for i, gl in enumerate(parse_result.gl_rows):
        excel_row = DATA_START_ROW + i

        # Col A — account # (None for MRI's empty string, so the cell stays
        # blank rather than holding an empty-string sentinel)
        ws.cell(row=excel_row, column=COL_ACCOUNT).value = (
            gl.account if gl.account else None
        )
        # Col B — description
        ws.cell(row=excel_row, column=COL_DESCRIPTION).value = gl.description
        # Cols C-N — 12 monthly amounts
        for j, v in enumerate(gl.monthly):
            ws.cell(row=excel_row, column=COL_MONTH_START + j).value = v
        # Col O — T12 total
        ws.cell(row=excel_row, column=COL_TOTAL).value = gl.total
        # Col P — UNTOUCHED. The Coverage Check formula already lives there.

    # --- Step 4: Append any UNMATCHED-resolution mappings to Description_Map
    if new_descmap_entries:
        _append_descmap_entries(wb, new_descmap_entries)

    # --- Step 5: Update Run_Info tab --------------------------------------
    _upsert_run_info(
        wb,
        parse_result=parse_result,
        source_filename=source_filename,
        t12_version=t12_version,
        t12_last_updated=t12_last_updated,
        new_descmap_count=len(new_descmap_entries) if new_descmap_entries else 0,
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Description_Map append (UNMATCHED-resolution writes)
# ---------------------------------------------------------------------------

def _append_descmap_entries(
    wb: openpyxl.Workbook,
    entries: Sequence[Dict[str, Any]],
) -> None:
    """Append new mapping rows to `Description_Map` after the last data row.

    The dynamic named ranges (`DescMap_Description`, `DescMap_Label`) auto-
    extend via their COUNTA expression, so no formula edits are needed.
    """
    if DESCMAP_SHEET not in wb.sheetnames:
        # If the substrate is somehow missing Description_Map, fail loud
        # rather than silently dropping the user's mapping work.
        raise ValueError(
            f"Cannot append mappings: workbook is missing '{DESCMAP_SHEET}' sheet."
        )
    ws = wb[DESCMAP_SHEET]

    # Find the last row with a non-empty description in col A starting from
    # the data start row. We can't trust ws.max_row because openpyxl sometimes
    # carries a phantom max_row from formatting that extends past the data.
    last_row = DESCMAP_DATA_START_ROW - 1
    for r in range(DESCMAP_DATA_START_ROW, ws.max_row + 1):
        v = ws.cell(r, DESCMAP_COL_DESC).value
        if v is not None and str(v).strip():
            last_row = r

    write_row = last_row + 1
    for entry in entries:
        desc = entry.get("description", "")
        if not desc or not str(desc).strip():
            continue  # skip blank rows defensively
        ws.cell(row=write_row, column=DESCMAP_COL_DESC).value = str(desc).strip()
        ws.cell(row=write_row, column=DESCMAP_COL_LABEL).value = entry.get("label") or None
        ws.cell(row=write_row, column=DESCMAP_COL_SECTION).value = entry.get("section") or None
        ws.cell(row=write_row, column=DESCMAP_COL_CARETYPE).value = entry.get("caretype") or "-"
        ws.cell(row=write_row, column=DESCMAP_COL_FLAG).value = entry.get("flag") or None
        write_row += 1


# ---------------------------------------------------------------------------
# Run_Info tab — append-or-create
# ---------------------------------------------------------------------------

def _upsert_run_info(
    wb: openpyxl.Workbook,
    *,
    parse_result: T12ParseResult,
    source_filename: str,
    t12_version: str,
    t12_last_updated: str,
    new_descmap_count: int,
) -> None:
    """Add T12-side run metadata to a `Run_Info` tab.

    If the tab already exists (e.g., RR side wrote it earlier in the session),
    append T12 keys after the existing content with a separator row. If absent,
    create it.
    """
    ts = dt.datetime.now().isoformat(timespec="seconds")

    rows: List[List[Any]] = [
        ["T12 Module Version",   t12_version or "(unset)"],
        ["T12 Last Updated",     t12_last_updated or "(unset)"],
        ["T12 Run Timestamp",    ts],
        ["T12 Source File",      source_filename or "(uploaded)"],
        ["T12 Format Detected",  parse_result.format_name],
        ["T12 Source Sheet",     parse_result.sheet_name],
        ["T12 GL Rows Written",  len(parse_result.gl_rows)],
        ["T12 Months Detected",  ", ".join(parse_result.month_labels)],
        ["T12 UNMATCHED at parse",  len(parse_result.unmatched)],
        ["T12 Description_Map appends",  new_descmap_count],
    ]

    if RUN_INFO_SHEET in wb.sheetnames:
        ws = wb[RUN_INFO_SHEET]
        # Find first empty row in col A; if sheet has any content, leave a
        # blank separator row before our T12 block.
        last_row = 0
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value is not None or ws.cell(r, 2).value is not None:
                last_row = r
        write_row = last_row + 2 if last_row > 0 else 1
        if last_row > 0:
            ws.cell(write_row - 1, 1).value = "—"
    else:
        ws = wb.create_sheet(RUN_INFO_SHEET)
        ws.cell(1, 1).value = "Key"
        ws.cell(1, 2).value = "Value"
        ws.cell(1, 1).font = openpyxl.styles.Font(bold=True)
        ws.cell(1, 2).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 60
        write_row = 2

    for key, val in rows:
        ws.cell(write_row, 1).value = key
        ws.cell(write_row, 2).value = val
        write_row += 1
